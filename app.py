import streamlit as st
import streamlit.components.v1 as _stc
import anthropic
import base64
import hashlib
import json
import math
import os
import re
import uuid
import pathlib
import datetime
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

try:
    from shapely.geometry import Polygon as ShapelyPolygon
    from shapely.validation import make_valid
    _SHAPELY_OK = True
except ImportError:
    _SHAPELY_OK = False

try:
    import ezdxf
    _EZDXF_OK = True
except ImportError:
    _EZDXF_OK = False

# Logo embebido como base64
_LOGO_PATH = pathlib.Path(__file__).parent / "logo.png"
_LOGO_B64 = base64.b64encode(_LOGO_PATH.read_bytes()).decode() if _LOGO_PATH.exists() else ""

# Wireframe imagen estado vacío
_WIRE_PATH = pathlib.Path(__file__).parent / "wireframe.png"
_WIRE_B64 = base64.b64encode(_WIRE_PATH.read_bytes()).decode() if _WIRE_PATH.exists() else ""

PROJECTS_DIR = pathlib.Path(__file__).parent / "projects"


class _Proyecto:
    """Lightweight reference to a saved project — works with Supabase or local files."""
    __slots__ = ("name", "_id", "_path")
    def __init__(self, name: str, id: str = None, path: pathlib.Path = None):
        self.name  = name
        self._id   = id
        self._path = path


def _get_supabase():
    if "_sb_client" not in st.session_state:
        try:
            from supabase import create_client
            _sb = st.secrets.get("supabase", {}) or {}
            _url = _sb.get("url", "")
            _key = _sb.get("key", "")
            st.session_state["_sb_client"] = create_client(_url, _key) if (_url and _key) else None
        except Exception:
            st.session_state["_sb_client"] = None
    return st.session_state.get("_sb_client")


def guardar_proyecto(nombre: str, estado: dict, tipo: str = "", zona: str = "") -> "_Proyecto":
    if not tipo:
        if "industrial_result" in estado:
            tipo = "industrial"
        elif "residencial_result" in estado:
            tipo = "residencial"
        else:
            tipo = "inmobiliario"
    if not zona:
        zona = str(estado.get("zona") or "")
    usuario = st.session_state.get("_username", "unknown")
    sb = _get_supabase()
    if sb:
        try:
            row = {
                "usuario": usuario,
                "nombre_proyecto": nombre.strip() or "sin_nombre",
                "tipo": tipo,
                "zona": zona,
                "datos": estado,
                "resumen": estado.get("resumen") or {},
            }
            resp = sb.table("proyectos").insert(row).execute()
            if resp.data:
                _id = resp.data[0]["id"]
                display = f"{nombre.strip() or 'sin_nombre'}  ·  {datetime.datetime.now().strftime('%d/%m/%Y')}"
                return _Proyecto(display, id=_id)
        except Exception:
            pass
    PROJECTS_DIR.mkdir(exist_ok=True)
    slug = re.sub(r"[^\w\-]", "_", nombre.strip())[:40]
    fp = PROJECTS_DIR / f"{slug}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.json"
    with open(fp, "w", encoding="utf-8") as f:
        json.dump({**estado, "nombre": nombre, "fecha": datetime.datetime.now().isoformat()}, f, ensure_ascii=False, indent=2)
    return _Proyecto(fp.name, path=fp)


def listar_proyectos(con_resumen: bool = False) -> list:
    usuario = st.session_state.get("_username", "unknown")
    sb = _get_supabase()
    if sb:
        try:
            cols = "id, nombre_proyecto, tipo, zona, creado_en" + (", resumen" if con_resumen else "")
            resp = (
                sb.table("proyectos")
                  .select(cols)
                  .eq("usuario", usuario)
                  .order("creado_en", desc=True)
                  .limit(100)
                  .execute()
            )
            result = []
            for row in (resp.data or []):
                fecha = (row.get("creado_en") or "")[:10]
                tipo_tag = row.get("tipo") or ""
                display = f"{row['nombre_proyecto']}  [{tipo_tag}  ·  {fecha}]"
                p = _Proyecto(display, id=row["id"])
                if con_resumen:
                    p._resumen = row.get("resumen") or {}
                    p._tipo    = tipo_tag
                    p._zona    = row.get("zona") or ""
                    p._fecha   = fecha
                    p._nombre  = row.get("nombre_proyecto") or display
                result.append(p)
            return result
        except Exception:
            pass
    if not PROJECTS_DIR.exists():
        return []
    return [
        _Proyecto(fp.name, path=fp)
        for fp in sorted(PROJECTS_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
    ]


def cargar_proyecto(ref) -> dict:
    if isinstance(ref, _Proyecto) and ref._id:
        sb = _get_supabase()
        if sb:
            try:
                resp = sb.table("proyectos").select("datos").eq("id", ref._id).single().execute()
                return (resp.data or {}).get("datos") or {}
            except Exception:
                pass
    if isinstance(ref, _Proyecto):
        path = ref._path
    elif isinstance(ref, pathlib.Path):
        path = ref
    else:
        path = pathlib.Path(str(ref))
    if path and path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def generar_link_compartido(proyecto_id: str) -> str:
    """Genera token único, lo guarda en Supabase y retorna la URL completa."""
    token = uuid.uuid4().hex          # 32 chars hex sin guiones
    sb = _get_supabase()
    if sb and proyecto_id:
        try:
            # Intenta update en columna top-level share_token
            sb.table("proyectos").update({"share_token": token}).eq("id", proyecto_id).execute()
        except Exception:
            try:
                # Fallback: guarda dentro del JSONB datos
                resp = sb.table("proyectos").select("datos").eq("id", proyecto_id).single().execute()
                datos = dict((resp.data or {}).get("datos") or {})
                datos["_share_token"] = token
                sb.table("proyectos").update({"datos": datos}).eq("id", proyecto_id).execute()
            except Exception:
                return ""
    base = (st.secrets.get("app") or {}).get("base_url", "http://localhost:8501")
    return f"{base}?share={token}"


def _irr_bisect(flujos, lo=-0.9999, hi=10.0, tol=1e-7, max_iter=300):
    """Bisection method to find monthly IRR."""
    def npv(r):
        return sum(f / (1 + r) ** i for i, f in enumerate(flujos))
    try:
        v_lo, v_hi = npv(lo), npv(hi)
        if v_lo * v_hi > 0:
            return None
        for _ in range(max_iter):
            mid = (lo + hi) / 2
            if abs(hi - lo) < tol:
                return mid
            if npv(mid) * v_lo < 0:
                hi = mid
            else:
                lo = mid
                v_lo = npv(lo)
        return (lo + hi) / 2
    except Exception:
        return None


# ═══════════════════════════════════════════════════════
# ANÁLISIS LOGÍSTICO / INDUSTRIAL
# ═══════════════════════════════════════════════════════

def calcular_industrial(inp: dict) -> dict:
    area = inp.get("area_terreno", 0)
    costo_terreno = inp.get("costo_terreno", 0)
    tipo_nave = inp.get("tipo_nave", "Almacén Logístico")
    zonificacion = inp.get("zonificacion", "I2")
    uso = inp.get("uso", "Uso directo")

    # Área techada / libre según porcentaje definido por usuario
    pct_techada = max(min(inp.get("pct_techada", 75), 95), 30) / 100
    area_nave = area * pct_techada          # nave techada
    area_libre = area * (1 - pct_techada)   # patios, maniobras, circulación

    # Costos de construcción — usuario puede sobreescribir
    # Referencia: Parque Logístico 47 (Lima, ~14,300 m², Clase A, 13.6m clara) = $291/m² all-in
    # Industrial es 3-4x más barato que residencial: estructura metálica, sin acabados
    _DEFAULTS_NAVE = {
        "Almacén Logístico":        280,   # 12-14m clara, estructura metálica, losa industrial
        "Nave Industrial":          300,   # 10-14m, estructura metálica + concreto, uso mixto
        "Cross-docking":            420,   # docks múltiples, mayor complejidad MEP
        "Producción / Manufactura": 380,   # refuerzo de losa, instalaciones especiales
    }
    _default_nave = _DEFAULTS_NAVE.get(tipo_nave, 300)
    _cn = inp.get("costo_nave_m2")
    costo_nave_m2 = _cn if _cn is not None else _default_nave
    _cp = inp.get("costo_piso_libre_m2")
    costo_piso_libre_m2 = _cp if _cp is not None else 80

    costo_nave_total = area_nave * costo_nave_m2
    costo_pisos_libres = area_libre * costo_piso_libre_m2
    costo_construccion = costo_nave_total + costo_pisos_libres
    pct_indirectos = inp.get("pct_indirectos", 5.0) / 100
    soft_costs = costo_construccion * pct_indirectos
    alcabala = costo_terreno * 0.03 if inp.get("include_alcabala", True) else 0
    costo_total = costo_terreno + alcabala + costo_construccion + soft_costs
    costo_por_m2_nave = costo_total / area_nave if area_nave > 0 else 0

    # ── Crédito Terreno ──────────────────────────────────────
    costo_terreno_alcabala = costo_terreno + alcabala
    dp_terreno_pct = max(0.0, min(100.0, inp.get("dp_terreno_pct", 40.0))) / 100
    capital_propio_terreno = costo_terreno_alcabala * dp_terreno_pct
    monto_credito_terreno  = costo_terreno_alcabala * (1 - dp_terreno_pct)
    tasa_terreno  = inp.get("tasa_terreno",  8.0) / 100
    plazo_terreno = max(inp.get("plazo_terreno", 10), 1)
    cuota_terreno = 0.0
    if monto_credito_terreno > 0 and tasa_terreno > 0:
        _rt = tasa_terreno / 12
        _nt = plazo_terreno * 12
        cuota_terreno = monto_credito_terreno * _rt * (1 + _rt)**_nt / ((1 + _rt)**_nt - 1)

    # ── Crédito Construcción ─────────────────────────────────
    costo_construccion_soft = costo_construccion + soft_costs
    dp_const_pct = max(0.0, min(100.0, inp.get("dp_const_pct", 30.0))) / 100
    capital_propio_const = costo_construccion_soft * dp_const_pct
    monto_credito_const  = costo_construccion_soft * (1 - dp_const_pct)
    tasa_const  = inp.get("tasa_const",  9.0) / 100
    plazo_const = max(inp.get("plazo_const", 8), 1)
    cuota_const = 0.0
    if monto_credito_const > 0 and tasa_const > 0:
        _rc = tasa_const / 12
        _nc = plazo_const * 12
        cuota_const = monto_credito_const * _rc * (1 + _rc)**_nc / ((1 + _rc)**_nc - 1)

    # ── Totales combinados ───────────────────────────────────
    capital_propio = capital_propio_terreno + capital_propio_const
    monto_credito  = monto_credito_terreno  + monto_credito_const
    cuota_mensual  = cuota_terreno          + cuota_const
    pct_credito    = (monto_credito / costo_total * 100) if costo_total > 0 else 0
    plazo_anos     = max(plazo_terreno, plazo_const)
    tasa_anual     = (tasa_terreno + tasa_const) / 2

    # Renta base sobre área techada (nave arrendable)
    renta_m2_mes = inp.get("renta_m2_mes", 0)
    renta_total_mes = renta_m2_mes * area_nave
    gastos_operacion = renta_total_mes * 12 * 0.08
    renta_neta_anual = max(renta_total_mes * 12 - gastos_operacion, 0)
    yield_bruto = (renta_total_mes * 12 / costo_total * 100) if costo_total > 0 else 0
    yield_neto = (renta_neta_anual / costo_total * 100) if costo_total > 0 else 0
    payback_anos = (costo_total / renta_neta_anual) if renta_neta_anual > 0 else None
    flujo_mensual = renta_total_mes - cuota_mensual if uso == "Inversión" else None
    dscr = (renta_total_mes / cuota_mensual) if cuota_mensual > 0 else None
    alquiler_vs_compra = renta_total_mes - cuota_mensual if uso == "Uso directo" else None

    # ── Indexación de renta por contrato plurianual ──
    tipo_contrato_ind    = inp.get("tipo_contrato", "Anual")
    ajuste_anual_ind_pct = inp.get("ajuste_anual_pct", 0.0) / 100
    inicio_ajuste_ind    = int(inp.get("inicio_ajuste_ano", 2))
    _es_plurianual_ind   = tipo_contrato_ind != "Anual" and ajuste_anual_ind_pct > 0

    def _renta_neta_ind(yr):
        """Renta neta anual del año yr con o sin ajuste contractual."""
        if not _es_plurianual_ind or yr < inicio_ajuste_ind:
            return renta_neta_anual
        factor = (1 + ajuste_anual_ind_pct) ** (yr - inicio_ajuste_ind + 1)
        return max(renta_total_mes * factor * 12 * (1 - 0.08), 0)

    # Proyecciones año 3 y 5 (renta mensual/m² y yield)
    _rn3 = _renta_neta_ind(3)
    _rn5 = _renta_neta_ind(5)
    renta_m2_ano3  = (_rn3 / 12 / area_nave) if area_nave > 0 else 0
    renta_m2_ano5  = (_rn5 / 12 / area_nave) if area_nave > 0 else 0
    yield_neto_ano3 = (_rn3 / costo_total * 100) if costo_total > 0 else 0
    yield_neto_ano5 = (_rn5 / costo_total * 100) if costo_total > 0 else 0

    # Payback indexado (años para recuperar inversión con renta creciente)
    payback_indexado_ind = None
    if _es_plurianual_ind and renta_neta_anual > 0:
        _acum = 0.0
        for _y in range(1, 41):
            _acum += _renta_neta_ind(_y)
            if _acum >= costo_total:
                payback_indexado_ind = _y
                break

    # ── Escudo fiscal por depreciación de la nave (Perú: 5%/año, 20 años, IR 29.5%) ──
    VIDA_UTIL_NAVE = 20        # años de vida útil tributaria (DS 122-94-EF)
    TASA_IR = 0.295            # Impuesto a la Renta corporativo Perú
    depreciacion_anual = costo_nave_total / VIDA_UTIL_NAVE   # base: solo la nave, no el terreno
    ahorro_fiscal_anual = depreciacion_anual * TASA_IR       # ahorro en IR por depreciación
    ahorro_fiscal_mensual = ahorro_fiscal_anual / 12
    # Costo efectivo mensual de compra para Uso directo (cuota – escudo fiscal)
    cuota_efectiva_mensual = max(cuota_mensual - ahorro_fiscal_mensual, 0)

    # ── IRR y flujo de caja anual (10 años, solo inversión) ──
    flujo_anual = []
    irr_anual = None
    van_10 = None
    tasa_desc = 0.10
    APRECIACION_IND = 0.03

    if uso == "Inversión" and renta_neta_anual > 0 and capital_propio > 0:
        flujo_anual = [-capital_propio]
        for yr in range(1, 11):
            cuota_yr = cuota_mensual * 12 if yr <= plazo_anos else 0
            flujo_yr = _renta_neta_ind(yr) - cuota_yr
            if yr == 10:
                saldo_deuda = 0
                if monto_credito > 0 and tasa_anual > 0 and yr <= plazo_anos:
                    r_m = tasa_anual / 12
                    n_p = yr * 12
                    saldo_deuda = max(
                        monto_credito * (1+r_m)**n_p - cuota_mensual * ((1+r_m)**n_p - 1) / r_m, 0)
                flujo_yr += costo_total * (1 + APRECIACION_IND)**10 - saldo_deuda
            flujo_anual.append(flujo_yr)
        irr_r = _irr_bisect(flujo_anual)
        irr_anual = round(irr_r * 100, 1) if irr_r is not None else None
        van_10 = sum(f / (1 + tasa_desc)**i for i, f in enumerate(flujo_anual))

    return {
        "area_terreno": area,
        "area_nave": area_nave, "area_libre": area_libre, "pct_techada": pct_techada * 100,
        "costo_terreno": costo_terreno, "alcabala": alcabala,
        "costo_nave_total": costo_nave_total, "costo_pisos_libres": costo_pisos_libres,
        "costo_construccion": costo_construccion, "soft_costs": soft_costs,
        "pct_indirectos": pct_indirectos * 100,
        "costo_total": costo_total, "costo_por_m2_nave": costo_por_m2_nave,
        "costo_nave_m2": costo_nave_m2, "costo_piso_libre_m2": costo_piso_libre_m2,
        "monto_credito": monto_credito, "capital_propio": capital_propio,
        "cuota_mensual": cuota_mensual, "pct_credito": pct_credito, "plazo_anos": plazo_anos,
        "tasa_anual": tasa_anual * 100,
        # Desglose terreno
        "costo_terreno_alcabala": costo_terreno_alcabala,
        "capital_propio_terreno": capital_propio_terreno,
        "monto_credito_terreno":  monto_credito_terreno,
        "cuota_terreno":          cuota_terreno,
        "dp_terreno_pct":         dp_terreno_pct * 100,
        "tasa_terreno":           tasa_terreno * 100,
        "plazo_terreno":          plazo_terreno,
        # Desglose construcción
        "costo_construccion_soft": costo_construccion_soft,
        "capital_propio_const":    capital_propio_const,
        "monto_credito_const":     monto_credito_const,
        "cuota_const":             cuota_const,
        "dp_const_pct":            dp_const_pct * 100,
        "tasa_const":              tasa_const * 100,
        "plazo_const":             plazo_const,
        "renta_m2_mes": renta_m2_mes, "renta_total_mes": renta_total_mes,
        "gastos_operacion": gastos_operacion, "renta_neta_anual": renta_neta_anual,
        "yield_bruto": yield_bruto, "yield_neto": yield_neto,
        "payback_anos": payback_anos, "flujo_mensual": flujo_mensual,
        "dscr": dscr, "alquiler_vs_compra": alquiler_vs_compra,
        "flujo_anual": flujo_anual, "irr_anual": irr_anual, "van_10": van_10,
        "tipo_nave": tipo_nave, "zonificacion": zonificacion, "uso": uso,
        "actividad_categoria": inp.get("actividad_categoria", ""),
        "actividad_descripcion": inp.get("actividad_descripcion", ""),
        # Escudo fiscal
        "depreciacion_anual": round(depreciacion_anual),
        "ahorro_fiscal_anual": round(ahorro_fiscal_anual),
        "ahorro_fiscal_mensual": round(ahorro_fiscal_mensual),
        "cuota_efectiva_mensual": round(cuota_efectiva_mensual),
        "APRECIACION_IND": APRECIACION_IND,
        # Indexación contractual
        "tipo_contrato": tipo_contrato_ind,
        "ajuste_anual_pct": round(ajuste_anual_ind_pct * 100, 2),
        "inicio_ajuste_ano": inicio_ajuste_ind,
        "renta_m2_ano3": round(renta_m2_ano3, 2),
        "renta_m2_ano5": round(renta_m2_ano5, 2),
        "yield_neto_ano3": round(yield_neto_ano3, 1),
        "yield_neto_ano5": round(yield_neto_ano5, 1),
        "payback_indexado": payback_indexado_ind,
    }


# ═══════════════════════════════════════════════════════
# ANÁLISIS RESIDENCIAL
# ═══════════════════════════════════════════════════════

def calcular_residencial(inp: dict) -> dict:
    precio = inp.get("precio", 0)
    pct_pie = max(min(inp.get("pct_pie", 20), 100), 0) / 100
    tasa_anual = inp.get("tasa_anual", 8.5) / 100
    plazo_anos = max(inp.get("plazo_anos", 20), 1)
    uso = inp.get("uso", "Vivienda propia")

    pie = precio * pct_pie
    monto_credito = precio * (1 - pct_pie)
    cuota_mensual = 0.0
    total_pagado = precio
    total_intereses = 0.0
    n_meses = plazo_anos * 12

    if monto_credito > 0 and tasa_anual > 0:
        r = tasa_anual / 12
        n = n_meses
        cuota_mensual = monto_credito * r * (1 + r)**n / ((1 + r)**n - 1)
        total_pagado = pie + cuota_mensual * n
        total_intereses = total_pagado - precio

    ingreso_minimo = cuota_mensual / 0.30 if cuota_mensual > 0 else 0

    alquiler_mes = inp.get("alquiler_mes", 0)
    gastos_mes = inp.get("gastos_mes", 0)
    renta_neta_mes = max(alquiler_mes - gastos_mes, 0)
    yield_bruto = (alquiler_mes * 12 / precio * 100) if precio > 0 else 0
    yield_neto = (renta_neta_mes * 12 / precio * 100) if precio > 0 else 0
    payback_anos = (precio / (renta_neta_mes * 12)) if renta_neta_mes > 0 else None
    flujo_mensual = renta_neta_mes - cuota_mensual if uso == "Inversión" else None
    alquiler_equilibrio = cuota_mensual + gastos_mes

    # ── Indexación de renta por contrato plurianual ──
    tipo_contrato_res    = inp.get("tipo_contrato", "Anual")
    ajuste_anual_res_pct = inp.get("ajuste_anual_pct", 0.0) / 100
    inicio_ajuste_res    = int(inp.get("inicio_ajuste_ano", 2))
    _es_plurianual_res   = tipo_contrato_res != "Anual" and ajuste_anual_res_pct > 0

    def _alquiler_ano_res(yr):
        if not _es_plurianual_res or yr < inicio_ajuste_res:
            return alquiler_mes
        return alquiler_mes * (1 + ajuste_anual_res_pct) ** (yr - inicio_ajuste_res + 1)

    def _renta_neta_ano_res(yr):
        return max(_alquiler_ano_res(yr) - gastos_mes, 0) * 12

    alquiler_ano3 = round(_alquiler_ano_res(3))
    alquiler_ano5 = round(_alquiler_ano_res(5))
    yield_neto_ano3_res = (_renta_neta_ano_res(3) / precio * 100) if precio > 0 else 0
    yield_neto_ano5_res = (_renta_neta_ano_res(5) / precio * 100) if precio > 0 else 0

    payback_indexado_res = None
    if _es_plurianual_res and renta_neta_mes > 0:
        _acum = 0.0
        for _y in range(1, 51):
            _acum += _renta_neta_ano_res(_y)
            if _acum >= precio:
                payback_indexado_res = _y
                break

    tasa_apreciacion = max(inp.get("variacion_anual_pct", 4.0), 0.0) / 100
    tasa_apreciacion = tasa_apreciacion if tasa_apreciacion > 0 else 0.04
    valor_5 = precio * ((1 + tasa_apreciacion)**5)
    valor_10 = precio * ((1 + tasa_apreciacion)**10)
    ganancia_capital_5 = valor_5 - precio
    ganancia_capital_10 = valor_10 - precio

    amort_tabla = []
    if cuota_mensual > 0 and monto_credito > 0:
        saldo = monto_credito
        r = tasa_anual / 12
        for yr in range(1, min(plazo_anos + 1, 11)):
            interes_yr = 0
            capital_yr = 0
            for _ in range(12):
                i = saldo * r
                k = cuota_mensual - i
                interes_yr += i
                capital_yr += k
                saldo = max(saldo - k, 0)
            amort_tabla.append({"año": yr, "capital": capital_yr, "interes": interes_yr, "saldo": saldo})

    return {
        "precio": precio, "pie": pie, "pct_pie": pct_pie * 100,
        "monto_credito": monto_credito, "cuota_mensual": cuota_mensual,
        "total_pagado": total_pagado, "total_intereses": total_intereses,
        "ingreso_minimo": ingreso_minimo, "plazo_anos": plazo_anos,
        "tasa_anual": tasa_anual * 100, "n_meses": n_meses,
        "alquiler_mes": alquiler_mes, "gastos_mes": gastos_mes,
        "renta_neta_mes": renta_neta_mes, "yield_bruto": yield_bruto,
        "yield_neto": yield_neto, "payback_anos": payback_anos,
        "flujo_mensual": flujo_mensual, "alquiler_equilibrio": alquiler_equilibrio,
        "valor_5": valor_5, "valor_10": valor_10,
        "ganancia_capital_5": ganancia_capital_5, "ganancia_capital_10": ganancia_capital_10,
        "amort_tabla": amort_tabla, "uso": uso,
        "tasa_apreciacion_pct": round(tasa_apreciacion * 100, 1),
        # Market / zone fields (populated by caller)
        "zona": inp.get("zona", ""),
        "dormitorios": inp.get("dormitorios", ""),
        "m2": inp.get("m2", 0),
        "antiguedad": inp.get("antiguedad", 0),
        "precio_m2": inp.get("precio_m2", 0),
        "precio_m2_mercado": inp.get("precio_m2_mercado", 0),
        "yield_mercado_pct": inp.get("yield_mercado_pct", 0),
        "alquiler_mercado_m2": inp.get("alquiler_mercado_m2", 0),
        "variacion_anual_pct": inp.get("variacion_anual_pct", 0),
        # Indexación contractual
        "tipo_contrato": tipo_contrato_res,
        "ajuste_anual_pct": round(ajuste_anual_res_pct * 100, 2),
        "inicio_ajuste_ano": inicio_ajuste_res,
        "alquiler_ano3": alquiler_ano3,
        "alquiler_ano5": alquiler_ano5,
        "yield_neto_ano3": round(yield_neto_ano3_res, 1),
        "yield_neto_ano5": round(yield_neto_ano5_res, 1),
        "payback_indexado": payback_indexado_res,
    }


# ═══════════════════════════════════════════════════════
# CALCULADORA INVERSA DE TERRENO
# ═══════════════════════════════════════════════════════

def calcular_terreno_maximo(inp: dict) -> dict:
    zona          = inp.get("zona", "")
    m             = MERCADO.get(zona, {})
    area_terreno  = inp.get("area_terreno", 0)
    area_vendible = inp.get("area_vendible", 0)
    area_techada  = inp.get("area_techada", 0)
    num_pisos     = max(inp.get("num_pisos", 7), 1)
    n_estac       = inp.get("n_estac", 0)
    n_depositos   = inp.get("n_depositos", 0)
    precio_m2     = inp.get("precio_m2", m.get("precio_2br", 0))
    costo_const   = inp.get("costo_construccion", 700)
    costo_sotano  = inp.get("costo_sotano", 450)
    fee_constr    = inp.get("fee_constructora", 10.0) / 100
    tasa_financ   = inp.get("tasa_financ", 9.0)
    tasa_ir       = inp.get("tasa_ir", 29.5) / 100

    ing_dptos    = area_vendible * precio_m2
    ing_estac    = n_estac * m.get("precio_estac", 0)
    ing_deposito = n_depositos * m.get("precio_deposito", 0)
    ing_brutos   = ing_dptos + ing_estac + ing_deposito

    c_obra_dptos   = area_techada * costo_const
    c_obra_sotanos = n_estac * 25 * costo_sotano   # 25 m² área techada por cochera (validado: Clemente X 335)
    c_construccion = (c_obra_dptos + c_obra_sotanos) * (1 + fee_constr)
    c_arq          = area_techada * inp.get("costo_arq_m2", 5.94)
    c_esp          = area_techada * inp.get("costo_esp_m2", 7.92)
    c_factib       = inp.get("costo_factibilidades", 17000)
    c_supervision  = c_construccion * 0.005
    c_permisos         = c_construccion * 0.015
    c_gerencia         = c_construccion * 0.05
    c_ventas_marketing = ing_brutos * 0.05
    c_due_dilig        = 10000
    _meses_obra    = 24 if num_pisos > 20 else (12 if num_pisos <= 5 else 16)
    c_financiero   = c_construccion * 0.75 * tasa_financ / 100 * (_meses_obra / 12)

    c_base_constr  = c_construccion + c_arq + c_esp + c_factib
    c_legales_base = (c_due_dilig + c_base_constr) * 0.005
    # factor de transacción: alcabala 3% + notarial 0.3% + registral 0.15%
    factor_trans   = 1 + 0.03 + 0.003 + 0.0015  # 1.0345
    # c_legales también sube 0.5% del terreno: factor_trans * 0.005 extra por cada $ de terreno
    factor_terreno = factor_trans * 1.005

    C_fixed = (c_base_constr + c_supervision + c_legales_base + c_permisos
               + c_gerencia + c_ventas_marketing
               + c_due_dilig + c_factib + c_financiero)

    k = max(1 - tasa_ir, 0.01)

    resultados = {}
    for mg in [0.10, 0.12, 0.15, 0.18, 0.20]:
        target_c_total = ing_brutos * (1 - mg / k)
        T = (target_c_total - C_fixed) / factor_terreno
        resultados[mg] = max(0, round(T))

    tc = float((st.secrets.get("mercado") or {}).get("tipo_cambio", 3.45))

    return {
        "ing_brutos":    round(ing_brutos),
        "C_fixed":       round(C_fixed),
        "c_construccion": round(c_construccion),
        "c_financiero":  round(c_financiero),
        "area_terreno":  area_terreno,
        "area_vendible": area_vendible,
        "area_techada":  area_techada,
        "precio_m2":     precio_m2,
        "zona":          zona,
        "resultados":    resultados,
        "tipo_cambio":   tc,
    }


# ═══════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════

st.set_page_config(
    page_title="FACTIS — Osterling Advisory",
    page_icon="🏛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ═══════════════════════════════════════════════════════
# LINK DE COMPARTIR — vista pública (sin login)
# ═══════════════════════════════════════════════════════

def _show_shared_view(token: str) -> None:
    logo_path = pathlib.Path(__file__).parent / "logo.png"
    logo_b64  = base64.b64encode(logo_path.read_bytes()).decode() if logo_path.exists() else ""
    st.markdown("""
    <style>
    html, body, .stApp { background: #F7F5F0 !important; }
    section[data-testid="stSidebar"], header[data-testid="stHeader"] { display:none !important; }
    .block-container { max-width:820px !important; margin:0 auto !important; padding-top:4vh !important; }
    </style>
    """, unsafe_allow_html=True)

    sb = _get_supabase()
    proyecto = None
    if sb:
        try:
            r = sb.table("proyectos").select("*").eq("share_token", token).single().execute()
            proyecto = r.data
        except Exception:
            pass
        if not proyecto:
            try:
                # Fallback: token guardado dentro del JSONB datos
                r2 = sb.table("proyectos").select("*").execute()
                for row in (r2.data or []):
                    if (row.get("datos") or {}).get("_share_token") == token:
                        proyecto = row
                        break
            except Exception:
                pass

    if not proyecto:
        st.error("Este enlace no es válido o ya no está disponible.")
        return

    datos  = proyecto.get("datos") or {}
    tipo   = proyecto.get("tipo", "inmobiliario")
    nombre = proyecto.get("nombre_proyecto", "Análisis")
    zona   = proyecto.get("zona", "")
    fecha  = (proyecto.get("creado_en") or "")[:10]

    logo_html = (f'<img src="data:image/png;base64,{logo_b64}" style="height:30px;display:block;">'
                 if logo_b64 else "FACTIS")
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:16px;margin-bottom:28px;
                padding-bottom:18px;border-bottom:2px solid #D4C9B4;">
        <div style="background:#1E2D3D;border-radius:8px;padding:8px 14px;">{logo_html}</div>
        <div>
            <div style="font-size:20px;font-weight:800;color:#1E2D3D;letter-spacing:-0.5px;">{nombre}</div>
            <div style="font-size:12px;color:#8A8070;margin-top:2px;">
                {tipo.capitalize()} · {zona} · {fecha}
            </div>
        </div>
        <div style="margin-left:auto;font-size:10px;color:#B0A090;background:#EDE9E2;
                    border-radius:20px;padding:4px 12px;font-weight:600;letter-spacing:1px;">
            SOLO LECTURA
        </div>
    </div>
    """, unsafe_allow_html=True)

    def _fmtv(v):
        try:
            return f"${float(v):,.0f}".replace(",","X").replace(".",",").replace("X",".")
        except Exception:
            return "—"
    def _fmtp(v):
        try:
            return f"{float(v):.1f}%"
        except Exception:
            return "—"

    if tipo == "inmobiliario":
        financ  = datos.get("financ") or {}
        resumen = financ.get("resumen") or {}
        cabida  = datos.get("cabida") or {}
        if resumen:
            st.markdown("#### Resumen Financiero")
            c1, c2, c3 = st.columns(3)
            c1.metric("Ingresos Brutos",   _fmtv(resumen.get("ingresos_brutos", 0)))
            c1.metric("Costo Total",        _fmtv(resumen.get("costo_total", 0)))
            c2.metric("Utilidad Neta",      _fmtv(resumen.get("utilidad_neta", 0)))
            c2.metric("Margen Neto",        _fmtp(resumen.get("margen_pct", 0)))
            c3.metric("TIR Anual",          _fmtp(resumen.get("tir_anual_pct", 0)))
            c3.metric("ROI",                _fmtp(resumen.get("roi_pct", 0)))
        if cabida:
            st.markdown("#### Cabida Arquitectónica")
            cc1, cc2, cc3 = st.columns(3)
            cc1.metric("Pisos",           cabida.get("num_pisos", "—"))
            cc2.metric("Unidades totales",cabida.get("total_unidades", "—"))
            cc3.metric("Área vendible",   f"{cabida.get('area_vendible_m2', 0):,.0f} m²")

    elif tipo == "industrial":
        r = datos.get("industrial_result") or {}
        if r:
            st.markdown("#### Análisis Industrial")
            c1, c2, c3 = st.columns(3)
            c1.metric("Costo Total",   _fmtv(r.get("costo_total", 0)))
            c1.metric("Área Nave",     f"{r.get('area_nave', 0):,.0f} m²")
            c2.metric("Yield Bruto",   _fmtp(r.get("yield_bruto", 0)))
            c2.metric("Yield Neto",    _fmtp(r.get("yield_neto", 0)))
            if r.get("irr_anual") is not None:
                c3.metric("TIR 10 años", _fmtp(r["irr_anual"]))
            if r.get("payback_anos") is not None:
                c3.metric("Payback",     f"{r['payback_anos']:.1f} años")

    elif tipo == "residencial":
        r = datos.get("residencial_result") or {}
        if r:
            st.markdown("#### Valorización del Inmueble")
            c1, c2 = st.columns(2)
            c1.metric("Valor estimado",  _fmtv(r.get("valor_mercado", 0)))
            c1.metric("Precio m²",       f"${r.get('precio_m2_mercado', 0):,.0f}/m²")
            c2.metric("Yield alquiler",  _fmtp(r.get("yield_alquiler_bruto", 0)))

    st.markdown("---")
    st.markdown(
        '<div style="text-align:center;font-size:11px;color:#8A8070;padding:10px 0;">'
        'Generado por <b>FACTIS</b> · Osterling Advisory · Este enlace es de solo lectura</div>',
        unsafe_allow_html=True)

# Verificar link compartido ANTES del login
_qt = st.query_params.get("share", "")
if _qt:
    _show_shared_view(str(_qt))
    st.stop()

# ═══════════════════════════════════════════════════════
# AUTENTICACIÓN
# ═══════════════════════════════════════════════════════

def _hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def _get_users() -> dict:
    try:
        return dict(st.secrets.get("users", {}))
    except Exception:
        return {}

def _show_login() -> None:
    import datetime as _dt_login, base64 as _b64l, pathlib as _pll
    _yr  = _dt_login.date.today().year
    _app = _pll.Path(__file__).parent

    def _enc(name):
        p = _app / name
        return _b64l.b64encode(p.read_bytes()).decode() if p.exists() else ""

    _bg  = _enc("bg_login.jpg")
    _lgw = _enc("logo_white.png")

    _bg_url  = f"url('data:image/jpeg;base64,{_bg}')"  if _bg  else "none"
    _lgw_url = f"url('data:image/png;base64,{_lgw}')"  if _lgw else "none"

    st.markdown(f"""
    <style>
    html, body {{
        background-image:{_bg_url} !important;
        background-size:100% auto !important;
        background-position:bottom center !important;
        background-repeat:no-repeat !important;
        background-color:#07111D !important;
        min-height:100vh;
    }}
    .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"],
    [data-testid="stMainBlockContainer"], .main, .appview-container {{
        background:transparent !important;
    }}
    section[data-testid="stSidebar"], header[data-testid="stHeader"] {{ display:none !important; }}
    .block-container {{
        max-width:100% !important; padding:0 !important; margin:0 !important;
        background:transparent !important;
    }}
    .block-container > div {{ background:transparent !important; border:none !important; padding:0 !important; }}
    [data-testid="stHorizontalBlock"] {{ gap:0 !important; align-items:stretch !important; }}
    [data-testid="column"]:first-child {{
        background:rgba(7,17,29,0.68) !important;
        min-height:100vh !important;
    }}
    [data-testid="column"]:first-child > div,
    [data-testid="column"]:first-child [data-testid="stVerticalBlockBorderWrapper"],
    [data-testid="column"]:first-child [data-testid="stVerticalBlock"] {{ min-height:100vh !important; padding:0 !important; }}
    [data-testid="column"]:last-child {{
        background:rgba(8,18,30,0.94) !important;
        border-left:1px solid rgba(184,144,74,0.20) !important;
        min-height:100vh !important;
        backdrop-filter:blur(14px) !important;
        -webkit-backdrop-filter:blur(14px) !important;
    }}
    [data-testid="column"]:last-child > div,
    [data-testid="column"]:last-child [data-testid="stVerticalBlockBorderWrapper"] {{ min-height:100vh !important; }}
    [data-testid="column"]:last-child [data-testid="stVerticalBlock"] {{
        padding:36vh 24% 6vh 24% !important;
        box-sizing:border-box !important; width:100% !important;
    }}
    .solum-hero {{
        width:90px; height:118px;
        background-image:{_lgw_url};
        background-size:contain; background-repeat:no-repeat; background-position:left center;
        margin:0 0 28px 0;
    }}
    .solum-sm {{
        width:52px; height:68px;
        background-image:{_lgw_url};
        background-size:contain; background-repeat:no-repeat; background-position:center;
        margin:0 auto 16px;
    }}
    [data-testid="column"]:last-child .stTextInput > label,
    [data-testid="column"]:last-child .stTextInput [data-testid="stWidgetLabel"] p {{
        color:rgba(184,200,216,0.48) !important; font-size:9px !important;
        font-weight:700 !important; letter-spacing:2px !important; text-transform:uppercase !important;
    }}
    [data-testid="column"]:last-child .stTextInput > div {{
        background:rgba(255,255,255,0.05) !important;
        border:1px solid rgba(255,255,255,0.12) !important; border-radius:8px !important;
    }}
    [data-testid="column"]:last-child .stTextInput > div:focus-within {{
        border-color:rgba(184,144,74,0.55) !important;
        box-shadow:0 0 0 3px rgba(184,144,74,0.08) !important;
    }}
    [data-testid="column"]:last-child .stTextInput input {{
        background:transparent !important; border:none !important;
        color:#E8EDF4 !important; font-size:14px !important; padding:12px 14px !important;
        width:100% !important; box-sizing:border-box !important;
    }}
    [data-testid="column"]:last-child .stTextInput input::placeholder {{ color:rgba(184,200,216,0.20) !important; }}
    [data-testid="column"]:last-child .stButton > button {{
        width:100% !important;
        background:linear-gradient(135deg,#B8904A 0%,#C9A055 100%) !important;
        color:#FFFFFF !important; border:none !important; border-radius:8px !important;
        font-weight:700 !important; font-size:11px !important; letter-spacing:3px !important;
        padding:15px !important; margin-top:10px !important; text-transform:uppercase !important;
        box-shadow:0 4px 24px rgba(184,144,74,0.30) !important;
    }}
    [data-testid="column"]:last-child .stButton > button:hover {{
        background:linear-gradient(135deg,#C9A055 0%,#D4A853 100%) !important;
        box-shadow:0 8px 32px rgba(184,144,74,0.50) !important;
        transform:translateY(-1px) !important;
    }}
    .stAlert {{ border-radius:8px !important; }}
    </style>
    """, unsafe_allow_html=True)

    col_left, col_right = st.columns([60, 40], gap="small")

    with col_left:
        st.markdown(f"""
        <div style="min-height:100vh;padding:52px 64px;display:flex;flex-direction:column;box-sizing:border-box;">
          <div>
            <div style="font-size:9px;font-weight:700;color:rgba(184,144,74,0.52);letter-spacing:4px;text-transform:uppercase;">SOLUM · Osterling Advisory</div>
          </div>
          <div style="flex:1;display:flex;flex-direction:column;justify-content:center;padding:36px 0;">
            <div class="solum-hero"></div>
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:18px;">
              <div style="width:28px;height:1px;background:rgba(184,144,74,0.46);flex-shrink:0;"></div>
              <div style="font-size:9px;color:rgba(184,144,74,0.76);letter-spacing:4px;text-transform:uppercase;font-weight:700;">Plataforma Analítica Inmobiliaria</div>
            </div>
            <div style="font-size:50px;font-weight:900;color:#FFFFFF;line-height:1.06;letter-spacing:-2px;margin-bottom:16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
              Análisis inmobiliario<br><span style="color:#B8904A;">para Lima.</span>
            </div>
            <div style="font-size:15px;color:rgba(184,200,216,0.52);line-height:1.72;margin-bottom:36px;max-width:400px;">
              IA para análisis inmobiliario integral — del certificado de parámetros al reporte para el banco.
            </div>
            <div style="border-top:1px solid rgba(184,144,74,0.18);padding-top:22px;max-width:420px;">
              <div style="font-size:10px;color:rgba(184,144,74,0.60);letter-spacing:3px;text-transform:uppercase;font-weight:600;line-height:2;">
                Cabida &nbsp;·&nbsp; Financiero &nbsp;·&nbsp; Legal &nbsp;·&nbsp; Asistente IA
              </div>
            </div>
          </div>
          <div style="padding-top:18px;border-top:1px solid rgba(255,255,255,0.08);">
            <div style="font-size:9px;color:rgba(184,200,216,0.18);letter-spacing:2px;text-transform:uppercase;">Osterling Advisory · Lima, Perú · {_yr}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

    with col_right:
        st.markdown("""
        <div style="margin-bottom:32px;">
          <div style="font-size:11px;color:rgba(184,144,74,0.80);letter-spacing:3px;text-transform:uppercase;font-weight:700;margin-bottom:6px;">Acceso a la Plataforma</div>
          <div style="font-size:13px;color:rgba(184,200,216,0.36);">Ingresa tus credenciales para continuar</div>
        </div>
        <div style="height:1px;background:linear-gradient(90deg,rgba(184,144,74,0.28),transparent);margin-bottom:28px;"></div>
        """, unsafe_allow_html=True)
        username = st.text_input("Usuario", placeholder="nombre de usuario", key="_login_user")
        password = st.text_input("Contraseña", type="password", placeholder="••••••••", key="_login_pw")
        if st.button("Ingresar →", key="_login_btn"):
            users = _get_users()
            user_cfg = users.get(username.strip().lower())
            if user_cfg and user_cfg.get("password") == _hash_pw(password):
                st.session_state["_authenticated"] = True
                st.session_state["_user_name"]     = user_cfg.get("name", username)
                st.session_state["_user_role"]     = user_cfg.get("role", "advisor")
                st.session_state["_username"]      = username.strip().lower()
                for _k in ("_login_pw", "_login_user", "_login_btn"):
                    st.session_state.pop(_k, None)
                st.session_state["_auth_loading"] = True
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")
        st.markdown(
            '<div style="text-align:center;margin-top:24px;">' +
            '<div style="font-size:10px;color:rgba(184,200,216,0.16);">🔒 Acceso restringido · Osterling Advisory</div>' +
            '</div>',
            unsafe_allow_html=True)


if not st.session_state.get("_authenticated"):
    _show_login()
    st.stop()

# ── Overlay de transición limpia ────────────────────────
# Se activa en el primer render post-login, cubre la pantalla
# mientras la app carga por debajo y se desvanece solo.
if st.session_state.pop("_auth_loading", False):
    st.markdown("""
    <style>
    #auth-overlay {
        position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
        background: linear-gradient(160deg,#0A1628 0%,#131F2E 55%,#0F1C2A 100%);
        display: flex; flex-direction: column; align-items: center; justify-content: center;
        z-index: 999999;
        animation: overlayFadeOut 0.45s ease 2.2s forwards;
        pointer-events: none;
    }
    @keyframes overlayFadeOut {
        from { opacity: 1; }
        to   { opacity: 0; visibility: hidden; }
    }
    .oa-dot {
        width: 8px; height: 8px; background: #B8904A; border-radius: 50%;
        display: inline-block; margin: 0 4px;
        animation: oaDot 1.1s ease-in-out infinite;
    }
    @keyframes oaDot {
        0%,100% { opacity:0.25; transform:scale(0.7); }
        50%      { opacity:1;    transform:scale(1.25); }
    }
    </style>
    <div id="auth-overlay">
        <div style="font-size:9px;color:#B8904A;letter-spacing:5px;text-transform:uppercase;
                    font-weight:600;margin-bottom:14px;">Osterling Advisory</div>
        <div style="font-size:30px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;">FACTIS</div>
        <div style="width:40px;height:2px;background:#B8904A;margin:16px auto 22px;"></div>
        <div style="font-size:11px;color:rgba(184,200,216,0.50);letter-spacing:1px;margin-bottom:22px;">
            Cargando plataforma…
        </div>
        <div>
            <span class="oa-dot" style="animation-delay:0s"></span>
            <span class="oa-dot" style="animation-delay:0.35s"></span>
            <span class="oa-dot" style="animation-delay:0.70s"></span>
        </div>
    </div>
    <script>
    // Remove overlay from DOM after animation completes
    (function() {
        var el = document.getElementById('auth-overlay');
        if (el) el.addEventListener('animationend', function() { el.remove(); });
    })();
    </script>
    """, unsafe_allow_html=True)

def _step_header(num: str, title: str) -> None:
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:10px;margin:18px 0 6px;">'
        f'<div style="background:#1E2D3D;color:#F5F2ED;font-size:10px;font-weight:700;'
        f'min-width:22px;height:22px;border-radius:50%;display:flex;align-items:center;'
        f'justify-content:center;">{num}</div>'
        f'<div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;'
        f'font-weight:700;color:#1E2D3D;">{title}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ── Base ── */
    html, body, [class*="css"], .stApp, .stMarkdown, .stMetric,
    section[data-testid="stSidebar"], .stTabs, button, input, select, textarea {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    }

    html { background-color: #F0EDE8 !important; }
    .stApp { background-color: #F0EDE8; }

    /* ── Expanders ── */
    [data-testid="stExpander"] summary,
    [data-testid="stExpander"] summary p,
    [data-testid="stExpander"] summary span,
    .streamlit-expanderHeader { color: #1E2D3D !important; }

    /* ── Loading spinner ── */
    [data-testid="stStatusWidget"] { color: #B8904A !important; }
    [data-testid="stStatusWidget"] svg { color: #B8904A !important; fill: #B8904A !important; }
    [data-testid="stStatusWidget"] label { color: #B8904A !important; }
    .stSpinner > div { border-top-color: #B8904A !important; }
    div[data-testid="stSpinner"] > div > div {
        border-color: #B8904A transparent #B8904A transparent !important;
    }

    /* ═══════════════════════════════════════
       SIDEBAR — dark navy, premium SaaS look
       ═══════════════════════════════════════ */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #131F2E 0%, #1A2D41 100%) !important;
        border-right: none !important;
        box-shadow: 4px 0 24px rgba(0,0,0,0.22) !important;
    }

    /* All sidebar text: light by default */
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stFileUploader label,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] small,
    section[data-testid="stSidebar"] div,
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stNumberInput label,
    section[data-testid="stSidebar"] .stTextInput label,
    section[data-testid="stSidebar"] .stTextArea label,
    section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
    section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p {
        color: #B8C8D8 !important;
    }

    /* Sidebar divider */
    section[data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.08) !important;
    }

    /* Sidebar markdown headers (h3) */
    section[data-testid="stSidebar"] h3 {
        color: #B8904A !important; font-size: 9px !important;
        letter-spacing: 2.5px !important; text-transform: uppercase !important;
        font-weight: 700 !important; padding-left: 10px !important;
        border-left: 2px solid #B8904A !important; line-height: 1.6 !important;
        margin-bottom: 8px !important;
    }

    /* Sidebar inputs — dark glass, cohesive with dark sidebar */
    /* Outer container: borde sutil, fondo oscuro, con padding para que el input no toque el borde */
    section[data-testid="stSidebar"] .stNumberInput > div,
    section[data-testid="stSidebar"] .stTextInput > div {
        background-color: rgba(255,255,255,0.08) !important;
        border: 1px solid rgba(255,255,255,0.14) !important;
        border-radius: 8px !important;
        overflow: hidden !important;
        padding: 0 !important;
        gap: 0 !important;
    }
    /* Input value area: sin separador interno — el borde del contenedor rodea todo */
    section[data-testid="stSidebar"] .stNumberInput input,
    section[data-testid="stSidebar"] .stTextInput input {
        background-color: transparent !important;
        color: #E8EDF2 !important;
        border: none !important;
        border-radius: 0 !important;
        margin: 0 !important;
        padding-left: 10px !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        box-sizing: border-box !important;
    }
    section[data-testid="stSidebar"] textarea,
    section[data-testid="stSidebar"] .stTextArea textarea {
        background-color: rgba(255,255,255,0.08) !important;
        color: #E8EDF2 !important;
        border: 1px solid rgba(255,255,255,0.14) !important;
        border-radius: 8px !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        padding: 8px 12px !important;
    }
    section[data-testid="stSidebar"] input::placeholder,
    section[data-testid="stSidebar"] textarea::placeholder {
        color: rgba(255,255,255,0.28) !important;
    }

    /* Number input +/− step buttons */
    section[data-testid="stSidebar"] .stNumberInput button,
    section[data-testid="stSidebar"] [data-testid="stNumberInput"] button {
        background-color: rgba(255,255,255,0.10) !important;
        color: #C8D8E8 !important;
        border: none !important;
        border-radius: 0 !important;
        font-weight: 700 !important;
        font-size: 15px !important;
        min-width: 30px !important;
        transition: background 0.15s !important;
        margin: 0 !important;
        align-self: stretch !important;
    }
    section[data-testid="stSidebar"] .stNumberInput button:hover,
    section[data-testid="stSidebar"] [data-testid="stNumberInput"] button:hover {
        background-color: rgba(184,144,74,0.28) !important;
        color: #E8C87A !important;
    }
    section[data-testid="stSidebar"] .stNumberInput button p,
    section[data-testid="stSidebar"] [data-testid="stNumberInput"] button p,
    section[data-testid="stSidebar"] .stNumberInput button span,
    section[data-testid="stSidebar"] [data-testid="stNumberInput"] button span {
        color: inherit !important;
    }

    /* Tooltip icons — todos los módulos (main content: gold, sidebar: blanco) */
    [data-testid="stTooltipIcon"],
    [data-testid="stTooltipIcon"] button,
    button[aria-label*="Learn more"],
    button[data-testid*="tooltip"],
    .stTooltipIcon {
        background: transparent !important;
        background-color: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }
    [data-testid="stTooltipIcon"] svg,
    [data-testid="stTooltipIcon"] button svg {
        color: #B8904A !important;
        fill: #B8904A !important;
    }
    section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] svg,
    section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button svg {
        color: rgba(255,255,255,0.40) !important;
        fill: rgba(255,255,255,0.40) !important;
    }

    /* Sidebar selectbox — dark glass */
    section[data-testid="stSidebar"] [data-baseweb="select"] > div,
    section[data-testid="stSidebar"] [data-baseweb="select"] div[class*="ValueContainer"],
    section[data-testid="stSidebar"] [data-baseweb="select"] div[class*="control"] {
        background-color: rgba(255,255,255,0.10) !important;
        border-color: rgba(255,255,255,0.16) !important;
        color: #E8EDF2 !important;
        border-radius: 6px !important;
    }
    section[data-testid="stSidebar"] [data-baseweb="select"] span,
    section[data-testid="stSidebar"] [data-baseweb="select"] div[class*="singleValue"] {
        color: #E8EDF2 !important;
        font-weight: 500 !important;
    }
    section[data-testid="stSidebar"] [data-baseweb="select"] svg {
        fill: #8AA8C8 !important;
    }

    /* Module selector radio — pill style on dark */
    section[data-testid="stSidebar"] [data-testid="stRadio"] > div {
        gap: 5px !important;
    }
    section[data-testid="stSidebar"] [data-testid="stRadio"] label {
        border: 1px solid rgba(255,255,255,0.12) !important;
        border-radius: 7px !important;
        padding: 9px 14px 9px 10px !important;
        background: rgba(255,255,255,0.05) !important;
        transition: border-color 0.15s, background 0.15s !important;
        width: 100% !important;
    }
    section[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
        border-color: rgba(184,144,74,0.55) !important;
        background: rgba(184,144,74,0.10) !important;
    }
    section[data-testid="stSidebar"] [data-testid="stRadio"] label p,
    section[data-testid="stSidebar"] [data-testid="stRadio"] label span {
        font-size: 12px !important;
        font-weight: 500 !important;
        color: #A8B8C8 !important;
        letter-spacing: 0.2px !important;
    }
    section[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) {
        border-color: #B8904A !important;
        background: rgba(184,144,74,0.15) !important;
    }
    section[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) p,
    section[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) span {
        color: #E8C87A !important;
        font-weight: 700 !important;
    }
    /* Radio circle: force white border + transparent fill on dark bg */
    section[data-testid="stSidebar"] [data-testid="stRadio"] [data-baseweb="radio"] div[role="radio"],
    section[data-testid="stSidebar"] [data-testid="stRadio"] div[data-baseweb="radio"] > div {
        border-color: rgba(255,255,255,0.35) !important;
        background-color: transparent !important;
    }
    section[data-testid="stSidebar"] [data-testid="stRadio"] [aria-checked="true"] div[role="radio"],
    section[data-testid="stSidebar"] [data-testid="stRadio"] [aria-checked="true"] > div > div {
        border-color: #B8904A !important;
        background-color: #B8904A !important;
    }
    /* Hide native radio circle SVG that renders black */
    section[data-testid="stSidebar"] [data-testid="stRadio"] [data-baseweb="radio"] svg {
        display: none !important;
    }

    /* Sidebar primary buttons — gold */
    section[data-testid="stSidebar"] .stButton > button[kind="primary"],
    section[data-testid="stSidebar"] .stButton > button[kind="primary"] span,
    section[data-testid="stSidebar"] .stButton > button[kind="primary"] p {
        background: linear-gradient(135deg, #C8A050 0%, #A87830 100%) !important;
        color: #FFFFFF !important;
        border: none !important;
        box-shadow: 0 4px 12px rgba(184,144,74,0.35) !important;
        letter-spacing: 1.5px !important;
        font-size: 11px !important;
        font-weight: 700 !important;
        border-radius: 6px !important;
        transition: transform 0.15s ease, box-shadow 0.2s ease !important;
    }
    section[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover,
    section[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover span {
        background: linear-gradient(135deg, #D4AC5C 0%, #B88838 100%) !important;
        color: #FFFFFF !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(184,144,74,0.45) !important;
    }

    /* Sidebar secondary buttons */
    section[data-testid="stSidebar"] .stButton > button:not([kind="primary"]),
    section[data-testid="stSidebar"] .stButton > button[kind="secondary"] {
        background-color: rgba(255,255,255,0.08) !important;
        color: #B8C8D8 !important;
        border: 1px solid rgba(255,255,255,0.15) !important;
        border-radius: 6px !important;
    }
    section[data-testid="stSidebar"] .stButton > button:not([kind="primary"]) span,
    section[data-testid="stSidebar"] .stButton > button[kind="secondary"] span {
        color: #B8C8D8 !important;
    }
    section[data-testid="stSidebar"] .stButton > button:not([kind="primary"]):hover {
        border-color: rgba(184,144,74,0.5) !important;
        background-color: rgba(184,144,74,0.1) !important;
    }
    section[data-testid="stSidebar"] .stButton > button:not([kind="primary"]):hover span {
        color: #E8C87A !important;
    }

    /* Sidebar form submit */
    section[data-testid="stSidebar"] .stFormSubmitButton > button,
    section[data-testid="stSidebar"] [data-testid="stFormSubmitButton"] > button {
        background: linear-gradient(135deg, #C8A050 0%, #A87830 100%) !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 6px !important;
        font-size: 11px !important;
        font-weight: 700 !important;
        letter-spacing: 1.5px !important;
    }
    section[data-testid="stSidebar"] .stFormSubmitButton > button span,
    section[data-testid="stSidebar"] [data-testid="stFormSubmitButton"] > button span { color: #FFFFFF !important; }

    /* Sidebar file uploader */
    /* File uploader — sin borde punteado, solo botón limpio */
    [data-testid="stFileUploaderDropzone"],
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
        border: none !important;
        background: transparent !important;
        padding: 0 !important;
        min-height: unset !important;
        box-shadow: none !important;
    }
    /* Ocultar instrucciones de drag-drop */
    [data-testid="stFileUploaderDropzoneInstructions"],
    [data-testid="stFileUploaderDropzone"] [data-testid="stFileUploaderDropzoneInstructions"] {
        display: none !important;
    }
    /* Botón Upload — compacto, ancho automático */
    [data-testid="stFileUploaderDropzone"] button,
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
        padding: 5px 14px !important;
        font-size: 11px !important;
        font-weight: 600 !important;
        height: 30px !important;
        min-height: unset !important;
        border-radius: 6px !important;
        white-space: nowrap !important;
        width: auto !important;
        letter-spacing: 0.5px !important;
    }
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
        background: rgba(255,255,255,0.10) !important;
        color: #C8D8E8 !important;
        border: 1px solid rgba(255,255,255,0.20) !important;
    }
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button:hover {
        background: rgba(184,144,74,0.20) !important;
        color: #E8C87A !important;
        border-color: rgba(184,144,74,0.45) !important;
    }
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button span {
        color: inherit !important;
    }
    /* Archivo subido — pill dorado compacto */
    [data-testid="stFileUploaderFile"],
    section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] {
        padding: 4px 10px !important;
        border-radius: 5px !important;
        margin-top: 4px !important;
    }
    section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] {
        background-color: rgba(184,144,74,0.14) !important;
        border: 1px solid rgba(184,144,74,0.35) !important;
    }
    section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] span,
    section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] p { color: #E8C87A !important; }

    /* Expanders área principal — texto navy */
    .main [data-testid="stExpander"] summary,
    .main [data-testid="stExpander"] summary p,
    .main [data-testid="stExpander"] summary span,
    .main [data-testid="stExpander"] details > summary,
    .main [data-testid="stExpander"] details > summary p,
    .main details > summary,
    .main details > summary p,
    .main details > summary span {
        color: #1E2D3D !important;
        font-weight: 600 !important;
    }
    .main [data-testid="stExpander"] details[open] > summary,
    .main [data-testid="stExpander"] details[open] > summary p,
    .main [data-testid="stExpander"] details[open] > summary span {
        color: #1E2D3D !important;
    }

    /* Sidebar expanders */
    [data-testid="stSidebar"] details,
    [data-testid="stSidebar"] details > summary,
    [data-testid="stSidebar"] [data-testid="stExpander"] {
        background-color: rgba(255,255,255,0.04) !important;
        border-color: rgba(255,255,255,0.10) !important;
        border-radius: 8px !important;
    }
    [data-testid="stSidebar"] details > summary {
        font-size: 12px !important; font-weight: 600 !important;
        letter-spacing: 0.3px !important; border-radius: 6px !important;
        padding: 9px 12px !important; color: #A8B8C8 !important;
    }
    [data-testid="stSidebar"] details > summary * { color: #A8B8C8 !important; }
    [data-testid="stSidebar"] details[open] > summary {
        border-bottom: 1px solid rgba(255,255,255,0.08) !important;
        color: #E8C87A !important;
    }
    [data-testid="stSidebar"] details[open] > summary * { color: #E8C87A !important; }

    /* Slider in sidebar */
    section[data-testid="stSidebar"] .stSlider label,
    section[data-testid="stSidebar"] .stSlider p,
    section[data-testid="stSidebar"] .stSlider span,
    section[data-testid="stSidebar"] [data-testid="stSlider"] p { color: #B8C8D8 !important; }

    /* Checkbox in sidebar */
    section[data-testid="stSidebar"] .stCheckbox label span { color: #B8C8D8 !important; }

    /* ── Header principal ── */
    .main-header {
        background: linear-gradient(135deg, #0F1C2A 0%, #1A2D41 50%, #0F1C2A 100%);
        padding: 24px 32px;
        border-radius: 12px;
        margin-bottom: 24px;
        box-shadow: 0 8px 32px rgba(10,20,35,0.28), 0 1px 0 rgba(184,144,74,0.2) inset;
        border-bottom: 2px solid rgba(184,144,74,0.25);
    }

    /* ── Cards de métricas ── */
    .metric-card {
        background: #FFFFFF;
        border: 1px solid #E4E0D8;
        border-top: 3px solid #B8904A;
        padding: 20px 20px 18px;
        border-radius: 10px;
        margin-bottom: 12px;
        box-shadow: 0 2px 12px rgba(30,45,61,0.07), 0 1px 3px rgba(30,45,61,0.05);
        transition: box-shadow 0.2s ease, transform 0.2s ease;
    }
    .metric-card:hover {
        box-shadow: 0 8px 24px rgba(30,45,61,0.12);
        transform: translateY(-2px);
    }
    .metric-card .label {
        font-size: 9px; color: #9A9080; letter-spacing: 2px;
        text-transform: uppercase; font-weight: 600;
    }
    .metric-card .value {
        font-size: 28px; color: #1E2D3D; font-weight: 800; margin-top: 8px;
        letter-spacing: -0.5px; font-variant-numeric: tabular-nums;
    }

    /* ── Sección titles ── */
    .section-title {
        color: #8A8078;
        font-size: 9px; font-weight: 700; letter-spacing: 3px;
        text-transform: uppercase;
        border-bottom: 1px solid #DDD9D0;
        padding-bottom: 8px; margin: 28px 0 16px 0;
        display: flex; align-items: center; gap: 8px;
    }
    .section-title::before {
        content: '';
        display: inline-block;
        width: 3px; height: 12px;
        background: #B8904A;
        border-radius: 2px;
        flex-shrink: 0;
    }

    /* ── KPI stat bar — rich metric display ── */
    .kpi-bar {
        display: grid; gap: 1px;
        background: #D8D4CC;
        border-radius: 10px; overflow: hidden;
        box-shadow: 0 2px 12px rgba(30,45,61,0.08);
        margin-bottom: 20px;
    }
    .kpi-cell {
        background: #FFFFFF; padding: 16px 18px;
        display: flex; flex-direction: column; gap: 4px;
    }
    .kpi-cell-label {
        font-size: 8px; color: #9A9080; letter-spacing: 2px;
        text-transform: uppercase; font-weight: 700;
    }
    .kpi-cell-value {
        font-size: 22px; color: #1E2D3D; font-weight: 800;
        letter-spacing: -0.5px; font-variant-numeric: tabular-nums;
        line-height: 1.1;
    }
    .kpi-cell-sub {
        font-size: 10px; color: #9A9080; font-weight: 500;
    }

    /* ── Alertas ── */
    .alert-gold {
        background: linear-gradient(135deg, #FFFBF3 0%, #FFF8EC 100%);
        border: 1px solid #DFC07A;
        border-left: 4px solid #B8904A; border-radius: 8px;
        padding: 14px 18px; color: #5C3D10; font-size: 13px;
        margin: 10px 0; line-height: 1.65;
        box-shadow: 0 2px 8px rgba(184,144,74,0.10);
    }
    .alert-legal {
        background: #F8F6F3; border: 1px solid #D0C8BC;
        border-left: 4px solid #8A9BAD; border-radius: 8px;
        padding: 14px 18px; color: #1E2D3D; font-size: 13px;
        margin: 10px 0; line-height: 1.65;
    }
    .alert-info {
        background: #E8F8F5; border: 1px solid #A8DDD6;
        border-left: 4px solid #1ABC9C; border-radius: 8px;
        padding: 14px 18px; color: #0D6B5E; font-size: 13px;
        margin: 10px 0; line-height: 1.65; font-weight: 500;
    }

    /* ── Tabs — underline style ── */
    .stTabs [data-baseweb="tab-list"] {
        background-color: transparent; gap: 0;
        border-bottom: 2px solid #D8D4CC;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: transparent; border: none;
        border-bottom: 3px solid transparent; margin-bottom: -2px;
        color: #8A8078; font-size: 10px; letter-spacing: 1.5px;
        text-transform: uppercase; padding: 12px 22px; font-weight: 600;
        transition: color 0.15s;
    }
    .stTabs [aria-selected="true"] {
        background-color: transparent !important;
        color: #B8904A !important;
        border-bottom: 3px solid #B8904A !important;
        font-weight: 800 !important;
    }

    /* ── Métricas nativas ── */
    div[data-testid="stMetricValue"],
    div[data-testid="stMetricValue"] > div {
        color: #1E2D3D !important; font-size: 24px !important;
        font-weight: 800 !important; letter-spacing: -0.5px !important;
        font-variant-numeric: tabular-nums !important;
    }
    div[data-testid="stMetricLabel"],
    div[data-testid="stMetricLabel"] > div,
    div[data-testid="stMetricLabel"] p,
    div[data-testid="stMetricLabel"] span,
    label[data-testid="stMetricLabel"] {
        color: #6A7888 !important; font-size: 9px !important;
        letter-spacing: 2px !important; text-transform: uppercase !important;
        font-weight: 700 !important;
    }
    div[data-testid="stMetricDelta"],
    div[data-testid="stMetricDelta"] > div {
        font-size: 11px !important;
        color: #7A8898 !important;
        background: transparent !important;
    }
    div[data-testid="stMetricDelta"] svg { display: none !important; }
    div[data-testid="stMetricDelta"] [data-testid="stMetricDeltaIcon"],
    div[data-testid="stMetricDelta"] span[class*="arrow"],
    div[data-testid="stMetricDelta"] span[class*="icon"] { display: none !important; }

    /* ── Metric container — add subtle card background ── */
    div[data-testid="metric-container"] {
        background: #FFFFFF !important;
        border: 1px solid #E8E4DC !important;
        border-radius: 10px !important;
        padding: 16px 20px !important;
        box-shadow: 0 1px 6px rgba(30,45,61,0.06) !important;
    }

    /* ── Botón primario ── */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #1E2D3D 0%, #243850 100%) !important;
        color: #FFFFFF !important;
        border: none !important; border-radius: 8px !important;
        letter-spacing: 1.5px; font-size: 11px; font-weight: 700;
        padding: 13px 22px; text-transform: uppercase;
        box-shadow: 0 4px 16px rgba(30,45,61,0.22);
        transition: background 0.2s ease, transform 0.15s ease, box-shadow 0.2s ease !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #B8904A 0%, #9A7030 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(184,144,74,0.40) !important;
    }
    .stButton > button[kind="primary"]:active {
        transform: translateY(0px) !important;
        box-shadow: 0 4px 16px rgba(30,45,61,0.22) !important;
    }
    /* Botón secundario */
    .stButton > button:not([kind="primary"]) {
        border: 1px solid #D8D4CC !important; color: #4A5568 !important;
        background: #FFFFFF !important; border-radius: 7px !important;
        font-size: 11px !important; font-weight: 600 !important;
        transition: border-color 0.15s, color 0.15s, background 0.15s !important;
    }
    .stButton > button:not([kind="primary"]):hover {
        border-color: #B8904A !important; color: #B8904A !important;
        background: #FDFAF6 !important;
    }

    /* ── Botón de descarga ── */
    [data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #1E2D3D 0%, #243850 100%) !important;
        color: #FFFFFF !important;
        border: none !important; border-radius: 8px !important;
        letter-spacing: 1px; font-size: 11px; font-weight: 700;
        padding: 13px 20px; width: 100%; opacity: 1 !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: linear-gradient(135deg, #B8904A 0%, #9A7030 100%) !important;
        opacity: 1 !important;
    }
    [data-testid="stDownloadButton"] > button:active,
    [data-testid="stDownloadButton"] > button:focus,
    [data-testid="stDownloadButton"] > button:disabled {
        background: linear-gradient(135deg, #1E2D3D 0%, #243850 100%) !important;
        color: #FFFFFF !important; opacity: 1 !important;
    }

    /* ── Spinner — rueda y texto en bronce ── */
    .stSpinner > div, .stSpinner p,
    [data-testid="stSpinner"] p,
    [data-testid="stSpinner"] > div { color: #B8904A !important; }
    [data-testid="stSpinner"] svg,
    [data-testid="stSpinner"] svg *,
    [data-testid="stSpinner"] svg circle,
    [data-testid="stSpinner"] svg path,
    [data-testid="stSpinner"] svg rect,
    [data-testid="stSpinner"] svg ellipse {
        stroke: #B8904A !important;
        color: #B8904A !important;
    }
    /* Streamlit spinner usa fill en algunos elementos */
    [data-testid="stSpinner"] svg [fill]:not([fill="none"]) { fill: #B8904A !important; }

    /* ── Alertas nativas ── */
    [data-testid="stWarning"], [data-testid="stWarning"] p,
    [data-testid="stWarning"] span { color: #5C4000 !important; }
    div[data-testid="stInfo"], div[data-testid="stInfo"] p,
    div[data-testid="stInfo"] span, div[data-testid="stInfo"] div {
        background-color: #EEF4FB !important; color: #1E2D3D !important;
        border-color: #8AA8C0 !important;
    }
    div[data-testid="stSuccess"], div[data-testid="stSuccess"] p,
    div[data-testid="stSuccess"] span, div[data-testid="stSuccess"] div {
        background-color: #EEF8F2 !important; color: #1A4731 !important;
        border-color: #6BAE90 !important;
    }
    div[data-testid="stError"], div[data-testid="stError"] p,
    div[data-testid="stError"] span { color: #7A1A1A !important; }
    div[data-testid="stInfo"] svg path { stroke: #1E2D3D !important; fill: none !important; }
    div[data-testid="stSuccess"] svg path { stroke: #1A4731 !important; fill: none !important; }

    /* ── Dataframe ── */
    [data-testid="stDataFrame"] div[class*="dvn-scroller"],
    [data-testid="stDataFrame"] div[class*="cell-wrapper"],
    [data-testid="stDataFrame"] div[class*="cell"],
    [data-testid="stDataFrame"] span { color: #1E2D3D !important; background-color: #FFFFFF !important; }
    [data-testid="stDataFrame"] div[class*="header"],
    [data-testid="stDataFrame"] div[class*="columnHeader"] {
        background-color: #F5F2ED !important; color: #1E2D3D !important; font-weight: 700 !important;
    }
    .stDataFrame {
        border: 1px solid #D8D4CC; border-radius: 8px; overflow: hidden;
        box-shadow: 0 2px 8px rgba(30,45,61,0.06);
    }
    .stDataFrame [data-testid="stDataFrameResizable"],
    .stDataFrame iframe { background-color: #FFFFFF !important; }
    .stAlert { border-radius: 8px; }

    /* ── Inputs (main area) ── */
    .stNumberInput input, .stTextInput input,
    .stSelectbox select, .stTextArea textarea {
        background-color: #FFFFFF !important; color: #1E2D3D !important;
        border: 1px solid #D8D4CC !important; border-radius: 6px !important;
        font-size: 13px !important;
    }
    .stNumberInput input::placeholder, .stTextInput input::placeholder,
    .stTextArea textarea::placeholder { color: #9A9080 !important; }
    .stNumberInput > div, .stTextInput > div {
        background-color: #FFFFFF !important;
        border: 1px solid #D8D4CC !important; border-radius: 6px !important;
    }
    .stTextArea label, .stTextInput label, .stNumberInput label,
    .stSelectbox label, .stMultiSelect label, .stCheckbox label,
    .stRadio label, [data-testid="stWidgetLabel"] { color: #1E2D3D !important; }
    .stSlider [data-baseweb="slider"] { background-color: transparent !important; }
    .stSlider label, .stSlider p, .stSlider span,
    [data-testid="stSlider"] label, [data-testid="stSlider"] p { color: #1E2D3D !important; }

    /* Radio buttons in main content — override Streamlit red with green */
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) [data-baseweb="radio"] div[role="radio"],
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) div[data-baseweb="radio"] > div {
        border-color: rgba(30,90,50,0.45) !important;
        background-color: transparent !important;
    }
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) [aria-checked="true"] div[role="radio"],
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) [aria-checked="true"] > div > div {
        border-color: #1E7A3C !important;
        background-color: #1E7A3C !important;
    }
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) label:has(input:checked) {
        background: rgba(30,122,60,0.08) !important;
    }
    /* Force dark text on all main-content radio labels — checked and unchecked */
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) label p,
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) label span,
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) label:has(input:checked) p,
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) label:has(input:checked) span {
        color: #1E2D3D !important;
    }
    /* Hide native black SVG dot inside radio circles */
    [data-testid="stRadio"]:not(section[data-testid="stSidebar"] *) [data-baseweb="radio"] svg {
        display: none !important;
    }

    /* ── Tablas markdown ── */
    .stMarkdown table { width:100%; border-collapse:collapse; background:#FFFFFF; border-radius:8px; overflow:hidden; }
    .stMarkdown thead th {
        background-color:#1E2D3D !important; color:#FFFFFF !important;
        padding:10px 14px; font-size:10px; letter-spacing:1.2px;
        text-transform:uppercase; border:none; font-weight:700;
    }
    .stMarkdown tbody td {
        background-color:#FFFFFF; color:#1E2D3D !important;
        padding:10px 14px; font-size:12px; border-bottom:1px solid #EEE;
    }
    .stMarkdown tbody tr:nth-child(even) td { background-color:#F9F7F4; }
    .stMarkdown p, .stMarkdown li, .stMarkdown strong { color:#1E2D3D !important; }

    /* Divs con fondo claro generados dinámicamente: forzar texto oscuro */
    .stMarkdown div[style*="background:#E8F5EE"],
    .stMarkdown div[style*="background:#FFF8EE"],
    .stMarkdown div[style*="background:#FFF8E6"],
    .stMarkdown div[style*="background:#FFFFFF"],
    .stMarkdown div[style*="background:#FDFAF6"],
    .stMarkdown div[style*="background:#F5F3EF"],
    .stMarkdown div[style*="background:#F7F5F1"],
    .stMarkdown div[style*="background:#FAFAF8"],
    .stMarkdown div[style*="background:#F0EDE8"],
    .stMarkdown div[style*="background:#FDECEA"],
    .stMarkdown div[style*="background:#FFF0F0"],
    .stMarkdown div[style*="background:#FFF0EE"] { color:#1A2233 !important; }

    /* ── Score card ── */
    .score-card {
        border-radius: 12px; padding: 32px 36px; text-align: center;
        border: 1px solid; margin-bottom: 20px;
        box-shadow: 0 6px 24px rgba(30,45,61,0.12);
    }

    /* ── Ocultar chrome de Streamlit ── */
    header[data-testid="stHeader"]  { display: none !important; }
    [data-testid="stDecoration"]    { display: none !important; }
    [data-testid="stToolbar"]       { display: none !important; }
    #MainMenu                       { display: none !important; }
    footer                          { display: none !important; }
    [data-testid="stStatusWidget"]  { display: none !important; }
    [data-testid="stSkeleton"],
    .stSpinner > div[data-testid],
    [class*="StatusWidget"]         { display: none !important; }
    .block-container { padding-top: 1.5rem !important; }

    /* ── DataTable th/td ── */
    .stDataFrame th {
        background-color: #1E2D3D !important; color: #FFFFFF !important;
        font-size: 10px !important; font-weight: 700 !important; letter-spacing: 0.8px !important;
    }
    .stDataFrame td { background-color: #FFFFFF !important; color: #1E2D3D !important; font-size: 12px !important; }

    /* ═══════════════════════════════════════════
       RESPONSIVE — tablet (≤900px) y móvil (≤600px)
       ═══════════════════════════════════════════ */

    /* ── Tablet: iPad, laptop pequeño ── */
    @media (max-width: 900px) {
        /* Hero banner: menos padding */
        .main-header {
            padding: 18px 20px !important;
            border-radius: 10px !important;
            margin-bottom: 16px !important;
        }
        .main-header h1, .main-header .factis-title {
            font-size: 20px !important;
        }
        /* KPI bar: 2 columnas en tablet */
        .kpi-bar {
            grid-template-columns: 1fr 1fr !important;
        }
        /* Metric card value: más pequeño */
        .metric-card .value { font-size: 22px !important; }
        /* Contenido principal: menos margen lateral */
        .block-container {
            padding-left: 1.2rem !important;
            padding-right: 1.2rem !important;
        }
        /* Tabs: scroll horizontal si no caben */
        .stTabs [data-baseweb="tab-list"] {
            overflow-x: auto !important;
            flex-wrap: nowrap !important;
            -webkit-overflow-scrolling: touch;
        }
        /* DataFrames: scroll horizontal */
        .stDataFrame { overflow-x: auto !important; }
        [data-testid="stDataFrame"] { overflow-x: auto !important; }
    }

    /* ── Móvil: smartphones (≤600px) ── */
    @media (max-width: 600px) {
        /* Hero banner: compacto */
        .main-header {
            padding: 14px 16px !important;
            min-height: unset !important;
            border-radius: 8px !important;
        }
        .main-header h1, .main-header .factis-title {
            font-size: 17px !important;
            letter-spacing: -0.3px !important;
        }
        /* KPI bar: columna única en móvil */
        .kpi-bar {
            grid-template-columns: 1fr !important;
        }
        .kpi-cell {
            padding: 12px 14px !important;
            border-right: none !important;
            border-bottom: 1px solid rgba(30,45,61,0.10) !important;
        }
        /* Metric card */
        .metric-card { padding: 14px 16px !important; }
        .metric-card .value { font-size: 20px !important; }
        /* Contenido: mínimo padding lateral */
        .block-container {
            padding-left: 0.75rem !important;
            padding-right: 0.75rem !important;
            padding-top: 0.75rem !important;
        }
        /* Columnas Streamlit: apilar en móvil */
        [data-testid="column"] {
            min-width: 100% !important;
            width: 100% !important;
        }
        /* Inputs: 16px mínimo para evitar zoom en iOS */
        input, select, textarea,
        .stNumberInput input,
        .stTextInput input,
        .stSelectbox select {
            font-size: 16px !important;
        }
        /* Botones: tap target mínimo 44px */
        .stButton button {
            min-height: 44px !important;
            font-size: 13px !important;
        }
        /* Tabs: más compactos */
        .stTabs [data-baseweb="tab"] {
            padding: 8px 12px !important;
            font-size: 11px !important;
        }
        .stTabs [data-baseweb="tab-list"] {
            overflow-x: auto !important;
            flex-wrap: nowrap !important;
            -webkit-overflow-scrolling: touch;
        }
        /* Expanders: menos padding */
        .streamlit-expanderHeader {
            padding: 10px 12px !important;
            font-size: 12px !important;
        }
        /* Section titles: más compactos */
        .section-title { font-size: 8px !important; letter-spacing: 2px !important; }
        /* Plotly charts: full width */
        .js-plotly-plot, .plotly { width: 100% !important; }
        /* Ocultar sidebar overlay click area en móvil no afecta contenido */
        .stApp > header { display: none !important; }
    }

    /* ── Touch: hover effects off en táctil (evita estados pegados) ── */
    @media (hover: none) {
        .metric-card:hover {
            transform: none !important;
            box-shadow: 0 2px 12px rgba(30,45,61,0.07), 0 1px 3px rgba(30,45,61,0.05) !important;
        }
        .stButton button:hover { opacity: 1 !important; }
    }

    /* ── Sidebar: colapsa por defecto en móvil ── */
    @media (max-width: 600px) {
        section[data-testid="stSidebar"] {
            transform: translateX(-100%) !important;
            transition: transform 0.25s ease !important;
        }
        section[data-testid="stSidebar"][aria-expanded="true"] {
            transform: translateX(0) !important;
        }
        /* Tablas HTML custom: scroll horizontal sin romper layout */
        .stMarkdown table,
        .stMarkdown div[style*="overflow"] {
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch !important;
            display: block !important;
            max-width: 100vw !important;
        }
        /* Matrices de sensibilidad y tablas de costos */
        div[style*="overflow-x:auto"] {
            max-width: calc(100vw - 24px) !important;
        }
        /* Métricas: 2 por fila en móvil */
        [data-testid="stMetric"] {
            min-width: 45% !important;
        }
        /* Download buttons: full width */
        [data-testid="stDownloadButton"] button {
            width: 100% !important;
            min-height: 44px !important;
        }
        /* Ocultar texto largo en tabs en móvil (solo icono si hay) */
        .stTabs [data-baseweb="tab"] span {
            font-size: 10px !important;
        }
        /* Plotly charts: no desbordar */
        .js-plotly-plot .plotly,
        .js-plotly-plot {
            max-width: 100% !important;
            overflow: hidden !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# Formato con comas eliminado — el overlay JS causaba color:transparent persistente
# tras re-renders de Streamlit, dejando los inputs invisibles al escribir.

# ═══════════════════════════════════════════════════════
# AUTENTICACIÓN (legacy block — eliminado, ver _show_login() arriba)
# ═══════════════════════════════════════════════════════

if False:
    st.markdown("""
    <style>
    section[data-testid="stSidebar"] { display: none !important; }
    .stApp { background: linear-gradient(135deg, #1A2737 0%, #1E2D3D 60%, #1A2737 100%) !important; }

    /* ── Contenedor del input ── */
    [data-testid="stTextInput"] { background: transparent !important; }
    [data-testid="stTextInput"] > div,
    [data-testid="stTextInput"] > div > div,
    [data-testid="stTextInput"] > div > div > div {
        background: rgba(255,255,255,0.08) !important;
        border: 1px solid rgba(184,144,74,0.4) !important;
        border-radius: 6px !important;
    }
    [data-testid="stTextInput"] > div > div:focus-within,
    [data-testid="stTextInput"] > div > div > div:focus-within {
        border-color: #B8904A !important;
        box-shadow: 0 0 0 2px rgba(184,144,74,0.2) !important;
    }

    /* ── Texto escrito y puntos de contraseña ── */
    [data-testid="stTextInput"] input,
    [data-testid="stTextInput"] input[type="password"],
    [data-testid="stTextInput"] input[type="text"] {
        background: transparent !important;
        color: #FFFFFF !important;
        -webkit-text-fill-color: #FFFFFF !important;
        caret-color: #B8904A !important;
        font-size: 15px !important;
    }
    [data-testid="stTextInput"] input::placeholder {
        color: rgba(255,255,255,0.38) !important;
        -webkit-text-fill-color: rgba(255,255,255,0.38) !important;
    }

    /* ── Botón ojo de Streamlit ── */
    [data-testid="stTextInput"] button,
    [data-testid="stTextInput"] button:focus,
    [data-testid="stTextInput"] button:hover {
        background: transparent !important;
        background-color: transparent !important;
        border: none !important;
        box-shadow: none !important;
        outline: none !important;
    }
    /* ── Div contenedor del botón ojo (quita el recuadro visible) ── */
    [data-testid="stTextInput"] div:has(> button),
    [data-testid="stTextInput"] > div > div > div > div,
    [data-testid="stTextInput"] > div > div > div > div:last-child {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        outline: none !important;
    }
    [data-testid="stTextInput"] button svg,
    [data-testid="stTextInput"] button svg path,
    [data-testid="stTextInput"] button svg circle,
    [data-testid="stTextInput"] button svg line,
    [data-testid="stTextInput"] button svg polyline,
    [data-testid="stTextInput"] button svg ellipse {
        fill: none !important;
        stroke: #B8904A !important;
        color: #B8904A !important;
    }
    [data-testid="stTextInput"] button:hover svg path,
    [data-testid="stTextInput"] button:hover svg circle,
    [data-testid="stTextInput"] button:hover svg line,
    [data-testid="stTextInput"] button:hover svg ellipse {
        stroke: #FFFFFF !important;
    }

    /* ── Suprimir iconos del gestor de contraseñas del browser ── */
    input[type="password"]::-webkit-credentials-auto-fill-button,
    input[type="password"]::-webkit-strong-password-auto-fill-button,
    input[type="password"]::-webkit-contacts-auto-fill-button,
    input::-webkit-credentials-auto-fill-button {
        display: none !important;
        visibility: hidden !important;
        pointer-events: none !important;
        width: 0 !important;
        height: 0 !important;
    }
    input[type="password"]::-ms-reveal,
    input[type="password"]::-ms-clear {
        display: none !important;
    }

    /* ── Label residual ── */
    [data-testid="stTextInput"] label {
        display: none !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    /* ── Suprimir "Press Enter to apply" ── */
    [data-testid="InputInstructions"],
    small[data-testid="InputInstructions"] {
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        line-height: 0 !important;
        overflow: hidden !important;
    }

    /* ── Form sin borde ni padding extra ── */
    [data-testid="stForm"] {
        border: none !important;
        padding: 0 !important;
        background: transparent !important;
    }

    /* ── Botón ACCEDER ── */
    [data-testid="stFormSubmitButton"] > button {
        background: #B8904A !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 6px !important;
        font-size: 12px !important;
        font-weight: 700 !important;
        letter-spacing: 3px !important;
        padding: 12px 0 !important;
        width: 100% !important;
        margin-top: 4px !important;
        transition: background 0.2s ease !important;
    }
    [data-testid="stFormSubmitButton"] > button:hover {
        background: #C9A05A !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Desactivar autocomplete del browser en el input de contraseña
    st.markdown("""
    <script>
    window.addEventListener('load', function() {
        const inputs = document.querySelectorAll('input[type="password"]');
        inputs.forEach(function(inp) {
            inp.setAttribute('autocomplete', 'new-password');
        });
    });
    </script>
    """, unsafe_allow_html=True)

    _l, _c, _r = st.columns([1, 1.1, 1])
    with _c:
        st.markdown("<div style='height:12vh'></div>", unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center;padding:44px 48px 36px;
                    background:rgba(255,255,255,0.04);
                    border:1px solid rgba(184,144,74,0.22);
                    border-radius:10px;">
            <div style="font-size:9px;color:#B8904A;letter-spacing:5px;text-transform:uppercase;
                        font-weight:600;margin-bottom:18px;">Osterling Advisory</div>
            <div style="font-size:30px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;">FACTIS</div>
            <div style="width:40px;height:2px;background:#B8904A;margin:16px auto 20px;"></div>
            <div style="font-size:12px;color:#8AA8C0;letter-spacing:0.3px;margin-bottom:8px;">
                Plataforma Analítica Inmobiliaria
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
        with st.form("login_form", border=False):
            _pwd = st.text_input("Contraseña", type="password", placeholder="Contraseña de acceso",
                                 label_visibility="collapsed", key="_login_pwd")
            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
            _submitted = st.form_submit_button("ACCEDER", use_container_width=True)
            if _submitted:
                if _pwd == _APP_PWD:
                    st.session_state._auth = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta.")
    st.stop()

# ═══════════════════════════════════════════════════════
# GEOMETRÍA DE LOTE
# ═══════════════════════════════════════════════════════

def _geo_poligono_tabular(frente: float, fondo: float, lado_izq: float, lado_der: float) -> "ShapelyPolygon | None":
    """
    Construye polígono shapely desde 4 medidas perimetrales.
    Modelo: trapezoide con frente en base y fondo en tope.
    Si los 4 lados son iguales → rectángulo exacto.
    """
    if not _SHAPELY_OK:
        return None
    if frente <= 0 or fondo <= 0 or lado_izq <= 0 or lado_der <= 0:
        return None
    # Altura del trapecio desde lado izquierdo
    offset = (frente - fondo) / 2.0
    h_izq = math.sqrt(max(0, lado_izq**2 - offset**2))
    h_der = math.sqrt(max(0, lado_der**2 - offset**2))
    h = (h_izq + h_der) / 2.0
    if h <= 0:
        return None
    # Vértices: base inferior izq → der, base superior der → izq
    coords = [
        (0.0,      0.0),
        (frente,   0.0),
        (frente - offset, h),
        (offset,   h),
    ]
    poly = ShapelyPolygon(coords)
    return make_valid(poly) if not poly.is_valid else poly


def _geo_poligono_dxf(fileobj) -> "ShapelyPolygon | None":
    """
    Extrae el polígono perimetral más grande de un archivo DXF/DWG.
    Busca LWPOLYLINE o POLYLINE cerradas; devuelve la de mayor área.
    """
    if not _SHAPELY_OK or not _EZDXF_OK:
        return None
    try:
        doc = ezdxf.read(fileobj)
        msp = doc.modelspace()
        best, best_area = None, 0.0
        for ent in msp:
            pts = None
            if ent.dxftype() == "LWPOLYLINE":
                pts = [(p[0], p[1]) for p in ent.get_points()]
            elif ent.dxftype() == "POLYLINE":
                try:
                    pts = [(v.dxf.location.x, v.dxf.location.y) for v in ent.vertices]
                except Exception:
                    continue
            if pts and len(pts) >= 3:
                try:
                    poly = ShapelyPolygon(pts)
                    if not poly.is_valid:
                        poly = make_valid(poly)
                    if poly.area > best_area:
                        best_area, best = poly.area, poly
                except Exception:
                    continue
        return best
    except Exception:
        return None


def _geo_aplicar_retiros(poly: "ShapelyPolygon", ret_frontal: float, ret_lat: float, ret_posterior: float) -> "ShapelyPolygon | None":
    """
    Aplica retiros direccionales al polígono del lote.
    Coordenadas de _geo_poligono_tabular: y=0 es el frente (calle), y=h es el posterior.
    Solo recorta las caras que tienen retiro > 0 — práctica Lima: ret_lat=0, ret_post=0.
    """
    if not _SHAPELY_OK or poly is None:
        return None
    from shapely.geometry import box as _box_geo
    x0, y0, x1, y1 = poly.bounds
    clip = _box_geo(
        x0 + ret_lat,
        y0 + ret_frontal,
        x1 - ret_lat,
        y1 - (ret_posterior if ret_posterior > 0 else 0),
    )
    huella = poly.intersection(clip)
    if huella.is_empty or not huella.is_valid:
        return None
    return huella


def _geo_validar(poly_lote, poly_huella, n_pisos: int, frente: float,
                 uso: str = "residencial", area_libre_min_pct: float = 0.0) -> dict:
    """
    Valida la geometría contra RNE y devuelve alertas + métricas.
    area_libre_min_pct: porcentaje de área libre mínima exigida por el CPU (ej. 40.0).
    """
    result = {"ok": True, "alertas": [], "metricas": {}}
    if poly_lote is None:
        result["ok"] = False
        result["alertas"].append("Sin geometría de lote definida.")
        return result

    area_lote   = poly_lote.area
    area_huella = poly_huella.area if poly_huella and not poly_huella.is_empty else 0.0
    cos_real    = round(area_huella / area_lote * 100, 1) if area_lote > 0 else 0.0

    # Restricción normativa de área libre: COS máximo = 100 − área_libre_min_pct
    cos_max_norma = 100.0 - area_libre_min_pct if area_libre_min_pct > 0 else 100.0
    area_huella_efectiva = min(area_huella, area_lote * cos_max_norma / 100)
    at_sobre = area_huella_efectiva * n_pisos

    # Frente mínimo
    frente_min = 13.0 if uso == "residencial" else 20.0
    if frente < frente_min:
        result["alertas"].append(f"Frente {frente:.1f}m < mínimo {frente_min:.0f}m para uso {uso}.")
        result["ok"] = False

    # Área de huella vs lote
    if area_huella <= 0:
        result["alertas"].append("Los retiros consumen todo el lote — revisar parámetros.")
        result["ok"] = False

    # Alerta si COS real supera el límite normativo
    if area_libre_min_pct > 0 and cos_real > cos_max_norma:
        result["alertas"].append(
            f"COS real ({cos_real:.1f}%) supera el máximo normativo ({cos_max_norma:.0f}%) "
            f"por área libre mínima exigida ({area_libre_min_pct:.0f}%). "
            f"Área techada efectiva limitada a {area_huella_efectiva:,.0f} m²."
        )

    # Pozos de luz RNE A.020: para >6 pisos exige pozo mínimo
    if n_pisos > 6 and uso == "residencial":
        h_edif = n_pisos * 3.0
        pozo_min = round(h_edif * 0.15, 1)
        result["alertas"].append(
            f"RNE A.020: edificio de {n_pisos} pisos ({h_edif:.0f}m) requiere pozos de luz ≥{pozo_min}m — verificar en planta.")

    result["metricas"] = {
        "area_lote_m2":          round(area_lote, 1),
        "area_huella_m2":        round(area_huella, 1),
        "area_huella_efectiva_m2": round(area_huella_efectiva, 1),
        "cos_real_pct":          cos_real,
        "cos_max_norma_pct":     cos_max_norma,
        "at_sobre_m2":           round(at_sobre, 0),
    }
    return result


def _geo_render_3d(poly_lote, poly_huella, n_pisos: int, n_sotanos: int = 0,
                   h_piso: float = 3.0, h_sotano: float = 2.6) -> go.Figure:
    """Genera visualización 3D del massing con Plotly."""
    GOLD, BLUE, RED = "#B8904A", "#4A90C4", "#C44A4A"

    def _ring_trace(poly, z_val, color, name="", width=2, showlegend=False):
        if poly is None or poly.is_empty:
            return None
        xs, ys = poly.exterior.xy
        return go.Scatter3d(
            x=list(xs), y=list(ys), z=[z_val] * len(xs),
            mode="lines", line=dict(color=color, width=width),
            name=name, showlegend=showlegend
        )

    traces = []

    # Límite del lote
    t = _ring_trace(poly_lote, 0, GOLD, "Lote", width=3, showlegend=True)
    if t:
        traces.append(t)

    if poly_huella and not poly_huella.is_empty:
        hxs, hys = list(poly_huella.exterior.xy[0]), list(poly_huella.exterior.xy[1])
        total_h = n_pisos * h_piso
        sotano_d = n_sotanos * h_sotano

        # Líneas de piso sobre rasante
        for p in range(n_pisos + 1):
            z = p * h_piso
            lbl = "Huella edificable" if p == 0 else ""
            show = p == 0
            t = _ring_trace(poly_huella, z, BLUE, lbl, width=2 if p == 0 else 1, showlegend=show)
            if t:
                traces.append(t)

        # Aristas verticales
        for cx, cy in zip(hxs[:-1], hys[:-1]):
            traces.append(go.Scatter3d(
                x=[cx, cx], y=[cy, cy], z=[-sotano_d, total_h],
                mode="lines", line=dict(color=BLUE, width=1), showlegend=False
            ))

        # Líneas de sótano
        for s in range(1, n_sotanos + 1):
            z = -s * h_sotano
            lbl = "Sótanos" if s == 1 else ""
            t = _ring_trace(poly_huella, z, GOLD, lbl, width=1, showlegend=(s == 1))
            if t:
                traces.append(t)

    fig = go.Figure(data=traces)
    fig.update_layout(
        scene=dict(
            xaxis=dict(title="m", showgrid=True, gridcolor="rgba(255,255,255,0.08)"),
            yaxis=dict(title="m", showgrid=True, gridcolor="rgba(255,255,255,0.08)"),
            zaxis=dict(title="altura (m)", showgrid=True, gridcolor="rgba(255,255,255,0.08)"),
            bgcolor="#0E1E2E", aspectmode="data",
            camera=dict(eye=dict(x=1.6, y=-1.6, z=1.2)),
        ),
        paper_bgcolor="#0A1628",
        font=dict(color="white", size=10),
        margin=dict(l=0, r=0, t=10, b=0),
        height=420,
        legend=dict(x=0.02, y=0.97, bgcolor="rgba(10,22,40,0.8)",
                    bordercolor="rgba(184,144,74,0.3)", font=dict(size=10)),
    )
    return fig


def _gen_seccion_alturas(n_pisos: int, n_sotanos: int = 0,
                         h_piso: float = 2.65, h_sotano: float = 2.80) -> go.Figure:
    """Esquema de alturas 2D — sección arquitectónica."""
    NAVY = "#1E2D3D"
    GOLD = "#B8904A"
    RED  = "#C44A4A"
    GREY = "#9A9590"
    BG   = "#FFFFFF"

    az_h  = 1.5
    BX0, BX1 = 1.5, 3.5
    DIM_X = BX1 + 0.40

    z_bot = -(n_sotanos * h_sotano)
    z_top = n_pisos * h_piso + az_h

    shapes = []
    annotations = []

    # Cuerpo sobre rasante
    shapes.append(dict(type="rect", x0=BX0, x1=BX1, y0=0, y1=n_pisos * h_piso,
                       line=dict(color=NAVY, width=2),
                       fillcolor="rgba(30,45,61,0.05)"))
    # Azotea (recuadro más estrecho)
    shapes.append(dict(type="rect", x0=BX0 + 0.22, x1=BX1 - 0.22,
                       y0=n_pisos * h_piso, y1=z_top,
                       line=dict(color=NAVY, width=2),
                       fillcolor="rgba(30,45,61,0.10)"))
    # Sótanos
    if n_sotanos > 0:
        shapes.append(dict(type="rect", x0=BX0, x1=BX1, y0=z_bot, y1=0,
                           line=dict(color=NAVY, width=2),
                           fillcolor="rgba(30,45,61,0.12)"))
        # Línea de terreno
        shapes.append(dict(type="line", x0=0.3, x1=5.2, y0=0, y1=0,
                           line=dict(color=GOLD, width=2.5)))
        annotations.append(dict(x=0.25, y=0, text="<b>N.0.00</b>",
                                 xanchor="right", yanchor="middle", showarrow=False,
                                 font=dict(size=9, color=GOLD, family="Courier New, monospace")))
    else:
        shapes.append(dict(type="line", x0=0.3, x1=5.2, y0=0, y1=0,
                           line=dict(color=GOLD, width=1.5, dash="dot")))
        annotations.append(dict(x=0.25, y=0, text="N.0.00",
                                 xanchor="right", yanchor="middle", showarrow=False,
                                 font=dict(size=9, color=GOLD, family="Courier New, monospace")))

    # Líneas de piso sobre rasante
    for p in range(1, n_pisos):
        shapes.append(dict(type="line", x0=BX0, x1=BX1, y0=p * h_piso, y1=p * h_piso,
                           line=dict(color=NAVY, width=0.7)))
    # Líneas de sótano
    for s in range(1, n_sotanos):
        shapes.append(dict(type="line", x0=BX0, x1=BX1,
                           y0=-s * h_sotano, y1=-s * h_sotano,
                           line=dict(color=NAVY, width=0.7, dash="dot")))
    # Rampa entre sótanos
    if n_sotanos >= 2:
        rx = BX0 + (BX1 - BX0) * 0.28
        for s in range(1, n_sotanos):
            ya = -(s - 1) * h_sotano - 0.25 if s > 1 else -0.25
            yb = -s * h_sotano + 0.25
            shapes.append(dict(type="line", x0=rx, x1=rx + 0.35, y0=ya, y1=yb,
                               line=dict(color=GREY, width=1, dash="dot")))

    # Etiquetas izquierda (nombre de nivel)
    for p in range(1, n_pisos + 1):
        annotations.append(dict(
            x=BX0 - 0.08, y=p * h_piso, text=f"P{p:02d}",
            xanchor="right", yanchor="middle", showarrow=False,
            font=dict(size=9, color=NAVY, family="Courier New, monospace")))
    annotations.append(dict(
        x=BX0 - 0.08, y=z_top, text="AZ",
        xanchor="right", yanchor="middle", showarrow=False,
        font=dict(size=9, color=NAVY, family="Courier New, monospace")))
    for s in range(1, n_sotanos + 1):
        annotations.append(dict(
            x=BX0 - 0.08, y=-s * h_sotano, text=f"S{s}",
            xanchor="right", yanchor="middle", showarrow=False,
            font=dict(size=9, color=NAVY, family="Courier New, monospace")))

    # Etiquetas derecha (cota)
    for p in range(1, n_pisos + 1):
        annotations.append(dict(
            x=BX1 + 0.08, y=p * h_piso, text=f"+{p * h_piso:.2f}",
            xanchor="left", yanchor="middle", showarrow=False,
            font=dict(size=9, color=GREY, family="Courier New, monospace")))
    annotations.append(dict(
        x=BX1 + 0.08, y=z_top, text=f"+{z_top:.2f}",
        xanchor="left", yanchor="middle", showarrow=False,
        font=dict(size=9, color=GREY, family="Courier New, monospace")))
    annotations.append(dict(
        x=BX1 + 0.08, y=0, text="±0.00",
        xanchor="left", yanchor="middle", showarrow=False,
        font=dict(size=9, color=GOLD, family="Courier New, monospace")))
    for s in range(1, n_sotanos + 1):
        annotations.append(dict(
            x=BX1 + 0.08, y=-s * h_sotano, text=f"{-s * h_sotano:.2f}",
            xanchor="left", yanchor="middle", showarrow=False,
            font=dict(size=9, color=GREY, family="Courier New, monospace")))

    # Cota de altura de piso (flecha roja entre P01 y P02)
    if n_pisos >= 2:
        for dz in [h_piso, 2 * h_piso]:
            shapes.append(dict(type="line", x0=DIM_X - 0.06, x1=DIM_X + 0.06, y0=dz, y1=dz,
                               line=dict(color=RED, width=1.5)))
        shapes.append(dict(type="line", x0=DIM_X, x1=DIM_X, y0=h_piso, y1=2 * h_piso,
                           line=dict(color=RED, width=1.5)))
        annotations.append(dict(
            x=DIM_X + 0.12, y=1.5 * h_piso, text=f"{h_piso:.2f}m",
            xanchor="left", yanchor="middle", showarrow=False,
            font=dict(size=9, color=RED, family="Courier New, monospace")))

    # Título y metadatos
    h_total = n_pisos * h_piso
    meta = f"CPU {n_pisos} PISOS: {h_total:.2f}ml  ·  ALT. PISO A PISO: {h_piso:.2f}m"
    if n_sotanos > 0:
        meta += f"  ·  {n_sotanos} SÓTANO{'S' if n_sotanos > 1 else ''}"
    annotations.append(dict(
        x=BX0, y=z_top + 0.85, text="<b>ESQUEMA DE ALTURAS</b>",
        xanchor="left", yanchor="bottom", showarrow=False,
        font=dict(size=12, color=NAVY, family="Inter, sans-serif")))
    annotations.append(dict(
        x=BX0, y=z_top + 0.35, text=meta,
        xanchor="left", yanchor="bottom", showarrow=False,
        font=dict(size=9, color=GREY, family="Courier New, monospace")))

    height_px = max(380, min(700, int((z_top - z_bot) * 30) + 130))
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=[None], y=[None], showlegend=False))
    fig.update_layout(
        shapes=shapes, annotations=annotations,
        paper_bgcolor=BG, plot_bgcolor=BG,
        height=height_px,
        margin=dict(l=60, r=90, t=70, b=30),
        xaxis=dict(range=[0, 5.8], showticklabels=False,
                   showgrid=False, zeroline=False, visible=False),
        yaxis=dict(range=[z_bot - 1.2, z_top + 1.8],
                   showticklabels=False, showgrid=False,
                   zeroline=False, visible=False),
    )
    return fig


def _gen_massing_3d_solid(poly_lote, poly_huella, n_pisos: int,
                           n_sotanos: int = 0, h_piso: float = 2.65,
                           h_sotano: float = 2.80,
                           unidades: list = None) -> go.Figure:
    """Massing 3D — tipologías por piso, márgenes compactos."""
    BGND = "#F0F2F6"
    GOLD = "#B8904A"
    TEXT = "#1E2D3D"
    GREY = "#5A6A7A"
    GAP  = 0.06

    # ── Distribución de tipologías por piso ──────────────────────────────
    _TIPO_COLOR = {"1D": "#4AB3D9", "2D": "#2A78C8", "3D": "#C8A050", "PH": "#9B59B6"}
    _TIPO_SHORT = {"1 Dorm.": "1D", "2 Dorm.": "2D", "3 Dorm.": "3D",
                   "1D": "1D", "2D": "2D", "3D": "3D",
                   "Dúplex": "PH", "PH": "PH"}

    def _build_floor_map(unids, n):
        if not unids or not n:
            return {}
        fmap = {}
        floor_idx = 0
        orden = ["1 Dorm.", "1D", "2 Dorm.", "2D", "3 Dorm.", "3D", "Dúplex", "PH"]
        seen = set()
        for key in orden:
            for u in unids:
                t = u.get("tipo", "")
                if t != key or t in seen:
                    continue
                seen.add(t)
                short = _TIPO_SHORT.get(t, t[:2])
                cant  = int(u.get("cantidad", 0))
                # 2 unidades/piso → cant/2 pisos (mín 1)
                n_floors = max(1, round(cant / 2))
                for _ in range(n_floors):
                    if floor_idx < n:
                        fmap[floor_idx] = short
                        floor_idx += 1
        # rellenar si sobran pisos con el último tipo asignado
        last = fmap.get(max(fmap.keys())) if fmap else None
        for i in range(floor_idx, n):
            fmap[i] = last or "—"
        return fmap

    _floor_map = _build_floor_map(unidades or [], n_pisos)

    def _floor_color(i, total):
        # Gradiente dramático: azul profundo abajo → azul acero arriba → toque celeste en cima
        f = i / max(total - 1, 1)
        r = int(0x12 + (0x4A - 0x12) * f)
        g = int(0x22 + (0x6A - 0x22) * f)
        b = int(0x3A + (0x90 - 0x3A) * f)
        return f"#{r:02X}{g:02X}{b:02X}"

    def _box(x0, y0, x1, y1, z0, z1, color, opacity=1.0, name="", show_leg=False):
        vx = [x0, x1, x1, x0, x0, x1, x1, x0]
        vy = [y0, y0, y1, y1, y0, y0, y1, y1]
        vz = [z0+GAP, z0+GAP, z0+GAP, z0+GAP, z1-GAP, z1-GAP, z1-GAP, z1-GAP]
        return go.Mesh3d(
            x=vx, y=vy, z=vz,
            i=[0,0,4,4, 0,1,5,4, 1,2,6,5, 2,3,7,6, 3,0,4,7, 4,5,6,4],
            j=[1,2,5,6, 1,5,5,0, 2,6,6,1, 3,7,7,2, 0,4,7,3, 5,6,7,6],
            k=[2,3,6,7, 5,4,1,4, 6,5,2,5, 7,6,3,6, 4,7,3,4, 6,7,4,7],
            color=color, opacity=opacity,
            flatshading=True,
            lighting=dict(ambient=0.55, diffuse=0.95, specular=0.45, roughness=0.4,
                          fresnel=0.3),
            lightposition=dict(x=100, y=200, z=300),
            name=name, showlegend=show_leg,
            hoverinfo="skip",
        )

    traces = []

    # Building bounds
    if poly_huella and not poly_huella.is_empty:
        hx0, hy0, hx1, hy1 = poly_huella.bounds
    else:
        hx0, hy0, hx1, hy1 = 0, 0, 20, 15
    if poly_lote and not poly_lote.is_empty:
        lx0, ly0, lx1, ly1 = poly_lote.bounds
    else:
        lx0, ly0, lx1, ly1 = hx0-2, hy0-2, hx1+2, hy1+2

    bw      = hx1 - hx0
    bd      = hy1 - hy0
    cy_y    = (hy0 + hy1) / 2
    lbl_off = max(bw * 0.55, 5.5)   # márgenes compactos para reducir fondo blanco
    lbl_y   = hy0 - max(bd * 0.10, 1.0)

    # Ground plane — terreno en verde visible
    traces.append(_box(lx0, ly0, lx1, ly1, -0.30, 0.0, "#27AE60",
                       opacity=0.60, name="Terreno", show_leg=True))

    # Basements
    for s in range(n_sotanos, 0, -1):
        z0 = -s * h_sotano
        z1 = -(s - 1) * h_sotano
        col = "#3C3C50" if (s % 2 == 0) else "#484860"
        traces.append(_box(hx0, hy0, hx1, hy1, z0, z1, col, opacity=0.88,
                           name="Sótanos" if s == n_sotanos else "",
                           show_leg=(s == n_sotanos)))

    # Floors — coloreados por tipología si hay datos, sino gradiente base
    _tipo_legend_seen = set()
    for p in range(n_pisos):
        tipo = _floor_map.get(p)
        if tipo and tipo in _TIPO_COLOR:
            col     = _TIPO_COLOR[tipo]
            leg_nm  = f"Tipo {tipo}"
            show_l  = leg_nm not in _tipo_legend_seen
            _tipo_legend_seen.add(leg_nm)
        else:
            col    = _floor_color(p, n_pisos)
            leg_nm = "Pisos"
            show_l = (p == 0 and not _floor_map)
        traces.append(_box(hx0, hy0, hx1, hy1,
                           p * h_piso, (p + 1) * h_piso,
                           col, opacity=0.92, name=leg_nm, show_leg=show_l))

    # Azotea cap
    ax_m  = bw * 0.08
    ay_m  = (hy1 - hy0) * 0.08
    z_az0 = n_pisos * h_piso
    z_az1 = z_az0 + max(h_piso * 0.5, 1.3)
    traces.append(_box(hx0+ax_m, hy0+ay_m, hx1-ax_m, hy1-ay_m,
                       z_az0, z_az1, "#C8A050", opacity=0.9,
                       name="Azotea", show_leg=True))

    # ── Floor labels — left of building (Y desplazado hacia cámara) ──
    lx = hx0 - lbl_off
    lbl_xs, lbl_ys, lbl_zs, lbl_ts = [], [], [], []

    for s in range(n_sotanos, 0, -1):
        lbl_xs.append(lx); lbl_ys.append(lbl_y)
        lbl_zs.append(-(s - 0.5) * h_sotano); lbl_ts.append(f"S{s}")

    for p in range(n_pisos):
        lbl_xs.append(lx); lbl_ys.append(lbl_y)
        lbl_zs.append((p + 0.5) * h_piso); lbl_ts.append(f"P{p+1:02d}")

    lbl_xs.append(lx); lbl_ys.append(lbl_y)
    lbl_zs.append(z_az0 + (z_az1 - z_az0) / 2); lbl_ts.append("AZ")

    traces.append(go.Scatter3d(
        x=lbl_xs, y=lbl_ys, z=lbl_zs, mode='text', text=lbl_ts,
        textfont=dict(size=10, color=TEXT, family="Courier New, monospace"),
        showlegend=False, hoverinfo="skip", name="",
    ))

    # ── Tipología por piso — columna junto a las etiquetas ──────────────
    if _floor_map:
        # Agrupar rangos consecutivos del mismo tipo para poner un solo label por banda
        _bands = []
        _prev_tipo = None
        _band_start = 0
        for _pi in range(n_pisos):
            _t = _floor_map.get(_pi, "—")
            if _t != _prev_tipo:
                if _prev_tipo is not None:
                    _bands.append((_prev_tipo, _band_start, _pi - 1))
                _prev_tipo = _t
                _band_start = _pi
        if _prev_tipo is not None:
            _bands.append((_prev_tipo, _band_start, n_pisos - 1))

        for _bt, _bf, _bl in _bands:
            _bz  = ((_bf + _bl) / 2 + 0.5) * h_piso
            _bx  = lx - lbl_off * 0.30  # columna a la derecha de las etiquetas P01..
            _col = _TIPO_COLOR.get(_bt, GREY)
            _lbl = f"[{_bt}]" if _bf == _bl else f"[{_bt}]\nP{_bf+1:02d}-P{_bl+1:02d}"
            traces.append(go.Scatter3d(
                x=[_bx], y=[lbl_y], z=[_bz],
                mode='text', text=[_lbl],
                textfont=dict(size=11, color=_col, family="Inter, sans-serif"),
                showlegend=False, hoverinfo="skip", name="",
            ))

    # ── Height cotas — right of building ─────────────
    cx = hx1 + lbl_off * 0.85
    cot_xs, cot_ys, cot_zs, cot_ts = [], [], [], []

    step = 1 if n_pisos <= 12 else 2
    cot_xs.append(cx); cot_ys.append(lbl_y); cot_zs.append(0.0); cot_ts.append("N.0.00")
    for p in range(n_pisos):
        z_top = (p + 1) * h_piso
        if (p + 1) % step == 0 or p == n_pisos - 1:
            cot_xs.append(cx); cot_ys.append(lbl_y)
            cot_zs.append(z_top); cot_ts.append(f"+{z_top:.2f}")
    for s in range(1, n_sotanos + 1):
        cot_xs.append(cx); cot_ys.append(lbl_y)
        cot_zs.append(-s * h_sotano); cot_ts.append(f"{-s*h_sotano:.2f}")

    traces.append(go.Scatter3d(
        x=cot_xs, y=cot_ys, z=cot_zs, mode='text', text=cot_ts,
        textfont=dict(size=9, color=GREY, family="Courier New, monospace"),
        showlegend=False, hoverinfo="skip", name="",
    ))

    # ── N.0.00 grade line (gold, spans full annotation width) ────
    traces.append(go.Scatter3d(
        x=[lx, cx], y=[lbl_y, lbl_y], z=[0.0, 0.0],
        mode='lines', line=dict(color=GOLD, width=4),
        showlegend=False, hoverinfo="skip", name="",
    ))

    # ── Piso-a-piso dimension indicator (red, between P01 and P02) ──
    if n_pisos >= 2:
        arr_x = hx1 + lbl_off * 0.28
        mid_z = h_piso * 1.5
        traces.append(go.Scatter3d(
            x=[arr_x, arr_x], y=[lbl_y, lbl_y], z=[h_piso, h_piso * 2],
            mode='lines', line=dict(color="#C04040", width=2),
            showlegend=False, hoverinfo="skip", name="",
        ))
        traces.append(go.Scatter3d(
            x=[arr_x + bw * 0.07], y=[lbl_y], z=[mid_z],
            mode='text', text=[f"{h_piso:.2f}m"],
            textfont=dict(size=8, color="#C04040", family="Courier New, monospace"),
            showlegend=False, hoverinfo="skip", name="",
        ))

    # ── Layout ───────────────────────────────────────
    h_total = n_pisos * h_piso
    z_base  = -n_sotanos * h_sotano
    x_extra = lbl_off * 1.05

    # Aspect ratio — limitar altura visual para que parezca edificio, no caja de cereales
    bld_w = hx1 - hx0
    bld_d = hy1 - hy0
    bld_h = h_total + n_sotanos * h_sotano
    _ax   = 1.0
    _ay   = max(0.5, min(round(bld_d / bld_w, 2), 1.5)) if bld_w > 0 else 1.0
    _az   = min(round(bld_h / bld_w, 2), 1.1) if bld_w > 0 else 1.0  # cap: evita caja de cereales

    fig = go.Figure(data=traces)
    fig.update_layout(
        scene=dict(
            xaxis=dict(visible=False, range=[hx0 - x_extra, hx1 + x_extra * 0.95]),
            yaxis=dict(visible=False, range=[ly0 - 5, ly1 + 5]),
            zaxis=dict(visible=False, range=[z_base - 2, h_total + 6]),
            bgcolor=BGND,
            aspectmode="manual",
            aspectratio=dict(x=_ax, y=_ay, z=_az),
            camera=dict(
                eye=dict(x=1.6, y=-1.2, z=0.65),   # vista frontal ancha — se ve el frente + volumen
                up=dict(x=0, y=0, z=1),
            ),
        ),
        paper_bgcolor=BGND,
        font=dict(color=TEXT, size=10, family="Inter, sans-serif"),
        margin=dict(l=10, r=10, t=10, b=10),
        height=650,
        legend=dict(
            x=0.01, y=0.97,
            bgcolor="rgba(240,242,246,0.92)",
            bordercolor="rgba(184,144,74,0.5)", borderwidth=1,
            font=dict(size=9, color=TEXT),
            itemsizing="constant",
        ),
        showlegend=True,
    )
    return fig


# ═══════════════════════════════════════════════════════
# DATOS DE MERCADO LIMA 2025-2026
# Fuente precios venta: Índice Urbania — Lima, Noviembre 2025
# Tipo de cambio: 3.45 S./USD (referencia interna Osterling Advisory)
# Base: mediana distrital S./m² convertida a USD/m²
# Ratios tipológicos (índice Urbania): 2br=+4%, 3br=-4%, 1br=+10% sobre mediana
# Estacionamientos / depósitos: estimados mercado 2024-2026
# ═══════════════════════════════════════════════════════

MERCADO = {
    # ── Fuente: Urbania Lima Index Nov-2025 · Tipo cambio: 3.45 S./USD
    # ── Tipología: 1br = base×1.10, 2br = base×1.04, 3br = base×0.96
    # ── Ordenado alfabéticamente por distrito
    "Ancón": {                                     # est. ~2,700 S./m² → $789/m²
        "precio_1br": 870,  "precio_2br": 820,  "precio_3br": 760,
        "precio_estac": 3500, "precio_deposito": 1000,
        "costo_construccion": 808,
        "velocidad_venta": 2.20, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.2, "yield_mercado_pct": 4.9, "variacion_anual_pct": 0.5,
    },
    "Ate": {                                       # Urbania: 4,531 S./m² → $1,313/m² (TC 3.45)
        "precio_1br": 1445, "precio_2br": 1365, "precio_3br": 1260,
        "precio_estac": 7000, "precio_deposito": 2500,
        "costo_construccion": 850,
        "velocidad_venta": 1.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 5.8, "yield_mercado_pct": 5.3, "variacion_anual_pct": -1.6,
    },
    "Barranco": {                                  # Urbania: 9,161 S./m² → $2,655/m² (TC 3.45)
        "precio_1br": 2920, "precio_2br": 2760, "precio_3br": 2550,
        "precio_estac": 14000, "precio_deposito": 5000,
        "costo_construccion": 1000,
        "velocidad_venta": 0.80, "duracion_base_meses": 30,
        "alquiler_m2_mes": 11.9, "yield_mercado_pct": 5.1, "variacion_anual_pct": -2.5,
    },
    "Breña": {                                     # Urbania: 5,222 S./m² → $1,514/m² (TC 3.45)
        "precio_1br": 1665, "precio_2br": 1575, "precio_3br": 1455,
        "precio_estac": 6000, "precio_deposito": 2000,
        "costo_construccion": 840,
        "velocidad_venta": 1.10, "duracion_base_meses": 26,
        "alquiler_m2_mes": 6.8, "yield_mercado_pct": 5.4, "variacion_anual_pct": 2.8,
    },
    "Callao": {                                    # Urbania: 3,239 S./m² → $939/m² (TC 3.45)
        "precio_1br": 1035, "precio_2br": 975,  "precio_3br": 900,
        "precio_estac": 5000, "precio_deposito": 1500,
        "costo_construccion": 820,
        "velocidad_venta": 1.30, "duracion_base_meses": 26,
        "alquiler_m2_mes": 4.1, "yield_mercado_pct": 5.3, "variacion_anual_pct": -9.1,
    },
    "Carabayllo": {                                # est. ~3,200 S./m² → $936/m²
        "precio_1br": 1030, "precio_2br": 975,  "precio_3br": 900,
        "precio_estac": 4000, "precio_deposito": 1400,
        "costo_construccion": 815,
        "velocidad_venta": 1.80, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.0, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.8,
    },
    "Cercado de Lima": {                           # Urbania: 6,253 S./m² → $1,813/m² (TC 3.45)
        "precio_1br": 1995, "precio_2br": 1885, "precio_3br": 1740,
        "precio_estac": 7000, "precio_deposito": 2500,
        "costo_construccion": 870,
        "velocidad_venta": 0.90, "duracion_base_meses": 28,
        "alquiler_m2_mes": 8.9, "yield_mercado_pct": 5.6, "variacion_anual_pct": 9.8,
    },
    "Chaclacayo": {                                # est. ~2,600 S./m² → $760/m²
        "precio_1br": 835,  "precio_2br": 790,  "precio_3br": 730,
        "precio_estac": 4000, "precio_deposito": 1200,
        "costo_construccion": 810,
        "velocidad_venta": 2.00, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.2, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.5,
    },
    "Chorrillos": {                                # Urbania: 5,718 S./m² → $1,658/m² (TC 3.45)
        "precio_1br": 1825, "precio_2br": 1725, "precio_3br": 1590,
        "precio_estac": 8500, "precio_deposito": 3000,
        "costo_construccion": 880,
        "velocidad_venta": 1.20, "duracion_base_meses": 26,
        "alquiler_m2_mes": 7.4, "yield_mercado_pct": 5.7, "variacion_anual_pct": -1.7,
    },
    "Cieneguilla": {                               # est. ~2,050 S./m² → $600/m²
        "precio_1br": 660,  "precio_2br": 625,  "precio_3br": 575,
        "precio_estac": 3500, "precio_deposito": 1000,
        "costo_construccion": 800,
        "velocidad_venta": 2.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 2.8, "yield_mercado_pct": 5.5, "variacion_anual_pct": 0.5,
    },
    "Comas": {                                     # est. ~3,490 S./m² → $1,020/m²
        "precio_1br": 1120, "precio_2br": 1060, "precio_3br": 980,
        "precio_estac": 4500, "precio_deposito": 1500,
        "costo_construccion": 820,
        "velocidad_venta": 1.70, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.5, "yield_mercado_pct": 5.1, "variacion_anual_pct": 1.0,
    },
    "El Agustino": {                               # est. ~3,933 S./m² → $1,150/m²
        "precio_1br": 1265, "precio_2br": 1195, "precio_3br": 1105,
        "precio_estac": 5000, "precio_deposito": 1700,
        "costo_construccion": 840,
        "velocidad_venta": 1.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 5.5, "yield_mercado_pct": 5.5, "variacion_anual_pct": 1.5,
    },
    "Independencia": {                             # est. ~3,762 S./m² → $1,100/m²
        "precio_1br": 1210, "precio_2br": 1145, "precio_3br": 1055,
        "precio_estac": 4500, "precio_deposito": 1500,
        "costo_construccion": 820,
        "velocidad_venta": 1.70, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.5, "yield_mercado_pct": 5.1, "variacion_anual_pct": 1.0,
    },
    "Jesús María": {                               # Urbania: 7,574 S./m² → $2,196/m² (TC 3.45)
        "precio_1br": 2415, "precio_2br": 2285, "precio_3br": 2110,
        "precio_estac": 9000, "precio_deposito": 3200,
        "costo_construccion": 880,
        "velocidad_venta": 1.30, "duracion_base_meses": 24,
        "alquiler_m2_mes": 9.5, "yield_mercado_pct": 5.2, "variacion_anual_pct": 7.4,
    },
    "La Molina": {                                 # Urbania: 5,337 S./m² → $1,547/m² (TC 3.45)
        "precio_1br": 1700, "precio_2br": 1610, "precio_3br": 1485,
        "precio_estac": 9000, "precio_deposito": 3200,
        "costo_construccion": 870,
        "velocidad_venta": 1.40, "duracion_base_meses": 24,
        "alquiler_m2_mes": 7.4, "yield_mercado_pct": 6.3, "variacion_anual_pct": -3.6,
    },
    "La Victoria": {                               # Urbania: 6,882 S./m² → $1,995/m² (TC 3.45) · +14.9%
        "precio_1br": 2195, "precio_2br": 2075, "precio_3br": 1915,
        "precio_estac": 6500, "precio_deposito": 2200,
        "costo_construccion": 860,
        "velocidad_venta": 1.00, "duracion_base_meses": 26,
        "alquiler_m2_mes": 10.5, "yield_mercado_pct": 6.3, "variacion_anual_pct": 14.9,
    },
    "Lince": {                                     # Urbania: 7,318 S./m² → $2,121/m² (TC 3.45)
        "precio_1br": 2335, "precio_2br": 2205, "precio_3br": 2035,
        "precio_estac": 8500, "precio_deposito": 3000,
        "costo_construccion": 875,
        "velocidad_venta": 1.20, "duracion_base_meses": 26,
        "alquiler_m2_mes": 9.9, "yield_mercado_pct": 5.6, "variacion_anual_pct": -1.4,
    },
    "Los Olivos": {                                # Urbania: 3,557 S./m² → $1,031/m² (TC 3.45)
        "precio_1br": 1135, "precio_2br": 1070, "precio_3br": 990,
        "precio_estac": 5000, "precio_deposito": 1600,
        "costo_construccion": 820,
        "velocidad_venta": 1.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.5, "yield_mercado_pct": 5.2, "variacion_anual_pct": -3.1,
    },
    "Lurigancho": {                                # est. ~3,078 S./m² → $900/m²
        "precio_1br": 990,  "precio_2br": 935,  "precio_3br": 865,
        "precio_estac": 4000, "precio_deposito": 1300,
        "costo_construccion": 810,
        "velocidad_venta": 1.90, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.8, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.5,
    },
    "Lurín": {                                     # est. ~2,736 S./m² → $800/m²
        "precio_1br": 880,  "precio_2br": 830,  "precio_3br": 770,
        "precio_estac": 4000, "precio_deposito": 1200,
        "costo_construccion": 810,
        "velocidad_venta": 2.00, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.5, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.0,
    },
    "Magdalena del Mar": {                         # Urbania: 6,908 S./m² → $2,002/m² (TC 3.45)
        "precio_1br": 2200, "precio_2br": 2080, "precio_3br": 1920,
        "precio_estac": 8500, "precio_deposito": 3000,
        "costo_construccion": 870,
        "velocidad_venta": 1.10, "duracion_base_meses": 26,
        "alquiler_m2_mes": 9.1, "yield_mercado_pct": 5.3, "variacion_anual_pct": 3.4,
    },
    "Miraflores": {                                # Urbania: 8,670 S./m² → $2,513/m² (TC 3.45)
        "precio_1br": 2765, "precio_2br": 2615, "precio_3br": 2415,
        "precio_estac": 15000, "precio_deposito": 5500,
        "costo_construccion": 1050,
        "velocidad_venta": 0.80, "duracion_base_meses": 30,
        "alquiler_m2_mes": 10.4, "yield_mercado_pct": 5.0, "variacion_anual_pct": -0.5,
    },
    "Pachacámac": {                                # est. ~2,905 S./m² → $850/m²
        "precio_1br": 935,  "precio_2br": 885,  "precio_3br": 815,
        "precio_estac": 4000, "precio_deposito": 1200,
        "costo_construccion": 810,
        "velocidad_venta": 2.00, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.5, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.0,
    },
    "Pueblo Libre": {                              # Urbania: 6,279 S./m² → $1,820/m² (TC 3.45)
        "precio_1br": 2000, "precio_2br": 1895, "precio_3br": 1750,
        "precio_estac": 8500, "precio_deposito": 3000,
        "costo_construccion": 880,
        "velocidad_venta": 1.20, "duracion_base_meses": 26,
        "alquiler_m2_mes": 8.6, "yield_mercado_pct": 5.7, "variacion_anual_pct": 1.3,
    },
    "Pucusana": {                                  # costa sur, balneario → est. $720/m²
        "precio_1br": 790,  "precio_2br": 750,  "precio_3br": 690,
        "precio_estac": 3500, "precio_deposito": 1000,
        "costo_construccion": 800,
        "velocidad_venta": 2.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.0, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.5,
    },
    "Puente Piedra": {                             # est. ~2,907 S./m² → $850/m²
        "precio_1br": 935,  "precio_2br": 885,  "precio_3br": 815,
        "precio_estac": 4500, "precio_deposito": 1500,
        "costo_construccion": 820,
        "velocidad_venta": 1.80, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.0, "yield_mercado_pct": 5.5, "variacion_anual_pct": 1.0,
    },
    "Punta Hermosa": {                             # balneario premium → est. $1,200/m²
        "precio_1br": 1320, "precio_2br": 1250, "precio_3br": 1150,
        "precio_estac": 6000, "precio_deposito": 2000,
        "costo_construccion": 850,
        "velocidad_venta": 1.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 5.5, "yield_mercado_pct": 5.3, "variacion_anual_pct": 1.5,
    },
    "Punta Negra": {                               # balneario costa sur → est. $950/m²
        "precio_1br": 1045, "precio_2br": 990,  "precio_3br": 910,
        "precio_estac": 5000, "precio_deposito": 1500,
        "costo_construccion": 830,
        "velocidad_venta": 1.80, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.0, "yield_mercado_pct": 5.1, "variacion_anual_pct": 1.0,
    },
    "La Perla": {                                  # Urbania: 4,393 S./m² → $1,274/m² (TC 3.45)
        "precio_1br": 1400, "precio_2br": 1325, "precio_3br": 1220,
        "precio_estac": 7000, "precio_deposito": 2500,
        "costo_construccion": 850,
        "velocidad_venta": 1.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 5.6, "yield_mercado_pct": 5.3, "variacion_anual_pct": 9.8,
    },
    "Bellavista": {                                # Urbania: 4,312 S./m² → $1,250/m² (TC 3.45)
        "precio_1br": 1375, "precio_2br": 1300, "precio_3br": 1200,
        "precio_estac": 7000, "precio_deposito": 2500,
        "costo_construccion": 850,
        "velocidad_venta": 1.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 5.5, "yield_mercado_pct": 5.3, "variacion_anual_pct": 3.6,
    },
    "Rímac": {                                     # est. ~4,959 S./m² → $1,450/m²
        "precio_1br": 1595, "precio_2br": 1510, "precio_3br": 1390,
        "precio_estac": 6000, "precio_deposito": 2000,
        "costo_construccion": 850,
        "velocidad_venta": 1.20, "duracion_base_meses": 26,
        "alquiler_m2_mes": 6.5, "yield_mercado_pct": 5.4, "variacion_anual_pct": 2.0,
    },
    "San Bartolo": {                               # balneario costa sur → est. $900/m²
        "precio_1br": 990,  "precio_2br": 935,  "precio_3br": 865,
        "precio_estac": 4500, "precio_deposito": 1300,
        "costo_construccion": 820,
        "velocidad_venta": 1.90, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.8, "yield_mercado_pct": 5.0, "variacion_anual_pct": 1.0,
    },
    "San Borja": {                                 # Urbania: 7,147 S./m² → $2,072/m² (TC 3.45)
        "precio_1br": 2280, "precio_2br": 2155, "precio_3br": 1990,
        "precio_estac": 13000, "precio_deposito": 4500,
        "costo_construccion": 960,
        "velocidad_venta": 0.90, "duracion_base_meses": 28,
        "alquiler_m2_mes": 8.0, "yield_mercado_pct": 4.9, "variacion_anual_pct": -2.1,
    },
    "San Isidro": {                                # Urbania: 9,231 S./m² → $2,676/m² (TC 3.45)
        "precio_1br": 2945, "precio_2br": 2785, "precio_3br": 2570,
        "precio_estac": 15000, "precio_deposito": 2500,
        "costo_construccion": 875,               # Boutique RDB mid-premium — referencia Clemente X / CAPECO 2025
        "velocidad_venta": 0.80, "duracion_base_meses": 24,
        "alquiler_m2_mes": 11.2, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.7,
    },
    "San Juan de Lurigancho": {                    # est. ~3,600 S./m² → $1,053/m²
        "precio_1br": 1160, "precio_2br": 1095, "precio_3br": 1010,
        "precio_estac": 5500, "precio_deposito": 2000,
        "costo_construccion": 830,
        "velocidad_venta": 1.60, "duracion_base_meses": 24,
        "alquiler_m2_mes": 5.0, "yield_mercado_pct": 5.5, "variacion_anual_pct": 0.0,
    },
    "San Juan de Miraflores": {                    # Urbania: 3,684 S./m² → $1,068/m² (TC 3.45)
        "precio_1br": 1175, "precio_2br": 1110, "precio_3br": 1025,
        "precio_estac": 5500, "precio_deposito": 1800,
        "costo_construccion": 830,
        "velocidad_venta": 1.60, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.7, "yield_mercado_pct": 5.2, "variacion_anual_pct": -1.1,
    },
    "San Luis": {                                  # est. ~5,814 S./m² → $1,700/m²
        "precio_1br": 1870, "precio_2br": 1770, "precio_3br": 1630,
        "precio_estac": 8000, "precio_deposito": 2800,
        "costo_construccion": 870,
        "velocidad_venta": 1.10, "duracion_base_meses": 26,
        "alquiler_m2_mes": 8.0, "yield_mercado_pct": 5.5, "variacion_anual_pct": 2.0,
    },
    "San Martín de Porres": {                      # Urbania: 3,002 S./m² → $870/m² (TC 3.45)
        "precio_1br": 955,  "precio_2br": 905,  "precio_3br": 835,
        "precio_estac": 4500, "precio_deposito": 1400,
        "costo_construccion": 810,
        "velocidad_venta": 1.60, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.8, "yield_mercado_pct": 5.2, "variacion_anual_pct": -1.8,
    },
    "San Miguel": {                                # Urbania: 6,147 S./m² → $1,782/m² (TC 3.45)
        "precio_1br": 1960, "precio_2br": 1855, "precio_3br": 1710,
        "precio_estac": 9500, "precio_deposito": 3500,
        "costo_construccion": 900,
        "velocidad_venta": 1.20, "duracion_base_meses": 26,
        "alquiler_m2_mes": 8.1, "yield_mercado_pct": 5.4, "variacion_anual_pct": 2.8,
    },
    "Santa Anita": {                               # est. ~4,800 S./m² → $1,404/m²
        "precio_1br": 1545, "precio_2br": 1460, "precio_3br": 1350,
        "precio_estac": 6000, "precio_deposito": 2000,
        "costo_construccion": 840,
        "velocidad_venta": 1.40, "duracion_base_meses": 24,
        "alquiler_m2_mes": 6.5, "yield_mercado_pct": 5.4, "variacion_anual_pct": 2.8,
    },
    "Santa María del Mar": {                       # balneario pequeño costa sur → est. $800/m²
        "precio_1br": 880,  "precio_2br": 830,  "precio_3br": 770,
        "precio_estac": 3500, "precio_deposito": 1000,
        "costo_construccion": 800,
        "velocidad_venta": 2.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.0, "yield_mercado_pct": 4.8, "variacion_anual_pct": 0.5,
    },
    "Santa Rosa": {                                # costa norte lejana → est. $700/m²
        "precio_1br": 770,  "precio_2br": 730,  "precio_3br": 670,
        "precio_estac": 3500, "precio_deposito": 1000,
        "costo_construccion": 800,
        "velocidad_venta": 2.50, "duracion_base_meses": 24,
        "alquiler_m2_mes": 3.0, "yield_mercado_pct": 5.0, "variacion_anual_pct": 0.5,
    },
    "Santiago de Surco": {                         # Urbania: 6,690 S./m² → $1,939/m² (TC 3.45)
        "precio_1br": 2135, "precio_2br": 2015, "precio_3br": 1860,
        "precio_estac": 11000, "precio_deposito": 4000,
        "costo_construccion": 930,
        "velocidad_venta": 1.10, "duracion_base_meses": 26,
        "alquiler_m2_mes": 7.9, "yield_mercado_pct": 5.0, "variacion_anual_pct": -1.1,
    },
    "Surquillo": {                                 # Urbania: 6,807 S./m² → $1,973/m² (TC 3.45)
        "precio_1br": 2170, "precio_2br": 2050, "precio_3br": 1895,
        "precio_estac": 9000, "precio_deposito": 3200,
        "costo_construccion": 890,
        "velocidad_venta": 1.20, "duracion_base_meses": 26,
        "alquiler_m2_mes": 10.2, "yield_mercado_pct": 6.1, "variacion_anual_pct": 3.1,
    },
    "Villa El Salvador": {                         # est. ~3,400 S./m² → $994/m²
        "precio_1br": 1095, "precio_2br": 1035, "precio_3br": 955,
        "precio_estac": 5000, "precio_deposito": 1800,
        "costo_construccion": 820,
        "velocidad_venta": 1.80, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.5, "yield_mercado_pct": 5.1, "variacion_anual_pct": 0.0,
    },
    "Villa María del Triunfo": {                   # est. ~3,249 S./m² → $950/m²
        "precio_1br": 1045, "precio_2br": 990,  "precio_3br": 910,
        "precio_estac": 4200, "precio_deposito": 1400,
        "costo_construccion": 818,
        "velocidad_venta": 1.80, "duracion_base_meses": 24,
        "alquiler_m2_mes": 4.2, "yield_mercado_pct": 5.2, "variacion_anual_pct": 0.5,
    },
    "Otro": {                                      # promedio Lima Index ~$1,700/m²
        "precio_1br": 1870, "precio_2br": 1770, "precio_3br": 1630,
        "precio_estac": 7000, "precio_deposito": 2500,
        "costo_construccion": 850,
        "velocidad_venta": 1.00, "duracion_base_meses": 24,
        "alquiler_m2_mes": 7.0, "yield_mercado_pct": 5.25, "variacion_anual_pct": 1.1,
    },
}

# ═══════════════════════════════════════════════════════
# PRECIOS DE MERCADO — carga desde Google Sheet (con fallback)
# ═══════════════════════════════════════════════════════
_MERCADO_COLS = [
    "precio_1br","precio_2br","precio_3br",
    "precio_estac","precio_deposito","costo_construccion",
    "velocidad_venta","duracion_base_meses",
    "alquiler_m2_mes","yield_mercado_pct","variacion_anual_pct",
]
_MERCADO_INT_COLS  = {"precio_1br","precio_2br","precio_3br","precio_estac","precio_deposito",
                      "costo_construccion","duracion_base_meses"}
_MERCADO_FLOAT_COLS= {"velocidad_venta","alquiler_m2_mes","yield_mercado_pct","variacion_anual_pct"}

@st.cache_data(ttl=3600, show_spinner=False)
def _cargar_mercado_sheet() -> tuple:
    """Lee precios desde Google Sheet publicado como CSV. TTL 1 hora.
    Soporta dos formatos:
      - Nuevo (precio_m2_soles + alquiler_soles_100m2): convierte a USD usando tipo_cambio
      - Clásico (precio_1br, precio_2br, precio_3br en USD): usa directamente
    Fila especial CONFIG con columna tipo_cambio actualiza el tipo de cambio dinámicamente.
    Retorna (precios_dict, tipo_cambio).
    """
    _tc_secrets = float((st.secrets.get("mercado") or {}).get("tipo_cambio", 3.45))
    try:
        sheet_url = (st.secrets.get("mercado", {}) or {}).get("sheet_url", "")
        if not sheet_url:
            return {}, _tc_secrets
        import urllib.request, io as _io
        with urllib.request.urlopen(sheet_url, timeout=8) as resp:
            raw = resp.read().decode("utf-8")
        df = pd.read_csv(_io.StringIO(raw))
        df.columns = [c.strip().lower() for c in df.columns]
        if "distrito" not in df.columns:
            return {}, _tc_secrets
        # Leer tipo_cambio desde fila CONFIG
        tc = _tc_secrets
        if "tipo_cambio" in df.columns:
            cfg = df[df["distrito"].astype(str).str.strip().str.upper() == "CONFIG"]
            if not cfg.empty:
                try:
                    v = float(str(cfg.iloc[0]["tipo_cambio"]).replace(",","").strip())
                    if v > 0:
                        tc = v
                except (ValueError, TypeError):
                    pass
        use_soles = "precio_m2_soles" in df.columns
        out = {}
        for _, row in df.iterrows():
            distrito = str(row.get("distrito", "")).strip()
            if not distrito or distrito.lower() == "nan" or distrito.upper() == "CONFIG":
                continue
            entry = {}
            if use_soles:
                try:
                    base_s = float(str(row.get("precio_m2_soles","")).replace(",","").strip())
                    base_u = base_s / tc
                    entry["precio_1br"] = int(base_u * 1.10)
                    entry["precio_2br"] = int(base_u * 1.04)
                    entry["precio_3br"] = int(base_u * 0.96)
                except (ValueError, TypeError):
                    pass
                try:
                    alq = float(str(row.get("alquiler_soles_100m2","")).replace(",","").strip())
                    entry["alquiler_m2_mes"] = round(alq / 100 / tc, 2)
                except (ValueError, TypeError):
                    pass
                otros = ["precio_estac","precio_deposito","costo_construccion",
                         "velocidad_venta","duracion_base_meses",
                         "yield_mercado_pct","variacion_anual_pct"]
                for col in otros:
                    if col not in row:
                        continue
                    try:
                        val = float(str(row[col]).replace(",","").strip())
                        entry[col] = int(val) if col in _MERCADO_INT_COLS else round(val, 4)
                    except (ValueError, TypeError):
                        pass
            else:
                for col in _MERCADO_COLS:
                    if col not in row:
                        continue
                    try:
                        val = float(str(row[col]).replace(",","").strip())
                        entry[col] = int(val) if col in _MERCADO_INT_COLS else round(val, 4)
                    except (ValueError, TypeError):
                        pass
            if entry:
                out[distrito] = entry
        return out, tc
    except Exception:
        return {}, _tc_secrets

def get_mercado() -> dict:
    """MERCADO actualizado: Sheet > hardcoded. Fusiona para preservar distritos sin Sheet."""
    global TIPO_CAMBIO
    live, tc = _cargar_mercado_sheet()
    TIPO_CAMBIO = tc
    if not live:
        return MERCADO
    merged = dict(MERCADO)
    for distrito, datos in live.items():
        if distrito in merged:
            merged[distrito] = {**merged[distrito], **datos}
        else:
            merged[distrito] = datos
    return merged

# Tipo de cambio S./USD — se actualiza desde Sheet (fila CONFIG) o secrets.toml
TIPO_CAMBIO = float((st.secrets.get("mercado") or {}).get("tipo_cambio", 3.45))
# Sobrescribe MERCADO con versión live (Sheet > hardcoded, con fallback automático)
MERCADO = get_mercado()

# ═══════════════════════════════════════════════════════
# CLAUDE API
# ═══════════════════════════════════════════════════════

def _sanitize_api_key(raw: str) -> str:
    """Strips non-ASCII characters that macOS autocorrect can silently inject."""
    return raw.encode("ascii", errors="ignore").decode("ascii").strip()


def get_client():
    api_key = (
        os.environ.get("ANTHROPIC_API_KEY")
        or (st.secrets.get("anthropic", {}) or {}).get("api_key", "")
        or st.session_state.get("api_key_input", "")
    )
    if not api_key:
        st.error("⚠️ Clave API no configurada. Contacta al administrador.")
        st.stop()
    api_key = _sanitize_api_key(api_key)
    if not api_key.startswith("sk-"):
        st.error("🔑 Clave API inválida. Contacta al administrador.")
        st.stop()
    return anthropic.Anthropic(api_key=api_key, max_retries=0, timeout=300.0)


def _run_with_retry(fn, spinner_msg, max_attempts=3):
    """Ejecuta fn() reintentando en errores de red, servidor ocupado o rate limit."""
    import time
    _status = st.empty()
    last_err = ""
    for attempt in range(1, max_attempts + 1):
        label = spinner_msg if attempt == 1 else f"{spinner_msg} — intento {attempt}/{max_attempts}"
        with st.spinner(label):
            try:
                result = fn()
                _status.empty()
                return result
            except Exception as e:
                last_err = str(e)
                err_lower = last_err.lower()

                # Auth error: clave inválida → limpiar y pedir que la re-ingresen
                if "401" in last_err or "authentication_error" in err_lower or "invalid x-api-key" in err_lower or "invalid api key" in err_lower:
                    _status.empty()
                    st.session_state.pop("api_key_input", None)
                    st.error("🔑 Clave de acceso inválida. Ingresa una clave API de Anthropic válida en el panel izquierdo (⚙ Configuración) e inténtalo de nuevo.")
                    st.stop()

                is_rate_limit = "429" in last_err or "rate_limit" in err_lower or "rate limit" in err_lower
                is_retriable = (
                    is_rate_limit
                    or "500" in last_err
                    or "529" in last_err
                    or "502" in last_err
                    or "503" in last_err
                    or "overloaded" in err_lower
                    or "bad gateway" in err_lower
                    or "internal server" in err_lower
                    or "connection" in err_lower
                    or "connect" in err_lower
                    or "timeout" in err_lower
                    or "timed out" in err_lower
                    or "network" in err_lower
                    or "remotedisconnected" in err_lower
                    or "read timeout" in err_lower
                    or "eoferror" in err_lower
                    or "reset by peer" in err_lower
                    or "ssl" in err_lower
                    or "json_parse_error" in err_lower
                )
                if not is_retriable:
                    _status.empty()
                    st.error(f"Error inesperado: {last_err[:400]}")
                    st.stop()

        if attempt < max_attempts:
            wait = 20 if is_rate_limit else 5
            for s in range(wait, 0, -1):
                msg = (f"Optimizando consulta — reintentando en {s}s… ({attempt + 1}/{max_attempts})"
                       if is_rate_limit else
                       f"Reintentando en {s}s… (intento {attempt + 1}/{max_attempts})")
                _status.info(msg)
                time.sleep(1)
            _status.empty()

    _status.empty()
    if "json_parse_error" in last_err.lower():
        st.error("El análisis no pudo completarse: Claude devolvió una respuesta incompleta. "
                 "Intenta con documentos más pequeños (máx. ~1MB cada uno) o reduce el número de archivos adjuntos.")
    elif "429" in last_err or "rate_limit" in last_err.lower():
        st.error("Se superó el límite de tokens por minuto de tu cuenta Anthropic. "
                 "Espera 1 minuto e inténtalo nuevamente, o reduce el número de documentos adjuntos.")
    else:
        st.error(f"No se pudo conectar con el servicio de análisis después de varios intentos. "
                 f"Verifica tu conexión a internet e inténtalo nuevamente.")
    st.stop()


def parse_json_safe(text: str) -> dict:
    """Extrae y parsea JSON de la respuesta de Claude. Lanza ValueError si falla."""
    # 1. Intentar extraer de bloque markdown ```json ... ```
    md_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if md_match:
        raw = md_match.group(1)
    else:
        # 2. Extraer el bloque JSON más externo { ... }
        brace_match = re.search(r'\{.*\}', text, re.DOTALL)
        if not brace_match:
            raise ValueError(f"json_parse_error: respuesta sin JSON — '{text[:120]}'")
        raw = brace_match.group()

    # Limpiar problemas comunes
    raw = re.sub(r'//[^\n]*', '', raw)
    raw = re.sub(r'/\*.*?\*/', '', raw, flags=re.DOTALL)
    raw = re.sub(r',\s*([\}\]])', r'\1', raw)
    raw = raw.replace('“', '"').replace('”', '"')
    raw = raw.replace('‘', "'").replace('’', "'")

    try:
        result = json.loads(raw)
        if not isinstance(result, dict) or len(result) == 0:
            raise ValueError(f"json_parse_error: JSON vacío o no es objeto")
        return result
    except json.JSONDecodeError as e:
        raise ValueError(f"json_parse_error: {e} — '{raw[:120]}'")


def pdf_block(pdf_bytes: bytes) -> dict:
    return {
        "type": "document",
        "source": {
            "type": "base64",
            "media_type": "application/pdf",
            "data": base64.standard_b64encode(pdf_bytes).decode("utf-8")
        }
    }


def image_block(img_bytes: bytes, media_type: str = "image/jpeg") -> dict:
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": media_type,
            "data": base64.standard_b64encode(img_bytes).decode("utf-8"),
        }
    }




# ── CARGA DE NORMATIVAS DESDE ARCHIVOS EXTERNOS ─────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def _load_norm(filename: str) -> str:
    """Carga un archivo de normativa desde la carpeta normativas/. Cacheado por sesión."""
    base = pathlib.Path(__file__).parent / "normativas"
    path = base / filename
    if path.exists():
        return path.read_text(encoding="utf-8")
    return f"[NORMATIVA NO ENCONTRADA: {filename}]"

RIN_SAN_ISIDRO = _load_norm("rin_san_isidro.txt")


# ── NORMATIVA ESPECÍFICA MIRAFLORES ──────────────────────────────────────────────────────────────
RIN_MIRAFLORES = _load_norm("rin_miraflores.txt")


# ── NORMATIVA ESPECÍFICA: LA VICTORIA ────────────────────────────────────────────────────────────
# Fuente: CPU N°000371-2021 (Av. Javier Prado/Palomar Norte), CPU N°000367-2021 (Av. Carlos Villarán/Palomar Norte),
#         CPU N°000132-2024 (Av. Javier Prado Este 1197-1199, caduca 01/04/2027)
RIN_LA_VICTORIA = _load_norm("rin_la_victoria.txt")


# ── NORMATIVA ESPECÍFICA: LINCE ───────────────────────────────────────────────────────────────────
# Fuente: CPU N°006-2021-MDL/GDU/SOPPUC (Av. Arequipa 2698 esq. Jr. Soledad, Ámbito C, caduca 14/01/2024)
RIN_LINCE = _load_norm("rin_lince.txt")


# ── NORMATIVA ESPECÍFICA: MAGDALENA DEL MAR ──────────────────────────────────────────────────────
# Fuente: CPU N°00377-2019 (Pque. Francisco Graña — RDB/Sector III), CPU N°00104-2020 (Jr. Flora Tristán — RDB/Sector IV),
#         CPU N°00100-2023 (Av. Javier Prado Oeste 2281-2291 — E3/Sector IV, caduca 29/03/2026)
RIN_MAGDALENA = _load_norm("rin_magdalena.txt")


# ── NORMATIVA ESPECÍFICA: JESÚS MARÍA ────────────────────────────────────────────────────────────
# Fuente: CPU N°391-2019-MDJM (Av. Garzón 550-572, RDA, caduca 03/10/2022)
#         CPU N°489-2018-MDJM (Av. Garzón 082, RDA+RDM, caduca 27/11/2021)
RIN_JESUS_MARIA = _load_norm("rin_jesus_maria.txt")

RIN_CERCADO_LIMA = _load_norm("rin_cercado_lima.txt")

RIN_SAN_BORJA = _load_norm("rin_san_borja.txt")

RIN_SANTA_ANITA = _load_norm("rin_santa_anita.txt")

RIN_SURCO = _load_norm("rin_surco.txt")

RIN_SURQUILLO = _load_norm("rin_surquillo.txt")

RIN_VILLA_EL_SALVADOR = _load_norm("rin_villa_el_salvador.txt")

RIN_SAN_JUAN_LURIGANCHO = _load_norm("rin_san_juan_lurigancho.txt")


# ── CONOCIMIENTO NORMATIVO: LIMA METROPOLITANA Y DISTRITOS ───────────────────────────────────────
REFERENCIAS_NORMATIVAS_LIMA = _load_norm("referencias_lima.txt")

# ── BENCHMARKS INDUSTRIALES (costos nave, Parque Logístico 47, retornos) ─────────────────────────
BENCHMARKS_INDUSTRIAL = _load_norm("benchmarks_industrial.txt")

# ── ÍNDICE DE USOS ATN-I (6,349 actividades, P=Permitido H=Compatible) ──────────────────────────
_INDICE_USOS_ATN_RAW = _load_norm("indice_usos_atni.txt")

def _parse_indice_usos_entries(raw: str) -> list:
    entries = []
    for line in raw.split('\n'):
        if line.count('|') < 2:
            continue
        parts = line.split('|')
        desc = parts[1].strip()
        if not desc or len(desc) < 4:
            continue
        zones: dict = {}
        for token in parts[2].strip().split():
            if '=' in token:
                z, v = token.split('=', 1)
                zones[z.strip()] = v.strip()
        if zones:
            entries.append({"desc": desc, "zones": zones})
    return entries

_INDICE_USOS_ENTRIES = _parse_indice_usos_entries(_INDICE_USOS_ATN_RAW)


def _ascii_lower(s: str) -> str:
    """Lowercase + strip accents for accent-insensitive search."""
    import unicodedata
    return unicodedata.normalize('NFKD', s.lower()).encode('ascii', 'ignore').decode()


def _buscar_actividad_indice(query: str, zona_sel: str, max_results: int = 12) -> list:
    """Keyword search across Índice de Usos ATN-I (Ord. 933-MML).
    Accent-insensitive + prefix match (≥7 chars) to handle plural/gender endings.
    zona_sel: 'I1'–'I4' or 'OU' or 'CZ'/'CM'/etc.
    Returns list of dicts: {desc, zones, compat} where compat='P','H' or None."""
    if not query or len(query.strip()) < 3:
        return []
    keywords = [_ascii_lower(w) for w in query.split() if len(w) >= 3]
    if not keywords:
        return []
    # Normalize sidebar zone ('I2') to file format ('I-2')
    if len(zona_sel) == 2 and zona_sel[0] == 'I' and zona_sel[1].isdigit():
        zona_norm = f"I-{zona_sel[1]}"
    else:
        zona_norm = zona_sel

    def _kw_match(kw: str, desc_l: str, desc_words: list) -> bool:
        if kw in desc_l:
            return True
        if len(kw) >= 7:
            pfx = kw[:7]
            return any(w.startswith(pfx) for w in desc_words)
        return False

    scored = []
    for entry in _INDICE_USOS_ENTRIES:
        desc_l = _ascii_lower(entry["desc"])
        desc_words = desc_l.split()
        score = sum(1 for kw in keywords if _kw_match(kw, desc_l, desc_words))
        if score > 0:
            compat = entry["zones"].get(zona_norm) or entry["zones"].get(zona_sel)
            scored.append((score, entry["desc"], entry["zones"], compat))
    scored.sort(key=lambda x: (-x[0], x[1]))
    return [{"desc": d, "zones": z, "compat": c} for _, d, z, c in scored[:max_results]]


# ── REGLAMENTO NACIONAL DE EDIFICACIONES (RNE) ───────────────────────────────────────────────────
# Norma A.010 (RM 191-2021-VIVIENDA) y Norma A.020 (RM 188-2021-VIVIENDA)
RNE_NACIONAL = _load_norm("rne_nacional.txt")


def extract_parameters(cert_bytes: bytes, norm_docs: list[bytes]) -> dict:
    client = get_client()

    content = [pdf_block(cert_bytes)]
    for doc in norm_docs:
        content.append(pdf_block(doc))

    docs_note = (
        f"\nAdemás del certificado, se adjuntan {len(norm_docs)} documento(s) normativos adicionales "
        f"(reglamentos, ordenanzas, cuadros de parámetros). Analízalos en conjunto para identificar "
        f"TODOS los beneficios, incentivos, mecanismos de densificación, excepciones o notas que puedan "
        f"incrementar la altura, reducir aportes, flexibilizar retiros u otorgar cualquier ventaja al proyecto."
        if norm_docs else ""
    )

    prompt = f"""Analiza este Certificado de Parámetros Urbanísticos y Edificatorios de Lima, Perú.{docs_note}

Devuelve ÚNICAMENTE el siguiente JSON, sin texto antes ni después, sin bloques de código markdown:

{{
  "zonificacion": "CZ",
  "distrito": "San Isidro",
  "ubicacion": "Ca. Las Camelias 180",
  "area_terreno_m2": 500.0,
  "frente_ml": 20.0,
  "pisos_max": 8,
  "pisos_por_via": [{{"via": "Ca. Las Camelias", "pisos": 8}}],
  "area_libre_min_pct": 30.0,
  "retiro_frontal_ml": 3.0,
  "retiro_lateral_ml": null,
  "retiro_posterior_ml": null,
  "coeficiente_edificacion": null,
  "usos_compatibles": ["Residencial", "Comercio"],
  "estacionamientos_norma": "2 est./vivienda (Ámbito B/C, Anexo N°02 RIN 523-2020)",
  "estac_visitas_norma": "texto exacto del CPU sobre estacionamiento de visitas, ej: '15% máximo del total' o '1 espacio' o '10% de residentes' — null si no especifica",
  "ambito_urbano": "B",
  "sector_urbano": "A",
  "fecha_caducidad": "2026-12-31",
  "notas_altura": ["Nota 12: lote esquina aplica promedio de alturas"],
  "ordenanzas_base": ["RIN 523-2020"],
  "beneficios_normativos": [
    {{
      "tipo": "mayor_altura",
      "descripcion": "Lote esquina permite promedio de alturas",
      "condicion": "Lote debe tener frente a dos vías",
      "impacto_estimado": "+2 pisos adicionales",
      "base_legal": "RIN 523-2020 Nota 12"
    }}
  ]
}}

IMPORTANTE: Reemplaza todos los valores del ejemplo con los datos reales del documento.
- area_libre_min_pct: si dice "no se exige área libre" usa 0
- pisos_max: el máximo entre todas las vías del certificado
- Si un dato no existe en el documento usa null
- ambito_urbano: identifica A, B, C o D según la ubicación del predio y el RIN del distrito correspondiente (San Isidro: Ámbitos A/B/C/D del Anexo N°02; otros distritos: usa null si no aplica)
- sector_urbano: si es Miraflores, extrae el Sector Urbano (A, B, C, D); para otros distritos usa null
- estacionamientos_norma: usa el ratio correcto según la normativa del distrito identificado — consulta el RIN correspondiente de la lista anterior; incluye la base legal (Ord./RIN y artículo)
- estac_visitas_norma: copia el texto EXACTO del certificado sobre visitas (ej: "15% máximo", "1 espacio", "10% del total"). Si el certificado no menciona visitas, usa null
- ordenanzas_base: cita las ordenanzas vigentes del certificado, no las derogadas
- beneficios_normativos: incluye lotes esquina, frente a parque o zona monumental, acumulación de lotes, bonificaciones de altura por vía, retiros compensados, usos mixtos, TDR, ATN, CZ, cualquier mecanismo que beneficie el proyecto según la normativa del distrito"""

    content.append({"type": "text", "text": prompt})

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        system=[{"type": "text",
                 "text": "Eres un arquitecto experto en normativa urbanística de Lima, Perú. Respondes únicamente con JSON válido, sin texto adicional.",
                 "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": content}]
    )

    if not response.content:
        raise ValueError("json_parse_error: API devolvió respuesta vacía")
    text = response.content[0].text.strip()
    return parse_json_safe(text)


def generate_cabida(params: dict, config: dict) -> dict:
    client = get_client()

    beneficios_txt = ""
    if params.get("beneficios_normativos"):
        beneficios_txt = (
            "\n\nBENEFICIOS NORMATIVOS POTENCIALES DETECTADOS (NO aplicar — reportar como observación):\n"
            + json.dumps(params["beneficios_normativos"], ensure_ascii=False, indent=2)
            + "\nINSTRUCCIÓN: estos beneficios NO deben modificar el análisis. "
            "Inclúyelos en 'observaciones' marcados como 'BENEFICIO POTENCIAL — requiere verificación del usuario'."
        )

    sugerencias_txt = f"\n\nNOTAS DEL ANALISTA:\n{config.get('sugerencias', '')}" if config.get("sugerencias") else ""

    # Regla de colindancia — calcular programáticamente y sobreescribir pisos_max
    colind_izq = config.get("colindante_izq_pisos")
    colind_der = config.get("colindante_der_pisos")
    colind_txt = ""
    params_cabida = dict(params)  # copia local — no mutar el original

    if colind_izq or colind_der:
        _colind_max  = max(colind_izq or 0, colind_der or 0)
        _base_pisos  = int(params_cabida.get("pisos_max") or 5)
        _pisos_calc  = int((_colind_max + _base_pisos) / 2)  # Art.6.3 Ord.523-MSI: floor — decimal no sube al entero superior
        # Sobreescribir en la copia que recibe Claude — no es sugerencia, es valor fijo
        params_cabida["pisos_max"] = _pisos_calc
        partes = []
        if colind_izq: partes.append(f"izquierdo: {colind_izq} pisos")
        if colind_der: partes.append(f"derecho: {colind_der} pisos")
        colind_txt = (
            f"\n\nCOLINDANTES VERIFICADOS EN CAMPO: {' | '.join(partes)}.\n"
            f"REGLA DE COLINDANCIA APLICADA (cálculo determinístico):\n"
            f"  colindante más alto = {_colind_max} pisos | altura base norma = {_base_pisos} pisos\n"
            f"  pisos_max = floor(({_colind_max} + {_base_pisos}) / 2) = {_pisos_calc} pisos\n"
            f"  *** pisos_max YA ha sido actualizado a {_pisos_calc} en los parámetros. "
            f"USA EXACTAMENTE {_pisos_calc} PISOS. NO uses la altura base original. ***"
        )
    else:
        params_cabida = params

    distrito = params.get("distrito", "")
    ambito = params.get("ambito_urbano", "")
    sector = params.get("sector_urbano", "")

    _visitas_norma = params.get("estac_visitas_norma") or ""
    if _visitas_norma:
        _pct = next((float(x.replace("%","").strip()) for x in _visitas_norma.split()
                     if x.replace("%","").strip().replace(".","").isdigit() and "%" in _visitas_norma), None)
        if _pct:
            estac_visitas_instruccion = (f"El CPU especifica visitas: '{_visitas_norma}'. "
                                         f"Calcula {_pct}% de estac_residentes redondeado al entero superior, mínimo 1.")
        else:
            estac_visitas_instruccion = f"El CPU especifica visitas: '{_visitas_norma}'. Aplica exactamente ese criterio, mínimo 1."
    else:
        estac_visitas_instruccion = "El CPU no especifica visitas. Usa estac_visitas = 4."
    # RNE_NACIONAL y REFERENCIAS_NORMATIVAS_LIMA van cacheados en el system prompt
    normativa_note = ""
    if "san isidro" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA SAN ISIDRO (Ord. 523-MSI):\n{RIN_SAN_ISIDRO}"
        if ambito:
            normativa_note += f"\n\nÁMBITO IDENTIFICADO: {ambito} — aplica reglas de estacionamiento y áreas mínimas del ámbito correspondiente."
    elif "miraflores" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA MIRAFLORES (Ord. 342-MM y modificatorias):\n{RIN_MIRAFLORES}"
        if sector:
            normativa_note += (
                f"\n\nSECTOR URBANO IDENTIFICADO: {sector} — aplica altura normativa del sector según "
                f"Ord. 226-MM. Si el sector es A con RDMA y lote ≥ normativo, altura hasta 17 pisos "
                f"(condición: área edificable total ≤ 0.60 × área_terreno × 12 pisos, Ord. 342-MM Art. 6° literal g)."
            )
    elif "la victoria" in distrito.lower() or "lavictoria" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA LA VICTORIA (Ord. N°355-MLV + Ord. N°1082-MML):\n{RIN_LA_VICTORIA}"
        normativa_note += (
            "\n\nCLAVE PARA CZ LA VICTORIA: la altura CZ = 1.5×(a+r) donde a=ancho vía y r=suma retiros ambos lados. "
            "No hay tope fijo de pisos — calcular con datos del certificado. "
            "El uso residencial compatible (RDM o RDA) determina áreas mínimas y estacionamientos."
        )
    elif "lince" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA LINCE (Ord. N°235-MDL y modificatorias):\n{RIN_LINCE}"
        normativa_note += (
            "\n\nCLAVES PARA LINCE: (1) Altura CM = 1.5×(a+r) — sin tope fijo de pisos; calcular con ancho de vía y retiros del certificado. "
            "(2) Si el ámbito es C (Parque Castilla), aplicar alturas especiales: 10/15/20 pisos según tamaño de lote. "
            "(3) Estacionamiento residencial = 1 cada 2 viviendas (más restrictivo que otros distritos). "
            "(4) Máx 40% de unidades de 1 dormitorio."
        )
    elif "magdalena" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA MAGDALENA DEL MAR (Ord. N°950-MML + Ord. N°017-2016-MDMM):\n{RIN_MAGDALENA}"
        normativa_note += (
            "\n\nCLAVES PARA MAGDALENA: (1) Altura máxima FIJA = 4 pisos = 12.00 ml — no hay excepciones confirmadas. "
            "(2) Estacionamiento = 1 est por vivienda — el más restrictivo de todos los distritos. "
            "(3) Zona predominante RDB — no hay RDA ni CM confirmados. "
            "(4) Lotes E3 sobre Av. Javier Prado pueden reconvertirse a residencial sin cambio de zonificación."
        )
    elif "jes" in distrito.lower() and "mar" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA JESÚS MARÍA (Ord. N°1017-MML + Ord. N°1076-MML):\n{RIN_JESUS_MARIA}"
        normativa_note += (
            "\n\nCLAVES PARA JESÚS MARÍA: (1) ATN II — normativa MML directa. "
            "(2) Estacionamiento DIFERENCIADO: RDA = 1/1 viv; RDM = 1/1.5 viv (Ord. N°586-MDJM). "
            "(3) Av. Garzón (20m) activa bonificación: RDA lotes ≥450m² hasta 15p; RDM +1 piso. "
            "(4) Área mínima 3D = 75m². Frente a pasaje: máx 3 pisos."
        )
    elif "cercado" in distrito.lower() or ("lima" in distrito.lower() and "cercado" in distrito.lower()):
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA CERCADO DE LIMA (Ord. N°893-MML + N°946-MML + N°1229-MML + N°2635-MML):\n{RIN_CERCADO_LIMA}"
        normativa_note += (
            "\n\nCLAVES PARA CERCADO DE LIMA: "
            "(1) ATN II — administrado directamente por MML (GDU-SPHU o ICL). "
            "(2) Altura CM/CZ comercial = 1.5×(a+r); RDA multifamiliar = 7 pisos; RDM = 3-6 pisos. "
            "(3) EXCEPCIÓN CRÍTICA: lote CR ≥1,000 m² en Santa Beatriz/Parque La Reserva → hasta 20 pisos "
            "(Ord. N°946-MML Art. 3°, Plano PA-02), área libre 50%. "
            "(4) Estacionamiento MUY FAVORABLE: 1 est./3 viviendas (Ord. N°1229-MML) para CM y CZ. "
            "(5) Retiros: 3m calles / 5m avenidas (DA N°127-1983). "
            "(6) Áreas mínimas: 3D=75m², 2D=55m², 1D=40m²."
        )
    elif "san borja" in distrito.lower() or "sanborja" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA SAN BORJA (Ord. N°491-MSB + TUO D.Alc. N°002-2025-MSB-A):\n{RIN_SAN_BORJA}"
        normativa_note += (
            "\n\nCLAVES PARA SAN BORJA: "
            "(1) ATN III — reglamento propio MSB, 5 Áreas Diferenciadas (A/B/C/D/E). "
            "(2) Altura máxima GENERAL: 8 pisos = 25.5 ml. "
            "EXCEPCIÓN: lotes >600 m² frente a Av. Javier Prado → hasta 12 pisos = 37.5 ml. "
            "(3) Estacionamiento SEGÚN ÁREA: A y B = 2/viv; C = 1.5/viv; D = 1/viv + 5% visitas; E = 1/1.5 viv. "
            "(4) Área mínima vivienda (70% de unidades): A=140m² · B=100m² · C=90m² · D=80m² · E=70m². "
            "(5) Área libre: A=40% · B/C=35% · D/E=30% (−5% en esquina). "
            "(6) Retiros: 3m calles / 5m avenidas. "
            "(7) Identificar el Área Diferenciada del predio — es la clave de todos los parámetros."
        )
    elif "santa anita" in distrito.lower() or "santaanita" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA SANTA ANITA (Ord. N°1025-MML + Ord. N°341-MML):\n{RIN_SANTA_ANITA}"
        normativa_note += (
            "\n\nCLAVES PARA SANTA ANITA: "
            "(1) ATN I — distrito industrial/periférico, normativa MML base, sin reglamento premium propio. "
            "(2) Zona I3 (Gran Industria): lote ≥2,500 m², retiro Carretera Central 10m, estac. 1/6 personas, patio maniobras obligatorio. "
            "(3) Zona CM (Comercio Metropolitano reconvertido): altura = 1.5×(a+r), área libre 0% comercial / 40% residencial (lote >200m²), "
            "estac. 1/50m² comercial o 1/2 viv residencial. "
            "(4) CM permite 100% del lote para uso residencial (RDA compatible). "
            "(5) No hay zonas residenciales puras premium — el potencial está en reconversión CM→residencial."
        )
    elif "surco" in distrito.lower() or "santiago de surco" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA SANTIAGO DE SURCO (Ord. N°1076-MML + Ord. N°912-MML + D.A. N°04-2023-MSS):\n{RIN_SURCO}"
        normativa_note += (
            "\n\nCLAVES PARA SURCO: "
            "(1) Surco tiene DOS sectores: ATN II (Ord. N°1076-MML, zonas intermedias/Paseo de la República) "
            "y ATN III (D.A. N°04-2023-MSS, zonas premium: Monterrico, Camacho, Chacarilla). Verificar en CPU. "
            "(2) Zona CZ ATN II: altura = 1.5×(a+r), área libre 0% comercial / 40% residencial (lote >200m²), "
            "100% del lote puede ser residencial. "
            "(3) RDA frente a avenida >20m: altura = 1.5×(a+r). "
            "(4) Estacionamiento: 1/1.5 viv residencial · 1/50m² comercial. "
            "(5) Altura piso a piso máx 3.00 ml en multifamiliar (Ord. N°1076-MML Anexo N°04, Item A.9). "
            "(6) Área mínima 3D = 75 m². Retiros: 5m Paseo de la República / 3m calles / 0m algunas calles locales."
        )
    elif "surquillo" in distrito.lower():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA SURQUILLO (Ord. N°1076-MML + Ord. N°501-MDS + Ord. N°533-2023-MDS):\n{RIN_SURQUILLO}"
        normativa_note += (
            "\n\nCLAVES PARA SURQUILLO: "
            "(1) ATN II — Ord. N°1076-MML para zonificación, Ord. N°501-MDS para estacionamientos vigentes. "
            "(2) Zona CZ: altura = 1.5×(a+r) · retiro Av. Angamos Este y Av. Tomas Marsano = 5.00 ml. "
            "(3) 100% del lote CZ puede ser residencial (RDA/RDM compatible). "
            "(4) Estacionamiento residencial (Ord. N°501-MDS vigente desde 2022): verificar en CPU reciente. "
            "Referencia histórica Ord. N°391-MDS (derogada): 2-3D = 1/viv; 1D = 2 est./3 unidades. "
            "(5) Área libre: 0% comercial / según uso residencial compatible en pisos residenciales. "
            "(6) ⚠ CPUs anteriores a jun 2023 citan Ord. N°391-MDS derogada — estacionamiento desactualizado."
        )
    elif "villa el salvador" in distrito.lower() or "ves" == distrito.lower().strip():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA VILLA EL SALVADOR (Ord. N°933-MML y modificatorias):\n{RIN_VILLA_EL_SALVADOR}"
        normativa_note += (
            "\n\nCLAVES PARA VILLA EL SALVADOR: "
            "(1) ATN I — normativa base MML Ord. N°933-MML, sin reglamento premium propio. "
            "(2) Zona CZ: altura FIJA 4 pisos (NO fórmula 1.5(a+r)) · retiro 1.50 ml · 100% residencial permitido. "
            "(3) Compatible RDM únicamente (no RDA) — menor densidad que distritos intermedios. "
            "(4) ⚠ Suelo arenoso: Estudio Geotécnico OBLIGATORIO para edificaciones >3 pisos. "
            "(5) Zona Industrial (ZIVS): ~800,000 m² almacenes · estac. 1/6 trabajadores · patio de maniobras. "
            "(6) Estacionamiento comercial: 1/50 m² · residencial: según Ord. N°933-MML Anexo 2. "
            "(7) Doble zonificación CZ+I2 posible (Ord. N°2220-MML Sector V/Litoral) — verificar con Informe de Compatibilidad de Uso vía GDEL. "
            "(8) Oportunidad residencial: crecimiento sur Lima, factibilidad servicios en expansión — "
            "verificar factibilidad SEDAPAL y Luz del Sur por sector antes de análisis financiero."
        )
    elif "san juan de lurigancho" in distrito.lower() or "sjl" == distrito.lower().strip():
        normativa_note += f"\n\nNORMATIVA ESPECÍFICA SAN JUAN DE LURIGANCHO (Ord. N°933-MML + Ord. N°284-MDSJL):\n{RIN_SAN_JUAN_LURIGANCHO}"
        normativa_note += (
            "\n\nCLAVES PARA SAN JUAN DE LURIGANCHO: "
            "(1) ATN I — distrito más poblado de Lima (~1.2M hab.), mercado residencial masivo. "
            "(2) Zona I2: lote ≥1,000 m² · frente 20m · retiro 3m · altura según proyecto/entorno · estac. 1/6 personas. "
            "(3) Estudios OBLIGATORIOS I2: Impacto Vial + Impacto Ambiental + Seguridad Integral (NRE A.60). "
            "(4) Zona RDM: altura 1.5(a+r) en vías locales (Ord. N°284-MDSJL) · área libre 30-35%. "
            "(5) Vías principales: Av. Próceres de la Independencia, Av. Gran Chimú, Av. Wiese, Av. Canto Grande. "
            "(6) Conectividad: Línea 1 Metro (eje Próceres) — activa bonus de demanda en predios cercanos. "
            "(7) Mercado objetivo: segmento B/C · precios USD 1,200-1,800/m² · tipologías 2D y 3D."
        )

    # ── Áreas mínimas obligatorias por distrito ──────────────────────────────
    _dl = distrito.lower()
    if "san isidro" in _dl:
        _area_min_nota = ""  # ya incluido en instrucción 4 con ámbito
    elif "miraflores" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (Miraflores — confirmar zona en certificado):\n"
            "  RDMA puro:        3D=130m²  2D=110m²  1D=80m²  (máx 50% unidades con 1D)\n"
            "  CM compatible:    3D=120m²  2D=100m²  1D=80m²  (máx 35% unidades con 1D)\n"
            "USA EXACTAMENTE estas áreas — son el valor fijo de diseño, no un piso mínimo."
        )
    elif "la victoria" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (La Victoria — Ord. N°355-MLV Art. 22°):\n"
            "  3D=75m²  2D=55m²  1D=40m²\n"
            "USA EXACTAMENTE estas áreas — son el valor fijo de diseño, no un piso mínimo."
        )
    elif "cercado" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (Cercado de Lima — Ord. N°2635-MML):\n"
            "  3D=75m²  2D=55m²  1D=40m²\n"
            "USA EXACTAMENTE estas áreas — son el valor fijo de diseño, no un piso mínimo."
        )
    elif "jesus maria" in _dl or "jesús maría" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (Jesús María):\n"
            "  3D=75m²  2D y 1D: a definir en proyecto (sin mínimo normativo fijo confirmado).\n"
            "Usar A.020 RNE como base (40m² mínimo multifamiliar). Optimizar según mercado."
        )
    elif "lince" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (Lince — Ord. N°235-MDL) según zona del certificado:\n"
            "  CM + RDA:       3D=60m²  2D=55m²  1D=45m²\n"
            "  CZ (RDA/RDM):   3D=60m²  2D=55m²  1D=45m²\n"
            "  RDA pura (Ámbito A): 3D=85m²  2D=75m²  1D=45m²\n"
            "  RDM pura (Ámbito B): 3D=75m²  2D=65m²  1D=45m²\n"
            "USA EXACTAMENTE estas áreas — son el valor fijo de diseño, no un piso mínimo."
        )
    elif "san borja" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (San Borja — Ord. N°491-MSB) por Área Diferenciada:\n"
            "  AD-A: 140m²  AD-B: 100m²  AD-C: 90m²  AD-D: 80m²  AD-E: 70m²\n"
            "  Restricción: el 70% de las unidades debe cumplir el área mínima del AD.\n"
            "Determina el AD desde el certificado de parámetros. "
            "USA EXACTAMENTE estas áreas — son el valor fijo de diseño, no un piso mínimo."
        )
    elif "san juan de lurigancho" in _dl or _dl.strip() == "sjl":
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (San Juan de Lurigancho — Ord. N°284-MDSJL):\n"
            "  2D y 3D: ≥75m²  1D: sin mínimo normativo fijo (usar ≥40m² según A.020 RNE).\n"
            "USA EXACTAMENTE 75m² para 2D y 3D — valor fijo de diseño."
        )
    elif "surco" in _dl or "santiago de surco" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS OBLIGATORIAS (Surco — ATN II):\n"
            "  3D=75m²  2D y 1D: sin mínimo normativo fijo confirmado (usar ≥40m² A.020 RNE).\n"
            "USA EXACTAMENTE 75m² para 3D — valor fijo de diseño."
        )
    elif "surquillo" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS (Surquillo): sin mínimo normativo distrital confirmado.\n"
            "Usar A.020 RNE como base: 40m² multifamiliar. Optimizar según mercado."
        )
    elif "magdalena" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS (Magdalena del Mar — inferidas Ord. N°950-MML):\n"
            "Usar A.020 RNE como base: 40m² mínimo multifamiliar. Optimizar según mercado."
        )
    elif "villa el salvador" in _dl or "ves" == _dl.strip():
        _area_min_nota = (
            "ÁREAS MÍNIMAS (Villa El Salvador — ATN I): sin mínimo normativo distrital específico.\n"
            "Usar A.020 RNE: 40m² multifamiliar. Mercado objetivo segmento C/D: tipologías 2D."
        )
    elif "santa anita" in _dl:
        _area_min_nota = (
            "ÁREAS MÍNIMAS (Santa Anita): sin mínimo normativo residencial específico confirmado.\n"
            "Usar A.020 RNE: 40m² mínimo multifamiliar."
        )
    else:
        _area_min_nota = (
            "ÁREAS MÍNIMAS: aplica A.020 RNE como base (40m² mínimo multifamiliar). "
            "Respeta los parámetros del certificado adjunto si especifica áreas mínimas."
        )

    prompt = f"""Eres un arquitecto especialista en desarrollo inmobiliario en Lima, Perú, con conocimiento experto de normativa municipal y el RNE.

PARÁMETROS NORMATIVOS (extraídos del Certificado de Parámetros):
{json.dumps(params_cabida, ensure_ascii=False, indent=2)}
{beneficios_txt}{colind_txt}{sugerencias_txt}{normativa_note}

CONFIGURACIÓN BASE:
- Uso: Residencial Multifamiliar (+ comercio en 1er piso si zonificación CZ/CM/CV)
- Sótanos para estacionamientos: determinar según normativa de estacionamientos del distrito

INSTRUCCIONES:
1. USA ÚNICAMENTE los parámetros explícitamente certificados en el documento adjunto. El valor de pisos_max ya incorpora la regla de colindancia calculada externamente — NO lo recalcules. NO apliques beneficios adicionales por cuenta propia (frente a parque, lote esquina, zonificación especial por cuadrante, etc.) aunque los identifiques en la normativa. Si detectas beneficios potenciales que podrían aplicar según la ubicación o normativa del distrito, agrégalos en "observaciones" con el texto: "BENEFICIO POTENCIAL — requiere verificación con la municipalidad antes de incorporar al proyecto: [descripción del beneficio]".
1b. RESTRICCIÓN GEOMÉTRICA (si geo_at_max > 0): el área techada total sobre rasante NO puede superar {st.session_state.get('geo_at_max', 0):,.0f} m² (calculado geométricamente desde las medidas reales del lote y los retiros). Si el cálculo normativo arroja más área, usa el valor geométrico como techo duro.
2. Calcula el área techada por piso: (area_terreno − área_perdida_retiro_frontal) × (1 − area_libre_min/100). IMPORTANTE: en Lima la práctica constructiva estándar es retiro lateral = 0 y retiro posterior = 0 (construcción pared con pared al límite de propiedad). Solo aplica el retiro frontal indicado en el certificado.
3. Si zona CZ/CV/CM y área libre = 0: área techada por piso ≈ área del lote − retiros
4. MAXIMIZA EL NÚMERO DE DEPARTAMENTOS. Para cada piso calcula cuántas unidades caben: num_unidades_piso = floor(area_techada_piso_vendible / area_promedio_del_mix). Para lotes ≤800 m² → mínimo 2 unidades/piso; para lotes >800 m² → mínimo 3 unidades/piso. VERIFICACIÓN OBLIGATORIA antes de devolver el JSON: Σ(cantidad_i × area_m2_i) debe ser ≈ area_vendible_m2 con tolerancia ±8%. Si no cuadra, ajusta las cantidades hacia arriba.
   REGLA CRÍTICA DE ÁREAS: USA EXACTAMENTE el área mínima normativa de cada tipología como valor fijo de diseño — NO como piso mínimo. area_m2 en el JSON debe ser IGUAL al mínimo normativo, no mayor. Unidades más grandes = menos departamentos = análisis incorrecto. La única forma de maximizar unidades es usar exactamente las áreas mínimas.
   {"" if "san isidro" not in _dl else f"""SAN ISIDRO (Anexo N°03 Ord. 523-MSI) — Ámbito: {ambito or 'verificar en certificado'}:
   - Ámbito A: area_m2 EXACTO → 3D=200m²  2D=150m²  1D=100m²
   - Ámbito B: area_m2 EXACTO → 3D=150m²  2D=120m²  1D=90m²
   - Ámbito C: area_m2 EXACTO → 3D=130m²  2D=110m²  1D=80m²
   - Ámbito D/CF: area_m2 EXACTO → 3D=110m²  2D=90m²  1D=70m²
   Mix obligatorio: 3D entre 50-100%, 2D máx 50%, 1D máx 20%."""}
   {_area_min_nota}
5. Calcula la eficiencia vendible: AV/AT (área vendible / área techada total sobre rasante) objetivo 75-80%; AV/AT total (incluyendo sótanos) = 60-74% en proyectos reales de Lima. Incluye en observaciones la eficiencia AV/AT total calculada.
   área_comunes_m2 = circulaciones, escaleras, lobby, cuartos de servicio por piso. En Lima promedio real = 8-12% del área techada sobre rasante (ref. validada: Los Fresnos 345 San Isidro = 8.9%, Clemente X = 10.2%). NO uses 15-25% — sobreestima el área no vendible y distorsiona los indicadores.
6. Calcula estacionamientos según normativa vigente con esta lógica EXACTA:
   estac_residentes = total_unidades × ratio_norma (usar el ratio que indica el CPU: San Isidro Ámbito A/B/C = 2 est./viv., Ámbito D/CF = 1 est./viv.; otros distritos según RNE o CPU adjunto).
   estac_visitas: aplica EXACTAMENTE lo que indica el parámetro estac_visitas_norma del CPU. {estac_visitas_instruccion}
   estac_total = estac_residentes + estac_visitas.
   CÁLCULO DE SÓTANOS — área techada por cochera = 25 m² (cajón + circulación directa, validado Clemente X 335: 900m²/36 coch.).
   area_neta_sotano = area_terreno × 0.87 (sótanos excavan el área total del lote; el 13% descuenta rampa de acceso, muros perimetrales y cuarto de máquinas).
   cocheras_por_sotano = floor(area_neta_sotano / 25).
   num_sotanos = ceil(estac_total / cocheras_por_sotano).
   IMPORTANTE: num_sotanos debe ser el número REAL necesario para alojar físicamente estac_total cocheras — no uses 1 o 2 por defecto sin verificar la capacidad.
7. NO apliques beneficios normativos adicionales que no estén explícitamente indicados en el certificado de parámetros adjunto. Si en la normativa del distrito o en los documentos detectas beneficios potenciales (por cuadrante de ubicación, frente a parque, lote esquina, zona especial, etc.), repórtalos ÚNICAMENTE en el campo "observaciones" con el prefijo "BENEFICIO POTENCIAL — requiere verificación". El usuario y su equipo técnico/legal decidirán si los incorporan.
8. DÚPLEX: PROHIBIDO incluir dúplex salvo instrucción explícita del usuario en NOTAS DEL ANALISTA. Si el JSON de unidades devuelve algún dúplex sin que el usuario lo haya pedido, el análisis es incorrecto. El objetivo es maximizar departamentos estándar (1D/2D/3D). Si se solicitan: cada dúplex ocupa la mitad del área del último piso + zona de azotea/terraza en nivel superior (zona techada superior ≤ 50% del piso inferior, RNE A.010 Art. 9); incluirlos como tipología "Dúplex" con area_m2 = área del piso/2 + zona techada superior.
9. Si es San Isidro: existe la posibilidad normativa de uso de azotea bajo régimen de propiedad exclusiva del último piso (30% del área utilizable después de retranques). NO lo apliques en el análisis — inclúyelo en observaciones como "BENEFICIO POTENCIAL — azotea exclusiva último piso: requiere verificación con Municipalidad de San Isidro antes de incorporar al proyecto".
10. Calcula depositos_total: número de depósitos/bodegas de almacenamiento del proyecto. En Lima el mercado compra depósitos como adicional, pero no todas las unidades incluyen uno. Usa: 40-50% de total_unidades en distritos premium (San Isidro, Miraflores, Barranco); 25-35% en distritos medios (San Borja, Surco, Jesús María, Magdalena, Lince, La Molina, San Miguel); 10-20% en distritos periféricos (SJL, VES, Santa Anita, La Victoria, Cercado); 0 si el mercado objetivo es económico. Redondea al entero más cercano.

Devuelve SOLO este JSON:
{{
  "area_techada_piso_m2": number,
  "area_techada_total_m2": number,
  "area_vendible_m2": number,
  "num_pisos": number,
  "num_sotanos": number,
  "unidades": [
    {{"tipo": "1 Dorm.", "cantidad": number, "area_m2": number, "area_total_m2": number}},
    {{"tipo": "2 Dorm.", "cantidad": number, "area_m2": number, "area_total_m2": number}},
    {{"tipo": "3 Dorm.", "cantidad": number, "area_m2": number, "area_total_m2": number}},
    {{"tipo": "Dúplex", "cantidad": number, "area_m2": number, "area_total_m2": number}}
  ],
  "total_unidades": number,
  "depositos_total": number,
  "estac_residentes": number,
  "estac_visitas": number,
  "estac_total": number,
  "area_libre_m2": number,
  "area_comunes_m2": number,
  "beneficios_aplicados": [{{"beneficio": "string", "impacto": "string"}}],
  "ordenanzas_mayor_altura": ["string"],
  "observaciones": ["string"],
  "metodologia": "string"
}}"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        system=[
            {
                "type": "text",
                "text": (
                    "Eres un arquitecto experto en normativa urbanística de Lima, Perú. "
                    "Respondes únicamente con JSON válido, sin texto adicional.\n\n"
                    f"REGLAMENTO NACIONAL DE EDIFICACIONES (RNE):\n{RNE_NACIONAL}\n\n"
                    f"MARCO NORMATIVO LIMA METROPOLITANA Y DISTRITOS:\n{REFERENCIAS_NORMATIVAS_LIMA}"
                ),
                "cache_control": {"type": "ephemeral"},
            }
        ],
        messages=[{"role": "user", "content": prompt}]
    )

    if not response.content:
        raise ValueError("json_parse_error: API devolvió respuesta vacía")
    text = response.content[0].text.strip()
    cab = parse_json_safe(text)

    # ── POST-PROCESO: corregir inconsistencias del output de la IA ──────────
    # 1. area_vendible = suma real de áreas de tipologías (ground truth)
    #    La IA a veces reporta area_vendible inconsistente con las unidades listadas.
    _sum_tip = sum(float(u.get("area_total_m2") or 0) for u in (cab.get("unidades") or []))
    if _sum_tip > 0:
        _ac = float(cab.get("area_comunes_m2") or 0)
        cab["area_vendible_m2"]    = round(_sum_tip, 1)
        cab["area_techada_total_m2"] = round(_sum_tip + _ac, 1)

    # 2. num_sotanos: recalcular con fórmula determinística (validada Clemente X 335)
    #    area_neta_sotano = area_terreno × 0.87  (sótanos excavan lote completo)
    #    cocheras_por_sotano = floor(area_neta / 25)
    _area_t  = float(params.get("area_terreno_m2") or 0)
    _estac_t = int(cab.get("estac_total") or 0)
    if _area_t > 0 and _estac_t > 0:
        _coch_x_sot = max(1, int(_area_t * 0.87 / 25))
        _sot_correcto = -(-_estac_t // _coch_x_sot)   # ceil sin importar
        if cab.get("num_sotanos", 0) != _sot_correcto:
            _sot_orig = cab.get("num_sotanos", "?")
            cab["num_sotanos"] = _sot_correcto
            _obs = cab.get("observaciones") or []
            _obs.append(
                f"Sótanos corregidos: {_area_t:.0f}m²×0.87/25m²/coch"
                f"={_coch_x_sot} coch/sótano → ceil({_estac_t}/{_coch_x_sot})="
                f"{_sot_correcto} (IA reportó {_sot_orig})"
            )
            cab["observaciones"] = _obs

    return cab


def analizar_legal(partida_bytes: bytes | None, puhr_bytes: bytes | None,
                   cert_params_bytes: bytes | None = None, planos_bytes: bytes | None = None,
                   sugerencias: str = "") -> dict:
    client = get_client()

    content = []
    docs_desc = []
    if partida_bytes:
        content.append(pdf_block(partida_bytes))
        docs_desc.append("Documento 1: Partida Registral (SUNARP)")
    if puhr_bytes:
        content.append(pdf_block(puhr_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: PU/HR (Predio Urbano / Hoja de Resumen - SAT/Municipalidad)")
    if cert_params_bytes:
        content.append(pdf_block(cert_params_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: Certificado de Parámetros Urbanísticos")
    if planos_bytes:
        content.append(pdf_block(planos_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: Planos del Inmueble")

    # Sistema estático (cacheado) — contiene el checklist completo y el schema JSON
    _legal_system = """Eres un abogado especialista en derecho registral e inmobiliario peruano y urbanista con expertise en normativa de Lima.
Respondes ÚNICAMENTE con JSON válido, sin texto antes ni después.

════════════════════════════════════════════════
REVISIÓN OBLIGATORIA — CHECKLIST DE 20 PUNTOS
════════════════════════════════════════════════
Revisa CADA UNO sin excepción. Para cada punto indica hallazgo preciso, severidad y — si es amarillo o rojo — procedimiento accionable para subsanarlo.

── PARTIDA REGISTRAL (Puntos 1–12) ──────────────
1. TITULARIDAD: Propietario(s) registrado(s) vs. vendedor declarado. ¿Coinciden exactamente (nombre + DNI/RUC)?
2. CADENA DE TITULARIDAD: Transferencias sucesivas. ¿Hay saltos, inscripciones incompletas o transferencias sin sustento?
3. HIPOTECAS: Cada hipoteca — acreedor, monto, fecha inscripción, asiento, estado (vigente/cancelada). Si cancelada, ¿figura el asiento de cancelación?
4. EMBARGOS: Embargos civiles, tributarios (SUNAT/MEF ejecutor coactivo) o penales. Estado actual.
5. MEDIDAS CAUTELARES: Tipo (inhibición, anotación demanda, etc.), expediente judicial, juzgado, estado.
6. SERVIDUMBRES: Tipo (paso, vista, acueducto, luz, etc.), predio dominante/sirviente, carácter (perpetuo/temporal/oneroso).
7. RESTRICCIONES DE DISPOSICIÓN: Cláusulas que limiten venta, arrendamiento, hipoteca u otro acto de disposición.
8. ANOTACIONES PREVENTIVAS: Vigentes (<3 años desde inscripción) vs. caducadas. Riesgo de cada una.
9. DOBLE INMATRICULACIÓN: Indicios de superposición con otro predio registrado o inconsistencias en linderos/ubicación.
10. ÁREA REGISTRAL: Área en partida. ¿Concuerda con el área ingresada por el usuario? Cuantificar discrepancia.
11. RÉGIMEN DE PROPIEDAD: Individual, copropiedad (% y DNI de cada copropietario), sociedad conyugal, persona jurídica.
12. ANTIGÜEDAD DE LA PARTIDA: Fecha de última actualización/impresión. Días transcurridos. < 30 días ideal; > 90 días riesgo.

── CERTIFICADO DE PARÁMETROS URBANÍSTICOS (Puntos 13–20) ──
13. ZONIFICACIÓN: Zonificación certificada vs. uso propuesto. ¿Compatible? ¿Requiere cambio de zonificación o ITT?
14. CUS MÁXIMO: Coeficiente de Utilización del Suelo. Valor exacto. ¿El proyecto lo respeta?
15. COS / ÁREA LIBRE MÍNIMA: Porcentaje de área libre exigido. ¿El diseño lo cumple?
16. ALTURA MÁXIMA: Pisos máximos o metros permitidos. ¿El proyecto los respeta?
17. RETIROS: Dimensiones exactas (frente, laterales izquierdo/derecho, posterior).
18. ESTACIONAMIENTOS MÍNIMOS: Ratio exigido por normativa distrital.
19. COMPATIBILIDAD DE USO: ¿El uso declarado está expresamente permitido según el certificado?
20. VIGENCIA DEL CERTIFICADO: Fecha de emisión. Los certificados tienen vigencia de 36 meses en Lima.

CRITERIOS DE SEVERIDAD:
- "verde": hallazgo favorable, no requiere acción
- "amarillo": observación menor; gestionar antes del cierre
- "rojo": riesgo crítico que bloquea o condiciona gravemente la operación
- "no_verificable": documento no adjuntado o información ilegible/ausente

SUBSANACIÓN (solo amarillo/rojo): procedimiento ESPECÍFICO Y ACCIONABLE paso a paso — qué gestionar, ante qué entidad, plazo estimado, condición para cierre.

IMPORTANTE — DNI: Extrae exactamente como figura. Si no figura, null. Nunca inferir ni inventar.
IMPORTANTE — PU/HR: Extrae autoavalúo, código predio/contribuyente, clasificación municipal, condición propietario (SAT o Municipalidad).

SCHEMA JSON DE RESPUESTA (devuelve EXACTAMENTE esta estructura):
{
  "propietarios_partida": [{"nombre": "...", "dni": "8 dígitos o RUC 11 dígitos o null", "porcentaje": "50% o null", "tipo_doc": "DNI/RUC/null"}],
  "propietarios_puhr": [{"nombre": "...", "dni": "8 dígitos o null", "condicion": "Propietario/Poseedor/null"}],
  "propietarios_coinciden": true,
  "diferencias_propietarios": null,
  "direccion_partida": null,
  "direccion_puhr": null,
  "direcciones_coinciden": null,
  "diferencias_direccion": null,
  "area_registral_m2": null,
  "area_puhr_m2": null,
  "areas_coinciden": null,
  "discrepancia_area_m2": null,
  "partida_numero": null,
  "numero_predio": null,
  "valor_autoavaluo": null,
  "moneda_autoavaluo": "PEN",
  "anio_autoavaluo": null,
  "clasificacion_municipal": null,
  "condicion_propietario_sat": null,
  "uso_predio": null,
  "cargas_vigentes": [],
  "hipotecas_vigentes": [],
  "medidas_cautelares": [],
  "anotaciones_diversas": [],
  "semaforo": "verde/amarillo/rojo",
  "alertas": [],
  "resumen_legal": "2-3 oraciones resumiendo el estado legal.",
  "hallazgos": [
    {"numero": 1, "punto": "Titularidad registral", "categoria": "partida", "hallazgo": "...", "severidad": "verde", "subsanacion": null},
    {"numero": 2, "punto": "Cadena de titularidad", "categoria": "partida", "hallazgo": "...", "severidad": "verde/amarillo/rojo/no_verificable", "subsanacion": "procedimiento o null"}
  ],
  "completitud": {"verificados": 18, "total": 20, "no_verificables": []}
}

REGLA: "hallazgos" debe tener EXACTAMENTE 20 objetos en orden 1–20. Puntos 13–20 usan "categoria": "parametros". Si el certificado no fue adjuntado → severidad "no_verificable".
SEMÁFORO: VERDE = todos verdes, sin cargas activas · AMARILLO = uno o más amarillos, ningún rojo · ROJO = cualquier punto rojo"""

    # Mensaje dinámico — solo los datos específicos del análisis
    cross_note = ("Compara y cruza la información entre todos los documentos disponibles."
                  if len(docs_desc) > 1
                  else "Extrae toda la información relevante del documento disponible.")
    sug_note = f"\n\nInstrucciones adicionales del analista:\n{sugerencias.strip()}" if sugerencias and sugerencias.strip() else ""
    user_prompt = (f"Analiza los siguientes documentos de un inmueble en Lima, Perú:\n"
                   f"{chr(10).join(docs_desc)}\n\n{cross_note}{sug_note}")
    content.append({"type": "text", "text": user_prompt})

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        system=[{"type": "text", "text": _legal_system, "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": content}]
    )
    if not response.content:
        raise ValueError("json_parse_error: API devolvió respuesta vacía")
    return parse_json_safe(response.content[0].text.strip())


def _extraer_texto_zip(file_bytes: bytes, filename: str) -> str:
    """Extract plain text from DOCX or PPTX (both are ZIP-based XML)."""
    import zipfile, io, re as _re
    ext = filename.lower().rsplit(".", 1)[-1]
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            if ext == "docx":
                xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
                return _re.sub(r"<[^>]+>", " ", xml)
            elif ext in ("pptx", "ppt"):
                slides = sorted(n for n in z.namelist()
                                if n.startswith("ppt/slides/slide") and n.endswith(".xml"))
                parts = []
                for s in slides:
                    xml = z.read(s).decode("utf-8", errors="ignore")
                    parts.append(_re.sub(r"<[^>]+>", " ", xml))
                return "\n".join(parts)
    except Exception:
        pass
    return ""


def extraer_datos_desde_doc(file_bytes: bytes, filename: str, modulo: str) -> dict:
    """Send document to Claude and extract structured property data for the given module."""
    client = get_client()
    ext = filename.lower().rsplit(".", 1)[-1]

    content: list = []
    if ext == "pdf":
        content.append(pdf_block(file_bytes))
    else:
        texto = _extraer_texto_zip(file_bytes, filename)
        if not texto.strip():
            return {"_error": "No se pudo extraer texto del documento."}
        content.append({"type": "text", "text": f"Contenido del documento:\n\n{texto[:15000]}"})

    PROMPTS = {
        "cabida": """Eres un analista inmobiliario. Del documento adjunto extrae los datos del proyecto o terreno.
Devuelve ÚNICAMENTE este JSON, sin texto adicional:
{
  "nombre_proyecto": "nombre del proyecto o null",
  "distrito": "distrito de Lima o null",
  "area_terreno_m2": null,
  "frente_ml": null,
  "fondo_ml": null,
  "costo_terreno_usd": null,
  "precio_venta_m2_usd": null,
  "costo_construccion_m2_usd": null,
  "num_pisos": null,
  "zonificacion": "RDA/RDM/RDB/CM/I2/otro o null"
}
Si un dato no aparece usa null. Solo extrae lo que está explícito.""",

        "residencial": """Eres un analista inmobiliario. Del documento adjunto extrae los datos del inmueble residencial.
Devuelve ÚNICAMENTE este JSON, sin texto adicional:
{
  "distrito": "distrito de Lima o null",
  "area_m2": null,
  "antiguedad_anios": null,
  "dormitorios": "1 Dormitorio / 2 Dormitorios / 3 Dormitorios / Dúplex / Otro — o null",
  "precio_usd": null,
  "alquiler_mes_usd": null,
  "descripcion": "breve descripción del inmueble o null"
}
Si un dato no aparece usa null. Solo extrae lo que está explícito.""",

        "industrial": """Eres un analista inmobiliario. Del documento adjunto extrae los datos del inmueble o proyecto industrial/logístico.
Devuelve ÚNICAMENTE este JSON, sin texto adicional:
{
  "distrito": "distrito de Lima o null",
  "area_terreno_m2": null,
  "frente_ml": null,
  "fondo_ml": null,
  "costo_terreno_usd": null,
  "tipo_nave": "Almacén Logístico / Nave Industrial / Cross-docking / Producción - Manufactura o null",
  "area_nave_m2": null,
  "renta_usd_m2_mes": null,
  "zonificacion": "I1/I2/I3/OU o null"
}
Si un dato no aparece usa null. Solo extrae lo que está explícito.""",
    }

    content.append({"type": "text", "text": PROMPTS.get(modulo, PROMPTS["residencial"])})

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": content}]
    )
    if not response.content:
        return {"_error": "Sin respuesta de la API."}
    return parse_json_safe(response.content[0].text.strip())


def extraer_precios_cierre(partidas_bytes: list) -> list:
    """Extrae precio de última compraventa inscrita de cada partida SUNARP."""
    client = get_client()
    resultados = []
    for i, pdf_bytes in enumerate(partidas_bytes):
        content = [
            pdf_block(pdf_bytes),
            {"type": "text", "text": """Eres un experto en derecho registral peruano.
Analiza esta partida registral SUNARP y extrae los datos de la ÚLTIMA compraventa (o transferencia onerosa) inscrita.

Devuelve ÚNICAMENTE este JSON, sin texto adicional:
{
  "partida_numero": "número de partida o null",
  "descripcion_predio": "dirección o descripción breve del inmueble",
  "area_m2": null,
  "ultima_transferencia": {
    "precio": null,
    "moneda": "USD o PEN o null",
    "fecha": "dd/mm/aaaa o null",
    "vendedor": "nombre completo o null",
    "comprador": "nombre completo o null",
    "tipo_acto": "Compraventa/Dación en pago/Adjudicación/Anticipo/otro"
  },
  "precio_m2_estimado": null,
  "observaciones": "nota si el precio no está explícito, si hay varios asientos, o si algo es relevante para la tasación"
}

Si el precio no figura explícitamente usa null en precio y explícalo en observaciones.
Si hay varias transferencias reporta solo la más reciente por fecha de inscripción."""},
        ]
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": content}]
        )
        if not response.content:
            raise ValueError("json_parse_error: API devolvió respuesta vacía en comparable")
        r = parse_json_safe(response.content[0].text.strip())
        r["_idx"] = i + 1
        resultados.append(r)
    return resultados


def analizar_factibilidad_industrial(
    partida_bytes: bytes | None,
    cert_params_bytes: bytes | None,
    cert_zon_bytes: bytes | None,
    tipo_nave: str = "Almacén Logístico",
    zonificacion_ref: str = "I2",
    uso: str = "Uso directo",
    planos_bytes: bytes | None = None,
    sugerencias: str = "",
) -> dict:
    client = get_client()

    content = []
    docs_desc = []
    if partida_bytes:
        content.append(pdf_block(partida_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: Partida Registral (SUNARP)")
    if cert_params_bytes:
        content.append(pdf_block(cert_params_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: Certificado de Parámetros Urbanísticos")
    if cert_zon_bytes:
        content.append(pdf_block(cert_zon_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: Certificado de Zonificación y Vías")
    if planos_bytes:
        content.append(pdf_block(planos_bytes))
        docs_desc.append(f"Documento {len(docs_desc)+1}: Planos del Inmueble")

    # Sistema estático (cacheado) — checklist + schema
    _ind_system = """Eres un especialista en derecho inmobiliario, derecho registral y normativa urbanística industrial peruana.
Respondes ÚNICAMENTE con JSON válido, sin texto antes ni después.

════════════════════════════════════════════════
REVISIÓN OBLIGATORIA — CHECKLIST DE 16 PUNTOS
════════════════════════════════════════════════
Revisa CADA punto sin excepción. Para amarillo/rojo: indica procedimiento específico y accionable para subsanar.

── ANÁLISIS TÉCNICO — CERTIFICADO (Puntos 1–8, categoria "tecnico") ──
1. ZONIFICACIÓN CERTIFICADA: Zona exacta según certificado vs. zonificación declarada por usuario. ¿Coinciden?
2. COMPATIBILIDAD DE ACTIVIDAD: ¿La actividad del proyecto está permitida, condicionada o prohibida según el índice de usos?
3. ALTURA MÁXIMA DE NAVE: Restricción de altura en metros según certificado o normativa aplicable.
4. RETIROS Y ALINEAMIENTO: Retiros frente, laterales y posterior. ¿Hay restricciones de alineamiento vial?
5. VÍAS DE FRENTE Y ACCESO PESADO: Nombre, tipo y ancho de vía(s) de frente. ¿Permiten maniobra de tráileres (>10m)?
6. ÁREA LIBRE Y COS INDUSTRIAL: Porcentaje de área libre o COS máximo aplicable a uso industrial.
7. CONDICIONANTES ESPECIALES: Restricciones ambientales, sanitarias, de seguridad industrial, zonas de amortiguamiento.
8. VIGENCIA DEL CERTIFICADO: Fecha de emisión. ¿Está dentro de los 36 meses de vigencia?

── ANÁLISIS LEGAL — PARTIDA REGISTRAL (Puntos 9–16, categoria "legal") ──
9. TITULARIDAD: Propietario(s) registrado(s) — nombre y DNI/RUC. ¿Coincide con quien ofrece el inmueble?
10. CADENA DE TITULARIDAD: Transferencias sucesivas. ¿Hay saltos o inscripciones incompletas?
11. HIPOTECAS: Cada hipoteca — acreedor, monto, fecha, asiento, estado vigente/cancelada.
12. EMBARGOS Y MEDIDAS CAUTELARES: Embargos civiles, tributarios (SUNAT), penales, inhibiciones, anotaciones.
13. SERVIDUMBRES Y RESTRICCIONES: Servidumbres de paso, restricciones de disposición o cláusulas especiales.
14. ÁREA REGISTRAL: Área en partida vs. área declarada para el proyecto. Discrepancia si existe.
15. RÉGIMEN DE PROPIEDAD: Individual, copropiedad (% y DNI), persona jurídica (poderes vigentes).
16. ANTIGÜEDAD DE LA PARTIDA: Días desde última actualización. < 30 días ideal; > 90 días riesgo.

CRITERIOS DE SEVERIDAD:
- "verde": sin observación, no requiere acción
- "amarillo": observación menor; gestionar antes del cierre
- "rojo": riesgo crítico; bloquea o condiciona gravemente la operación
- "no_verificable": documento no adjuntado o información ilegible

SUBSANACIÓN (solo amarillo/rojo): procedimiento paso a paso — qué gestionar, ante qué entidad, plazo estimado.

SCHEMA JSON DE RESPUESTA (devuelve EXACTAMENTE esta estructura):
{
  "docs_analizados": [],
  "semaforo_tecnico": "verde/amarillo/rojo",
  "semaforo_legal": "verde/amarillo/rojo",
  "semaforo_global": "verde/amarillo/rojo",
  "zonificacion_certificada": null,
  "compatible_actividad": null,
  "nota_compatibilidad": "explicación detallada",
  "actividades_permitidas": [],
  "actividades_condicionadas": [],
  "actividades_prohibidas": [],
  "restricciones_altura_m": null,
  "restricciones_especiales": [],
  "vias_frente": [{"nombre": "...", "tipo": "arterial/colectora/local", "ancho_ml": null}],
  "acceso_vehiculos_pesados": null,
  "area_registral_m2": null,
  "alertas_tecnicas": [],
  "alertas_legales": [],
  "propietarios_partida": [],
  "direccion_partida": null,
  "partida_numero": null,
  "cargas_vigentes": [],
  "hipotecas_vigentes": [],
  "medidas_cautelares": [],
  "resumen_tecnico": "2-3 oraciones sobre compatibilidad y restricciones técnicas.",
  "resumen_legal": "2-3 oraciones sobre estado registral y alertas legales.",
  "hallazgos": [
    {"numero": 1, "punto": "Zonificación certificada", "categoria": "tecnico", "hallazgo": "...", "severidad": "verde", "subsanacion": null}
  ],
  "completitud": {"verificados": 14, "total": 16, "no_verificables": []}
}

REGLA: "hallazgos" EXACTAMENTE 16 objetos en orden numérico. Puntos 1–8: categoria "tecnico". Puntos 9–16: categoria "legal".
SEMÁFORO TÉCNICO: VERDE=compatible sin restricciones críticas · AMARILLO=condicionado o trámite · ROJO=incompatible o sin acceso pesado
SEMÁFORO LEGAL: VERDE=sin cargas activas, propietarios claros · AMARILLO=cargas canceladas, info parcial · ROJO=hipotecas/embargos vigentes, cautelares, titularidad en cuestión
SEMÁFORO GLOBAL: el más restrictivo entre técnico y legal. Si un dato no existe usa null."""

    # Mensaje dinámico — datos específicos del proyecto
    sug_note_ind = f"\n\nInstrucciones adicionales del analista:\n{sugerencias.strip()}" if sugerencias and sugerencias.strip() else ""
    user_prompt_ind = (
        f"Analiza los siguientes documentos de un inmueble industrial/logístico en Lima, Perú:\n"
        f"{chr(10).join(docs_desc)}\n\n"
        f"Contexto del proyecto:\n"
        f"- Actividad a desarrollar: {tipo_nave}\n"
        f"- Zonificación de referencia declarada: {zonificacion_ref}\n"
        f"- Propósito: {uso}\n\n"
        f"Aplica los dos análisis (A) Factibilidad Técnica y (B) Legal al 100% de los documentos adjuntos.{sug_note_ind}"
    )
    content.append({"type": "text", "text": user_prompt_ind})

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        system=[{"type": "text", "text": _ind_system, "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": content}]
    )
    if not response.content:
        raise ValueError("json_parse_error: API devolvió respuesta vacía")
    return parse_json_safe(response.content[0].text.strip())


# ═══════════════════════════════════════════════════════
# MODELO FINANCIERO
# ═══════════════════════════════════════════════════════

def calcular_financiero(cabida: dict, fin: dict, zona: str) -> dict:
    m  = MERCADO.get(zona, {})
    av = cabida.get("area_vendible_m2", 0)
    at = cabida.get("area_techada_total_m2", 0)

    # ── Ingresos ────────────────────────────────────────
    precio_m2    = fin.get("precio_venta_m2") if fin.get("precio_venta_m2", 0) > 0 else m.get("precio_2br", 0)
    ing_dptos    = av * (precio_m2 or 0)
    _pe = fin.get("precio_estac") or m.get("precio_estac", 0)
    _pd = fin.get("precio_deposito") or m.get("precio_deposito", 0)
    ing_estac    = cabida.get("estac_residentes", 0) * _pe
    ing_deposito = cabida.get("depositos_total", 0) * _pd
    ing_brutos   = ing_dptos + ing_estac + ing_deposito

    # ── Terreno ─────────────────────────────────────────
    c_terreno_base  = fin.get("costo_terreno", 0) or 0
    c_alcabala      = c_terreno_base * 0.03 if fin.get("include_alcabala", True) else 0
    c_notarial      = c_terreno_base * 0.003     # Gastos notariales: 0.3%
    c_registral     = c_terreno_base * 0.0015    # Gastos registrales: 0.15%
    c_due_dilig     = 10000 if fin.get("include_dd", True) else 0
    c_demolicion    = fin.get("costo_demolicion", 0) or 0
    c_terreno_total = c_terreno_base + c_alcabala + c_notarial + c_registral + c_due_dilig + c_demolicion

    # ── Construcción ────────────────────────────────────
    c_obra_dptos   = at * (fin.get("costo_construccion", 0) or 0)
    c_sotanos_area = cabida.get("estac_total", 0) * 25   # 25 m² área techada por cochera (validado: Clemente X 335, 900m²/36 coch.)
    c_obra_sotanos = c_sotanos_area * fin.get("costo_sotano_m2", 450)
    c_constructora = (c_obra_dptos + c_obra_sotanos) * fin.get("fee_constructora", 10.0) / 100
    c_construccion = c_obra_dptos + c_obra_sotanos + c_constructora

    # ── Costos Inmobiliarios — calibrado Clemente X 335 (UP Advisory Board 2025) ─
    c_arq          = at * (fin.get("costo_arq_m2") or 5.94)    # $5.94/m² → $15,000 en 2,525m² (ref. NÓMENA)
    c_esp          = at * (fin.get("costo_esp_m2") or 7.92)    # $7.92/m² → $20,000 en 2,525m² (sanitario+eléct.+gas)
    c_factib       = fin.get("costo_factibilidades") or 17000  # Estudios pre-proyecto (≠ DD terreno)
    c_supervision  = c_construccion * 0.005      # Supervisión técnica: 0.5% costo directo
    c_costos_base  = c_terreno_total + c_construccion + c_arq + c_esp
    c_legales      = c_costos_base * 0.005       # Legales y contabilidad: 0.5%
    c_permisos        = c_construccion * 0.015    # Permisos y licencias: 1.5% costo construcción
    c_gerenciamiento  = c_construccion * 0.05    # Gerenciamiento inmobiliario: 5% costo construcción
    c_ventas_marketing = ing_brutos * 0.05       # Ventas y marketing consolidado: 5% ingresos

    # ── Financiamiento (75% del costo construcción × tasa × período obra) ─
    # Refleja línea de crédito constructor estándar Lima: ~75% del costo de obra
    _n_pisos_prelim  = cabida.get("num_pisos", 7)
    _meses_obra_prel = 24 if _n_pisos_prelim > 20 else (12 if _n_pisos_prelim <= 5 else 16)
    c_financiero   = c_construccion * 0.75 * fin.get("tasa_financ", 9.0) / 100 * (_meses_obra_prel / 12)

    # ── Totales ─────────────────────────────────────────
    c_sin_financ = (c_terreno_total + c_construccion + c_arq + c_esp + c_factib +
                    c_supervision + c_legales + c_permisos + c_gerenciamiento + c_ventas_marketing)
    c_total      = c_sin_financ + c_financiero

    utilidad_bruta = ing_brutos - c_total
    tasa_ir        = max(fin.get("tasa_ir", 29.5), 0) / 100
    c_ir           = utilidad_bruta * tasa_ir if utilidad_bruta > 0 else 0
    utilidad_neta  = utilidad_bruta - c_ir

    margen_bruto   = (utilidad_bruta / ing_brutos * 100) if ing_brutos else 0
    margen_neto    = (utilidad_neta  / ing_brutos * 100) if ing_brutos else 0
    roi            = (utilidad_neta  / c_total      * 100) if c_total else 0

    # Escenario sin financiamiento
    util_sin_f      = ing_brutos - c_sin_financ
    ir_sin_f        = util_sin_f * tasa_ir if util_sin_f > 0 else 0
    util_neta_sin_f = util_sin_f - ir_sin_f
    margen_sin_f    = (util_neta_sin_f / ing_brutos * 100) if ing_brutos else 0

    # Métricas estratégicas
    be_precio_m2    = round(c_sin_financ / av) if av > 0 else 0
    _otros_costos = (c_construccion + c_arq + c_esp + c_factib + c_supervision + c_legales
                     + c_permisos + c_gerenciamiento + c_ventas_marketing
                     + c_due_dilig + c_notarial + c_registral + c_demolicion)
    _ir_factor      = max(1 - tasa_ir, 0.50)
    max_terreno_20  = max(0, round(ing_brutos * (1 - 0.20 / _ir_factor) - _otros_costos))
    max_terreno_15  = max(0, round(ing_brutos * (1 - 0.15 / _ir_factor) - _otros_costos))
    max_terreno_12  = max(0, round(ing_brutos * (1 - 0.12 / _ir_factor) - _otros_costos))
    ratio_terreno   = round(c_terreno_base / ing_brutos * 100, 1) if ing_brutos > 0 else 0

    # TIT: Tasa de Incidencia del Terreno (professional KPI)
    tit_pct         = round(c_terreno_base / ing_brutos * 100, 1) if ing_brutos > 0 else 0

    vel             = m.get("velocidad_venta", 1.0)
    n_unidades      = cabida.get("total_unidades", 0)
    n_pisos         = cabida.get("num_pisos", 7)
    _obra_auto      = 24 if n_pisos > 20 else (12 if n_pisos <= 5 else 16)
    meses_obra      = int(fin.get("meses_obra_override") or _obra_auto)
    meses_obra      = max(1, min(meses_obra, 60))
    _meses_venta_raw = round(n_unidades / vel) if (vel and n_unidades) else 0
    # Cap plazo de ventas al window real del proyecto: obra + 6 m post-obra
    meses_venta      = min(_meses_venta_raw, meses_obra + 6)
    # Preventa → obra → post-obra ventas (máx 6 m adicionales)
    _post_obra_sales = min(6, max(0, meses_venta - meses_obra))
    meses_proyecto   = 2 + meses_obra + _post_obra_sales

    if c_total > 0 and utilidad_neta > 0 and meses_proyecto > 0:
        # Factor 0.65: capital no está 100% inmovilizado desde el día 0 (se desembolsa en S-curve
        # durante obra y se recupera gradualmente en ventas — período efectivo ≈ 65% del total)
        tir_aprox = round(((1 + utilidad_neta / c_total) ** (12 / (meses_proyecto * 0.65)) - 1) * 100, 1)
    else:
        tir_aprox = 0.0

    return {
        "resumen": {
            "departamentos":        cabida.get("total_unidades", 0),
            "estacionamientos":     cabida.get("estac_total", 0),
            "m2_construibles":      round(at),
            "m2_vendibles":         round(av),
            "ingresos_brutos":      round(ing_brutos),
            "costo_total_sin_financ": round(c_sin_financ),
            "costo_total":          round(c_total),   # alias para compatibilidad
            "utilidad_bruta":       round(utilidad_bruta),
            "margen_bruto_pct":     round(margen_bruto, 1),
            "ir_pct":               round(tasa_ir * 100, 1),
            "costo_ir":             round(c_ir),
            "utilidad_neta":        round(utilidad_neta),
            "margen_pct":           round(margen_neto, 1),
            "utilidad_neta_sin_f":  round(util_neta_sin_f),
            "margen_sin_f_pct":     round(margen_sin_f, 1),
            "costo_financiero":     round(c_financiero),
            "costo_total_con_financ": round(c_total),
            "utilidad_con_financ":  round(utilidad_neta),
            "margen_con_financ_pct": round(margen_neto, 1),
            "roi_pct":              round(roi, 1),
            "tir_anual_pct":        tir_aprox,
            "be_precio_m2":         be_precio_m2,
            "max_terreno_20pct":    max_terreno_20,
            "max_terreno_15pct":    max_terreno_15,
            "max_terreno_12pct":    max_terreno_12,
            "ratio_terreno_pct":    ratio_terreno,
            "tit_pct":              tit_pct,
            "meses_obra":           meses_obra,
            "meses_venta":          meses_venta,
            "meses_venta_calc":     _meses_venta_raw,
            "meses_proyecto":       meses_proyecto,
        },
        "detalle_ingresos": {
            "Departamentos":        round(ing_dptos),
            "Estacionamientos":     round(ing_estac),
            "Depósitos":            round(ing_deposito),
            "TOTAL INGRESOS":       round(ing_brutos),
        },
        "detalle_costos": {
            "── TERRENO ──────────────────────": 0,
            "Precio del terreno":              round(c_terreno_base),
            "Alcabala (3%)":                   round(c_alcabala),
            "Gastos notariales (0.3%)":        round(c_notarial),
            "Gastos registrales (0.15%)":      round(c_registral),
            "Due diligence":                   round(c_due_dilig),
            **({"Demolición": round(c_demolicion)} if c_demolicion > 0 else {}),
            "── CONSTRUCCIÓN ─────────────────": 0,
            "Obra civil":                      round(c_obra_dptos),
            "Sótanos":                         round(c_obra_sotanos),
            f"Fee constructora ({(fin.get('fee_constructora') or 10):.0f}%)": round(c_constructora),
            "── COSTOS TÉCNICOS ───────────────": 0,
            f"Arquitectura (${fin.get('costo_arq_m2') or 5.94:.2f}/m²)":  round(c_arq),
            f"Especialidades (${fin.get('costo_esp_m2') or 7.92:.2f}/m²)": round(c_esp),
            "Supervisión técnica (0.5%)":        round(c_supervision),
            "Permisos y licencias (1.5%)":       round(c_permisos),
            "── COSTOS INMOBILIARIOS ──────────": 0,
            "Gerenciamiento (5% construcción)": round(c_gerenciamiento),
            "Ventas y marketing (5%)":          round(c_ventas_marketing),
            "Legales y contabilidad (0.5%)":    round(c_legales),
            "SUBTOTAL SIN FINANCIAMIENTO":      round(c_sin_financ),
            "Gasto financiero (banco)":         round(c_financiero),
            "TOTAL EGRESOS":                    round(c_total),
            "── RESULTADO ────────────────────": 0,
            f"IR ({(fin.get('tasa_ir') or 29.5):.1f}%)": round(c_ir),
        },
        "_raw": {
            "c_terreno_base":    c_terreno_base,
            "c_alcabala":        c_alcabala,
            "c_notarial":        c_notarial,
            "c_registral":       c_registral,
            "c_due_dilig":       c_due_dilig,
            "c_terreno_total":   c_terreno_total,
            "c_obra_dptos":      c_obra_dptos,
            "c_obra_sotanos":    c_obra_sotanos,
            "c_constructora":    c_constructora,
            "c_arq":             c_arq,
            "c_esp":             c_esp,
            "c_supervision":     c_supervision,
            "c_legales":           c_legales,
            "c_permisos":          c_permisos,
            "c_gerenciamiento":    c_gerenciamiento,
            "c_ventas_marketing":  c_ventas_marketing,
            "c_financiero":        c_financiero,
            "c_ir":              c_ir,
            "ing_brutos":        ing_brutos,
        },
    }


# ═══════════════════════════════════════════════════════
# GENERADOR DE REPORTE EXCEL
# ═══════════════════════════════════════════════════════

def generar_excel_factis(result: dict, cabida: dict, params: dict,
                         fin_inputs: dict, zona: str) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter

    GOLD   = "B8904A"
    DARK   = "0A1628"
    LIGHT  = "F5F2ED"
    WHITE  = "FFFFFF"
    GRAY   = "E8E4DC"

    def _hdr_fill(ws, row, col, value, bg=DARK, fg=WHITE, bold=True, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=bold, color=fg, size=size)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            bottom=Side(style="thin", color=GOLD),
            right=Side(style="thin", color="DDDDDD"))
        return c

    def _val_cell(ws, row, col, value, fmt=None, bold=False, bg=WHITE):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=10, color=DARK)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = Border(right=Side(style="thin", color="DDDDDD"),
                          bottom=Side(style="thin", color="EEEEEE"))
        if fmt:
            c.number_format = fmt
        return c

    def _lbl_cell(ws, row, col, value, bold=False, bg=LIGHT):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=10, color=DARK)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(right=Side(style="thin", color="DDDDDD"),
                          bottom=Side(style="thin", color="EEEEEE"))
        return c

    wb  = Workbook()
    r   = result.get("resumen", {})
    det = result.get("detalle_costos", {})
    ing = result.get("detalle_ingresos", {})

    # ── Hoja 1: Resumen Ejecutivo ────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    ws1.row_dimensions[1].height = 36

    # Título
    t = ws1.cell(row=1, column=1, value=f"FACTIS — Reporte Financiero · {zona}")
    t.font = Font(bold=True, size=14, color=GOLD)
    t.fill = PatternFill("solid", fgColor=DARK)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws1.merge_cells("A1:C1")
    ws1.cell(row=2, column=1, value=f"Proyecto: {fin_inputs.get('nombre_proyecto', zona)}").font = Font(size=10, color="888888")
    ws1.cell(row=2, column=2, value=f"Zona: {zona}").font = Font(size=10, color="888888")
    ws1.cell(row=2, column=3, value=f"Fecha: {datetime.date.today().strftime('%d/%m/%Y')}").font = Font(size=10, color="888888")

    # Cabeceras bloque KPIs
    _hdr_fill(ws1, 4, 1, "INDICADOR")
    _hdr_fill(ws1, 4, 2, "VALOR (USD)")
    _hdr_fill(ws1, 4, 3, "REFERENCIA")

    kpis = [
        ("Ingresos Brutos",     r.get("ingresos_brutos", 0),      "$#,##0",  "100%"),
        ("Costo Total",         r.get("costo_total_con_financ", r.get("costo_total_sin_financ", 0)), "$#,##0",  f"{100-r.get('margen_bruto_pct',0):.1f}% ing."),
        ("Utilidad Bruta",      r.get("utilidad_bruta", 0),        "$#,##0",  f"{r.get('margen_bruto_pct',0):.1f}% bruto"),
        ("Impuesto a la Renta", r.get("costo_ir", 0),             "$#,##0",  f"{r.get('ir_pct',29.5):.1f}%"),
        ("Utilidad Neta",       r.get("utilidad_neta", 0),        "$#,##0",  f"{r.get('margen_pct',0):.1f}% neto"),
        (None, None, None, None),
        ("TIR Anual (aprox.)",  r.get("tir_anual_pct", 0),        "0.0\"%\"", "Objetivo >15%"),
        ("ROI",                 r.get("roi_pct", 0),               "0.0\"%\"", "Utilidad/Costo"),
        ("TIT (terreno/ingr.)", r.get("tit_pct", 0),              "0.0\"%\"", "Ideal <20%"),
        ("Margen bruto",        r.get("margen_bruto_pct", 0),     "0.0\"%\"", "Pre-IR"),
        ("Margen neto",         r.get("margen_pct", 0),           "0.0\"%\"", "Post-IR"),
    ]
    for i, row_data in enumerate(kpis, start=5):
        if row_data[0] is None:
            for col in range(1, 4):
                ws1.cell(row=i, column=col).fill = PatternFill("solid", fgColor=GRAY)
            continue
        lbl, val, fmt, ref = row_data
        _lbl_cell(ws1, i, 1, lbl, bold=(lbl in ("Utilidad Neta", "TIR Anual (aprox.)")))
        _val_cell(ws1, i, 2, val, fmt=fmt, bold=(lbl in ("Utilidad Neta",)))
        ws1.cell(row=i, column=3, value=ref).font = Font(size=9, color="888888", italic=True)

    # Cabida resumen
    ws1.cell(row=17, column=1, value="CABIDA").font = Font(bold=True, size=11, color=GOLD)
    cabida_kpis = [
        ("Unidades totales",  cabida.get("total_unidades", 0)),
        ("Pisos",             cabida.get("num_pisos", 0)),
        ("Área techada (m²)", cabida.get("area_techada_total_m2", 0)),
        ("Área vendible (m²)",cabida.get("area_vendible_m2", 0)),
        ("Estacionamientos",  cabida.get("estac_residentes", 0)),
        ("Depósitos",         cabida.get("depositos_total", 0)),
    ]
    for i, (lbl, val) in enumerate(cabida_kpis, start=18):
        _lbl_cell(ws1, i, 1, lbl)
        _val_cell(ws1, i, 2, val, fmt="#,##0")

    # ── Hoja 2: Estructura de Costos ────────────────────
    ws2 = wb.create_sheet("Estructura de Costos")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 14

    _hdr_fill(ws2, 1, 1, "RUBRO DE COSTO")
    _hdr_fill(ws2, 1, 2, "MONTO (USD)")
    _hdr_fill(ws2, 1, 3, "% TOTAL")
    total_cost = r.get("costo_total_con_financ", r.get("costo_total_sin_financ", 1)) or 1
    for i, (lbl, val) in enumerate(det.items(), start=2):
        bg = LIGHT if i % 2 == 0 else WHITE
        _lbl_cell(ws2, i, 1, lbl, bg=bg)
        _val_cell(ws2, i, 2, val if val else None, fmt="$#,##0", bg=bg)
        _pct = round((val or 0) / total_cost * 100, 1) if total_cost and val else None
        _val_cell(ws2, i, 3, _pct, fmt='0.0"%"', bg=bg)
    # Total
    n = len(det) + 2
    _lbl_cell(ws2, n, 1, "TOTAL COSTOS", bold=True, bg=GRAY)
    _val_cell(ws2, n, 2, sum(det.values()), fmt="$#,##0", bold=True, bg=GRAY)
    _val_cell(ws2, n, 3, 100.0, fmt='0.0"%"', bold=True, bg=GRAY)

    # ── Hoja 3: Ingresos ────────────────────────────────
    ws3 = wb.create_sheet("Ingresos")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 14
    _hdr_fill(ws3, 1, 1, "FUENTE DE INGRESO")
    _hdr_fill(ws3, 1, 2, "MONTO (USD)")
    _hdr_fill(ws3, 1, 3, "% TOTAL")
    total_ing = r.get("ingresos_brutos", 1) or 1
    for i, (lbl, val) in enumerate(ing.items(), start=2):
        bg = LIGHT if i % 2 == 0 else WHITE
        _lbl_cell(ws3, i, 1, lbl, bg=bg)
        _val_cell(ws3, i, 2, val, fmt="$#,##0", bg=bg)
        _pct3 = round((val or 0) / total_ing * 100, 1) if total_ing and val else None
        _val_cell(ws3, i, 3, _pct3, fmt='0.0"%"', bg=bg)
    n3 = len(ing) + 2
    _lbl_cell(ws3, n3, 1, "TOTAL INGRESOS", bold=True, bg=GRAY)
    _val_cell(ws3, n3, 2, sum(ing.values()), fmt="$#,##0", bold=True, bg=GRAY)
    _val_cell(ws3, n3, 3, 100.0, fmt='0.0"%"', bold=True, bg=GRAY)

    # ── Hoja 4: Parámetros del Terreno ──────────────────
    ws4 = wb.create_sheet("Parámetros")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 24
    _hdr_fill(ws4, 1, 1, "PARÁMETRO")
    _hdr_fill(ws4, 1, 2, "VALOR")
    param_rows = [
        ("Ubicación",            params.get("ubicacion", zona)),
        ("Distrito",             zona),
        ("Área del terreno",     f"{params.get('area_terreno_m2','—')} m²"),
        ("Frente",               f"{params.get('frente_ml','—')} ml"),
        ("Zonificación",         params.get("zonificacion", "—")),
        ("Pisos máx.",           params.get("pisos_max", "—")),
        ("Área libre mín.",      params.get("area_libre_min", "—")),
        ("Precio terreno",       f"${fin_inputs.get('costo_terreno',0):,.0f}"),
        ("Precio venta m²",      f"${fin_inputs.get('precio_venta_m2',0):,.0f}"),
        ("Costo construcción m²",f"${fin_inputs.get('costo_construccion',0):,.0f}"),
    ]
    for i, (lbl, val) in enumerate(param_rows, start=2):
        bg = LIGHT if i % 2 == 0 else WHITE
        _lbl_cell(ws4, i, 1, lbl, bg=bg)
        _val_cell(ws4, i, 2, val, bg=bg)
        ws4.cell(row=i, column=2).alignment = Alignment(horizontal="left")

    # ── Hoja 5: Sensibilidad Precio × Terreno ─────────────────────
    ws5 = wb.create_sheet("Sensibilidad")
    _sens_res = calcular_sensibilidad_terreno(cabida, fin_inputs, zona)
    if _sens_res:
        _df_mg, _df_tir, _s_precios, _s_terrenos, _s_p0, _s_t0 = _sens_res
        ws5.column_dimensions["A"].width = 18
        # Título
        _tc = ws5.cell(row=1, column=1, value="MATRIZ ESTRATÉGICA — PRECIO DE VENTA × PRECIO DEL TERRENO")
        _tc.font = Font(bold=True, size=11, color=WHITE)
        _tc.fill = PatternFill("solid", fgColor=DARK)
        _tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_s_precios)+1)
        ws5.row_dimensions[1].height = 24
        ws5.cell(row=2, column=1, value="Celda: Margen neto % · Verde ≥18% · Amarillo 12–18% · Rojo <12%").font = Font(size=8, italic=True, color="888888")
        ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_s_precios)+1)
        # Cabecera de columnas (precio de venta)
        _hdr_fill(ws5, 3, 1, "Terreno ↓  /  Precio m² →", bg=DARK, fg="B8904A")
        for ci, p in enumerate(_s_precios):
            _is_base = abs(p - _s_p0) == min(abs(x - _s_p0) for x in _s_precios)
            _bg = "1E3A5A" if _is_base else DARK
            _hdr_fill(ws5, 3, ci+2, f"${p:,}/m²", bg=_bg)
            ws5.column_dimensions[get_column_letter(ci+2)].width = 13
        # Filas de datos
        for ri, t in enumerate(_s_terrenos):
            _is_base_row = abs(t - _s_t0) == min(abs(x - _s_t0) for x in _s_terrenos)
            _rh_bg = "1E3A5A" if _is_base_row else DARK
            _rh_fg = "B8904A" if _is_base_row else WHITE
            _hdr_fill(ws5, ri+4, 1, f"${t:,.0f}", bg=_rh_bg, fg=_rh_fg)
            for ci in range(len(_s_precios)):
                mg  = float(_df_mg.iloc[ri, ci])
                tir = float(_df_tir.iloc[ri, ci])
                _is_base_col = abs(_s_precios[ci] - _s_p0) == min(abs(x - _s_p0) for x in _s_precios)
                cell = ws5.cell(row=ri+4, column=ci+2,
                                value=f"{mg:.0f}% mg / {tir:.0f}% TIR")
                if mg >= 18:
                    _fc, _bc = "1B5E20", "C8E6C9"
                elif mg >= 12:
                    _fc, _bc = "7A5500", "FFF9C4"
                else:
                    _fc, _bc = "B71C1C", "FFCDD2"
                cell.fill = PatternFill("solid", fgColor=_bc)
                cell.font = Font(bold=(_is_base_row and _is_base_col), color=_fc, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if _is_base_row and _is_base_col:
                    from openpyxl.styles import Border as XlBorder, Side as XlSide
                    _gold_side = XlSide(style="medium", color="B8904A")
                    cell.border = XlBorder(top=_gold_side, bottom=_gold_side,
                                           left=_gold_side, right=_gold_side)
        ws5.row_dimensions[3].height = 18
        for ri in range(len(_s_terrenos)):
            ws5.row_dimensions[ri+4].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════
# GENERADOR DE REPORTE PDF
# ═══════════════════════════════════════════════════════

def generar_pdf_factis(result: dict, cabida: dict, params: dict,
                       fin_inputs: dict, zona: str,
                       legal: dict | None = None) -> bytes:
    """Genera el reporte PDF ejecutivo de Factis."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm, cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable, KeepTogether, PageBreak,
                                    NextPageTemplate, BaseDocTemplate, Frame, PageTemplate)
    from reportlab.pdfgen import canvas as pdfgen_canvas
    from reportlab.graphics.shapes import Drawing, Rect, String, Line, Group
    from reportlab.graphics import renderPDF

    # ── Paleta ──────────────────────────────────────────────────
    NAV   = colors.HexColor("#1E2D3D")
    GOLD  = colors.HexColor("#B8904A")
    GRN   = colors.HexColor("#1A4731")
    GRN_L = colors.HexColor("#E8F5EE")
    AMB   = colors.HexColor("#7A5500")
    AMB_L = colors.HexColor("#FFF8E6")
    RED   = colors.HexColor("#7A1A1A")
    RED_L = colors.HexColor("#FDECEA")
    GREY  = colors.HexColor("#7A7268")
    LGREY = colors.HexColor("#F5F2ED")
    BORD  = colors.HexColor("#D8D4CC")
    WHITE = colors.white

    W, H = A4
    M = 20 * mm  # márgenes

    r   = result.get("resumen", {})
    det = result.get("detalle_costos", {})
    ing = result.get("detalle_ingresos", {})

    today  = datetime.date.today().strftime("%d/%m/%Y")
    distrito = zona
    direccion = params.get("ubicacion") or params.get("direccion") or "—"

    # ── Clasificación viabilidad ─────────────────────────────────
    _mg  = r.get("margen_pct", 0)
    _tir = r.get("tir_anual_pct", 0)
    _tit = r.get("tit_pct", 0)
    _nombre_proy = fin_inputs.get("nombre_proyecto", "") or zona
    if _mg >= 20 and _tir >= 15:
        _perfil_txt, _perfil_col = "RETORNOS SÓLIDOS", GRN
    elif _mg > 0 and _tir > 0:
        _perfil_txt, _perfil_col = "RETORNOS MODERADOS", AMB
    else:
        _perfil_txt, _perfil_col = "RETORNOS NEGATIVOS", RED

    # ── Valor terreno clasificación ─────────────────────────────
    _pc = fin_inputs.get("costo_terreno", 0)
    _v20 = r.get("max_terreno_20pct", 0)
    _v15 = r.get("max_terreno_15pct", 0)
    _v12 = r.get("max_terreno_12pct", 0)
    if _pc <= _v20:
        _zona_v, _zona_c = "ZONA ÓPTIMA", GRN
    elif _pc <= _v15:
        _zona_v, _zona_c = "ZONA ACEPTABLE", AMB
    elif _pc <= _v12:
        _zona_v, _zona_c = "ZONA DE RIESGO", RED
    else:
        _zona_v, _zona_c = "PRECIO ELEVADO", RED

    # ── Buffer y canvas ─────────────────────────────────────────
    buf = BytesIO()

    HEADER_H = 16 * mm   # altura reservada para el encabezado en páginas internas

    def _content_page(canvas_obj, doc):
        canvas_obj.saveState()
        # Banda dorada lateral (igual que portada)
        canvas_obj.setFillColor(GOLD)
        canvas_obj.rect(0, 0, 6 * mm, H, fill=1, stroke=0)
        # Franja de encabezado (fondo blanco implícito — solo línea dorada inferior)
        canvas_obj.setStrokeColor(GOLD)
        canvas_obj.setLineWidth(1.0)
        canvas_obj.line(6 * mm, H - HEADER_H, W, H - HEADER_H)
        # Mini-wordmark FACTIS
        canvas_obj.setFillColor(NAV)
        canvas_obj.setFont("Helvetica-Bold", 11)
        canvas_obj.drawString(M + 2 * mm, H - HEADER_H * 0.55, "FACTIS")
        # Sub-etiqueta
        canvas_obj.setFillColor(GREY)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.drawString(M + 2 * mm + 46, H - HEADER_H * 0.55,
                              "·  Osterling Advisory")
        # Pie de página — línea separadora
        canvas_obj.setStrokeColor(colors.HexColor("#E8E0D4"))
        canvas_obj.setLineWidth(0.3)
        canvas_obj.line(M, 15 * mm, W - M, 15 * mm)
        # Pie de página — identificación (encima de la línea)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(GREY)
        canvas_obj.drawString(M, 11.5 * mm,
            f"FACTIS — Análisis de Cabida y Factibilidad Financiera  ·  {today}")
        canvas_obj.drawRightString(W - M, 11.5 * mm,
            f"Preparado por Osterling Advisory  ·  Pág. {doc.page}")
        # Pie de página — disclaimer IA (debajo de la línea)
        canvas_obj.setFont("Helvetica", 5.5)
        canvas_obj.setFillColor(colors.HexColor("#A89880"))
        _disclaimer = (
            "NOTA: Esta IA de Análisis Inmobiliario debe utilizarse como herramienta complementaria al criterio profesional, "
            "permitiendo obtener resultados preliminares de manera rápida. El profesional podrá definir tipologías, "
            "distribución por plantas y modificaciones pertinentes. La IA irá alineándose con la visión del profesional."
        )
        canvas_obj.drawCentredString(W / 2, 7.5 * mm, _disclaimer)
        canvas_obj.restoreState()

    def _cover_page(canvas_obj, doc):
        # Fondo navy completo
        canvas_obj.setFillColor(NAV)
        canvas_obj.rect(0, 0, W, H, fill=1, stroke=0)
        # Banda gold lateral
        canvas_obj.setFillColor(GOLD)
        canvas_obj.rect(0, 0, 6 * mm, H, fill=1, stroke=0)
        # Wordmark FACTIS
        canvas_obj.setFillColor(WHITE)
        canvas_obj.setFont("Helvetica-Bold", 52)
        canvas_obj.drawString(M + 6 * mm, H - 55 * mm, "FACTIS")
        # Línea dorada
        canvas_obj.setStrokeColor(GOLD)
        canvas_obj.setLineWidth(1.2)
        canvas_obj.line(M + 6 * mm, H - 63 * mm, W - M, H - 63 * mm)
        # Subtítulo
        canvas_obj.setFont("Helvetica", 13)
        canvas_obj.setFillColor(colors.HexColor("#B8C8D8"))
        canvas_obj.drawString(M + 6 * mm, H - 72 * mm,
                              "Análisis de Cabida y Factibilidad Financiera")
        # Datos del proyecto
        canvas_obj.setFillColor(GOLD)
        canvas_obj.setFont("Helvetica-Bold", 10)
        canvas_obj.drawString(M + 6 * mm, H * 0.52, "PROYECTO")
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.setFillColor(WHITE)
        canvas_obj.drawString(M + 6 * mm, H * 0.52 - 14, f"Proyecto:   {_nombre_proy}")
        canvas_obj.drawString(M + 6 * mm, H * 0.52 - 27, f"Distrito:    {distrito}")
        canvas_obj.drawString(M + 6 * mm, H * 0.52 - 40, f"Dirección:  {direccion}")
        canvas_obj.drawString(M + 6 * mm, H * 0.52 - 53, f"Fecha:       {today}")
        # KPIs grandes en portada
        _kpis = [
            ("MARGEN NETO", f"{_mg:.1f}%"),
            ("TIR ANUAL",   f"{_tir:.1f}%"),
            ("ROI",         f"{r.get('roi_pct',0):.1f}%"),
        ]
        _kx = M + 6 * mm
        for _lbl, _val in _kpis:
            canvas_obj.setFillColor(colors.HexColor("#2A3D4D"))
            canvas_obj.rect(_kx, H * 0.25, 52 * mm, 28 * mm, fill=1, stroke=0)
            canvas_obj.setFillColor(GOLD)
            canvas_obj.setFont("Helvetica-Bold", 7)
            canvas_obj.drawString(_kx + 4 * mm, H * 0.25 + 22 * mm, _lbl)
            canvas_obj.setFillColor(WHITE)
            canvas_obj.setFont("Helvetica-Bold", 22)
            canvas_obj.drawString(_kx + 4 * mm, H * 0.25 + 9 * mm, _val)
            _kx += 56 * mm
        # Perfil de inversión
        canvas_obj.setFillColor(colors.HexColor("#2A3D4D"))
        canvas_obj.rect(M + 6 * mm, H * 0.15, 160 * mm, 18 * mm, fill=1, stroke=0)
        canvas_obj.setFillColor(GOLD)
        canvas_obj.setFont("Helvetica-Bold", 7)
        canvas_obj.drawString(M + 10 * mm, H * 0.15 + 12 * mm, "PERFIL DE INVERSIÓN")
        canvas_obj.setFillColor(WHITE)
        canvas_obj.setFont("Helvetica-Bold", 14)
        canvas_obj.drawString(M + 10 * mm, H * 0.15 + 4 * mm, _perfil_txt)
        canvas_obj.setFillColor(colors.HexColor("#8A9BAD"))
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.drawString(M + 70 * mm, H * 0.15 + 4 * mm,
                              f"TIT terreno: {_tit:.1f}% · Utilidad neta: {_fmt(r.get('utilidad_neta',0))}")
        # Footer portada
        canvas_obj.setFillColor(GREY)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.drawString(M + 6 * mm, 12 * mm,
                              "Preparado por Osterling Advisory  ·  factis.pe")
        canvas_obj.drawRightString(W - M, 12 * mm, "Confidencial")

    def _fmt(v):
        v = v or 0
        if abs(v) >= 1_000_000:
            return f"${v/1_000_000:.2f}M"
        return f"${v:,.0f}"

    # ── Estilos ──────────────────────────────────────────────────
    styles = getSampleStyleSheet()

    def _style(name, parent="Normal", **kw):
        s = ParagraphStyle(name, parent=styles[parent], **kw)
        return s

    S_TITLE   = _style("s_title",  fontSize=14, textColor=NAV,
                        fontName="Helvetica-Bold", spaceAfter=4)
    S_SECTION = _style("s_sec",    fontSize=8,  textColor=GOLD,
                        fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4,
                        letterSpacing=2)
    S_BODY    = _style("s_body",   fontSize=9,  textColor=NAV,
                        fontName="Helvetica", leading=13)
    S_SMALL   = _style("s_small",  fontSize=7.5, textColor=GREY,
                        fontName="Helvetica", leading=11)
    S_NUM     = _style("s_num",    fontSize=9,  textColor=NAV,
                        fontName="Helvetica-Bold", alignment=TA_RIGHT)
    S_CENTER  = _style("s_ctr",    fontSize=9,  textColor=NAV,
                        fontName="Helvetica", alignment=TA_CENTER)

    def _section(txt):
        return [Paragraph((txt or "").upper(), S_SECTION),
                HRFlowable(width="100%", thickness=0.5, color=BORD, spaceAfter=4)]

    def _kpi_table(items):
        """items = list of (label, value, ref) tuples — 4 columnas principales.
        Each KPI is a nested 3-row table to avoid paragraph overlap."""
        if not items:
            return Spacer(1, 1)
        col_w = (W - 2 * M) / len(items)
        inner_w = col_w - 8  # subtract left+right padding

        def _kpi_shorten(v_str):
            """$649,000 → $649K to avoid line wrap in narrow cells."""
            if v_str.startswith("$") and "," in v_str and not v_str.endswith("M"):
                try:
                    num = float(v_str[1:].replace(",", ""))
                    if abs(num) >= 100_000:
                        return f"${num / 1_000:.0f}K"
                except ValueError:
                    pass
            return v_str

        cells = []
        for i, (l, v, ref) in enumerate(items):
            v_disp = _kpi_shorten(v)
            p_lbl = Paragraph(l, _style(f"kl{i}", fontSize=7, fontName="Helvetica-Bold",
                                         textColor=GOLD, alignment=TA_CENTER, leading=9))
            p_val = Paragraph(v_disp, _style(f"kv{i}", fontSize=20, fontName="Helvetica-Bold",
                                              textColor=NAV, alignment=TA_CENTER, leading=24))
            p_ref = Paragraph(ref, _style(f"kr{i}", fontSize=7, fontName="Helvetica",
                                           textColor=GREY, alignment=TA_CENTER, leading=9))
            inner = Table(
                [[p_lbl], [p_val], [p_ref]],
                colWidths=[inner_w],
                rowHeights=[10, 22, 9],   # total 41pt — fits in 16mm cell
            )
            inner.setStyle(TableStyle([
                ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",   (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
                ("LEFTPADDING",  (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ]))
            cells.append(inner)

        t = Table([cells], colWidths=[col_w] * len(items), rowHeights=[16 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, -1), LGREY),
            ("GRID",         (0, 0), (-1, -1), 0.5, BORD),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    def _cost_table(rows_dict):
        """rows_dict from detalle_costos"""
        if not rows_dict:
            return Spacer(1, 1)
        data, styles_ts = [], []
        row_idx = 0
        for k, v in rows_dict.items():
            if k.startswith("──"):
                lbl = k.replace("──", "").replace("─", "").strip()
                data.append([Paragraph(lbl, _style(f"sh{row_idx}", fontSize=7.5,
                              textColor=GOLD, fontName="Helvetica-Bold")), ""])
                styles_ts += [
                    ("BACKGROUND",  (0, row_idx), (-1, row_idx), NAV),
                    ("TEXTCOLOR",   (0, row_idx), (-1, row_idx), GOLD),
                    ("TOPPADDING",  (0, row_idx), (-1, row_idx), 5),
                    ("BOTTOMPADDING",(0, row_idx),(-1, row_idx), 5),
                ]
            elif "TOTAL" in k or "SUBTOTAL" in k:
                data.append([Paragraph(f"<b>{k}</b>",
                              _style(f"st{row_idx}", fontSize=9, textColor=NAV,
                                     fontName="Helvetica-Bold")),
                             Paragraph(f"<b>{_fmt(v)}</b>",
                              _style(f"sv{row_idx}", fontSize=9, textColor=NAV,
                                     fontName="Helvetica-Bold", alignment=TA_RIGHT))])
                styles_ts += [
                    ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#E8EDF3")),
                    ("LINEABOVE",  (0, row_idx), (-1, row_idx), 0.8, NAV),
                ]
            else:
                data.append([Paragraph(k, S_BODY),
                             Paragraph(_fmt(v), S_NUM)])
                if row_idx % 2 == 0:
                    styles_ts.append(("BACKGROUND", (0, row_idx), (-1, row_idx),
                                      colors.HexColor("#FDFAF6")))
            row_idx += 1

        col_w = W - 2 * M
        t = Table(data, colWidths=[col_w * 0.68, col_w * 0.32])
        t.setStyle(TableStyle([
            ("GRID",         (0, 0), (-1, -1), 0.3, BORD),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ] + styles_ts))
        return t

    def _chart_financiero() -> Drawing:
        """Barras verticales: Ingresos / Costo / Utilidad Neta."""
        cw  = float(W - 2 * M)
        ch  = 58.0 * mm
        d   = Drawing(cw, ch)

        items = [
            ("Ingresos Brutos",  r.get("ingresos_brutos", 0), "#1A4731"),
            ("Costo Total",      r.get("costo_total_con_financ", r.get("costo_total_sin_financ", 0)), "#1E2D3D"),
            ("Utilidad Neta",    r.get("utilidad_neta",    0), "#B8904A"),
        ]
        n       = len(items)
        max_v   = max(v for _, v, _ in items) or 1
        pad_l   = 8.0
        pad_r   = 8.0
        axis_y  = 18.0
        bar_top = ch - 14.0
        bar_area_h = bar_top - axis_y
        total_w = cw - pad_l - pad_r
        bar_w   = total_w / (n * 2.2)
        spacing = (total_w - n * bar_w) / (n + 1)

        # fondo gris claro
        d.add(Rect(0, 0, cw, ch,
                   fillColor=colors.HexColor("#F5F2ED"), strokeColor=None))

        # línea base
        d.add(Line(pad_l, axis_y, cw - pad_r, axis_y,
                   strokeColor=colors.HexColor("#D8D4CC"), strokeWidth=0.6))

        for i, (lbl, val, hex_col) in enumerate(items):
            x  = pad_l + spacing + i * (bar_w + spacing)
            bh = max(0.0, (val / max_v) * bar_area_h) if max_v > 0 else 0
            fc = colors.HexColor(hex_col)

            # barra
            d.add(Rect(x, axis_y, bar_w, bh, fillColor=fc, strokeColor=None))

            # valor encima de la barra
            val_txt = _fmt(val)
            d.add(String(x + bar_w / 2, axis_y + bh + 3, val_txt,
                         fontName="Helvetica-Bold", fontSize=7.5,
                         fillColor=colors.HexColor("#1E2D3D"),
                         textAnchor="middle"))

            # etiqueta debajo del eje
            d.add(String(x + bar_w / 2, 5, lbl,
                         fontName="Helvetica", fontSize=7,
                         fillColor=colors.HexColor("#7A7268"),
                         textAnchor="middle"))

        # título del gráfico
        d.add(String(pad_l, ch - 10, "RESUMEN FINANCIERO",
                     fontName="Helvetica-Bold", fontSize=7,
                     fillColor=colors.HexColor("#B8904A"),
                     textAnchor="start"))
        return d

    def _chart_costos() -> Drawing:
        """Barras horizontales: top 8 rubros de costo."""
        _items = [(k, v) for k, v in det.items()
                  if "Impuesto" not in k and v and v > 0]
        _items.sort(key=lambda x: x[1], reverse=True)
        _items = _items[:8]
        if not _items:
            return Drawing(1, 1)

        cw     = float(W - 2 * M)
        row_h  = 9.0 * mm
        pad_t  = 10.0
        n      = len(_items)
        ch     = row_h * n + pad_t + 8.0
        d      = Drawing(cw, ch)

        # fondo
        d.add(Rect(0, 0, cw, ch,
                   fillColor=colors.HexColor("#F5F2ED"), strokeColor=None))

        # título
        d.add(String(4, ch - 8, "PRINCIPALES RUBROS DE COSTO",
                     fontName="Helvetica-Bold", fontSize=7,
                     fillColor=colors.HexColor("#B8904A"),
                     textAnchor="start"))

        lbl_w   = cw * 0.36
        bar_max = cw * 0.45
        val_x   = lbl_w + bar_max + 6
        max_v   = _items[0][1] or 1

        for i, (lbl, val) in enumerate(_items):
            y  = ch - pad_t - (i + 1) * row_h
            bw = (val / max_v) * bar_max

            # fila alterna
            bg = colors.HexColor("#ECEAE6") if i % 2 == 0 else colors.HexColor("#F5F2ED")
            d.add(Rect(0, y, cw, row_h, fillColor=bg, strokeColor=None))

            # etiqueta izquierda
            lbl_s = (lbl[:30] + "…") if len(lbl) > 30 else lbl
            d.add(String(6, y + row_h * 0.32, lbl_s,
                         fontName="Helvetica", fontSize=7.5,
                         fillColor=colors.HexColor("#2C2C2C"),
                         textAnchor="start"))

            # barra
            bar_y = y + row_h * 0.18
            bar_h = row_h * 0.60
            d.add(Rect(lbl_w, bar_y, bw, bar_h,
                       fillColor=colors.HexColor("#1E2D3D"), strokeColor=None))

            # valor a la derecha
            d.add(String(lbl_w + bw + 4, y + row_h * 0.32, _fmt(val),
                         fontName="Helvetica-Bold", fontSize=7.5,
                         fillColor=colors.HexColor("#1E2D3D"),
                         textAnchor="start"))

        # línea separadora título
        d.add(Line(0, ch - pad_t, cw, ch - pad_t,
                   strokeColor=colors.HexColor("#B8904A"), strokeWidth=0.6))
        return d

    # ── Documento con portada especial ───────────────────────────
    doc = BaseDocTemplate(buf, pagesize=A4,
                          leftMargin=M, rightMargin=M,
                          topMargin=M, bottomMargin=18 * mm)

    cover_frame   = Frame(0, 0, W, H, id="cover")
    content_frame = Frame(M, 18 * mm, W - 2 * M,
                          H - HEADER_H - 18 * mm, id="content")

    doc.addPageTemplates([
        PageTemplate(id="cover",   frames=[cover_frame],
                     onPage=_cover_page),
        PageTemplate(id="content", frames=[content_frame],
                     onPage=_content_page),
    ])

    story = []

    # ── Página 1: Portada (vacía — _cover_page la pinta) ────────
    story.append(NextPageTemplate("content"))
    story.append(PageBreak())

    # ── Página 2: Resumen Ejecutivo ───────────────────────────────
    story += _section("Resumen Ejecutivo")
    story.append(Paragraph(
        f"<b>Distrito:</b> {distrito} &nbsp;&nbsp; "
        f"<b>Dirección:</b> {direccion} &nbsp;&nbsp; "
        f"<b>Fecha:</b> {today}", S_BODY))
    story.append(Spacer(1, 8))

    # KPI panel 4 columnas
    story.append(_kpi_table([
        ("MARGEN NETO POST-IR",  f"{_mg:.1f}%",  "Ref. óptimo ≥ 20%"),
        ("TIR ANUAL ESTIMADA",   f"{_tir:.1f}%", "Ref. óptimo ≥ 15%"),
        ("ROI",                  f"{r.get('roi_pct',0):.1f}%", "Ref. óptimo ≥ 20%"),
        ("UTILIDAD NETA",        _fmt(r.get("utilidad_neta",0)), "post impuestos"),
    ]))
    story.append(Spacer(1, 8))

    # Perfil inversión + clasificación terreno (2 columnas)
    _perf_bg = {GRN: GRN_L, AMB: AMB_L, RED: RED_L, NAV: colors.HexColor("#EEF2F7")}
    _p_bg = _perf_bg.get(_perfil_col, LGREY)
    _cell_w2 = (W - 2 * M) * 0.5

    def _info_cell(label, value, sub, val_color=NAV):
        """Nested 3-row table for label / large-value / sub — spacing controlled via rowHeights."""
        _iw = _cell_w2 - 32   # inner width minus left+right padding
        _p_lbl = Paragraph(label, _style(f"il{label[:4]}", fontSize=7,
                            fontName="Helvetica-Bold", textColor=GOLD,
                            alignment=TA_CENTER, leading=9))
        _p_val = Paragraph(value, _style(f"iv{label[:4]}", fontSize=16,
                            fontName="Helvetica-Bold", textColor=val_color,
                            alignment=TA_CENTER, leading=20))
        _p_sub = Paragraph(sub,   _style(f"is{label[:4]}", fontSize=7,
                            fontName="Helvetica", textColor=GREY,
                            alignment=TA_CENTER, leading=9))
        _inner = Table([[_p_lbl], [_p_val], [_p_sub]],
                       colWidths=[_iw], rowHeights=[14, 24, 14])
        _inner.setStyle(TableStyle([
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        return _inner

    perf_cell = _info_cell(
        "PERFIL DE INVERSIÓN",
        _perfil_txt,
        f"TIT terreno: {_tit:.1f}% · Ref. óptima: 15–25%",
        val_color=_perfil_col,
    )
    zona_cell = _info_cell(
        "PRECIO TERRENO INGRESADO",
        _fmt(_pc),
        f"{_zona_v} · Óptimo: {_fmt(_v20)}",
        val_color=_zona_c,
    )
    t2 = Table([[perf_cell, zona_cell]],
               colWidths=[_cell_w2] * 2, rowHeights=[44 * mm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), _p_bg),
        ("BACKGROUND", (1, 0), (1, 0), LGREY),
        ("GRID",       (0, 0), (-1, -1), 0.5, BORD),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story.append(t2)
    story.append(Spacer(1, 8))

    # Indicadores secundarios
    _sec_data = [
        ["Ingresos brutos", _fmt(r.get("ingresos_brutos", 0)),
         "Meses de obra", f"{r.get('meses_obra',0)} meses"],
        ["Costo total s/financ.", _fmt(r.get("costo_total_sin_financ", 0)),
         "Meses de ventas", f"{r.get('meses_venta',0)} meses"],
        ["Costo financiero", _fmt(r.get("costo_financiero", 0)),
         "Duración total", f"{r.get('meses_proyecto',0)} meses"],
        [f"IR ({r.get('ir_pct',29.5):.1f}%)", _fmt(r.get("costo_ir", 0)),
         "Break-even m²", f"${r.get('be_precio_m2',0):,}/m²"],
    ]
    cw = (W - 2 * M) / 4
    sec_t = Table([[Paragraph(cell, S_BODY if i % 2 == 0 else S_NUM)
                    for i, cell in enumerate(row)]
                   for row in _sec_data],
                  colWidths=[cw * 1.4, cw * 0.9, cw * 1.1, cw * 0.6])
    sec_t.setStyle(TableStyle([
        ("GRID",        (0, 0), (-1, -1), 0.3, BORD),
        ("BACKGROUND",  (0, 0), (-1, -1), colors.HexColor("#FDFAF6")),
        ("BACKGROUND",  (2, 0), (3, -1), LGREY),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(sec_t)

    # ── Página 3: Cabida Arquitectónica ──────────────────────────
    story.append(PageBreak())
    story += _section("Programa Arquitectónico")

    # Parámetros normativos resumidos
    _param_rows = []
    for k in ["pisos_max", "area_libre_min", "retiro_frontal", "retiro_lateral",
              "coeficiente_edificacion", "densidad_neta", "uso_suelo", "zonificacion"]:
        v = params.get(k)
        if v is not None and v != "" and v != 0:
            _label = k.replace("_", " ").title()
            _param_rows.append([Paragraph(_label, S_BODY), Paragraph(str(v), S_NUM)])
    if _param_rows:
        p_cw = (W - 2 * M) / 2
        p_tbl = Table(_param_rows, colWidths=[p_cw * 1.3, p_cw * 0.7])
        p_tbl.setStyle(TableStyle([
            ("GRID",        (0, 0), (-1, -1), 0.3, BORD),
            ("BACKGROUND",  (0, 0), (-1, -1), colors.HexColor("#FDFAF6")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",(0, 0), (-1, -1), 8),
        ]))
        story.append(p_tbl)
        story.append(Spacer(1, 8))

    # Programa volumétrico
    story += _section("Volúmenes y Superficies")
    _vol_data = [
        ["M² construibles totales", f"{cabida.get('area_techada_total_m2',0):,.0f} m²",
         "Pisos",                    str(cabida.get("num_pisos", "—"))],
        ["M² vendibles",            f"{cabida.get('area_vendible_m2',0):,.0f} m²",
         "Sótanos",                  str(cabida.get("num_sotanos", 0))],
        ["Área libre",               f"{cabida.get('area_libre_m2',0):,.0f} m²",
         "Depósitos",                str(cabida.get("depositos_total", 0))],
        ["Estac. residentes",         str(cabida.get("estac_residentes", 0)),
         "Estac. visitas",            str(cabida.get("estac_visitas", 0))],
    ]
    cw2 = (W - 2 * M) / 4
    vol_t = Table([[Paragraph(cell, S_BODY if i % 2 == 0 else S_NUM)
                    for i, cell in enumerate(row)]
                   for row in _vol_data],
                  colWidths=[cw2 * 1.3, cw2 * 0.9, cw2 * 1.0, cw2 * 0.8])
    vol_t.setStyle(TableStyle([
        ("GRID",        (0, 0), (-1, -1), 0.3, BORD),
        ("BACKGROUND",  (0, 0), (-1, -1), colors.HexColor("#FDFAF6")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(vol_t)
    story.append(Spacer(1, 8))

    # Mix tipológico
    unidades = cabida.get("unidades", [])
    _pvm = fin_inputs.get("precio_venta_m2", 0) or 0
    if unidades:
        _num_pisos = cabida.get("num_pisos", 1) or 1
        _total_u   = cabida.get("total_unidades", 1) or 1
        _dptos_piso = round(_total_u / _num_pisos, 1)

        story += _section("Mix de Tipologías y Precio de Venta")
        # Nota dptos/piso
        story.append(Paragraph(
            f"Total: <b>{_total_u} departamentos</b> en <b>{_num_pisos} pisos</b> "
            f"(~{_dptos_piso} dptos/piso · precio de venta: <b>${_pvm:,.0f}/m²</b>)",
            S_SMALL))
        story.append(Spacer(1, 5))

        _hdr_cols = ["Tipología", "Cant.", "m²/und", "m² total",
                     "Precio/und", "Subtotal", "% mix"]
        u_header = [Paragraph(h, _style("uh", fontSize=7.5, textColor=WHITE,
                                         fontName="Helvetica-Bold", alignment=TA_CENTER))
                    for h in _hdr_cols]
        u_data = [u_header]
        for u in unidades:
            pct  = round(u.get("cantidad", 0) / _total_u * 100, 1)
            _am2 = u.get("area_m2", 0)
            _cant = u.get("cantidad", 0)
            _p_und  = _am2 * _pvm
            _p_tot  = _cant * _p_und
            u_data.append([
                Paragraph(u.get("tipo", "—"), S_BODY),
                Paragraph(str(_cant), S_NUM),
                Paragraph(f"{_am2:.0f}", S_NUM),
                Paragraph(f"{u.get('area_total_m2', 0):,.0f}", S_NUM),
                Paragraph(f"${_p_und:,.0f}" if _pvm else "—", S_NUM),
                Paragraph(f"${_p_tot:,.0f}" if _pvm else "—", S_NUM),
                Paragraph(f"{pct}%", S_NUM),
            ])
        # Fila totales
        _ing_dpts = sum(u.get("cantidad",0) * u.get("area_m2",0) * _pvm for u in unidades)
        u_data.append([
            Paragraph("<b>TOTAL</b>", _style("utot", fontSize=9, fontName="Helvetica-Bold", textColor=NAV)),
            Paragraph(f"<b>{_total_u}</b>", _style("utn", fontSize=9, fontName="Helvetica-Bold", textColor=NAV, alignment=TA_RIGHT)),
            Paragraph("", S_NUM),
            Paragraph(f"<b>{cabida.get('area_vendible_m2',0):,.0f}</b>", _style("utav", fontSize=9, fontName="Helvetica-Bold", textColor=NAV, alignment=TA_RIGHT)),
            Paragraph("", S_NUM),
            Paragraph(f"<b>${_ing_dpts:,.0f}</b>" if _pvm else "<b>—</b>",
                      _style("utit", fontSize=9, fontName="Helvetica-Bold", textColor=GRN, alignment=TA_RIGHT)),
            Paragraph("<b>100%</b>", _style("utpct", fontSize=9, fontName="Helvetica-Bold", textColor=NAV, alignment=TA_RIGHT)),
        ])

        _uw = W - 2 * M
        u_tbl = Table(u_data, colWidths=[_uw*0.22, _uw*0.08, _uw*0.09, _uw*0.11,
                                          _uw*0.16, _uw*0.18, _uw*0.10])
        u_tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  NAV),
            ("BACKGROUND",     (0, -1),(-1, -1), colors.HexColor("#E8EDF3")),
            ("LINEABOVE",      (0, -1),(-1, -1), 0.8, NAV),
            ("GRID",           (0, 0), (-1, -1), 0.3, BORD),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2),
             [colors.HexColor("#FDFAF6"), colors.HexColor("#F0EDE8")]),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
            ("LEFTPADDING",    (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
        ]))
        story.append(u_tbl)

    # Observaciones
    obs = cabida.get("observaciones", [])
    if obs:
        story.append(Spacer(1, 8))
        story += _section("Observaciones Normativas")
        for o in obs:
            story.append(Paragraph(f"• {o}", S_SMALL))

    # ── Página 4: Análisis Financiero ────────────────────────────
    story.append(PageBreak())
    story += _section("Análisis Financiero")

    # Panel ingresos vs costos
    _ing_rows = [(k, v) for k, v in ing.items() if v != 0]
    _ING_total = r.get("ingresos_brutos", 1) or 1
    ing_header = [Paragraph(h, _style("ih", fontSize=8, textColor=WHITE,
                                       fontName="Helvetica-Bold"))
                  for h in ["INGRESOS", "Monto", "% del total"]]
    ing_data = [ing_header]
    for k, v in _ing_rows:
        ing_data.append([Paragraph(k, S_BODY),
                          Paragraph(_fmt(v), S_NUM),
                          Paragraph(f"{v/_ING_total*100:.1f}%", S_NUM)])
    i_cw = (W - 2 * M) / 3
    ing_tbl = Table(ing_data, colWidths=[i_cw * 1.6, i_cw * 0.8, i_cw * 0.6])
    ing_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GRN),
        ("GRID",       (0, 0), (-1, -1), 0.3, BORD),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [GRN_L, colors.HexColor("#F5FAF7")]),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING",(0, 0), (-1, -1), 8),
        ("RIGHTPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(ing_tbl)
    story.append(Spacer(1, 10))

    story += _section("Estructura de Costos")
    story.append(_cost_table(det))
    story.append(Spacer(1, 14))

    story += _section("Análisis Gráfico")
    story.append(_chart_financiero())
    story.append(Spacer(1, 10))
    story.append(_chart_costos())
    story.append(Spacer(1, 10))

    # ── Matriz estratégica Precio × Terreno ─────────────────────────
    _mx_res = calcular_sensibilidad_terreno(cabida, fin_inputs, zona)
    if _mx_res:
        _mx_mg, _mx_tir, _mx_precios, _mx_terrenos, _mx_p0, _mx_t0 = _mx_res
        story += _section("Matriz Estratégica — Precio de Venta × Precio del Terreno")
        story.append(Paragraph(
            "Margen neto % en cada combinación. "
            "Verde ≥18% · Amarillo 12–18% · Rojo &lt;12%  |  Celda con borde dorado = escenario actual.",
            S_SMALL))
        story.append(Spacer(1, 6))
        _NAV_C  = colors.HexColor("#0A1628")
        _GOLD_C = colors.HexColor("#B8904A")
        _GRN_C  = colors.HexColor("#1B5E20")
        _YEL_C  = colors.HexColor("#7A5500")
        _RED_C  = colors.HexColor("#B71C1C")
        _GRN_BG = colors.HexColor("#C8E6C9")
        _YEL_BG = colors.HexColor("#FFF9C4")
        _RED_BG = colors.HexColor("#FFCDD2")

        _col0 = min(range(len(_mx_precios)), key=lambda i: abs(_mx_precios[i] - _mx_p0))
        _row0 = min(range(len(_mx_terrenos)), key=lambda i: abs(_mx_terrenos[i] - _mx_t0))

        _mx_data  = []
        _mx_style = [
            ("FONTSIZE",    (0, 0), (-1, -1), 7),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#2A3D52")),
        ]
        # Header row
        _hdr_style = ParagraphStyle("mxh", fontSize=7, fontName="Helvetica-Bold",
                                    textColor=colors.white, alignment=TA_CENTER)
        _hdr_data = [Paragraph("Terreno / Precio m²", _hdr_style)]
        for ci, p in enumerate(_mx_precios):
            _fw = "Helvetica-Bold"
            _fg = colors.HexColor("#B8904A") if ci == _col0 else colors.white
            _bg_hdr = colors.HexColor("#1E3A5A") if ci == _col0 else _NAV_C
            _mx_style.append(("BACKGROUND", (ci+1, 0), (ci+1, 0), _bg_hdr))
            _hdr_data.append(Paragraph(f"${p:,}", ParagraphStyle(
                f"mxhc{ci}", fontSize=7, fontName=_fw, textColor=_fg, alignment=TA_CENTER)))
        _mx_data.append(_hdr_data)
        _mx_style.append(("BACKGROUND", (0, 0), (0, 0), _NAV_C))
        _mx_style.append(("BACKGROUND", (1, 0), (len(_mx_precios), 0), _NAV_C))

        for ri, t in enumerate(_mx_terrenos):
            _is_br = (ri == _row0)
            _rh_bg = colors.HexColor("#1E3A5A") if _is_br else _NAV_C
            _rh_fg = _GOLD_C if _is_br else colors.white
            _mx_style.append(("BACKGROUND", (0, ri+1), (0, ri+1), _rh_bg))
            _row_data = [Paragraph(f"${t:,.0f}", ParagraphStyle(
                f"mxrl{ri}", fontSize=7, fontName="Helvetica-Bold",
                textColor=_rh_fg, alignment=TA_CENTER))]
            for ci in range(len(_mx_precios)):
                mg  = float(_mx_mg.iloc[ri, ci])
                tir = float(_mx_tir.iloc[ri, ci])
                _is_bc = (ci == _col0)
                if mg >= 18:   _bg_c, _fg_c = _GRN_BG, _GRN_C
                elif mg >= 12: _bg_c, _fg_c = _YEL_BG, _YEL_C
                else:          _bg_c, _fg_c = _RED_BG, _RED_C
                _mx_style.append(("BACKGROUND", (ci+1, ri+1), (ci+1, ri+1), _bg_c))
                _cell_txt = f"{mg:.0f}%\nTIR {tir:.0f}%"
                if _is_br and _is_bc:
                    _mx_style.append(("BOX", (ci+1, ri+1), (ci+1, ri+1), 1.5, _GOLD_C))
                _row_data.append(Paragraph(_cell_txt, ParagraphStyle(
                    f"mxc{ri}{ci}", fontSize=6.5, fontName="Helvetica-Bold",
                    textColor=_fg_c, alignment=TA_CENTER, leading=8)))
            _mx_data.append(_row_data)

        _avail_w = W - 2 * M
        _col_w   = _avail_w / (len(_mx_precios) + 1)
        _mx_tbl  = Table(_mx_data, colWidths=[_col_w] * (len(_mx_precios) + 1),
                         rowHeights=[16] * (len(_mx_terrenos) + 1))
        _mx_tbl.setStyle(TableStyle(_mx_style))
        story.append(_mx_tbl)
        story.append(Spacer(1, 10))

    # Resultado final
    story += _section("Resumen de Resultado")
    _res_rows = [
        ("Ingresos brutos",     _fmt(r.get("ingresos_brutos", 0)),      ""),
        ("Costo total s/financ.", _fmt(r.get("costo_total_sin_financ", 0)), ""),
        ("Utilidad bruta",      _fmt(r.get("utilidad_bruta", 0)),
         f"{r.get('margen_bruto_pct',0):.1f}% bruto"),
        (f"Impuesto a la Renta ({r.get('ir_pct',29.5):.1f}%)",
         _fmt(r.get("costo_ir", 0)), ""),
        ("UTILIDAD NETA",       _fmt(r.get("utilidad_neta", 0)),
         f"{_mg:.1f}% neto"),
        ("",  "",  ""),
        ("Gasto financiero banco", _fmt(r.get("costo_financiero", 0)),  ""),
        ("Margen s/financiamiento", "",
         f"{r.get('margen_sin_f_pct',0):.1f}%"),
    ]
    r_cw = (W - 2 * M) / 3
    res_data = []
    res_styles_ts = []
    for idx, (a, b, c) in enumerate(_res_rows):
        if not a:
            continue
        is_total = "UTILIDAD NETA" in a
        sty = _style(f"rs{idx}", fontSize=9,
                     fontName="Helvetica-Bold" if is_total else "Helvetica",
                     textColor=NAV)
        res_data.append([Paragraph(a, sty),
                          Paragraph(b, _style(f"rv{idx}", fontSize=9,
                                              fontName="Helvetica-Bold" if is_total else "Helvetica",
                                              textColor=NAV, alignment=TA_RIGHT)),
                          Paragraph(c, _style(f"rc{idx}", fontSize=8,
                                              fontName="Helvetica",
                                              textColor=GRN if is_total else GREY,
                                              alignment=TA_RIGHT))])
        if is_total:
            res_styles_ts += [
                ("BACKGROUND", (0, len(res_data) - 1), (-1, len(res_data) - 1), GRN_L),
                ("LINEABOVE",  (0, len(res_data) - 1), (-1, len(res_data) - 1), 1.0, GRN),
            ]

    res_tbl = Table(res_data, colWidths=[r_cw * 1.5, r_cw * 0.9, r_cw * 0.6])
    res_tbl.setStyle(TableStyle([
        ("GRID",        (0, 0), (-1, -1), 0.3, BORD),
        ("BACKGROUND",  (0, 0), (-1, -1), colors.HexColor("#FDFAF6")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",(0, 0), (-1, -1), 8),
    ] + res_styles_ts))
    story.append(res_tbl)

    # ── Página: Due Diligence Legal ──────────────────────────────
    if legal:
        story.append(PageBreak())
        story += _section("Due Diligence Legal — Análisis Registral")

        _sem = (legal.get("semaforo") or "amarillo").lower()
        _sem_map = {
            "verde":    (GRN,  GRN_L,  "#1A4731", "● SIN ALERTAS CRÍTICAS"),
            "amarillo": (AMB,  AMB_L,  "#7A5500", "▲ OBSERVACIONES MENORES"),
            "rojo":     (RED,  RED_L,  "#7A1A1A", "✕ ALERTAS CRÍTICAS"),
        }
        _sem_col, _sem_bg, _sem_hex, _sem_label = _sem_map.get(
            _sem, (colors.HexColor("#1E2D3D"), LGREY, "#1E2D3D", "○ INDETERMINADO"))

        # Semáforo banner
        sem_data = [[
            Paragraph(
                f'<para alignment="center">'
                f'<font name="Helvetica-Bold" size="11" color="{_sem_hex}">'
                f'{_sem_label}'
                f'</font></para>',
                _style("sem_lbl", fontSize=11, fontName="Helvetica-Bold",
                       textColor=_sem_col, alignment=TA_CENTER)),
        ]]
        sem_tbl = Table(sem_data, colWidths=[W - 2 * M])
        sem_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), _sem_bg),
            ("BOX",           (0, 0), (-1, -1), 1.5, _sem_col),
            ("LEFTPADDING",   (0, 0), (-1, -1), 14),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
            ("TOPPADDING",    (0, 0), (-1, -1), 14),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ]))
        story.append(sem_tbl)
        story.append(Spacer(1, 8))

        # Resumen legal
        _resumen = legal.get("resumen_legal") or ""
        if _resumen:
            story.append(Paragraph(_resumen,
                _style("leg_res", fontSize=10, textColor=NAV, leading=15)))
            story.append(Spacer(1, 10))

        # Datos registrales clave
        _reg_items = [
            ("Propietario registral",    legal.get("propietario") or legal.get("titular") or "—"),
            ("Área registral",           legal.get("area_registral") or "—"),
            ("Partida registral N°",     legal.get("partida") or legal.get("numero_partida") or "—"),
            ("Cargas / gravámenes",      legal.get("cargas") or "—"),
            ("Hipotecas",                legal.get("hipotecas") or "—"),
            ("Medidas cautelares",       legal.get("medidas_cautelares") or "—"),
            ("Consistencia de área",     legal.get("consistencia_area") or "—"),
        ]
        # Solo mostrar filas con datos reales
        _reg_rows_filtered = [(k, v) for k, v in _reg_items if v and v != "—"]
        if _reg_rows_filtered:
            _reg_header = [
                Paragraph("DATO REGISTRAL", _style("rh0", fontSize=8, fontName="Helvetica-Bold",
                                                    textColor=WHITE)),
                Paragraph("DETALLE",        _style("rh1", fontSize=8, fontName="Helvetica-Bold",
                                                    textColor=WHITE)),
            ]
            _reg_data = [_reg_header] + [
                [Paragraph(k, _style(f"rk{i}", fontSize=9, textColor=NAV)),
                 Paragraph(str(v), _style(f"rv{i}", fontSize=9, textColor=NAV))]
                for i, (k, v) in enumerate(_reg_rows_filtered)
            ]
            _reg_cw = W - 2 * M
            _reg_tbl = Table(_reg_data, colWidths=[_reg_cw * 0.38, _reg_cw * 0.62])
            _reg_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), NAV),
                ("GRID",          (0, 0), (-1, -1), 0.3, BORD),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [LGREY, WHITE]),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",    (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ]))
            story.append(_reg_tbl)
            story.append(Spacer(1, 10))

        # Alertas
        _alertas = legal.get("alertas") or []
        if _alertas:
            story += _section("Alertas y Observaciones")
            for _al in _alertas:
                _al_data = [[
                    Paragraph("⚠",   _style("al_ic", fontSize=10, textColor=AMB,
                                             alignment=TA_CENTER)),
                    Paragraph(str(_al), _style("al_tx", fontSize=9, textColor=NAV, leading=14)),
                ]]
                _al_tbl = Table(_al_data, colWidths=[10 * mm, W - 2 * M - 10 * mm])
                _al_tbl.setStyle(TableStyle([
                    ("BACKGROUND",    (0, 0), (-1, -1), AMB_L),
                    ("BOX",           (0, 0), (-1, -1), 0.5, AMB),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                    ("TOPPADDING",    (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(_al_tbl)
                story.append(Spacer(1, 5))

        # Recomendación final
        _recom = legal.get("recomendacion") or legal.get("conclusion") or ""
        if _recom:
            story.append(Spacer(1, 4))
            story += _section("Recomendación")
            story.append(Paragraph(_recom,
                _style("leg_rec", fontSize=10, textColor=NAV, leading=15)))

        # Nota de alcance
        story.append(Spacer(1, 14))
        story.append(Paragraph(
            "Nota: El análisis legal se basa exclusivamente en los documentos adjuntados "
            "(Partida Registral SUNARP y/o PU/HR). No sustituye la revisión de un abogado "
            "especialista en derecho registral y no cubre aspectos tributarios, ambientales "
            "ni de zonificación.",
            _style("leg_nota", fontSize=8, textColor=GREY, leading=12)))

    # ── Conclusión y Recomendación Estratégica ───────────────────
    story.append(PageBreak())
    story += _section("Conclusión y Recomendación Estratégica")

    _dur_txt = f"{r.get('meses_proyecto', 0)} meses" if r.get("meses_proyecto") else "—"
    _unds    = cabida.get("total_unidades", 0) or 0
    _area_v  = cabida.get("area_vendible_m2", 0) or 0

    # Párrafo de contexto
    story.append(Paragraph(
        f"El presente análisis evalúa el proyecto <b>{_nombre_proy}</b> ubicado en <b>{distrito}</b>, "
        f"sobre un terreno de <b>{params.get('area_terreno_m2', 0):,.0f} m²</b> con potencial de "
        f"<b>{_unds} unidades</b> y <b>{_area_v:,.0f} m² vendibles</b>. "
        f"La duración proyectada del ciclo completo (obra + ventas) es de <b>{_dur_txt}</b>.",
        _style("con_ctx", fontSize=9.5, textColor=NAV, leading=15)))
    story.append(Spacer(1, 6))

    # Diagnóstico financiero
    _diag_lines = []
    if _mg >= 20 and _tir >= 15:
        _diag_lines.append(
            f"Margen neto de <b>{_mg:.1f}%</b> y TIR anual de <b>{_tir:.1f}%</b> — "
            f"retornos en rango alto para el mercado de Lima. "
            f"El usuario evalúa si se alinean con su perfil y expectativas.")
    elif _mg > 0 and _tir > 0:
        _diag_lines.append(
            f"Margen neto de <b>{_mg:.1f}%</b> y TIR de <b>{_tir:.1f}%</b> — "
            f"el proyecto genera retornos positivos. "
            f"El usuario evalúa si los márgenes se ajustan a su objetivo de inversión.")
    else:
        _diag_lines.append(
            f"Margen de <b>{_mg:.1f}%</b> y TIR de <b>{_tir:.1f}%</b> — "
            f"el proyecto presenta pérdidas en las condiciones actuales. "
            f"Revisar estructura de costos, precio de venta o valor del terreno.")

    # Diagnóstico del precio del terreno
    if _pc > 0:
        if _pc <= _v20:
            _diag_lines.append(
                f"El precio del terreno (<b>{_fmt(_pc)}</b>) se encuentra en <b>zona óptima</b>: "
                f"por debajo del máximo compatible con margen ≥ 20% ({_fmt(_v20)}). "
                f"Hay capacidad negociadora y cushion financiero.")
        elif _pc <= _v15:
            _diag_lines.append(
                f"El precio del terreno (<b>{_fmt(_pc)}</b>) está en <b>zona aceptable</b> "
                f"(entre los umbrales de 15% y 20% de margen). "
                f"El proyecto es viable pero sin holgura; cualquier sobrecosto impactará el resultado.")
        else:
            _diag_lines.append(
                f"El precio del terreno (<b>{_fmt(_pc)}</b>) <b>supera</b> el umbral de viabilidad óptima "
                f"({_fmt(_v20)}). Se recomienda negociar a la baja o ajustar el programa arquitectónico "
                f"para incrementar la densidad vendible.")

    for _dl in _diag_lines:
        story.append(Paragraph(_dl, _style("con_diag", fontSize=9.5, textColor=NAV, leading=15)))
        story.append(Spacer(1, 5))

    # Recomendación ejecutiva en cuadro destacado
    if _mg >= 20 and _tir >= 15:
        _rec_color, _rec_bg = GRN, GRN_L
        _rec_icon  = "✔"
        _rec_title = "RECOMENDACIÓN: PROCEDER"
        _rec_body  = (
            f"El proyecto presenta condiciones financieras favorables. Se recomienda <b>avanzar con el proceso "
            f"de adquisición del terreno</b> sujeto a la confirmación del due diligence legal y a la obtención "
            f"de la certificación de parámetros urbanísticos vigentes. Asegurar el financiamiento bancario "
            f"(40% del costo de construcción) antes del inicio de obra para "
            f"optimizar el retorno sobre capital propio.")
    elif _mg >= 12 and _tir >= 10:
        _rec_color, _rec_bg = AMB, AMB_L
        _rec_icon  = "▲"
        _rec_title = "RECOMENDACIÓN: NEGOCIAR ANTES DE PROCEDER"
        _rec_body  = (
            f"El proyecto es viable bajo las condiciones actuales, pero <b>se recomienda negociar el precio "
            f"del terreno</b> para ampliar el margen de seguridad. Una reducción del 5-10% en el costo del "
            f"terreno mejoraría el margen neto en ~2-3 puntos porcentuales. Evaluar también la posibilidad "
            f"de ampliar el precio de venta por m² o reducir el área de estacionamientos si la norma lo permite.")
    else:
        _rec_color, _rec_bg = RED, RED_L
        _rec_icon  = "✕"
        _rec_title = "RECOMENDACIÓN: REFORMULAR EL PROYECTO"
        _rec_body  = (
            f"Las condiciones actuales no justifican el avance. <b>Se recomienda reformular el programa "
            f"antes de comprometer capital</b>: revisar el precio del terreno, explorar mayor densidad "
            f"(si la norma lo permite), optimizar el mix tipológico hacia unidades de mayor valor por m², "
            f"o reconsiderar la viabilidad del distrito en el contexto del mercado actual.")

    _rec_data = [[
        Paragraph(_rec_icon, _style("ri", fontSize=16, textColor=_rec_color,
                                    alignment=TA_CENTER, fontName="Helvetica-Bold")),
        Table([
            [Paragraph(_rec_title, _style("rt", fontSize=9, textColor=_rec_color,
                                           fontName="Helvetica-Bold", leading=12))],
            [Paragraph(_rec_body,  _style("rb", fontSize=9, textColor=NAV, leading=13))],
        ], colWidths=[W - 2*M - 18*mm],
           style=TableStyle([("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
                              ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)])),
    ]]
    _rec_tbl = Table(_rec_data, colWidths=[14*mm, W - 2*M - 14*mm])
    _rec_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _rec_bg),
        ("BOX",           (0, 0), (-1, -1), 1.2, _rec_color),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(_rec_tbl)
    story.append(Spacer(1, 8))

    # Nota de alcance
    story.append(Paragraph(
        "Este análisis es de carácter referencial y no sustituye el estudio de mercado detallado, "
        "el due diligence legal completo ni la opinión de un arquitecto sobre la cabida definitiva. "
        "Los resultados dependen de los supuestos ingresados y pueden variar con la variación de "
        "precios de mercado, costos de construcción y normativa municipal.",
        _style("con_nota", fontSize=7.5, textColor=GREY, leading=11)))

    # ── Build ────────────────────────────────────────────────────
    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════
# HELPERS UI
# ═══════════════════════════════════════════════════════

def fmt_usd(v):
    v = v or 0
    if abs(v) >= 1_000_000:
        return f"${v/1_000_000:.2f}M"
    return f"${v:,.0f}"

def card(label, value, color="#1E2D3D"):
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">{label}</div>
        <div class="value" style="color:{color}">{value}</div>
    </div>""", unsafe_allow_html=True)

def row_item(label, value, highlight=False):
    bg    = "#F5F0E8" if highlight else "#FFFFFF"
    fg    = "#1E2D3D"
    val_c = "#B8904A" if highlight else "#1E2D3D"
    bdr   = "#C8A86A" if highlight else "#E8E4DC"
    lw    = "600" if highlight else "400"
    st.markdown(f"""
    <div style="display:flex;justify-content:space-between;align-items:center;
                padding:9px 14px;background:{bg};border-radius:3px;
                margin-bottom:3px;border:1px solid {bdr};
                {'border-left:3px solid #B8904A;' if highlight else ''}">
        <span style="color:{fg};font-size:12px;letter-spacing:0.3px;font-weight:{lw}">{label}</span>
        <span style="color:{val_c};font-weight:700;font-size:13px">{value}</span>
    </div>""", unsafe_allow_html=True)


def score_viabilidad(r: dict) -> tuple:
    """Devuelve (pts, score_10, etiqueta, color_txt, color_bg, recomendacion, items)."""
    margen = r.get("margen_pct", 0)
    tir    = r.get("tir_anual_pct", 0)
    roi    = r.get("roi_pct", 0)
    pts, items = 0, []

    if margen >= 25:   pts += 4; items.append(("Margen neto",   f"{margen:.0f}%", 4, 4))
    elif margen >= 20: pts += 3; items.append(("Margen neto",   f"{margen:.0f}%", 3, 4))
    elif margen >= 15: pts += 2; items.append(("Margen neto",   f"{margen:.0f}%", 2, 4))
    elif margen >= 10: pts += 1; items.append(("Margen neto",   f"{margen:.0f}%", 1, 4))
    else:                         items.append(("Margen neto",   f"{margen:.0f}%", 0, 4))

    if tir >= 20:   pts += 3; items.append(("TIR anual",   f"{tir:.0f}%",    3, 3))
    elif tir >= 15: pts += 2; items.append(("TIR anual",   f"{tir:.0f}%",    2, 3))
    elif tir >= 10: pts += 1; items.append(("TIR anual",   f"{tir:.0f}%",    1, 3))
    else:                      items.append(("TIR anual",   f"{tir:.0f}%",    0, 3))

    if roi >= 20:   pts += 2; items.append(("ROI",         f"{roi:.0f}%",    2, 2))
    elif roi >= 15: pts += 1; items.append(("ROI",         f"{roi:.0f}%",    1, 2))
    else:                      items.append(("ROI",         f"{roi:.0f}%",    0, 2))

    score_10 = round(pts / 9 * 10, 1)

    if pts >= 8:
        return (pts, score_10, "RETORNOS SÓLIDOS",    "#1A4731", "#E8F5EE",
                "Margen y TIR en rango alto. Evalúa si se alinea con tu perfil de inversión.", items)
    elif pts >= 3:
        return (pts, score_10, "RETORNOS MODERADOS",  "#7A4F1A", "#FFF8EE",
                "El proyecto genera retornos positivos. Analiza si los márgenes se ajustan a tu objetivo.", items)
    else:
        return (pts, score_10, "RETORNOS AJUSTADOS",  "#4A4A5A", "#F4F4F6",
                "Retornos por debajo del promedio de mercado. El usuario evalúa si se ajusta a su perfil.", items)


@st.cache_data(show_spinner=False)
def calcular_sensibilidad(cabida: dict, fin_base: dict, zona: str) -> pd.DataFrame:
    """Matriz de márgenes % para ±20% en precio y costo."""
    variaciones = [-20, -10, 0, 10, 20]
    filas = []
    for dpct in variaciones:
        fila = []
        for cpct in variaciones:
            p_adj = fin_base.get("precio_venta_m2", 0) * (1 + dpct / 100)
            c_adj = fin_base.get("costo_construccion", 0) * (1 + cpct / 100)
            fin_adj = {**fin_base, "precio_venta_m2": p_adj, "costo_construccion": c_adj}
            r_adj = calcular_financiero(cabida, fin_adj, zona)["resumen"]
            fila.append(f"{r_adj['margen_pct']:.0f}%")
        filas.append(fila)
    cols = [f"Costo {x:+d}%" for x in variaciones]
    idx  = [f"Precio {x:+d}%" for x in variaciones]
    return pd.DataFrame(filas, columns=cols, index=idx)


def calcular_sensibilidad_terreno(cabida: dict, fin_base: dict, zona: str):
    """Matriz margen% + TIR% cruzando precio de venta (cols) × precio de terreno (filas).
    Returns (df_mg, df_tir, precios_list, terrenos_list, precio_base, terreno_base).
    """
    precio_base  = fin_base.get("precio_venta_m2", 0)
    terreno_base = fin_base.get("costo_terreno", 0)
    if precio_base <= 0 or terreno_base <= 0:
        return None

    # 7 columnas: precio de venta -15%…+15% redondeado a $50
    pct_p = [-15, -10, -5, 0, 5, 10, 15]
    precios   = [max(500, round(precio_base * (1 + p / 100) / 50) * 50) for p in pct_p]

    # 7 filas: terreno -30%…+30% redondeado a $5,000
    pct_t = [-30, -20, -10, 0, 10, 20, 30]
    terrenos  = [max(10_000, round(terreno_base * (1 + t / 100) / 5_000) * 5_000) for t in pct_t]

    mg_grid  = []
    tir_grid = []
    for t in terrenos:
        mg_row  = []
        tir_row = []
        for p in precios:
            fin_adj = {**fin_base, "precio_venta_m2": p, "costo_terreno": t}
            r = calcular_financiero(cabida, fin_adj, zona)["resumen"]
            mg_row.append(r["margen_pct"])
            tir_row.append(r["tir_anual_pct"])
        mg_grid.append(mg_row)
        tir_grid.append(tir_row)

    cols = [f"${p:,}" for p in precios]
    idx  = [f"${t:,.0f}" for t in terrenos]
    df_mg  = pd.DataFrame(mg_grid,  columns=cols, index=idx)
    df_tir = pd.DataFrame(tir_grid, columns=cols, index=idx)
    return df_mg, df_tir, precios, terrenos, precio_base, terreno_base


def _s_curve_weights(n: int) -> list:
    """Bell-shaped marginal weights (slow-fast-slow disbursement), sum = 1.0.
    Uses midpoints t=(i+0.5)/n so no endpoint is ever zero (avoids $0 first/last month).
    """
    if n <= 0:
        return []
    if n == 1:
        return [1.0]
    raw = [6.0 * ((i + 0.5) / n) * (1.0 - (i + 0.5) / n) for i in range(n)]
    s = sum(raw) or 1.0
    return [w / s for w in raw]


def generar_dcf_excel(df_fl: "pd.DataFrame", result_financiero: dict,
                      fin: dict, escenarios: dict, nombre_proyecto: str = "") -> bytes:
    """Genera un Excel profesional del DCF mensual."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Mensual"

    # ── Paleta ──────────────────────────────────────────────────────────────
    DARK   = "0A1628"; GOLD   = "B8904A"; WHITE  = "FFFFFF"
    GRN_D  = "1A4731"; GRN_L  = "E8F5EE"; RED_D  = "7A1A1A"; RED_L  = "FDECEA"
    NAV    = "1E2D3D"; LIGHT  = "F5F2ED"; ALT    = "EAE6DF"

    def _fill(hex_): return PatternFill("solid", fgColor=hex_)
    def _font(hex_="000000", bold=False, sz=10):
        return Font(color=hex_, bold=bold, size=sz, name="Calibri")
    def _side(): return Side(style="thin", color="D8D4CC")
    def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    _align_c = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _align_r = Alignment(horizontal="right",  vertical="center")
    _align_l = Alignment(horizontal="left",   vertical="center")
    _usd_fmt = '#,##0;[Red]-#,##0'

    r = result_financiero.get("resumen", {})
    sb = escenarios.get("sin_banco", {})
    cb = escenarios.get("con_banco", {})

    # ── Fila 1: Título ───────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"].value = f"FLUJO DE CAJA MENSUAL — {nombre_proyecto.upper() or 'PROYECTO'}"
    ws["A1"].fill  = _fill(DARK)
    ws["A1"].font  = Font(color=GOLD, bold=True, size=13, name="Calibri")
    ws["A1"].alignment = _align_c
    ws.row_dimensions[1].height = 24

    # ── Filas 2-3: KPIs clave ────────────────────────────────────────────────
    _kpi_headers = ["Margen Neto", "TIR Anual", "Inversión Máxima", "Breakeven",
                    "Ingresos Brutos", "Utilidad Neta", "Meses Obra", "Meses Venta"]
    _kpi_values  = [
        f"{r.get('margen_pct', 0):.1f}%",
        f"{r.get('tir_anual_pct', 0):.1f}%",
        f"${abs(sb.get('max_exp', 0)):,.0f}",
        f"Mes {sb.get('mes_be', '—')}",
        f"${r.get('ingresos_brutos', 0):,.0f}",
        f"${r.get('utilidad_neta', 0):,.0f}",
        f"{r.get('meses_obra', 0)} meses",
        f"{r.get('meses_venta', 0)} meses",
    ]
    for ci, (hdr, val) in enumerate(zip(_kpi_headers, _kpi_values), start=1):
        ws.cell(row=2, column=ci).value     = hdr
        ws.cell(row=2, column=ci).fill      = _fill(NAV)
        ws.cell(row=2, column=ci).font      = _font(GOLD, bold=True, sz=9)
        ws.cell(row=2, column=ci).alignment = _align_c
        ws.cell(row=3, column=ci).value     = val
        ws.cell(row=3, column=ci).fill      = _fill(LIGHT)
        ws.cell(row=3, column=ci).font      = _font(NAV, bold=True, sz=10)
        ws.cell(row=3, column=ci).alignment = _align_c
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 18

    # ── Fila 4: espacio ──────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 8

    # ── Fila 5: Cabeceras de columna ─────────────────────────────────────────
    _cols = [
        ("Mes",                    10),
        ("Fase / Actividad",       22),
        ("Flujo Mensual\nSin Banco", 17),
        ("Flujo Acum.\nSin Banco",   17),
        ("Flujo Mensual\nCon Banco", 17),
        ("Flujo Acum.\nCon Banco",   17),
        ("Saldo Deuda",            15),
        ("Var. Mensual %",         14),
        ("Semáforo",                9),
    ]
    for ci, (label, width) in enumerate(_cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        c = ws.cell(row=5, column=ci)
        c.value = label; c.fill = _fill(NAV)
        c.font  = _font(WHITE, bold=True, sz=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[5].height = 28

    # ── Fases del proyecto ───────────────────────────────────────────────────
    n_meses = len(df_fl)
    mo     = r.get("meses_obra", 16)
    mv     = r.get("meses_venta", 12)
    inicio_o = 2   # preventa 2 m → banco → obra
    fin_o  = inicio_o + mo
    fin_v  = fin_o + min(6, max(0, mv - mo))

    def _fase(mes):
        if mes <= 2:             return "Due Diligence / Preventa"
        if mes <= fin_o:         return "Construcción activa"
        if mes <= fin_o + 3:     return "Entrega y cierre"
        return "Post-entrega"

    # ── Datos ────────────────────────────────────────────────────────────────
    flujo_sin = df_fl["Flujo Sin Banco"].tolist()
    flujo_con = df_fl["Flujo Con Banco"].tolist()
    acum_sin  = df_fl["Acum. Sin Banco"].tolist()
    acum_con  = df_fl["Acum. Con Banco"].tolist()
    saldo     = df_fl["Saldo Deuda"].tolist()

    for i, mes in enumerate(df_fl["Mes"].tolist()):
        row = 6 + i
        bg  = ALT if i % 2 == 0 else WHITE
        fs  = flujo_sin[i]; fa = acum_sin[i]
        fc_ = flujo_con[i]; fb = acum_con[i]
        sd  = saldo[i]
        var = (fs - flujo_sin[i-1]) / abs(flujo_sin[i-1]) * 100 if i > 0 and flujo_sin[i-1] != 0 else 0
        sem = "🟢" if fs >= 0 else ("🔴" if fs < -50000 else "🟡")
        f_bg_s = GRN_L if fs >= 0 else RED_L
        f_bg_c = GRN_L if fc_ >= 0 else RED_L
        a_bg_s = GRN_L if fa >= 0 else RED_L
        a_bg_c = GRN_L if fb >= 0 else RED_L

        vals = [mes, _fase(mes), fs, fa, fc_, fb, sd if sd > 0 else None, var if i > 0 else None, sem]
        fills= [bg, bg, f_bg_s, a_bg_s, f_bg_c, a_bg_c, bg, bg, bg]
        fmts = [None, None, _usd_fmt, _usd_fmt, _usd_fmt, _usd_fmt, _usd_fmt, '0.0"%"', None]
        aligns= [_align_c, _align_l, _align_r, _align_r, _align_r, _align_r, _align_r, _align_r, _align_c]

        for ci, (v, fl, fmt, aln) in enumerate(zip(vals, fills, fmts, aligns), start=1):
            c = ws.cell(row=row, column=ci)
            c.value = v; c.fill = _fill(fl)
            c.font  = _font(GRN_D if (ci in (3,4) and isinstance(v, (int,float)) and v >= 0) else
                            RED_D  if (ci in (3,4) and isinstance(v, (int,float)) and v < 0) else NAV,
                            bold=(ci == 1), sz=10)
            c.alignment = aln; c.border = _border()
            if fmt and isinstance(v, (int, float)): c.number_format = fmt
        ws.row_dimensions[row].height = 15

    # ── Fila de totales ───────────────────────────────────────────────────────
    tot_row = 6 + len(df_fl)
    ws.merge_cells(f"A{tot_row}:B{tot_row}")
    ws.cell(row=tot_row, column=1).value     = "TOTAL / RESULTADO FINAL"
    ws.cell(row=tot_row, column=1).fill      = _fill(DARK)
    ws.cell(row=tot_row, column=1).font      = _font(GOLD, bold=True, sz=10)
    ws.cell(row=tot_row, column=1).alignment = _align_c
    for ci, val in enumerate([None, None, sum(flujo_sin), acum_sin[-1],
                               sum(flujo_con), acum_con[-1], None, None, None], start=1):
        c = ws.cell(row=tot_row, column=ci)
        if val is not None:
            c.value = val; c.number_format = _usd_fmt
            c.fill  = _fill(GRN_L if val >= 0 else RED_L)
            c.font  = _font(GRN_D if val >= 0 else RED_D, bold=True, sz=10)
        else:
            c.fill = _fill(DARK)
        c.border = _border(); c.alignment = _align_r
    ws.row_dimensions[tot_row].height = 18

    # ── Freeze panes ────────────────────────────────────────────────────────
    ws.freeze_panes = "C6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_flujo(cabida: dict, result_financiero: dict, fin: dict, zona: str):
    """DCF mensual completo: S-curve construcción, IGV crédito fiscal, sin-banco vs con-banco.
    Returns: (df, flujo_list, tir_anual, mes_be, max_exp, escenarios)
      escenarios = {
        "sin_banco": {"tir": float, "max_exp": int, "mes_be": int, "acum": list, "flujo": list},
        "con_banco":  {"tir": float, "max_exp": int, "mes_be": int, "acum": list, "flujo": list,
                       "saldo_deuda": list, "interes_total": float, "igv_credito": float},
      }
    """
    m   = MERCADO.get(zona, {})
    r   = result_financiero["resumen"]
    raw = result_financiero["_raw"]

    meses_obra   = int(fin.get("meses_obra_override") or r["meses_obra"])
    meses_obra   = max(1, min(meses_obra, 60))
    meses_venta  = max(r["meses_venta"], 1)
    n_unidades   = max(cabida.get("total_unidades", 1) or 1, 1)
    vel          = m.get("velocidad_venta", 1.0) or 1.0

    # ── Preventa: tiempo para alcanzar el % mínimo exigido por banco ─
    # Escala con el tamaño del proyecto: meses = ⌈(n_unidades × pct_preventa) / vel⌉
    import math as _math
    _pct_pv        = fin.get("pct_preventa_banco", 30.0) / 100
    _unid_req      = max(1, _math.ceil(n_unidades * _pct_pv))
    _meses_pv_auto = max(1, _math.ceil(_unid_req / vel))
    meses_preventa = int(fin.get("meses_preventa_override") or _meses_pv_auto)
    meses_preventa = max(1, min(meses_preventa, 36))    # clamp 1–36 meses

    inicio_obra  = meses_preventa   # construcción arranca después de preventa
    fin_obra     = inicio_obra + meses_obra
    total_meses  = fin_obra + 6
    n_months     = total_meses + 1

    # ── S-curve construcción ──────────────────────────
    c_obra       = raw["c_obra_dptos"] + raw["c_obra_sotanos"] + raw["c_constructora"]
    s_weights    = _s_curve_weights(meses_obra)
    obra_mensual = [c_obra * w for w in s_weights]

    # ── IGV crédito fiscal ────────────────────────────
    # 18% IGV pagado en construcción = crédito fiscal recuperable al saldo de obra
    igv_credito  = c_obra * 0.18

    # ── Banco — un crédito con dos esquemas posibles ──
    # Estándar:    terreno = 100% equity → banco financia solo obra
    # Con terreno: promotor aporta X% en minuta → banco financia saldo terreno + obra
    estructura      = fin.get("estructura_financ", "estandar")
    aporte_pct      = fin.get("aporte_propio_pct", 20.0) / 100
    c_terreno_total = raw.get("c_terreno_total", 0.0)
    banco_terreno   = 0.0 if estructura == "estandar" else c_terreno_total * (1.0 - aporte_pct)
    tasa_mes_banco  = fin.get("tasa_financ", 9.0) / 100 / 12

    # Pre-computa saldo: arranca con banco_terreno (mes 0),
    # acumula interés durante preventa, luego suma armadas de obra
    saldo_banco      = banco_terreno
    saldo_deuda_list = [0.0] * n_months
    saldo_deuda_list[0] = banco_terreno
    # Interés sobre saldo terreno durante fase de preventa (antes de inicio de obra)
    for _pv in range(meses_preventa):
        saldo_banco  += saldo_banco * tasa_mes_banco
        if _pv < n_months:
            saldo_deuda_list[_pv] = saldo_banco
    for j, costo_mes in enumerate(obra_mensual):
        i = inicio_obra + j
        if i >= n_months:
            break
        interes_i    = saldo_banco * tasa_mes_banco
        saldo_banco += costo_mes + interes_i     # banco financia 100% obra en armadas
        saldo_deuda_list[i] = saldo_banco
    principal_total = banco_terreno + c_obra
    interes_total   = max(0.0, saldo_banco - principal_total)

    # ── Constructor de flujo genérico ─────────────────
    def _build_flujo(con_banco: bool) -> list:
        fl = [0.0] * n_months

        # Terreno: mes 0
        # Estándar → equity paga 100% | Con terreno → equity paga solo aporte (banco paga saldo)
        equity_terreno = c_terreno_total if (estructura == "estandar" or not con_banco) else c_terreno_total * aporte_pct
        fl[0] -= equity_terreno

        # Diseño (arq + esp + permisos): spread durante fase de preventa
        soft = raw["c_arq"] + raw["c_esp"] + raw["c_permisos"]
        for i in range(meses_preventa):
            fl[i] -= soft / meses_preventa

        # Legales: spread uniforme sobre todo el proyecto
        c_leg = raw["c_legales"] / n_months
        for i in range(n_months):
            fl[i] -= c_leg

        # Marketing de preventa (equity): sala de ventas, web, imagen — antes de inicio de obra
        pct_mktg_pv = fin.get("pct_mktg_preventa", 2.0) / 100
        c_mktg_pv   = (raw.get("c_ventas_marketing", 0) + raw.get("c_gerenciamiento", 0)) * pct_mktg_pv
        for i in range(meses_preventa):
            fl[i] -= c_mktg_pv / meses_preventa

        # Construcción + supervisión técnica con S-curve (arranca en inicio_obra = meses_preventa)
        # con_banco → banco paga 100% obra en armadas; equity paga solo supervisión
        c_sup = raw.get("c_supervision", 0.0)
        for j, costo_mes in enumerate(obra_mensual):
            i = inicio_obra + j
            if i >= n_months:
                break
            sup_mes = c_sup * s_weights[j]
            if con_banco:
                fl[i] -= sup_mes          # solo supervisión; banco paga la obra
            else:
                fl[i] -= costo_mes + sup_mes   # equity paga 100%

        # Repago banco (principal + interés capitalizado) al fin de obra
        if con_banco:
            fl[min(fin_obra, total_meses)] -= saldo_banco

        # Costos de ventas + gerenciamiento restantes (excl. marketing preventa ya aplicado)
        c_vtas = (raw["c_ventas_marketing"] + raw["c_gerenciamiento"]) * (1.0 - pct_mktg_pv)
        for i in range(min(meses_venta, n_months)):
            fl[i] -= c_vtas / meses_venta

        # Ingresos: PIE 10% | 20 cuotas 30% | saldo 60% al fin de obra
        precio_u      = raw["ing_brutos"] / n_unidades
        unidades_acum = 0.0
        for mes in range(n_months):
            restantes = n_unidades - unidades_acum
            if restantes < 0.001:
                break
            u_mes          = min(vel, restantes)
            unidades_acum += u_mes
            fl[mes] += precio_u * 0.10 * u_mes
            cuota = precio_u * 0.30 / 20 * u_mes
            for k in range(1, 21):
                if mes + k < n_months:
                    fl[mes + k] += cuota
            fl[min(fin_obra, total_meses)] += precio_u * 0.60 * u_mes

        # IR
        fl[min(fin_obra + 2, total_meses)] -= raw["c_ir"]

        return fl

    flujo_sin = _build_flujo(con_banco=False)
    flujo_con = _build_flujo(con_banco=True)

    def _stats(fl):
        tir_m  = _irr_bisect(fl)
        tir_a  = round(((1 + tir_m) ** 12 - 1) * 100, 1) if tir_m is not None else None
        acum, acc = 0.0, []
        for f in fl:
            acum += f
            acc.append(acum)
        mes_be  = next((i for i, a in enumerate(acc) if a >= 0 and i > 0), None)
        max_exp = min(acc)
        return tir_a, mes_be, round(max_exp), acc

    tir_sin, mes_be_sin, max_exp_sin, acum_sin = _stats(flujo_sin)
    tir_con, mes_be_con, max_exp_con, acum_con = _stats(flujo_con)

    escenarios = {
        "sin_banco": {
            "tir": tir_sin, "max_exp": max_exp_sin, "mes_be": mes_be_sin,
            "acum": acum_sin, "flujo": flujo_sin,
        },
        "con_banco": {
            "tir": tir_con, "max_exp": max_exp_con, "mes_be": mes_be_con,
            "acum": acum_con, "flujo": flujo_con,
            "saldo_deuda": saldo_deuda_list,
            "interes_total": round(interes_total),
            "igv_credito": round(igv_credito),
            "obra_mensual": [round(x) for x in obra_mensual],
        },
    }

    # Backward-compatible base (sin banco)
    tir_anual = tir_sin
    mes_be    = mes_be_sin
    max_exp   = max_exp_sin

    df = pd.DataFrame({
        "Mes":               list(range(n_months)),
        "Flujo Sin Banco":   [round(f) for f in flujo_sin],
        "Flujo Con Banco":   [round(f) for f in flujo_con],
        "Acum. Sin Banco":   [round(f) for f in acum_sin],
        "Acum. Con Banco":   [round(f) for f in acum_con],
        "Saldo Deuda":       [round(f) for f in saldo_deuda_list],
    })
    return df, flujo_sin, tir_anual, mes_be, max_exp, escenarios


def generar_resumen_ejecutivo_ia(tipo: str, datos: dict) -> dict:
    client = get_client()

    if tipo == "industrial":
        _dscr_str = (f"{datos['dscr']:.2f}x" if datos.get('dscr') else 'sin financiamiento')
        _payback_str = (f"{datos['payback_anos']:.1f} años" if datos.get('payback_anos') else 'N/A')
        _irr_str = (f"{datos['irr_anual']:.1f}%" if datos.get('irr_anual') is not None else 'N/A')
        ctx = (
            f"Activo: {datos.get('tipo_nave')} · Zonificación {datos.get('zonificacion')} · Uso: {datos.get('uso')}\n"
            f"Área nave: {datos.get('area_nave',0):,.0f} m² · Área libre: {datos.get('area_libre',0):,.0f} m²\n"
            f"Costo total: ${datos.get('costo_total',0):,.0f} · Costo/m² nave: ${datos.get('costo_por_m2_nave',0):,.0f}\n"
            f"Yield bruto: {datos.get('yield_bruto',0):.1f}% · Yield neto: {datos.get('yield_neto',0):.1f}%\n"
            f"DSCR: {_dscr_str}\n"
            f"Payback: {_payback_str}\n"
            f"TIR equity 10a: {_irr_str}\n"
            f"Capital propio: ${datos.get('capital_propio',0):,.0f} · Financiado: {datos.get('pct_credito',0):.0f}%\n"
            f"Renta mercado: ${datos.get('renta_m2_mes',0):.2f}/m²/mes"
        )
        tipo_label = "industrial / logístico en Lima, Perú"
        ref = ("Referencia Lima 2025-2026: yield neto target 6–8%, TIR equity mínima 12%, "
               "DSCR ≥ 1.20x, renta logística $5.5–7.0/m²/mes. "
               "Costo construcción nave industrial (estructura metálica, SIN acabados residenciales): "
               "Logística Clase A 12-14m clara $270–300/m², estándar $220–260/m², básica $180–220/m², "
               "cross-docking $380–500/m², manufactura $350–450/m². "
               "Patios/maniobras: $60–90/m². "
               "Proyecto de referencia real: Parque Logístico Lima 14,315 m² nave, 13.6m clara, "
               "inversión $291/m² all-in, renta $6.5/m²/mes, payback 3.74 años, yield bruto 26.8%. "
               "IMPORTANTE: costos industriales son 3-4x más bajos que construcción residencial "
               "por ser estructura metálica sin particiones ni acabados interiores.")
    else:
        _payback_str = (f"{datos['payback_anos']:.1f} años" if datos.get('payback_anos') else 'N/A')
        _flujo_str = (f"${datos['flujo_mensual']:,.0f}" if datos.get('flujo_mensual') is not None else 'N/A')
        ctx = (
            f"Inmueble: {datos.get('uso')} · Precio: ${datos.get('precio',0):,.0f}\n"
            f"Cuota mensual: ${datos.get('cuota_mensual',0):,.0f} · Plazo: {datos.get('plazo_anos')} años\n"
            f"Ingreso mínimo recomendado: ${datos.get('ingreso_minimo',0):,.0f}/mes\n"
            f"Yield bruto: {datos.get('yield_bruto',0):.1f}% · Yield neto: {datos.get('yield_neto',0):.1f}%\n"
            f"Payback: {_payback_str}\n"
            f"Flujo mensual neto: {_flujo_str}\n"
            f"Total intereses crédito: ${datos.get('total_intereses',0):,.0f}\n"
            f"Apreciación est. 5 años: +${datos.get('ganancia_capital_5',0):,.0f}"
        )
        tipo_label = "residencial en Lima, Perú"
        ref = ("Referencia Lima 2025-2026: yield neto residencial 4–6%, payback típico 18–25 años, "
               "cuota recomendada máx. 30% del ingreso mensual, apreciación histórica ~4%/año.")

    _bench_ctx = f"\n\nBENCHMARKS INDUSTRIALES DE REFERENCIA:\n{BENCHMARKS_INDUSTRIAL}" if tipo_label.startswith("industrial") else ""

    prompt = f"""Eres Enrique Osterling, director de Osterling Advisory, con 20 años de experiencia en activos comerciales e industriales en Lima. Eres directo, preciso y orientado a la decisión.

Analiza este activo {tipo_label}:

{ctx}

{ref}{_bench_ctx}
Tasa libre de riesgo Perú: ~7.5% (bonos soberanos PEN).

Devuelve ÚNICAMENTE este JSON sin texto adicional:
{{
  "recomendacion": "comprar/evaluar_con_condiciones/no_recomendado",
  "titulo": "Frase de 6-9 palabras resumiendo la oportunidad de inversión",
  "resumen": "2-3 oraciones describiendo el activo, su posicionamiento y contexto de mercado.",
  "argumentos_favor": ["argumento concreto 1", "argumento concreto 2", "argumento concreto 3"],
  "riesgos": ["riesgo concreto 1", "riesgo concreto 2"],
  "conclusion": "1-2 oraciones de recomendación final. Directo y accionable."
}}"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}]
    )
    if not response.content:
        raise ValueError("json_parse_error: API devolvió respuesta vacía")
    return parse_json_safe(response.content[0].text.strip())


def generar_memorandum_advisory_ind(datos: dict) -> dict:
    client = get_client()

    _dscr_str    = (f"{datos['dscr']:.2f}x" if datos.get('dscr') else 'sin financiamiento')
    _payback_str = (f"{datos['payback_anos']:.1f} años" if datos.get('payback_anos') else 'N/A')
    _irr_str     = (f"{datos['irr_anual']:.1f}%" if datos.get('irr_anual') is not None else 'N/A')

    _renta_m2 = datos.get('renta_m2_mes', 0)
    _costo_m2 = datos.get('costo_m2_mes', 0)
    _renta_ctx = ""
    if _renta_m2 > 0 and _costo_m2 > 0:
        _dif = _renta_m2 - _costo_m2
        _renta_ctx = (
            f"\nComparativa compra vs. arrendamiento:"
            f"\n  Costo efectivo compra: ${_costo_m2:.2f}/m²/mes"
            f"\n  Renta de mercado equivalente: ${_renta_m2:.2f}/m²/mes"
            f"\n  Diferencial: ${_dif:+.2f}/m²/mes ({'favor compra' if _dif > 0 else 'favor arrendamiento'})"
        )

    ctx = (
        f"PARÁMETROS DEL PROYECTO:\n"
        f"Tipo: {datos.get('tipo_nave')} · Zonificación: {datos.get('zonificacion')} · Propósito: {datos.get('uso')}\n"
        f"Área nave techada: {datos.get('area_nave',0):,.0f} m²  ·  Área libre/maniobra: {datos.get('area_libre',0):,.0f} m²\n"
        f"Actividad declarada: {datos.get('actividad_desc', 'No especificada')}\n\n"
        f"ESTRUCTURA DE COSTOS:\n"
        f"Costo terreno: ${datos.get('costo_terreno',0):,.0f}  (${datos.get('costo_terreno_m2',0):,.0f}/m² de terreno)\n"
        f"Costo nave techada: ${datos.get('costo_nave_total',0):,.0f}  (${datos.get('costo_nave_m2',0):,.0f}/m²)\n"
        f"Costo pisos libres: ${datos.get('costo_pisos_libres',0):,.0f}\n"
        f"Costos indirectos ({datos.get('pct_indirectos',5):.0f}%): ${datos.get('soft_costs',0):,.0f}\n"
        f"COSTO TOTAL: ${datos.get('costo_total',0):,.0f}  ·  Costo/m² nave all-in: ${datos.get('costo_por_m2_nave',0):,.0f}\n\n"
        f"FINANCIAMIENTO:\n"
        f"Capital propio: ${datos.get('capital_propio',0):,.0f}\n"
        f"Crédito terreno: ${datos.get('monto_credito_terreno',0):,.0f}  · cuota ${datos.get('cuota_terreno',0):,.0f}/mes\n"
        f"Crédito construcción: ${datos.get('monto_credito_const',0):,.0f}  · cuota ${datos.get('cuota_const',0):,.0f}/mes\n"
        f"Cuota mensual total: ${datos.get('cuota_mensual',0):,.0f}  ·  DSCR: {_dscr_str}\n\n"
        f"INDICADORES DE RETORNO:\n"
        f"Yield bruto: {datos.get('yield_bruto',0):.1f}%  ·  Yield neto: {datos.get('yield_neto',0):.1f}%\n"
        f"Payback: {_payback_str}  ·  TIR equity 10 años: {_irr_str}"
        f"{_renta_ctx}"
    )

    ref = (
        "Benchmarks Lima 2025-2026:\n"
        "- Renta logística Clase A (VES / Lurín / Callao): $5.50–$7.50/m²/mes\n"
        "- Yield neto target operador institucional: 6–8% anual\n"
        "- TIR equity referencia proyecto nuevo: 12–18%\n"
        "- DSCR bancario mínimo: ≥1.20x\n"
        "- Costo nave logística 12-14m clara (estructura metálica): $270–$300/m²\n"
        "- Terreno industrial VES/Lurín build-to-rent: $140–$180/m² (dato real Aldea Logística 2023-2024)\n"
        "- Payback típico proyecto logístico Lima: 8–15 años según renta y uso\n"
        "- Tasa libre de riesgo Perú: ~7.5% (bonos soberanos PEN)"
    )

    prompt = f"""Eres consultor senior de Osterling Advisory, especializado en activos logísticos e industriales en Lima, Perú.

Tu trabajo es elaborar un Memorandum de Advisory Board: un documento de análisis objetivo que consolida toda la información del proyecto para que el cliente tome su propia decisión según sus criterios, estrategia y necesidades específicas.

PRINCIPIO FUNDAMENTAL: No emitas recomendación de compra, venta ni decisión. Presenta la data analítica, crítica, verificable y exacta. Los inversionistas tienen objetivos distintos — generación de flujo, acumulación de patrimonio, uso operativo propio, escudo fiscal — y solo ellos pueden evaluar si este proyecto se alinea con su estrategia.

Datos del proyecto:
{ctx}

Referencias de mercado:
{ref}

Devuelve ÚNICAMENTE este JSON sin texto adicional ni markdown:
{{
  "titulo": "Título descriptivo del proyecto en 6-10 palabras",
  "perfil_activo": "2-3 oraciones describiendo el activo: tipo, superficie, zonificación, propósito y actividad. Solo hechos.",
  "indicadores_clave": "2-3 oraciones con los números más relevantes: inversión total, costo/m², yield, payback, TIR. Solo datos, sin calificativos positivos ni negativos.",
  "posicionamiento_mercado": "2-3 oraciones sobre cómo se posiciona este proyecto frente al mercado: comparativa de costos vs. renta, benchmarks. Factual y neutro.",
  "estructura_financiera": "2 oraciones describiendo el esquema de financiamiento: capital propio, estructura de créditos, cuota mensual, DSCR.",
  "factores_relevantes": ["Dato concreto 1 con cifra verificable", "Dato concreto 2 con cifra verificable", "Dato concreto 3 con cifra verificable", "Dato concreto 4 con cifra verificable"],
  "consideraciones": ["Aspecto a evaluar 1 (dato verificable, sin juicio)", "Aspecto a evaluar 2 (dato verificable, sin juicio)"],
  "sintesis": "2-3 oraciones de síntesis objetiva: qué es el proyecto, qué retorno ofrece bajo los supuestos ingresados, y qué implica financieramente. Sin sesgos. El cliente evaluará su alineación con su estrategia de inversión."
}}"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1400,
        messages=[{"role": "user", "content": prompt}]
    )
    if not response.content:
        raise ValueError("json_parse_error: API devolvió respuesta vacía")
    return parse_json_safe(response.content[0].text.strip())


def generar_informe_industrial_html(r: dict, factibilidad: dict | None, fecha: str) -> str:
    NAV = "#1E2D3D"; GLD = "#B8904A"; LGT = "#F5F2ED"; BRD = "#D8D4CC"
    sem_col = {"verde": "#1A4731", "amarillo": "#7A4F1A", "rojo": "#7A1A1A"}
    sem_bg  = {"verde": "#E8F5EE", "amarillo": "#FFF8EE", "rojo": "#FFF0F0"}

    def kpi(label, value, sub=""):
        sub_html = f'<div style="font-size:10px;color:#7A7268;margin-top:3px;">{sub}</div>' if sub else ""
        return (f'<div style="background:#FFFFFF;border:1px solid {BRD};border-top:3px solid {GLD};'
                f'border-radius:5px;padding:14px 16px;min-width:140px;flex:1;">'
                f'<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;">{label}</div>'
                f'<div style="font-size:22px;font-weight:700;color:{NAV};margin-top:6px;">{value}</div>'
                f'{sub_html}</div>')

    kpis_html = (
        kpi("Costo Total", f"${r.get('costo_total', 0) or 0:,.0f}") +
        kpi("Costo / m² nave", f"${r.get('costo_por_m2_nave', 0) or 0:,.0f}") +
        (kpi("Yield Neto", f"{r.get('yield_neto', 0) or 0:.1f}%", "anual s/ costo total") if r.get('yield_neto') else "") +
        (kpi("Payback", f"{r.get('payback_anos', 0) or 0:.1f} años") if r.get('payback_anos') else "") +
        (kpi("TIR Equity 10a", f"{r.get('irr_anual', 0) or 0:.1f}%" if r.get('irr_anual') is not None else "—") if r.get('uso') == "Inversión" else "")
    )

    _ct = r.get('costo_total') or 1
    costo_rows = "".join(
        f'<tr><td style="padding:8px 12px;color:{NAV};font-size:12px;">{lbl}</td>'
        f'<td style="padding:8px 12px;color:{NAV};font-size:12px;text-align:right;font-weight:600;">${val:,.0f}</td>'
        f'<td style="padding:8px 12px;color:#7A7268;font-size:11px;text-align:right;">{pct}</td></tr>'
        for lbl, val, pct in [
            ("Terreno", r.get('costo_terreno', 0), f"{r.get('costo_terreno',0)/_ct*100:.1f}%"),
            ("Alcabala (3%)", r.get('alcabala', 0), f"{r.get('alcabala',0)/_ct*100:.1f}%"),
            (f"Nave techada ({r.get('area_nave',0):,.0f} m² × ${r.get('costo_nave_m2',0):,.0f}/m²)", r.get('costo_nave_total', 0), f"{r.get('costo_nave_total',0)/_ct*100:.1f}%"),
            (f"Piso área libre ({r.get('area_libre',0):,.0f} m² × ${r.get('costo_piso_libre_m2',0):,.0f}/m²)", r.get('costo_pisos_libres', 0), f"{r.get('costo_pisos_libres',0)/_ct*100:.1f}%"),
            (f"Costos Indirectos ({r.get('pct_indirectos', 5):.0f}%)", r.get('soft_costs', 0), f"{r.get('soft_costs',0)/_ct*100:.1f}%"),
            ("<strong>TOTAL</strong>", r.get('costo_total', 0), "100%"),
        ]
    )

    flujo_rows = ""
    if r.get('flujo_anual') and len(r['flujo_anual']) > 1:
        for i, f in enumerate(r['flujo_anual']):
            yr = f"Año {i}" if i > 0 else "Inversión inicial"
            col = "#1A4731" if f >= 0 else "#7A1A1A"
            flujo_rows += (f'<tr><td style="padding:7px 12px;font-size:12px;color:{NAV};">{yr}</td>'
                          f'<td style="padding:7px 12px;font-size:12px;color:{col};text-align:right;font-weight:600;">${f:,.0f}</td></tr>')

    fac_html = ""
    if factibilidad:
        sg = factibilidad.get("semaforo_global", "amarillo").lower()
        fac_html = (
            f'<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;'
            f'margin:28px 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Factibilidad Técnica y Legal</h3>'
            f'<div style="background:{sem_bg.get(sg,"#F5F2ED")};border-left:4px solid {sem_col.get(sg,NAV)};'
            f'border-radius:5px;padding:14px 18px;margin-bottom:12px;">'
            f'<div style="font-size:12px;color:{sem_col.get(sg,NAV)};font-weight:700;">'
            f'{"SIN ALERTAS CRÍTICAS" if sg=="verde" else ("OBSERVACIONES" if sg=="amarillo" else "ALERTAS CRÍTICAS")}</div>'
            f'<div style="font-size:12px;color:{sem_col.get(sg,NAV)};margin-top:6px;">{factibilidad.get("resumen_tecnico","")}</div>'
            f'<div style="font-size:12px;color:{sem_col.get(sg,NAV)};margin-top:4px;">{factibilidad.get("resumen_legal","")}</div>'
            f'</div>'
        )

    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Informe Industrial — Osterling Advisory</title>
<style>body{{font-family:'Segoe UI',Arial,sans-serif;background:#EDEAE4;margin:0;padding:32px;color:{NAV};}}
.page{{background:#FFFFFF;max-width:860px;margin:0 auto;padding:48px 56px;border-radius:6px;}}
table{{width:100%;border-collapse:collapse;}}thead th{{background:{NAV};color:#FFFFFF;padding:9px 12px;font-size:10px;letter-spacing:1px;text-transform:uppercase;}}
tbody tr:nth-child(even) td{{background:#F9F7F4;}}
</style></head><body><div class="page">
<div style="border-bottom:2px solid {GLD};padding-bottom:20px;margin-bottom:28px;display:flex;justify-content:space-between;align-items:flex-end;">
  <div>
    <div style="font-size:9px;color:{GLD};letter-spacing:4px;text-transform:uppercase;font-weight:600;">Osterling Advisory</div>
    <div style="font-size:22px;font-weight:700;color:{NAV};margin-top:6px;">Análisis Logístico / Industrial</div>
    <div style="font-size:12px;color:#7A7268;margin-top:4px;">{r.get('tipo_nave','—')} · {r.get('zonificacion','—')} · {r.get('uso','—')}</div>
  </div>
  <div style="text-align:right;font-size:11px;color:#9A9080;">{fecha}<br>Lima, Perú</div>
</div>

{"" if not (r.get("actividad_categoria") or r.get("actividad_descripcion")) else
f'<div style="background:#F5F2ED;border-left:4px solid {GLD};border-radius:5px;padding:12px 16px;margin-bottom:20px;">'
f'<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;margin-bottom:4px;">Actividad a Realizar</div>'
+ (f'<div style="font-size:13px;font-weight:700;color:{NAV};">{r["actividad_categoria"]}</div>' if r.get("actividad_categoria") else "")
+ (f'<div style="font-size:12px;color:#4A5870;margin-top:4px;">{r["actividad_descripcion"]}</div>' if r.get("actividad_descripcion") else "")
+ f'<div style="font-size:10px;color:#7A7268;margin-top:6px;">Zonificación: {r.get("zonificacion","—")} (RNE A.060)</div>'
+ '</div>'}

<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Indicadores Clave</h3>
<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:28px;">{kpis_html}</div>

<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Distribución del Área</h3>
<table style="margin-bottom:24px;"><thead><tr><th>Componente</th><th style="text-align:right;">Área (m²)</th><th style="text-align:right;">% del Terreno</th></tr></thead>
<tbody>
<tr><td style="padding:8px 12px;font-size:12px;">Área total terreno</td><td style="padding:8px 12px;font-size:12px;text-align:right;">{r.get('area_terreno', 0) or 0:,.0f} m²</td><td style="padding:8px 12px;font-size:12px;text-align:right;color:#7A7268;">100%</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Nave techada</td><td style="padding:8px 12px;font-size:12px;text-align:right;font-weight:600;">{r.get('area_nave', 0) or 0:,.0f} m²</td><td style="padding:8px 12px;font-size:12px;text-align:right;">{r.get('pct_techada', 0) or 0:.0f}%</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Patios y maniobras</td><td style="padding:8px 12px;font-size:12px;text-align:right;">{r.get('area_libre', 0) or 0:,.0f} m²</td><td style="padding:8px 12px;font-size:12px;text-align:right;">{100 - (r.get('pct_techada', 0) or 0):.0f}%</td></tr>
</tbody></table>

<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Estructura de Costos</h3>
<table style="margin-bottom:24px;"><thead><tr><th>Concepto</th><th style="text-align:right;">Monto USD</th><th style="text-align:right;">% del Total</th></tr></thead>
<tbody>{costo_rows}</tbody></table>

<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Financiamiento</h3>
<table style="margin-bottom:24px;"><thead><tr><th>Concepto</th><th style="text-align:right;">Valor</th></tr></thead>
<tbody>
<tr><td style="padding:8px 12px;font-size:12px;">Capital propio</td><td style="padding:8px 12px;font-size:12px;text-align:right;font-weight:600;">${r.get('capital_propio', 0) or 0:,.0f} ({100 - (r.get('pct_credito', 0) or 0):.0f}%)</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Monto financiado</td><td style="padding:8px 12px;font-size:12px;text-align:right;">${r.get('monto_credito', 0) or 0:,.0f} ({r.get('pct_credito', 0) or 0:.0f}%)</td></tr>
{"<tr><td style='padding:8px 12px;font-size:12px;'>Cuota mensual</td><td style='padding:8px 12px;font-size:12px;text-align:right;'>$"+f"{r.get('cuota_mensual', 0) or 0:,.0f}"+"</td></tr>" if (r.get('cuota_mensual') or 0) > 0 else ""}
{"<tr><td style='padding:8px 12px;font-size:12px;'>Plazo</td><td style='padding:8px 12px;font-size:12px;text-align:right;'>"+str(r.get('plazo_anos', '—'))+" años</td></tr>" if (r.get('cuota_mensual') or 0) > 0 else ""}
{"<tr><td style='padding:8px 12px;font-size:12px;'>DSCR</td><td style='padding:8px 12px;font-size:12px;text-align:right;font-weight:600;'>"+f"{r.get('dscr', 0) or 0:.2f}x"+"</td></tr>" if r.get('dscr') else ""}
</tbody></table>

{"<h3 style='color:"+NAV+";font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid "+BRD+";padding-bottom:6px;'>Flujo de Caja Proyectado (10 años)</h3><table style='margin-bottom:24px;'><thead><tr><th>Período</th><th style='text-align:right;'>Flujo USD</th></tr></thead><tbody>"+flujo_rows+"</tbody></table>" if flujo_rows else ""}

{fac_html}

<div style="margin-top:48px;border-top:1px solid {BRD};padding-top:20px;">
<p style="font-size:11px;font-weight:700;color:{NAV};margin:0;">Enrique Osterling</p>
<p style="font-size:10px;color:#555;margin:3px 0;">Gerente General — Osterling Advisory · Inmobiliaria Corporativa</p>
<p style="font-size:10px;color:#555;margin:3px 0;">+51 950 891 995 · eosterling@grupoosterling.com · Lima, Perú</p>
<p style="font-size:9px;color:#AAA;margin-top:12px;">Análisis referencial basado en los parámetros ingresados. No constituye asesoría legal ni financiera formal.</p>
</div></div></body></html>"""


def generar_informe_industrial_pdf(r: dict, factibilidad, fecha: str, altura_nave: float = 0) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable, KeepTogether)
    import io

    NAV   = colors.HexColor("#1E2D3D")
    GOLD  = colors.HexColor("#B8904A")
    CREAM = colors.HexColor("#F8F5F0")
    BORD  = colors.HexColor("#D8D4CC")
    GRAY  = colors.HexColor("#7A7268")
    BLUE  = colors.HexColor("#4A90C4")
    WHITE = colors.white

    r = r or {}
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    sty = getSampleStyleSheet()
    def _s(name, **kw):
        return ParagraphStyle(name, parent=sty["Normal"], **kw)

    S_body  = _s("ibody",  fontSize=9,  leading=13, textColor=NAV)
    S_small = _s("ismall", fontSize=8,  leading=11, textColor=GRAY)
    S_label = _s("ilabel", fontSize=8,  leading=10, textColor=NAV, fontName="Helvetica-Bold")
    S_gray  = _s("igray",  fontSize=9,  leading=13, textColor=GRAY)
    S_gold  = _s("igold",  fontSize=9,  leading=13, textColor=GOLD, fontName="Helvetica-Bold")
    S_sec   = _s("isec",   fontSize=7,  leading=9,  textColor=NAV, fontName="Helvetica-Bold",
                 charSpace=2, spaceAfter=4)
    S_note  = _s("inote",  fontSize=8,  leading=11, textColor=GRAY, fontName="Helvetica-Oblique")

    W = doc.width
    story = []

    # ── Helpers ──────────────────────────────────────────────────────────────────
    def _section_title(txt):
        return KeepTogether([
            HRFlowable(width=W, thickness=0.75, color=GOLD, spaceAfter=3),
            Paragraph((txt or "").upper(), S_sec),
        ])

    def _data_table(header_row, data_rows, col_widths, bold_last=False):
        """3-column data table with NAV header and alternating CREAM/WHITE rows."""
        hdr = [Paragraph(h, _s(f"dh{i}", fontSize=7, leading=9, textColor=WHITE,
                                fontName="Helvetica-Bold", alignment=(TA_RIGHT if i > 0 else TA_LEFT)))
               for i, h in enumerate(header_row)]
        tbl_data = [hdr]
        for row in data_rows:
            tbl_data.append([Paragraph(str(c), S_body) for c in row])
        ts = TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  NAV),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.4, BORD),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 7),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
            ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ])
        for i in range(len(data_rows)):
            bg = CREAM if i % 2 == 0 else WHITE
            ts.add("BACKGROUND", (0, i + 1), (-1, i + 1), bg)
        if bold_last and data_rows:
            last = len(data_rows)
            ts.add("FONTNAME",   (0, last), (-1, last), "Helvetica-Bold")
            ts.add("BACKGROUND", (0, last), (-1, last), colors.HexColor("#E8E4DC"))
        return Table(tbl_data, colWidths=col_widths, style=ts)

    # ── 1. Header ─────────────────────────────────────────────────────────────────
    tipo_nave   = r.get("tipo_nave", "—")
    zonificacion = r.get("zonificacion", "—")
    uso         = r.get("uso", "—")

    hdr_left = [
        [Paragraph("OSTERLING ADVISORY", _s("hl1", fontSize=7, leading=9, textColor=GOLD,
                                             fontName="Helvetica-Bold", charSpace=3))],
        [Paragraph("FACTIS", _s("hl2", fontSize=20, leading=22, textColor=NAV,
                                 fontName="Helvetica-Bold"))],
        [Paragraph("IA DE ANÁLISIS INMOBILIARIO", _s("hl3", fontSize=6, leading=8,
                                                              textColor=GRAY, charSpace=1.5))],
    ]
    badge_tbl = Table(
        [[Paragraph("ANÁLISIS INDUSTRIAL", _s("badge", fontSize=7, leading=9,
                                               textColor=WHITE, fontName="Helvetica-Bold",
                                               charSpace=1.5, alignment=TA_CENTER))]],
        colWidths=[W * 0.38],
        style=TableStyle([("BACKGROUND",    (0,0), (-1,-1), GOLD),
                          ("TOPPADDING",    (0,0), (-1,-1), 5),
                          ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                          ("LEFTPADDING",   (0,0), (-1,-1), 8),
                          ("RIGHTPADDING",  (0,0), (-1,-1), 8)])
    )
    hdr_right = [
        [Paragraph(f"{fecha}  ·  Lima, Perú", S_small)],
        [Paragraph("eosterling@grupoosterling.com", S_small)],
        [badge_tbl],
    ]
    tbl_hdr = Table(
        [[Table(hdr_left,  colWidths=[W*0.55],
                style=TableStyle([("LEFTPADDING",(0,0),(-1,-1),0),("TOPPADDING",(0,0),(-1,-1),2),
                                   ("BOTTOMPADDING",(0,0),(-1,-1),2)])),
          Table(hdr_right, colWidths=[W*0.45],
                style=TableStyle([("ALIGN",(0,0),(-1,-1),"RIGHT"),("LEFTPADDING",(0,0),(-1,-1),0),
                                   ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)]))]],
        colWidths=[W*0.55, W*0.45],
        style=TableStyle([("BOTTOMPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),0)])
    )
    story.append(tbl_hdr)
    story.append(HRFlowable(width=W, thickness=1, color=GOLD, spaceAfter=6))

    # Title row
    title_tbl = Table(
        [[Paragraph("Análisis Logístico / Industrial",
                    _s("itit", fontSize=14, leading=17, textColor=NAV, fontName="Helvetica-Bold")),
          Paragraph(f"{tipo_nave}  ·  {zonificacion}  ·  {uso}",
                    _s("isub2", fontSize=9, leading=12, textColor=GRAY, alignment=TA_RIGHT))]],
        colWidths=[W*0.58, W*0.42],
        style=TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                          ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
                          ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),6)])
    )
    story.append(title_tbl)
    story.append(Spacer(1, 4*mm))

    # ── 2. Actividad ──────────────────────────────────────────────────────────────
    act_cat  = r.get("actividad_categoria", "") or ""
    act_desc = r.get("actividad_descripcion", "") or ""
    if act_cat or act_desc:
        act_inner = []
        act_inner.append(Paragraph("ACTIVIDAD A REALIZAR",
                                   _s("alab", fontSize=7, leading=9, textColor=GOLD,
                                      fontName="Helvetica-Bold", charSpace=2)))
        if act_cat:
            act_inner.append(Paragraph(act_cat,
                                        _s("acat", fontSize=10, leading=13, textColor=NAV,
                                           fontName="Helvetica-Bold")))
        if act_desc:
            act_inner.append(Paragraph(act_desc, S_gray))
        act_inner.append(Paragraph(f"Zonificación: {zonificacion}  ·  RNE A.060", S_small))
        act_tbl = Table(
            [[act_inner]],
            colWidths=[W],
            style=TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), CREAM),
                ("LINEBEFORE",   (0,0), (0,-1),  3, GOLD),
                ("TOPPADDING",   (0,0), (-1,-1), 8),
                ("BOTTOMPADDING",(0,0), (-1,-1), 8),
                ("LEFTPADDING",  (0,0), (-1,-1), 10),
                ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ])
        )
        story.append(act_tbl)
        story.append(Spacer(1, 4*mm))

    # ── 3. Indicadores Clave ──────────────────────────────────────────────────────
    story.append(_section_title("Indicadores Clave"))
    story.append(Spacer(1, 2*mm))

    _at  = r.get("area_terreno", 0) or 0
    _ct  = r.get("costo_terreno", 0) or 0
    _cta = r.get("costo_terreno_alcabala", 0) or 0
    _rna = r.get("renta_neta_anual", 0) or 0
    _pay = r.get("payback_anos", 0) or 0
    _cpm = r.get("costo_por_m2_nave", 0) or 0
    _ctot= r.get("costo_total", 0) or 0

    def _kpi_cell(label, value, note):
        return [
            Paragraph(label, _s(f"kl{label[:4]}", fontSize=7, leading=9, textColor=GRAY,
                                 fontName="Helvetica-Bold", charSpace=1.5)),
            Paragraph(value, _s(f"kv{label[:4]}", fontSize=16, leading=19, textColor=NAV,
                                 fontName="Helvetica-Bold")),
            Paragraph(note,  _s(f"kn{label[:4]}", fontSize=7, leading=9,  textColor=GRAY)),
        ]

    kpi1 = _kpi_cell("COSTO TOTAL TERRENO", f"${_cta:,.0f}", "incl. alcabala")
    _cpm_terreno = (_ct / _at) if _at > 0 else 0
    kpi2 = _kpi_cell("COSTO POR m² TERRENO", f"${_cpm_terreno:,.0f}", "USD/m²")
    if _rna > 0 and _cta > 0:
        kpi3 = _kpi_cell("YIELD SOBRE TERRENO", f"{(_rna/_cta*100):.1f}%",
                          "renta neta anual / costo terreno")
    else:
        kpi3 = _kpi_cell("COSTO POR m² NAVE", f"${_cpm:,.0f}", "costo total / m² nave")
    if _pay > 0:
        kpi4 = _kpi_cell("PAYBACK", f"{_pay:.1f} años", "sobre costo total proyecto")
    else:
        kpi4 = _kpi_cell("COSTO TOTAL PROYECTO", f"${_ctot:,.0f}", "terreno + construcción")

    def _kpi_block(items):
        return Table(items, colWidths=[W*0.25]*4,
                     style=TableStyle([
                         ("BACKGROUND",    (0,0), (-1,-1), WHITE),
                         ("BOX",           (0,0), (0,-1),  0.5, BORD),
                         ("BOX",           (1,0), (1,-1),  0.5, BORD),
                         ("BOX",           (2,0), (2,-1),  0.5, BORD),
                         ("BOX",           (3,0), (3,-1),  0.5, BORD),
                         ("LINEABOVE",     (0,0), (0,0),   2, GOLD),
                         ("LINEABOVE",     (1,0), (1,0),   2, GOLD),
                         ("LINEABOVE",     (2,0), (2,0),   2, GOLD),
                         ("LINEABOVE",     (3,0), (3,0),   2, GOLD),
                         ("TOPPADDING",    (0,0), (-1,-1), 8),
                         ("BOTTOMPADDING", (0,0), (-1,-1), 8),
                         ("LEFTPADDING",   (0,0), (-1,-1), 8),
                         ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                         ("VALIGN",        (0,0), (-1,-1), "TOP"),
                         ("INNERGRID",     (0,0), (-1,-1), 0,   WHITE),
                     ]))

    # Each KPI column is a nested 3-row table (label, value, note)
    def _kpi_col_tbl(items):
        return Table([[p] for p in items],
                     colWidths=[W*0.25 - 16],
                     style=TableStyle([
                         ("LEFTPADDING",   (0,0), (-1,-1), 0),
                         ("RIGHTPADDING",  (0,0), (-1,-1), 0),
                         ("TOPPADDING",    (0,0), (-1,-1), 2),
                         ("BOTTOMPADDING", (0,0), (-1,-1), 2),
                     ]))

    kpi_row = Table(
        [[_kpi_col_tbl(kpi1), _kpi_col_tbl(kpi2), _kpi_col_tbl(kpi3), _kpi_col_tbl(kpi4)]],
        colWidths=[W*0.25]*4,
        style=TableStyle([
            ("BOX",        (0,0), (0,0), 0.5, BORD),
            ("BOX",        (1,0), (1,0), 0.5, BORD),
            ("BOX",        (2,0), (2,0), 0.5, BORD),
            ("BOX",        (3,0), (3,0), 0.5, BORD),
            ("LINEABOVE",  (0,0), (0,0), 2,   GOLD),
            ("LINEABOVE",  (1,0), (1,0), 2,   GOLD),
            ("LINEABOVE",  (2,0), (2,0), 2,   GOLD),
            ("LINEABOVE",  (3,0), (3,0), 2,   GOLD),
            ("TOPPADDING", (0,0), (-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ("LEFTPADDING",(0,0), (-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 6),
            ("BACKGROUND", (0,0), (-1,-1), WHITE),
            ("VALIGN",     (0,0), (-1,-1), "TOP"),
        ])
    )
    story.append(kpi_row)
    story.append(Spacer(1, 5*mm))

    # ── 4. Estructura de Costos ───────────────────────────────────────────────────
    story.append(_section_title("Estructura de Costos"))
    story.append(Spacer(1, 2*mm))

    _alc   = r.get("alcabala", 0) or 0
    _an    = r.get("area_nave", 0) or 0
    _al    = r.get("area_libre", 0) or 0
    _cnm2  = r.get("costo_nave_m2", 0) or 0
    _cnt   = r.get("costo_nave_total", 0) or 0
    _cplm2 = r.get("costo_piso_libre_m2", 0) or 0
    _cpl   = r.get("costo_pisos_libres", 0) or 0
    _sc    = r.get("soft_costs", 0) or 0
    _pind  = r.get("pct_indirectos", 5) or 5
    _ccs   = r.get("costo_construccion_soft", 0) or 0
    _ctot_safe = _ctot if _ctot > 0 else 1

    def _pct(v):
        return f"{v/_ctot_safe*100:.1f}%"

    altura_str = f"  ·  {altura_nave:.0f}m al hombro" if altura_nave and altura_nave > 0 else ""

    story.append(Paragraph("A  ·  TERRENO",
                            _s("subA", fontSize=8, leading=10, textColor=GOLD,
                               fontName="Helvetica-Bold")))
    story.append(Spacer(1, 1*mm))
    terreno_rows = [
        ["Terreno", f"${_ct:,.0f}", _pct(_ct)],
        ["Alcabala (3%)", f"${_alc:,.0f}", _pct(_alc)],
        ["TOTAL TERRENO", f"${_cta:,.0f}", _pct(_cta)],
    ]
    story.append(_data_table(["Concepto", "Monto USD", "% Total"],
                              terreno_rows,
                              [W*0.55, W*0.25, W*0.20],
                              bold_last=True))
    story.append(Spacer(1, 3*mm))

    story.append(Paragraph("B  ·  CONSTRUCCIÓN E IMPLEMENTACIÓN",
                            _s("subB", fontSize=8, leading=10, textColor=BLUE,
                               fontName="Helvetica-Bold")))
    story.append(Spacer(1, 1*mm))
    const_rows = [
        [f"Nave techada  {_an:,.0f} m²  ×  ${_cnm2:,.0f}/m²{altura_str}",
         f"${_cnt:,.0f}", _pct(_cnt)],
        [f"Piso descubierto / patios  {_al:,.0f} m²  ×  ${_cplm2:,.0f}/m²",
         f"${_cpl:,.0f}", _pct(_cpl)],
        [f"Costos indirectos ({_pind:.0f}%)",
         f"${_sc:,.0f}", _pct(_sc)],
        ["TOTAL CONSTRUCCIÓN", f"${_ccs:,.0f}", _pct(_ccs)],
    ]
    story.append(_data_table(["Concepto", "Monto USD", "% Total"],
                              const_rows,
                              [W*0.55, W*0.25, W*0.20],
                              bold_last=True))
    story.append(Spacer(1, 2*mm))

    # Summary total row
    total_tbl = Table(
        [[Paragraph("COSTO TOTAL PROYECTO (A + B)",
                    _s("totlbl", fontSize=8, leading=10, textColor=WHITE,
                       fontName="Helvetica-Bold")),
          Paragraph(f"${_ctot:,.0f}",
                    _s("totval", fontSize=9, leading=11, textColor=GOLD,
                       fontName="Helvetica-Bold", alignment=TA_RIGHT)),
          Paragraph("100%",
                    _s("totpct", fontSize=8, leading=10, textColor=WHITE,
                       alignment=TA_RIGHT))]],
        colWidths=[W*0.55, W*0.25, W*0.20],
        style=TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), NAV),
            ("TOPPADDING",    (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING",   (0,0), (-1,-1), 7),
            ("RIGHTPADDING",  (0,0), (-1,-1), 7),
            ("ALIGN",         (1,0), (-1,-1), "RIGHT"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ])
    )
    story.append(total_tbl)
    story.append(Spacer(1, 5*mm))

    # ── 5. Financiamiento ────────────────────────────────────────────────────────
    story.append(_section_title("Estructura de Financiamiento"))
    story.append(Spacer(1, 2*mm))

    _cpt  = r.get("capital_propio_terreno", 0) or 0
    _mct  = r.get("monto_credito_terreno", 0) or 0
    _qut  = r.get("cuota_terreno", 0) or 0
    _dpt  = r.get("dp_terreno_pct", 0) or 0
    _tast = r.get("tasa_terreno", 0) or 0
    _plzt = r.get("plazo_terreno", 0) or 0

    _cpc  = r.get("capital_propio_const", 0) or 0
    _mcc  = r.get("monto_credito_const", 0) or 0
    _quc  = r.get("cuota_const", 0) or 0
    _dpc  = r.get("dp_const_pct", 0) or 0
    _tasc = r.get("tasa_const", 0) or 0
    _plzc = r.get("plazo_const", 0) or 0

    def _fin_inner_rows(cap, cred, cuota, dp_pct, tasa, plazo, border_color):
        rows_data = [
            [Paragraph("Downpayment",    S_label),
             Paragraph(f"${cap:,.0f}",   S_body),
             Paragraph(f"{dp_pct:.0f}% al contado", S_small)],
            [Paragraph("Crédito",        S_label),
             Paragraph(f"${cred:,.0f}",  S_body),
             Paragraph(f"{100-dp_pct:.0f}% financiado", S_small)],
            [Paragraph("Cuota mensual",  S_label),
             Paragraph(f"${cuota:,.0f}/mes", S_body),
             Paragraph(f"{tasa:.1f}%  ·  {plazo} años", S_small)],
        ]
        ts = TableStyle([
            ("LINEBELOW",     (0,0), (-1,-1), 0.4, BORD),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("BACKGROUND",    (0,0), (-1,-1), CREAM),
            ("LINEBEFORE",    (0,0), (0,-1),  2, border_color),
        ])
        return Table(rows_data, colWidths=[(W*0.5-12)*f for f in [0.38,0.30,0.32]], style=ts)

    fin_box_A = Table(
        [[Paragraph("A  ·  TERRENO",
                    _s("faH", fontSize=8, leading=10, textColor=GOLD,
                       fontName="Helvetica-Bold"))],
         [_fin_inner_rows(_cpt, _mct, _qut, _dpt, _tast, _plzt, GOLD)]],
        colWidths=[W*0.5 - 4],
        style=TableStyle([
            ("BOX",           (0,0), (-1,-1), 0.5, BORD),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ])
    )
    fin_box_B = Table(
        [[Paragraph("B  ·  CONSTRUCCIÓN",
                    _s("fbH", fontSize=8, leading=10, textColor=BLUE,
                       fontName="Helvetica-Bold"))],
         [_fin_inner_rows(_cpc, _mcc, _quc, _dpc, _tasc, _plzc, BLUE)]],
        colWidths=[W*0.5 - 4],
        style=TableStyle([
            ("BOX",           (0,0), (-1,-1), 0.5, BORD),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ])
    )
    fin_pair = Table(
        [[fin_box_A, fin_box_B]],
        colWidths=[W*0.5 - 2, W*0.5 - 2],
        style=TableStyle([
            ("LEFTPADDING",  (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING",   (0,0), (-1,-1), 0),
            ("BOTTOMPADDING",(0,0), (-1,-1), 0),
            ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ])
    )
    story.append(fin_pair)
    story.append(Spacer(1, 5*mm))

    # ── 6. Gran Total ─────────────────────────────────────────────────────────────
    story.append(_section_title("Gran Total"))
    story.append(Spacer(1, 2*mm))

    _cp   = r.get("capital_propio", 0) or 0
    _mc   = r.get("monto_credito", 0) or 0
    _qm   = r.get("cuota_mensual", 0) or 0
    _pctc = r.get("pct_credito", 0) or 0

    def _gt_row(label, value, note, bg_row):
        return [
            Paragraph(label, _s(f"gtl{label[:3]}", fontSize=9, leading=11,
                                 textColor=WHITE, fontName="Helvetica-Bold")),
            Paragraph(value, _s(f"gtv{label[:3]}", fontSize=12, leading=14,
                                 textColor=GOLD, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
            Paragraph(note,  _s(f"gtn{label[:3]}", fontSize=8, leading=10,
                                 textColor=colors.HexColor("#AAAAAA"), alignment=TA_RIGHT)),
        ]

    gt_data = [
        _gt_row("CAPITAL PROPIO TOTAL", f"${_cp:,.0f}", "A + B downpayments", NAV),
        _gt_row("DEUDA TOTAL",          f"${_mc:,.0f}", f"{_pctc:.0f}% del proyecto", NAV),
        _gt_row("CUOTA MENSUAL TOTAL",  f"${_qm:,.0f}/mes", "Terreno + Obra", NAV),
    ]
    gt_tbl = Table(gt_data, colWidths=[W*0.42, W*0.28, W*0.30],
                   style=TableStyle([
                       ("BACKGROUND",    (0,0), (-1,-1), NAV),
                       ("LINEBELOW",     (0,0), (-1,-2), 0.5, colors.HexColor("#3A4D5F")),
                       ("TOPPADDING",    (0,0), (-1,-1), 8),
                       ("BOTTOMPADDING", (0,0), (-1,-1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 10),
                       ("RIGHTPADDING",  (0,0), (-1,-1), 10),
                       ("ALIGN",         (1,0), (-1,-1), "RIGHT"),
                       ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                   ]))
    story.append(gt_tbl)
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "El financiamiento puede estructurarse de forma independiente: solo terreno, solo construcción, "
        "o ambos según la estrategia del inversionista.",
        S_note))
    story.append(Spacer(1, 5*mm))

    # ── 7. Factibilidad ───────────────────────────────────────────────────────────
    if factibilidad:
        story.append(_section_title("Factibilidad Técnica y Legal"))
        story.append(Spacer(1, 2*mm))
        sg = (factibilidad.get("semaforo_global") or "amarillo").lower()
        sem_bg_map  = {"verde": "#E8F5EE", "amarillo": "#FFF8EE", "rojo": "#FFF0F0"}
        sem_col_map = {"verde": "#1A4731", "amarillo": "#7A5500", "rojo": "#8B1A1A"}
        sem_border  = {"verde": colors.HexColor("#1A4731"),
                       "amarillo": colors.HexColor("#7A5500"),
                       "rojo":     colors.HexColor("#8B1A1A")}
        label_map   = {"verde": "SIN ALERTAS CRÍTICAS", "amarillo": "OBSERVACIONES",
                       "rojo": "ALERTAS CRÍTICAS"}
        fac_bg  = colors.HexColor(sem_bg_map.get(sg, "#FFF8EE"))
        fac_brd = sem_border.get(sg, colors.HexColor("#7A5500"))
        fac_txt = colors.HexColor(sem_col_map.get(sg, "#7A5500"))

        fac_inner = [
            Paragraph(label_map.get(sg, "OBSERVACIONES"),
                      _s("faclbl", fontSize=9, leading=11, textColor=fac_txt,
                         fontName="Helvetica-Bold")),
        ]
        rt = factibilidad.get("resumen_tecnico") or ""
        rl = factibilidad.get("resumen_legal") or ""
        if rt:
            fac_inner.append(Paragraph(rt, _s("facrt", fontSize=9, leading=13, textColor=fac_txt)))
        if rl:
            fac_inner.append(Paragraph(rl, _s("facrl", fontSize=9, leading=13, textColor=fac_txt)))

        fac_tbl = Table(
            [[fac_inner]],
            colWidths=[W],
            style=TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), fac_bg),
                ("LINEBEFORE",   (0,0), (0,-1),  3, fac_brd),
                ("TOPPADDING",   (0,0), (-1,-1), 8),
                ("BOTTOMPADDING",(0,0), (-1,-1), 8),
                ("LEFTPADDING",  (0,0), (-1,-1), 10),
                ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ])
        )
        story.append(fac_tbl)
        story.append(Spacer(1, 5*mm))

    # ── 8. Footer ─────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width=W, thickness=0.5, color=BORD, spaceAfter=4))
    footer_tbl = Table(
        [[Paragraph("© Osterling Advisory  ·  Acceso restringido  ·  Confidencial",
                    _s("ftl", fontSize=7, leading=9, textColor=GRAY)),
          Paragraph("FACTIS — Plataforma Analítica Inmobiliaria",
                    _s("ftr", fontSize=7, leading=9, textColor=GRAY, alignment=TA_RIGHT))]],
        colWidths=[W*0.55, W*0.45],
        style=TableStyle([("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
                          ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)])
    )
    story.append(footer_tbl)
    story.append(Spacer(1, 5))
    story.append(Paragraph(
        "<b>NOTA:</b> Esta IA de Análisis Inmobiliario debe utilizarse como herramienta complementaria "
        "al criterio profesional, permitiendo obtener resultados preliminares de manera rápida. "
        "El profesional podrá definir tipologías, distribución y modificaciones pertinentes. "
        "La IA irá alineándose con la visión del profesional a medida que se retroalimenta con sus decisiones.",
        _s("disc_ind", fontSize=6, leading=9, textColor=colors.HexColor("#A89880"),
           fontName="Helvetica", leftIndent=0, rightIndent=0)))

    doc.build(story)
    return buf.getvalue()


def generar_propuesta_html(
    tipo: str,
    propietario: str,
    params: dict,
    financ: dict | None,
    legal: dict | None,
    comps_sunarp: list,
    precio_oferta: float,
    moneda_oferta: str,
    condiciones: str,
    plazo_respuesta: int,
    fecha: str,
    representante: str = "Enrique Osterling",
    cargo: str = "Director General",
    # Estructura de compra/pago
    tiene_opcion: bool = True,
    dias_opcion: int = 90,
    pct_opcion: float = 0.0,
    pct_minuta: float = 20.0,
    condicion_minuta: str = "Aprobación del anteproyecto por la Municipalidad",
    condicion_escritura: str = "Desocupación y entrega del inmueble libre de cargas",
) -> str:
    p  = params or {}
    r  = (financ or {}).get("resumen", {})
    lg = legal or {}
    NAV  = "#1E2D3D"
    GOLD = "#B8904A"
    tc   = TIPO_CAMBIO

    precio_pen = round(precio_oferta * tc, 0) if moneda_oferta == "USD" else precio_oferta
    precio_usd = precio_oferta if moneda_oferta == "USD" else round(precio_oferta / tc, 0)
    area       = p.get("area_terreno_m2") or p.get("area_m2") or 0
    pm2_usd    = round(precio_usd / area, 0) if area > 0 else 0
    ubicacion  = p.get("ubicacion") or p.get("direccion") or "—"
    distrito   = p.get("distrito") or ""
    zona_res   = p.get("zona_residencial") or p.get("zonificacion") or "—"
    partida    = lg.get("partida_numero") or "—"
    _props_raw = lg.get("propietarios_partida") or []
    propietario_reg = (", ".join(
        (x.get("nombre", str(x)) if isinstance(x, dict) else str(x))
        for x in _props_raw if x
    )) or propietario

    # Comparables SUNARP
    comp_rows = ""
    precios_cierre = []
    for rc in comps_sunarp:
        ut  = rc.get("ultima_transferencia") or {}
        p_v = ut.get("precio")
        mon = ut.get("moneda", "USD")
        pm2 = rc.get("precio_m2_estimado")
        # Guard: ensure p_v and pm2 are numeric before arithmetic
        try:
            p_v = float(p_v) if p_v is not None else None
        except (TypeError, ValueError):
            p_v = None
        try:
            pm2 = float(pm2) if pm2 is not None else None
        except (TypeError, ValueError):
            pm2 = None
        if p_v and mon == "PEN":
            p_v = round(p_v / tc, 0)
            pm2 = round(pm2 / tc, 0) if pm2 else None
        if pm2:
            precios_cierre.append(pm2)
        _pv_str  = f"${p_v:,.0f}"  if p_v  else "—"
        _pm2_str = f"${pm2:,.0f}/m²" if pm2 else "—"
        comp_rows += (
            "<tr>"
            f'<td style="padding:6px 10px;border:1px solid #E0DDD8;font-size:11px;">{rc.get("descripcion_predio","—")}</td>'
            f'<td style="padding:6px 10px;border:1px solid #E0DDD8;font-size:11px;text-align:center;">{rc.get("area_m2","—")}</td>'
            f'<td style="padding:6px 10px;border:1px solid #E0DDD8;font-size:11px;text-align:center;">{_pv_str}</td>'
            f'<td style="padding:6px 10px;border:1px solid #E0DDD8;font-size:11px;text-align:center;">{_pm2_str}</td>'
            f'<td style="padding:6px 10px;border:1px solid #E0DDD8;font-size:11px;text-align:center;">{ut.get("fecha","—")}</td>'
            "</tr>"
        )
    med_cierre = round(sum(precios_cierre) / len(precios_cierre), 0) if precios_cierre else None
    comp_section = ""
    if comp_rows:
        comp_section = f"""
        <div style="margin:24px 0 0;">
          <div style="font-size:10px;font-weight:700;color:{NAV};letter-spacing:2px;text-transform:uppercase;
                      border-bottom:2px solid {GOLD};padding-bottom:4px;margin-bottom:12px;">
            III. SUSTENTO DE VALOR — COMPARABLES SUNARP
          </div>
          <p style="font-size:11px;color:#555;margin-bottom:10px;">
            Los siguientes precios de cierre, obtenidos de partidas registrales SUNARP, respaldan el valor propuesto:
          </p>
          <table style="width:100%;border-collapse:collapse;margin-bottom:10px;">
            <thead>
              <tr style="background:{NAV};">
                <th style="padding:7px 10px;color:#fff;font-size:10px;font-weight:600;text-align:left;">Predio comparable</th>
                <th style="padding:7px 10px;color:#fff;font-size:10px;font-weight:600;text-align:center;">Área m²</th>
                <th style="padding:7px 10px;color:#fff;font-size:10px;font-weight:600;text-align:center;">Precio cierre</th>
                <th style="padding:7px 10px;color:#fff;font-size:10px;font-weight:600;text-align:center;">USD/m²</th>
                <th style="padding:7px 10px;color:#fff;font-size:10px;font-weight:600;text-align:center;">Fecha</th>
              </tr>
            </thead>
            <tbody>{comp_rows}</tbody>
          </table>
          {"<p style='font-size:11px;font-weight:700;color:" + NAV + ";'>Precio de cierre promedio (SUNARP): $" + f"{med_cierre:,.0f}/m²</p>" if med_cierre else ""}
        </div>"""

    # ── Estructura de compra / pago ───────────────────────────────────────────
    _pct_escritura = round(100.0 - pct_minuta - (pct_opcion if tiene_opcion else 0.0), 1)
    _monto_opcion   = round(precio_usd * pct_opcion / 100) if tiene_opcion and pct_opcion > 0 else 0
    _monto_minuta   = round(precio_usd * pct_minuta / 100)
    _monto_escritura = round(precio_usd * _pct_escritura / 100)

    if tipo == "Compra":
        _pago_rows = ""
        _num = 1
        if tiene_opcion:
            _pago_rows += f"""
              <tr style="background:#F8F5F0;">
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:700;color:{NAV};">{_num}. Opción de Compra</td>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;text-align:center;font-weight:600;color:{GOLD};">
                  {f"USD {_monto_opcion:,.0f}" if _monto_opcion > 0 else "Sin cargo"}{f" ({pct_opcion:.0f}% del precio)" if _monto_opcion > 0 else ""}
                </td>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;color:#555;">
                  A la firma del contrato de opción. Plazo: <strong>{dias_opcion} días calendario</strong>
                  {"para aprobación del anteproyecto municipal." if dias_opcion > 0 else "."}
                  {" Monto imputable al precio total." if _monto_opcion > 0 else ""}
                </td>
              </tr>"""
            _num += 1
        _pago_rows += f"""
              <tr{"" if _num % 2 == 0 else ' style="background:#F8F5F0;"'}>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:700;color:{NAV};">{_num}. Pago Inicial — Minuta</td>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;text-align:center;font-weight:600;color:{GOLD};">
                  USD {_monto_minuta:,.0f} ({pct_minuta:.0f}%)
                </td>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;color:#555;">
                  A la firma de la Minuta de Compraventa ante Notario. Condición: <strong>{condicion_minuta}</strong>.
                </td>
              </tr>"""
        _num += 1
        _pago_rows += f"""
              <tr style="background:#F8F5F0;">
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:700;color:{NAV};">{_num}. Saldo — Escritura Pública</td>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;text-align:center;font-weight:700;color:{NAV};">
                  USD {_monto_escritura:,.0f} ({_pct_escritura:.0f}%)
                </td>
                <td style="padding:9px 12px;border:1px solid #E0DDD8;font-size:11px;color:#555;">
                  A la firma de la Escritura Pública e inscripción en SUNARP. Condición: <strong>{condicion_escritura}</strong>.
                </td>
              </tr>"""

        pago_section = f"""
        <div style="margin:24px 0 0;">
          <div style="font-size:10px;font-weight:700;color:{NAV};letter-spacing:2px;text-transform:uppercase;
                      border-bottom:2px solid {GOLD};padding-bottom:4px;margin-bottom:12px;">
            {"IV" if comp_section else "III"}. ESTRUCTURA DE PAGO
          </div>
          <p style="font-size:11px;color:#555;margin-bottom:10px;">
            La presente oferta contempla la siguiente estructura de pago sobre el precio total de
            <strong>USD {precio_usd:,.0f}</strong>:
          </p>
          <table style="width:100%;border-collapse:collapse;">
            <thead>
              <tr style="background:{NAV};">
                <th style="padding:8px 12px;color:#fff;font-size:10px;font-weight:600;width:22%;">Etapa</th>
                <th style="padding:8px 12px;color:#fff;font-size:10px;font-weight:600;text-align:center;width:22%;">Monto</th>
                <th style="padding:8px 12px;color:#fff;font-size:10px;font-weight:600;width:56%;">Condición / Oportunidad de pago</th>
              </tr>
            </thead>
            <tbody>
              {_pago_rows}
            </tbody>
            <tfoot>
              <tr style="background:{NAV};">
                <td style="padding:8px 12px;font-size:11px;font-weight:700;color:#fff;border:1px solid #334;">TOTAL</td>
                <td style="padding:8px 12px;font-size:12px;font-weight:800;color:{GOLD};text-align:center;border:1px solid #334;">
                  USD {precio_usd:,.0f}
                </td>
                <td style="padding:8px 12px;font-size:10px;color:#aaa;border:1px solid #334;">
                  TC referencial: S/. {tc} por USD
                </td>
              </tr>
            </tfoot>
          </table>
        </div>"""
    else:
        pago_section = ""
    fin_section = pago_section

    tipo_label = "COMPRA" if tipo == "Compra" else "ARRENDAMIENTO"
    renta_note = ""
    if tipo == "Arrendamiento":
        renta_note = f"""
        <tr style="background:#F8F5F0;">
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Renta mensual ofertada</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;" colspan="3">
            USD {precio_usd:,.0f} + IGV &nbsp;|&nbsp; USD {pm2_usd:,.2f}/m²/mes
          </td>
        </tr>"""
    else:
        renta_note = f"""
        <tr style="background:#F8F5F0;">
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Precio ofertado</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;" colspan="3">
            <strong>USD {precio_usd:,.0f}</strong> &nbsp;|&nbsp; USD {pm2_usd:,.0f}/m²
          </td>
        </tr>"""

    condiciones_html = "".join(
        f'<li style="margin-bottom:5px;font-size:11px;color:#444;">{c.strip()}</li>'
        for c in (condiciones or "").split("\n") if c.strip()
    ) or '<li style="font-size:11px;color:#888;">—</li>'

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Propuesta de {tipo_label} — {ubicacion}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:'Inter',sans-serif; background:#F4F1EC; color:#1E2D3D; }}
  .page {{ max-width:780px; margin:40px auto; background:#fff; padding:52px 56px;
           box-shadow:0 4px 32px rgba(0,0,0,0.10); }}
  @media print {{ body{{background:#fff}} .page{{box-shadow:none;margin:0;padding:40px}} }}
</style>
</head>
<body>
<div class="page">

  <!-- Header -->
  <div style="display:flex;justify-content:space-between;align-items:flex-start;
              border-bottom:3px solid {NAV};padding-bottom:20px;margin-bottom:28px;">
    <div>
      <div style="font-size:9px;letter-spacing:4px;text-transform:uppercase;color:{GOLD};
                  font-weight:700;margin-bottom:6px;">Osterling Advisory</div>
      <div style="font-size:26px;font-weight:800;color:{NAV};letter-spacing:-0.5px;">FACTIS</div>
      <div style="font-size:9px;letter-spacing:2px;text-transform:uppercase;color:#888;margin-top:2px;">
        IA de Análisis Inmobiliario
      </div>
    </div>
    <div style="text-align:right;">
      <div style="font-size:10px;color:#888;">Lima, Perú &nbsp;|&nbsp; {fecha}</div>
      <div style="font-size:10px;color:#888;margin-top:4px;">eosterling@grupoosterling.com</div>
      <div style="display:inline-block;margin-top:10px;background:{GOLD};color:#fff;
                  font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;
                  padding:5px 14px;border-radius:2px;">
        PROPUESTA DE {tipo_label}
      </div>
    </div>
  </div>

  <!-- Destinatario -->
  <div style="margin-bottom:24px;">
    <div style="font-size:12px;font-weight:600;color:{NAV};">Señor(es)</div>
    <div style="font-size:13px;font-weight:700;color:{NAV};margin-top:2px;">{propietario or propietario_reg}</div>
    <div style="font-size:11px;color:#666;margin-top:2px;">Propietario(s) del inmueble</div>
    <div style="margin-top:12px;font-size:11px;color:#444;line-height:1.7;">
      Por medio de la presente, <strong>Osterling Advisory</strong> — en representación de su cliente —
      tiene el agrado de presentar su propuesta formal de <strong>{tipo.lower()}</strong> para el
      inmueble de su propiedad, con las condiciones que se detallan a continuación.
    </div>
  </div>

  <!-- I. Identificación -->
  <div style="margin:24px 0 0;">
    <div style="font-size:10px;font-weight:700;color:{NAV};letter-spacing:2px;text-transform:uppercase;
                border-bottom:2px solid {GOLD};padding-bottom:4px;margin-bottom:12px;">
      I. IDENTIFICACIÓN DEL INMUEBLE
    </div>
    <table style="width:100%;border-collapse:collapse;">
      <tbody>
        <tr style="background:#F8F5F0;">
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;
                     color:{NAV};width:35%;">Ubicación</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;">{ubicacion}{(', ' + distrito) if distrito else ''}</td>
        </tr>
        <tr>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Partida Registral</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;">{partida}</td>
        </tr>
        <tr style="background:#F8F5F0;">
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Área del terreno</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;">{area:,.1f} m²</td>
        </tr>
        <tr>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Zonificación</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;">{zona_res}</td>
        </tr>
        <tr style="background:#F8F5F0;">
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Propietario registral</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;">{propietario_reg}</td>
        </tr>
      </tbody>
    </table>
  </div>

  <!-- II. Propuesta económica -->
  <div style="margin:24px 0 0;">
    <div style="font-size:10px;font-weight:700;color:{NAV};letter-spacing:2px;text-transform:uppercase;
                border-bottom:2px solid {GOLD};padding-bottom:4px;margin-bottom:12px;">
      II. PROPUESTA ECONÓMICA Y CONDICIONES
    </div>
    <table style="width:100%;border-collapse:collapse;margin-bottom:14px;">
      <tbody>
        {renta_note}
        <tr{"" if tipo == "Arrendamiento" else ' style="background:#F8F5F0;"'}>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;font-weight:600;color:{NAV};">Plazo de respuesta</td>
          <td style="padding:8px 12px;border:1px solid #E0DDD8;font-size:11px;" colspan="3">{plazo_respuesta} días calendario desde la recepción de la presente</td>
        </tr>
      </tbody>
    </table>
    <div style="font-size:10px;font-weight:700;color:{NAV};margin-bottom:8px;text-transform:uppercase;
                letter-spacing:1px;">Condiciones de la operación</div>
    <ul style="padding-left:18px;margin:0;">{condiciones_html}</ul>
  </div>

  {comp_section}
  {fin_section}

  <!-- Cierre -->
  <div style="margin-top:40px;padding-top:20px;border-top:1px solid #E0DDD8;">
    <p style="font-size:11px;color:#444;line-height:1.7;margin-bottom:24px;">
      Quedamos a su disposición para cualquier consulta o coordinación adicional.
      La presente propuesta tiene carácter indicativo y no genera obligación legal hasta la
      suscripción de un documento de compraventa / arrendamiento definitivo.
    </p>
    <div style="display:flex;justify-content:space-between;align-items:flex-end;">
      <div>
        <div style="font-size:12px;font-weight:700;color:{NAV};">{representante}</div>
        <div style="font-size:10px;color:#666;">{cargo}</div>
        <div style="font-size:10px;color:{GOLD};margin-top:2px;">Osterling Advisory</div>
      </div>
      <div style="text-align:right;font-size:9px;color:#AAA;">
        Generado con FACTIS · Osterling Advisory<br>
        eosterling@grupoosterling.com · Lima, Perú
      </div>
    </div>
  </div>

</div>
</body>
</html>"""


def generar_propuesta_pdf(
    tipo: str,
    propietario: str,
    params: dict,
    financ: dict | None,
    legal: dict | None,
    comps_sunarp: list,
    precio_oferta: float,
    moneda_oferta: str,
    condiciones: str,
    plazo_respuesta: int,
    fecha: str,
    representante: str = "Enrique Osterling",
    cargo: str = "Director General",
    tiene_opcion: bool = True,
    dias_opcion: int = 90,
    pct_opcion: float = 0.0,
    pct_minuta: float = 20.0,
    condicion_minuta: str = "Aprobación del anteproyecto por la Municipalidad",
    condicion_escritura: str = "Desocupación y entrega del inmueble libre de cargas",
) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable, KeepTogether)
    import io

    NAV  = colors.HexColor("#1E2D3D")
    GOLD = colors.HexColor("#B8904A")
    CREAM= colors.HexColor("#F8F5F0")
    BORD = colors.HexColor("#E0DDD8")
    GRAY = colors.HexColor("#666666")

    p   = params or {}
    r   = (financ or {}).get("resumen", {})
    lg  = legal or {}
    tc  = TIPO_CAMBIO

    precio_usd  = precio_oferta if moneda_oferta == "USD" else round(precio_oferta / tc, 0)
    precio_pen  = round(precio_usd * tc, 0)
    area        = p.get("area_terreno_m2") or p.get("area_m2") or 0
    pm2_usd     = round(precio_usd / area, 0) if area > 0 else 0
    ubicacion   = p.get("ubicacion") or p.get("direccion") or "—"
    distrito    = p.get("distrito") or ""
    zona_res    = p.get("zona_residencial") or p.get("zonificacion") or "—"
    partida     = lg.get("partida_numero") or "—"
    _props_raw2 = lg.get("propietarios_partida") or []
    propietario_reg = (", ".join(
        (x.get("nombre", str(x)) if isinstance(x, dict) else str(x))
        for x in _props_raw2 if x
    )) or propietario
    tipo_label  = "COMPRA" if tipo == "Compra" else "ARRENDAMIENTO"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=18*mm)

    sty = getSampleStyleSheet()
    def _s(name, **kw):
        base = sty["Normal"]
        return ParagraphStyle(name, parent=base, **kw)

    S_body   = _s("body",   fontSize=9,  leading=13, textColor=NAV)
    S_small  = _s("small",  fontSize=8,  leading=11, textColor=GRAY)
    S_label  = _s("label",  fontSize=8,  leading=10, textColor=NAV, fontName="Helvetica-Bold")
    S_gold   = _s("gold",   fontSize=9,  leading=13, textColor=GOLD, fontName="Helvetica-Bold")
    S_h1     = _s("h1",     fontSize=22, leading=26, textColor=NAV, fontName="Helvetica-Bold")
    S_kicker = _s("kicker", fontSize=7,  leading=9,  textColor=GOLD, fontName="Helvetica-Bold",
                  charSpace=3)
    S_sub    = _s("sub",    fontSize=7,  leading=9,  textColor=GRAY, charSpace=2)
    S_sec    = _s("sec",    fontSize=7,  leading=9,  textColor=NAV,  fontName="Helvetica-Bold",
                  charSpace=2, spaceAfter=4)
    S_intro  = _s("intro",  fontSize=9,  leading=14, textColor=colors.HexColor("#444444"))
    S_cond   = _s("cond",   fontSize=8,  leading=12, textColor=colors.HexColor("#444444"),
                  leftIndent=8, bulletIndent=0)

    W = doc.width

    story = []

    # ── Header ───────────────────────────────────────────────────────────────────
    hdr_left = [
        [Paragraph("OSTERLING ADVISORY", _s("kl", fontSize=7, leading=9, textColor=GOLD,
                                             fontName="Helvetica-Bold", charSpace=3))],
        [Paragraph("FACTIS", _s("fct", fontSize=20, leading=22, textColor=NAV,
                                 fontName="Helvetica-Bold"))],
        [Paragraph("IA DE ANÁLISIS INMOBILIARIO", _s("sub2", fontSize=6, leading=8,
                                                              textColor=GRAY, charSpace=1.5))],
    ]
    hdr_right = [
        [Paragraph(f"Lima, Perú  |  {fecha}", S_small)],
        [Paragraph("eosterling@grupoosterling.com", S_small)],
        [Table([[Paragraph(f"PROPUESTA DE {tipo_label}", _s("badge", fontSize=7, leading=9,
                                                             textColor=colors.white,
                                                             fontName="Helvetica-Bold", charSpace=1.5,
                                                             alignment=TA_CENTER))]],
               colWidths=[W*0.38],
               style=TableStyle([("BACKGROUND", (0,0), (-1,-1), GOLD),
                                  ("TOPPADDING",  (0,0), (-1,-1), 5),
                                  ("BOTTOMPADDING",(0,0),(-1,-1), 5),
                                  ("LEFTPADDING", (0,0), (-1,-1), 8),
                                  ("RIGHTPADDING",(0,0), (-1,-1), 8)]))],
    ]
    tbl_hdr = Table(
        [[Table(hdr_left,  colWidths=[W*0.55], style=TableStyle([("LEFTPADDING",(0,0),(-1,-1),0),("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)])),
          Table(hdr_right, colWidths=[W*0.45], style=TableStyle([("ALIGN",(0,0),(-1,-1),"RIGHT"),("LEFTPADDING",(0,0),(-1,-1),0),("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)]))]],
        colWidths=[W*0.55, W*0.45],
        style=TableStyle([
            ("LINEBELOW", (0,0), (-1,-1), 2, NAV),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
        ])
    )
    story.append(tbl_hdr)
    story.append(Spacer(1, 10*mm))

    # ── Destinatario ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Señor(es)", S_small))
    story.append(Paragraph(propietario or propietario_reg,
                            _s("rcpt", fontSize=11, leading=14, textColor=NAV, fontName="Helvetica-Bold")))
    story.append(Paragraph("Propietario(s) del inmueble", S_small))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(
        f"Por medio de la presente, <b>Osterling Advisory</b> — en representación de su cliente — "
        f"tiene el agrado de presentar su propuesta formal de <b>{tipo.lower()}</b> para el "
        f"inmueble de su propiedad, con las condiciones que se detallan a continuación.",
        S_intro))
    story.append(Spacer(1, 6*mm))

    def _section_title(txt):
        return KeepTogether([
            Paragraph((txt or "").upper(), _s("stit", fontSize=7, leading=9, textColor=NAV,
                                              fontName="Helvetica-Bold", charSpace=2)),
            HRFlowable(width=W, thickness=1.5, color=GOLD, spaceAfter=6),
        ])

    def _kv_table(rows, col_w=None):
        if not rows:
            return Spacer(1, 1)
        cw = col_w or [W*0.32, W*0.68]
        data = []
        for i, (k, v) in enumerate(rows):
            data.append([Paragraph(k, S_label), Paragraph(str(v), S_body)])
        ts = TableStyle([
            ("BACKGROUND",    (0,0), (0,-1), CREAM),
            ("LINEBELOW",     (0,0), (-1,-1), 0.5, BORD),
            ("LINEABOVE",     (0,0), (-1,0),  0.5, BORD),
            ("LINEBEFORE",    (0,0), (0,-1),  0.5, BORD),
            ("LINEAFTER",     (1,0), (1,-1),  0.5, BORD),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("RIGHTPADDING",  (0,0), (-1,-1), 8),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ])
        for i in range(len(rows)):
            if i % 2 == 0:
                ts.add("BACKGROUND", (1, i), (1, i), colors.white)
            else:
                ts.add("BACKGROUND", (1, i), (1, i), CREAM)
        return Table(data, colWidths=cw, style=ts, repeatRows=0)

    # ── I. Identificación ────────────────────────────────────────────────────────
    story.append(_section_title("I. Identificación del Inmueble"))
    ub_txt = ubicacion + (f", {distrito}" if distrito else "")
    story.append(_kv_table([
        ("Ubicación",          ub_txt),
        ("Partida Registral",  partida),
        ("Área del terreno",   f"{area:,.1f} m²"),
        ("Zonificación",       zona_res),
        ("Propietario registral", propietario_reg),
    ]))
    story.append(Spacer(1, 6*mm))

    # ── II. Propuesta Económica ────────────────────────────────────────────────
    story.append(_section_title("II. Propuesta Económica y Condiciones"))
    if tipo == "Compra":
        precio_str = f"USD {precio_usd:,.0f}  |  USD {pm2_usd:,.0f}/m²"
        econ_rows  = [("Precio ofertado", precio_str)]
    else:
        precio_str = f"USD {precio_usd:,.0f} + IGV  |  USD {pm2_usd:,.2f}/m²/mes"
        econ_rows  = [("Renta mensual ofertada", precio_str)]
    econ_rows.append(("Plazo de respuesta",
                       f"{plazo_respuesta} días calendario desde la recepción de la presente"))
    story.append(_kv_table(econ_rows))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Condiciones de la operación", S_label))
    story.append(Spacer(1, 2*mm))
    for c in (condiciones or "").split("\n"):
        if c.strip():
            story.append(Paragraph(f"• {c.strip()}", S_cond))
    story.append(Spacer(1, 6*mm))

    # ── III. Comparables SUNARP (optional) ───────────────────────────────────────
    comp_rows_data = []
    precios_cierre = []
    for rc in (comps_sunarp or []):
        ut  = rc.get("ultima_transferencia") or {}
        p_v = ut.get("precio")
        mon = ut.get("moneda", "USD")
        pm2 = rc.get("precio_m2_estimado")
        try:
            p_v = float(p_v) if p_v is not None else None
        except (TypeError, ValueError):
            p_v = None
        try:
            pm2 = float(pm2) if pm2 is not None else None
        except (TypeError, ValueError):
            pm2 = None
        if p_v and mon == "PEN":
            p_v = round(p_v / tc, 0)
            pm2 = round(pm2 / tc, 0) if pm2 else None
        if pm2:
            precios_cierre.append(pm2)
        pv_s  = f"${p_v:,.0f}"  if p_v  else "—"
        pm2_s = f"${pm2:,.0f}/m²" if pm2 else "—"
        comp_rows_data.append([
            Paragraph(rc.get("descripcion_predio", "—"), S_small),
            Paragraph(str(rc.get("area_m2", "—")), S_small),
            Paragraph(pv_s,  S_small),
            Paragraph(pm2_s, S_small),
            Paragraph(ut.get("fecha", "—"), S_small),
        ])
    if comp_rows_data:
        story.append(_section_title("III. Sustento de Valor — Comparables SUNARP"))
        story.append(Paragraph(
            "Los siguientes precios de cierre, obtenidos de partidas registrales SUNARP, respaldan el valor propuesto:",
            S_intro))
        story.append(Spacer(1, 3*mm))
        hdr_comp = [[Paragraph(h, _s("ch", fontSize=7, leading=9, textColor=colors.white,
                                      fontName="Helvetica-Bold"))
                     for h in ["Predio comparable", "Área m²", "Precio cierre", "USD/m²", "Fecha"]]]
        comp_tbl = Table(hdr_comp + comp_rows_data,
                         colWidths=[W*0.38, W*0.1, W*0.16, W*0.16, W*0.2],
                         style=TableStyle([
                             ("BACKGROUND",    (0,0), (-1,0),  NAV),
                             ("LINEBELOW",     (0,0), (-1,-1), 0.5, BORD),
                             ("LINEBEFORE",    (0,0), (0,-1),  0.5, BORD),
                             ("LINEAFTER",    (-1,0), (-1,-1), 0.5, BORD),
                             ("TOPPADDING",    (0,0), (-1,-1), 5),
                             ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                             ("LEFTPADDING",   (0,0), (-1,-1), 7),
                             ("RIGHTPADDING",  (0,0), (-1,-1), 7),
                             ("ALIGN",         (1,0), (-1,-1), "CENTER"),
                             ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, CREAM]),
                         ]))
        story.append(comp_tbl)
        if precios_cierre:
            med = round(sum(precios_cierre)/len(precios_cierre), 0)
            story.append(Spacer(1, 3*mm))
            story.append(Paragraph(
                f"Precio de cierre promedio (SUNARP): <b>${med:,.0f}/m²</b>", S_body))
        story.append(Spacer(1, 6*mm))
        sec_num = "IV"
    else:
        sec_num = "III"

    # ── Estructura de Pago (solo Compra) ─────────────────────────────────────────
    if tipo == "Compra":
        _pct_escritura  = round(100.0 - pct_minuta - (pct_opcion if tiene_opcion else 0.0), 1)
        _monto_opcion   = round(precio_usd * pct_opcion / 100) if tiene_opcion and pct_opcion > 0 else 0
        _monto_minuta   = round(precio_usd * pct_minuta / 100)
        _monto_escritura= round(precio_usd * _pct_escritura / 100)

        story.append(_section_title(f"{sec_num}. Estructura de Pago"))
        story.append(Paragraph(
            f"La presente oferta contempla la siguiente estructura de pago sobre el precio total de "
            f"<b>USD {precio_usd:,.0f}</b>:", S_intro))
        story.append(Spacer(1, 3*mm))

        pago_hdr = [[Paragraph(h, _s("ph", fontSize=7, leading=9, textColor=colors.white,
                                      fontName="Helvetica-Bold"))
                     for h in ["Etapa", "Monto", "Condición / Oportunidad de pago"]]]
        pago_data = []
        _num = 1
        if tiene_opcion:
            monto_s = (f"USD {_monto_opcion:,.0f}\n({pct_opcion:.0f}% del precio)"
                       if _monto_opcion > 0 else "Sin cargo")
            cond_s = (f"A la firma del contrato de opción. Plazo: {dias_opcion} días calendario."
                      + (" Monto imputable al precio total." if _monto_opcion > 0 else ""))
            pago_data.append([
                Paragraph(f"{_num}. Opción de Compra", S_label),
                Paragraph(monto_s, S_gold),
                Paragraph(cond_s, S_small),
            ])
            _num += 1
        pago_data.append([
            Paragraph(f"{_num}. Pago Inicial — Minuta", S_label),
            Paragraph(f"USD {_monto_minuta:,.0f}\n({pct_minuta:.0f}%)", S_gold),
            Paragraph(f"A la firma de la Minuta de Compraventa ante Notario. Condición: <b>{condicion_minuta}</b>.", S_small),
        ])
        _num += 1
        pago_data.append([
            Paragraph(f"{_num}. Saldo — Escritura Pública", S_label),
            Paragraph(f"USD {_monto_escritura:,.0f}\n({_pct_escritura:.0f}%)",
                      _s("esc", fontSize=9, leading=13, textColor=NAV, fontName="Helvetica-Bold")),
            Paragraph(f"A la firma de la Escritura Pública e inscripción en SUNARP. Condición: <b>{condicion_escritura}</b>.", S_small),
        ])
        total_row = [
            Paragraph("TOTAL", _s("tot", fontSize=9, leading=11, textColor=colors.white,
                                   fontName="Helvetica-Bold")),
            Paragraph(f"USD {precio_usd:,.0f}", _s("totv", fontSize=10, leading=12,
                                                     textColor=GOLD, fontName="Helvetica-Bold")),
            Paragraph(f"TC referencial: S/. {tc} por USD",
                      _s("tots", fontSize=7, leading=9, textColor=colors.HexColor("#AAAAAA"))),
        ]
        pago_tbl_style = TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  NAV),
            ("BACKGROUND",    (0,-1),(-1,-1), NAV),
            ("LINEBELOW",     (0,0), (-1,-2), 0.5, BORD),
            ("LINEBEFORE",    (0,0), (0,-1),  0.5, BORD),
            ("LINEAFTER",    (-1,0), (-1,-1), 0.5, BORD),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("RIGHTPADDING",  (0,0), (-1,-1), 8),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ])
        for i in range(len(pago_data)):
            if i % 2 == 0:
                pago_tbl_style.add("BACKGROUND", (0, i+1), (-1, i+1), colors.white)
            else:
                pago_tbl_style.add("BACKGROUND", (0, i+1), (-1, i+1), CREAM)
        pago_tbl = Table(pago_hdr + pago_data + [total_row],
                         colWidths=[W*0.28, W*0.22, W*0.50],
                         style=pago_tbl_style)
        story.append(pago_tbl)
        story.append(Spacer(1, 6*mm))

    # ── Closing ───────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width=W, thickness=0.5, color=BORD, spaceAfter=5))
    story.append(Paragraph(
        "Quedamos a su disposición para cualquier consulta o coordinación adicional. "
        "La presente propuesta tiene carácter indicativo y no genera obligación legal hasta la "
        "suscripción de un documento de compraventa / arrendamiento definitivo.",
        S_intro))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        "<i>NOTA: Esta propuesta ha sido elaborada con el apoyo de la IA de Análisis Inmobiliario FACTIS como "
        "herramienta complementaria al criterio profesional. Los valores indicados son preliminares y están "
        "sujetos a verificación técnica y legal. La propuesta definitiva deberá ser suscrita por las partes.</i>",
        _s("disc_prop", fontSize=6.5, leading=10, textColor=colors.HexColor("#AAAAAA"),
           fontName="Helvetica-Oblique")))
    story.append(Spacer(1, 5*mm))
    sign_tbl = Table(
        [[Paragraph(representante, _s("sig", fontSize=10, leading=13, textColor=NAV,
                                       fontName="Helvetica-Bold")),
          Paragraph("Generado con FACTIS · Osterling Advisory<br/>eosterling@grupoosterling.com · Lima, Perú",
                    _s("ft", fontSize=7, leading=10, textColor=colors.HexColor("#AAAAAA"),
                       alignment=TA_RIGHT))],
         [Paragraph(cargo, _s("cgo", fontSize=8, leading=10, textColor=GRAY)), Paragraph("", S_small)],
         [Paragraph("Osterling Advisory", _s("oa", fontSize=8, leading=10, textColor=GOLD)), Paragraph("", S_small)]],
        colWidths=[W*0.55, W*0.45],
        style=TableStyle([("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
                           ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
                           ("VALIGN",(1,0),(1,0),"MIDDLE")])
    )
    story.append(sign_tbl)

    doc.build(story)
    return buf.getvalue()


def generar_informe_residencial_html(r: dict, legal: dict | None, fecha: str,
                                      distrito: str = "", m2: int = 0, antiguedad: int = 0, fotos: list = None) -> str:
    NAV = "#1E2D3D"; GLD = "#B8904A"; BRD = "#D8D4CC"
    sem_col = {"verde": "#1A4731", "amarillo": "#7A4F1A", "rojo": "#7A1A1A"}
    sem_bg  = {"verde": "#E8F5EE", "amarillo": "#FFF8EE", "rojo": "#FFF0F0"}

    def kpi(label, value, sub=""):
        sub_html = f'<div style="font-size:10px;color:#7A7268;margin-top:3px;">{sub}</div>' if sub else ""
        return (f'<div style="background:#FFFFFF;border:1px solid {BRD};border-top:3px solid {GLD};'
                f'border-radius:5px;padding:14px 16px;min-width:130px;flex:1;">'
                f'<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;">{label}</div>'
                f'<div style="font-size:20px;font-weight:700;color:{NAV};margin-top:6px;">{value}</div>'
                f'{sub_html}</div>')

    kpis = (kpi("Precio", f"${r.get('precio', 0) or 0:,.0f}") +
            kpi("Pago inicial", f"${r.get('pie', 0) or 0:,.0f}", f"{r.get('pct_pie', 0) or 0:.0f}%") +
            kpi("Cuota Mensual", f"${r.get('cuota_mensual', 0) or 0:,.0f}" if (r.get('cuota_mensual') or 0) > 0 else "Al contado") +
            kpi("Ingreso Mínimo", f"${r.get('ingreso_minimo', 0) or 0:,.0f}" if (r.get('ingreso_minimo') or 0) > 0 else "—", "recomendado") +
            (kpi("Yield Neto", f"{r.get('yield_neto', 0) or 0:.1f}%") if r.get('uso') == "Inversión" else "") +
            (kpi("Payback", f"{r.get('payback_anos', 0) or 0:.1f} años") if r.get('payback_anos') else ""))

    amort_rows = "".join(
        f'<tr><td style="padding:7px 12px;font-size:12px;text-align:center;">{row["año"]}</td>'
        f'<td style="padding:7px 12px;font-size:12px;text-align:right;color:#1A4731;font-weight:600;">${row["capital"]:,.0f}</td>'
        f'<td style="padding:7px 12px;font-size:12px;text-align:right;color:#7A4F1A;">${row["interes"]:,.0f}</td>'
        f'<td style="padding:7px 12px;font-size:12px;text-align:right;">${row["saldo"]:,.0f}</td></tr>'
        for row in (r.get('amort_tabla') or [])
    )

    legal_html = ""
    if legal:
        sg = legal.get("semaforo", "amarillo").lower()
        legal_html = (
            f'<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;'
            f'margin:28px 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Análisis Legal</h3>'
            f'<div style="background:{sem_bg.get(sg,"#F5F2ED")};border-left:4px solid {sem_col.get(sg,NAV)};'
            f'border-radius:5px;padding:14px 18px;margin-bottom:12px;">'
            f'<div style="font-size:12px;color:{sem_col.get(sg,NAV)};font-weight:700;">'
            f'{"SIN ALERTAS CRÍTICAS" if sg=="verde" else ("OBSERVACIONES MENORES" if sg=="amarillo" else "ALERTAS CRÍTICAS")}</div>'
            f'<div style="font-size:12px;color:{sem_col.get(sg,NAV)};margin-top:6px;">{legal.get("resumen_legal","")}</div>'
            f'</div>'
        )
        for al in (legal.get("alertas") or []):
            legal_html += f'<div style="background:#FFF8EE;border-left:3px solid {GLD};border-radius:4px;padding:10px 14px;margin-bottom:6px;font-size:12px;color:{NAV};">⚠ {al}</div>'

    prop_section = ""
    if r.get('uso') in ["Inversión para alquilar", "Evaluación para venta"]:
        prop_section = f"""
<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:24px 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Análisis de Inversión</h3>
<table style="margin-bottom:20px;"><thead><tr><th>Indicador</th><th style="text-align:right;">Valor</th></tr></thead><tbody>
<tr><td style="padding:8px 12px;font-size:12px;">Alquiler mensual estimado</td><td style="padding:8px 12px;font-size:12px;text-align:right;">${r.get('alquiler_mes', 0) or 0:,.0f}</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Renta neta mensual</td><td style="padding:8px 12px;font-size:12px;text-align:right;font-weight:600;">${r.get('renta_neta_mes', 0) or 0:,.0f}</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Yield bruto anual</td><td style="padding:8px 12px;font-size:12px;text-align:right;">{r.get('yield_bruto', 0) or 0:.1f}%</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Yield neto anual</td><td style="padding:8px 12px;font-size:12px;text-align:right;font-weight:600;">{r.get('yield_neto', 0) or 0:.1f}%</td></tr>
{'<tr><td style="padding:8px 12px;font-size:12px;">Payback</td><td style="padding:8px 12px;font-size:12px;text-align:right;">'+f"{r.get('payback_anos', 0) or 0:.1f} años"+'</td></tr>' if r.get('payback_anos') else ''}
{'<tr><td style="padding:8px 12px;font-size:12px;">Flujo mensual neto</td><td style="padding:8px 12px;font-size:12px;text-align:right;font-weight:600;">$'+f"{r.get('flujo_mensual', 0) or 0:,.0f}"+'</td></tr>' if r.get('flujo_mensual') is not None else ''}
</tbody></table>
<table style="margin-bottom:20px;"><thead><tr><th>Apreciación estimada (4%/año)</th><th style="text-align:right;">Valor</th><th style="text-align:right;">Ganancia</th></tr></thead><tbody>
<tr><td style="padding:8px 12px;font-size:12px;">Valor a 5 años</td><td style="padding:8px 12px;font-size:12px;text-align:right;">${r.get('valor_5', 0) or 0:,.0f}</td><td style="padding:8px 12px;font-size:12px;text-align:right;color:#1A4731;">+${r.get('ganancia_capital_5', 0) or 0:,.0f}</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Valor a 10 años</td><td style="padding:8px 12px;font-size:12px;text-align:right;">${r.get('valor_10', 0) or 0:,.0f}</td><td style="padding:8px 12px;font-size:12px;text-align:right;color:#1A4731;font-weight:600;">+${r.get('ganancia_capital_10', 0) or 0:,.0f}</td></tr>
</tbody></table>"""

    _fotos_html = ""
    if fotos:
        import base64
        _fotos_html = '<div style="page-break-before:always;margin-top:40px;"><div class="section-title">Fotografías del Inmueble</div><div style="display:flex;flex-wrap:wrap;gap:12px;margin-top:16px;">'
        for _fb in fotos[:6]:
            try:
                _b64 = base64.b64encode(_fb).decode()
                _fotos_html += f'<img src="data:image/jpeg;base64,{_b64}" style="max-width:48%;max-height:300px;object-fit:cover;border-radius:6px;border:1px solid #E4E0D8;">'
            except Exception:
                pass
        _fotos_html += '</div></div>'

    _rpt_precio    = r.get("precio", 0) or 0
    _rpt_ppm2      = r.get("precio_m2", 0) or 0
    _rpt_yield     = r.get("yield_bruto", 0) or 0
    _rpt_dorm      = r.get("dormitorios") or "—"
    _rpt_precio_s  = f"{int(_rpt_precio):,}"
    _rpt_ppm2_s    = f"{int(_rpt_ppm2):,}"
    _rpt_yield_s   = f"{_rpt_yield:.1f}%" if _rpt_yield > 0 else "—"

    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Informe Residencial — Osterling Advisory</title>
<style>body{{font-family:'Segoe UI',Arial,sans-serif;background:#EDEAE4;margin:0;padding:32px;color:{NAV};}}
.page{{background:#FFFFFF;max-width:860px;margin:0 auto;padding:48px 56px;border-radius:6px;}}
table{{width:100%;border-collapse:collapse;}}thead th{{background:{NAV};color:#FFFFFF;padding:9px 12px;font-size:10px;letter-spacing:1px;text-transform:uppercase;}}
tbody tr:nth-child(even) td{{background:#F9F7F4;}}
</style></head><body><div class="page">
<div style="background:linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%);
            border-radius:12px;padding:36px 40px;margin-bottom:32px;color:#FFFFFF;">
    <div style="font-size:8px;letter-spacing:4px;text-transform:uppercase;
                color:rgba(255,255,255,0.5);margin-bottom:12px;">
        FACTIS · Informe de Análisis Residencial · {fecha}
    </div>
    <div style="font-size:32px;font-weight:800;line-height:1.2;margin-bottom:8px;">
        {distrito if distrito else r.get("zona", "—")} &nbsp;·&nbsp; {_rpt_dorm}
    </div>
    <div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:24px;">
        {str(m2) + " m²" if m2 else ""}{(" &nbsp;·&nbsp; " + str(antiguedad) + " años de antigüedad") if antiguedad else ""}
    </div>
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:16px;">
        <div style="background:rgba(255,255,255,0.10);border-radius:8px;padding:16px;">
            <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                        letter-spacing:1px;margin-bottom:4px;">Precio de compra</div>
            <div style="font-size:20px;font-weight:700;">${_rpt_precio_s}</div>
        </div>
        <div style="background:rgba(255,255,255,0.10);border-radius:8px;padding:16px;">
            <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                        letter-spacing:1px;margin-bottom:4px;">Precio / m²</div>
            <div style="font-size:20px;font-weight:700;">${_rpt_ppm2_s}/m²</div>
        </div>
        <div style="background:rgba(255,255,255,0.10);border-radius:8px;padding:16px;">
            <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                        letter-spacing:1px;margin-bottom:4px;">Yield bruto</div>
            <div style="font-size:20px;font-weight:700;">{_rpt_yield_s}</div>
        </div>
    </div>
</div>
<div style="border-bottom:2px solid {GLD};padding-bottom:20px;margin-bottom:28px;display:flex;justify-content:space-between;align-items:flex-end;">
  <div>
    <div style="font-size:9px;color:{GLD};letter-spacing:4px;text-transform:uppercase;font-weight:600;">Osterling Advisory</div>
    <div style="font-size:22px;font-weight:700;color:{NAV};margin-top:6px;">Evaluación Inmueble Residencial</div>
    <div style="font-size:12px;color:#7A7268;margin-top:4px;">{r.get('uso','—')}{(" · "+distrito) if distrito else ""}{(" · "+str(m2)+" m²") if m2 else ""}{(" · "+str(antiguedad)+" años") if antiguedad else ""}</div>
  </div>
  <div style="text-align:right;font-size:11px;color:#9A9080;">{fecha}<br>Lima, Perú</div>
</div>

<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Indicadores Clave</h3>
<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:28px;">{kpis}</div>

<h3 style="color:{NAV};font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid {BRD};padding-bottom:6px;">Estructura de Crédito Hipotecario</h3>
<table style="margin-bottom:24px;"><thead><tr><th>Concepto</th><th style="text-align:right;">Valor</th></tr></thead><tbody>
<tr><td style="padding:8px 12px;font-size:12px;">Precio de compra</td><td style="padding:8px 12px;font-size:12px;text-align:right;font-weight:600;">${r.get('precio', 0) or 0:,.0f}</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Pago inicial ({r.get('pct_pie', 0) or 0:.0f}%)</td><td style="padding:8px 12px;font-size:12px;text-align:right;">${r.get('pie', 0) or 0:,.0f}</td></tr>
<tr><td style="padding:8px 12px;font-size:12px;">Monto del crédito</td><td style="padding:8px 12px;font-size:12px;text-align:right;">${r.get('monto_credito', 0) or 0:,.0f}</td></tr>
{"<tr><td style='padding:8px 12px;font-size:12px;'>Tasa de interés</td><td style='padding:8px 12px;font-size:12px;text-align:right;'>"+f"{r.get('tasa_anual', 0) or 0:.2f}% TEA"+"</td></tr>" if (r.get('cuota_mensual') or 0)>0 else ""}
{"<tr><td style='padding:8px 12px;font-size:12px;'>Plazo</td><td style='padding:8px 12px;font-size:12px;text-align:right;'>"+str(r.get('plazo_anos', '—'))+" años ("+str(r.get('n_meses', '—'))+" cuotas)"+"</td></tr>" if (r.get('cuota_mensual') or 0)>0 else ""}
{"<tr><td style='padding:8px 12px;font-size:12px;'>Cuota mensual</td><td style='padding:8px 12px;font-size:12px;text-align:right;font-weight:600;'>$"+f"{r.get('cuota_mensual', 0) or 0:,.0f}"+"</td></tr>" if (r.get('cuota_mensual') or 0)>0 else ""}
{"<tr><td style='padding:8px 12px;font-size:12px;'>Total pagado al banco</td><td style='padding:8px 12px;font-size:12px;text-align:right;'>$"+f"{r.get('total_pagado', 0) or 0:,.0f}"+" (intereses: $"+f"{r.get('total_intereses', 0) or 0:,.0f}"+")"+"</td></tr>" if (r.get('cuota_mensual') or 0)>0 else ""}
{"<tr><td style='padding:8px 12px;font-size:12px;'>Ingreso mínimo recomendado</td><td style='padding:8px 12px;font-size:12px;text-align:right;font-weight:600;'>$"+f"{r.get('ingreso_minimo', 0) or 0:,.0f}"+"/mes</td></tr>" if (r.get('ingreso_minimo') or 0)>0 else ""}
</tbody></table>

{prop_section}

{"<h3 style='color:"+NAV+";font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;border-bottom:1px solid "+BRD+";padding-bottom:6px;'>Tabla de Amortización (primeros 10 años)</h3><table><thead><tr><th style='text-align:center;'>Año</th><th style='text-align:right;'>Capital Pagado</th><th style='text-align:right;'>Intereses</th><th style='text-align:right;'>Saldo</th></tr></thead><tbody>"+amort_rows+"</tbody></table>" if amort_rows else ""}

{legal_html}

<div style="margin-top:48px;border-top:1px solid {BRD};padding-top:20px;">
<p style="font-size:11px;font-weight:700;color:{NAV};margin:0;">Enrique Osterling</p>
<p style="font-size:10px;color:#555;margin:3px 0;">Gerente General — Osterling Advisory · Inmobiliaria Corporativa</p>
<p style="font-size:10px;color:#555;margin:3px 0;">+51 950 891 995 · eosterling@grupoosterling.com · Lima, Perú</p>
<p style="font-size:9px;color:#AAA;margin-top:12px;">Análisis referencial. No constituye asesoría legal ni financiera formal.</p>
</div>
{_fotos_html}
</div></body></html>"""


def generar_informe_html(params, cabida, financ, legal, zona, financ_inputs, fecha):
    """Genera informe HTML descargable con encabezado fijo en impresión y control de saltos."""
    r   = (financ or {}).get("resumen", {}) if financ else {}
    fi  = financ_inputs or {}
    p   = params or {}
    c   = cabida or {}
    leg = legal  or {}

    # ── Flujo de caja ────────────────────────────────
    tir_real = mes_be = max_exp = None
    if cabida and financ:
        m_data = MERCADO.get(zona, {})
        fin_fl = {
            "costo_terreno":      fi.get("costo_terreno", 0),
            "costo_construccion": fi.get("costo_construccion", m_data.get("costo_construccion", 0)),
            "costo_sotano_m2":    fi.get("costo_sotano_m2", 450),
            "fee_constructora":   fi.get("fee_constructora", 10.0),
            "tasa_ir":            fi.get("tasa_ir", 29.5),
            "include_alcabala":   fi.get("include_alcabala", True),
            "include_dd":         fi.get("include_dd", True),
            "precio_venta_m2":    fi.get("precio_venta_m2", m_data.get("precio_2br", 0)),
            "precio_estac":       m_data.get("precio_estac", 0),
            "precio_deposito":    m_data.get("precio_deposito", 0),
            "tasa_financ":        fi.get("tasa_financ", 9.0),
        }
        result_fl = calcular_financiero(cabida, fin_fl, zona)
        try:
            _, _, tir_real, mes_be, max_exp, _ = generar_flujo(cabida, result_fl, fin_fl, zona)
        except Exception:
            pass

    _score_r = score_viabilidad(r) if r else (0, 0, "—", "", "", "—", [])
    pts, score_10, etiqueta, _, _, recomendacion, score_items = _score_r

    NAV  = "#1E2D3D"
    GOLD = "#B8904A"
    ALT  = "#F4F1EC"
    BORD = "#D8D4CC"
    GREY = "#6B7280"

    ubicacion    = p.get("ubicacion", "—")
    zonificacion = p.get("zonificacion", "—")
    area_t       = f"{p.get('area_terreno_m2', '—')} m²"
    pisos_max    = f"{p.get('pisos_max', '—')} pisos"
    area_libre   = f"{p.get('area_libre_min_pct', '—')}%"
    retiro       = f"{p.get('retiro_frontal_ml', p.get('retiro_frontal', '—'))} ml"
    coef         = p.get("coeficiente_edificacion", "—")

    at_total = c.get("area_techada_total_m2", 0)
    av_total = c.get("area_vendible_m2", 0)
    efic     = f"{round(av_total/at_total*100,1)}%" if at_total else "—"

    semaforo_color = {"verde": "#1A7A4A", "amarillo": "#B8862E", "rojo": "#8B1A1A"}.get(
        leg.get("semaforo", ""), NAV)
    semaforo_label = {"verde": "SIN ALERTAS CRÍTICAS", "amarillo": "ALERTAS MENORES",
                      "rojo": "ALERTAS CRÍTICAS"}.get(leg.get("semaforo", ""), "No ejecutado")

    # viabilidad
    _mg  = r.get("margen_pct", 0)
    _tir = r.get("tir_anual_pct", 0)
    if _mg >= 20 and _tir >= 15:
        _viab_col, _viab_bg, _viab_txt = "#1A7A4A", "#E8F5EE", "RETORNOS SÓLIDOS"
    elif _mg >= 12 and _tir >= 10:
        _viab_col, _viab_bg, _viab_txt = "#B8862E", "#FFF8EE", "RETORNOS MODERADOS"
    else:
        _viab_col, _viab_bg, _viab_txt = "#4A4A5A", "#F4F4F6", "RETORNOS AJUSTADOS"

    # ── Helpers ──────────────────────────────────────
    def kv_row(i, k, v):
        bg = ALT if i % 2 == 0 else "#fff"
        return (f'<tr>'
                f'<td style="background:{bg};padding:7px 12px;font-size:11px;font-weight:600;'
                f'color:{NAV};width:42%;border-bottom:1px solid {BORD};">{k}</td>'
                f'<td style="background:{bg};padding:7px 12px;font-size:11px;color:#333;'
                f'border-bottom:1px solid {BORD};">{v}</td>'
                f'</tr>')

    def kv_table(rows):
        body = "".join(kv_row(i, k, v) for i, (k, v) in enumerate(rows))
        return (f'<table style="width:100%;border-collapse:collapse;margin-bottom:0;">'
                f'{body}</table>')

    def section(title, force_break=False):
        pb = 'page-break-before:always;' if force_break else ''
        return (f'<div class="section-block" style="{pb}padding-top:4px;">'
                f'<p class="sec-title" style="font-size:12px;font-weight:700;color:{NAV};'
                f'margin:0 0 2px 0;letter-spacing:0.3px;">{title}</p>'
                f'<div style="border-top:1.5px solid {GOLD};margin-bottom:10px;"></div>')

    def end_section():
        return '</div>'

    def _leg_list(val, empty="Ninguna registrada"):
        if not val or val in ([], "[]"): return empty
        if isinstance(val, list):
            items = [str(v.get("descripcion", v) if isinstance(v, dict) else v) for v in val if v]
            return "; ".join(items) if items else empty
        return str(val) if val and val != "—" else empty

    tir_str = f"{tir_real}%" if tir_real is not None else "—"
    exp_str = fmt_usd(abs(max_exp)) if max_exp is not None else "—"
    be_str  = f"Mes {mes_be}" if mes_be else "—"

    nombre_proyecto = fi.get("nombre_proyecto", "") or zona

    # ── Unidades mix ─────────────────────────────────
    unidades_rows = []
    for u in (c.get("unidades") or []):
        unidades_rows.append((f"Dpto {u.get('tipo','—')}",
                              f"{u.get('cantidad',0)} und · {u.get('area_m2',0):.0f} m²/und"))

    # ── Beneficios normativos ─────────────────────────
    benef_items = ""
    for b in p.get("beneficios_normativos", []):
        benef_items += (f'<li style="font-size:10px;color:{GOLD};margin:3px 0;">'
                        f'<strong>{b.get("descripcion","")}</strong> — {b.get("impacto_estimado","")}</li>')
    benef_block = (f'<ul style="margin:6px 0 0 16px;padding:0;">{benef_items}</ul>'
                   if benef_items else "")

    # ── Alertas legales ───────────────────────────────
    alertas_html = ""
    for a in leg.get("alertas", []):
        alertas_html += (f'<p style="font-size:10px;color:#8B1A1A;margin:4px 0;'
                         f'padding:4px 8px;background:#FFF0F0;border-left:3px solid #8B1A1A;">'
                         f'&#9888; {a}</p>')

    # ── HTML ─────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>FACTIS — Informe — {ubicacion}</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{
    font-family: 'Helvetica Neue', Arial, sans-serif;
    color: #2C2C2C;
    font-size: 11px;
    line-height: 1.65;
    background: #fff;
    max-width: 760px;
    margin: 0 auto;
    padding: 32px 40px 40px;
  }}
  table {{ border-collapse:collapse; width:100%; }}
  .section-block {{ margin-bottom:20px; }}
  .sec-title {{ page-break-after: avoid; break-after: avoid; }}
  .section-block table {{ page-break-inside: auto; }}
  .section-block tr {{ page-break-inside: avoid; break-inside: avoid; }}

  /* ── Screen header ── */
  .screen-header {{ display:block; }}
  .print-header  {{ display:none; }}

  @media print {{
    @page {{
      size: A4;
      margin: 22mm 18mm 16mm;
    }}
    body {{
      padding: 0;
      padding-top: 16mm;  /* reserve space for fixed header */
      max-width: 100%;
    }}
    .screen-header {{ display:none !important; }}
    .print-header {{
      display: block !important;
      position: fixed;
      top: -22mm;        /* reach into the @page top margin */
      left: -18mm;
      right: -18mm;
      background: white;
      padding: 6mm 18mm 4mm;
      z-index: 9999;
    }}
    .print-header-line {{
      border-bottom: 1.5px solid {GOLD};
      margin-top: 3mm;
    }}
    .section-block {{ page-break-inside: avoid; break-inside: avoid; }}
    .sec-title {{ page-break-after: avoid; break-after: avoid; }}
    .no-break {{ page-break-inside: avoid; break-inside: avoid; }}
    a {{ text-decoration: none; color: inherit; }}
  }}
</style>
</head>
<body>

<!-- ── Print header (fixed, repeats on every page) ── -->
<div class="print-header">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <div>
      <span style="font-size:16px;font-weight:800;color:{NAV};letter-spacing:-0.5px;">FACTIS</span>
      <span style="font-size:8px;color:{GOLD};letter-spacing:3px;text-transform:uppercase;
                   margin-left:10px;font-weight:600;">IA de Análisis Inmobiliario</span>
    </div>
    <span style="font-size:8px;color:{GREY};">Osterling Advisory &nbsp;·&nbsp; {fecha}</span>
  </div>
  <div class="print-header-line"></div>
</div>

<!-- ── Screen header (only on screen) ── -->
<div class="screen-header" style="margin-bottom:6px;">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <div>
      <span style="font-size:22px;font-weight:800;color:{NAV};letter-spacing:-1px;">FACTIS</span>
      <span style="font-size:9px;color:{GOLD};letter-spacing:3px;text-transform:uppercase;
                   margin-left:12px;font-weight:600;">IA de Análisis Inmobiliario</span>
    </div>
    <div style="text-align:right;">
      <div style="font-size:9px;color:#555;font-weight:600;">Osterling Advisory</div>
      <div style="font-size:9px;color:#777;">{fecha}</div>
    </div>
  </div>
  <div style="border-top:1.5px solid {GOLD};margin-top:6px;margin-bottom:16px;"></div>
</div>

<!-- ── Título ── -->
<div style="text-align:center;margin-bottom:20px;">
  <p style="font-size:15px;font-weight:700;color:{NAV};letter-spacing:0.5px;">
    INFORME DE ANÁLISIS DE CABIDA E INVERSIÓN
  </p>
  <p style="font-size:11px;color:{GREY};margin-top:4px;">
    {ubicacion} &nbsp;·&nbsp; {zona} &nbsp;·&nbsp; {fecha}
  </p>
</div>

<!-- ── Viabilidad banner ── -->
<div class="no-break" style="background:{_viab_bg};border:1px solid {_viab_col};border-left:4px solid {_viab_col};
     border-radius:4px;padding:10px 16px;margin-bottom:20px;display:flex;
     justify-content:space-between;align-items:center;">
  <div>
    <div style="font-size:8px;color:{_viab_col};letter-spacing:2px;font-weight:700;text-transform:uppercase;">
      Perfil de Inversión
    </div>
    <div style="font-size:16px;font-weight:800;color:{_viab_col};margin-top:2px;">{_viab_txt}</div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:9px;color:{_viab_col};opacity:0.8;">Margen neto &nbsp;·&nbsp; TIR anual</div>
    <div style="font-size:15px;font-weight:700;color:{_viab_col};">{_mg:.1f}% &nbsp;·&nbsp; {_tir:.1f}%</div>
  </div>
</div>

<!-- ── I. Inmueble ── -->
{section("I. Datos del Inmueble")}
{kv_table([
    ("Ubicación",           ubicacion),
    ("Zonificación",        zonificacion),
    ("Área del terreno",    area_t),
    ("Altura máxima",       pisos_max),
    ("Área libre mínima",   area_libre),
    ("Retiro frontal",      retiro),
    ("Coeficiente edific.", str(coef)),
])}
{benef_block}
{end_section()}

<!-- ── II. Cabida ── -->
{section("II. Programa Arquitectónico")}
{kv_table([
    ("Pisos / sótanos",    f"{c.get('num_pisos','—')} pisos · {c.get('num_sotanos',0)} sótanos"),
    ("Área techada total", f"{at_total:,.0f} m²"),
    ("Área vendible",      f"{av_total:,.0f} m² (eficiencia {efic})"),
    ("Departamentos",      str(c.get("total_unidades","—"))),
    *unidades_rows,
    ("Estacionamientos",   str(c.get("estac_total","—"))),
    ("Depósitos",          str(c.get("depositos_total", 0))),
])}
{end_section()}

<!-- ── III. Financiero ── -->
{section("III. Análisis Financiero")}
{kv_table([
    ("Ingresos brutos",           fmt_usd(r.get("ingresos_brutos", 0))),
    ("Costo total (s/financ.)",   fmt_usd(r.get("costo_total_sin_financ", 0))),
    ("Gasto financiero banco",    fmt_usd(r.get("costo_financiero", 0))),
    ("Utilidad neta",             fmt_usd(r.get("utilidad_neta", 0))),
    ("Margen neto post-IR",       f"{r.get('margen_pct', 0):.1f}%"),
    ("ROI",                       f"{r.get('roi_pct', 0):.1f}%"),
    ("TIR anual estimada",        f"{r.get('tir_anual_pct', 0):.1f}%"),
    ("Duración del proyecto",     f"{r.get('meses_proyecto', 0)} meses"),
    ("Precio máx. terreno (20%)", fmt_usd(r.get("max_terreno_20pct", 0))),
    ("Break-even precio/m²",      f"${r.get('be_precio_m2', 0):,}/m²"),
])}
{end_section()}

<!-- ── IV. Flujo de Caja ── -->
{section("IV. Flujo de Caja", force_break=True)}
{kv_table([
    ("TIR anual (bisección numérica)", tir_str),
    ("Exposición máxima de capital",   exp_str),
    ("Mes de break-even",              be_str),
])}
{end_section()}

<!-- ── V. Due Diligence Legal ── -->
{section("V. Due Diligence Legal", force_break=(bool(leg)))}
<div class="no-break" style="background:{semaforo_color}15;border-left:4px solid {semaforo_color};
     padding:8px 14px;margin-bottom:10px;border-radius:0 4px 4px 0;">
  <span style="font-size:12px;font-weight:700;color:{semaforo_color};">{semaforo_label}</span>
</div>
{kv_table([
    ("Propietario(s) Partida",  ", ".join(
        (x.get("nombre", str(x)) if isinstance(x, dict) else str(x))
        for x in (leg.get("propietarios_partida") or []) if x
    ) or "—"),
    ("Área registral",          f"{leg.get('area_registral_m2','—')} m²" if leg.get('area_registral_m2') else "—"),
    ("Cargas vigentes",         _leg_list(leg.get("cargas_vigentes"))),
    ("Hipotecas vigentes",      _leg_list(leg.get("hipotecas_vigentes"))),
    ("Medidas cautelares",      _leg_list(leg.get("medidas_cautelares"))),
]) if leg else kv_table([("Estado", "Análisis legal no ejecutado")])}
{alertas_html}
{"<p style='font-size:10px;color:#555;margin-top:8px;font-style:italic;'>" + (leg.get("resumen_legal") or "") + "</p>" if leg.get("resumen_legal") else ""}
{end_section()}

<!-- ── VI. Recomendación ── -->
{section("VI. Recomendación Estratégica")}
<div class="no-break" style="background:{_viab_bg};border:1px solid {_viab_col};border-radius:4px;
     padding:12px 16px;line-height:1.7;">
  <p style="font-size:11px;color:{NAV};">{recomendacion}</p>
</div>
{end_section()}

<!-- ── Firma ── -->
<div style="margin-top:32px;padding-top:16px;border-top:1px solid {BORD};
     display:flex;justify-content:space-between;align-items:flex-end;">
  <div>
    <p style="font-size:11px;font-weight:700;color:{NAV};">Enrique Osterling</p>
    <p style="font-size:10px;color:{GREY};">Director General · Osterling Advisory</p>
    <p style="font-size:10px;color:{GREY};">eosterling@grupoosterling.com</p>
  </div>
  <p style="font-size:8px;color:#AAA;text-align:right;">
    Resultados referenciales — validar con asesores especializados.<br>
    Generado con FACTIS &copy; 2026 · Osterling Advisory
  </p>
</div>

</body></html>"""
    return html


# ═══════════════════════════════════════════════════════
# APP PRINCIPAL
# ═══════════════════════════════════════════════════════

_module_subtitles = {
    "Proyecto Inmobiliario":          "IA de Análisis Inmobiliario",
    "Proyecto Logístico / Industrial": "Análisis de Activos Logísticos e Industriales",
    "Inmueble Residencial":            "Evaluación de Inmuebles Residenciales",
}
_active_module = st.session_state.get("tipo_operacion", "Proyecto Inmobiliario")
_subtitle = _module_subtitles.get(_active_module, "IA de Análisis Inmobiliario")
_module_tags = {
    "Proyecto Inmobiliario":          "Cabida · Financiero · Legal · Asistente",
    "Proyecto Logístico / Industrial": "Costo · Yield · DSCR · Comparativa",
    "Inmueble Residencial":            "Crédito · Inversión · Amortización",
}
_tag = _module_tags.get(_active_module, "")

st.markdown(
    '<div class="main-header">'
    '<div style="display:flex;align-items:center;justify-content:space-between;">'
    '<div>'
    '<div style="font-size:9px;color:#B8904A;letter-spacing:4px;text-transform:uppercase;font-weight:600;margin-bottom:10px;">Osterling Advisory</div>'
    '<div style="display:flex;align-items:center;gap:18px;">'
    '<span style="font-size:22px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;">FACTIS</span>'
    '<span style="width:1px;height:18px;background:#B8904A;opacity:0.4;display:inline-block;flex-shrink:0;"></span>'
    f'<span style="font-size:10px;color:#8AA8C0;letter-spacing:2.5px;text-transform:uppercase;font-weight:500;">{_subtitle}</span>'
    '</div>'
    f'<div style="margin-top:10px;font-size:9px;color:#B8904A;letter-spacing:1.5px;opacity:0.8;">{_tag}</div>'
    '</div>'
    '<div style="text-align:right;">'
    '<div style="font-size:8px;color:#B8904A;letter-spacing:2px;text-transform:uppercase;font-weight:600;opacity:0.7;">Lima, Perú</div>'
    f'<div style="font-size:9px;color:rgba(184,144,74,0.5);margin-top:6px;letter-spacing:1px;">{_active_module}</div>'
    '</div>'
    '</div></div>',
    unsafe_allow_html=True
)

# ── SIDEBAR ──────────────────────────────────────────

run = False
run_industrial = False
run_residencial = False
run_ind_docs = False
run_res_docs = False

with st.sidebar:
    # ── Sidebar brand header ──────────────────────────
    _user_display = st.session_state.get("_user_name", "")
    _user_role    = st.session_state.get("_user_role", "advisor")
    if _LOGO_B64:
        st.markdown(f"""
        <div style="padding:16px 16px 10px;text-align:center;
                    border-bottom:1px solid rgba(255,255,255,0.09);margin-bottom:6px;">
            <div style="background:#FFFFFF;border-radius:8px;padding:10px 16px;
                        display:inline-block;max-width:90%;">
                <img src="data:image/png;base64,{_LOGO_B64}"
                     style="height:36px;max-width:100%;object-fit:contain;display:block;">
            </div>
            <div style="font-size:7px;color:rgba(184,144,74,0.75);letter-spacing:3px;
                        text-transform:uppercase;font-weight:600;margin-top:9px;">
                Plataforma Analítica Inmobiliaria
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="padding:24px 20px 14px;border-bottom:1px solid rgba(255,255,255,0.09);
                    margin-bottom:6px;text-align:center;">
            <div style="font-size:22px;font-weight:800;color:#FFFFFF;letter-spacing:-0.5px;">
                FACTIS
            </div>
            <div style="font-size:7px;color:rgba(184,144,74,0.80);letter-spacing:4px;
                        text-transform:uppercase;font-weight:600;margin-top:6px;">
                Osterling Advisory
            </div>
        </div>""", unsafe_allow_html=True)

    # ── Usuario activo + logout ───────────────────────
    _role_label = "Admin" if _user_role == "admin" else "Asesor"
    st.markdown(
        f'<div style="display:flex;align-items:center;justify-content:space-between;'
        f'padding:8px 2px 4px;">'
        f'<div>'
        f'<div style="font-size:11px;font-weight:600;color:#C8D8E8;">{_user_display}</div>'
        f'<div style="font-size:9px;color:rgba(184,144,74,0.75);letter-spacing:1px;'
        f'text-transform:uppercase;font-weight:600;">{_role_label}</div>'
        f'</div></div>',
        unsafe_allow_html=True)
    if st.button("Cerrar sesión", key="_logout_btn", use_container_width=True):
        for k in ["_authenticated","_user_name","_user_role","_username"]:
            st.session_state.pop(k, None)
        st.rerun()

    st.markdown("### MÓDULO DE ANÁLISIS")
    tipo_op = st.radio(
        "tipo_op_radio",
        ["Proyecto Inmobiliario", "Proyecto Logístico / Industrial", "Inmueble Residencial", "Portfolio"],
        key="tipo_operacion",
        label_visibility="collapsed",
    )

    # ── Descripción contextual del módulo seleccionado ───
    _mod_ctx = {
        "Proyecto Inmobiliario": (
            "🏗",
            "Desarrollo residencial multifamiliar",
            "Cabida normativa · Financiero · RIN · PDF",
            "Úsalo cuando tienes un terreno y quieres proyectar pisos, unidades, costos y rentabilidad de un edificio de departamentos.",
        ),
        "Proyecto Logístico / Industrial": (
            "🏭",
            "Nave logística · Almacén · Industria",
            "Costo · Yield · DSCR · Comparativa",
            "Úsalo para evaluar la construcción o adquisición de una nave industrial, almacén logístico o planta de manufactura.",
        ),
        "Inmueble Residencial": (
            "🏠",
            "Valuación de inmueble existente",
            "Precio · Renta · Yield · Comparables",
            "Úsalo cuando tienes un departamento, casa o unidad residencial ya construida y quieres analizar su valor o rentabilidad.",
        ),
        "Portfolio": (
            "📁",
            "Proyectos guardados",
            "KPIs · Historial · Comparativa",
            "Vista consolidada de todos los proyectos analizados y guardados. Compara KPIs entre proyectos y revisa el historial de versiones.",
        ),
    }
    _mico, _mtit, _mtag, _mdesc = _mod_ctx.get(tipo_op, _mod_ctx["Proyecto Inmobiliario"])
    st.markdown(
        f'<div style="background:rgba(255,255,255,0.04);border-radius:9px;padding:11px 13px;'
        f'border:1px solid rgba(255,255,255,0.08);margin:6px 0 4px;">'
        f'<div style="font-size:12px;font-weight:700;color:#C8D8E8;margin-bottom:2px;">'
        f'{_mico} {_mtit}</div>'
        f'<div style="font-size:9px;color:rgba(184,144,74,0.80);letter-spacing:1px;'
        f'text-transform:uppercase;font-weight:600;margin-bottom:6px;">{_mtag}</div>'
        f'<div style="font-size:11px;color:rgba(184,200,216,0.70);line-height:1.5;">{_mdesc}</div>'
        f'</div>',
        unsafe_allow_html=True)
    st.markdown("---")

    # ── Helper compartido: render de paso de progreso ────
    def _sp(idx, lbl, sub, done, is_cur, is_last):
        if done:
            cbg, cco, ctxt, lco, sco, lnc = "#6BCEA0","#0A1628","✓","#6BCEA0","rgba(107,206,160,0.6)","rgba(107,206,160,0.35)"
        elif is_cur:
            cbg, cco, ctxt, lco, sco, lnc = "#B8904A","#FFF",str(idx+1),"#D4A853","rgba(184,200,216,0.75)","rgba(255,255,255,0.10)"
        else:
            cbg, cco, ctxt, lco, sco, lnc = "rgba(255,255,255,0.08)","rgba(184,200,216,0.35)",str(idx+1),"rgba(184,200,216,0.40)","rgba(184,200,216,0.28)","rgba(255,255,255,0.06)"
        ln = "" if is_last else f'<div style="width:1px;height:13px;background:{lnc};margin:2px auto;"></div>'
        return (f'<div style="display:flex;align-items:flex-start;gap:10px;margin-bottom:2px;">'
                f'<div style="display:flex;flex-direction:column;align-items:center;flex-shrink:0;">'
                f'<div style="width:22px;height:22px;border-radius:50%;background:{cbg};'
                f'display:flex;align-items:center;justify-content:center;'
                f'font-size:10px;font-weight:700;color:{cco};">{ctxt}</div>{ln}</div>'
                f'<div style="padding-top:3px;">'
                f'<div style="font-size:11px;font-weight:600;color:{lco};line-height:1.2;">{lbl}</div>'
                f'<div style="font-size:10px;color:{sco};line-height:1.3;">{sub}</div>'
                f'</div></div>')

    # ── MÓDULO 1: PROYECTO INMOBILIARIO ──────────────────
    if tipo_op == "Proyecto Inmobiliario":
        # ── Indicador de progreso guiado ─────────────────
        _step1_done = bool(st.session_state.get("nombre_proyecto", "").strip())
        _step2_done = int(st.session_state.get("cab_precio_compra", 0) or 0) > 0
        _step3_done = st.session_state.get("cabida") is not None
        _step4_done = _step3_done
        _steps_prog = [
            ("Terreno & Proyecto", "Nombre, ubicación, área", _step1_done),
            ("Datos Financieros",  "Precio compra y venta/m²", _step2_done),
            ("Análisis IA",        "Cabida, RIN, Factibilidad", _step3_done),
            ("Reporte PDF",        "Descarga informe ejecutivo", _step4_done),
        ]
        _current_step = next((i for i, (_, _, d) in enumerate(_steps_prog) if not d), 4)
        _prog_html = "".join([_sp(i,l,s,d,i==_current_step,i==3) for i,(l,s,d) in enumerate(_steps_prog)])
        st.markdown(
            f'<div style="background:rgba(255,255,255,0.04);border-radius:10px;padding:12px 14px 10px;'
            f'border:1px solid rgba(255,255,255,0.08);margin-bottom:10px;">'
            f'<div style="font-size:9px;font-weight:700;color:rgba(184,200,216,0.45);'
            f'letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">FLUJO DE TRABAJO</div>'
            f'{_prog_html}</div>',
            unsafe_allow_html=True)
        st.markdown("### DOCUMENTOS")
        pdf_cert    = st.file_uploader("Certificado de Parámetros ✱",       type=["pdf"], key="cert")
        pdf_plano   = st.file_uploader("Planos (perimetral / topográfico)",  type=["pdf"], key="plano",
                                       accept_multiple_files=True)
        pdf_partida = st.file_uploader("Partida Registral",                  type=["pdf"], key="partida")
        pdf_puhr    = st.file_uploader("PU / HR",                            type=["pdf"], key="puhr")
        pdf_norms   = st.file_uploader("Ordenanzas y normativa",             type=["pdf"], key="norms",
                                       accept_multiple_files=True)

        st.markdown("---")
        st.markdown("### PROYECTO")
        # ── Estado vacío guiado ────────────────────────────
        _no_nombre = not bool(st.session_state.get("nombre_proyecto", "").strip())
        _no_area   = float(st.session_state.get("cab_override_area", 0) or 0) == 0
        _no_precio = int(st.session_state.get("cab_precio_compra", 0) or 0) == 0
        if _no_nombre and _no_area and _no_precio:
            st.markdown(
                '<div style="background:rgba(184,144,74,0.08);border-radius:8px;padding:11px 13px;'
                'border:1px solid rgba(184,144,74,0.22);margin-bottom:10px;">'
                '<div style="font-size:10px;font-weight:700;color:#D4A853;letter-spacing:0.5px;margin-bottom:7px;">¿POR DÓNDE EMPEZAR?</div>'
                '<div style="font-size:11px;color:rgba(184,200,216,0.82);line-height:1.75;">'
                '① Ingresa el <b style="color:#C8D8E8;">nombre</b> del proyecto<br>'
                '② Selecciona el <b style="color:#C8D8E8;">distrito</b><br>'
                '③ Sube el <b style="color:#C8D8E8;">certificado de parámetros</b><br>'
                '④ Completa el <b style="color:#C8D8E8;">precio de compra</b><br>'
                '⑤ Presiona <b style="color:#D4A853;">Analizar →</b>'
                '</div></div>',
                unsafe_allow_html=True)
        nombre_proyecto = st.text_input(
            "Nombre del proyecto",
            placeholder="Ej: Torres Las Camelias — Miraflores",
            key="nombre_proyecto")
        _cab_zona_keys = list(MERCADO.keys())
        _cab_zona_saved = st.session_state.get("cab_zona_sel", "Miraflores")
        _cab_zona_default = _cab_zona_keys.index(_cab_zona_saved) if _cab_zona_saved in _cab_zona_keys else min(20, len(_cab_zona_keys) - 1)
        zona = st.selectbox("Ubicación", _cab_zona_keys, index=_cab_zona_default, key="cab_zona_sel_widget")

        st.markdown("---")
        st.markdown("### INFORMACIÓN COMPLEMENTARIA DEL INMUEBLE")
        with st.expander("📄 Completar desde documento"):
            st.caption("Sube el certificado de parámetros, plano o cualquier documento del predio. La IA extrae automáticamente el área, frente, fondo, distrito y precios.")
            _cab_doc_up = st.file_uploader(
                "PDF, PPTX o DOCX",
                type=["pdf", "pptx", "ppt", "docx"],
                key="cab_import_doc",
            )
            if _cab_doc_up and st.button("EXTRAER DATOS", key="btn_cab_extract", use_container_width=True):
                _cab_bytes = _cab_doc_up.read()
                _cab_ext = _run_with_retry(
                    lambda _b=_cab_bytes, _n=_cab_doc_up.name: extraer_datos_desde_doc(_b, _n, "cabida"),
                    "Analizando documento…"
                )
                if _cab_ext.get("_error"):
                    st.error(_cab_ext["_error"])
                else:
                    if _cab_ext.get("nombre_proyecto"):
                        st.session_state["nombre_proyecto"] = _cab_ext["nombre_proyecto"]
                    if _cab_ext.get("area_terreno_m2"):
                        st.session_state["cab_override_area"] = float(_cab_ext["area_terreno_m2"])
                    if _cab_ext.get("frente_ml"):
                        st.session_state["cab_override_frente"] = float(_cab_ext["frente_ml"])
                    if _cab_ext.get("fondo_ml"):
                        st.session_state["cab_override_fondo"] = float(_cab_ext["fondo_ml"])
                    if _cab_ext.get("costo_terreno_usd"):
                        st.session_state["cab_precio_compra_inp"] = int(_cab_ext["costo_terreno_usd"])
                    if _cab_ext.get("precio_venta_m2_usd"):
                        st.session_state["cab_pventa_inp"] = int(_cab_ext["precio_venta_m2_usd"])
                    if _cab_ext.get("costo_construccion_m2_usd"):
                        st.session_state["cab_cconst_inp"] = int(_cab_ext["costo_construccion_m2_usd"])
                    _ext_dist = _cab_ext.get("distrito") or ""
                    for _k in MERCADO.keys():
                        if _ext_dist.lower() in _k.lower() or _k.lower() in _ext_dist.lower():
                            st.session_state["cab_zona_sel"] = _k
                            break
                    st.success("Datos extraídos. Revisa y ajusta antes de ejecutar.")
                    st.rerun()
        with st.expander("Medidas del lote", expanded=False):
            st.caption("Completa o adiciona información del inmueble para enriquecer el análisis.")
            col_fr, col_fo = st.columns(2)
            override_frente = col_fr.number_input("Frente (ml)", min_value=0.0, max_value=500.0,
                                                  value=float(st.session_state.get("cab_override_frente", 0.0)), step=0.5,
                                                  key="cab_frente_inp")
            override_fondo  = col_fo.number_input("Fondo (ml)",  min_value=0.0, max_value=500.0,
                                                  value=float(st.session_state.get("cab_override_fondo", 0.0)), step=0.5,
                                                  key="cab_fondo_inp")

            _area_calc = round(override_frente * override_fondo, 1) if override_frente > 0 and override_fondo > 0 else 0.0
            if _area_calc > 0:
                st.markdown(f"Área calculada: **{_area_calc:,.1f} m²** (frente × fondo)")

            override_area = st.number_input("Área del terreno (m²)", min_value=0.0, max_value=50000.0,
                                            value=float(st.session_state.get("cab_override_area", _area_calc)), step=10.0, key="cab_area_inp")
            override_al   = st.number_input("Área libre mínima (%)", min_value=0.0, max_value=80.0,
                                            value=0.0, step=5.0,
                                            key="cab_override_al_inp")

        with st.expander("Colindantes", expanded=False):
            st.caption("Norma por Colindancia — Se activa cuando las edificaciones colindantes superan la altura permitida, otorgando beneficio de altura al proyecto.")
            _col_izq, _col_der = st.columns(2)
            colind_izq = _col_izq.number_input("Colindante izq. (pisos)", min_value=0, max_value=40,
                                                value=0, step=1,
                                                key="cab_colind_izq_inp")
            colind_der = _col_der.number_input("Colindante der. (pisos)", min_value=0, max_value=40,
                                                value=0, step=1,
                                                key="cab_colind_der_inp")
            if colind_izq > 0 or colind_der > 0:
                _col_max = max(colind_izq, colind_der)
                st.caption(f"Regla colindancia: edificio a analizar podrá alcanzar hasta el promedio "
                           f"con el colindante más alto ({_col_max} pisos). Claude calculará la altura permitida.")

        st.markdown("---")
        st.markdown("### DATOS FINANCIEROS")
        _financ_inputs_ss = st.session_state.get("financ_inputs") or {}

        # Inicializar keys de widgets solo si no existen (evita conflicto value/key en Streamlit)
        if "cab_precio_compra_inp" not in st.session_state:
            _pcompra_init = int(
                ((_financ_inputs_ss.get("costo_terreno") or 0) if isinstance(_financ_inputs_ss, dict) else 0)
                or 1
            )
            st.session_state["cab_precio_compra_inp"] = max(1, _pcompra_init)
        if "cab_pventa_inp" not in st.session_state:
            st.session_state["cab_pventa_inp"] = 0
        if "cab_cconst_inp" not in st.session_state:
            st.session_state["cab_cconst_inp"] = 700

        precio_compra   = st.number_input("Precio de compra del inmueble (USD)",
                                          min_value=1, max_value=50_000_000,
                                          step=10_000, format="%d", key="cab_precio_compra_inp")
        precio_venta_m2 = st.number_input("Precio de venta / m² (USD)",
                                          min_value=0, max_value=15_000,
                                          step=100, key="cab_pventa_inp")
        _ref_p1 = MERCADO.get(zona, {}).get("precio_1br", 0)
        _ref_p2 = MERCADO.get(zona, {}).get("precio_2br", 0)
        _ref_p3 = MERCADO.get(zona, {}).get("precio_3br", 0)
        _m_zona = MERCADO.get(zona, {})
        _def_estac = int((_financ_inputs_ss.get("precio_estac") or 0) or _m_zona.get("precio_estac", 0))
        _def_dep   = int((_financ_inputs_ss.get("precio_deposito") or 0) or _m_zona.get("precio_deposito", 0))
        _col_pe, _col_pd = st.columns(2)
        with _col_pe:
            precio_estac_inp = st.number_input(
                "Precio cochera (USD)", min_value=0, max_value=100_000,
                value=_def_estac, step=500, key="cab_precio_estac_inp")
        with _col_pd:
            precio_deposito_inp = st.number_input(
                "Precio depósito (USD)", min_value=0, max_value=50_000,
                value=_def_dep, step=250, key="cab_precio_deposito_inp")
        costo_const_m2  = st.number_input("Costo construcción dptos / m² (USD)",
                                          min_value=300, max_value=3_000,
                                          step=25, key="cab_cconst_inp")
        with st.expander("Costos avanzados"):
            costo_sotano_m2 = st.number_input("Costo sótano / m² (USD)", 200, 1000, 450, 25,
                                                key="cab_costo_sotano_inp")
            _fi_arq = float((_financ_inputs_ss.get("costo_arq_m2") or 5.94))
            _fi_esp = float((_financ_inputs_ss.get("costo_esp_m2") or 7.92))
            _fi_fac = int((_financ_inputs_ss.get("costo_factibilidades") or 17000))
            _col_arq, _col_esp = st.columns(2)
            with _col_arq:
                costo_arq_m2 = st.number_input(
                    "Planos arquitectura / m² (USD)", 1.0, 50.0, _fi_arq, 0.5,
                    key="cab_costo_arq_inp")
            with _col_esp:
                costo_esp_m2 = st.number_input(
                    "Especialidades / m² (USD)", 1.0, 50.0, _fi_esp, 0.5,
                    key="cab_costo_esp_inp")
            costo_factibilidades = st.number_input(
                "Factibilidades (USD)", 0, 100_000, _fi_fac, 1000,
                key="cab_costo_factib_inp")
            _estructura_opts = ["Terreno como aporte (estándar)", "Banco financia terreno (con track record)"]
            _estructura_sel  = st.radio(
                "Estructura de financiamiento",
                _estructura_opts, index=0,
                key="cab_estructura_financ_inp")
            estructura_financ = "estandar" if _estructura_sel == _estructura_opts[0] else "con_terreno"
            if estructura_financ == "con_terreno":
                aporte_propio_pct = st.number_input(
                    "Aporte inicial terreno (%)", 10.0, 40.0, 20.0, 5.0,
                    key="cab_aporte_propio_inp")
            else:
                aporte_propio_pct = 100.0
                st.caption("Terreno 100% equity → sirve de garantía hipotecaria para la línea de construcción")
            st.markdown("---")
            # Preventa
            import math as _m_sb
            _n_unid_sb   = int((st.session_state.get("cabida") or {}).get("total_unidades", 20) or 20)
            _vel_sb      = float(MERCADO.get(st.session_state.get("zona", ""), {}).get("velocidad_venta", 1.5) or 1.5)
            _fi_sb       = st.session_state.get("financ_inputs") or {}
            pct_preventa_banco = st.number_input(
                "Preventa mínima exigida por banco (%)", 10.0, 50.0,
                float(_fi_sb.get("pct_preventa_banco", 30.0)), 5.0,
                key="cab_pct_preventa_banco_inp")
            _req_sb  = max(1, _m_sb.ceil(_n_unid_sb * pct_preventa_banco / 100))
            _auto_pv = max(1, _m_sb.ceil(_req_sb / _vel_sb))
            st.caption(f"Preventa requerida: {_req_sb} unidades ({pct_preventa_banco:.0f}%) · velocidad {_vel_sb:.1f}/mes → {_auto_pv} meses calculado")
            _pv_override = st.number_input(
                "Meses de preventa (0 = automático)", 0, 36,
                int(_fi_sb.get("meses_preventa_override") or 0),
                1, key="cab_meses_preventa_inp")
            meses_preventa_override = int(_pv_override) if _pv_override > 0 else None
            _n_pisos_sb   = int((st.session_state.get("cabida") or {}).get("num_pisos", 7) or 7)
            _obra_auto_sb = 24 if _n_pisos_sb > 20 else (12 if _n_pisos_sb <= 5 else 16)
            _obra_override = st.number_input(
                "Plazo de construcción (0 = automático)", 0, 60,
                int(_fi_sb.get("meses_obra_override") or 0),
                1, key="cab_meses_obra_inp")
            st.caption(f"Plazo auto: {_obra_auto_sb} meses · {_n_pisos_sb} piso{'s' if _n_pisos_sb != 1 else ''}")
            meses_obra_override = int(_obra_override) if _obra_override > 0 else None
            pct_mktg_preventa = st.number_input(
                "Costo marketing preventa (% de ventas+gerenc.)", 0.5, 10.0,
                float(_fi_sb.get("pct_mktg_preventa", 2.0)), 0.5,
                key="cab_pct_mktg_pv_inp")
            fee_constructora = st.number_input("Fee constructora (%)", 0.0, 20.0, 10.0, 0.5,
                                                key="cab_fee_constructora_inp")
            tasa_ir          = st.number_input("Impuesto a la Renta (%)", 0.0, 40.0, 29.5, 0.5,
                                                key="cab_tasa_ir_inp")
            include_alcabala = st.checkbox("Incluir Alcabala (3%)", value=True,
                                            key="cab_include_alcabala_inp")
            include_dd       = st.checkbox("Incluir Due Diligence ($10,000)", value=True,
                                            key="cab_include_dd_inp")
            include_demo     = st.checkbox("Incluir costo demolición", value=False,
                                            key="cab_include_demo_inp")
            if include_demo:
                _area_t_demo = int((st.session_state.params or {}).get("area_terreno", 500)) if st.session_state.get("params") else 500
                costo_demolicion = st.number_input(
                    "Costo demolición (USD)",
                    min_value=0, max_value=500_000,
                    value=int(_area_t_demo * 60),
                    step=1_000,
                    key="cab_costo_demo_inp",
                )
            else:
                costo_demolicion = 0


        st.markdown("---")
        st.markdown("### DOCUMENTOS DE REFERENCIA")
        st.caption("Sube partidas, fichas técnicas, capturas o información de inmuebles alternativos que quieras considerar.")
        _ref_files = st.file_uploader(
            "PDF, imágenes o documentos",
            type=["pdf", "jpg", "jpeg", "png", "webp", "xlsx", "xls"],
            accept_multiple_files=True,
            key="ref_docs_inm",
            label_visibility="collapsed",
        )
        if _ref_files:
            st.session_state["ref_docs_inm_bytes"] = [
                {"name": f.name, "size": len(f.read())} for f in _ref_files
            ]
            st.markdown(
                f'<div style="font-size:10px;color:rgba(107,206,160,0.85);'
                f'background:rgba(107,206,160,0.08);border-radius:6px;'
                f'padding:6px 10px;border-left:2px solid rgba(107,206,160,0.4);">'
                f'✓ {len(_ref_files)} documento(s) adjunto(s)</div>',
                unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### INSTRUCCIONES AL ANÁLISIS")
        st.caption("Lo que escribas aquí guía a la IA — tipologías, restricciones, preferencias del cliente.")
        sugerencias = st.text_area(
            label="instrucciones",
            placeholder="Ej: uso mixto en primer piso, acumulación de lotes, maximizar unidades de 2 dormitorios...",
            height=90,
            label_visibility="collapsed",
            key="cab_sugerencias_inp"
        )

        st.markdown("---")
        run = st.button("GENERAR ANÁLISIS", use_container_width=True, type="primary")

        # ── RETROALIMENTACIÓN ─────────────────────────────
        if st.session_state.get("params") or st.session_state.get("cabida"):
            st.markdown("---")
            st.markdown(
                '<div style="font-size:9px;color:rgba(184,144,74,0.90);letter-spacing:2px;'
                'text-transform:uppercase;font-weight:700;margin-bottom:8px;">Retroalimentar a Solum</div>'
                '<div style="font-size:10px;color:rgba(200,216,232,0.75);line-height:1.6;margin-bottom:10px;">'
                '¿Ya consolidaste este análisis con tu criterio profesional? Compártelo con Solum — '
                'así aprendo tu forma de trabajar y el próximo análisis saldrá más alineado a tu visión desde el inicio.'
                '</div>', unsafe_allow_html=True)
            _fb_file = st.file_uploader(
                "Excel, PDF, DWG o imagen",
                type=["xlsx", "xls", "pdf", "dwg", "dxf", "jpg", "jpeg", "png"],
                key="fb_upload_inm",
                label_visibility="collapsed",
            )
            if _fb_file:
                _fb_bytes = _fb_file.read()
                _fb_proyecto = st.session_state.get("_last_inm_proy")
                _fb_id = getattr(_fb_proyecto, "_id", None) if _fb_proyecto else None
                _sb_fb = _get_supabase()
                if _sb_fb and _fb_id:
                    try:
                        _sb_fb.table("proyectos").update({
                            "feedback_filename": _fb_file.name,
                            "feedback_size":     len(_fb_bytes),
                            "feedback_at":       datetime.datetime.utcnow().isoformat(),
                        }).eq("id", _fb_id).execute()
                    except Exception:
                        pass
                st.markdown(
                    '<div style="background:rgba(107,206,160,0.10);border-left:3px solid '
                    'rgba(107,206,160,0.50);border-radius:6px;padding:8px 12px;'
                    'font-size:11px;color:rgba(107,206,160,0.90);font-weight:600;margin-top:6px;">'
                    f'✓ Recibido: {_fb_file.name}</div>', unsafe_allow_html=True)

        # ── GUARDAR / CARGAR PROYECTO ─────────────────────
        st.markdown("---")
        st.markdown("### PROYECTOS")
        proyectos_saved = listar_proyectos()
        if proyectos_saved:
            nombres_saved = [p.name for p in proyectos_saved]
            sel_proy = st.selectbox("Cargar proyecto guardado", ["— seleccionar —"] + nombres_saved,
                                    label_visibility="collapsed")
            if sel_proy != "— seleccionar —":
                if st.button("CARGAR", use_container_width=True):
                    _pref = next((p for p in proyectos_saved if p.name == sel_proy), None)
                    datos = cargar_proyecto(_pref or sel_proy)
                    st.session_state.params        = datos.get("params") or None
                    st.session_state.cabida        = datos.get("cabida") or None
                    st.session_state.financ_inputs = datos.get("financ_inputs") or {}
                    st.session_state.zona          = datos.get("zona") or None
                    st.rerun()

        nombre_proy = st.text_input("Nombre del proyecto", placeholder="Ej: Torres Las Camelias",
                                    label_visibility="collapsed", key="guardar_nombre_proy")
        if st.button("GUARDAR PROYECTO", use_container_width=True, key="btn_guardar_inm"):
            if st.session_state.get("params"):
                fp = guardar_proyecto(nombre_proy or "sin_nombre", {
                    "params":        st.session_state.params,
                    "cabida":        st.session_state.cabida,
                    "financ_inputs": st.session_state.financ_inputs,
                    "financ":        st.session_state.financ,
                    "zona":          st.session_state.zona,
                })
                st.session_state["_last_inm_proy"] = fp
                st.session_state.pop("_share_url_inm", None)
                st.markdown('<div style="background:rgba(107,206,160,0.12);border-left:3px solid rgba(107,206,160,0.50);border-radius:6px;padding:8px 12px;color:rgba(107,206,160,0.90);font-size:12px;font-weight:600;margin-top:8px;">✓ ' + f"Guardado: {fp.name}" + '</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="alert-info">ℹ️ ' + "Genera el análisis primero." + '</div>', unsafe_allow_html=True)
        _lp_inm = st.session_state.get("_last_inm_proy")
        if _lp_inm and _lp_inm._id:
            if st.button("GENERAR LINK DE COMPARTIR", use_container_width=True, key="btn_share_inm"):
                _tok = str(uuid.uuid4())
                _sb2 = _get_supabase()
                if _sb2:
                    _sb2.table("proyectos").update({"share_token": _tok}).eq("id", _lp_inm._id).execute()
                    _base = (st.secrets.get("app", {}) or {}).get("base_url", "http://localhost:8501")
                    st.session_state["_share_url_inm"] = f"{_base}?share={_tok}"
            if st.session_state.get("_share_url_inm"):
                st.code(st.session_state["_share_url_inm"])
                st.caption("Comparte con tu cliente — no requiere contraseña")

        st.markdown("---")
        st.markdown("### FOTOS DEL INMUEBLE")
        st.caption("Las fotos se incluyen en el reporte PDF")
        _cab_fotos = st.file_uploader(
            "Sube fotos del inmueble / terreno",
            type=["jpg", "jpeg", "png", "webp"],
            accept_multiple_files=True,
            key="cab_fotos_upload",
        )
        if _cab_fotos:
            st.session_state["cab_fotos_bytes"] = [f.read() for f in _cab_fotos]
            st.session_state["cab_fotos_nombres"] = [f.name for f in _cab_fotos]
            cols_f = st.columns(min(len(_cab_fotos), 3))
            for _fi, _fup in enumerate(_cab_fotos[:3]):
                cols_f[_fi].image(_fup, use_container_width=True)

        st.markdown("---")
        st.markdown("### ANÁLISIS LEGAL")
        st.caption("Verifica titularidad, cargas, hipotecas y estado registral")

        _sb_has_partida = pdf_partida is not None
        _sb_has_puhr    = pdf_puhr is not None
        _sb_has_session_p = st.session_state.get("partida_bytes") is not None
        _sb_has_session_u = st.session_state.get("puhr_bytes") is not None
        _sb_can_run = _sb_has_partida or _sb_has_puhr or _sb_has_session_p or _sb_has_session_u

        if not _sb_can_run:
            st.markdown(
                '<div style="font-size:11px;color:rgba(255,255,255,0.45);'
                'background:rgba(255,255,255,0.06);border-radius:6px;'
                'padding:8px 12px;border-left:2px solid rgba(184,144,74,0.40);">Adjunta la '
                '<b style="color:rgba(255,255,255,0.65);">Partida Registral</b> y/o '
                '<b style="color:rgba(255,255,255,0.65);">PU/HR</b> en la sección '
                '<b style="color:rgba(184,144,74,0.80);">DOCUMENTOS</b> para habilitar este análisis.</div>',
                unsafe_allow_html=True)
        else:
            _sb_docs_list = []
            if _sb_has_partida or _sb_has_session_p: _sb_docs_list.append("Partida Registral")
            if _sb_has_puhr    or _sb_has_session_u: _sb_docs_list.append("PU / HR")
            st.markdown(
                f'<div style="font-size:11px;color:#7BCFA0;background:rgba(107,206,160,0.12);border-radius:6px;'
                f'padding:6px 10px;border-left:2px solid rgba(107,206,160,0.50);margin-bottom:6px;">'
                f'● {" · ".join(_sb_docs_list)}</div>',
                unsafe_allow_html=True)

            if st.button("ANALIZAR DOCUMENTOS LEGALES", use_container_width=True,
                         key="btn_legal_sidebar", type="secondary"):
                if pdf_partida is not None:
                    st.session_state.partida_bytes = pdf_partida.read()
                if pdf_puhr is not None:
                    st.session_state.puhr_bytes = pdf_puhr.read()
                _p_bytes = st.session_state.get("partida_bytes")
                _u_bytes = st.session_state.get("puhr_bytes")
                _c_bytes = pdf_cert.read() if pdf_cert is not None else None
                if _p_bytes or _u_bytes:
                    st.session_state.legal = _run_with_retry(
                        lambda _p=_p_bytes, _u=_u_bytes, _c=_c_bytes: analizar_legal(_p, _u, _c),
                        "Analizando documentos registrales y parámetros urbanísticos…",
                    )

        if st.session_state.get("legal"):
            _sb_lg  = st.session_state.legal
            _sb_sem = _sb_lg.get("semaforo", "amarillo").lower()
            _sb_sem_map = {
                "verde":    ("rgba(107,206,160,0.90)", "rgba(107,206,160,0.12)", "rgba(107,206,160,0.40)", "✓ SIN ALERTAS CRÍTICAS"),
                "amarillo": ("rgba(232,200,122,0.90)", "rgba(184,144,74,0.12)",  "rgba(184,144,74,0.40)",  "⚠ OBSERVACIONES MENORES"),
                "rojo":     ("rgba(232,100,100,0.90)", "rgba(200,60,60,0.12)",   "rgba(200,80,80,0.40)",   "✕ ALERTAS CRÍTICAS"),
            }
            _sb_sc, _sb_sbg, _sb_bdr, _sb_sl = _sb_sem_map.get(_sb_sem, ("rgba(200,210,220,0.7)", "rgba(255,255,255,0.06)", "rgba(255,255,255,0.2)", "— INDETERMINADO"))
            st.markdown(
                f'<div style="background:{_sb_sbg};border-left:3px solid {_sb_bdr};border-radius:6px;'
                f'padding:8px 10px;margin-top:4px;font-size:11px;color:{_sb_sc};font-weight:700;">'
                f'{_sb_sl}</div>',
                unsafe_allow_html=True)
            st.caption("Ver detalles completos en la pestaña Legal →")

        st.markdown("---")
        st.markdown("### PROPUESTA")
        st.caption("Genera una propuesta formal de compra o arrendamiento")
        prop_tipo        = st.selectbox("Tipo de propuesta", ["Compra", "Arrendamiento"],
                                        key="prop_tipo")
        prop_propietario = st.text_input("Propietario(s)", placeholder="Nombre del vendedor/arrendador",
                                         key="prop_propietario")
        prop_precio      = st.number_input(
            "Precio ofertado (USD)" if prop_tipo == "Compra" else "Renta mensual ofertada (USD)",
            min_value=0, max_value=50_000_000, value=0, step=5_000, format="%d",
            key="prop_precio"
        )
        _fin_ss = st.session_state.get("financ")
        if _fin_ss:
            _r_ss  = _fin_ss.get("resumen", {})
            _v20_ss = _r_ss.get("max_terreno_20pct", 0)
            _v15_ss = _r_ss.get("max_terreno_15pct", 0)

        # ── Estructura de compra / pago ─────────────────
        with st.expander("Estructura de Compra / Pago", expanded=True):
            _tiene_opcion = st.checkbox("Incluir opción de compra", value=True, key="prop_opcion")
            if _tiene_opcion:
                _dias_opcion = st.number_input("Plazo opción (días)", 1, 360, 90, 15, key="prop_dias_opcion")
                _pct_opcion  = st.number_input("Monto opción (% del precio)", 0.0, 20.0, 0.0, 0.5,
                                               format="%.1f", key="prop_pct_opcion")
            else:
                _dias_opcion = 0
                _pct_opcion  = 0.0

            st.caption("Estructura de pagos al precio total")
            _pct_minuta   = st.number_input("Pago inicial a la firma de minuta (%)", 0.0, 100.0, 20.0, 5.0,
                                            format="%.1f", key="prop_pct_minuta")
            _condicion_minuta = st.text_input("Condición para la minuta",
                                              value="Aprobación del anteproyecto por la Municipalidad",
                                              key="prop_cond_minuta")
            _pct_escritura = round(100.0 - _pct_minuta - _pct_opcion, 1)
            st.info(f"Saldo a escritura pública: **{_pct_escritura:.1f}%**")
            _condicion_escritura = st.text_input("Condición para escritura pública",
                                                  value="Desocupación y entrega del inmueble libre de cargas",
                                                  key="prop_cond_escritura")

        prop_plazo       = st.number_input("Plazo de respuesta (días)", 1, 90, 10, 1, key="prop_plazo")
        prop_condiciones = st.text_area(
            "Condiciones adicionales",
            placeholder="Ej:\nDue diligence de 30 días\nInmueble libre de cargas e hipotecas\nServicios al día",
            height=80, key="prop_condiciones"
        )


    # ── MÓDULO 2: LOGÍSTICO / INDUSTRIAL ─────────────────
    elif tipo_op == "Proyecto Logístico / Industrial":
        run = False
        run_residencial = False

        # ── Indicador de progreso guiado Industrial ───────
        _ind1_done = int(st.session_state.get("ind_costo_terreno", 0) or 0) > 0
        _ind2_done = float(st.session_state.get("ind_area_nave_m2", 0) or 0) > 0
        _ind3_done = st.session_state.get("ind_analizado") is True
        _ind4_done = _ind3_done
        _steps_ind = [
            ("Terreno",        "Ubicación, área, costo",       _ind1_done),
            ("Proyecto",       "Nave, actividad, % techada",   _ind2_done),
            ("Análisis",       "Uso, renta, yield, DSCR",      _ind3_done),
            ("Financiamiento", "Terreno + obra",               _ind3_done),
            ("Reporte",        "Resumen ejecutivo PDF",        _ind4_done),
        ]
        _cur_ind = next((i for i, (_, _, d) in enumerate(_steps_ind) if not d), 5)
        _ind_html = "".join([_sp(i,l,s,d,i==_cur_ind,i==4) for i,(l,s,d) in enumerate(_steps_ind)])
        st.markdown(
            f'<div style="background:rgba(255,255,255,0.04);border-radius:10px;padding:12px 14px 10px;'
            f'border:1px solid rgba(255,255,255,0.08);margin-bottom:10px;">'
            f'<div style="font-size:9px;font-weight:700;color:rgba(184,200,216,0.45);'
            f'letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">FLUJO DE TRABAJO</div>'
            f'{_ind_html}</div>',
            unsafe_allow_html=True)

        # ── Estado vacío guiado Industrial ────────────────
        _ind_empty = (not _ind1_done and not _ind2_done)
        if _ind_empty:
            st.markdown(
                '<div style="background:rgba(184,144,74,0.08);border-radius:8px;padding:11px 13px;'
                'border:1px solid rgba(184,144,74,0.22);margin-bottom:10px;">'
                '<div style="font-size:10px;font-weight:700;color:#D4A853;letter-spacing:0.5px;margin-bottom:7px;">¿CUÁNDO USAR ESTE MÓDULO?</div>'
                '<div style="font-size:11px;color:rgba(184,200,216,0.82);line-height:1.75;">'
                '✓ <b style="color:#C8D8E8;">Nave logística</b> o almacén a construir<br>'
                '✓ <b style="color:#C8D8E8;">Planta industrial</b> o manufactura<br>'
                '✓ <b style="color:#C8D8E8;">Cross-docking</b> o centro de distribución<br>'
                '✗ No para departamentos — usa <b style="color:#D4A853;">Proyecto Inmobiliario</b>'
                '</div>'
                '<div style="margin-top:9px;padding-top:8px;border-top:1px solid rgba(255,255,255,0.08);'
                'font-size:10px;font-weight:700;color:#D4A853;letter-spacing:0.5px;margin-bottom:5px;">¿POR DÓNDE EMPEZAR?</div>'
                '<div style="font-size:11px;color:rgba(184,200,216,0.82);line-height:1.75;">'
                '① Selecciona la <b style="color:#C8D8E8;">zona industrial</b> de Lima<br>'
                '② Ingresa el <b style="color:#C8D8E8;">área de la nave</b> proyectada<br>'
                '③ Completa el <b style="color:#C8D8E8;">costo del terreno</b><br>'
                '④ Presiona <b style="color:#D4A853;">Ejecutar análisis →</b>'
                '</div></div>',
                unsafe_allow_html=True)

        # ── 1 · TERRENO ─────────────────────────────────────
        _step_header("1", "Terreno")
        with st.expander("📄 Importar datos desde documento"):
            st.caption("Sube una ficha técnica, brochure o plano del terreno — la IA extrae área, ubicación y tipo de nave para pre-llenar los campos automáticamente.")
            _ind_doc_up = st.file_uploader(
                "PDF, PPTX o DOCX",
                type=["pdf", "pptx", "ppt", "docx"],
                key="ind_import_doc",
            )
            if _ind_doc_up and st.button("EXTRAER DATOS", key="btn_ind_extract", use_container_width=True):
                _ind_bytes = _ind_doc_up.read()
                _ind_ext = _run_with_retry(
                    lambda _b=_ind_bytes, _n=_ind_doc_up.name: extraer_datos_desde_doc(_b, _n, "industrial"),
                    "Analizando documento…"
                )
                if _ind_ext.get("_error"):
                    st.error(_ind_ext["_error"])
                else:
                    if _ind_ext.get("area_terreno_m2"):
                        st.session_state["ind_area_val"] = float(_ind_ext["area_terreno_m2"])
                        st.session_state["ind_area"] = float(_ind_ext["area_terreno_m2"])
                        st.session_state["ind_area_auto"] = False
                    if _ind_ext.get("frente_ml"):
                        st.session_state["ind_frente"] = float(_ind_ext["frente_ml"])
                    if _ind_ext.get("fondo_ml"):
                        st.session_state["ind_fondo"] = float(_ind_ext["fondo_ml"])
                    if _ind_ext.get("costo_terreno_usd"):
                        st.session_state["ind_costo_terreno"] = int(_ind_ext["costo_terreno_usd"])
                    if _ind_ext.get("tipo_nave"):
                        _tipo_opts = ["Almacén Logístico", "Nave Industrial", "Cross-docking", "Producción / Manufactura"]
                        _ext_tipo = _ind_ext["tipo_nave"]
                        for _t in _tipo_opts:
                            if any(w in _ext_tipo for w in _t.split()[:2]):
                                st.session_state["ind_tipo"] = _t
                                break
                    if _ind_ext.get("renta_usd_m2_mes"):
                        st.session_state["ind_renta"] = float(_ind_ext["renta_usd_m2_mes"])
                    if _ind_ext.get("zonificacion"):
                        _zon_opts = ["I1", "I2", "I3", "I4", "OU"]
                        if _ind_ext["zonificacion"] in _zon_opts:
                            st.session_state["ind_zona_ind"] = _ind_ext["zonificacion"]
                    st.success("Datos extraídos. Revisa y ajusta antes de ejecutar.")
                    st.rerun()

        ind_ubicacion = st.text_input("Dirección / Zona",
            placeholder="Ej: Av. Las Torres 1240, Lurín",
            key="ind_ubicacion")
        ind_zona_lima = st.selectbox("Zona industrial Lima",
            [
                "Lurín / Pachacámac",
                "Villa El Salvador",
                "Chilca",
                "Callao / Bellavista",
                "Ventanilla",
                "Huachipa / Ate",
                "Ate Vitarte",
                "Santa Anita",
                "San Juan de Lurigancho (SJL)",
                "Lurigancho / Chosica",
                "Puente Piedra",
                "Carabayllo",
                "Los Olivos / Independencia",
                "La Victoria / Cercado de Lima",
                "Chorrillos",
                "Otro",
            ],
            key="ind_zona_lima")

        ind_col1, ind_col2 = st.columns(2)
        ind_frente = ind_col1.number_input("Frente (ml)", 0.0, 1000.0, 0.0, 0.5, key="ind_frente")
        ind_fondo  = ind_col2.number_input("Fondo (ml)",  0.0, 1000.0, 0.0, 0.5, key="ind_fondo")
        _ind_area_calc = round(ind_frente * ind_fondo, 1) if ind_frente > 0 and ind_fondo > 0 else 0.0
        if _ind_area_calc > 0:
            st.markdown(f'<div style="font-size:11px;color:#B8904A;font-weight:600;margin-top:-4px;">= {_ind_area_calc:,.0f} m²</div>', unsafe_allow_html=True)
            if st.session_state.get("ind_area_auto", True):
                st.session_state["ind_area_val"] = _ind_area_calc
        _ind_area_default = st.session_state.get("ind_area_val", 5000.0)
        ind_area = st.number_input("Área total del terreno (m²)", 1.0, 500_000.0,
                                   max(1.0, _ind_area_default), 50.0, key="ind_area")
        if abs(ind_area - _ind_area_calc) > 1 and _ind_area_calc > 0:
            st.session_state["ind_area_auto"] = False
        elif _ind_area_calc == 0:
            st.session_state["ind_area_auto"] = True

        ind_costo_terreno = st.number_input("Costo del terreno (USD)", 1, 100_000_000,
                                            max(1, int(st.session_state.get("ind_costo_terreno") or 1_000_000)),
                                            50_000, key="ind_costo_terreno")
        st.markdown(f'<div style="font-size:11px;color:#B8904A;font-weight:600;margin-top:-6px;">= ${ind_costo_terreno:,.0f} USD</div>', unsafe_allow_html=True)

        # ── 2 · PROYECTO ────────────────────────────────────
        _step_header("2", "Proyecto")
        ind_pct_techada = st.number_input(
            "% Área techada (nave)",
            min_value=30.0, max_value=95.0, value=75.0, step=5.0, key="ind_pct_techada")
        _ind_nave = ind_area * ind_pct_techada / 100
        _ind_libre = ind_area * (1 - ind_pct_techada / 100)
        st.markdown(
            f'<div style="font-size:11px;color:#A8C0D8;margin-top:-4px;margin-bottom:4px;">'
            f'Nave: <strong>{_ind_nave:,.0f} m²</strong> · Patios/maniobras: <strong>{_ind_libre:,.0f} m²</strong>'
            f'</div>', unsafe_allow_html=True)

        ind_tipo = st.selectbox("Tipo de nave",
            ["Almacén Logístico", "Nave Industrial", "Cross-docking", "Producción / Manufactura"],
            key="ind_tipo")

        # ── Actividad — campo único: descripción + búsqueda automática en Índice ATN-I ──
        _TIPO_ZONA_MIN = {
            "Almacén Logístico":        "I1",
            "Nave Industrial":          "I1",
            "Cross-docking":            "I1",
            "Producción / Manufactura": "I2",
        }
        _ZONA_ORDEN = {"I1": 1, "I2": 2, "I3": 3, "I4": 4, "OU": 0}

        ind_actividad_desc = st.text_area(
            "Actividad a desarrollar",
            value=st.session_state.get("ind_actividad_desc", ""),
            placeholder="Ej: Almacenamiento y distribución de productos farmacéuticos refrigerados…",
            height=68, key="ind_actividad_desc")

        # Semáforo de compatibilidad (derivado de tipo de nave)
        _zona_sel = st.session_state.get("ind_zona_ind", "I2")
        _zona_min = _TIPO_ZONA_MIN.get(st.session_state.get("ind_tipo", "Almacén Logístico"), "I1")
        _ord_sel  = _ZONA_ORDEN.get(_zona_sel, 0)
        _ord_min  = _ZONA_ORDEN.get(_zona_min, 1)
        if _zona_sel == "OU":
            _norm_icon, _norm_label, _norm_color, _norm_bg = "⛔", "INCOMPATIBLE", "#C44A4A", "rgba(196,74,74,0.12)"
            _norm_sub = "Zonificación OU no admite uso industrial — verificar certificado."
        elif _ord_sel >= _ord_min:
            _norm_icon, _norm_label, _norm_color, _norm_bg = "✓", "COMPATIBLE", "#1A7A4A", "rgba(26,122,74,0.12)"
            _norm_sub = f"Zona {_zona_sel} · mínimo requerido {_zona_min} · Ord. 933-MML ATN-I"
        elif _ord_sel == _ord_min - 1:
            _norm_icon, _norm_label, _norm_color, _norm_bg = "⚠", "CONDICIONADO", "#B8862E", "rgba(184,134,46,0.12)"
            _norm_sub = f"Zona {_zona_sel} — requiere mínimo {_zona_min}. Verificar PDU distrital."
        else:
            _norm_icon, _norm_label, _norm_color, _norm_bg = "⛔", "INCOMPATIBLE", "#C44A4A", "rgba(196,74,74,0.12)"
            _norm_sub = f"Requiere {_zona_min} mínimo — zona {_zona_sel} no es suficiente."
        st.markdown(
            f'<div style="background:{_norm_bg};border-radius:6px;padding:8px 12px;margin:6px 0 10px;'
            f'display:flex;align-items:flex-start;gap:10px;">'
            f'<span style="font-size:15px;line-height:1.2;">{_norm_icon}</span>'
            f'<div>'
            f'<div style="font-size:10px;font-weight:700;letter-spacing:1px;color:{_norm_color};">{_norm_label}</div>'
            f'<div style="font-size:10px;color:#A8B8C8;margin-top:1px;">{_norm_sub}</div>'
            f'</div></div>', unsafe_allow_html=True)
        ind_act_categoria = st.session_state.get("ind_tipo", "Almacén Logístico")

        _COSTO_DEFAULTS = {
            "Almacén Logístico":        (280, "Estructura metálica portal frame, 12–14m clara, losa industrial. "
                                              "Ref. real: Parque Logístico 47 Lima (14,300 m², 13.6m, Clase A) = $270–300/m²"),
            "Nave Industrial":          (300, "Estructura metálica + elementos de concreto, uso mixto. "
                                              "Ref. Lima: $280–350/m² según especificación técnica"),
            "Cross-docking":            (420, "Mayor complejidad: múltiples docks, corredores internos, MEP intensivo. "
                                              "Ref. Lima: $380–500/m²"),
            "Producción / Manufactura": (380, "Losa reforzada (cargas pesadas), instalaciones especiales eléctricas/agua. "
                                              "Ref. Lima: $350–450/m²"),
        }
        _def_nave, _help_nave = _COSTO_DEFAULTS.get(
            st.session_state.get("ind_tipo", "Almacén Logístico"),
            (280, ""))
        st.caption("⚠ Industrial ≠ Residencial: estructura metálica sin acabados → **3-4x más barato por m²**")

        with st.expander("Costos de construcción"):
            ind_zona_ind = st.selectbox("Zonificación (referencia)",
                ["I1", "I2", "I3", "I4", "OU"], index=1, key="ind_zona_ind")
            ind_costo_nave = st.number_input(
                "Costo nave (USD/m²)",
                min_value=1, max_value=2000, value=max(1, _def_nave), step=25, key="ind_costo_nave")
            ind_costo_piso = st.number_input(
                "Costo patios / maniobras (USD/m²)",
                min_value=0, max_value=500, value=80, step=10, key="ind_costo_piso")
            ind_pct_indirectos = st.number_input(
                "Costos indirectos (%)",
                min_value=0.0, max_value=30.0, value=5.0, step=0.5, key="ind_pct_indirectos")

        # ── 3 · ANÁLISIS DE USO ──────────────────────────────
        _step_header("3", "Análisis de Uso")
        ind_uso = st.radio("Propósito del activo",
            ["Uso directo", "Inversión"], key="ind_uso")
        ind_renta = st.number_input(
            "Renta de mercado (USD/m²/mes)",
            0.0, 50.0, 6.5, 0.25, key="ind_renta")
        ind_tipo_contrato = st.radio(
            "Tipo de contrato", ["Anual", "Plurianual (3+ años)"],
            horizontal=True, key="ind_tipo_contrato")
        ind_ajuste_pct   = 0.0
        ind_inicio_ajuste = 2
        if ind_tipo_contrato == "Plurianual (3+ años)":
            _ica1, _ica2 = st.columns(2)
            ind_ajuste_pct    = _ica1.number_input(
                "Ajuste anual (%)", 0.0, 10.0, 3.0, 0.5, key="ind_ajuste_pct")
            ind_inicio_ajuste = _ica2.selectbox(
                "Año de inicio del ajuste", [2, 3], key="ind_inicio_ajuste")

        # ── 4 · FINANCIAMIENTO ───────────────────────────────
        _step_header("4", "Financiamiento")
        ind_alcabala = st.checkbox("Incluir Alcabala (3%)", value=True, key="ind_alcabala")
        with st.expander("Crédito terreno", expanded=False):
            ind_dp_terreno = st.number_input(
                "Downpayment (%)", 0.0, 100.0,
                float(st.session_state.get("ind_dp_terreno", 40.0)), 5.0,
                key="ind_dp_terreno")
            ind_tasa_terreno = st.number_input(
                "Tasa (% anual)", 0.0, 30.0,
                float(st.session_state.get("ind_tasa_terreno", 8.0)), 0.25,
                key="ind_tasa_terreno")
            ind_plazo_terreno = st.number_input(
                "Plazo (años)", 1, 30,
                int(st.session_state.get("ind_plazo_terreno", 10)), 1,
                key="ind_plazo_terreno")
        with st.expander("Crédito obra", expanded=False):
            ind_dp_const = st.number_input(
                "Downpayment (%)", 0.0, 100.0,
                float(st.session_state.get("ind_dp_const", 30.0)), 5.0,
                key="ind_dp_const")
            ind_tasa_const = st.number_input(
                "Tasa (% anual)", 0.0, 30.0,
                float(st.session_state.get("ind_tasa_const", 9.0)), 0.25,
                key="ind_tasa_const")
            ind_plazo_const = st.number_input(
                "Plazo (años)", 1, 20,
                int(st.session_state.get("ind_plazo_const", 8)), 1,
                key="ind_plazo_const")

        # ── 5 · DOCUMENTOS & FOTOS ──────────────────────────
        _step_header("5", "Documentos & Fotos")
        st.caption("Sube los documentos del proyecto para análisis de factibilidad")
        _ind_secrets_key = (st.secrets.get("anthropic", {}) or {}).get("api_key", "")
        if not _ind_secrets_key and not st.session_state.get("api_key_input"):
            with st.expander("⚙ Configuración API", expanded=True):
                _ind_api_k = st.text_input("Clave de acceso", type="password",
                                           placeholder="••••••••••••••••", key="ind_api_key_inp",
                                           value=st.session_state.get("api_key_input", ""))
                if _ind_api_k:
                    st.session_state["api_key_input"] = _sanitize_api_key(_ind_api_k)
        if _ind_secrets_key or st.session_state.get("api_key_input"):
            st.markdown(
                '<div style="font-size:10px;color:#1A4731;background:#E8F5EE;'
                'border-radius:4px;padding:5px 10px;margin-bottom:6px;'
                'border-left:2px solid #1A4731;">● Sistema activo</div>',
                unsafe_allow_html=True)
        ind_doc_partida = st.file_uploader("Partida Registral (SUNARP)", type="pdf", key="ind_doc_partida")
        ind_doc_params  = st.file_uploader("Certificado de Parámetros",  type="pdf", key="ind_doc_params")
        ind_doc_zon     = st.file_uploader("Cert. Zonificación y Vías",  type="pdf", key="ind_doc_zon")
        ind_doc_planos  = st.file_uploader("Planos del Inmueble (PDF, DWG o DXF)", type=["pdf", "dwg", "dxf"], key="ind_doc_planos")
        if ind_doc_planos and ind_doc_planos.name.lower().endswith((".dwg", ".dxf")):
            st.info("DWG/DXF cargado como referencia. Para incluirlo en el análisis de documentos, expórtalo como PDF desde AutoCAD: Archivo → Exportar → PDF. El polígono geométrico se procesa en el paso de Cabida Industrial.")
        _ind_has_docs = any([ind_doc_partida, ind_doc_params, ind_doc_zon, ind_doc_planos])
        run_ind_docs = False  # se dispara junto con EJECUTAR ANÁLISIS si hay docs

        with st.expander("Fotos del inmueble", expanded=False):
            st.caption("Las fotos se incluyen en el reporte")
            _ind_fotos = st.file_uploader(
                "Sube fotos del terreno / nave",
                type=["jpg", "jpeg", "png", "webp"],
                accept_multiple_files=True,
                key="ind_fotos_upload",
            )
            if _ind_fotos:
                st.session_state["ind_fotos_bytes"] = [f.read() for f in _ind_fotos]
                st.session_state["ind_fotos_nombres"] = [f.name for f in _ind_fotos]
                cols_if = st.columns(min(len(_ind_fotos), 3))
                for _ifi, _ifup in enumerate(_ind_fotos[:3]):
                    cols_if[_ifi].image(_ifup, use_container_width=True)

        st.markdown("---")
        st.caption("Instrucciones al análisis — tipo de nave, restricciones, preferencias del cliente.")
        ind_sugerencias = st.text_area(
            label="instrucciones_ind",
            placeholder="Ej: priorizar altura libre de 12m, acceso para tráileres, cámara frigorífica en sector norte...",
            height=90,
            label_visibility="collapsed",
            key="ind_sugerencias_inp"
        )

        # ── EJECUTAR ─────────────────────────────────────────
        st.markdown("---")
        run_industrial = st.button("EJECUTAR ANÁLISIS", use_container_width=True, type="primary")

        # ── PROYECTOS GUARDADOS ──────────────────────────────
        with st.expander("💾 Proyectos guardados"):
            _ind_saved = listar_proyectos()
            _ind_nombres = [p.name for p in _ind_saved]
            if _ind_nombres:
                _ind_sel = st.selectbox("Cargar proyecto", ["— seleccionar —"] + _ind_nombres,
                                        label_visibility="collapsed", key="ind_sel_proy")
                if _ind_sel != "— seleccionar —":
                    if st.button("CARGAR", use_container_width=True, key="btn_ind_cargar"):
                        _ind_pref = next((p for p in _ind_saved if p.name == _ind_sel), None)
                        _d = cargar_proyecto(_ind_pref or _ind_sel)
                        if _d.get("industrial_result"):
                            st.session_state.industrial_result = _d["industrial_result"]
                            st.rerun()
            _ind_nombre_proy = st.text_input("Nombre del proyecto", placeholder="Ej: Nave Lurín",
                                              label_visibility="collapsed", key="ind_nombre_proy")
            if st.button("GUARDAR PROYECTO", use_container_width=True, key="btn_ind_guardar"):
                if st.session_state.get("industrial_result"):
                    fp = guardar_proyecto(_ind_nombre_proy or "industrial_sin_nombre",
                                          {"industrial_result": st.session_state.industrial_result})
                    st.session_state["_last_ind_proy"] = fp
                    st.session_state.pop("_share_url_ind", None)
                    st.markdown(f'<div style="background:rgba(107,206,160,0.12);border-left:3px solid rgba(107,206,160,0.50);'
                                f'border-radius:6px;padding:8px 12px;color:rgba(107,206,160,0.90);font-size:12px;font-weight:600;margin-top:6px;">'
                                f'✓ {fp.name}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="alert-info">ℹ️ ' + "Ejecuta el análisis primero." + '</div>', unsafe_allow_html=True)
            _lp_ind = st.session_state.get("_last_ind_proy")
            if _lp_ind and _lp_ind._id:
                if st.button("GENERAR LINK DE COMPARTIR", use_container_width=True, key="btn_share_ind"):
                    _tok = str(uuid.uuid4())
                    _sb2 = _get_supabase()
                    if _sb2:
                        _sb2.table("proyectos").update({"share_token": _tok}).eq("id", _lp_ind._id).execute()
                        _base = (st.secrets.get("app", {}) or {}).get("base_url", "http://localhost:8501")
                        st.session_state["_share_url_ind"] = f"{_base}?share={_tok}"
                if st.session_state.get("_share_url_ind"):
                    st.code(st.session_state["_share_url_ind"])
                    st.caption("Comparte con tu cliente — no requiere contraseña")

        # ── RETROALIMENTACIÓN ─────────────────────────────
        if st.session_state.get("industrial_result"):
            st.markdown("---")
            st.markdown(
                '<div style="font-size:9px;color:rgba(184,144,74,0.90);letter-spacing:2px;'
                'text-transform:uppercase;font-weight:700;margin-bottom:8px;">Retroalimentar a Solum</div>'
                '<div style="font-size:10px;color:rgba(200,216,232,0.75);line-height:1.6;margin-bottom:10px;">'
                '¿Ya consolidaste este análisis con tu criterio profesional? Compártelo con Solum — '
                'así aprendo tu forma de trabajar y el próximo análisis saldrá más alineado a tu visión desde el inicio.'
                '</div>', unsafe_allow_html=True)
            _fb_file_ind = st.file_uploader(
                "Excel, PDF, DWG o imagen",
                type=["xlsx", "xls", "pdf", "dwg", "dxf", "jpg", "jpeg", "png"],
                key="fb_upload_ind",
                label_visibility="collapsed",
            )
            if _fb_file_ind:
                _fb_bytes_ind = _fb_file_ind.read()
                _fb_proy_ind = st.session_state.get("_last_ind_proy")
                _fb_id_ind = getattr(_fb_proy_ind, "_id", None) if _fb_proy_ind else None
                _sb_fb_ind = _get_supabase()
                if _sb_fb_ind and _fb_id_ind:
                    try:
                        _sb_fb_ind.table("proyectos").update({
                            "feedback_filename": _fb_file_ind.name,
                            "feedback_size":     len(_fb_bytes_ind),
                            "feedback_at":       datetime.datetime.utcnow().isoformat(),
                        }).eq("id", _fb_id_ind).execute()
                    except Exception:
                        pass
                st.markdown(
                    '<div style="background:rgba(107,206,160,0.10);border-left:3px solid '
                    'rgba(107,206,160,0.50);border-radius:6px;padding:8px 12px;'
                    'font-size:11px;color:rgba(107,206,160,0.90);font-weight:600;margin-top:6px;">'
                    f'✓ Recibido: {_fb_file_ind.name}</div>', unsafe_allow_html=True)

    # ── MÓDULO 3: INMUEBLE RESIDENCIAL ────────────────────
    elif tipo_op == "Inmueble Residencial":
        run = False
        run_industrial = False

        # ── 1 · UBICACIÓN ──────────────────────────────────
        _step_header("1", "Ubicación")
        with st.expander("📄 Importar datos desde documento"):
            st.caption("Sube una ficha técnica, tasación o brochure del inmueble — la IA extrae automáticamente área, antigüedad, precio y ubicación para pre-llenar los campos.")
            _res_doc_up = st.file_uploader(
                "PDF, PPTX o DOCX",
                type=["pdf", "pptx", "ppt", "docx"],
                key="res_import_doc",
            )
            if _res_doc_up and st.button("EXTRAER DATOS", key="btn_res_extract", use_container_width=True):
                _res_bytes = _res_doc_up.read()
                _res_ext = _run_with_retry(
                    lambda _b=_res_bytes, _n=_res_doc_up.name: extraer_datos_desde_doc(_b, _n, "residencial"),
                    "Analizando documento…"
                )
                if _res_ext.get("_error"):
                    st.error(_res_ext["_error"])
                else:
                    if _res_ext.get("area_m2"):
                        st.session_state["res_m2_k"] = int(_res_ext["area_m2"])
                    if _res_ext.get("antiguedad_anios") is not None:
                        st.session_state["res_antig_k"] = int(_res_ext["antiguedad_anios"])
                    if _res_ext.get("precio_usd"):
                        st.session_state["res_precio_k"] = int(_res_ext["precio_usd"])
                    if _res_ext.get("alquiler_mes_usd"):
                        st.session_state["res_alq_k"] = int(_res_ext["alquiler_mes_usd"])
                    _ext_dorm = _res_ext.get("dormitorios") or ""
                    _dorm_opts_ext = ["1 Dormitorio", "2 Dormitorios", "3 Dormitorios", "Dúplex / Otro"]
                    for _d in _dorm_opts_ext:
                        if any(x in _ext_dorm for x in (_d[:2], _d.split()[0])):
                            st.session_state["res_dorm_k"] = _d
                            break
                    _ext_dist = _res_ext.get("distrito") or ""
                    for _k in MERCADO.keys():
                        if _ext_dist.lower() in _k.lower() or _k.lower() in _ext_dist.lower():
                            st.session_state["res_zona_sel"] = _k
                            break
                    st.success("Datos extraídos. Revisa y ajusta antes de ejecutar.")
                    st.rerun()

        _res_zona_keys = list(MERCADO.keys())
        _res_zona_saved = st.session_state.get("res_zona_sel", "")
        _res_zona_idx = _res_zona_keys.index(_res_zona_saved) if _res_zona_saved in _res_zona_keys else min(20, len(_res_zona_keys) - 1)
        res_zona = st.selectbox("Ubicación", _res_zona_keys,
                                 index=_res_zona_idx, key="res_zona_sel")
        _m_res = MERCADO.get(res_zona, {})

        # ── 2 · INMUEBLE ───────────────────────────────────
        _step_header("2", "Inmueble")
        res_col1, res_col2 = st.columns(2)
        res_m2 = res_col1.number_input("Área (m²)", 1, 2000, max(1, int(st.session_state.get("res_m2_k", 80))), 5, key="res_m2_k")
        res_antiguedad = res_col2.number_input("Antigüedad (años)", 0, 100, int(st.session_state.get("res_antig_k", 5)), 1, key="res_antig_k")
        _dorm_opts = ["1 Dormitorio", "2 Dormitorios", "3 Dormitorios", "Dúplex / Otro"]
        try:
            _dorm_saved = st.session_state.get("res_dorm_k", "")
            _dorm_default_idx = _dorm_opts.index(_dorm_saved) if _dorm_saved in _dorm_opts else 1
        except (ValueError, KeyError):
            _dorm_default_idx = 1
        res_dormitorios = st.selectbox("Tipología", _dorm_opts, index=_dorm_default_idx, key="res_dorm_k")

        # ── 3 · PRECIO ─────────────────────────────────────
        _step_header("3", "Precio")
        _precio_ref = (_m_res.get("precio_1br", 0) if "1" in res_dormitorios else
                       _m_res.get("precio_2br", 0) if "2" in res_dormitorios else
                       _m_res.get("precio_3br", 0)) * res_m2
        _ref_m2_display = (_m_res.get("precio_2br", 0) if "2" in res_dormitorios else
                           (_m_res.get("precio_1br", 0) if "1" in res_dormitorios else _m_res.get("precio_3br", 0)))
        st.markdown(f'<div style="font-size:11px;color:#B8904A;font-weight:600;margin-top:-4px;">Mercado: ${_ref_m2_display:,}/m² · Ref. total ${_precio_ref:,.0f}</div>', unsafe_allow_html=True)

        res_precio = st.number_input("Precio de compra (USD)", 1, 10_000_000,
                                      max(1, int(st.session_state.get("res_precio_k", max(int(_precio_ref / 10000) * 10000, 50000)))), 5_000, format="%d", key="res_precio_k")

        _ppm2 = res_precio / res_m2 if res_m2 > 0 else 0
        _ref_m2 = (_m_res.get("precio_2br", 0) if "2" in res_dormitorios else
                   (_m_res.get("precio_1br", 0) if "1" in res_dormitorios else _m_res.get("precio_3br", 0)))
        _diff_pct = ((_ppm2 - _ref_m2) / _ref_m2 * 100) if _ref_m2 > 0 else 0
        if abs(_diff_pct) <= 8:
            _pos_color, _pos_text = "#1A4731", f"En línea con mercado ({_diff_pct:+.1f}%)"
        elif _diff_pct > 8:
            _pos_color, _pos_text = "#7A1A1A", f"Sobre el mercado ({_diff_pct:+.1f}%) — negociar"
        else:
            _pos_color, _pos_text = "#1A4731", f"Por debajo del mercado ({_diff_pct:+.1f}%) — oportunidad"
        st.markdown(f'<div style="background:rgba(255,255,255,0.06);border-left:3px solid {_pos_color};border-radius:6px;'
                    f'padding:6px 12px;font-size:11px;color:{_pos_color};font-weight:600;margin-bottom:4px;">'
                    f'${_ppm2:,.0f}/m² — {_pos_text}</div>', unsafe_allow_html=True)

        # ── 4 · FINANCIAMIENTO ──────────────────────────────
        _step_header("4", "Financiamiento")
        res_pct_pie = st.number_input("Pago inicial / Down Payment (%)", 0.0, 100.0, 20.0, 1.0,
                                       key="res_pct_pie_inp")
        res_tasa = st.number_input("Tasa de interés anual (%)", 0.0, 30.0, 8.5, 0.25,
                                    key="res_tasa_inp")
        res_plazo = st.number_input("Plazo del crédito (años)", 1, 30, 20, 1,
                                    key="res_plazo_inp")

        # ── 5 · PROPÓSITO ───────────────────────────────────
        _step_header("5", "Propósito")
        res_uso = st.radio("¿Para qué?",
            ["Vivienda propia", "Inversión para alquilar", "Evaluación para venta"],
            key="res_uso_radio")

        if res_uso in ["Inversión para alquilar", "Evaluación para venta"]:
            _alq_sugerido = round(_m_res.get("alquiler_m2_mes", 0) * res_m2 / 50) * 50
            if res_uso == "Inversión para alquilar":
                st.markdown(f'<div style="font-size:11px;color:#B8904A;font-weight:600;margin-top:-4px;">Mercado: ${_alq_sugerido:,}/mes · ${_m_res.get("alquiler_m2_mes", 0):.1f}/m²/mes</div>', unsafe_allow_html=True)
                res_alquiler = st.number_input("Renta mensual (USD)", 0, 20_000,
                                                int(st.session_state.get("res_alq_k", int(_alq_sugerido))), 50, key="res_alq_inp")
                res_tipo_contrato = st.radio(
                    "Tipo de contrato", ["Anual", "Plurianual (3+ años)"],
                    horizontal=True, key="res_tipo_contrato")
                res_ajuste_pct    = 0.0
                res_inicio_ajuste = 2
                if res_tipo_contrato == "Plurianual (3+ años)":
                    _rca1, _rca2 = st.columns(2)
                    res_ajuste_pct    = _rca1.number_input(
                        "Ajuste anual (%)", 0.0, 10.0, 3.0, 0.5, key="res_ajuste_pct")
                    res_inicio_ajuste = _rca2.selectbox(
                        "Año de inicio del ajuste", [2, 3], key="res_inicio_ajuste")
            else:
                res_alquiler      = 0
                res_tipo_contrato = "Anual"
                res_ajuste_pct    = 0.0
                res_inicio_ajuste = 2
            res_gastos = st.number_input("Gastos mensuales (USD)", 0, 5_000,
                                          max(int(res_precio * 0.004 / 12), 50), 10,
                                          key="res_gastos_inp")
        else:
            res_alquiler      = 0
            res_gastos        = 0
            res_tipo_contrato = "Anual"
            res_ajuste_pct    = 0.0
            res_inicio_ajuste = 2

        # ── 6 · COMPARATIVA ─────────────────────────────────
        _step_header("6", "Comparativa de Inmuebles")
        st.caption("Agrega hasta 3 inmuebles alternativos para comparar")
        if "res_inmuebles_comp" not in st.session_state:
            st.session_state.res_inmuebles_comp = []

        with st.expander("Agregar inmueble a comparar"):
            _rc_nombre = st.text_input("Nombre / referencia", placeholder="Ej: Depto Av. Larco 320", key="rc_nombre_inp")
            _rc_col1, _rc_col2 = st.columns(2)
            _rc_precio = _rc_col1.number_input("Precio USD", 0, 10_000_000, 200_000, 5_000, format="%d", key="rc_precio_inp")
            _rc_m2 = _rc_col2.number_input("m²", 10, 2000, 80, 5, key="rc_m2")
            _rc_col3, _rc_col4 = st.columns(2)
            _rc_alq = _rc_col3.number_input("Alquiler/mes USD", 0, 20_000, 800, 50, key="rc_alq")
            _rc_dorm = _rc_col4.selectbox("Dorm.", ["1D", "2D", "3D"], key="rc_dorm")
            _rc_zona = st.selectbox("Distrito", list(MERCADO.keys()), key="rc_zona_sel")
            _rc_img = st.file_uploader("Foto del inmueble (opcional)", type=["jpg", "jpeg", "png", "webp"], key="rc_img_upload")
            if st.button("AGREGAR", use_container_width=True, key="btn_rc_agregar", type="primary"):
                if _rc_nombre.strip():
                    _rc_img_bytes = _rc_img.read() if _rc_img else None
                    st.session_state.res_inmuebles_comp.append({
                        "nombre": _rc_nombre.strip(),
                        "precio": _rc_precio, "m2": _rc_m2,
                        "alquiler": _rc_alq, "dormitorios": _rc_dorm,
                        "zona": _rc_zona,
                        "precio_m2": round(_rc_precio / _rc_m2) if _rc_m2 > 0 else 0,
                        "yield_bruto": round(_rc_alq * 12 / _rc_precio * 100, 1) if _rc_precio > 0 and _rc_alq > 0 else 0,
                        "imagen_bytes": _rc_img_bytes,
                        "imagen_nombre": _rc_img.name if _rc_img else None,
                    })
                    st.rerun()

        if st.session_state.res_inmuebles_comp:
            for _i, _ci in enumerate(st.session_state.res_inmuebles_comp):
                _cc1, _cc2 = st.columns([5, 1])
                _cc1.caption(f"**{_ci['nombre']}** — ${_ci['precio']:,} · ${_ci['precio_m2']:,}/m²")
                if _cc2.button("✕", key=f"del_ri_{_i}"):
                    st.session_state.res_inmuebles_comp.pop(_i)
                    st.rerun()
            if st.button("Limpiar comparativa", use_container_width=True, key="btn_clear_ri"):
                st.session_state.res_inmuebles_comp = []
                st.rerun()

        # ── 7 · DOCUMENTOS ──────────────────────────────────
        _step_header("7", "Documentos")
        st.caption("Sube los documentos del inmueble para análisis registral y técnico")
        _res_secrets_key = (st.secrets.get("anthropic", {}) or {}).get("api_key", "")
        if not _res_secrets_key and not st.session_state.get("api_key_input"):
            with st.expander("⚙ Configuración API", expanded=True):
                _res_api_k = st.text_input("Clave de acceso", type="password",
                                           placeholder="••••••••••••••••", key="res_api_key_inp",
                                           value=st.session_state.get("api_key_input", ""))
                if _res_api_k:
                    st.session_state["api_key_input"] = _sanitize_api_key(_res_api_k)
        if _res_secrets_key or st.session_state.get("api_key_input"):
            st.markdown(
                '<div style="font-size:10px;color:#1A4731;background:#E8F5EE;'
                'border-radius:4px;padding:5px 10px;margin-bottom:6px;'
                'border-left:2px solid #1A4731;">● Sistema activo</div>',
                unsafe_allow_html=True)
        res_doc_partida = st.file_uploader("Partida Registral (SUNARP)", type="pdf", key="res_doc_partida")
        res_doc_puhr    = st.file_uploader("PU / HR",                    type="pdf", key="res_doc_puhr")
        res_doc_params  = st.file_uploader("Certificado de Parámetros",  type="pdf", key="res_doc_params")
        res_doc_planos  = st.file_uploader("Planos del Inmueble",        type="pdf", key="res_doc_planos")
        _res_has_docs = any([res_doc_partida, res_doc_puhr, res_doc_params, res_doc_planos])
        if _res_has_docs:
            run_res_docs = st.button("ANALIZAR DOCUMENTOS", use_container_width=True, key="btn_res_docs")
        else:
            run_res_docs = False
            st.caption("Adjunta al menos un documento para habilitar el análisis.")

        # ── 8 · FOTOS ───────────────────────────────────────
        _step_header("8", "Fotos del Inmueble")
        st.caption("Las fotos se incluyen en el informe")
        _res_fotos = st.file_uploader(
            "Sube fotos del inmueble",
            type=["jpg", "jpeg", "png", "webp"],
            accept_multiple_files=True,
            key="res_fotos_upload",
        )
        if _res_fotos:
            st.session_state["res_fotos_bytes"] = [f.read() for f in _res_fotos]
            st.session_state["res_fotos_nombres"] = [f.name for f in _res_fotos]
            cols_rf = st.columns(min(len(_res_fotos), 3))
            for _rfi, _rfup in enumerate(_res_fotos[:3]):
                cols_rf[_rfi].image(_rfup, use_container_width=True)

        st.markdown("---")
        st.markdown("### INSTRUCCIONES AL ANÁLISIS")
        st.caption("Lo que escribas aquí guía a la IA — uso del inmueble, condiciones especiales, perfil del comprador.")
        res_sugerencias = st.text_area(
            label="instrucciones_res",
            placeholder="Ej: cliente busca inversión para alquiler, priorizar zonas con alta plusvalía, presupuesto máximo $250K...",
            height=90,
            label_visibility="collapsed",
            key="res_sugerencias_inp"
        )

        # ── EJECUTAR ─────────────────────────────────────────
        st.markdown("---")
        run_residencial = st.button("EJECUTAR ANÁLISIS", use_container_width=True, type="primary")

        # ── PROYECTOS GUARDADOS ──────────────────────────────
        with st.expander("💾 Proyectos guardados"):
            _res_saved = listar_proyectos()
            _res_nombres = [p.name for p in _res_saved]
            if _res_nombres:
                _res_sel = st.selectbox("Cargar proyecto", ["— seleccionar —"] + _res_nombres,
                                        label_visibility="collapsed", key="res_sel_proy")
                if _res_sel != "— seleccionar —":
                    if st.button("CARGAR", use_container_width=True, key="btn_res_cargar"):
                        _res_pref = next((p for p in _res_saved if p.name == _res_sel), None)
                        _d = cargar_proyecto(_res_pref or _res_sel)
                        if _d.get("residencial_result"):
                            st.session_state.residencial_result = _d["residencial_result"]
                            st.rerun()
            _res_nombre_proy = st.text_input("Nombre del proyecto", placeholder="Ej: Depto Miraflores",
                                              label_visibility="collapsed", key="res_nombre_proy")
            if st.button("GUARDAR PROYECTO", use_container_width=True, key="btn_res_guardar"):
                if st.session_state.get("residencial_result"):
                    fp = guardar_proyecto(_res_nombre_proy or "residencial_sin_nombre",
                                          {"residencial_result": st.session_state.residencial_result})
                    st.session_state["_last_res_proy"] = fp
                    st.session_state.pop("_share_url_res", None)
                    st.markdown(f'<div style="background:rgba(107,206,160,0.12);border-left:3px solid rgba(107,206,160,0.50);'
                                f'border-radius:6px;padding:8px 12px;color:rgba(107,206,160,0.90);font-size:12px;font-weight:600;margin-top:6px;">'
                                f'✓ {fp.name}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="alert-info">ℹ️ ' + "Ejecuta el análisis primero." + '</div>', unsafe_allow_html=True)
            _lp_res = st.session_state.get("_last_res_proy")
            if _lp_res and _lp_res._id:
                if st.button("GENERAR LINK DE COMPARTIR", use_container_width=True, key="btn_share_res"):
                    _tok = str(uuid.uuid4())
                    _sb2 = _get_supabase()
                    if _sb2:
                        _sb2.table("proyectos").update({"share_token": _tok}).eq("id", _lp_res._id).execute()
                        _base = (st.secrets.get("app", {}) or {}).get("base_url", "http://localhost:8501")
                        st.session_state["_share_url_res"] = f"{_base}?share={_tok}"
                if st.session_state.get("_share_url_res"):
                    st.code(st.session_state["_share_url_res"])
                    st.caption("Comparte con tu cliente — no requiere contraseña")

        # ── RETROALIMENTACIÓN ─────────────────────────────
        if st.session_state.get("residencial_result"):
            st.markdown("---")
            st.markdown(
                '<div style="font-size:9px;color:rgba(184,144,74,0.90);letter-spacing:2px;'
                'text-transform:uppercase;font-weight:700;margin-bottom:8px;">Retroalimentar a Solum</div>'
                '<div style="font-size:10px;color:rgba(200,216,232,0.75);line-height:1.6;margin-bottom:10px;">'
                '¿Ya consolidaste este análisis con tu criterio profesional? Compártelo con Solum — '
                'así aprendo tu forma de trabajar y el próximo análisis saldrá más alineado a tu visión desde el inicio.'
                '</div>', unsafe_allow_html=True)
            _fb_file_res = st.file_uploader(
                "Excel, PDF, DWG o imagen",
                type=["xlsx", "xls", "pdf", "dwg", "dxf", "jpg", "jpeg", "png"],
                key="fb_upload_res",
                label_visibility="collapsed",
            )
            if _fb_file_res:
                _fb_bytes_res = _fb_file_res.read()
                _fb_proy_res = st.session_state.get("_last_res_proy")
                _fb_id_res = getattr(_fb_proy_res, "_id", None) if _fb_proy_res else None
                _sb_fb_res = _get_supabase()
                if _sb_fb_res and _fb_id_res:
                    try:
                        _sb_fb_res.table("proyectos").update({
                            "feedback_filename": _fb_file_res.name,
                            "feedback_size":     len(_fb_bytes_res),
                            "feedback_at":       datetime.datetime.utcnow().isoformat(),
                        }).eq("id", _fb_id_res).execute()
                    except Exception:
                        pass
                st.markdown(
                    '<div style="background:rgba(107,206,160,0.10);border-left:3px solid '
                    'rgba(107,206,160,0.50);border-radius:6px;padding:8px 12px;'
                    'font-size:11px;color:rgba(107,206,160,0.90);font-weight:600;margin-top:6px;">'
                    f'✓ Recibido: {_fb_file_res.name}</div>', unsafe_allow_html=True)

# ── SESSION STATE ────────────────────────────────────

for k in ("params", "cabida", "financ", "zona", "legal",
          "partida_bytes", "puhr_bytes", "cert_bytes",
          "industrial_result", "residencial_result",
          "industrial_factibilidad", "residencial_legal", "ind_resumen", "res_resumen"):
    if k not in st.session_state:
        st.session_state[k] = None
if "financ_inputs" not in st.session_state:
    st.session_state.financ_inputs = {}
if "ind_comparativa" not in st.session_state:
    st.session_state.ind_comparativa = []
if "res_comparativa" not in st.session_state:
    st.session_state.res_comparativa = []
if "comps_sunarp" not in st.session_state:
    st.session_state.comps_sunarp = []

tipo_op = st.session_state.get("tipo_operacion", "Proyecto Inmobiliario")

# ── EJECUCIÓN ────────────────────────────────────────

if tipo_op == "Proyecto Logístico / Industrial" and run_industrial:
    _ind_inp = {
        "area_terreno":        ind_area,
        "costo_terreno":       ind_costo_terreno,
        "tipo_nave":           ind_tipo,
        "zonificacion":        ind_zona_ind,
        "pct_techada":         ind_pct_techada,
        "costo_nave_m2":       ind_costo_nave,
        "costo_piso_libre_m2": ind_costo_piso,
        "pct_indirectos":      ind_pct_indirectos,
        "include_alcabala":    ind_alcabala,
        "renta_m2_mes":        ind_renta,
        "uso":                 ind_uso,
        "tipo_contrato":       ind_tipo_contrato,
        "ajuste_anual_pct":    ind_ajuste_pct,
        "inicio_ajuste_ano":   ind_inicio_ajuste,
        "actividad_categoria": st.session_state.get("ind_act_categoria", ""),
        "actividad_descripcion": st.session_state.get("ind_actividad_desc", ""),
        # Dual credit
        "dp_terreno_pct":   st.session_state.get("ind_dp_terreno", 40.0),
        "tasa_terreno":     st.session_state.get("ind_tasa_terreno", 8.0),
        "plazo_terreno":    int(st.session_state.get("ind_plazo_terreno", 10)),
        "dp_const_pct":     st.session_state.get("ind_dp_const", 30.0),
        "tasa_const":       st.session_state.get("ind_tasa_const", 9.0),
        "plazo_const":      int(st.session_state.get("ind_plazo_const", 8)),
    }
    st.session_state.industrial_result = calcular_industrial(_ind_inp)
    st.session_state.ind_analizado = True
    # Si hay documentos cargados, dispara el análisis de factibilidad en el mismo flujo
    if _ind_has_docs:
        run_ind_docs = True

if tipo_op == "Proyecto Logístico / Industrial" and run_ind_docs:
    _ip  = ind_doc_partida.read() if ind_doc_partida else None
    _ic  = ind_doc_params.read()  if ind_doc_params  else None
    _iz  = ind_doc_zon.read()     if ind_doc_zon      else None
    _ipl = (ind_doc_planos.read()
            if ind_doc_planos and not ind_doc_planos.name.lower().endswith((".dwg", ".dxf"))
            else None)
    _it  = st.session_state.get("ind_tipo", "Almacén Logístico")
    _iz2 = st.session_state.get("ind_zona_ind", "I2")
    _iu  = st.session_state.get("ind_uso", "Uso directo")
    _isug = st.session_state.get("ind_sugerencias_inp", "")
    st.session_state.industrial_factibilidad = _run_with_retry(
        lambda _ip=_ip, _ic=_ic, _iz=_iz, _it=_it, _iz2=_iz2, _iu=_iu, _ipl=_ipl, _isug=_isug: analizar_factibilidad_industrial(_ip, _ic, _iz, _it, _iz2, _iu, _ipl, _isug),
        "Analizando factibilidad técnica y documentos registrales…",
    )
    st.rerun()

if tipo_op == "Inmueble Residencial" and run_res_docs:
    _rp = res_doc_partida.read() if res_doc_partida else None
    _ru = res_doc_puhr.read()    if res_doc_puhr    else None
    _rc = res_doc_params.read() if res_doc_params else None
    _rl = res_doc_planos.read() if res_doc_planos  else None
    _rsug = st.session_state.get("res_sugerencias_inp", "")
    st.session_state.residencial_legal = _run_with_retry(
        lambda _p=_rp, _u=_ru, _c=_rc, _l=_rl, _rsug=_rsug: analizar_legal(_p, _u, _c, _l, _rsug),
        "Analizando documentos legales…",
    )
    st.rerun()

elif tipo_op == "Inmueble Residencial" and run_residencial:
    _m_res_run = MERCADO.get(res_zona, {})
    _res_precio_m2_mercado = (_m_res_run.get("precio_2br", 0) if "2" in res_dormitorios
                               else _m_res_run.get("precio_1br", 0) if "1" in res_dormitorios
                               else _m_res_run.get("precio_3br", 0))
    st.session_state.residencial_result = calcular_residencial({
        "precio":     res_precio,
        "pct_pie":    res_pct_pie,
        "tasa_anual": res_tasa,
        "plazo_anos": res_plazo,
        "uso":        res_uso,
        "alquiler_mes":      res_alquiler,
        "gastos_mes":        res_gastos,
        "tipo_contrato":     res_tipo_contrato,
        "ajuste_anual_pct":  res_ajuste_pct,
        "inicio_ajuste_ano": res_inicio_ajuste,
        "zona":       res_zona,
        "dormitorios": res_dormitorios,
        "m2":         res_m2,
        "antiguedad": res_antiguedad,
        "precio_m2":  res_precio / res_m2 if res_m2 > 0 else 0,
        "precio_m2_mercado": _res_precio_m2_mercado,
        "yield_mercado_pct": _m_res_run.get("yield_mercado_pct", 0),
        "alquiler_mercado_m2": _m_res_run.get("alquiler_m2_mes", 0),
        "variacion_anual_pct": _m_res_run.get("variacion_anual_pct", 0),
    })
    st.session_state.res_analizado = True
    st.session_state["_res_zona_val"]      = res_zona
    st.session_state["_res_m2_val"]        = res_m2
    st.session_state["_res_antiguedad_val"] = res_antiguedad

elif run:
    if not pdf_cert:
        st.markdown('<div class="alert-info">ℹ️ ' + "⚠️ Adjunta el Certificado de Parámetros para continuar." + '</div>', unsafe_allow_html=True)
        st.stop()

    # Resetear restricción geométrica — solo aplica si el usuario corre el módulo de geometría
    # con las medidas reales del lote en esta misma sesión
    st.session_state["geo_at_max"] = 0

    # Recopilar todos los documentos adicionales
    extra_docs = []
    if pdf_plano:
        for f in (pdf_plano if isinstance(pdf_plano, list) else [pdf_plano]):
            extra_docs.append(f.read())
    if pdf_partida:
        _partida_bytes = pdf_partida.read()
        extra_docs.append(_partida_bytes)
        st.session_state.partida_bytes = _partida_bytes
    if pdf_puhr:
        _puhr_bytes = pdf_puhr.read()
        extra_docs.append(_puhr_bytes)
        st.session_state.puhr_bytes = _puhr_bytes
    if pdf_norms:
        for f in pdf_norms:
            extra_docs.append(f.read())

    _cert_bytes = pdf_cert.read()
    st.session_state.cert_bytes = _cert_bytes
    st.session_state.params = _run_with_retry(
        lambda _cb=_cert_bytes, _ed=list(extra_docs): extract_parameters(_cb, _ed),
        "Extrayendo información del documento…",
    )

    # Aplicar overrides manuales sobre los parámetros extraídos
    if override_area > 0:
        st.session_state.params["area_terreno_m2"] = override_area
    if override_al > 0:
        st.session_state.params["area_libre_min_pct"] = override_al

    config = {
        "sugerencias": sugerencias or "",
        "colindante_izq_pisos": colind_izq if colind_izq > 0 else None,
        "colindante_der_pisos": colind_der if colind_der > 0 else None,
    }

    _params_snap = st.session_state.params
    _config_snap = config
    st.session_state.cabida = _run_with_retry(
        lambda _ps=dict(_params_snap), _cs=dict(_config_snap): generate_cabida(_ps, _cs),
        "Analizando data y elaborando cabida…",
    )

    # Aplicar regla de colindancia al params real (generate_cabida modifica solo la copia local)
    if (colind_izq or 0) > 0 or (colind_der or 0) > 0:
        _base_p  = int(st.session_state.params.get("pisos_max") or 5)
        _max_col = max(colind_izq or 0, colind_der or 0)
        _pisos_col = int((_max_col + _base_p) / 2)  # Art.6.3 Ord.523-MSI: floor — decimal no sube al entero superior
        st.session_state.params["pisos_max"] = _pisos_col
        # Inyectar nota normativa con base legal y fórmula
        _dist_col = str(st.session_state.params.get("distrito", "")).lower()
        if "san isidro" in _dist_col:
            _base_legal_col = "Ord. 523-MSI Art.6.3 — Reglamento de Parámetros Urbanísticos y Edificatorios de San Isidro, Norma de Colindancia"
        else:
            _base_legal_col = "PDU del distrito / CPU — Regla de Colindancia (el proyecto puede alcanzar el promedio con la edificación colindante más alta)"
        _nota_colind = (
            f"Regla de Colindancia aplicada: colindante más alto = {_max_col} pisos | "
            f"altura base del certificado = {_base_p} pisos → "
            f"pisos permitidos = floor(({_max_col} + {_base_p}) ÷ 2) = {_pisos_col} pisos. "
            f"Base legal: {_base_legal_col}."
        )
        _notas_prev = [n for n in st.session_state.params.get("notas_altura", []) if "Colindancia" not in str(n)]
        _notas_prev.append(_nota_colind)
        st.session_state.params["notas_altura"] = _notas_prev
    # Sincronizar widget de pisos del módulo geo
    st.session_state["geo_n_pisos"] = int(st.session_state.params.get("pisos_max") or 8)

    m = MERCADO.get(zona, {})
    st.session_state.financ_inputs = {
        "costo_terreno":      precio_compra,
        "costo_construccion": costo_const_m2,
        "costo_sotano_m2":    costo_sotano_m2,
        "fee_constructora":   fee_constructora,
        "tasa_ir":            tasa_ir,
        "include_alcabala":   include_alcabala,
        "include_dd":         include_dd,
        "costo_demolicion":   costo_demolicion,
        "precio_venta_m2":    precio_venta_m2,
        "precio_1br":         m.get("precio_1br", 0),
        "precio_2br":         m.get("precio_2br", 0),
        "precio_3br":         m.get("precio_3br", 0),
        "precio_estac":       precio_estac_inp if precio_estac_inp > 0 else m.get("precio_estac", 0),
        "precio_deposito":    precio_deposito_inp if precio_deposito_inp > 0 else m.get("precio_deposito", 0),
        "costo_arq_m2":       costo_arq_m2,
        "costo_esp_m2":       costo_esp_m2,
        "costo_factibilidades": costo_factibilidades,
        "tasa_financ":        9.0,
        "estructura_financ":  estructura_financ,
        "aporte_propio_pct":  aporte_propio_pct,
        "pct_preventa_banco": pct_preventa_banco,
        "meses_preventa_override": meses_preventa_override,
        "meses_obra_override": meses_obra_override,
        "pct_mktg_preventa":  pct_mktg_preventa,
    }
    st.session_state.zona = zona

# ── TABS ─────────────────────────────────────────────

if tipo_op == "Proyecto Inmobiliario":
    if st.session_state.params:
        p        = st.session_state.params
        c        = st.session_state.cabida
        zona_sel = st.session_state.zona or zona

        # Persistir tab activa en localStorage para sobrevivir reruns de Streamlit
        st.components.v1.html("""<script>
        (function() {
            var STORE_KEY = 'solum_proj_tab';
            function getTabs() {
                return Array.from(window.parent.document.querySelectorAll('[data-baseweb="tab-list"] [role="tab"]'));
            }
            function restoreTab(attempt) {
                var saved = localStorage.getItem(STORE_KEY);
                if (!saved) return;
                var tabs = getTabs();
                if (!tabs.length && attempt < 8) { setTimeout(function(){ restoreTab(attempt+1); }, 200); return; }
                var target = tabs.find(function(t){ return t.textContent.trim() === saved; });
                if (target && target.getAttribute('aria-selected') !== 'true') { target.click(); }
            }
            // Guardar tab activa al hacer click
            function attachListeners(attempt) {
                var tabs = getTabs();
                if (!tabs.length && attempt < 8) { setTimeout(function(){ attachListeners(attempt+1); }, 200); return; }
                tabs.forEach(function(t) {
                    t.addEventListener('click', function() { localStorage.setItem(STORE_KEY, t.textContent.trim()); });
                });
            }
            setTimeout(function(){ restoreTab(0); attachListeners(0); }, 300);
        })();
        </script>""", height=0)

        if st.session_state.get("_goto_tab_name"):
            _tab_name = st.session_state.pop("_goto_tab_name")
            st.components.v1.html(f"""<script>
            localStorage.setItem('solum_proj_tab', {repr(_tab_name)});
            (function click(n, a) {{
                var tabs = Array.from(window.parent.document.querySelectorAll('[data-baseweb="tab-list"] [role="tab"]'));
                var t = tabs.find(function(x){{ return x.textContent.trim() === n; }});
                if (t) {{ t.click(); }} else if (a < 6) {{ setTimeout(function(){{ click(n, a+1); }}, 250); }}
            }})({repr(_tab_name)}, 0);
            </script>""", height=0)

        _cab_ubicacion = p.get("ubicacion") or p.get("direccion") or zona_sel or "—"
        _cab_zona_txt  = p.get("zonificacion", "—")
        _cab_pisos     = p.get("pisos_max", "—")

        # ── Pre-compute financiero eagerly so all tabs (Gantt, Sensibilidad) have data ──
        if c and not st.session_state.financ:
            _fi_e = st.session_state.financ_inputs or {}
            _md_e = MERCADO.get(zona_sel, {})
            _fin_eager = {
                "costo_terreno":      _fi_e.get("costo_terreno", 0),
                "costo_construccion": _fi_e.get("costo_construccion", _md_e.get("costo_construccion", 850)),
                "costo_sotano_m2":    _fi_e.get("costo_sotano_m2", 450),
                "fee_constructora":   _fi_e.get("fee_constructora", 10.0),
                "tasa_ir":            _fi_e.get("tasa_ir", 29.5),
                "include_alcabala":   _fi_e.get("include_alcabala", True),
                "include_dd":         _fi_e.get("include_dd", True),
                "precio_venta_m2":    _fi_e.get("precio_venta_m2", _md_e.get("precio_2br", 0)),
                "precio_1br":         _md_e.get("precio_1br", 0),
                "precio_2br":         _md_e.get("precio_2br", 0),
                "precio_3br":         _md_e.get("precio_3br", 0),
                "precio_estac":       _md_e.get("precio_estac", 0),
                "precio_deposito":    _md_e.get("precio_deposito", 0),
                "tasa_financ":        _fi_e.get("tasa_financ", 9.0),
                "estructura_financ":  _fi_e.get("estructura_financ", "estandar"),
                "aporte_propio_pct":  _fi_e.get("aporte_propio_pct", 100.0),
                "pct_preventa_banco": _fi_e.get("pct_preventa_banco", 30.0),
                "meses_preventa_override": _fi_e.get("meses_preventa_override"),
                "meses_obra_override": _fi_e.get("meses_obra_override"),
                "pct_mktg_preventa":  _fi_e.get("pct_mktg_preventa", 2.0),
                "nombre_proyecto":    st.session_state.get("nombre_proyecto", ""),
            }
            try:
                st.session_state.financ = calcular_financiero(c, _fin_eager, zona_sel)
            except Exception:
                pass

        # ── Project Hero Banner ───────────────────────────────────────
        import base64 as _b64c
        _cab_hero_fotos = st.session_state.get("cab_fotos_bytes") or []
        if _cab_hero_fotos:
            try:
                _b64_cab = _b64c.b64encode(_cab_hero_fotos[0]).decode()
                _cab_photo_css = f"url('data:image/jpeg;base64,{_b64_cab}') center/cover no-repeat"
            except Exception:
                _cab_photo_css = "linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%)"
        else:
            _cab_photo_css = "linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%)"

        _fr = st.session_state.financ.get("resumen") if st.session_state.financ else None
        _cab_kpi1_label, _cab_kpi1_val = ("Terreno", f"{p.get('area_terreno_m2','—')} m²")
        _cab_kpi2_label, _cab_kpi2_val = ("Zonificación", _cab_zona_txt)
        _cab_kpi3_label, _cab_kpi3_val = ("Altura máx.", f"{_cab_pisos} pisos")
        _cab_kpi4_tuple = ("Margen neto", f"{_fr['margen_pct']:.1f}%") if _fr else ("Área libre mín.", f"{p.get('area_libre_min_pct','—')}%")
        _cab_kpi4_label, _cab_kpi4_val = _cab_kpi4_tuple
        _cab_fin_extra = (
            f'<div><div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">TIR Anual</div>'
            f'<div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_fr["tir_anual_pct"]:.1f}%</div></div>'
            if _fr else ""
        )
        _cab_nombre = st.session_state.get("nombre_proyecto") or _cab_ubicacion

        st.markdown(f"""
        <div style="position:relative;border-radius:16px;overflow:hidden;margin-bottom:20px;
                    box-shadow:0 6px 30px rgba(30,45,61,0.22);">
            <div style="background:{_cab_photo_css};height:220px;"></div>
            <div style="position:absolute;inset:0;background:linear-gradient(to bottom,
                        rgba(0,0,0,0.0) 0%,rgba(0,0,0,0.75) 100%);
                        display:flex;flex-direction:column;justify-content:flex-end;
                        padding:24px 28px;">
                <div style="font-size:9px;color:rgba(255,255,255,0.60);letter-spacing:3px;
                            text-transform:uppercase;margin-bottom:6px;">
                    Análisis de Proyecto Inmobiliario · FACTIS
                </div>
                <div style="font-size:28px;font-weight:800;color:#FFFFFF;line-height:1.15;
                            text-shadow:0 2px 8px rgba(0,0,0,0.5);">
                    {_cab_nombre}
                </div>
                <div style="display:flex;gap:20px;margin-top:12px;flex-wrap:wrap;">
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">{_cab_kpi1_label}</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_cab_kpi1_val}</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">{_cab_kpi2_label}</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_cab_kpi2_val}</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">{_cab_kpi3_label}</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_cab_kpi3_val}</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">{_cab_kpi4_label}</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_cab_kpi4_val}</div>
                    </div>
                    {_cab_fin_extra}
                </div>
            </div>
        </div>""", unsafe_allow_html=True)

        tabs = st.tabs(["Parámetros", "Cabida", "Financiero", "Flujo de Caja", "Legal", "Resumen", "Propuesta", "Renta / Holding"], key="proj_tabs")

        # ── TAB 1: PARÁMETROS ────────────────────────────
        with tabs[0]:
            st.markdown('<div class="section-title">Datos Extraídos del Certificado</div>', unsafe_allow_html=True)

            c1, c2, c3 = st.columns(3)
            with c1:
                card("Zonificación",    p.get("zonificacion",    "—"))
                card("Área del terreno", f"{p.get('area_terreno_m2','—')} m²")
                card("Frente",          f"{p.get('frente_ml','—')} ml")
            with c2:
                card("Altura máxima",    f"{p.get('pisos_max','—')} pisos")
                card("Área libre mínima", f"{p.get('area_libre_min_pct','—')}%")
                card("Retiro frontal",   f"{p.get('retiro_frontal_ml','—')} ml")
            with c3:
                card("Retiro lateral",   f"{p.get('retiro_lateral_ml','—') or '—'} ml")
                card("Retiro posterior",  f"{p.get('retiro_posterior_ml','—') or '—'} ml")
                card("Caduca",           p.get("fecha_caducidad","—"), color="#8B3A2A")

            if p.get("pisos_por_via"):
                st.markdown('<div class="section-title">Alturas por Vía</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame(p["pisos_por_via"]), use_container_width=True, hide_index=True)

            if p.get("notas_altura"):
                st.markdown('<div class="section-title">Notas Normativas de Altura</div>', unsafe_allow_html=True)
                for nota in p["notas_altura"]:
                    st.markdown(f'<div class="alert-gold">📌 {nota}</div>', unsafe_allow_html=True)

            # ── BENEFICIOS NORMATIVOS ─────────────────────
            beneficios = p.get("beneficios_normativos", [])
            if beneficios:
                st.markdown('<div class="section-title">Beneficios Normativos Identificados</div>', unsafe_allow_html=True)

                tipo_icons = {
                    "mayor_altura":             "🔺",
                    "reduccion_aportes":        "📉",
                    "flexibilizacion_retiros":  "↔️",
                    "incentivo_densificacion":  "🏙️",
                    "otro":                     "⚖️",
                }
                for b in beneficios:
                    icon = tipo_icons.get(b.get("tipo", "otro"), "⚖️")
                    st.markdown(f"""
                    <div class="alert-legal">
                        <strong>{icon} {b.get('descripcion','')}</strong><br>
                        <span style="font-size:12px;color:#4A6080">
                            <em>Condición:</em> {b.get('condicion','—')} &nbsp;|&nbsp;
                            <em>Impacto:</em> {b.get('impacto_estimado','—')} &nbsp;|&nbsp;
                            <em>Base legal:</em> {b.get('base_legal','—')}
                        </span>
                    </div>""", unsafe_allow_html=True)
            else:
                st.markdown('<div class="section-title">Base Legal</div>', unsafe_allow_html=True)

            if p.get("ordenanzas_base"):
                for ord_ in p["ordenanzas_base"]:
                    st.markdown(f"• {ord_}")

        # ── TAB 2: CABIDA ────────────────────────────────
        with tabs[1]:
            # ── GEOMETRÍA DEL LOTE ────────────────────────────────
            st.markdown('<div class="section-title">Geometría del Lote</div>', unsafe_allow_html=True)

            # Pre-poblar medidas del lote desde params la primera vez que se carga
            _p_params_geo = p or {}
            _p_frente_geo = float(_p_params_geo.get("frente_ml") or 0)
            _p_area_geo   = float(_p_params_geo.get("area_terreno_m2") or 0)
            if "geo_frente" not in st.session_state and _p_frente_geo > 0:
                st.session_state["geo_frente"] = _p_frente_geo
                _p_fondo_explicit = float(_p_params_geo.get("fondo_ml") or 0)
                if _p_fondo_explicit > 0:
                    _p_fondo_geo = round(_p_fondo_explicit, 1)
                else:
                    # Sin fondo explícito: asumir fondo = frente (lote regular).
                    # Usuario debe corregir subiendo plano o tabulando medidas.
                    _p_fondo_geo = round(_p_frente_geo, 1)
                # Lados: área ÷ frente (profundidad real). Si no hay área, usar fondo como fallback.
                _p_lado_geo = round(_p_area_geo / _p_frente_geo, 1) if _p_area_geo > 0 else _p_fondo_geo
                st.session_state["geo_fondo"] = _p_fondo_geo
                st.session_state["geo_izq"]   = _p_lado_geo
                st.session_state["geo_der"]   = _p_lado_geo

            _geo_modo = st.radio(
                "Fuente de medidas",
                ["Tabular medidas", "Adjuntar plano (DXF/DWG)"],
                horizontal=True, key="geo_modo"
            )

            _poly_lote = st.session_state.get("geo_poly_lote")

            if _geo_modo == "Tabular medidas":
                _gc1, _gc2, _gc3, _gc4 = st.columns(4)
                _g_frente  = _gc1.number_input("Frente (ml)",        min_value=1.0, value=float(st.session_state.get("geo_frente", 15.0)),  step=0.5, key="geo_frente")
                _g_fondo   = _gc2.number_input("Fondo (ml)",         min_value=1.0, value=float(st.session_state.get("geo_fondo",  20.0)),  step=0.5, key="geo_fondo")
                _g_izq     = _gc3.number_input("Lado izquierdo (ml)", min_value=1.0, value=float(st.session_state.get("geo_izq",   20.0)),  step=0.5, key="geo_izq")
                _g_der     = _gc4.number_input("Lado derecho (ml)",   min_value=1.0, value=float(st.session_state.get("geo_der",   20.0)),  step=0.5, key="geo_der")
                _g_esquina = st.checkbox("Lote en esquina", value=st.session_state.get("geo_esquina", False), key="geo_esquina")

                if st.button("Calcular geometría", key="geo_calc_btn", type="primary"):
                    if _SHAPELY_OK:
                        _poly_lote = _geo_poligono_tabular(_g_frente, _g_fondo, _g_izq, _g_der)
                        st.session_state["geo_poly_lote"]  = _poly_lote
                        st.session_state["geo_frente_val"] = _g_frente
                    else:
                        st.markdown('<div class="alert-info">ℹ️ ' + "Librería shapely no disponible. Ejecuta: pip install shapely" + '</div>', unsafe_allow_html=True)

            else:
                _geo_file = st.file_uploader("Cargar plano perimétrico (.dxf)", type=["dxf"], key="geo_dxf_file")
                if _geo_file:
                    if _SHAPELY_OK and _EZDXF_OK:
                        import io
                        _poly_lote = _geo_poligono_dxf(io.TextIOWrapper(io.BytesIO(_geo_file.read()), encoding="utf-8", errors="ignore"))
                        if _poly_lote:
                            st.session_state["geo_poly_lote"] = _poly_lote
                            st.session_state["geo_frente_val"] = _g_frente if "geo_frente" in st.session_state else 0.0
                            st.success(f"Polígono extraído — área: {_poly_lote.area:,.1f} m²")
                        else:
                            st.error("No se encontró perímetro en el DXF. Verifica que el archivo contenga una polilínea cerrada.")
                    else:
                        st.markdown('<div class="alert-info">ℹ️ ' + "Instala ezdxf y shapely para usar esta función." + '</div>', unsafe_allow_html=True)

            # Retiros y Esquema de Alturas
            if _poly_lote and not _poly_lote.is_empty:
                st.markdown('<div class="section-title">Retiros y Esquema de Alturas</div>', unsafe_allow_html=True)
                _rg1, _rg2, _rg3, _rg4, _rg5 = st.columns(5)

                # Pre-llenar desde el certificado (params) si están disponibles
                _p_now = st.session_state.get("params") or {}
                _ret_f_def  = float(_p_now.get("retiro_frontal_ml")   or 3.0)
                _ret_l_def  = float(_p_now.get("retiro_lateral_ml")   or 0.0)  # práctica Lima: pared con pared
                _ret_po_def = float(_p_now.get("retiro_posterior_ml") or 0.0)  # práctica Lima: pared con pared

                _g_ret_f  = _rg1.number_input("Retiro frontal (m)",   min_value=0.0, value=_ret_f_def,  step=0.5, key="geo_ret_frontal")
                _g_ret_l  = _rg2.number_input("Retiro lateral (m)",   min_value=0.0, value=_ret_l_def,  step=0.5, key="geo_ret_lateral")
                _g_ret_p  = _rg3.number_input("Retiro posterior (m)", min_value=0.0, value=_ret_po_def, step=0.5, key="geo_ret_posterior")
                _p_pisos_def = int(_p_now.get("pisos_max") or 8)
                _g_pisos  = _rg4.number_input("N° pisos",             min_value=1,   value=int(st.session_state.get("geo_n_pisos",  _p_pisos_def) or _p_pisos_def), step=1, key="geo_n_pisos")
                # Auto-sync sótanos: recalculo local (cocheras / capacidad por nivel)
                _estac_ms  = int((c.get("estac_total") or 0) if c else 0)
                _huella_ms = float((c.get("area_techada_piso_m2") or
                                    st.session_state.get("geo_huella", 0)) if c else
                                   st.session_state.get("geo_huella", 0))
                _area_lote_ms = float(_poly_lote.area if _poly_lote else 0)
                if _estac_ms > 0 and _area_lote_ms > 0:
                    _coch_ms  = max(1, int(_area_lote_ms * 0.87 / 25))
                    _sot_calc = -(-_estac_ms // _coch_ms)
                else:
                    _sot_calc = int((c.get("num_sotanos") or 0) if c else 0)
                if _sot_calc > 0 and _sot_calc != st.session_state.get("_geo_sot_synced"):
                    st.session_state["geo_n_sotanos"] = _sot_calc
                    st.session_state["_geo_sot_synced"] = _sot_calc
                _g_sotanos = _rg5.number_input("N° sótanos",          min_value=0,   value=int(st.session_state.get("geo_n_sotanos", max(_sot_calc, 1)) or 1), step=1, key="geo_n_sotanos")

                _g_h_piso = st.number_input("Altura piso a piso (m)", min_value=2.40, max_value=4.00, value=2.65, step=0.05, key="geo_h_piso")

                _poly_huella = _geo_aplicar_retiros(_poly_lote, _g_ret_f, _g_ret_l, _g_ret_p)
                _frente_val  = st.session_state.get("geo_frente_val", 0.0)
                _uso_geo     = "residencial"
                _al_min_pct  = float((st.session_state.get("params") or {}).get("area_libre_min_pct") or 0)
                _val_geo     = _geo_validar(_poly_lote, _poly_huella, _g_pisos, _frente_val, _uso_geo, _al_min_pct)

                # Métricas geométricas
                _mg = _val_geo["metricas"]
                _cos_real = _mg.get('cos_real_pct', 0)
                _cos_max  = _mg.get('cos_max_norma_pct', 100)
                _cos_delta = f"máx. norma {_cos_max:.0f}%" if _al_min_pct > 0 else None
                _mc1, _mc2, _mc3, _mc4 = st.columns(4)
                _mc1.metric("Área del lote",            f"{_mg.get('area_lote_m2', 0):,.0f} m²")
                _mc2.metric("Huella edificable",        f"{_mg.get('area_huella_efectiva_m2', 0):,.0f} m²")
                _mc3.metric("Coef. de Ocupación (COS)", f"{_cos_real:.1f}%", delta=_cos_delta,
                            delta_color="inverse")
                _mc4.metric("Área techada máx.",        f"{_mg.get('at_sobre_m2', 0):,.0f} m²")

                # Guardar techo duro en session_state para pasarlo a la IA
                st.session_state["geo_at_max"] = _mg.get("at_sobre_m2", 0)
                st.session_state["geo_huella"]  = _mg.get("area_huella_efectiva_m2", 0)

                # Alertas normativas
                for _alerta in _val_geo["alertas"]:
                    st.markdown(f'<div class="alert-gold">⚠ {_alerta}</div>', unsafe_allow_html=True)

                # ── Massing 3D — ancho completo con etiquetas ──
                st.markdown('<div style="font-size:11px;color:#7A7268;letter-spacing:1.5px;'
                            'text-transform:uppercase;font-weight:600;margin-bottom:6px;">'
                            'Massing 3D</div>', unsafe_allow_html=True)
                _unids_3d = (c.get("unidades") or []) if c else []
                _fig_3d = _gen_massing_3d_solid(
                    _poly_lote, _poly_huella, _g_pisos, _g_sotanos, _g_h_piso,
                    unidades=_unids_3d)
                st.plotly_chart(_fig_3d, use_container_width=True)

                # ── Leyenda de tipologías legible debajo del chart ──
                if _unids_3d:
                    _tipo_colors_html = {"1D":"#4AB3D9","2D":"#2A78C8","3D":"#C8A050","PH":"#9B59B6"}
                    _tipo_short_map   = {"1 Dorm.":"1D","2 Dorm.":"2D","3 Dorm.":"3D","Dúplex":"PH"}
                    _leg_items = []
                    for _u in _unids_3d:
                        _ts = _tipo_short_map.get(_u.get("tipo",""), _u.get("tipo","")[:2])
                        _tc = _tipo_colors_html.get(_ts, "#5A6A7A")
                        _cant = _u.get("cantidad", 0)
                        _am   = _u.get("area_m2", 0)
                        _leg_items.append(
                            f'<span style="display:inline-flex;align-items:center;gap:6px;'
                            f'margin-right:18px;font-size:13px;color:#1E2D3D;">'
                            f'<span style="width:14px;height:14px;border-radius:3px;'
                            f'background:{_tc};display:inline-block;flex-shrink:0;"></span>'
                            f'<strong style="color:{_tc};">{_ts}</strong>'
                            f' {_cant} unid. · {_am:.0f} m²</span>'
                        )
                    st.markdown(
                        '<div style="display:flex;flex-wrap:wrap;gap:4px;padding:10px 4px;'
                        'border-top:1px solid rgba(184,144,74,0.2);margin-top:-8px;">'
                        + "".join(_leg_items) + '</div>',
                        unsafe_allow_html=True
                    )

            else:
                # Sin polígono calculado — mostrar placeholder en lugar del 3D con datos por defecto
                st.markdown(
                    '<div style="border:1.5px dashed rgba(184,144,74,0.4);border-radius:10px;'
                    'padding:32px 24px;text-align:center;background:rgba(184,144,74,0.04);margin:8px 0;">'
                    '<div style="font-size:28px;margin-bottom:10px;">📐</div>'
                    '<div style="font-size:13px;font-weight:600;color:#1E2D3D;margin-bottom:6px;">'
                    'Massing 3D no disponible</div>'
                    '<div style="font-size:12px;color:#7A8A9A;line-height:1.6;">'
                    'Ingresa las medidas del lote (Frente · Fondo · Laterales) y presiona<br>'
                    '<strong>Calcular geometría</strong> para activar la visualización 3D del proyecto.</div>'
                    '</div>',
                    unsafe_allow_html=True
                )

            st.markdown("---")

            # ── PROGRAMA ARQUITECTÓNICO (cabida IA) ──────────────
            if not c:
                st.markdown('<div class="alert-legal">Genera el análisis primero.</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="section-title">Programa Arquitectónico</div>', unsafe_allow_html=True)
                col1, col2, col3, col4 = st.columns(4)
                # Recalculo local de sótanos: siempre consistente con estac_total actual
                _estac_sot  = int(c.get("estac_total", 0) or 0)
                _huella_sot = float(c.get("area_techada_piso_m2", 0) or 0)
                if _huella_sot > 0 and _estac_sot > 0:
                    _coch_x_sot = max(1, int(_huella_sot * 0.78 // 20))
                    _n_sot = -(-_estac_sot // _coch_x_sot)   # ceil division
                else:
                    _n_sot = int(c.get("num_sotanos", 0) or 0)
                col1.metric("M² construibles",  f"{c.get('area_techada_total_m2',0):,.0f}")
                col2.metric("M² vendibles",      f"{c.get('area_vendible_m2',0):,.0f}")
                col3.metric("Departamentos",     str(c.get('total_unidades', 0)))
                col4.metric("Estacionamientos",  str(_estac_sot),
                            delta=f"{_n_sot} sótano{'s' if _n_sot != 1 else ''}" if _n_sot else None,
                            delta_color="off")

                _al_m2 = c.get("area_libre_m2", 0) or 0
                _ac_m2 = c.get("area_comunes_m2", 0) or 0
                _at_m2 = c.get("area_techada_total_m2", 0) or 1
                _av_m2 = c.get("area_vendible_m2", 0) or 0
                _pct_vend = round(_av_m2 / _at_m2 * 100, 1) if _at_m2 > 0 else 0
                _lote_m2  = float((p or {}).get("area_terreno_m2", 0) or 0)
                _pct_libre = round(_al_m2 / _lote_m2 * 100, 1) if _lote_m2 > 0 else 0
                colA, colB, colC, colD = st.columns(4)
                colA.metric("Área libre", f"{_al_m2:,.0f} m²",
                            delta=f"{_pct_libre:.0f}% del lote")
                colB.metric("Área común", f"{_ac_m2:,.0f} m²")
                colC.metric("Eficiencia vendible", f"{_pct_vend:.1f}%")
                colD.metric("Área lote", f"{_lote_m2:,.0f} m²")

                st.markdown('<div class="section-title">Mix de Tipologías</div>', unsafe_allow_html=True)
                unidades = c.get("unidades", [])
                if unidades:
                    # ── Tabla HTML tipologías + cocheras + depósitos ──────
                    _estac_r = c.get("estac_residentes", 0)
                    _estac_v = c.get("estac_visitas", 0)
                    _estac_t = c.get("estac_total", 0)
                    _dep_t   = c.get("depositos_total", 0)
                    _tot_u   = c.get("total_unidades", 0)

                    _rows_tipol = ""
                    for _u in unidades:
                        _tip  = _u[0] if isinstance(_u, (list, tuple)) else _u.get("tipologia", _u.get("tipo", "—"))
                        _cant = _u[1] if isinstance(_u, (list, tuple)) else _u.get("cantidad", 0)
                        _area = _u[2] if isinstance(_u, (list, tuple)) else _u.get("area_m2", 0)
                        _atot = _u[3] if isinstance(_u, (list, tuple)) else _u.get("area_total_m2", _u.get("area_total", 0))
                        _pct  = round(_cant / _tot_u * 100) if _tot_u else 0
                        _rows_tipol += (
                            f'<tr>'
                            f'<td style="padding:9px 14px;color:#1E2D3D;font-weight:600;">{_tip}</td>'
                            f'<td style="padding:9px 14px;color:#1E2D3D;text-align:center;font-weight:700;font-size:15px;">{_cant}</td>'
                            f'<td style="padding:9px 14px;color:#7A7268;text-align:center;">{_pct}%</td>'
                            f'<td style="padding:9px 14px;color:#1E2D3D;text-align:right;">{float(_area):,.1f} m²</td>'
                            f'<td style="padding:9px 14px;color:#1E2D3D;text-align:right;">{float(_atot):,.1f} m²</td>'
                            f'</tr>'
                        )

                    _rows_extra = (
                        f'<tr style="border-top:2px solid #D8D4CC;">'
                        f'<td style="padding:9px 14px;color:#1E2D3D;font-weight:600;">Cocheras residentes</td>'
                        f'<td style="padding:9px 14px;color:#1E2D3D;text-align:center;font-weight:700;font-size:15px;">{_estac_r}</td>'
                        f'<td colspan="3" style="padding:9px 14px;color:#7A7268;font-size:11px;">de {_estac_t} totales ({_estac_v} visitas)</td>'
                        f'</tr>'
                    )
                    if _dep_t > 0:
                        _rows_extra += (
                            f'<tr>'
                            f'<td style="padding:9px 14px;color:#1E2D3D;font-weight:600;">Depósitos / bodegas</td>'
                            f'<td style="padding:9px 14px;color:#1E2D3D;text-align:center;font-weight:700;font-size:15px;">{_dep_t}</td>'
                            f'<td colspan="3" style="padding:9px 14px;color:#7A7268;font-size:11px;">'
                            f'{round(_dep_t/_tot_u*100) if _tot_u else 0}% de las unidades</td>'
                            f'</tr>'
                        )

                    st.markdown(f"""
                    <table style="width:100%;border-collapse:collapse;background:#FFFFFF;
                                  border:1px solid #D8D4CC;border-radius:8px;overflow:hidden;
                                  font-family:Inter,sans-serif;font-size:13px;">
                      <thead>
                        <tr style="background:#1E2D3D;">
                          <th style="padding:9px 14px;color:#B8C8D8;font-weight:600;font-size:10px;
                                     letter-spacing:1.5px;text-transform:uppercase;text-align:left;">Tipología</th>
                          <th style="padding:9px 14px;color:#B8C8D8;font-weight:600;font-size:10px;
                                     letter-spacing:1.5px;text-transform:uppercase;text-align:center;">Unidades</th>
                          <th style="padding:9px 14px;color:#B8C8D8;font-weight:600;font-size:10px;
                                     letter-spacing:1.5px;text-transform:uppercase;text-align:center;">Mix</th>
                          <th style="padding:9px 14px;color:#B8C8D8;font-weight:600;font-size:10px;
                                     letter-spacing:1.5px;text-transform:uppercase;text-align:right;">Área/unid.</th>
                          <th style="padding:9px 14px;color:#B8C8D8;font-weight:600;font-size:10px;
                                     letter-spacing:1.5px;text-transform:uppercase;text-align:right;">Área total</th>
                        </tr>
                      </thead>
                      <tbody>
                        {_rows_tipol}
                        {_rows_extra}
                      </tbody>
                    </table>
                    """, unsafe_allow_html=True)

                    st.markdown('<div style="height:16px;"></div>', unsafe_allow_html=True)

                    df_u = pd.DataFrame([{
                        "Tipología": u.get("tipologia", u.get("tipo", "—")) if isinstance(u, dict) else (u[0] if isinstance(u, (list,tuple)) else "—"),
                        "Cantidad":  u.get("cantidad", 0) if isinstance(u, dict) else (u[1] if isinstance(u, (list,tuple)) else 0),
                    } for u in unidades])
                    fig = px.pie(df_u, values="Cantidad", names="Tipología",
                                 color_discrete_sequence=["#1E2D3D", "#B8904A", "#8A9BAD", "#6B7F8E"], hole=0.5)
                    fig.update_traces(textfont_size=12, textfont_family="Inter, sans-serif",
                                      marker=dict(line=dict(color="#EDEAE4", width=3)))
                    fig.update_layout(
                        height=280, margin=dict(t=10, b=10, l=10, r=10),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(family="Inter, sans-serif", color="#1E2D3D"),
                        legend=dict(font=dict(size=12), bgcolor="rgba(0,0,0,0)"),
                    )
                    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

                col1, col2, col3 = st.columns(3)
                col1.metric("Estac. residentes", c.get("estac_residentes", 0))
                col2.metric("Estac. visitas",    c.get("estac_visitas",    0))
                col3.metric("Pisos",             c.get("num_pisos", "—"))

                # ── KPIs de eficiencia (benchmarks U. del Pacífico) ─
                _at_sobre  = c.get("area_techada_total_m2", 0)
                _at_sotano = c.get("num_sotanos", 0) * c.get("area_techada_piso_m2", 0)
                _at_total  = _at_sobre + _at_sotano
                _av        = c.get("area_vendible_m2", 0)
                _unidades  = c.get("total_unidades", 1) or 1
                _estac     = c.get("estac_total", 0)
                _av_at_sobre = round(_av / _at_sobre * 100, 1) if _at_sobre > 0 else 0
                _av_at_total = round(_av / _at_total * 100, 1) if _at_total > 0 else 0
                _estac_res   = c.get("estac_residentes", _estac)
                _autos_viv   = round(_estac_res / _unidades, 2) if _unidades > 0 else 0

                st.markdown('<div class="section-title">KPIs de Eficiencia</div>', unsafe_allow_html=True)
                _kc1, _kc2, _kc3, _kc4 = st.columns(4)
                _kc1.metric("AV/AT sobre rasante", f"{_av_at_sobre}%")
                _kc2.metric("AV/AT total (c/sótanos)", f"{_av_at_total}%")
                _kc3.metric("Autos / vivienda", f"{_autos_viv}")
                _kc4.metric("Depósitos / vivienda", f"{round(c.get('depositos_total',0)/_unidades,2)}")

                # ── Cuadro normativo: aforo / residuos / bicicletas ──
                if unidades:
                    _aforo_map = {"1 dorm": 2, "1dorm": 2, "1 d": 2,
                                  "2 dorm": 3, "2dorm": 3, "2 d": 3,
                                  "3 dorm": 4, "3dorm": 4, "3 d": 4,
                                  "dúplex": 5, "duplex": 5, "penthouse": 5}
                    def _personas_por_tipo(tipo_str):
                        t = tipo_str.lower()
                        for k, v in _aforo_map.items():
                            if k in t:
                                return v
                        return 3

                    _reg_rows = ""
                    _tot_aforo = 0
                    _tot_resid = 0
                    for _u in unidades:
                        _tip  = _u.get("tipologia", _u.get("tipo", "—")) if isinstance(_u, dict) else (_u[0] if isinstance(_u, (list,tuple)) else "—")
                        _cant = _u.get("cantidad", 0) if isinstance(_u, dict) else (_u[1] if isinstance(_u, (list,tuple)) else 0)
                        _area = _u.get("area_m2", 0) if isinstance(_u, dict) else (_u[2] if isinstance(_u, (list,tuple)) else 0)
                        _prs  = _personas_por_tipo(_tip)
                        _afor = _cant * _prs
                        _res  = _afor * 4.0
                        _tot_aforo += _afor
                        _tot_resid += _res
                        _pct2 = round(_cant / _tot_u * 100) if _tot_u else 0
                        _reg_rows += (
                            f'<tr>'
                            f'<td style="padding:7px 12px;color:#1E2D3D;">{_tip}</td>'
                            f'<td style="padding:7px 12px;text-align:center;color:#1E2D3D;">{_area:.0f} m²</td>'
                            f'<td style="padding:7px 12px;text-align:center;font-weight:700;color:#1E2D3D;">{_cant}</td>'
                            f'<td style="padding:7px 12px;text-align:center;color:#7A7268;">{_pct2}%</td>'
                            f'<td style="padding:7px 12px;text-align:center;color:#1E2D3D;">{_afor}</td>'
                            f'<td style="padding:7px 12px;text-align:center;color:#1E2D3D;">{_res:.0f} L/día</td>'
                            f'</tr>'
                        )
                    _bicicletas = max(1, round(_tot_u / 3))
                    _reg_rows += (
                        f'<tr style="border-top:2px solid #D8D4CC;background:#F5F3EF;">'
                        f'<td style="padding:7px 12px;color:#1E2D3D;font-weight:700;">TOTAL</td>'
                        f'<td style="padding:7px 12px;"></td>'
                        f'<td style="padding:7px 12px;text-align:center;font-weight:700;color:#1E2D3D;">{_tot_u}</td>'
                        f'<td style="padding:7px 12px;text-align:center;color:#7A7268;">100%</td>'
                        f'<td style="padding:7px 12px;text-align:center;font-weight:700;color:#1E2D3D;">{_tot_aforo}</td>'
                        f'<td style="padding:7px 12px;text-align:center;font-weight:700;color:#1E2D3D;">{_tot_resid:.0f} L/día</td>'
                        f'</tr>'
                    )
                    st.markdown('<div class="section-title">Cuadro Normativo — Aforo y Residuos</div>', unsafe_allow_html=True)
                    st.markdown(f"""
                    <table style="width:100%;border-collapse:collapse;background:#FFFFFF;
                                  border:1px solid #D8D4CC;border-radius:8px;overflow:hidden;
                                  font-family:Inter,sans-serif;font-size:12px;margin-bottom:8px;">
                      <thead>
                        <tr style="background:#1E2D3D;">
                          <th style="padding:7px 12px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:left;">Tipología</th>
                          <th style="padding:7px 12px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:center;">Área/unid.</th>
                          <th style="padding:7px 12px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:center;">Cant.</th>
                          <th style="padding:7px 12px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:center;">Mix</th>
                          <th style="padding:7px 12px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:center;">Aforo</th>
                          <th style="padding:7px 12px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:center;">Residuos (4 L/p/d)</th>
                        </tr>
                      </thead>
                      <tbody>{_reg_rows}</tbody>
                    </table>
                    <p style="font-family:Inter,sans-serif;font-size:11px;color:#9A9590;margin:4px 0 16px;">
                      Bicicletas requeridas (A.020 — 1 c/3 unidades): <b style="color:#1E2D3D;">{_bicicletas}</b> &nbsp;·&nbsp;
                      Aforo total: <b style="color:#1E2D3D;">{_tot_aforo} personas</b> &nbsp;·&nbsp;
                      Generación residuos: <b style="color:#1E2D3D;">{_tot_resid:.0f} L/día</b>
                    </p>
                    """, unsafe_allow_html=True)

                _ord_alt = c.get("ordenanzas_mayor_altura", [])
                _obs_all = c.get("observaciones", [])
                _obs_benefit = [o for o in _obs_all if "BENEFICIO POTENCIAL" in o]
                _obs_regular = [o for o in _obs_all if "BENEFICIO POTENCIAL" not in o]

                if _ord_alt:
                    with st.expander(f"⚖️ Ordenanzas para Mayor Altura ({len(_ord_alt)})"):
                        for o in _ord_alt:
                            st.markdown(f'<div class="alert-gold">⚖️ {o}</div>', unsafe_allow_html=True)

                if _obs_benefit:
                    for o in _obs_benefit:
                        st.markdown(f'<div class="alert-gold">✨ {o}</div>', unsafe_allow_html=True)

                if _obs_regular:
                    with st.expander(f"Observaciones normativas ({len(_obs_regular)})"):
                        for obs in _obs_regular:
                            st.markdown(f"• {obs}")

                if c.get("metodologia"):
                    with st.expander("Ver metodología de cálculo"):
                        st.write(c["metodologia"])

        # ── TAB 3: FINANCIERO ────────────────────────────
        with tabs[2]:
            if not c:
                st.markdown('<div class="alert-legal">Genera el análisis primero.</div>', unsafe_allow_html=True)
            else:
                fi     = st.session_state.financ_inputs or {}
                m_data = MERCADO.get(zona_sel, {})

                col_s1, col_s2 = st.columns([3, 1])
                with col_s1:
                    tasa = st.slider("Tasa financiamiento bancario (%)", 0.0, 15.0,
                                     float(fi.get("tasa_financ", 9.0)), step=0.5)
                with col_s2:
                    st.metric("Velocidad de mercado", f"{m_data.get('velocidad_venta', 1.0):.2f} und/mes")

                fin_run = {
                    "costo_terreno":      fi.get("costo_terreno", 0),
                    "costo_construccion": fi.get("costo_construccion", 700),
                    "costo_sotano_m2":    fi.get("costo_sotano_m2", 450),
                    "fee_constructora":   fi.get("fee_constructora", 10.0),
                    "tasa_ir":            fi.get("tasa_ir", 29.5),
                    "include_alcabala":   fi.get("include_alcabala", True),
                    "include_dd":         fi.get("include_dd", True),
                    "costo_demolicion":   fi.get("costo_demolicion", 0),
                    "precio_venta_m2":    fi.get("precio_venta_m2", m_data.get("precio_2br", 0)),
                    "precio_estac":       m_data.get("precio_estac", 0),
                    "precio_deposito":    m_data.get("precio_deposito", 0),
                    "tasa_financ":        tasa,
                    "estructura_financ":  fi.get("estructura_financ", "estandar"),
                    "aporte_propio_pct":  fi.get("aporte_propio_pct", 100.0),
                    "pct_preventa_banco": fi.get("pct_preventa_banco", 30.0),
                    "meses_preventa_override": fi.get("meses_preventa_override"),
                    "meses_obra_override": fi.get("meses_obra_override"),
                    "pct_mktg_preventa":  fi.get("pct_mktg_preventa", 2.0),
                    "nombre_proyecto":    st.session_state.get("nombre_proyecto", ""),
                }
                result = calcular_financiero(c, fin_run, zona_sel)
                st.session_state.financ = result
                r = result.get("resumen", {})

                # ── Métricas clave destacadas ─────────────
                _mg  = r.get("margen_pct", 0) or 0
                _tir = r.get("tir_anual_pct", 0) or 0
                _roi = r.get("roi_pct", 0) or 0
                _mg_c  = "#1A4731" if _mg  >= 20 else ("#7A4F1A" if _mg  >= 12 else "#7A1A1A")
                _tir_c = "#1A4731" if _tir >= 15 else ("#7A4F1A" if _tir >= 10 else "#7A1A1A")
                _roi_c = "#1A4731" if _roi >= 20 else ("#7A4F1A" if _roi >= 12 else "#7A1A1A")
                st.markdown(f"""
                <div style="background:#E8F5EE;border:1px solid #6BAE90;border-left:4px solid #1A4731;
                            border-radius:8px;padding:16px 24px;margin-bottom:16px;">
                  <div style="font-size:9px;color:#1A4731;letter-spacing:3px;text-transform:uppercase;
                              font-weight:700;margin-bottom:12px;">Métricas Determinantes del Proyecto</div>
                  <div style="display:flex;gap:32px;flex-wrap:wrap;align-items:flex-end;">
                    <div>
                      <div style="font-size:10px;color:#4A5568;font-weight:600;text-transform:uppercase;
                                  letter-spacing:1px;margin-bottom:2px;">Margen Neto post-IR</div>
                      <div style="font-size:32px;font-weight:800;color:{_mg_c};letter-spacing:-1px;">{_mg:.1f}%</div>
                      <div style="font-size:10px;color:#7A9A80;">Ref. óptimo: ≥ 20%</div>
                    </div>
                    <div style="width:1px;background:#6BAE90;opacity:0.4;align-self:stretch;"></div>
                    <div>
                      <div style="font-size:10px;color:#4A5568;font-weight:600;text-transform:uppercase;
                                  letter-spacing:1px;margin-bottom:2px;">TIR Anual Estimada</div>
                      <div style="font-size:32px;font-weight:800;color:{_tir_c};letter-spacing:-1px;">{_tir:.1f}%</div>
                      <div style="font-size:10px;color:#7A9A80;">Ref. óptimo: ≥ 15%</div>
                    </div>
                    <div style="width:1px;background:#6BAE90;opacity:0.4;align-self:stretch;"></div>
                    <div>
                      <div style="font-size:10px;color:#4A5568;font-weight:600;text-transform:uppercase;
                                  letter-spacing:1px;margin-bottom:2px;">ROI</div>
                      <div style="font-size:32px;font-weight:800;color:{_roi_c};letter-spacing:-1px;">{_roi:.1f}%</div>
                      <div style="font-size:10px;color:#7A9A80;">Ref. óptimo: ≥ 20%</div>
                    </div>
                    <div style="width:1px;background:#6BAE90;opacity:0.4;align-self:stretch;"></div>
                    <div>
                      <div style="font-size:10px;color:#4A5568;font-weight:600;text-transform:uppercase;
                                  letter-spacing:1px;margin-bottom:2px;">Utilidad Neta</div>
                      <div style="font-size:28px;font-weight:800;color:#1A4731;letter-spacing:-0.5px;">{fmt_usd(r["utilidad_neta"])}</div>
                      <div style="font-size:10px;color:#7A9A80;">post-impuestos</div>
                    </div>
                  </div>
                </div>""", unsafe_allow_html=True)

                # ── KPIs secundarios ──────────────────────
                st.markdown('<div class="section-title">Detalle de Indicadores</div>', unsafe_allow_html=True)
                col1, col2, col3, col4, col5, col6 = st.columns(6)
                col1.metric("Ingresos brutos",   fmt_usd(r.get("ingresos_brutos", 0)))
                col2.metric("Utilidad bruta",    fmt_usd(r.get("utilidad_bruta", 0)),  delta=f"{r.get('margen_bruto_pct', 0)}% bruto")
                col3.metric(f"IR ({r.get('ir_pct', 29.5)}%)", fmt_usd(r.get("costo_ir", 0)))
                col4.metric("Utilidad neta",     fmt_usd(r.get("utilidad_neta", 0)),  delta=f"{r.get('margen_pct', 0)}% neto")
                col5.metric("ROI / TIR",         f"{r.get('roi_pct', 0)}% / {r.get('tir_anual_pct', 0)}%")
                col6.metric("Break-even m²",     f"${r.get('be_precio_m2', 0):,}")

                # ── Viabilidad del proyecto ──────────────────────────
                _tit = r.get("tit_pct", 0) or 0
                if _mg >= 20 and _tir >= 15:
                    _perfil = ("RETORNOS SÓLIDOS",   "#1A4731", "#E8F5EE", "#6BAE90",
                               "Margen y TIR en rango alto — evalúa si se alinea con tu perfil de inversión")
                elif _mg > 0 and _tir > 0:
                    _perfil = ("RETORNOS MODERADOS", "#7A5500", "#FFF8E6", "#E8C55A",
                               "El proyecto genera retornos positivos — analiza si los márgenes se ajustan a tu objetivo")
                else:
                    _perfil = ("RETORNOS NEGATIVOS", "#4A4A5A", "#F4F4F6", "#9A9AAA",
                               "El proyecto presenta pérdidas en las condiciones actuales")
                st.markdown(
                    f'<div style="background:{_perfil[2]};border:1px solid {_perfil[3]};border-left:4px solid {_perfil[1]};'
                    f'border-radius:6px;padding:12px 20px;margin:8px 0 16px;display:flex;align-items:center;gap:20px;">'
                    f'<div><div style="font-size:9px;color:{_perfil[1]};letter-spacing:2px;font-weight:700;text-transform:uppercase;">Perfil de Inversión</div>'
                    f'<div style="font-size:18px;font-weight:800;color:{_perfil[1]};letter-spacing:1px;margin-top:2px;">{_perfil[0]}</div></div>'
                    f'<div style="width:1px;background:{_perfil[3]};opacity:0.5;align-self:stretch;"></div>'
                    f'<div style="font-size:11px;color:{_perfil[1]};opacity:0.9;">{_perfil[4]}</div>'
                    f'<div style="margin-left:auto;text-align:right;">'
                    f'<div style="font-size:9px;color:{_perfil[1]};opacity:0.7;letter-spacing:1px;">TIT (Incidencia Terreno)</div>'
                    f'<div style="font-size:16px;font-weight:700;color:{_perfil[1]};">{_tit:.1f}% / ingresos</div>'
                    f'</div></div>',
                    unsafe_allow_html=True
                )

                # ── Escenario con banco ───────────────────
                st.markdown('<div class="section-title">Escenario con Financiamiento Bancario</div>', unsafe_allow_html=True)
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Gasto financiero",     fmt_usd(r.get("costo_financiero", 0)))
                col2.metric("Costo total c/banco",  fmt_usd(r.get("costo_total_con_financ", 0)))
                col3.metric("Utilidad neta c/banco", fmt_usd(r.get("utilidad_con_financ", 0)), delta=f"{r.get('margen_con_financ_pct', 0)}% neto")
                col4.metric("TIR anual est.",        f"{r.get('tir_anual_pct', 0)}%")

                # ── Timeline ─────────────────────────────
                st.markdown('<div class="section-title">Timeline Estimado</div>', unsafe_allow_html=True)
                col1, col2, col3 = st.columns(3)
                col1.metric("Meses de obra",          f"{r.get('meses_obra', 0)} meses")
                col2.metric("Meses de ventas",         f"{r.get('meses_venta', 0)} meses")
                col3.metric("Duración total proyecto", f"{r.get('meses_proyecto', 0)} meses")

                # ── Detalle ingresos / costos + tipologías ─
                col1, col2 = st.columns(2)
                _pvm = fin_run.get("precio_venta_m2", m_data.get("precio_2br", 0))
                _pe  = m_data.get("precio_estac", 0)
                _pd  = m_data.get("precio_deposito", 0)
                with col1:
                    st.markdown('<div class="section-title">Ingresos</div>', unsafe_allow_html=True)
                    for k, v in result["detalle_ingresos"].items():
                        row_item(k, fmt_usd(v), highlight="TOTAL" in k)

                    # ── Desglose de tipologías ────────────
                    _unidades = c.get("unidades") or []
                    if _unidades:
                        NAV = "#1E2D3D"; BRD = "#D8D4CC"
                        _th = f'style="background:{NAV};color:#fff;padding:7px 10px;font-size:10px;font-weight:700;letter-spacing:0.5px;text-align:left;border:1px solid {BRD};"'
                        _tr = f'style="background:{NAV};color:#fff;padding:7px 10px;font-size:10px;font-weight:700;letter-spacing:0.5px;text-align:right;border:1px solid {BRD};"'
                        _td = f'style="padding:6px 10px;font-size:11px;color:{NAV};border:1px solid {BRD};"'
                        _tv = f'style="padding:6px 10px;font-size:11px;color:{NAV};text-align:right;border:1px solid {BRD};"'
                        _tf = f'style="padding:7px 10px;font-size:11px;font-weight:700;color:{NAV};border:1px solid {BRD};"'
                        _tfv = f'style="padding:7px 10px;font-size:11px;font-weight:700;color:{NAV};text-align:right;border:1px solid {BRD};"'
                        _tbl = (f'<div style="margin-top:14px;">'
                                f'<div style="font-size:9px;color:#B8904A;letter-spacing:2px;font-weight:700;'
                                f'text-transform:uppercase;padding:10px 0 6px;">Desglose por Tipología</div>'
                                f'<table style="border-collapse:collapse;width:100%;">'
                                f'<thead><tr>'
                                f'<th {_th}>Tipología</th>'
                                f'<th {_tr}>Unds.</th>'
                                f'<th {_tr}>Área m²</th>'
                                f'<th {_tr}>Precio/und</th>'
                                f'<th {_tr}>Subtotal</th>'
                                f'</tr></thead><tbody>')
                        for _u in _unidades:
                            _tipo = _u.get("tipo", "—")
                            _cant = _u.get("cantidad", 0)
                            _am2  = _u.get("area_m2", 0)
                            _pund = _am2 * _pvm
                            _sub  = _cant * _pund
                            _tbl += (f'<tr>'
                                     f'<td {_td}>{_tipo}</td>'
                                     f'<td {_tv}>{_cant}</td>'
                                     f'<td {_tv}>{_am2:.0f}</td>'
                                     f'<td {_tv}>${_pund:,.0f}</td>'
                                     f'<td {_tv}>${_sub:,.0f}</td>'
                                     f'</tr>')
                        _n_estac = c.get("estac_total", 0)
                        _n_dep   = c.get("depositos_total", 0)
                        if _n_estac:
                            _tbl += (f'<tr style="background:#F5F2ED !important;">'
                                     f'<td {_td}>Estacionamientos</td>'
                                     f'<td {_tv}>{_n_estac}</td>'
                                     f'<td {_tv}>—</td>'
                                     f'<td {_tv}>${_pe:,.0f}</td>'
                                     f'<td {_tv}>${_n_estac*_pe:,.0f}</td>'
                                     f'</tr>')
                        if _n_dep:
                            _tbl += (f'<tr style="background:#F5F2ED !important;">'
                                     f'<td {_td}>Depósitos</td>'
                                     f'<td {_tv}>{_n_dep}</td>'
                                     f'<td {_tv}>—</td>'
                                     f'<td {_tv}>${_pd:,.0f}</td>'
                                     f'<td {_tv}>${_n_dep*_pd:,.0f}</td>'
                                     f'</tr>')
                        _tbl += (f'<tr>'
                                 f'<td {_tf}>TOTAL</td>'
                                 f'<td {_tfv}>{sum(u.get("cantidad",0) for u in _unidades)}</td>'
                                 f'<td {_tfv}>—</td>'
                                 f'<td {_tfv}>—</td>'
                                 f'<td {_tfv}>${r["ingresos_brutos"]:,.0f}</td>'
                                 f'</tr>')
                        _tbl += '</tbody></table></div>'
                        st.markdown(_tbl, unsafe_allow_html=True)

                with col2:
                    st.markdown('<div class="section-title">Costos</div>', unsafe_allow_html=True)
                    for k, v in result["detalle_costos"].items():
                        if k.startswith("──"):
                            st.markdown(f'<div style="font-size:9px;color:#B8904A;letter-spacing:2px;font-weight:700;padding:10px 0 4px 14px;text-transform:uppercase;">{k.replace("──","").strip()}</div>', unsafe_allow_html=True)
                        else:
                            row_item(k, fmt_usd(v), highlight="TOTAL" in k or "SUBTOTAL" in k)

                # ── Distribución de costos (pie chart) ──
                st.markdown('<div class="section-title">Distribución de Costos</div>', unsafe_allow_html=True)
                _raw = result.get("_raw", {})
                _pie_labels = ["Terreno", "Construcción", "Costos Técnicos", "Costos Inmobiliarios", "Gasto Financiero"]
                _pie_values = [
                    _raw.get("c_terreno_total", 0),
                    _raw.get("c_obra_dptos", 0) + _raw.get("c_obra_sotanos", 0) + _raw.get("c_constructora", 0),
                    _raw.get("c_arq", 0) + _raw.get("c_esp", 0) + _raw.get("c_supervision", 0) + _raw.get("c_permisos", 0),
                    _raw.get("c_gerenciamiento", 0) + _raw.get("c_ventas_marketing", 0) + _raw.get("c_legales", 0),
                    _raw.get("c_financiero", 0),
                ]
                _pie_colors = ["#4A6B8A", "#1E2D3D", "#B8904A", "#8A9BAD", "#D4A853"]
                _pie_total = sum(_pie_values) or 1
                _pie_fig = px.pie(
                    values=_pie_values, names=_pie_labels,
                    color_discrete_sequence=_pie_colors,
                    hole=0.0,
                )
                _pie_fig.update_traces(
                    texttemplate="<b>%{label}</b><br>US$ %{value:,.0f}<br>(%{percent})",
                    textposition="outside",
                    textfont_size=11, textfont_family="Inter, sans-serif",
                    marker=dict(line=dict(color="#EDEAE4", width=2)),
                    outsidetextfont=dict(size=11, family="Inter, sans-serif", color="#1E2D3D"),
                )
                _pie_fig.update_layout(
                    height=420, margin=dict(t=30, b=30, l=120, r=120),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Inter, sans-serif", color="#1E2D3D"),
                    showlegend=False,
                    uniformtext=dict(minsize=9, mode="hide"),
                )
                st.plotly_chart(_pie_fig, use_container_width=True, config={"displayModeBar": False})

                # ── Valor estimado de compra del terreno ──
                st.markdown('<div class="section-title">Valor Estimado de Compra del Terreno</div>', unsafe_allow_html=True)
                _precio_actual = fi.get("costo_terreno", 0)
                _v20 = r.get("max_terreno_20pct", 0)
                _v15 = r.get("max_terreno_15pct", 0)
                _v12 = r.get("max_terreno_12pct", 0)
                # Determinar en qué zona está el precio actual
                if _precio_actual <= _v20:
                    _zona_actual = ("ZONA ÓPTIMA", "#1A4731", "#E8F5EE", "#6BAE90")
                elif _precio_actual <= _v15:
                    _zona_actual = ("ZONA ACEPTABLE", "#7A5500", "#FFF8E6", "#E8C55A")
                elif _precio_actual <= _v12:
                    _zona_actual = ("ZONA DE RIESGO", "#7A1A1A", "#FDECEA", "#E87070")
                else:
                    _zona_actual = ("PRECIO ELEVADO — INVIABLE", "#5A1A1A", "#FDE8E8", "#C84040")

                _ratio = r["ratio_terreno_pct"]
                st.markdown(
                    f'<div style="background:#FDFAF6;border:1px solid #D8D4CC;border-radius:8px;padding:20px 24px;margin-bottom:16px;">'
                    f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px;">'
                    # Escenario 01
                    f'<div style="background:#FFFFFF;border:1px solid #D8D4CC;border-left:4px solid #1E2D3D;border-radius:6px;padding:14px 16px;">'
                    f'<div style="font-size:9px;color:#1E2D3D;letter-spacing:2px;font-weight:700;text-transform:uppercase;margin-bottom:6px;">Escenario 01</div>'
                    f'<div style="font-size:9px;color:#7A7268;margin-bottom:8px;">Margen neto ≥ 20%</div>'
                    f'<div style="font-size:20px;font-weight:800;color:#1E2D3D;">{fmt_usd(_v20)}</div>'
                    f'<div style="font-size:10px;color:#9A9590;margin-top:4px;">Precio máximo al margen objetivo</div>'
                    f'</div>'
                    # Escenario 02
                    f'<div style="background:#FFFFFF;border:1px solid #D8D4CC;border-left:4px solid #B8904A;border-radius:6px;padding:14px 16px;">'
                    f'<div style="font-size:9px;color:#1E2D3D;letter-spacing:2px;font-weight:700;text-transform:uppercase;margin-bottom:6px;">Escenario 02</div>'
                    f'<div style="font-size:9px;color:#7A7268;margin-bottom:8px;">Margen neto 15–20%</div>'
                    f'<div style="font-size:20px;font-weight:800;color:#1E2D3D;">{fmt_usd(_v15)}</div>'
                    f'<div style="font-size:10px;color:#9A9590;margin-top:4px;">Precio máximo negociable</div>'
                    f'</div>'
                    # Escenario 03
                    f'<div style="background:#FFFFFF;border:1px solid #D8D4CC;border-left:4px solid #9A9590;border-radius:6px;padding:14px 16px;">'
                    f'<div style="font-size:9px;color:#1E2D3D;letter-spacing:2px;font-weight:700;text-transform:uppercase;margin-bottom:6px;">Escenario 03</div>'
                    f'<div style="font-size:9px;color:#7A7268;margin-bottom:8px;">Margen neto 12–15%</div>'
                    f'<div style="font-size:20px;font-weight:800;color:#1E2D3D;">{fmt_usd(_v12)}</div>'
                    f'<div style="font-size:10px;color:#9A9590;margin-top:4px;">Precio máximo permisible</div>'
                    f'</div>'
                    f'</div>'
                    # Bottom info bar — neutral, sin semáforo
                    f'<div style="background:#F5F3EF;border:1px solid #D8D4CC;border-radius:6px;padding:10px 16px;display:flex;align-items:center;gap:16px;">'
                    f'<div style="font-size:11px;color:#1E2D3D;font-weight:700;">Precio ingresado: {fmt_usd(_precio_actual)}</div>'
                    f'<div style="width:1px;background:#D8D4CC;height:18px;"></div>'
                    f'<div style="margin-left:auto;font-size:10px;color:#7A7268;">Ratio terreno/ingresos: {_ratio}%&nbsp;·&nbsp;Ref. óptima: 15–25%</div>'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

                # ── Matriz de sensibilidad ────────────────
                st.markdown('<div class="section-title">Matriz de Sensibilidad — Margen %</div>', unsafe_allow_html=True)
                st.caption("Impacto en margen neto ante variaciones de precio de venta (filas) y costo de construcción (columnas)")
                df_sens = calcular_sensibilidad(c, fin_run, zona_sel)

                # Renderizar como HTML con colores por umbral
                def _cell_color(val_str):
                    try:
                        v = float(val_str.strip("%"))
                        if v >= 20:   return "#E8F5EE", "#1A5C32"  # verde
                        elif v >= 12: return "#FFF8E6", "#7A5500"  # amarillo
                        else:         return "#FDECEA", "#7A1A1A"  # rojo
                    except Exception:
                        return "#FFFFFF", "#1E2D3D"

                NAV_S  = "#1E2D3D"
                BORD_S = "#D8D4CC"
                hdr_style = f'style="background:{NAV_S};color:#fff;padding:8px 14px;font-size:11px;font-weight:600;text-align:center;border:1px solid {BORD_S};"'
                lbl_style = f'style="background:#F0EDE8;color:{NAV_S};padding:8px 14px;font-size:11px;font-weight:700;border:1px solid {BORD_S};white-space:nowrap;"'

                sens_html = f'<table style="border-collapse:collapse;width:100%;margin-bottom:8px;"><thead><tr><th {hdr_style}></th>'
                for col in df_sens.columns:
                    sens_html += f'<th {hdr_style}>{col}</th>'
                sens_html += '</tr></thead><tbody>'
                for row_lbl, row in df_sens.iterrows():
                    sens_html += f'<tr><td {lbl_style}>{row_lbl}</td>'
                    for val in row:
                        bg, fg = _cell_color(str(val))
                        sens_html += (f'<td style="background:{bg} !important;color:{fg} !important;padding:8px 14px;'
                                      f'font-size:12px;font-weight:700;text-align:center;'
                                      f'border:1px solid {BORD_S};">{val}</td>')
                    sens_html += '</tr>'
                sens_html += '</tbody></table>'
                st.markdown(sens_html, unsafe_allow_html=True)
                st.caption("Verde = margen ≥ 20% · Amarillo = 12–20% · Rojo = < 12%")

                # ── Matriz Precio de Venta × Terreno ──────
                st.markdown("---")
                st.markdown(
                    '<div class="section-title">Matriz Estratégica — '
                    'Precio de Venta ($/m²) × Precio del Terreno ($)</div>',
                    unsafe_allow_html=True)
                st.caption(
                    "Margen neto % en cada combinación · "
                    "⭐ = escenario actual · "
                    "Verde ≥ 18% · Amarillo 12–18% · Rojo < 12%")

                _st_result = calcular_sensibilidad_terreno(c, fin_run, zona_sel)
                if _st_result:
                    _df_mg, _df_tir, _st_precios, _st_terrenos, _st_p0, _st_t0 = _st_result

                    # Precio base y terreno base más cercanos en la grilla
                    _st_col0 = min(range(len(_st_precios)), key=lambda i: abs(_st_precios[i] - _st_p0))
                    _st_row0 = min(range(len(_st_terrenos)), key=lambda i: abs(_st_terrenos[i] - _st_t0))

                    def _cell_color_mg(v):
                        if v >= 18:  return "#1B5E20", "#A5D6A7"   # verde oscuro / texto claro
                        if v >= 12:  return "#F57F17", "#FFF9C4"   # naranja / amarillo claro
                        return "#B71C1C", "#FFCDD2"                 # rojo oscuro / rosa claro

                    _NAV = "#0A1628"
                    _BRD = "#2A3D52"
                    # Cabecera de columnas (precio de venta)
                    _st_html = (
                        f'<div style="overflow-x:auto;margin-bottom:4px;">'
                        f'<table style="border-collapse:collapse;min-width:100%;">'
                        f'<thead><tr>'
                        f'<th style="background:{_NAV};color:#B8904A;padding:8px 12px;font-size:10px;'
                        f'font-weight:700;text-align:left;border:1px solid {_BRD};white-space:nowrap;">'
                        f'Terreno ↓ / Precio m² →</th>'
                    )
                    for ci, _sp in enumerate(_st_precios):
                        _is_base_col = (ci == _st_col0)
                        _ch_bg = "#1E3A5A" if _is_base_col else _NAV
                        _ch_fw = "800" if _is_base_col else "600"
                        _st_html += (
                            f'<th style="background:{_ch_bg};color:#FFFFFF;padding:8px 10px;'
                            f'font-size:10px;font-weight:{_ch_fw};text-align:center;'
                            f'border:1px solid {_BRD};white-space:nowrap;">'
                            f'{"⭐ " if _is_base_col else ""}${_sp:,}</th>'
                        )
                    _st_html += '</tr></thead><tbody>'

                    for ri, t in enumerate(_st_terrenos):
                        _is_base_row = (ri == _st_row0)
                        _rh_bg = "#0E2236" if _is_base_row else "#0A1628"
                        _st_html += (
                            f'<tr><td style="background:{_rh_bg};color:{"#B8904A" if _is_base_row else "#8AA8C0"};'
                            f'padding:8px 12px;font-size:10px;font-weight:{"800" if _is_base_row else "600"};'
                            f'border:1px solid {_BRD};white-space:nowrap;">'
                            f'{"⭐ " if _is_base_row else ""}${t:,.0f}</td>'
                        )
                        for ci in range(len(_st_precios)):
                            mg_val  = float(_df_mg.iloc[ri, ci])
                            tir_val = float(_df_tir.iloc[ri, ci])
                            _bg_cell, _txt_cell = _cell_color_mg(mg_val)
                            _is_cur = (_is_base_row and ci == _st_col0)
                            _border_extra = f"outline:2px solid #B8904A;outline-offset:-2px;" if _is_cur else ""
                            _st_html += (
                                f'<td style="background:{_bg_cell};color:{_txt_cell};'
                                f'padding:6px 10px;font-size:11px;font-weight:700;'
                                f'text-align:center;border:1px solid {_BRD};{_border_extra}">'
                                f'{mg_val:.0f}%'
                                f'<div style="font-size:9px;font-weight:400;margin-top:1px;">'
                                f'TIR {tir_val:.0f}%</div></td>'
                            )
                        _st_html += '</tr>'
                    _st_html += '</tbody></table></div>'
                    st.markdown(_st_html, unsafe_allow_html=True)
                    st.caption(
                        f"Base actual: terreno ${fin_run['costo_terreno']:,.0f} · "
                        f"precio ${fin_run['precio_venta_m2']:,}/m² → "
                        f"margen {(st.session_state.financ or {}).get('resumen', {}).get('margen_pct', 0):.1f}%"
                    )
                else:
                    st.info("Ingresa precio de venta y costo de terreno para ver la matriz.")

                # ── Comparador de Escenarios ──────────────
                st.markdown("---")
                st.markdown('<div class="section-title">Comparador de Escenarios</div>', unsafe_allow_html=True)
                st.caption("Modifica precio de venta, costo de construcción o terreno para ver el impacto lado a lado")

                _ecc1, _ecc2, _ecc3 = st.columns(3)
                with _ecc1:
                    st.markdown(
                        '<div style="background:#1E2D3D;border-radius:8px;padding:10px 14px;margin-bottom:6px;">'
                        '<div style="font-size:11px;font-weight:700;color:#B8904A;letter-spacing:1.5px;text-transform:uppercase;">Escenario 1 — BASE</div>'
                        '<div style="font-size:10px;color:rgba(255,255,255,0.45);margin-top:2px;">Análisis actual</div>'
                        '</div>', unsafe_allow_html=True)
                    st.markdown(f'<div style="font-size:12px;color:#5A6A7A;margin-top:2px;">Precio: <b style="color:#1E2D3D;">${fin_run["precio_venta_m2"]:,}/m²</b></div>', unsafe_allow_html=True)
                    st.markdown(f'<div style="font-size:12px;color:#5A6A7A;">Construcción: <b style="color:#1E2D3D;">${fin_run["costo_construccion"]:,}/m²</b></div>', unsafe_allow_html=True)
                    st.markdown(f'<div style="font-size:12px;color:#5A6A7A;">Terreno: <b style="color:#1E2D3D;">${fin_run["costo_terreno"]:,.0f}</b></div>', unsafe_allow_html=True)

                with _ecc2:
                    _e2n = st.text_input("Nombre", value="Optimista", key="esc2_n", label_visibility="collapsed",
                                         placeholder="Escenario 2 — nombre")
                    st.caption(f"**{_e2n or 'Escenario 2'}**")
                    _e2p = st.number_input("Precio venta m²", 800, 8000,
                                           min(max(int(fin_run["precio_venta_m2"] * 1.10), 800), 8000), 50,
                                           key="esc2_p")
                    _e2c = st.number_input("Costo construcción m²", 400, 3000,
                                           int(fin_run["costo_construccion"]), 50, key="esc2_c")
                    _e2t = st.number_input("Terreno ($)", 0, 50_000_000,
                                           int(fin_run["costo_terreno"]), 5000, format="%d", key="esc2_t")

                with _ecc3:
                    _e3n = st.text_input("Nombre", value="Conservador", key="esc3_n", label_visibility="collapsed",
                                         placeholder="Escenario 3 — nombre")
                    st.caption(f"**{_e3n or 'Escenario 3'}**")
                    _e3p = st.number_input("Precio venta m²", 800, 8000,
                                           min(max(int(fin_run["precio_venta_m2"] * 0.90), 800), 8000), 50,
                                           key="esc3_p")
                    _e3c = st.number_input("Costo construcción m²", 400, 3000,
                                           min(int(fin_run["costo_construccion"] * 1.05), 3000), 50, key="esc3_c")
                    _e3t = st.number_input("Terreno ($)", 0, 50_000_000,
                                           int(fin_run["costo_terreno"]), 5000, format="%d", key="esc3_t")

                # Calcular los 3 escenarios
                _esc_all = [
                    ("Base",             fin_run["precio_venta_m2"], fin_run["costo_construccion"], fin_run["costo_terreno"]),
                    (_e2n or "Optimista",    _e2p, _e2c, _e2t),
                    (_e3n or "Conservador",  _e3p, _e3c, _e3t),
                ]
                _esc_res = []
                for _en, _ep, _ecv, _et in _esc_all:
                    _fin_e = {**fin_run, "precio_venta_m2": _ep, "costo_construccion": _ecv, "costo_terreno": _et}
                    _r_e   = calcular_financiero(c, _fin_e, zona_sel)["resumen"]
                    _esc_res.append((_en, _ep, _ecv, _et, _r_e))

                # Tabla comparativa
                def _esc_bg(val, t1, t2):
                    if val >= t1: return "#E8F5EE", "#1A5C32"
                    if val >= t2: return "#FFF8E6", "#7A5500"
                    return "#FDECEA", "#7A1A1A"

                _BRD = "#D8D4CC"
                _NH  = "#1E2D3D"
                _tbl = (f'<table style="border-collapse:collapse;width:100%;margin-top:14px;">'
                        f'<thead><tr>'
                        f'<th style="background:{_NH};color:#fff;padding:8px 14px;font-size:11px;font-weight:600;text-align:left;border:1px solid {_BRD};min-width:170px;">Métrica</th>')
                for _en, *_ in _esc_res:
                    _is_base = _en == "Base"
                    _hbg = "#2A3D52" if _is_base else _NH
                    _tbl += f'<th style="background:{_hbg};color:#fff;padding:8px 14px;font-size:11px;font-weight:600;text-align:center;border:1px solid {_BRD};">{_en}</th>'
                _tbl += '</tr></thead><tbody>'

                _metricas = [
                    ("Precio venta m²",    lambda e: f"${e[1]:,}",                                    None,  None),
                    ("Costo construcción", lambda e: f"${e[2]:,}/m²",                                 None,  None),
                    ("Terreno",            lambda e: f"${e[3]:,.0f}",                                  None,  None),
                    (None, None, None, None),
                    ("Ingresos brutos",    lambda e: f"${e[4]['ingresos_brutos']:,.0f}",               None,  None),
                    ("Costo total",        lambda e: f"${e[4]['costo_total_con_financ']:,.0f}",        None,  None),
                    ("Utilidad neta",      lambda e: f"${e[4]['utilidad_neta']:,.0f}",                 None,  None),
                    ("Margen neto",        lambda e: f"{e[4]['margen_pct']:.1f}%",                    "margen_pct",    (20, 12)),
                    ("TIR anual",          lambda e: f"{e[4]['tir_anual_pct']:.1f}%",                 "tir_anual_pct", (15, 10)),
                    ("ROI",                lambda e: f"{e[4]['roi_pct']:.1f}%",                       "roi_pct",       (20, 12)),
                    ("Break-even m²",      lambda e: f"${e[4]['be_precio_m2']:,}",                    None,  None),
                    ("TIT (terreno/ingr.)",lambda e: f"{e[4]['tit_pct']:.1f}%",                       None,  None),
                ]
                _lbl_sty = f'style="background:#F0EDE8;color:{_NH};padding:7px 14px;font-size:11px;font-weight:700;border:1px solid {_BRD};"'
                _sep_sty = f'colspan="{len(_esc_res)+1}" style="background:#E8E4DC;height:1px;padding:0;border:1px solid {_BRD};"'
                for _row in _metricas:
                    _rl, _rfn, _rfield, _rthresh = _row
                    if _rl is None:
                        _tbl += f'<tr><td {_sep_sty}></td></tr>'
                        continue
                    _tbl += f'<tr><td {_lbl_sty}>{_rl}</td>'
                    for _e in _esc_res:
                        _vs = _rfn(_e)
                        if _rfield and _rthresh:
                            _raw = float(_e[4].get(_rfield, 0))
                            _bg, _fg = _esc_bg(_raw, _rthresh[0], _rthresh[1])
                        else:
                            _bg, _fg = "#FFFFFF", _NH
                        _tbl += (f'<td style="background:{_bg};color:{_fg};padding:7px 14px;'
                                 f'font-size:12px;font-weight:700;text-align:center;border:1px solid {_BRD};">{_vs}</td>')
                    _tbl += '</tr>'
                _tbl += '</tbody></table>'
                st.markdown(_tbl, unsafe_allow_html=True)
                st.caption("Margen/TIR/ROI: Verde = óptimo · Amarillo = aceptable · Rojo = revisar")

                # ── Botones de descarga PDF + Excel ──────────
                st.markdown("---")
                _nombre_dl = st.session_state.get("nombre_proyecto") or zona_sel
                # Cache key: regenera solo cuando cambia el análisis
                _dl_key = f"{zona_sel}|{r.get('margen_pct',0):.2f}|{r.get('utilidad_neta',0):.0f}|{fin_run.get('precio_venta_m2',0)}"
                if st.session_state.get("_dl_key") != _dl_key:
                    st.session_state["_dl_key"]  = _dl_key
                    st.session_state["_pdf_dl"]  = None
                    st.session_state["_xl_dl"]   = None

                # Generar PDF si no está en caché
                if st.session_state.get("_pdf_dl") is None:
                    try:
                        st.session_state["_pdf_dl"] = generar_pdf_factis(
                            result=st.session_state.financ,
                            cabida=st.session_state.cabida,
                            params=st.session_state.params,
                            fin_inputs=fin_run,
                            zona=zona_sel,
                            legal=st.session_state.get("legal"),
                        )
                    except Exception as _pdf_err:
                        st.session_state["_pdf_dl"] = None

                # Generar Excel si no está en caché
                if st.session_state.get("_xl_dl") is None:
                    try:
                        st.session_state["_xl_dl"] = generar_excel_factis(
                            result=st.session_state.financ,
                            cabida=st.session_state.cabida,
                            params=st.session_state.params,
                            fin_inputs=fin_run,
                            zona=zona_sel,
                        )
                    except Exception as _xl_err:
                        st.session_state["_xl_dl"] = None

                _dl_c1, _dl_c2 = st.columns(2)
                with _dl_c1:
                    if st.session_state.get("_pdf_dl"):
                        st.download_button(
                            label="⬇ DESCARGAR PDF",
                            data=st.session_state["_pdf_dl"],
                            file_name=f"Reporte Financiero - {_nombre_dl}.pdf",
                            mime="application/octet-stream",
                            use_container_width=True,
                            type="primary",
                            key="btn_pdf_dl",
                        )
                    else:
                        st.button("⬇ DESCARGAR PDF", use_container_width=True, type="primary",
                                  key="btn_pdf_na", disabled=True)
                with _dl_c2:
                    if st.session_state.get("_xl_dl"):
                        st.download_button(
                            label="⬇ DESCARGAR EXCEL",
                            data=st.session_state["_xl_dl"],
                            file_name=f"Reporte Financiero - {_nombre_dl}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="btn_xl_dl",
                        )
                    else:
                        st.button("⬇ DESCARGAR EXCEL", use_container_width=True,
                                  key="btn_xl_na", disabled=True)

        # ── TAB 4: FLUJO DE CAJA ─────────────────────────
        with tabs[3]:
            if not c or not st.session_state.financ:
                st.markdown('<div class="alert-legal">Genera el análisis primero.</div>', unsafe_allow_html=True)
            else:
                fi        = st.session_state.financ_inputs or {}
                result_fl = st.session_state.financ
                fin_fl    = fi

                with st.spinner("Calculando DCF mensual…"):
                    df_fl, _, _, _, _, esc = generar_flujo(c, result_fl, fin_fl, zona_sel)

                sb  = esc["sin_banco"]
                cb  = esc["con_banco"]
                NAV_FL  = "#1E2D3D"
                BORD_FL = "#D8D4CC"
                ALT_FL  = "#F2EFE9"
                GRN     = "#1A5C32"
                RED     = "#C0392B"
                meses   = df_fl["Mes"].tolist()

                # ── Cronograma Gantt ──────────────────────────────
                st.markdown('<div class="section-title">Cronograma del Proyecto</div>', unsafe_allow_html=True)
                _r_fl    = (result_fl or {}).get("resumen", {})
                _mo      = _r_fl.get("meses_obra", 16)
                _mv      = _r_fl.get("meses_venta", 12)
                # Preventa: recalcular con la misma fórmula que generar_flujo
                import math as _math_g
                _n_unid_g   = max(c.get("total_unidades", 1) or 1, 1)
                _vel_g      = float((MERCADO.get(zona_sel, {}) or {}).get("velocidad_absorcion", 1.5) or 1.5)
                _pct_pv_g   = fi.get("pct_preventa_banco", 30.0) / 100
                _unid_req_g = max(1, _math_g.ceil(_n_unid_g * _pct_pv_g))
                _pv_auto_g  = max(1, _math_g.ceil(_unid_req_g / _vel_g))
                _preventa_m = int(fi.get("meses_preventa_override") or _pv_auto_g)
                _preventa_m = max(1, min(_preventa_m, 36))
                _pct_pv_lbl = fi.get("pct_preventa_banco", 30.0)
                _inicio_o   = _preventa_m
                _fin_o      = _inicio_o + _mo
                _post_obra  = min(6, max(0, _mv - _mo))   # ventas que quedan tras fin de obra
                _fin_v      = _fin_o + _post_obra
                _total_g    = max(_fin_v, _fin_o + 3)      # entrega siempre 3 m post-obra

                import plotly.express as px
                _hoy    = datetime.date.today().replace(day=1)

                def _m2d(m):
                    yr  = _hoy.year + (_hoy.month + m - 1) // 12
                    mo_ = (_hoy.month + m - 1) % 12 + 1
                    return datetime.date(yr, mo_, 1).isoformat()

                # Gantt con eje en meses relativos (mes 0 = inicio proyecto)
                _g_phases = [
                    ("Due Diligence",        0,           2,               "Gestión"),
                    ("Permisos / Licencias", 0,           3,               "Gestión"),
                    (f"Preventa ({_pct_pv_lbl:.0f}% obj.)", 0, _preventa_m,  "Comercial"),
                    ("Construcción",         _inicio_o,   _fin_o,          "Obra"),
                    ("Ventas activas",       _preventa_m, _fin_v,          "Comercial"),
                    ("Entrega y cierre",     _fin_o,      _fin_o + 3,      "Entrega"),
                ]
                # Alerta: unidades restantes post-preventa vs window de obra+post-obra
                _unid_post_pv   = _n_unid_g * (1 - _pct_pv_g)  # unidades a vender tras preventa
                _mv_post_pv     = (_math_g.ceil(_unid_post_pv / _vel_g) if _vel_g > 0 else 0)
                _mv_calc        = _r_fl.get("meses_venta_calc", _mv)
                if _mv_post_pv > _mo + 6:
                    st.markdown(
                        f'<div class="alert-gold">⚠ Tras la preventa quedan {_unid_post_pv:.0f} unidades '
                        f'que al ritmo actual ({_vel_g:.1f} und/mes) tardarían {_mv_post_pv} meses en colocarse, '
                        f'superando el window de obra + post-obra ({_mo} m + 6 m = {_mo+6} m). '
                        f'Revisar velocidad de ventas o ajustar precio/mix de tipologías.</div>',
                        unsafe_allow_html=True
                    )
                _color_map = {"Gestión":"#4A90C4","Obra":"#B8904A","Comercial":"#2E7D32","Entrega":"#7A5500"}
                _fig_g_rows = []
                for _gf, _gs, _ge, _get in _g_phases:
                    _dur = _ge - _gs
                    _fig_g_rows.append({
                        "Fase": _gf, "Inicio": _gs, "Fin": _ge,
                        "Etapa": _get,
                        "Duración": f"{_dur} mes{'es' if _dur != 1 else ''}",
                        "Centro": (_gs + _ge) / 2,
                    })

                _NAV_G = "#0A1628"; _PLT_G = "#0E1E2E"
                _seen_gantt_legends = set()
                import plotly.graph_objects as _go_gantt
                _fig_g = _go_gantt.Figure()
                for _row in reversed(_fig_g_rows):
                    _show_leg = _row["Etapa"] not in _seen_gantt_legends
                    _seen_gantt_legends.add(_row["Etapa"])
                    _fig_g.add_trace(_go_gantt.Bar(
                        x=[_row["Fin"] - _row["Inicio"]],
                        y=[_row["Fase"]],
                        base=[_row["Inicio"]],
                        orientation="h",
                        marker_color=_color_map[_row["Etapa"]],
                        name=_row["Etapa"],
                        text=[_row["Duración"]],
                        textposition="inside",
                        insidetextanchor="middle",
                        textfont=dict(color="white", size=11, family="Inter"),
                        hovertemplate=(
                            f"<b>{_row['Fase']}</b><br>"
                            f"Inicio: mes {_row['Inicio']}<br>"
                            f"Fin: mes {_row['Fin']}<br>"
                            f"Duración: {_row['Duración']}<extra></extra>"
                        ),
                        showlegend=_show_leg,
                    ))

                # Línea vertical en mes 0
                _fig_g.add_vline(x=0, line_color="#B8904A", line_width=1.5, line_dash="dot")
                _fig_g.update_layout(
                    barmode="overlay",
                    paper_bgcolor=_NAV_G,
                    plot_bgcolor=_PLT_G,
                    font=dict(color="white", size=10, family="Inter"),
                    margin=dict(l=0, r=20, t=10, b=30),
                    height=280,
                    xaxis=dict(
                        title="Mes del proyecto",
                        tickmode="linear", tick0=0, dtick=3,
                        gridcolor="#1E2D3D", gridwidth=1,
                        tickfont=dict(size=10),
                        range=[-0.5, _total_g + 0.5],
                    ),
                    yaxis=dict(showgrid=False, tickfont=dict(size=11)),
                    legend=dict(
                        orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                        font=dict(size=10), bgcolor="rgba(0,0,0,0)",
                        itemsizing="constant",
                    ),
                    showlegend=True,
                )
                # Desduplicar leyenda
                _seen_etapas = set()
                for _tr in _fig_g.data:
                    if _tr.name in _seen_etapas:
                        _tr.showlegend = False
                    else:
                        _seen_etapas.add(_tr.name)
                        _tr.showlegend = True
                        _tr.marker.color = _color_map.get(_tr.name, "#888")
                st.plotly_chart(_fig_g, use_container_width=True)

                _gc1, _gc2, _gc3, _gc4, _gc5 = st.columns(5)
                _gc1.metric("Preventa",        f"{_preventa_m} meses")
                _gc2.metric("Obra",            f"{_mo} meses")
                _gc3.metric("Plazo ventas",    f"{_mv} meses")
                _gc4.metric("Post-obra",       f"{_post_obra} meses")
                _gc5.metric("Duración total",  f"{_total_g} meses")
                st.markdown("---")

                # ── Bloque comparativo escenarios ──────────
                st.markdown('<div class="section-title">Escenarios — Sin Banco vs Con Banco</div>',
                            unsafe_allow_html=True)

                _be_sin = f"Mes {sb['mes_be']}" if sb["mes_be"] else "—"
                _be_con = f"Mes {cb['mes_be']}" if cb["mes_be"] else "—"
                _tir_sin = f"{sb['tir']}%" if sb["tir"] is not None else "—"
                _tir_con = f"{cb['tir']}%" if cb["tir"] is not None else "—"
                _exp_sin = fmt_usd(abs(sb["max_exp"]))
                _exp_con = fmt_usd(abs(cb["max_exp"]))
                _delta_tir = (
                    f"+{cb['tir'] - sb['tir']:.1f}pp apalancamiento"
                    if (cb["tir"] is not None and sb["tir"] is not None and cb["tir"] > sb["tir"])
                    else ""
                )
                # Pre-build span to avoid nested f-string quote conflict (Python < 3.12)
                _delta_span = (
                    '<span style="font-size:10px;color:#B8904A;">(' + _delta_tir + ")</span>"
                    if _delta_tir else ""
                )

                _td  = f'padding:7px 14px;font-size:12px;border:1px solid {BORD_FL};'
                _interes_str  = fmt_usd(cb["interes_total"])
                _igv_str      = fmt_usd(cb["igv_credito"])
                _horizonte    = len(df_fl) - 1
                _cmp_html = (
                    '<table style="border-collapse:collapse;width:100%;margin-bottom:16px;"><thead><tr>'
                    + f'<th style="background:{NAV_FL};color:#fff;padding:8px 14px;font-size:11px;font-weight:600;text-align:left;border:1px solid {BORD_FL};">Métrica</th>'
                    + f'<th style="background:{NAV_FL};color:#fff;padding:8px 14px;font-size:11px;font-weight:600;text-align:right;border:1px solid {BORD_FL};">Sin Banco (100% Equity)</th>'
                    + f'<th style="background:{NAV_FL};color:#fff;padding:8px 14px;font-size:11px;font-weight:600;text-align:right;border:1px solid {BORD_FL};">Con Banco (Linea 40% obra)</th>'
                    + '</tr></thead><tbody>'
                    + f'<tr style="background:{ALT_FL};">'
                    + f'<td style="{_td}">TIR Equity anual</td>'
                    + f'<td style="{_td}font-weight:700;text-align:right;color:{GRN};">{_tir_sin}</td>'
                    + f'<td style="{_td}font-weight:700;text-align:right;color:{GRN};">{_tir_con} {_delta_span}</td>'
                    + '</tr>'
                    + '<tr>'
                    + f'<td style="{_td}">Maxima exposicion equity</td>'
                    + f'<td style="{_td}text-align:right;color:{RED};">{_exp_sin}</td>'
                    + f'<td style="{_td}text-align:right;color:{RED};">{_exp_con}</td>'
                    + '</tr>'
                    + f'<tr style="background:{ALT_FL};">'
                    + f'<td style="{_td}">Breakeven (flujo acum.)</td>'
                    + f'<td style="{_td}text-align:right;">{_be_sin}</td>'
                    + f'<td style="{_td}text-align:right;">{_be_con}</td>'
                    + '</tr>'
                    + '<tr>'
                    + f'<td style="{_td}">Gasto financiero banco</td>'
                    + f'<td style="{_td}text-align:right;">--</td>'
                    + f'<td style="{_td}text-align:right;">{_interes_str}</td>'
                    + '</tr>'
                    + f'<tr style="background:{ALT_FL};">'
                    + f'<td style="{_td}">IGV credito fiscal (recuperado)</td>'
                    + f'<td style="{_td}text-align:right;color:{GRN};">{_igv_str}</td>'
                    + f'<td style="{_td}text-align:right;color:{GRN};">{_igv_str}</td>'
                    + '</tr>'
                    + '<tr>'
                    + f'<td style="{_td}">Horizonte del modelo</td>'
                    + f'<td style="{_td}text-align:right;" colspan="2">{_horizonte} meses</td>'
                    + '</tr>'
                    + '</tbody></table>'
                )
                st.markdown(_cmp_html, unsafe_allow_html=True)

                # ── Gráfico: Curva S construcción ──────────
                st.markdown('<div class="section-title">Curva S — Desembolso Construcción</div>',
                            unsafe_allow_html=True)
                obra_m = cb["obra_mensual"]
                obra_cumsum = []
                _s = 0
                total_obra = sum(obra_m) or 1
                for v in obra_m:
                    _s += v
                    obra_cumsum.append(_s / total_obra * 100)
                obra_x = list(range(3, 3 + len(obra_m)))

                fig_s = go.Figure()
                fig_s.add_trace(go.Bar(
                    x=obra_x, y=[v / 1000 for v in obra_m],
                    name="Desembolso mensual ($K)",
                    marker_color="rgba(184,144,74,0.65)",
                    hovertemplate="Mes %{x}<br>Desembolso: $%{y:,.0f}K<extra></extra>",
                    yaxis="y",
                ))
                fig_s.add_trace(go.Scatter(
                    x=obra_x, y=obra_cumsum,
                    name="Avance acumulado (%)",
                    mode="lines+markers",
                    line=dict(color="#1E2D3D", width=2.5),
                    marker=dict(size=4),
                    hovertemplate="Mes %{x}<br>Avance: %{y:.1f}%<extra></extra>",
                    yaxis="y2",
                ))
                fig_s.update_layout(
                    height=300,
                    barmode="overlay",
                    yaxis=dict(title="Desembolso ($K)", tickformat="$,.0f",
                               showgrid=True, gridcolor="#E8E3DA",
                               tickfont=dict(color="#4A5568", size=10)),
                    yaxis2=dict(title="% Avance acumulado", overlaying="y", side="right",
                                range=[0, 105], ticksuffix="%",
                                tickfont=dict(color="#4A5568", size=10),
                                showgrid=False),
                    xaxis=dict(title="Mes de proyecto", tickfont=dict(color="#4A5568", size=10), showgrid=False),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0,
                                font=dict(size=10, color="#4A5568")),
                    margin=dict(t=30, b=40, l=70, r=70),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#4A5568", size=11, family="Inter, sans-serif"),
                )
                st.plotly_chart(fig_s, use_container_width=True, config={"displayModeBar": False})

                # ── Gráfico: Flujo acumulado comparativo ───
                st.markdown('<div class="section-title">Flujo de Caja Acumulado — Comparativa</div>',
                            unsafe_allow_html=True)

                acum_sin = sb["acum"]
                acum_con = cb["acum"]

                fig_fl = go.Figure()
                fig_fl.add_trace(go.Scatter(
                    x=meses, y=acum_sin,
                    mode="lines", name="Sin banco",
                    line=dict(color="#1E2D3D", width=2.5),
                    hovertemplate="Mes %{x}<br>Sin banco: $%{y:,.0f}<extra></extra>",
                ))
                fig_fl.add_trace(go.Scatter(
                    x=meses, y=acum_con,
                    mode="lines", name="Con banco",
                    line=dict(color="#B8904A", width=2, dash="dash"),
                    hovertemplate="Mes %{x}<br>Con banco: $%{y:,.0f}<extra></extra>",
                ))
                fig_fl.add_hline(y=0, line_color="#AAAAAA", line_width=1.2)
                if sb["mes_be"]:
                    fig_fl.add_vline(x=sb["mes_be"], line_dash="dot", line_color="#1E2D3D",
                                     line_width=1.2,
                                     annotation_text=f" BE sin banco — mes {sb['mes_be']}",
                                     annotation_font=dict(size=10, color="#1E2D3D"),
                                     annotation_position="top left")
                if cb["mes_be"] and cb["mes_be"] != sb["mes_be"]:
                    fig_fl.add_vline(x=cb["mes_be"], line_dash="dot", line_color="#B8904A",
                                     line_width=1.2,
                                     annotation_text=f" BE con banco — mes {cb['mes_be']}",
                                     annotation_font=dict(size=10, color="#B8904A"),
                                     annotation_position="top right")
                fig_fl.update_layout(
                    height=360,
                    yaxis=dict(tickformat="$,.0f", zeroline=False,
                               showgrid=True, gridcolor="#E8E3DA",
                               title=dict(text="USD acumulado", font=dict(color="#4A5568", size=11)),
                               tickfont=dict(color="#4A5568", size=10)),
                    xaxis=dict(title=dict(text="Mes", font=dict(color="#4A5568", size=11)),
                               showgrid=False, tickfont=dict(color="#4A5568", size=10)),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0,
                                font=dict(size=11, color="#4A5568")),
                    margin=dict(t=30, b=40, l=80, r=20),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#4A5568", size=11, family="Inter, sans-serif"),
                )
                st.plotly_chart(fig_fl, use_container_width=True, config={"displayModeBar": False})

                # ── Tabla mensual detallada ─────────────────
                st.markdown('<div class="section-title">Detalle Mensual</div>', unsafe_allow_html=True)
                _esc_sel = st.radio("Escenario:", ["Sin Banco", "Con Banco"],
                                    horizontal=True, label_visibility="collapsed")
                _flujo_col = "Flujo Sin Banco" if _esc_sel == "Sin Banco" else "Flujo Con Banco"
                _acum_col  = "Acum. Sin Banco"  if _esc_sel == "Sin Banco" else "Acum. Con Banco"

                tbl = ('<table style="border-collapse:collapse;width:100%;margin-bottom:12px;">'
                       '<thead><tr>'
                       + ''.join(
                           f'<th style="background:{NAV_FL};color:#fff;padding:8px 12px;font-size:11px;'
                           f'font-weight:600;text-align:{"right" if idx > 0 else "center"};border:1px solid {BORD_FL};">{h}</th>'
                           for idx, h in enumerate(["Mes", "Flujo Mensual", "Flujo Acumulado",
                                                    "Saldo Deuda" if _esc_sel == "Con Banco" else ""])
                           if h
                       )
                       + '</tr></thead><tbody>')
                for i, row in df_fl.iterrows():
                    bg = ALT_FL if i % 2 == 0 else "#FAFAF8"
                    fm = row[_flujo_col]
                    fa = row[_acum_col]
                    fc = RED if fm < 0 else GRN
                    ac = RED if fa < 0 else GRN
                    sd = row["Saldo Deuda"]
                    tbl += (f'<tr>'
                            f'<td style="background:{bg};padding:6px 12px;font-size:11px;text-align:center;border:1px solid {BORD_FL};color:{NAV_FL};font-weight:600;">{int(row["Mes"])}</td>'
                            f'<td style="background:{bg};padding:6px 12px;font-size:11px;text-align:right;border:1px solid {BORD_FL};color:{fc};font-weight:600;">${fm:,.0f}</td>'
                            f'<td style="background:{bg};padding:6px 12px;font-size:11px;text-align:right;border:1px solid {BORD_FL};color:{ac};font-weight:600;">${fa:,.0f}</td>'
                            + (f'<td style="background:{bg};padding:6px 12px;font-size:11px;text-align:right;border:1px solid {BORD_FL};color:{NAV_FL};">{"$"+f"{sd:,.0f}" if sd > 0 else "—"}</td>'
                               if _esc_sel == "Con Banco" else "")
                            + '</tr>')
                tbl += '</tbody></table>'
                st.markdown(tbl, unsafe_allow_html=True)

                try:
                    _dcf_xlsx = generar_dcf_excel(
                        df_fl, result_fl,
                        fin_fl, esc,
                        nombre_proyecto=st.session_state.get("nombre_proyecto", "") or
                                        p.get("ubicacion", "Proyecto") if isinstance(p, dict) else "Proyecto"
                    )
                    _dcf_nombre_proy = st.session_state.get("nombre_proyecto") or p.get("ubicacion", "Proyecto") if isinstance(p, dict) else "Proyecto"
                    _dcf_fname = f"Flujo de Caja - {_dcf_nombre_proy}.xlsx"
                    st.download_button(
                        "⬇ Descargar DCF completo (.xlsx)",
                        _dcf_xlsx, _dcf_fname,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as _exc_dcf:
                    csv = df_fl.to_csv(index=False).encode("utf-8")
                    st.download_button("Descargar DCF (.csv)", csv, "dcf_mensual.csv", "text/csv",
                                       use_container_width=True)
                    st.caption(f"Excel no disponible: {_exc_dcf}")

        # ── TAB 5: LEGAL ─────────────────────────────────
        with tabs[4]:
            st.markdown('<div class="section-title">Due Diligence Legal — Partida Registral &amp; PU/HR</div>',
                        unsafe_allow_html=True)

            tiene_partida = st.session_state.get("partida_bytes") is not None
            tiene_puhr    = st.session_state.get("puhr_bytes") is not None

            if not tiene_partida and not tiene_puhr:
                st.markdown("""
                <div class="alert-legal">
                    Para el análisis legal adjunta la <strong>Partida Registral</strong> y/o el <strong>PU/HR</strong>
                    en el panel izquierdo y luego ejecuta <strong>GENERAR ANÁLISIS</strong>.
                </div>""", unsafe_allow_html=True)
            else:
                docs_disponibles = []
                if tiene_partida: docs_disponibles.append("Partida Registral")
                if tiene_puhr:    docs_disponibles.append("PU/HR")
                st.caption(f"Documentos disponibles: {' · '.join(docs_disponibles)}")

                if st.button("ANALIZAR DOCUMENTOS LEGALES", use_container_width=True, type="primary"):
                    _p_bytes = st.session_state.get("partida_bytes")
                    _u_bytes = st.session_state.get("puhr_bytes")
                    _c_bytes = st.session_state.get("cert_bytes")
                    st.session_state.legal = _run_with_retry(
                        lambda _p=_p_bytes, _u=_u_bytes, _c=_c_bytes: analizar_legal(_p, _u, _c),
                        "Analizando documentos registrales y parámetros urbanísticos…",
                    )

                lg = st.session_state.legal
                if lg:
                    # ── Semáforo ─────────────────────────────
                    sem = lg.get("semaforo", "amarillo").lower()
                    sem_cfg = {
                        "verde":    ("#1A4731", "#E8F5EE", "SIN ALERTAS CRÍTICAS",   "El inmueble presenta un estado legal favorable para la adquisición."),
                        "amarillo": ("#7A4F1A", "#FFF8EE", "OBSERVACIONES MENORES",  "Existen observaciones que deben verificarse antes de proceder."),
                        "rojo":     ("#7A1A1A", "#FFF0F0", "ALERTAS CRÍTICAS",       "Se detectaron riesgos legales que requieren atención inmediata."),
                    }.get(sem, ("#1E2D3D", "#F5F2ED", "INDETERMINADO", ""))
                    sc, sbg, setiq, _ = sem_cfg

                    st.markdown(f"""
                    <div style="background:{sbg};border:1px solid {sc};border-left:4px solid {sc};
                                border-radius:8px;padding:20px 24px;margin-bottom:20px;
                                box-shadow:0 2px 10px rgba(0,0,0,0.06);">
                        <div style="font-size:9px;letter-spacing:3px;color:{sc};text-transform:uppercase;
                                    font-weight:700;opacity:0.7;margin-bottom:6px;">Estado Legal del Inmueble</div>
                        <div style="font-size:20px;font-weight:700;color:{sc};margin-bottom:10px;">{setiq}</div>
                        <div style="font-size:13px;color:{sc};opacity:0.85;line-height:1.6;">
                            {lg.get('resumen_legal','—')}
                        </div>
                    </div>""", unsafe_allow_html=True)

                    # ── Checklist de hallazgos ───────────────
                    hallazgos = lg.get("hallazgos") or []
                    completitud = lg.get("completitud") or {}
                    if hallazgos:
                        _cv = completitud.get("verificados", len(hallazgos))
                        _ct = completitud.get("total", 20)
                        _cnv = completitud.get("no_verificables") or []
                        _cv_pct = round(_cv / _ct * 100) if _ct else 0
                        _cv_color = "#1A4731" if _cv_pct >= 80 else ("#7A4F1A" if _cv_pct >= 60 else "#7A1A1A")
                        st.markdown(
                            f'<div style="display:flex;justify-content:space-between;align-items:center;'
                            f'background:#F7F5F1;border:1px solid #E4E0D8;border-radius:8px;'
                            f'padding:12px 20px;margin-bottom:16px;">'
                            f'<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:700;">Completitud del análisis</div>'
                            f'<div style="font-size:16px;font-weight:700;color:{_cv_color};">{_cv}/{_ct} puntos verificados</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                        _SEV_CFG = {
                            "verde":         ("🟢", "#1A4731", "#E8F5EE", "#C3E6CB"),
                            "amarillo":      ("🟡", "#7A4F1A", "#FFF8EE", "#FFE0A0"),
                            "rojo":          ("🔴", "#7A1A1A", "#FFF0F0", "#F5C6CB"),
                            "no_verificable":("⚪", "#6A6A6A", "#F5F5F5", "#D8D4CC"),
                        }
                        _CAT_LABEL = {
                            "partida":    ("PARTIDA REGISTRAL", "#1E2D3D"),
                            "parametros": ("PARÁMETROS URBANOS", "#4A6078"),
                            "tecnico":    ("TÉCNICO", "#2A5070"),
                            "legal":      ("LEGAL", "#1E2D3D"),
                            "cruce":      ("CRUCE", "#5A4020"),
                        }

                        # Group by category for display
                        _partida_h = [h for h in hallazgos if h.get("categoria") in ("partida", "legal")]
                        _params_h  = [h for h in hallazgos if h.get("categoria") in ("parametros", "tecnico")]
                        _other_h   = [h for h in hallazgos if h.get("categoria") not in ("partida", "legal", "parametros", "tecnico")]

                        def _render_hallazgo_list(h_list, section_label):
                            if not h_list:
                                return
                            st.markdown(f'<div class="section-title">{section_label}</div>', unsafe_allow_html=True)
                            for h in h_list:
                                sev = (h.get("severidad") or "no_verificable").lower()
                                icon, tcol, bg, brd = _SEV_CFG.get(sev, _SEV_CFG["no_verificable"])
                                cat = h.get("categoria", "")
                                cat_lbl, cat_col = _CAT_LABEL.get(cat, (cat.upper(), "#6A6A6A"))
                                num = h.get("numero", "")
                                punto = h.get("punto", "—")
                                hallazgo_txt = h.get("hallazgo", "—")
                                subsanacion = h.get("subsanacion")
                                sev_label = {"verde": "OK", "amarillo": "VERIFICAR", "rojo": "CRÍTICO", "no_verificable": "NO VERIFICABLE"}.get(sev, sev.upper())

                                sub_html = ""
                                if subsanacion:
                                    sub_html = (
                                        f'<div style="margin-top:10px;padding-top:10px;'
                                        f'border-top:1px solid {brd};">'
                                        f'<div style="font-size:9px;color:{tcol};letter-spacing:1.5px;'
                                        f'text-transform:uppercase;font-weight:700;margin-bottom:4px;">Procedimiento de subsanación</div>'
                                        f'<div style="font-size:12px;color:#2A3A4A;line-height:1.6;">{subsanacion}</div>'
                                        f'</div>'
                                    )

                                st.markdown(
                                    f'<div style="background:{bg};border:1px solid {brd};border-left:4px solid {tcol};'
                                    f'border-radius:6px;padding:14px 18px;margin-bottom:10px;">'
                                    f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px;">'
                                    f'<div style="font-size:9px;color:{tcol};letter-spacing:1.5px;text-transform:uppercase;font-weight:700;">'
                                    f'{icon} {num}. {punto}</div>'
                                    f'<div style="font-size:8px;font-weight:700;color:{tcol};background:rgba(0,0,0,0.07);'
                                    f'padding:2px 8px;border-radius:3px;white-space:nowrap;">{sev_label}</div>'
                                    f'</div>'
                                    f'<div style="font-size:12px;color:#1E2D3D;line-height:1.6;">{hallazgo_txt}</div>'
                                    f'{sub_html}'
                                    f'</div>',
                                    unsafe_allow_html=True
                                )

                        _render_hallazgo_list(_partida_h, "Análisis Registral — Partida SUNARP")
                        _render_hallazgo_list(_params_h,  "Análisis Urbanístico — Certificado de Parámetros")
                        if _other_h:
                            _render_hallazgo_list(_other_h, "Verificación Cruzada")

                        if _cnv:
                            with st.expander(f"Puntos no verificables ({len(_cnv)}) — documento no adjunto o ilegible"):
                                for nv in _cnv:
                                    st.markdown(f'<div style="font-size:12px;color:#9A9080;padding:3px 0;">⚪ {nv}</div>', unsafe_allow_html=True)
                    else:
                        # Fallback: legacy alertas display
                        alertas = lg.get("alertas", [])
                        if alertas:
                            st.markdown('<div class="section-title">Alertas</div>', unsafe_allow_html=True)
                            for al in alertas:
                                icon = "🔴" if sem == "rojo" else ("🟡" if sem == "amarillo" else "🟢")
                                st.markdown(f'<div class="alert-gold">{icon} {al}</div>', unsafe_allow_html=True)

                    # ── Verificación cruzada ─────────────────
                    st.markdown('<div class="section-title">Verificación Cruzada — Documentos</div>', unsafe_allow_html=True)

                    def _check(ok):
                        if ok is True:  return "✓", "#1A4731"
                        if ok is False: return "✗", "#7A1A1A"
                        return "—", "#9A9080"

                    def _vcard(label, row1_lbl, row1_val, row2_lbl, row2_val, note, icon, col):
                        note_html = ('<div style="font-size:11px;color:#7A4F1A;margin-top:8px;font-style:italic;">' + note + '</div>') if note else ''
                        return (
                            '<div style="background:#FFFFFF;border:1px solid #E4E0D8;border-radius:6px;padding:16px 20px;margin-bottom:10px;box-shadow:0 1px 4px rgba(30,45,61,0.05);">'
                            '<div style="display:flex;justify-content:space-between;align-items:flex-start;">'
                            '<div style="flex:1;">'
                            '<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;margin-bottom:8px;">' + label + '</div>'
                            '<div style="font-size:12px;color:#1E2D3D;margin-bottom:4px;"><strong>' + row1_lbl + ':</strong> ' + str(row1_val) + '</div>'
                            '<div style="font-size:12px;color:#1E2D3D;"><strong>' + row2_lbl + ':</strong> ' + str(row2_val) + '</div>'
                            + note_html
                            + '</div>'
                            '<div style="font-size:28px;color:' + col + ';font-weight:700;margin-left:16px;">' + icon + '</div>'
                            '</div></div>'
                        )

                    # Propietarios con DNI
                    prop_p_raw = lg.get("propietarios_partida") or []
                    prop_h_raw = lg.get("propietarios_puhr") or []

                    def _fmt_propietario(p, fuente):
                        if not p or not isinstance(p, dict):
                            return str(p) if p else "—"
                        nombre = p.get("nombre") or "—"
                        dni    = p.get("dni")
                        pct    = p.get("porcentaje")
                        cond   = p.get("condicion")
                        tipo   = p.get("tipo_doc", "DNI")
                        parts  = [nombre]
                        if pct:  parts.append(f"({pct})")
                        if cond: parts.append(f"[{cond}]")
                        result = " ".join(parts)
                        if dni:
                            result += f' &nbsp;<span style="background:#E8F5EE;color:#1A4731;font-size:10px;font-weight:700;padding:2px 6px;border-radius:3px;border:1px solid #1A4731;">{tipo}: {dni}</span>'
                        else:
                            result += ' &nbsp;<span style="background:#FFF8EE;color:#7A4F1A;font-size:10px;padding:2px 6px;border-radius:3px;border:1px solid #B8904A;">DNI no encontrado</span>'
                        return result

                    if prop_p_raw and isinstance(prop_p_raw[0], dict):
                        prop_p_str = "<br>".join(_fmt_propietario(p, "Partida") for p in prop_p_raw)
                    else:
                        prop_p_str = ", ".join(str(x) for x in prop_p_raw) if prop_p_raw else "—"

                    if prop_h_raw and isinstance(prop_h_raw[0], dict):
                        prop_h_str = "<br>".join(_fmt_propietario(p, "PU/HR") for p in prop_h_raw)
                    else:
                        prop_h_str = ", ".join(str(x) for x in prop_h_raw if x) if prop_h_raw else "—"

                    icon_p, col_p = _check(lg.get("propietarios_coinciden"))
                    note_p = lg.get("diferencias_propietarios", "") or ""
                    note_p_html = ('<div style="font-size:11px;color:#7A4F1A;margin-top:8px;font-style:italic;">' + note_p + '</div>') if note_p else ""
                    st.markdown(
                        '<div style="background:#FFFFFF;border:1px solid #E4E0D8;border-radius:6px;padding:16px 20px;margin-bottom:10px;box-shadow:0 1px 4px rgba(30,45,61,0.05);">'
                        '<div style="display:flex;justify-content:space-between;align-items:flex-start;">'
                        '<div style="flex:1;">'
                        '<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;margin-bottom:8px;">Titularidad / Propietarios</div>'
                        '<div style="font-size:12px;color:#1E2D3D;margin-bottom:6px;"><strong>Partida:</strong><br>' + prop_p_str + '</div>'
                        '<div style="font-size:12px;color:#1E2D3D;"><strong>PU/HR:</strong><br>' + prop_h_str + '</div>'
                        + note_p_html +
                        '</div>'
                        '<div style="font-size:28px;color:' + col_p + ';font-weight:700;margin-left:16px;">' + icon_p + '</div>'
                        '</div></div>',
                        unsafe_allow_html=True
                    )

                    # Dirección
                    dir_p = lg.get("direccion_partida", "—") or "—"
                    dir_h = lg.get("direccion_puhr", "—") or "—"
                    icon_d, col_d = _check(lg.get("direcciones_coinciden"))
                    st.markdown(_vcard(
                        "Ubicación / Dirección",
                        "Partida", dir_p,
                        "PU/HR",   dir_h,
                        lg.get("diferencias_direccion", ""),
                        icon_d, col_d,
                    ), unsafe_allow_html=True)

                    # Área
                    area_r = lg.get("area_registral_m2")
                    area_h_v = lg.get("area_puhr_m2")
                    disc   = lg.get("discrepancia_area_m2")
                    icon_a, col_a = _check(lg.get("areas_coinciden"))
                    disc_note = ("Discrepancia: <strong>" + f"{disc:+.2f} m²</strong>") if disc else ""
                    st.markdown(_vcard(
                        "Área del Inmueble",
                        "Partida", f"{area_r:,.2f} m²" if area_r else "—",
                        "PU/HR",   f"{area_h_v:,.2f} m²" if area_h_v else "—",
                        disc_note,
                        icon_a, col_a,
                    ), unsafe_allow_html=True)

                    # ── Datos PU/HR municipales ───────────────
                    _autoavaluo   = lg.get("valor_autoavaluo")
                    _moneda_av    = lg.get("moneda_autoavaluo") or "PEN"
                    _anio_av      = lg.get("anio_autoavaluo")
                    _clasif_muni  = lg.get("clasificacion_municipal")
                    _cond_sat     = lg.get("condicion_propietario_sat")
                    _uso_predio   = lg.get("uso_predio")
                    if any([_autoavaluo, _clasif_muni, _cond_sat, _uso_predio]):
                        st.markdown('<div class="section-title">Datos Municipales — PU/HR</div>', unsafe_allow_html=True)
                        _mu_cols = st.columns(2)
                        if _autoavaluo:
                            _av_label = f"Autoavalúo {_anio_av}" if _anio_av else "Autoavalúo"
                            _av_val   = f"{_moneda_av} {_autoavaluo:,.0f}" if isinstance(_autoavaluo, (int, float)) else f"{_moneda_av} {_autoavaluo}"
                            _mu_cols[0].metric(_av_label, _av_val)
                        if _clasif_muni:
                            _mu_cols[1].metric("Clasificación Municipal", _clasif_muni)
                        if _cond_sat or _uso_predio:
                            _mu_cols2 = st.columns(2)
                            if _cond_sat:   _mu_cols2[0].metric("Condición (SAT)", _cond_sat)
                            if _uso_predio: _mu_cols2[1].metric("Uso del Predio", _uso_predio)

                    # ── Cargas e Hipotecas ────────────────────
                    hipotecas = lg.get("hipotecas_vigentes", []) or []
                    cargas    = lg.get("cargas_vigentes", []) or []
                    medidas   = lg.get("medidas_cautelares", []) or []
                    anotac    = lg.get("anotaciones_diversas", []) or []

                    if hipotecas or cargas or medidas:
                        st.markdown('<div class="section-title">Cargas, Gravámenes e Hipotecas</div>',
                                    unsafe_allow_html=True)

                        for h in hipotecas:
                            estado = h.get("estado", "vigente")
                            bg = "#FFF0F0" if estado == "vigente" else "#F7F5F1"
                            bl = "#7A1A1A" if estado == "vigente" else "#8A9BAD"
                            st.markdown(f"""
                            <div style="background:{bg};border-left:3px solid {bl};border-radius:0 6px 6px 0;
                                        padding:12px 16px;margin-bottom:8px;">
                                <div style="font-size:10px;color:{bl};letter-spacing:1.5px;text-transform:uppercase;
                                            font-weight:700;margin-bottom:4px;">Hipoteca — {estado.upper()}</div>
                                <div style="font-size:12px;color:#1E2D3D;">
                                    <strong>Acreedor:</strong> {h.get('acreedor','—')} &nbsp;·&nbsp;
                                    <strong>Monto:</strong> {h.get('monto','—')} &nbsp;·&nbsp;
                                    <strong>Inscripción:</strong> {h.get('fecha_inscripcion','—')}
                                    {f" &nbsp;·&nbsp; <strong>Asiento:</strong> {h.get('asiento','')}" if h.get('asiento') else ""}
                                </div>
                            </div>""", unsafe_allow_html=True)

                        for cg in cargas:
                            st.markdown(f"""
                            <div style="background:#FFF8EE;border-left:3px solid #B8904A;border-radius:0 6px 6px 0;
                                        padding:12px 16px;margin-bottom:8px;">
                                <div style="font-size:10px;color:#7A4F1A;letter-spacing:1.5px;text-transform:uppercase;
                                            font-weight:700;margin-bottom:4px;">Carga — {cg.get('tipo','').upper()}</div>
                                <div style="font-size:12px;color:#1E2D3D;">
                                    {cg.get('descripcion','—')}
                                    {f" &nbsp;·&nbsp; <strong>Acreedor:</strong> {cg.get('acreedor','')}" if cg.get('acreedor') else ""}
                                    {f" &nbsp;·&nbsp; <strong>Asiento:</strong> {cg.get('asiento','')}" if cg.get('asiento') else ""}
                                </div>
                            </div>""", unsafe_allow_html=True)

                        for mc in medidas:
                            st.markdown(f"""
                            <div style="background:#FFF0F0;border-left:3px solid #7A1A1A;border-radius:0 6px 6px 0;
                                        padding:12px 16px;margin-bottom:8px;">
                                <div style="font-size:10px;color:#7A1A1A;letter-spacing:1.5px;text-transform:uppercase;
                                            font-weight:700;margin-bottom:4px;">Medida Cautelar — {mc.get('tipo','').upper()}</div>
                                <div style="font-size:12px;color:#1E2D3D;">
                                    {mc.get('descripcion','—')}
                                    {f" &nbsp;·&nbsp; Exp. {mc.get('expediente','')}" if mc.get('expediente') else ""}
                                    {f" &nbsp;·&nbsp; {mc.get('fecha','')}" if mc.get('fecha') else ""}
                                </div>
                            </div>""", unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="section-title">Cargas, Gravámenes e Hipotecas</div>',
                                    unsafe_allow_html=True)
                        st.markdown('<div class="alert-legal">Sin cargas, gravámenes ni hipotecas identificados en los documentos analizados.</div>',
                                    unsafe_allow_html=True)

                    # ── Datos registrales ─────────────────────
                    st.markdown('<div class="section-title">Identificación Registral</div>', unsafe_allow_html=True)
                    col1, col2 = st.columns(2)
                    col1.metric("N° Partida Registral", lg.get("partida_numero") or "—")
                    col2.metric("N° Predio / Contribuyente", lg.get("numero_predio") or "—")

                    if anotac:
                        st.markdown('<div class="section-title">Anotaciones Adicionales</div>', unsafe_allow_html=True)
                        for an in anotac:
                            st.markdown(f'<div class="alert-legal">• {an}</div>', unsafe_allow_html=True)

        # ── TAB 6: RESUMEN ───────────────────────────────
        with tabs[5]:
            # Score de viabilidad
            if c and st.session_state.financ:
                r = (st.session_state.financ or {}).get("resumen", {})
                pts, score_10, etiqueta, color_txt, color_bg, recomendacion, score_items = score_viabilidad(r)

                # Score bar helper
                def _bar(v, mx):
                    filled = round(v / mx * 5)
                    return "●" * filled + "○" * (5 - filled)

                items_html = "".join(
                    '<div style="display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px solid rgba(0,0,0,0.06);">'
                    + f'<span style="font-size:11px;color:{color_txt};font-weight:500;">{it[0]}</span>'
                    + f'<span style="font-size:11px;color:{color_txt};opacity:0.7;letter-spacing:1px;font-variant-numeric:tabular-nums;">{it[1]}</span>'
                    + f'<span style="font-size:10px;color:{color_txt};opacity:0.5;letter-spacing:2px;">{_bar(it[2], it[3])}</span>'
                    + '</div>'
                    for it in score_items
                )

                st.markdown(f"""
                <div class="score-card" style="background:{color_bg};border-color:{color_txt};
                            display:grid;grid-template-columns:1fr auto;gap:32px;text-align:left;">
                    <div>
                        <div style="font-size:9px;letter-spacing:3px;color:{color_txt};
                                    text-transform:uppercase;font-weight:700;opacity:0.6;margin-bottom:10px;">
                            Evaluación de la Oportunidad
                        </div>
                        <div style="font-size:22px;color:{color_txt};font-weight:700;
                                    letter-spacing:-0.3px;margin-bottom:10px;">
                            {etiqueta}
                        </div>
                        <div style="font-size:12px;color:{color_txt};opacity:0.75;
                                    line-height:1.7;margin-bottom:16px;">
                            {recomendacion}
                        </div>
                        {items_html}
                    </div>
                    <div style="display:flex;flex-direction:column;align-items:center;
                                justify-content:center;min-width:110px;">
                        <div style="font-size:56px;font-weight:700;color:{color_txt};
                                    letter-spacing:-2px;line-height:1;font-variant-numeric:tabular-nums;">
                            {score_10}
                        </div>
                        <div style="font-size:11px;color:{color_txt};opacity:0.5;
                                    letter-spacing:1px;margin-top:4px;">/ 10</div>
                        <div style="font-size:8px;color:{color_txt};opacity:0.4;
                                    letter-spacing:2px;text-transform:uppercase;margin-top:8px;">Score</div>
                    </div>
                </div>""", unsafe_allow_html=True)

            st.markdown('<div class="section-title">Resumen Ejecutivo</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                st.markdown('<div class="section-title" style="font-size:8px">Datos del Inmueble</div>', unsafe_allow_html=True)
                data_inm = [
                    ("Ubicación",          p.get("ubicacion", "—")),
                    ("Zonificación",       p.get("zonificacion", "—")),
                    ("Área del terreno",   f"{p.get('area_terreno_m2','—')} m²"),
                    ("Frente",             f"{p.get('frente_ml','—')} ml"),
                    ("Altura máxima",      f"{p.get('pisos_max','—')} pisos"),
                    ("Área libre mínima",  f"{p.get('area_libre_min_pct','—')}%"),
                    ("Certificado caduca", p.get("fecha_caducidad","—")),
                ]
                for lbl, val in data_inm:
                    row_item(lbl, str(val))

            with col2:
                if c:
                    at_total = c.get("area_techada_total_m2", 0)
                    av_total = c.get("area_vendible_m2", 0)
                    at_piso  = c.get("area_techada_piso_m2", 0)
                    area_t   = p.get("area_terreno_m2", 1) or 1
                    cos = round(at_piso / area_t * 100, 1) if area_t else 0
                    cus = round(at_total / area_t, 2) if area_t else 0

                    st.markdown('<div class="section-title" style="font-size:8px">Programa Arquitectónico</div>', unsafe_allow_html=True)
                    data_cab = [
                        ("Pisos",              c.get("num_pisos", "—")),
                        ("Sótanos",            c.get("num_sotanos", 0)),
                        ("M² construibles",    f"{at_total:,.0f} m²"),
                        ("M² vendibles",       f"{av_total:,.0f} m²"),
                        ("Eficiencia vendible",f"{round(av_total/at_total*100,1) if at_total else 0}%"),
                        ("Departamentos",      c.get("total_unidades", "—")),
                        ("Estacionamientos",   c.get("estac_total", "—")),
                        ("COS (ocupación)",    f"{cos}%"),
                        ("CUS (uso del suelo)",f"{cus}"),
                    ]
                    for lbl, val in data_cab:
                        row_item(lbl, str(val))

            if c and st.session_state.financ:
                r = (st.session_state.financ or {}).get("resumen", {})
                st.markdown('<div class="section-title">Resultados Financieros</div>', unsafe_allow_html=True)

                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Ingresos brutos",  fmt_usd(r.get("ingresos_brutos", 0)))
                col2.metric("Costo total",      fmt_usd(r.get("costo_total_sin_financ", 0)))
                col3.metric("Utilidad neta",    fmt_usd(r.get("utilidad_neta", 0)))
                col4.metric("Margen",           f"{r.get('margen_pct', 0):.1f}%")
                col5.metric("ROI / TIR",        f"{r.get('roi_pct', 0):.1f}% / {r.get('tir_anual_pct', 0):.1f}%")

                st.markdown('<div class="section-title">Referencia del Terreno</div>', unsafe_allow_html=True)
                col1, col2, col3 = st.columns(3)
                col1.metric("Precio pagado",              fmt_usd(r.get("costo_total_sin_financ", 0)))
                col2.metric("Precio máx. para 20% mg.",   fmt_usd(r.get("max_terreno_20pct", 0)))
                col3.metric("Break-even precio/m²",       f"${r.get('be_precio_m2', 0):,}")

                st.markdown('<div class="section-title">Con Financiamiento Bancario</div>', unsafe_allow_html=True)
                col1, col2, col3 = st.columns(3)
                col1.metric("Costo financiero",   fmt_usd(r.get("costo_financiero", 0)))
                col2.metric("Utilidad neta",      fmt_usd(r.get("utilidad_con_financ", 0)),
                            delta=f"{r.get('margen_con_financ_pct', 0):.1f}% margen")
                col3.metric("Duración proyecto",  f"{r.get('meses_proyecto', 0)} meses")

            beneficios_all = p.get("beneficios_normativos", [])
            if beneficios_all:
                st.markdown('<div class="section-title">Beneficios Normativos Aplicables</div>', unsafe_allow_html=True)
                for b in beneficios_all:
                    st.markdown(f'<div class="alert-legal">⚖️ <strong>{b.get("descripcion","")}</strong> — {b.get("impacto_estimado","")}</div>', unsafe_allow_html=True)

            st.markdown("---")

            # ── Descarga del informe completo ─────────────
            if c and st.session_state.financ:
                _nombre_proy_inf = (st.session_state.get("nombre_proyecto") or
                                    p.get("ubicacion", "Proyecto"))
                try:
                    _pdf_inf = generar_pdf_factis(
                        result      = st.session_state.financ,
                        cabida      = c,
                        params      = p,
                        fin_inputs  = st.session_state.get("financ_inputs") or {},
                        zona        = zona_sel,
                        legal       = st.session_state.get("legal"),
                    )
                    st.download_button(
                        label             = "Descargar Informe Completo (PDF)",
                        data              = _pdf_inf,
                        file_name         = f"Informe de Análisis Cabida - {_nombre_proy_inf}.pdf",
                        mime              = "application/pdf",
                        use_container_width = True,
                    )
                except Exception as _e_pdf_inf:
                    from datetime import date as _date
                    informe_html = generar_informe_html(
                        params=p, cabida=c, financ=st.session_state.financ,
                        legal=st.session_state.get("legal"), zona=zona_sel,
                        financ_inputs=st.session_state.get("financ_inputs"),
                        fecha=_date.today().strftime("%d/%m/%Y"),
                    )
                    st.download_button(
                        label="Descargar Informe Completo (HTML)",
                        data=informe_html.encode("utf-8"),
                        file_name=f"Informe de Análisis Cabida - {_nombre_proy_inf}.html",
                        mime="text/html",
                        use_container_width=True,
                    )

            st.markdown("""
<div style="border-top:1px solid #E8E0D4;margin-top:32px;padding-top:14px;">
  <div style="font-size:9px;color:#9A8A7A;line-height:1.7;text-align:center;">
    <span style="color:#B8904A;font-weight:700;letter-spacing:1px;">NOTA · </span>
    Esta IA de Análisis Inmobiliario debe utilizarse como herramienta complementaria al criterio profesional,
    permitiendo obtener resultados preliminares de manera rápida. Como paso final, el profesional podrá
    terminar de definir las tipologías, distribución por plantas y las modificaciones que considere pertinentes.
    La IA irá volviéndose más responsiva y alineada con la visión del profesional a medida que se retroalimenta
    con sus decisiones.
  </div>
</div>""", unsafe_allow_html=True)
            st.caption("Osterling Advisory — Inmobiliaria Corporativa | eosterling@grupoosterling.com | Lima, Perú")

        # ── TAB 7: PROPUESTA ─────────────────────────────
        with tabs[6]:
            st.markdown('<div class="section-title">Propuesta de Compra / Arrendamiento</div>',
                        unsafe_allow_html=True)

            _prop_precio     = st.session_state.get("prop_precio", 0)
            _prop_tipo       = st.session_state.get("prop_tipo", "Compra")
            _prop_propietario = st.session_state.get("prop_propietario", "")
            _prop_plazo      = st.session_state.get("prop_plazo", 10)
            _prop_cond       = st.session_state.get("prop_condiciones", "")

            if not p:
                st.markdown('<div class="alert-legal">Genera el análisis primero para completar los datos del inmueble en la propuesta.</div>',
                            unsafe_allow_html=True)
            elif _prop_precio == 0:
                _financ_ss = st.session_state.get("financ")
                if _financ_ss:
                    _rr = _financ_ss.get("resumen", {})
                    _v20p = _rr.get("max_terreno_20pct", 0)
                    _v15p = _rr.get("max_terreno_15pct", 0)
                    _v12p = _rr.get("max_terreno_12pct", 0)
                    _ct   = (st.session_state.get("financ_inputs") or {}).get("costo_terreno", 0) or 0
                    st.markdown(
                        '<div class="alert-legal">'
                        '<b>Precio ofertado no ingresado.</b> El modelo financiero sugiere los siguientes precios '
                        'máximos de adquisición del terreno según margen objetivo:<br><br>'
                        + (f'<b>Precio ingresado en modelo:</b> ${_ct:,}<br><br>' if _ct > 0 else '')
                        + f'• <b>${_v20p:,}</b> → margen neto ≥ 20% &nbsp;<span style="color:#1A4731;font-weight:700;">●</span> óptimo<br>'
                        + f'• <b>${_v15p:,}</b> → margen neto ≥ 15% &nbsp;<span style="color:#7A5500;font-weight:700;">●</span> aceptable<br>'
                        + f'• <b>${_v12p:,}</b> → margen neto ≥ 12% &nbsp;<span style="color:#7A1A1A;font-weight:700;">●</span> mínimo<br><br>'
                        + 'Ingresa el precio ofertado en el panel izquierdo → sección <strong>PROPUESTA</strong>.'
                        '</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown('<div class="alert-legal">Ingresa el precio ofertado en el panel izquierdo (sección <strong>PROPUESTA</strong>) para generar el documento.</div>',
                                unsafe_allow_html=True)
            else:
                from datetime import date as _date2
                _fecha_prop = _date2.today().strftime("%d de %B de %Y").replace(
                    "January","enero").replace("February","febrero").replace("March","marzo"
                    ).replace("April","abril").replace("May","mayo").replace("June","junio"
                    ).replace("July","julio").replace("August","agosto").replace("September","septiembre"
                    ).replace("October","octubre").replace("November","noviembre").replace("December","diciembre")

                _prop_html = generar_propuesta_html(
                    tipo           = _prop_tipo,
                    propietario    = _prop_propietario,
                    params         = p,
                    financ         = st.session_state.get("financ"),
                    legal          = st.session_state.get("legal"),
                    comps_sunarp   = st.session_state.get("comps_sunarp", []),
                    precio_oferta  = float(_prop_precio),
                    moneda_oferta  = "USD",
                    condiciones    = _prop_cond,
                    plazo_respuesta = int(_prop_plazo),
                    fecha          = _fecha_prop,
                    tiene_opcion        = st.session_state.get("prop_opcion", True),
                    dias_opcion         = int(st.session_state.get("prop_dias_opcion", 90)),
                    pct_opcion          = float(st.session_state.get("prop_pct_opcion", 0.0)),
                    pct_minuta          = float(st.session_state.get("prop_pct_minuta", 20.0)),
                    condicion_minuta    = st.session_state.get("prop_cond_minuta", "Aprobación del anteproyecto por la Municipalidad"),
                    condicion_escritura = st.session_state.get("prop_cond_escritura", "Desocupación y entrega del inmueble libre de cargas"),
                )

                # Preview (HTML)
                st.components.v1.html(_prop_html, height=820, scrolling=True)

                # PDF download
                try:
                    _prop_pdf = generar_propuesta_pdf(
                        tipo           = _prop_tipo,
                        propietario    = _prop_propietario,
                        params         = p,
                        financ         = st.session_state.get("financ"),
                        legal          = st.session_state.get("legal"),
                        comps_sunarp   = st.session_state.get("comps_sunarp", []),
                        precio_oferta  = float(_prop_precio),
                        moneda_oferta  = "USD",
                        condiciones    = _prop_cond,
                        plazo_respuesta = int(_prop_plazo),
                        fecha          = _fecha_prop,
                        tiene_opcion        = st.session_state.get("prop_opcion", True),
                        dias_opcion         = int(st.session_state.get("prop_dias_opcion", 90)),
                        pct_opcion          = float(st.session_state.get("prop_pct_opcion", 0.0)),
                        pct_minuta          = float(st.session_state.get("prop_pct_minuta", 20.0)),
                        condicion_minuta    = st.session_state.get("prop_cond_minuta", "Aprobación del anteproyecto por la Municipalidad"),
                        condicion_escritura = st.session_state.get("prop_cond_escritura", "Desocupación y entrega del inmueble libre de cargas"),
                    )
                    _nombre_prop = f"Propuesta_{_prop_tipo}_{(p.get('ubicacion') or 'Inmueble').replace(' ','_')[:35]}.pdf"
                    st.download_button(
                        label               = f"⬇ Descargar Propuesta de {_prop_tipo} (PDF)",
                        data                = _prop_pdf,
                        file_name           = _nombre_prop,
                        mime                = "application/pdf",
                        use_container_width = True,
                    )
                except Exception as _pdf_err:
                    st.markdown(f'<div class="alert-info">ℹ️ No se pudo generar el PDF: {_pdf_err}</div>', unsafe_allow_html=True)

        # ── TAB 8: Renta / Holding ────────────────────────────────
        with tabs[7]:
            st.markdown('<div class="section-title">Análisis de Renta y Holding</div>', unsafe_allow_html=True)
            st.caption("¿Conviene más vender o rentar? Compara la estrategia de venta contra mantener el activo y cobrar alquiler.")

            _fin_ok = st.session_state.get("financ")
            if not _fin_ok:
                st.info("Completa primero el análisis financiero (tab Financiero) para ver este análisis.")
            else:
                _hr = (_fin_ok or {}).get("resumen", {})
                _hraw = _fin_ok.get("_raw", {})
                _m_hold = MERCADO.get(zona_sel, {})
                _tc_hold = float((st.secrets.get("mercado") or {}).get("tipo_cambio", 3.45))

                # ── Parámetros de renta ────────────────────────────
                st.markdown("#### Parámetros de la estrategia holding")
                _hc1, _hc2, _hc3 = st.columns(3)
                with _hc1:
                    _h_alq_m2 = st.number_input(
                        "Alquiler mercado ($/m²/mes)",
                        min_value=1.0, max_value=50.0,
                        value=float(_m_hold.get("alquiler_m2_mes", 8.0)),
                        step=0.5, key="h_alq_m2")
                    _h_vac = st.number_input(
                        "Vacancia estimada (%)", 0.0, 30.0, 7.0, 1.0, key="h_vac")
                with _hc2:
                    _h_opex = st.number_input(
                        "Gastos operativos (%)", 0.0, 40.0, 22.0, 1.0, key="h_opex")
                    _h_apre = st.number_input(
                        "Apreciación anual (%)", -10.0, 15.0,
                        float(max(-10.0, min(15.0, _m_hold.get("variacion_anual_pct", 4.0)))),
                        0.5, key="h_apre")
                with _hc3:
                    _h_horizon = st.selectbox(
                        "Horizonte de análisis", [5, 7, 10, 15], index=1, key="h_horizon")
                    _h_wacc = st.number_input(
                        "Tasa descuento / WACC (%)", 5.0, 20.0, 9.0, 0.5, key="h_wacc")

                # ── Cálculos ───────────────────────────────────────
                _av_hold  = _hr.get("m2_vendibles", 0)
                _inv_hold = _hr.get("costo_total_sin_financ", 0)   # inversión total sin financiamiento
                _util_vta = _hr.get("utilidad_neta", 0)
                _ing_brut = _hr.get("ingresos_brutos", 0)

                # Renta bruta anual
                _rent_brut_anual = _av_hold * _h_alq_m2 * 12 * (1 - _h_vac / 100)
                # NOI (Net Operating Income)
                _noi          = _rent_brut_anual * (1 - _h_opex / 100)
                # Cap rate
                _cap_rate     = (_noi / _inv_hold * 100) if _inv_hold else 0
                # Yield bruto sobre inversión
                _yield_bruto  = (_rent_brut_anual / _inv_hold * 100) if _inv_hold else 0
                # Payback (años para recuperar inversión solo con renta)
                _payback_rent = (_inv_hold / _noi) if _noi > 0 else 999

                # Valor terminal del activo
                _val_terminal = _inv_hold * ((1 + _h_apre / 100) ** _h_horizon)
                # Ganancia de capital al año N
                _gan_capital  = _val_terminal - _inv_hold

                # NPV holding: suma NOI descontado + valor terminal descontado
                _wacc_d = _h_wacc / 100
                _npv_hold = sum(_noi / (1 + _wacc_d) ** t for t in range(1, _h_horizon + 1))
                _npv_hold += _val_terminal / (1 + _wacc_d) ** _h_horizon
                _npv_hold -= _inv_hold

                # NPV venta: utilidad neta ya realizada en año 1
                _npv_venta = _util_vta / (1 + _wacc_d) - _inv_hold

                # ── KPIs comparativos ──────────────────────────────
                st.markdown("---")
                _kh1, _kh2, _kh3, _kh4 = st.columns(4)
                def _kpi_hold(col, label, value, suffix="", delta=None):
                    col.metric(label, f"{value:,.1f}{suffix}" if isinstance(value, float) else f"{value:,}{suffix}", delta)

                _kh1.metric("Renta bruta anual", f"${_rent_brut_anual:,.0f}")
                _kh2.metric("NOI anual", f"${_noi:,.0f}")
                _kh3.metric("Cap Rate", f"{_cap_rate:.1f}%",
                            delta="Óptimo >7%" if _cap_rate >= 7 else "Bajo vs. Lima 6–9%")
                _kh4.metric("Yield bruto s/inv.", f"{_yield_bruto:.1f}%")

                _kh5, _kh6, _kh7, _kh8 = st.columns(4)
                _kh5.metric("Payback renta", f"{_payback_rent:.1f} años")
                _kh6.metric(f"Valor activo año {_h_horizon}", f"${_val_terminal:,.0f}")
                _kh7.metric("Ganancia de capital", f"${_gan_capital:,.0f}",
                            delta=f"+{_h_apre:.1f}%/año")
                _kh8.metric(f"NPV holding {_h_horizon}a", f"${_npv_hold:,.0f}")

                # ── Tabla comparativa Venta vs. Holding ───────────
                st.markdown("---")
                st.markdown("#### Venta inmediata vs. Holding")
                _GOLD_H = "#B8904A"
                _NAV_H  = "#0A1628"
                _BRD_H  = "#2A3D52"
                _comp_h = [
                    ("Inversión total",             f"${_inv_hold:,.0f}",      f"${_inv_hold:,.0f}"),
                    ("Retorno total",                f"${_util_vta:,.0f}",      f"${(_noi * _h_horizon + _gan_capital):,.0f}"),
                    ("Margen / yield total",         f"{_hr.get('margen_pct',0):.1f}%", f"{(_noi*_h_horizon+_gan_capital)/_inv_hold*100:.1f}%"),
                    ("NPV",                          f"${_npv_venta:,.0f}",     f"${_npv_hold:,.0f}"),
                    ("Flujo año 1",                  f"${_util_vta:,.0f}",      f"${_noi:,.0f}"),
                    (f"Flujo año {_h_horizon}",      "—",                       f"${_noi + _gan_capital:,.0f}"),
                    ("Liquidez",                     "Alta (se vende)",         "Baja (activo ilíquido)"),
                    ("Riesgo",                       "Bajo (ya ejecutado)",     "Mercado alquiler + vacancia"),
                ]
                _TH_BG = "#1E2D3D"
                _TH_TXT = "#FFFFFF"
                _TR_ODD = "#FFFFFF"
                _TR_EVN = "#F9F7F4"
                _TD_LBL = "#5A7A8A"
                _TD_VAL = "#1E2D3D"
                _BRD_H2 = "#E0DDD8"
                _tbl_h = (
                    f'<table style="border-collapse:collapse;width:100%;margin-top:10px;border-radius:8px;overflow:hidden;">'
                    f'<thead><tr>'
                    f'<th style="background:{_TH_BG};color:{_TH_TXT};padding:10px 16px;font-size:10px;font-weight:600;text-align:left;border:1px solid {_TH_BG};">MÉTRICA</th>'
                    f'<th style="background:{_TH_BG};color:{_GOLD_H};padding:10px 16px;font-size:10px;font-weight:700;text-align:center;border:1px solid {_TH_BG};">VENDER</th>'
                    f'<th style="background:{_TH_BG};color:{_GOLD_H};padding:10px 16px;font-size:10px;font-weight:700;text-align:center;border:1px solid {_TH_BG};">HOLDING {_h_horizon}A</th>'
                    f'</tr></thead><tbody>'
                )
                for _i_h, (_lh, _vv, _vh) in enumerate(_comp_h):
                    _row_bg = _TR_ODD if _i_h % 2 == 0 else _TR_EVN
                    _tbl_h += (
                        f'<tr>'
                        f'<td style="background:{_row_bg};color:{_TD_LBL};padding:9px 16px;font-size:11px;font-weight:600;border:1px solid {_BRD_H2};">{_lh}</td>'
                        f'<td style="background:{_row_bg};color:{_TD_VAL};padding:9px 16px;font-size:12px;font-weight:700;text-align:center;border:1px solid {_BRD_H2};">{_vv}</td>'
                        f'<td style="background:{_row_bg};color:{_TD_VAL};padding:9px 16px;font-size:12px;font-weight:700;text-align:center;border:1px solid {_BRD_H2};">{_vh}</td>'
                        f'</tr>'
                    )
                _tbl_h += '</tbody></table>'
                st.markdown(_tbl_h, unsafe_allow_html=True)

                # ── Recomendación ──────────────────────────────────
                st.markdown("---")
                if _npv_hold > _npv_venta and _cap_rate >= 6:
                    _rec_txt = f"El holding a {_h_horizon} años genera mayor NPV (${_npv_hold:,.0f} vs. ${_npv_venta:,.0f}). Con un cap rate de {_cap_rate:.1f}%, el activo produce renta competitiva mientras se aprecia."
                    _rec_color = "#1B5E20"
                    _rec_bg = "rgba(27,94,32,0.1)"
                elif _cap_rate >= 6:
                    _rec_txt = f"El cap rate de {_cap_rate:.1f}% es atractivo pero la venta genera NPV más alto en el horizonte analizado. Considera el holding si el objetivo es flujo de caja recurrente."
                    _rec_color = "#7A5500"
                    _rec_bg = "rgba(122,85,0,0.1)"
                else:
                    _rec_txt = f"El cap rate de {_cap_rate:.1f}% está por debajo del mínimo recomendado (6%). La estrategia de venta genera mejor retorno ajustado por riesgo."
                    _rec_color = "#B71C1C"
                    _rec_bg = "rgba(183,28,28,0.1)"
                st.markdown(
                    f'<div style="padding:14px 18px;background:{_rec_bg};'
                    f'border-left:4px solid {_rec_color};border-radius:0 8px 8px 0;margin-top:8px;">'
                    f'<span style="font-size:12px;font-weight:700;color:{_rec_color};">Recomendación: </span>'
                    f'<span style="font-size:12px;color:#1E2D3D;">{_rec_txt}</span>'
                    f'</div>', unsafe_allow_html=True)

                # ── Proyección anual ───────────────────────────────
                st.markdown("---")
                st.markdown('<div class="section-title">Proyección año a año</div>', unsafe_allow_html=True)
                _proj_rows = []
                _val_acum = _inv_hold
                for _yr in range(1, _h_horizon + 1):
                    _val_acum *= (1 + _h_apre / 100)
                    _noi_yr   = _noi
                    _cash_acum = _noi * _yr
                    _total_ret = _cash_acum + (_val_acum - _inv_hold)
                    _ret_pct   = _total_ret / _inv_hold * 100 if _inv_hold else 0
                    _proj_rows.append((_yr, _noi_yr, _cash_acum, _val_acum, _total_ret, _ret_pct))

                _proj_header = (
                    '<tr style="background:#1E2D3D;">'
                    '<th style="padding:8px 14px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;">Año</th>'
                    '<th style="padding:8px 14px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:right;">NOI anual</th>'
                    '<th style="padding:8px 14px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:right;">Renta acum.</th>'
                    '<th style="padding:8px 14px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:right;">Valor activo</th>'
                    '<th style="padding:8px 14px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:right;">Retorno total</th>'
                    '<th style="padding:8px 14px;color:#B8C8D8;font-size:10px;letter-spacing:1.2px;text-transform:uppercase;text-align:right;">Retorno %</th>'
                    '</tr>'
                )
                _proj_body = ""
                for _yr, _noi_yr, _ca, _va, _rt, _rp in _proj_rows:
                    _bg = "#F5F3EF" if _yr % 2 == 0 else "#FFFFFF"
                    _proj_body += (
                        f'<tr style="background:{_bg};">'
                        f'<td style="padding:7px 14px;color:#1E2D3D;font-weight:700;">Año {_yr}</td>'
                        f'<td style="padding:7px 14px;color:#1E2D3D;text-align:right;">${_noi_yr:,.0f}</td>'
                        f'<td style="padding:7px 14px;color:#1E2D3D;text-align:right;">${_ca:,.0f}</td>'
                        f'<td style="padding:7px 14px;color:#1E2D3D;text-align:right;">${_va:,.0f}</td>'
                        f'<td style="padding:7px 14px;color:#B8904A;text-align:right;font-weight:700;">${_rt:,.0f}</td>'
                        f'<td style="padding:7px 14px;color:#B8904A;text-align:right;font-weight:700;">{_rp:.0f}%</td>'
                        f'</tr>'
                    )
                st.markdown(
                    f'<table style="width:100%;border-collapse:collapse;background:#FFFFFF;'
                    f'border:1px solid #D8D4CC;border-radius:8px;overflow:hidden;'
                    f'font-family:Inter,sans-serif;font-size:12px;">'
                    f'<thead>{_proj_header}</thead>'
                    f'<tbody>{_proj_body}</tbody>'
                    f'</table>',
                    unsafe_allow_html=True
                )


    else:
        st.markdown(
            '<div style="border-radius:8px;min-height:460px;'
            'background:linear-gradient(160deg,#1A2737 0%,#1E2D3D 60%,#1A2737 100%);'
            'display:flex;align-items:center;justify-content:center;'
            'box-shadow:0 8px 32px rgba(30,45,61,0.18);padding:64px 48px;">'

            '<div style="display:grid;grid-template-columns:1fr 1px 1fr;gap:0;max-width:820px;width:100%;">'

            # Left: descriptor
            '<div style="padding-right:48px;display:flex;flex-direction:column;justify-content:center;gap:16px;">'
            '<div style="font-size:9px;color:#B8904A;letter-spacing:4px;text-transform:uppercase;'
            'font-weight:700;font-family:Inter,sans-serif;">Potenciado por IA</div>'
            '<div style="font-size:22px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;line-height:1.25;'
            'font-family:Inter,sans-serif;">IA de Análisis Inmobiliario</div>'
            '<div style="font-size:13px;color:#8AA8C0;line-height:1.7;font-family:Inter,sans-serif;">'
            'Integra en una sesión de trabajo; cabida arquitectónica, análisis financiero '
            'y due diligence legal. Optimizando tu tiempo en el análisis de Proyectos:'
            '</div>'
            '<div style="border-left:2px solid #B8904A;padding-left:14px;margin-top:4px;">'
            '<span style="font-size:13px;color:#FFFFFF;font-family:Inter,sans-serif;font-weight:500;">'
            'Multifamiliares, Logísticos e Industriales.'
            '</span>'
            '</div>'
            '<div style="font-size:11px;color:#B8904A;letter-spacing:1.5px;text-transform:uppercase;'
            'font-weight:600;font-family:Inter,sans-serif;">Acelerador de decisiones · Optimización de análisis</div>'
            '</div>'

            # Separator
            '<div style="background:rgba(184,144,74,0.3);"></div>'

            # Right: steps
            '<div style="padding-left:48px;display:flex;flex-direction:column;justify-content:center;gap:24px;">'
            '<div style="font-size:9px;color:#B8904A;letter-spacing:3px;text-transform:uppercase;'
            'font-weight:700;font-family:Inter,sans-serif;">Cómo comenzar</div>'

            '<div style="display:flex;align-items:flex-start;gap:16px;">'
            '<div style="min-width:24px;height:24px;border-radius:50%;background:rgba(184,144,74,0.2);'
            'border:1px solid #B8904A;display:flex;align-items:center;justify-content:center;'
            'font-size:11px;color:#B8904A;font-weight:700;">1</div>'
            '<div style="font-size:13px;color:#E8EDF2;line-height:1.65;font-family:Inter,sans-serif;">'
            'Adjunta el Certificado de Parámetros y documentos en el panel izquierdo'
            '</div></div>'

            '<div style="display:flex;align-items:flex-start;gap:16px;">'
            '<div style="min-width:24px;height:24px;border-radius:50%;background:rgba(184,144,74,0.2);'
            'border:1px solid #B8904A;display:flex;align-items:center;justify-content:center;'
            'font-size:11px;color:#B8904A;font-weight:700;">2</div>'
            '<div style="font-size:13px;color:#E8EDF2;line-height:1.65;font-family:Inter,sans-serif;">'
            'Configura la zona de mercado y los parámetros financieros del proyecto'
            '</div></div>'

            '<div style="display:flex;align-items:flex-start;gap:16px;">'
            '<div style="min-width:24px;height:24px;border-radius:50%;background:rgba(184,144,74,0.2);'
            'border:1px solid #B8904A;display:flex;align-items:center;justify-content:center;'
            'font-size:11px;color:#B8904A;font-weight:700;">3</div>'
            '<div style="font-size:13px;color:#E8EDF2;line-height:1.65;font-family:Inter,sans-serif;">'
            'Presiona GENERAR ANÁLISIS — cabida, financiero, legal y flujo de caja en segundos'
            '</div></div>'

            '</div>'  # end right
            '</div>'  # end grid
            '</div>',  # end hero
            unsafe_allow_html=True
        )

        # ── Capabilities bar ─────────────────────────────
        # icon: small gold monogram circle
        def _cap_icon(letter):
            return (
                '<div style="width:36px;height:36px;border-radius:50%;'
                'border:1.5px solid #B8904A;display:flex;align-items:center;'
                'justify-content:center;margin:0 auto 12px auto;">'
                f'<span style="font-size:13px;font-weight:700;color:#B8904A;'
                f'font-family:Inter,sans-serif;">{letter}</span>'
                '</div>'
            )
        def _cap(letter, title, desc):
            return (
                '<div style="display:flex;flex-direction:column;align-items:center;'
                'text-align:center;padding:28px 20px;">'
                + _cap_icon(letter)
                + f'<div style="font-size:12px;font-weight:700;color:#1E2D3D;letter-spacing:0.3px;'
                f'font-family:Inter,sans-serif;margin-bottom:6px;">{title}</div>'
                f'<div style="font-size:11px;color:#7A8A99;line-height:1.55;font-family:Inter,sans-serif;">{desc}</div>'
                '</div>'
            )
        cap_sep = '<div style="width:1px;background:#E0DAD0;margin:16px 0;align-self:stretch;"></div>'
        st.markdown(
            '<div style="background:#FAFAF8;border:1px solid #E8E3DA;border-radius:8px;'
            'margin-top:16px;display:grid;grid-template-columns:1fr 1px 1fr 1px 1fr 1px 1fr;'
            'align-items:stretch;">'
            + _cap("CA", "Cabida Arquitectónica", "Área techada, unidades, pisos y programa óptimo según normativa")
            + cap_sep
            + _cap("AF", "Análisis Financiero", "TIR, utilidad, margen y estructura de costos del proyecto")
            + cap_sep
            + _cap("FC", "Flujo de Caja", "Curva S mensual, breakeven y exposición máxima de capital")
            + cap_sep
            + _cap("DL", "Due Diligence Legal", "Partida registral, PU/HR, cargas, hipotecas y alertas registrales")
            + '</div>',
            unsafe_allow_html=True
        )

        # ── Footer disclaimer ────────────────────────────
        st.markdown(
            '<div style="margin-top:32px;padding:20px 32px;border-top:1px solid #E0DAD0;'
            'text-align:center;">'
            '<p style="font-size:11px;color:#8A8A8A;line-height:1.7;font-family:Inter,sans-serif;'
            'max-width:680px;margin:0 auto;">'
            'Osterling Advisory está comprometido con brindar información rigurosa y actualizada para la '
            'toma de decisiones inmobiliarias. Los resultados generados por FACTIS tienen carácter '
            'referencial y se basan en los parámetros ingresados por el usuario. Se recomienda validar '
            'los resultados con asesores legales, financieros y técnicos antes de tomar decisiones de inversión. '
            'Para consultas o soporte, escríbenos a '
            '<a href="mailto:eosterling@grupoosterling.com" style="color:#B8904A;text-decoration:none;">'
            'eosterling@grupoosterling.com</a>.'
            '</p>'
            '<p style="font-size:10px;color:#AAAAAA;margin-top:12px;font-family:Inter,sans-serif;">'
            '© 2026 Osterling Advisory — Lima, Perú. Todos los derechos reservados.'
            '</p>'
            '</div>',
            unsafe_allow_html=True
        )

# ═══════════════════════════════════════════════════════
# MÓDULO 2: PROYECTO LOGÍSTICO / INDUSTRIAL
# ═══════════════════════════════════════════════════════

elif tipo_op == "Proyecto Logístico / Industrial":
    r = st.session_state.get("industrial_result")

    if r:
        # ── Controles de comparativa ─────────────────────
        _ic1, _ic2, _ic3 = st.columns([2, 1, 1])
        with _ic1:
            _ind_label = st.text_input("Nombre del escenario", placeholder="Ej: Nave Lurigancho — Opción A",
                                        label_visibility="collapsed", key="ind_cmp_label")
        with _ic2:
            if st.button("GUARDAR EN COMPARATIVA", use_container_width=True, key="btn_ind_cmp"):
                _lbl = _ind_label.strip() or f"Escenario {len(st.session_state.ind_comparativa)+1}"
                _entry = {"label": _lbl, "r": dict(r)}
                existing = [e["label"] for e in st.session_state.ind_comparativa]
                if _lbl in existing:
                    idx = existing.index(_lbl)
                    st.session_state.ind_comparativa[idx] = _entry
                elif len(st.session_state.ind_comparativa) < 3:
                    st.session_state.ind_comparativa.append(_entry)
                else:
                    st.session_state.ind_comparativa[2] = _entry
                st.toast(f"✓ '{_lbl}' guardado en comparativa")
        with _ic3:
            if st.session_state.ind_comparativa and st.button("LIMPIAR COMPARATIVA", use_container_width=True, key="btn_ind_cmp_clear"):
                st.session_state.ind_comparativa = []
                st.rerun()

        # ── Industrial Hero Banner ────────────────────────────────────
        import base64 as _b64i
        _ind_hero_fotos = st.session_state.get("ind_fotos_bytes") or []
        if _ind_hero_fotos:
            try:
                _b64_ind = _b64i.b64encode(_ind_hero_fotos[0]).decode()
                _ind_photo_css = f"url('data:image/jpeg;base64,{_b64_ind}') center/cover no-repeat"
            except Exception:
                _ind_photo_css = "linear-gradient(135deg,#1A2A1A 0%,#243824 100%)"
        else:
            _ind_photo_css = "linear-gradient(135deg,#1A2A1A 0%,#243824 100%)"

        _ind_ubicacion_label = st.session_state.get("ind_ubicacion") or (r.get("zonificacion", "") + " — Lima")
        _ind_yield_bruto = r.get("yield_bruto", 0)
        _ind_dscr        = r.get("dscr")
        _ind_payback     = r.get("payback_anos")
        _ind_kpi4 = (
            f'<div><div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">DSCR</div>'
            f'<div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_ind_dscr:.2f}x</div></div>'
            if _ind_dscr else (
            f'<div><div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">Payback</div>'
            f'<div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_ind_payback:.1f} años</div></div>'
            if _ind_payback else ""
            )
        )

        st.markdown(f"""
        <div style="position:relative;border-radius:16px;overflow:hidden;margin-bottom:20px;
                    box-shadow:0 6px 30px rgba(30,45,61,0.22);">
            <div style="background:{_ind_photo_css};height:220px;"></div>
            <div style="position:absolute;inset:0;background:linear-gradient(to bottom,
                        rgba(0,0,0,0.0) 0%,rgba(0,0,0,0.75) 100%);
                        display:flex;flex-direction:column;justify-content:flex-end;
                        padding:24px 28px;">
                <div style="font-size:9px;color:rgba(255,255,255,0.60);letter-spacing:3px;
                            text-transform:uppercase;margin-bottom:6px;">
                    Análisis Logístico / Industrial · FACTIS
                </div>
                <div style="font-size:28px;font-weight:800;color:#FFFFFF;line-height:1.15;
                            text-shadow:0 2px 8px rgba(0,0,0,0.5);">
                    {_ind_ubicacion_label}
                </div>
                <div style="display:flex;gap:20px;margin-top:12px;flex-wrap:wrap;">
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">Tipo de Nave</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{r.get("tipo_nave","—")}</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">Área Nave</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{r.get("area_nave",0):,.0f} m²</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">Costo Total</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">${r.get("costo_total",0):,.0f}</div>
                    </div>
                    {"" if _ind_yield_bruto == 0 else f'<div><div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">Yield Bruto</div><div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_ind_yield_bruto:.1f}%</div></div>'}
                    {_ind_kpi4}
                </div>
            </div>
        </div>""", unsafe_allow_html=True)

        if st.session_state.get("_goto_tab_name_ind"):
            _ind_tab_name = st.session_state.pop("_goto_tab_name_ind")
            st.components.v1.html(f"""<script>
            setTimeout(function(){{
                var tabs = Array.from(window.parent.document.querySelectorAll('[role="tab"]'));
                var target = tabs.find(function(t){{
                    return t.textContent.trim() === {repr(_ind_tab_name)};
                }});
                if(target){{ target.click(); }}
            }}, 350);
            </script>""", height=0)

        ind_tabs = st.tabs(["Resumen Ejecutivo", "Parámetros", "Financiero", "Flujo de Caja", "Comparativa", "Factibilidad", "Análisis IA"])

        # TAB 0: RESUMEN EJECUTIVO
        with ind_tabs[0]:
            _NAV = "#1E2D3D"; _GOLD = "#B8904A"; _BRD = "#2E3F52"

            # ── Semáforo global ───────────────────────────────────
            fac_re = st.session_state.get("industrial_factibilidad") or {}
            _sem_g = fac_re.get("semaforo_global", "")
            _SEM_CFG = {
                "verde":    ("#107040", "rgba(16,112,64,0.12)",   "✓ SIN ALERTAS CRÍTICAS"),
                "amarillo": ("#9A6E10", "rgba(154,110,16,0.12)",  "⚠ CON OBSERVACIONES"),
                "rojo":     ("#8B1A1A", "rgba(139,26,26,0.12)",   "⛔ ALERTAS CRÍTICAS"),
            }
            if _sem_g in _SEM_CFG:
                _sc, _sbg, _slbl = _SEM_CFG[_sem_g]
                st.markdown(
                    f'<div style="background:{_sbg};border-left:4px solid {_sc};border-radius:8px;'
                    f'padding:14px 20px;margin-bottom:18px;display:flex;align-items:center;gap:14px;">'
                    f'<div style="font-size:24px;line-height:1;">{_slbl.split()[0]}</div>'
                    f'<div><div style="font-size:11px;font-weight:700;letter-spacing:1.5px;color:{_sc};">'
                    f'{" ".join(_slbl.split()[1:])}</div>'
                    f'<div style="font-size:11px;color:{_sc};opacity:0.75;margin-top:2px;">'
                    f'{fac_re.get("resumen_tecnico","")}</div></div></div>',
                    unsafe_allow_html=True)

            # ── Parámetros clave ──────────────────────────────────
            st.markdown('<div class="section-title">Parámetros del Activo</div>', unsafe_allow_html=True)
            _re_cols = st.columns(4)
            _re_cols[0].metric("Tipo de nave",    r.get("tipo_nave", "—"))
            _re_cols[1].metric("Zonificación",     r.get("zonificacion", "—"))
            _re_cols[2].metric("Área nave",        f"{r.get('area_nave',0):,.0f} m²")
            _re_cols[3].metric("Área lote",        f"{r.get('area_terreno',0):,.0f} m²")
            _re_cols2 = st.columns(4)
            _re_cols2[0].metric("Costo total",     f"${r.get('costo_total',0):,.0f}")
            _re_cols2[1].metric("Costo/m² nave",   f"${r.get('costo_por_m2_nave',0):,.0f}/m²")
            _re_cols2[2].metric("Yield bruto",     f"{r.get('yield_bruto',0):.1f}%")
            _re_cols2[3].metric("Payback",         f"{r.get('payback_anos',0):.1f} años" if r.get('payback_anos',0) > 0 else "—")
            if r.get("actividad_descripcion"):
                st.markdown(
                    f'<div style="font-size:11px;color:#8A9BAD;margin-top:4px;">'
                    f'Actividad declarada: <em>{r["actividad_descripcion"]}</em></div>',
                    unsafe_allow_html=True)

            # ── Sustento normativo ────────────────────────────────
            st.markdown('<div class="section-title" style="margin-top:18px;">Sustento Normativo y Técnico</div>', unsafe_allow_html=True)

            def _norm_row(icon, titulo, estado, base_legal, detalle=""):
                _col_ic = "#1A7A4A" if icon == "✓" else ("#C44A4A" if icon == "⛔" else "#B8862E")
                st.markdown(
                    f'<div style="border-left:3px solid {_col_ic};padding:10px 14px;margin-bottom:10px;'
                    f'background:rgba(255,255,255,0.03);border-radius:0 6px 6px 0;">'
                    f'<div style="display:flex;align-items:baseline;gap:8px;">'
                    f'<span style="font-size:14px;color:{_col_ic};">{icon}</span>'
                    f'<span style="font-size:12px;font-weight:700;color:#C8D8E8;">{titulo}</span>'
                    f'<span style="font-size:11px;color:#8A9BAD;margin-left:auto;">{estado}</span></div>'
                    + (f'<div style="font-size:10px;color:#6B8098;margin-top:4px;">'
                       f'Base legal: <strong style="color:#8A9BAD;">{base_legal}</strong></div>' if base_legal else "")
                    + (f'<div style="font-size:11px;color:#A8B8C8;margin-top:3px;">{detalle}</div>' if detalle else "")
                    + '</div>', unsafe_allow_html=True)

            # 1. Zonificación vs. tipo de nave
            _zona_r = r.get("zonificacion", "I2")
            _tipo_r = r.get("tipo_nave", "Almacén Logístico")
            _zona_ok = _zona_r not in ("OU",) and not (_zona_r == "I1" and _tipo_r == "Producción / Manufactura")
            _norm_row(
                "✓" if _zona_ok else "⛔",
                f"Zonificación {_zona_r} — {_tipo_r}",
                "COMPATIBLE" if _zona_ok else "INCOMPATIBLE",
                "Ord. 933-MML-2006 · Índice de Usos ATN-I (RDM/RDA/VT/CV/CZ/CM/I-1/I-2/I-3/I-4)",
                f"La zonificación {_zona_r} permite almacenes y naves logísticas per Índice Usos ATN-I." if _zona_ok
                else f"Actividad requiere zonificación superior — verificar PDU distrital.")

            # 2. Área libre / patios de maniobra
            _area_libre_r = r.get("area_libre", 0)
            _patio_ok = _area_libre_r >= 900
            _norm_row(
                "✓" if _patio_ok else "⚠",
                "Patio de maniobras y circulación",
                f"{_area_libre_r:,.0f} m² disponibles",
                "RNE A.060 Art. 10 · Radio de giro tráiler = 30×30m mínimo (900 m²)",
                f"{'Cumple' if _patio_ok else 'Insuficiente —'} el radio de giro de tráileres requiere mínimo 900 m² de área libre.")

            # 3. EIV — Estudio de Impacto Vial
            _area_nave_r = r.get("area_nave", 0)
            _eiv_req = _area_nave_r >= 1500
            _norm_row(
                "⚠" if _eiv_req else "✓",
                "Estudio de Impacto Vial (EIV)",
                "REQUERIDO" if _eiv_req else "No requerido",
                "RNE A.011 · Resolución Ministerial 167-2023-MTC/14",
                f"Nave de {_area_nave_r:,.0f} m² {'supera el umbral — EIV obligatorio ante municipio y MTC antes de licencia.' if _eiv_req else 'está bajo el umbral de 1,500 m² — sin exigencia de EIV.'}")

            # 4. DSCR bancario
            _dscr_r = r.get("dscr", 0)
            _dscr_ok = _dscr_r == 0 or _dscr_r >= 1.20
            _norm_row(
                "✓" if _dscr_ok else "⛔",
                f"DSCR (cobertura del servicio de deuda): {_dscr_r:.2f}x" if _dscr_r > 0 else "DSCR — sin financiamiento",
                "APROBABLE" if _dscr_ok else "OBSERVADO",
                "Estándar banca comercial Lima · BCP, BBVA, Scotiabank: DSCR mínimo 1.20x",
                f"{'Supera el mínimo bancario de 1.20x.' if _dscr_ok and _dscr_r > 0 else ('Sin crédito declarado.' if _dscr_r == 0 else 'Por debajo del mínimo — banco observará la operación.')}")

            # 5. Yield vs. mercado
            _yld_r = r.get("yield_bruto", 0)
            _yld_ok = _yld_r >= 8.0
            _norm_row(
                "✓" if _yld_ok else ("⚠" if _yld_r >= 5 else "⛔"),
                f"Yield bruto: {_yld_r:.1f}%",
                "SOBRE MERCADO" if _yld_r >= 10 else ("EN MERCADO" if _yld_ok else "BAJO MERCADO"),
                "Benchmarks Lima 2025 · Cushman & Wakefield / Colliers — Lurín Clase A: 8–10% bruto",
                f"Ref. Parque Logístico 47 (Lima VES, 14,315 m², Clase A): yield bruto 26.8% · Lurín mercado: 8–10%.")

            # 6. Acceso pesado (si hay factibilidad)
            _acceso = fac_re.get("restricciones_acceso", "")
            if _acceso:
                _norm_row("⚠", "Acceso y circulación pesada", "VER NOTA",
                          "RNE A.060 Art. 8 · Vías habilitadas para carga mayor",
                          _acceso)

            # 7. Si hay factibilidad registral
            _est_reg = fac_re.get("estado_registral", "")
            if _est_reg:
                _norm_row("✓", "Estado registral (SUNARP)", "VERIFICADO", "SUNARP — partida registral", _est_reg)

            # ── Conclusión ejecutiva (del Análisis IA si ya fue generado) ──
            _memo_re = st.session_state.get("ind_resumen") or {}
            if _memo_re.get("perfil_activo") or _memo_re.get("conclusion"):
                st.markdown('<div class="section-title" style="margin-top:18px;">Conclusión del Análisis</div>', unsafe_allow_html=True)
                _concl = _memo_re.get("conclusion") or _memo_re.get("perfil_activo", "")
                st.markdown(
                    f'<div style="background:rgba(184,144,74,0.07);border-left:3px solid #B8904A;'
                    f'border-radius:0 8px 8px 0;padding:14px 18px;font-size:12px;color:#C8D8E8;line-height:1.7;">'
                    f'{_concl}</div>', unsafe_allow_html=True)
            else:
                st.markdown(
                    '<div style="font-size:11px;color:#6B8098;margin-top:16px;font-style:italic;">'
                    'Genera el Análisis IA (última pestaña) para obtener la conclusión narrativa del advisory board.</div>',
                    unsafe_allow_html=True)

        # TAB 1: PARÁMETROS / GEOMETRÍA + RESUMEN
        with ind_tabs[1]:
            # ── GEOMETRÍA DEL LOTE INDUSTRIAL ────────────────────
            st.markdown('<div class="section-title">Geometría del Lote</div>', unsafe_allow_html=True)

            _ind_geo_modo = st.radio(
                "Fuente de medidas",
                ["Tabular medidas", "Adjuntar plano (DXF/DWG)"],
                horizontal=True, key="ind_geo_modo"
            )

            _ind_poly_lote = st.session_state.get("ind_geo_poly_lote")

            if _ind_geo_modo == "Tabular medidas":
                _igc1, _igc2, _igc3, _igc4 = st.columns(4)
                _ig_frente = _igc1.number_input("Frente (ml)",         min_value=1.0, value=float(st.session_state.get("ind_geo_frente", 30.0)), step=1.0, key="ind_geo_frente")
                _ig_fondo  = _igc2.number_input("Fondo (ml)",          min_value=1.0, value=float(st.session_state.get("ind_geo_fondo",  50.0)), step=1.0, key="ind_geo_fondo")
                _ig_izq    = _igc3.number_input("Lado izquierdo (ml)", min_value=1.0, value=float(st.session_state.get("ind_geo_izq",   50.0)), step=1.0, key="ind_geo_izq")
                _ig_der    = _igc4.number_input("Lado derecho (ml)",   min_value=1.0, value=float(st.session_state.get("ind_geo_der",   50.0)), step=1.0, key="ind_geo_der")

                if st.button("Calcular geometría", key="ind_geo_calc_btn", type="primary"):
                    if _SHAPELY_OK:
                        _ind_poly_lote = _geo_poligono_tabular(_ig_frente, _ig_fondo, _ig_izq, _ig_der)
                        # Si el área declarada en Parámetros difiere >3%, recalcular
                        # la altura del trapezoide para que el área coincida.
                        # Caso típico: lote irregular donde los lados laterales no son
                        # paralelos entre sí (izq y der muy diferentes), lo que invalida
                        # el modelo de altura promedio.
                        _area_decl_fix = float(st.session_state.get("ind_area", 0) or 0)
                        if (_area_decl_fix > 0 and _ind_poly_lote and
                                abs(_ind_poly_lote.area - _area_decl_fix) / _area_decl_fix > 0.03):
                            _h_adj  = 2 * _area_decl_fix / max(_ig_frente + _ig_fondo, 1)
                            _off_adj = (_ig_frente - _ig_fondo) / 2.0
                            from shapely.geometry import Polygon as _SPfix
                            _ind_poly_lote = _SPfix([
                                (0, 0), (_ig_frente, 0),
                                (_ig_frente - _off_adj, _h_adj),
                                (_off_adj, _h_adj),
                            ])
                            st.session_state["ind_geo_adjusted"] = True
                        else:
                            st.session_state["ind_geo_adjusted"] = False
                        st.session_state["ind_geo_poly_lote"]  = _ind_poly_lote
                        st.session_state["ind_geo_frente_val"] = _ig_frente
                        st.session_state["ind_geo_fondo_val"]  = _ig_fondo
                        st.session_state["_goto_tab_name_ind"] = "Parámetros"
                        st.rerun()
                    else:
                        st.markdown('<div class="alert-info">ℹ️ ' + "Librería shapely no disponible." + '</div>', unsafe_allow_html=True)

                if st.session_state.get("ind_geo_adjusted"):
                    _area_decl_note = float(st.session_state.get("ind_area", 0) or 0)
                    st.markdown(
                        f'<div style="background:#E8F5EE;border-left:3px solid #1A4731;border-radius:4px;'
                        f'padding:7px 11px;font-size:11px;color:#1A4731;margin-top:4px;">'
                        f'✓ Polígono ajustado para coincidir con el área declarada en Parámetros '
                        f'<strong>({_area_decl_note:,.0f} m²)</strong>. '
                        f'Los lados laterales desiguales (izq ≠ der) indican un lote irregular — '
                        f'para máxima precisión sube el DXF del levantamiento topográfico.'
                        f'</div>', unsafe_allow_html=True)

            else:
                _ind_geo_file = st.file_uploader(
                    "Cargar plano perimétrico (.dxf / .dwg / .pdf)",
                    type=["dxf", "dwg", "pdf"], key="ind_geo_dxf_file")
                if _ind_geo_file:
                    if _ind_geo_file.name.lower().endswith(".pdf"):
                        st.info("PDF cargado como referencia. Para análisis 3D automático sube el archivo DXF exportado desde AutoCAD o adjunta las medidas en 'Tabular medidas'.")
                    elif _ind_geo_file.name.lower().endswith(".dwg"):
                        st.warning(
                            "**DWG es formato propietario binario** — SOLUM no puede extraer el polígono directamente. "
                            "Exporta como DXF desde AutoCAD: **Archivo → Guardar como → AutoCAD DXF (\\*.dxf)**. "
                            "Una vez en DXF, SOLUM extrae el perímetro automáticamente.")
                    elif _SHAPELY_OK and _EZDXF_OK:
                        import io as _io
                        _ind_poly_lote = _geo_poligono_dxf(_io.TextIOWrapper(_io.BytesIO(_ind_geo_file.read()), encoding="utf-8", errors="ignore"))
                        if _ind_poly_lote:
                            st.session_state["ind_geo_poly_lote"] = _ind_poly_lote
                            st.success(f"Polígono extraído — área geométrica: {_ind_poly_lote.area:,.1f} m²")
                        else:
                            st.error("No se encontró perímetro en el DXF. Verifica que el archivo tenga una LWPOLYLINE o POLYLINE cerrando el lote.")
                    else:
                        st.markdown('<div class="alert-info">ℹ️ ' + "Instala ezdxf y shapely para usar esta función." + '</div>', unsafe_allow_html=True)

            if _ind_poly_lote and not _ind_poly_lote.is_empty:
                st.markdown('<div class="section-title">Actividad y Altura de Nave</div>', unsafe_allow_html=True)

                # Actividades con altura recomendada al hombro
                _IND_ACTIVIDADES = {
                    "Almacén / Centro de distribución / 3PL":  (12.0, 14.0, "Inventario vertical con racks — maximiza m³/m²"),
                    "Producción / Línea de manufactura":        (8.0,  10.0, "Equipos de proceso, ventilación y iluminación industrial"),
                    "Taller / Maestranza / Metalmecánica":      (6.0,   8.0, "Maquinaria pesada, altura moderada suficiente"),
                    "Cámara frigorífica / Frío":                (10.0, 12.0, "Estructura de aislamiento + racks refrigerados"),
                    "Material apilable a granel":               (12.0, 16.0, "Máxima utilización vertical del volumen"),
                    "Con puente grúa (overhead crane)":         (14.0, 16.0, "Altura libre de gancho + estructura del puente ~2-3m"),
                    "Otro / Especificar":                       (6.0,  20.0, "Ingresa la altura requerida por la operación"),
                }
                _act_opts = list(_IND_ACTIVIDADES.keys())
                _act_idx  = _act_opts.index(st.session_state.get("ind_actividad", _act_opts[0])) if st.session_state.get("ind_actividad") in _act_opts else 0
                _ig_actividad = st.selectbox("Actividad a desarrollar", _act_opts, index=_act_idx, key="ind_actividad")
                _h_min, _h_max, _h_nota = _IND_ACTIVIDADES[_ig_actividad]
                _h_default = (_h_min + _h_max) / 2.0

                st.markdown(
                    f'<div style="background:rgba(184,144,74,0.10);border-left:3px solid #B8904A;'
                    f'padding:8px 12px;border-radius:4px;font-size:11px;color:#C8A86A;margin-bottom:10px;">'
                    f'Altura recomendada al hombro: <strong>{_h_min:.0f}–{_h_max:.0f}m</strong> · {_h_nota}</div>',
                    unsafe_allow_html=True)

                st.markdown('<div class="section-title">Massing 3D</div>', unsafe_allow_html=True)
                _irg_h, _irg_sp = st.columns([1, 3])
                _ig_h_nave = _irg_h.number_input(
                    "Altura al hombro (m)", min_value=3.0, max_value=25.0,
                    value=float(st.session_state.get("ind_geo_h_nave_val", _h_default)),
                    step=0.5, key="ind_geo_h_nave")
                st.session_state["ind_geo_h_nave_val"] = _ig_h_nave

                # Retiros fijos en 0 — la huella declarada ES la ocupación de la nave
                _ig_ret_f, _ig_ret_l, _ig_ret_p = 0.0, 0.0, 0.0
                _ind_poly_huella = _geo_aplicar_retiros(_ind_poly_lote, _ig_ret_f, _ig_ret_l, _ig_ret_p)
                _ig_frente_val   = st.session_state.get("ind_geo_frente_val", 0.0)

                # Validación industrial (área libre industrial: patio de maniobras ~30%)
                _ind_al_pct = float((st.session_state.get("ind_params") or {}).get("area_libre_min_pct") or 0)
                _ind_val = _geo_validar(_ind_poly_lote, _ind_poly_huella, 1, _ig_frente_val, "industrial", _ind_al_pct)

                _img = _ind_val["metricas"]
                _imc1, _imc2, _imc3, _imc4 = st.columns(4)
                _imc1.metric("Área del lote",            f"{_img.get('area_lote_m2', 0):,.0f} m²")
                _imc2.metric("Huella edificable",        f"{_img.get('area_huella_efectiva_m2', 0):,.0f} m²")
                _imc3.metric("Coef. de Ocupación (COS)", f"{_img.get('cos_real_pct', 0):.1f}%")
                _imc4.metric("Área libre",               f"{max(0, _img.get('area_lote_m2',0) - _img.get('area_huella_efectiva_m2',0)):,.0f} m²")

                # Validación radio de giro tráileres
                _ig_frente_check = st.session_state.get("ind_geo_frente_val", 0.0) or (_img.get("area_lote_m2", 0) ** 0.5)
                if _img.get("area_huella_m2", 0) > 0:
                    _area_libre_ind = _img.get("area_lote_m2", 0) - _img.get("area_huella_m2", 0)
                    if _area_libre_ind < 900:
                        st.markdown('<div class="alert-gold">⚠ RNE A.060: patio de maniobras mínimo 30×30m (900m²) para radio de giro de tráileres — área libre insuficiente.</div>', unsafe_allow_html=True)

                for _alerta_ind in _ind_val["alertas"]:
                    st.markdown(f'<div class="alert-gold">⚠ {_alerta_ind}</div>', unsafe_allow_html=True)

                # ── Layout: 3D a la izquierda, tabla de áreas a la derecha ──
                _col_3d, _col_tbl = st.columns([3, 2])

                with _col_3d:
                    # 3D industrial: nave de 1 planta con altura real al hombro
                    _fig_3d_ind = _geo_render_3d(_ind_poly_lote, _ind_poly_huella, 1, 0, _ig_h_nave)
                    st.plotly_chart(_fig_3d_ind, use_container_width=True)
                    st.caption(f"Dorado: límite del lote · Azul: huella de nave · Altura: {_ig_h_nave:.1f}m al hombro")

                with _col_tbl:
                    _area_lote_geo  = _img.get("area_lote_m2", 0)
                    _area_nave_geo  = _img.get("area_huella_m2", 0)
                    _area_libre_geo = max(0, _area_lote_geo - _area_nave_geo)
                    _vol_nave       = _area_nave_geo * _ig_h_nave
                    _area_declarada = st.session_state.get("ind_area", 0)

                    # Discrepancy warning
                    if _area_declarada > 0 and _area_lote_geo > 0:
                        _diff_pct = abs(_area_lote_geo - _area_declarada) / _area_declarada * 100
                        if _diff_pct > 5:
                            st.markdown(
                                f'<div style="background:#FFF8EE;border-left:3px solid #B8904A;'
                                f'padding:7px 10px;border-radius:4px;font-size:10px;color:#7A4F1A;margin-bottom:8px;">'
                                f'⚠ Área geométrica ({_area_lote_geo:,.0f} m²) difiere {_diff_pct:.0f}% del área '
                                f'declarada en Parámetros ({_area_declarada:,.0f} m²). '
                                f'Los cálculos financieros usan el área de Parámetros.</div>',
                                unsafe_allow_html=True)

                    _tbl_rows = [
                        ("Área del lote (geo)",      f"{_area_lote_geo:,.0f} m²",  "Calculada del polígono"),
                        ("Área declarada (Paráms.)", f"{_area_declarada:,.0f} m²", "Usada en finanzas"),
                        ("Área de nave (huella)",    f"{_area_nave_geo:,.0f} m²",  f"{_img.get('cos_real_pct',0):.1f}% del lote"),
                        ("Área de operaciones",      f"{_area_libre_geo:,.0f} m²", "Patios + maniobras"),
                        ("Altura de nave",           f"{_ig_h_nave:.1f} m",         "Al hombro (gotera)"),
                        ("Volumen operativo",        f"{_vol_nave:,.0f} m³",        "Nave × altura"),
                    ]
                    _tbl_html = (
                        '<table style="width:100%;border-collapse:collapse;font-size:11px;">'
                        '<thead><tr>'
                        '<th style="background:#1E2D3D;color:#B8904A;padding:6px 8px;text-align:left;font-size:10px;">Parámetro</th>'
                        '<th style="background:#1E2D3D;color:#FFFFFF;padding:6px 8px;text-align:right;">Valor</th>'
                        '<th style="background:#1E2D3D;color:#9A9080;padding:6px 8px;text-align:left;font-size:9px;">Nota</th>'
                        '</tr></thead><tbody>'
                    )
                    for _i, (_lbl, _val, _nota) in enumerate(_tbl_rows):
                        _bg = "#F5F2ED" if _i % 2 == 0 else "#FFFFFF"
                        _tbl_html += (
                            f'<tr style="background:{_bg};">'
                            f'<td style="padding:6px 8px;color:#1E2D3D;">{_lbl}</td>'
                            f'<td style="padding:6px 8px;color:#1E2D3D;text-align:right;font-weight:600;">{_val}</td>'
                            f'<td style="padding:6px 8px;color:#7A7268;font-size:10px;">{_nota}</td>'
                            f'</tr>'
                        )
                    _tbl_html += '</tbody></table>'
                    st.markdown(_tbl_html, unsafe_allow_html=True)

                st.session_state["ind_geo_huella"] = _img.get("area_huella_m2", 0)

            st.markdown("---")
            st.markdown('<div class="section-title">Resumen del Proyecto</div>', unsafe_allow_html=True)
            ci1, ci2, ci3, ci4 = st.columns(4)
            ci1.metric("Área del Terreno", f"{r.get('area_terreno', 0):,.0f} m²")
            ci2.metric("Área Nave (techada)", f"{r.get('area_nave', 0):,.0f} m²", f"{r.get('pct_techada', 0):.0f}% techada · {r.get('area_libre', 0):,.0f} m² libre")
            ci3.metric("Costo Total Proyecto", f"${r.get('costo_total', 0):,.0f}")
            ci4.metric("Costo por m² (nave)", f"${r.get('costo_por_m2_nave', 0):,.0f}/m²")

            st.markdown('<div class="section-title">Estructura de Costos</div>', unsafe_allow_html=True)
            ci5, ci6 = st.columns(2)
            with ci5:
                ci5.metric("Terreno", f"${r.get('costo_terreno', 0):,.0f}")
                ci5.metric("Alcabala (3%)", f"${r.get('alcabala', 0):,.0f}")
                ci5.metric("Nave techada", f"${r.get('costo_nave_total', 0):,.0f}", f"${r.get('costo_nave_m2', 0):,.0f}/m² × {r.get('area_nave', 0):,.0f} m²")
                ci5.metric("Piso área libre", f"${r.get('costo_pisos_libres', 0):,.0f}", f"${r.get('costo_piso_libre_m2', 0):,.0f}/m² × {r.get('area_libre', 0):,.0f} m²")
                ci5.metric(f"Costos Indirectos ({r.get('pct_indirectos', 5):.0f}%)", f"${r['soft_costs']:,.0f}")
            with ci6:
                costo_items = [
                    ("Terreno", r.get('costo_terreno', 0)),
                    ("Nave techada", r.get('costo_nave_total', 0)),
                    ("Piso libre", r.get('costo_pisos_libres', 0)),
                    (f"Costos Indirectos", r.get('soft_costs', 0)),
                ]
                fig_costs = go.Figure(go.Bar(
                    x=[x[0] for x in costo_items],
                    y=[x[1] for x in costo_items],
                    marker_color=["#1E2D3D", "#B8904A", "#8A9BAD", "#C8A86A"],
                    text=[f"${x[1]:,.0f}" for x in costo_items],
                    textposition="outside",
                ))
                fig_costs.update_layout(
                    height=300, margin=dict(t=20, b=20, l=10, r=10),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    yaxis=dict(title=dict(text="USD", font=dict(color="#4A5568", size=11)), tickfont=dict(color="#4A5568", size=10), showgrid=True, gridcolor="#E8E4DC"),
                    xaxis=dict(tickfont=dict(color="#4A5568", size=10)),
                    font=dict(family="Inter", color="#4A5568"),
                    showlegend=False,
                )
                st.plotly_chart(fig_costs, use_container_width=True)

            st.markdown(
                '<div class="alert-gold">'
                f'<strong>Costos de construcción:</strong> Nave: ${r.get("costo_nave_m2", 0):,.0f}/m² ({r.get("tipo_nave", "")}) · '
                f'Piso área libre: ${r.get("costo_piso_libre_m2", 0):,.0f}/m². '
                f'Costos indirectos ({r.get("pct_indirectos", 5):.0f}%): permisos, licencias municipales, supervisión y gestión de obra.'
                '</div>',
                unsafe_allow_html=True
            )

            # ── Validador precio de terreno (referencia orientativa) ──
            _pt_m2 = r.get('costo_terreno', 0) / r.get('area_terreno', 1) if r.get('area_terreno', 0) > 0 else 0
            _zona_ind_sel = st.session_state.get("ind_zona_lima", "")
            _es_ves_lurin = any(z in _zona_ind_sel for z in ["Villa El Salvador", "Lurín"])
            if _pt_m2 > 0:
                _uso_ind = r.get("uso", "Inversión")
                if _es_ves_lurin:
                    _max_t = 180 if _uso_ind == "Inversión" else 300
                    _ref_txt = ("build-to-rent · ref. Aldea Logística VES 2023-2024: $140–$172/m²"
                                if _uso_ind == "Inversión" else "uso propio · umbral orientativo")
                    if _pt_m2 <= _max_t:
                        _tc, _tb = "#E8F5EE", "#1A4731"
                        _tm = f"✓ Terreno a <b>${_pt_m2:,.0f}/m²</b> — dentro del rango para {_ref_txt} en VES/Lurín."
                    elif _pt_m2 <= _max_t * 1.25:
                        _tc, _tb = "#FFF8E6", "#7A5500"
                        _tm = (f"⚡ Terreno a <b>${_pt_m2:,.0f}/m²</b> — ligeramente sobre el umbral orientativo "
                               f"(${_max_t}/m²) para {_ref_txt}. Verificar márgenes.")
                    else:
                        _tc, _tb = "#FDECEA", "#7A1A1A"
                        _tm = (f"⚠ Terreno a <b>${_pt_m2:,.0f}/m²</b> — supera el umbral orientativo "
                               f"(${_max_t}/m²) para {_ref_txt}. Evaluar ajuste de precio o renta requerida.")
                else:
                    _tc, _tb = "#F0EDE8", "#4A5568"
                    _zona_rng = "$450–$700/m²" if any(z in _zona_ind_sel for z in ["Callao", "SJL", "Cercado", "Huachipa"]) else "variable según zona"
                    _tm = (f"Terreno a <b>${_pt_m2:,.0f}/m²</b>. "
                           f"En zonas céntricas ({_zona_rng}) la viabilidad depende principalmente "
                           f"de la capacidad de pago de la industria/inquilino.")
                st.markdown(
                    f'<div style="background:{_tc};border-left:3px solid {_tb};border-radius:6px;'
                    f'padding:10px 14px;margin-top:10px;font-size:12px;color:#1A2233;">'
                    f'{_tm}<br>'
                    f'<span style="font-size:10px;color:#4A5568;">Referencia orientativa — el profesional aplica su criterio según la operación específica.</span>'
                    f'</div>',
                    unsafe_allow_html=True)

        # TAB 2: FINANCIERO
        with ind_tabs[2]:
            # ── Compra vs. Arrendamiento ──────────────────────
            st.markdown('<div class="section-title">Compra vs. Arrendamiento</div>', unsafe_allow_html=True)
            st.caption("Referencia orientativa — el profesional aplica su criterio según la industria, perfil del inquilino y estructura del deal.")

            _an = r.get('area_nave', 0) or 0
            _cuota_ef = r.get('cuota_efectiva_mensual', 0)
            if _cuota_ef > 0 and r.get('cuota_mensual', 0) > 0:
                _costo_m2_mes = _cuota_ef / _an if _an > 0 else 0
                _metodo_lbl   = f"Cuota efectiva a {r.get('plazo_anos', 0)} años (neta escudo fiscal)"
            else:
                _costo_m2_mes = r.get('costo_total', 0) / (10 * 12 * _an) if _an > 0 else 0
                _metodo_lbl   = "Amortización lineal 10 años (sin financiamiento)"
            _renta_be10 = r.get('costo_total', 0) / (10 * 12 * _an) if _an > 0 else 0

            _PRIME_LO, _PRIME_HI = 5.50, 7.50
            if _costo_m2_mes <= _PRIME_LO:
                _bvr_bg = "#1A4731"
            elif _costo_m2_mes <= _PRIME_HI:
                _bvr_bg = "#7A5500"
            else:
                _bvr_bg = "#8B1A1A"

            bvr1, bvr2 = st.columns(2)
            with bvr1:
                st.metric("Costo efectivo compra", f"${_costo_m2_mes:.2f}/m²/mes", _metodo_lbl)
                st.markdown(
                    '<div style="font-size:10px;color:#7A7268;margin-top:2px;line-height:1.5;">'
                    'Ref. renta de mercado Lima: <strong>$5.50–$7.50/m²/mes</strong> '
                    '· Naves logísticas Clase A · VES, Lurín, Callao</div>',
                    unsafe_allow_html=True)
            bvr2.metric("Break-even renta 10 años", f"${_renta_be10:.2f}/m²/mes",
                        "Renta mínima para recuperar la inversión en 10 años")

            st.markdown(
                f'<div style="background:{_bvr_bg};height:6px;border-radius:3px;margin-bottom:16px;"></div>',
                unsafe_allow_html=True)

            # ── Proyección indexada (solo contratos plurianuales) ──
            if r.get("tipo_contrato") == "Plurianual (3+ años)" and r.get("ajuste_anual_pct", 0) > 0:
                _aj  = r["ajuste_anual_pct"]
                _ini = r["inicio_ajuste_ano"]
                st.markdown(
                    f'<div style="background:#F0F4FF;border-left:3px solid #3B5BDB;border-radius:6px;'
                    f'padding:10px 14px;margin-bottom:8px;font-size:12px;">'
                    f'<strong>Contrato plurianual · ajuste +{_aj:.1f}% anual desde año {_ini}</strong><br>'
                    f'La renta crece con el contrato; el costo de compra (cuota) permanece fijo — '
                    f'el diferencial favorable <b>se amplía año a año</b>.'
                    f'</div>', unsafe_allow_html=True)
                _pi_cols = st.columns(4)
                _pi_cols[0].metric("Renta/m² año 1", f"${r['renta_m2_mes']:.2f}/mes")
                _pi_cols[1].metric(f"Renta/m² año 3", f"${r['renta_m2_ano3']:.2f}/mes",
                                   f"+{((r['renta_m2_ano3']/r['renta_m2_mes']-1)*100):.1f}%" if r['renta_m2_mes'] > 0 else "")
                _pi_cols[2].metric(f"Renta/m² año 5", f"${r['renta_m2_ano5']:.2f}/mes",
                                   f"+{((r['renta_m2_ano5']/r['renta_m2_mes']-1)*100):.1f}%" if r['renta_m2_mes'] > 0 else "")
                if r.get("payback_indexado") and r.get("payback_anos"):
                    _delta_pb = r["payback_anos"] - r["payback_indexado"]
                    _pi_cols[3].metric("Payback indexado", f"{r['payback_indexado']} años",
                                       f"{_delta_pb:.1f} años menos vs renta fija")
                _yield_data = {
                    "Año": ["Año 1 (base)", f"Año 3 (+{_aj:.1f}% × {max(3-_ini+1,0)})", f"Año 5 (+{_aj:.1f}% × {max(5-_ini+1,0)})"],
                    "Yield neto": [f"{r['yield_neto']:.1f}%", f"{r['yield_neto_ano3']:.1f}%", f"{r['yield_neto_ano5']:.1f}%"],
                }
                st.table(pd.DataFrame(_yield_data).set_index("Año"))

            st.markdown("---")
            st.markdown('<div class="section-title">Estructura de Financiamiento</div>', unsafe_allow_html=True)

            # ── Bloque A: Terreno ──
            st.markdown(
                '<div style="font-size:10px;font-weight:700;color:#B8904A;letter-spacing:1.5px;'
                'text-transform:uppercase;margin:10px 0 6px;">A · Adquisición del Terreno</div>',
                unsafe_allow_html=True)
            _fA1, _fA2, _fA3, _fA4 = st.columns(4)
            _fA1.metric("Costo Terreno + Alcabala", f"${r.get('costo_terreno_alcabala', 0):,.0f}")
            _fA2.metric("Downpayment",
                        f"${r.get('capital_propio_terreno', 0):,.0f}",
                        f"{r.get('dp_terreno_pct', 0):.0f}% al contado")
            _fA3.metric("Crédito Terreno",
                        f"${r.get('monto_credito_terreno', 0):,.0f}",
                        f"{100 - r.get('dp_terreno_pct', 0):.0f}% financiado")
            _fA4.metric("Cuota Terreno / mes",
                        f"${r.get('cuota_terreno', 0):,.0f}" if (r.get('cuota_terreno') or 0) > 0 else "—",
                        f"{r.get('tasa_terreno', 0):.1f}% · {r.get('plazo_terreno', 0)} años")

            # ── Bloque B: Obra ──
            st.markdown(
                '<div style="font-size:10px;font-weight:700;color:#4A90C4;letter-spacing:1.5px;'
                'text-transform:uppercase;margin:14px 0 6px;">B · Construcción e Implementación</div>',
                unsafe_allow_html=True)
            _fB1, _fB2, _fB3, _fB4 = st.columns(4)
            _fB1.metric("Costo Total Construcción", f"${r.get('costo_construccion_soft', 0):,.0f}",
                        f"${r.get('costo_por_m2_nave', 0):,.0f}/m² nave")
            _fB2.metric("Downpayment Obra",
                        f"${r.get('capital_propio_const', 0):,.0f}",
                        f"{r.get('dp_const_pct', 0):.0f}% al contado")
            _fB3.metric("Crédito Construcción",
                        f"${r.get('monto_credito_const', 0):,.0f}",
                        f"{100 - r.get('dp_const_pct', 0):.0f}% financiado")
            _fB4.metric("Cuota Obra / mes",
                        f"${r.get('cuota_const', 0):,.0f}" if (r.get('cuota_const') or 0) > 0 else "—",
                        f"{r.get('tasa_const', 0):.1f}% · {r.get('plazo_const', 0)} años")

            # ── Bloque C: Totales ──
            st.markdown(
                '<div style="font-size:10px;font-weight:700;color:#1E2D3D;letter-spacing:1.5px;'
                'text-transform:uppercase;margin:14px 0 6px;border-top:1px solid #D8D4CC;padding-top:10px;">'
                'C · Resumen Total del Financiamiento</div>',
                unsafe_allow_html=True)
            cf1, cf2, cf3 = st.columns(3)
            cf1.metric("Capital Propio Total", f"${r.get('capital_propio', 0):,.0f}",
                       f"A+B downpayments")
            cf2.metric("Deuda Total", f"${r.get('monto_credito', 0):,.0f}",
                       f"{r.get('pct_credito', 0):.0f}% del proyecto")
            cf3.metric("Cuota Total / mes", f"${r.get('cuota_mensual', 0):,.0f}" if (r.get('cuota_mensual') or 0) > 0 else "Sin crédito",
                       "Terreno + Obra")

            if r.get('uso') == "Inversión":
                st.markdown('<div class="section-title">Métricas de Retorno</div>', unsafe_allow_html=True)
                cr1, cr2, cr3, cr4 = st.columns(4)
                cr1.metric("Renta Mensual Total", f"${r.get('renta_total_mes', 0):,.0f}", f"${r.get('renta_m2_mes', 0):.2f}/m²/mes")
                cr2.metric("Yield Bruto Anual", f"{r.get('yield_bruto', 0):.1f}%")
                cr3.metric("Yield Neto Anual", f"{r.get('yield_neto', 0):.1f}%", "8% gastos op.")
                if r.get('payback_anos'):
                    cr4.metric("Payback", f"{r['payback_anos']:.1f} años")
                else:
                    cr4.metric("Payback", "N/A")

                if (r.get('cuota_mensual') or 0) > 0:
                    st.markdown('<div class="section-title">Flujo con Financiamiento</div>', unsafe_allow_html=True)
                    cd1, cd2, cd3 = st.columns(3)
                    flujo = r.get('flujo_mensual') or 0
                    flujo_label = "Flujo mensual neto" if flujo >= 0 else "Déficit mensual"
                    cd1.metric(flujo_label, f"${abs(flujo):,.0f}/mes")
                    if r.get('dscr'):
                        cd2.metric("DSCR", f"{r['dscr']:.2f}x", "Cobertura deuda")
                    cd3.metric("Escudo fiscal nave", f"${r.get('ahorro_fiscal_mensual', 0):,.0f}/mes",
                               f"${r.get('ahorro_fiscal_anual', 0):,.0f}/año · IR 29.5%")

                    _dscr_val = r.get('dscr') or 0
                    if _dscr_val >= 1.2:
                        _dscr_msg = (f"DSCR {_dscr_val:.2f}x — Cobertura adecuada. La renta cubre el servicio de deuda con margen.")
                        _dscr_cls = "alert-legal"
                    elif _dscr_val >= 1.0:
                        _dscr_msg = (f"DSCR {_dscr_val:.2f}x — Cobertura ajustada. La renta apenas cubre la deuda; "
                                     f"el retorno se evalúa por TIR equity, no solo por DSCR. "
                                     f"En proyectos de desarrollo con financiamiento &gt;60%, DSCR &lt; 1.20x es normal "
                                     f"al inicio; la vacancia en estabilización y la apreciación del activo compensan.")
                        _dscr_cls = "alert-gold"
                    else:
                        _dscr_msg = (f"DSCR {_dscr_val:.2f}x — La renta no cubre el servicio de deuda. "
                                     f"Esto es frecuente en proyectos de desarrollo nuevos donde el costo total de obra "
                                     f"es la base del financiamiento. El indicador determinante es la <strong>TIR equity</strong> "
                                     f"y la <strong>plusvalía del activo</strong> al horizonte de salida.")
                        _dscr_cls = "alert-gold"
                    st.markdown(f'<div class="{_dscr_cls}">{_dscr_msg}</div>', unsafe_allow_html=True)

            elif r.get('uso') == "Uso directo":
                st.markdown('<div class="section-title">Análisis vs. Arrendamiento</div>', unsafe_allow_html=True)
                ca1, ca2, ca3 = st.columns(3)
                ca1.metric("Renta de Mercado (ref.)", f"${r.get('renta_total_mes', 0):,.0f}/mes", f"${r.get('renta_m2_mes', 0):.2f}/m²/mes")
                ca2.metric("Cuota Mensual", f"${r.get('cuota_mensual', 0):,.0f}/mes" if (r.get('cuota_mensual') or 0) > 0 else "Compra al contado")
                ahorro = r.get('alquiler_vs_compra') or 0
                ca3.metric("Ahorro vs. Alquilar", f"${ahorro:,.0f}/mes", "Compra vs renta mensual")

                if (r.get('cuota_mensual') or 0) > 0 and (r.get('renta_total_mes') or 0) > 0:
                    anos_breakeven = r.get('capital_propio', 0) / (ahorro * 12) if ahorro > 0 else None
                    if anos_breakeven:
                        st.markdown(
                            '<div class="alert-gold">'
                            f'<strong>Punto de equilibrio estimado:</strong> El ahorro acumulado vs. arrendamiento '
                            f'recupera el capital propio invertido en aproximadamente <strong>{anos_breakeven:.1f} años</strong>.'
                            '</div>',
                            unsafe_allow_html=True
                        )

        # TAB 3: FLUJO DE CAJA
        with ind_tabs[3]:
            if r.get('uso') == "Inversión" and r.get('flujo_anual'):
                st.markdown('<div class="section-title">Score de Viabilidad</div>', unsafe_allow_html=True)

                # Score compuesto
                _score_pts = 0
                _score_max = 4
                if r.get('yield_neto', 0) >= 7:    _score_pts += 1
                if r.get('dscr') and r['dscr'] >= 1.2: _score_pts += 1
                if r.get('payback_anos') and r['payback_anos'] <= 12: _score_pts += 1
                if r.get('irr_anual') and r['irr_anual'] >= 12: _score_pts += 1
                _score_pct = _score_pts / _score_max
                _score_color = "#1A4731" if _score_pct >= 0.75 else ("#7A4F1A" if _score_pct >= 0.5 else "#7A1A1A")
                _score_bg = "#E8F5EE" if _score_pct >= 0.75 else ("#FFF8EE" if _score_pct >= 0.5 else "#FFF0F0")
                _score_etiq = "INVERSIÓN VIABLE" if _score_pct >= 0.75 else ("INVERSIÓN CON RESERVAS" if _score_pct >= 0.5 else "REVISAR ESTRUCTURA")

                st.markdown(f"""
                <div style="background:{_score_bg};border:1px solid {_score_color};border-left:4px solid {_score_color};
                            border-radius:8px;padding:18px 24px;margin-bottom:20px;display:flex;align-items:center;gap:24px;">
                    <div style="font-size:48px;font-weight:800;color:{_score_color};min-width:64px;text-align:center;">{_score_pts}/{_score_max}</div>
                    <div>
                        <div style="font-size:9px;letter-spacing:3px;color:{_score_color};text-transform:uppercase;font-weight:700;opacity:0.7;">Score de Viabilidad Industrial</div>
                        <div style="font-size:18px;font-weight:700;color:{_score_color};margin-top:4px;">{_score_etiq}</div>
                        <div style="font-size:12px;color:{_score_color};opacity:0.8;margin-top:6px;line-height:1.5;">
                            {"✓" if r.get('yield_neto',0)>=7 else "✗"} Yield neto ≥ 7% &nbsp;·&nbsp;
                            {"✓" if r.get('dscr') and r['dscr']>=1.2 else "✗"} DSCR ≥ 1.20x &nbsp;·&nbsp;
                            {"✓" if r.get('payback_anos') and r['payback_anos']<=12 else "✗"} Payback ≤ 12 años &nbsp;·&nbsp;
                            {"✓" if r.get('irr_anual') and r['irr_anual']>=12 else "✗"} TIR equity ≥ 12%
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)

                st.markdown('<div class="section-title">Indicadores de Retorno</div>', unsafe_allow_html=True)
                cf1, cf2, cf3, cf4 = st.columns(4)
                cf1.metric("TIR Equity 10 años", f"{r.get('irr_anual', 0):.1f}%" if r.get('irr_anual') is not None else "—")
                cf2.metric("VAN 10 años (10%)", f"${r.get('van_10', 0):,.0f}" if r.get('van_10') is not None else "—")
                cf3.metric("Yield Neto", f"{r.get('yield_neto', 0):.1f}%")
                cf4.metric("Payback", f"{r.get('payback_anos', 0):.1f} a." if r.get('payback_anos') else "—")

                st.markdown('<div class="section-title">Flujo de Caja Proyectado (10 años)</div>', unsafe_allow_html=True)
                fa = r['flujo_anual']
                anos_fc = list(range(len(fa)))
                colores_fc = ["#7A1A1A" if v < 0 else "#1A4731" for v in fa]
                fig_fc = go.Figure(go.Bar(
                    x=[f"Año {i}" if i > 0 else "Inv. Inicial" for i in anos_fc],
                    y=fa,
                    marker_color=colores_fc,
                    text=[f"${v:,.0f}" for v in fa],
                    textposition="outside",
                ))
                fig_fc.update_layout(
                    height=300, margin=dict(t=30, b=20, l=10, r=10),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    yaxis=dict(tickfont=dict(color="#4A5568", size=10), showgrid=True, gridcolor="#E8E4DC", tickformat="$,.0f"),
                    xaxis=dict(tickfont=dict(color="#4A5568", size=10)),
                    font=dict(family="Inter", color="#4A5568"),
                    showlegend=False,
                )
                st.plotly_chart(fig_fc, use_container_width=True)

                # Tabla flujo
                _fc_th = "".join(
                    f'<th style="background:#1E2D3D;color:#FFFFFF;padding:9px 12px;font-size:10px;'
                    f'letter-spacing:1px;text-transform:uppercase;font-weight:700;border:1px solid #2A3D51;text-align:right;">{h}</th>'
                    for h in ["Período", "Renta Neta", "Cuota Deuda", "Flujo Libre", "Flujo Acumulado"]
                )
                _fc_rows = ""
                _acum = 0
                cuota_anual = (r.get('cuota_mensual') or 0) * 12
                for i, fv in enumerate(fa):
                    bg = "#FFFFFF" if i % 2 == 0 else "#F9F7F4"
                    yr_lbl = "Inversión Inicial" if i == 0 else f"Año {i}"
                    if i == 0:
                        _acum += fv
                        _fc_rows += (f'<tr><td style="background:{bg};color:#1E2D3D;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;">{yr_lbl}</td>'
                                     f'<td style="background:{bg};color:#1E2D3D;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">—</td>'
                                     f'<td style="background:{bg};color:#1E2D3D;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">—</td>'
                                     f'<td style="background:{bg};color:#7A1A1A;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;font-weight:700;">${fv:,.0f}</td>'
                                     f'<td style="background:{bg};color:#7A1A1A;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">${_acum:,.0f}</td></tr>')
                    else:
                        yr_actual = i
                        _cuota_yr = cuota_anual if yr_actual <= r.get('plazo_anos', 0) else 0
                        _renta = r.get('renta_neta_anual', 0) or 0
                        _flujo_libre = _renta - _cuota_yr
                        _flujo_display = _flujo_libre if i < len(fa) - 1 else fv
                        _acum += _flujo_display
                        _col_libre = "#1A4731" if _flujo_display >= 0 else "#7A1A1A"
                        _col_acum  = "#1A4731" if _acum >= 0 else "#7A1A1A"
                        _yr_note = " + Venta" if i == len(fa) - 1 else ""
                        _fc_rows += (f'<tr><td style="background:{bg};color:#1E2D3D;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;">{yr_lbl}{_yr_note}</td>'
                                     f'<td style="background:{bg};color:#1E2D3D;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">${_renta:,.0f}</td>'
                                     f'<td style="background:{bg};color:#1E2D3D;padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">${_cuota_yr:,.0f}</td>'
                                     f'<td style="background:{bg};color:{_col_libre};padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;font-weight:600;">${_flujo_display:,.0f}</td>'
                                     f'<td style="background:{bg};color:{_col_acum};padding:8px 12px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">${_acum:,.0f}</td></tr>')

                st.markdown(
                    f'<div style="overflow-x:auto;margin-bottom:20px;">'
                    f'<table style="width:100%;border-collapse:collapse;">'
                    f'<thead><tr>{_fc_th}</tr></thead><tbody>{_fc_rows}</tbody></table></div>',
                    unsafe_allow_html=True
                )

                st.markdown('<div class="section-title">Descargar Informe</div>', unsafe_allow_html=True)
                _ind_pdf_bytes = generar_informe_industrial_pdf(
                    r, st.session_state.get("industrial_factibilidad"),
                    datetime.datetime.now().strftime("%d/%m/%Y"),
                    altura_nave=float(st.session_state.get("ind_geo_h_nave_val") or 0))
                st.download_button(
                    "DESCARGAR INFORME PDF",
                    data=_ind_pdf_bytes,
                    file_name=f"informe_industrial_{datetime.datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
                st.caption("El archivo .html se abre en cualquier navegador y puede imprimirse como PDF.")

            else:
                st.markdown(
                    '<div class="alert-legal">El flujo de caja e IRR están disponibles en modo <strong>Inversión</strong> '
                    'con renta de mercado ingresada. Cambia el propósito del activo en el panel izquierdo.</div>',
                    unsafe_allow_html=True
                )
                # Siempre mostrar botón de descarga aunque no haya IRR
                st.markdown('<div class="section-title">Descargar Informe</div>', unsafe_allow_html=True)
                _ind_pdf_bytes = generar_informe_industrial_pdf(
                    r, st.session_state.get("industrial_factibilidad"),
                    datetime.datetime.now().strftime("%d/%m/%Y"),
                    altura_nave=float(st.session_state.get("ind_geo_h_nave_val") or 0))
                st.download_button(
                    "DESCARGAR INFORME PDF",
                    data=_ind_pdf_bytes,
                    file_name=f"informe_industrial_{datetime.datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

        # TAB 4: COMPARATIVA
        with ind_tabs[4]:
            # Multi-project comparison
            cmp_list = st.session_state.ind_comparativa
            if cmp_list:
                st.markdown('<div class="section-title">Comparativa de Escenarios</div>', unsafe_allow_html=True)
                _cmp_metrics = ["costo_total", "costo_por_m2_nave", "yield_neto", "dscr", "payback_anos", "irr_anual", "capital_propio", "cuota_mensual"]
                _cmp_labels  = ["Costo Total", "Costo/m² Nave", "Yield Neto %", "DSCR", "Payback (años)", "TIR Equity %", "Capital Propio", "Cuota Mensual"]
                _cmp_fmt     = ["${:,.0f}", "${:,.0f}", "{:.1f}%", "{:.2f}x", "{:.1f} a.", "{:.1f}%", "${:,.0f}", "${:,.0f}"]

                _th = '<th style="background:#1E2D3D;color:#FFFFFF;padding:9px 14px;font-size:10px;letter-spacing:1px;text-transform:uppercase;">Indicador</th>'
                for e in cmp_list:
                    _th += f'<th style="background:#1E2D3D;color:#B8904A;padding:9px 14px;font-size:10px;letter-spacing:1px;text-transform:uppercase;text-align:right;">{e["label"]}</th>'

                _rows = ""
                for lbl, key, fmt in zip(_cmp_labels, _cmp_metrics, _cmp_fmt):
                    bg_row = "#FFFFFF" if _cmp_metrics.index(key) % 2 == 0 else "#F9F7F4"
                    _row = f'<td style="background:{bg_row};color:#4A5568;padding:8px 14px;font-size:11px;font-weight:600;border:1px solid #E0DAD0;">{lbl}</td>'
                    vals = [e["r"].get(key) for e in cmp_list]
                    # Determine best value (highest or lowest depending on metric)
                    _higher_is_better = key in ["yield_neto", "dscr", "irr_anual"]
                    _lower_is_better  = key in ["costo_total", "costo_por_m2_nave", "payback_anos", "cuota_mensual", "capital_propio"]
                    valid_vals = [v for v in vals if v is not None]
                    best = max(valid_vals) if _higher_is_better and valid_vals else (min(valid_vals) if _lower_is_better and valid_vals else None)
                    for v in vals:
                        try:
                            display = fmt.format(v) if v is not None else "—"
                        except Exception:
                            display = str(v) if v is not None else "—"
                        is_best = (v == best and best is not None)
                        _cell_color = "#1A4731" if is_best else "#1E2D3D"
                        _cell_weight = "700" if is_best else "400"
                        _row += f'<td style="background:{bg_row};color:{_cell_color};padding:8px 14px;font-size:12px;font-weight:{_cell_weight};border:1px solid #E0DAD0;text-align:right;">{display}</td>'
                    _rows += f"<tr>{_row}</tr>"

                st.markdown(
                    f'<div style="overflow-x:auto;margin-bottom:20px;">'
                    f'<table style="width:100%;border-collapse:collapse;">'
                    f'<thead><tr>{_th}</tr></thead><tbody>{_rows}</tbody></table></div>',
                    unsafe_allow_html=True
                )
                st.caption("Verde = mejor valor del escenario para ese indicador.")

                # Bar chart: yield neto comparison
                if len(cmp_list) > 1:
                    fig_cmp_multi = go.Figure(data=[
                        go.Bar(name=e["label"],
                               x=["Yield Neto %", "TIR Equity %", "DSCR x10"],
                               y=[e["r"].get("yield_neto", 0),
                                  e["r"].get("irr_anual") or 0,
                                  (e["r"].get("dscr") or 0) * 10],
                               text=[f"{e['r'].get('yield_neto',0):.1f}%",
                                     f"{e['r'].get('irr_anual') or 0:.1f}%",
                                     f"{(e['r'].get('dscr') or 0):.2f}x"],
                               textposition="outside")
                        for e in cmp_list
                    ])
                    fig_cmp_multi.update_layout(
                        barmode="group", height=300, margin=dict(t=20, b=20, l=10, r=10),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        yaxis=dict(tickfont=dict(color="#4A5568", size=10), showgrid=True, gridcolor="#E8E4DC"),
                        xaxis=dict(tickfont=dict(color="#4A5568", size=11)),
                        legend=dict(font=dict(color="#4A5568", size=11)),
                        font=dict(family="Inter", color="#4A5568"),
                        colorway=["#1E2D3D", "#B8904A", "#8A9BAD"],
                    )
                    st.plotly_chart(fig_cmp_multi, use_container_width=True)

            st.markdown('<div class="section-title">Comprar vs. Arrendar</div>', unsafe_allow_html=True)
            if (r.get('renta_m2_mes') or 0) > 0:
                APRECIACION = r.get('APRECIACION_IND', 0.03)
                plazo = max(r.get('plazo_anos', 0) or 0, 5)
                renta_anual = (r.get('renta_total_mes') or 0) * 12
                ahorro_fis_anual = r.get('ahorro_fiscal_anual') or 0
                ahorro_fis_mens  = r.get('ahorro_fiscal_mensual') or 0
                cuota_ef         = r.get('cuota_efectiva_mensual') or 0

                # Costo acumulado arrendamiento (solo egreso puro)
                cum_alq = [renta_anual * y for y in range(1, plazo + 1)]

                # Costo acumulado compra bruto (capital propio + servicio deuda)
                cum_compra_bruta = []
                for y in range(1, plazo + 1):
                    if (r.get('cuota_mensual') or 0) > 0:
                        cum_compra_bruta.append((r.get('capital_propio') or 0) + (r.get('cuota_mensual') or 0) * 12 * y)
                    else:
                        cum_compra_bruta.append(r.get('costo_total') or 0)

                # Costo neto compra (– escudo fiscal 20 años) + valor del activo al año y
                cum_compra_neta  = []
                cum_valor_activo = []
                for y in range(1, plazo + 1):
                    escudo = ahorro_fis_anual * min(y, 20)      # hasta 20 años de depreciación
                    cum_compra_neta.append(cum_compra_bruta[y-1] - escudo)
                    cum_valor_activo.append((r.get('costo_total') or 0) * (1 + APRECIACION) ** y)

                anos_range = list(range(1, plazo + 1))

                # ── Métricas clave ────────────────────────────────────────────────
                cv1, cv2, cv3, cv4 = st.columns(4)
                cv1.metric("Cuota mensual", f"${r.get('cuota_mensual', 0):,.0f}/mes" if (r.get('cuota_mensual') or 0) > 0 else "Al contado")
                cv2.metric("Cuota efectiva (post-impuestos)", f"${cuota_ef:,.0f}/mes",
                           f"−${ahorro_fis_mens:,.0f}/mes escudo fiscal")
                cv3.metric("Renta de mercado equiv.", f"${r.get('renta_total_mes', 0):,.0f}/mes")
                dif = (r.get('renta_total_mes') or 0) - cuota_ef
                cv4.metric("Ahorro vs. arrendar", f"${dif:,.0f}/mes" if dif >= 0 else f"−${abs(dif):,.0f}/mes",
                           "comprar es más barato" if dif >= 0 else "arrendar es más barato aún")

                # ── Panel escudo fiscal ───────────────────────────────────────────
                st.markdown(f"""
                <div style="background:#F7F5F1;border:1px solid #D8D4CC;border-left:4px solid #B8904A;
                            border-radius:6px;padding:14px 20px;margin:12px 0;">
                    <div style="font-size:9px;color:#B8904A;letter-spacing:2px;text-transform:uppercase;
                                font-weight:700;margin-bottom:8px;">Beneficio Fiscal de la Propiedad</div>
                    <div style="display:flex;gap:32px;flex-wrap:wrap;">
                        <div><div style="font-size:10px;color:#7A7268;">Base depreciable (nave)</div>
                             <div style="font-size:14px;font-weight:700;color:#1E2D3D;">${r.get('costo_nave_total', 0):,.0f}</div></div>
                        <div><div style="font-size:10px;color:#7A7268;">Depreciación anual (5%/año · 20 años)</div>
                             <div style="font-size:14px;font-weight:700;color:#1E2D3D;">${r.get('depreciacion_anual', 0):,.0f}/año</div></div>
                        <div><div style="font-size:10px;color:#7A7268;">Ahorro IR anual (29.5%)</div>
                             <div style="font-size:14px;font-weight:700;color:#1A4731;">${ahorro_fis_anual:,.0f}/año</div></div>
                        <div><div style="font-size:10px;color:#7A7268;">Ahorro fiscal acumulado (20 años)</div>
                             <div style="font-size:14px;font-weight:700;color:#1A4731;">${ahorro_fis_anual * 20:,.0f}</div></div>
                    </div>
                    <div style="font-size:11px;color:#7A7268;margin-top:8px;line-height:1.5;">
                        El terreno no deprecia. Base: DS 122-94-EF · Tasa industrial 5% anual · IR corporativo 29.5%.
                        La compra del inmueble genera además <strong>plusvalía del activo</strong>,
                        <strong>capital societario respaldado</strong> y posibilidad de <strong>crédito fiscal</strong>.
                    </div>
                </div>""", unsafe_allow_html=True)

                # ── Gráfico comparativo ───────────────────────────────────────────
                fig_cmp = go.Figure()
                fig_cmp.add_trace(go.Scatter(
                    x=anos_range, y=cum_alq,
                    name="Arrendamiento acumulado",
                    line=dict(color="#B8904A", width=2.5), mode="lines",
                    hovertemplate="Año %{x}: $%{y:,.0f}<extra>Arrendamiento</extra>"))
                fig_cmp.add_trace(go.Scatter(
                    x=anos_range, y=cum_compra_bruta,
                    name="Compra (egreso bruto)",
                    line=dict(color="#8A9BAD", width=2, dash="dot"), mode="lines",
                    hovertemplate="Año %{x}: $%{y:,.0f}<extra>Compra bruta</extra>"))
                fig_cmp.add_trace(go.Scatter(
                    x=anos_range, y=cum_compra_neta,
                    name="Compra (neto escudo fiscal)",
                    line=dict(color="#1E2D3D", width=2.5), mode="lines",
                    hovertemplate="Año %{x}: $%{y:,.0f}<extra>Compra neta</extra>"))
                fig_cmp.add_trace(go.Scatter(
                    x=anos_range, y=cum_valor_activo,
                    name="Valor del activo (3% apreciación)",
                    line=dict(color="#1A4731", width=2, dash="dash"), mode="lines",
                    hovertemplate="Año %{x}: $%{y:,.0f}<extra>Valor activo</extra>"))
                fig_cmp.update_layout(
                    height=340, margin=dict(t=20, b=20, l=10, r=10),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(title=dict(text="Año", font=dict(color="#4A5568", size=11)),
                               tickfont=dict(color="#4A5568")),
                    yaxis=dict(title=dict(text="USD", font=dict(color="#4A5568", size=11)),
                               tickfont=dict(color="#4A5568"), tickformat="$,.0f"),
                    legend=dict(font=dict(color="#4A5568", size=11), bgcolor="rgba(255,255,255,0.8)"),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_cmp, use_container_width=True)

                # ── Resumen patrimonial al final del plazo ────────────────────────
                val_final = cum_valor_activo[-1]
                egreso_neto = cum_compra_neta[-1]
                patrimonio_neto = val_final - egreso_neto
                st.markdown(f"""
                <div style="background:#1E2D3D;border-radius:6px;padding:14px 20px;margin-top:4px;">
                    <div style="font-size:9px;color:#B8904A;letter-spacing:2px;text-transform:uppercase;
                                font-weight:700;margin-bottom:10px;">Posición Patrimonial al Año {plazo}</div>
                    <div style="display:flex;gap:32px;flex-wrap:wrap;">
                        <div><div style="font-size:10px;color:#8AA8C0;">Valor del activo</div>
                             <div style="font-size:16px;font-weight:700;color:#FFFFFF;">${val_final:,.0f}</div></div>
                        <div><div style="font-size:10px;color:#8AA8C0;">Egreso neto (compra − escudo fiscal)</div>
                             <div style="font-size:16px;font-weight:700;color:#FFFFFF;">${egreso_neto:,.0f}</div></div>
                        <div><div style="font-size:10px;color:#8AA8C0;">Arrendamiento pagado (sin activo)</div>
                             <div style="font-size:16px;font-weight:700;color:#B8904A;">${cum_alq[-1]:,.0f}</div></div>
                        <div><div style="font-size:10px;color:#8AA8C0;">Patrimonio neto generado</div>
                             <div style="font-size:16px;font-weight:700;color:#{'6BAE90' if patrimonio_neto > 0 else 'E07070'};">${patrimonio_neto:,.0f}</div></div>
                    </div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown('<div class="alert-legal">Ingresa la renta de mercado equivalente en el panel izquierdo para activar la comparativa Comprar vs. Arrendar.</div>', unsafe_allow_html=True)

        # TAB 6: MEMORANDUM ADVISORY BOARD
        with ind_tabs[6]:
            memo = st.session_state.get("ind_resumen")
            if not memo:
                st.markdown(
                    '<div style="background:#F7F5F1;border:1px solid #D8D4CC;border-radius:8px;'
                    'padding:36px 32px;text-align:center;margin-top:8px;">'
                    '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                    'font-weight:600;margin-bottom:12px;">Advisory Board</div>'
                    '<div style="font-size:16px;font-weight:600;color:#1E2D3D;margin-bottom:8px;">'
                    'Memorandum de Análisis</div>'
                    '<div style="width:36px;height:2px;background:#B8904A;margin:12px auto;"></div>'
                    '<div style="font-size:13px;color:#7A7268;line-height:1.7;max-width:520px;margin:0 auto 24px;">'
                    'Genera el memorandum de Advisory Board: consolidación de la data analítica del proyecto — '
                    'indicadores, estructura de costos, financiamiento y contexto de mercado — '
                    'para que el cliente evalúe la oportunidad según sus propios criterios y estrategia.'
                    '</div>',
                    unsafe_allow_html=True
                )
                st.markdown("</div>", unsafe_allow_html=True)
            if st.button("GENERAR MEMORANDUM", use_container_width=True, type="primary", key="btn_ind_rsm"):
                _r_copy = dict(r)
                st.session_state.ind_resumen = _run_with_retry(
                    lambda _rc=_r_copy: generar_memorandum_advisory_ind(_rc),
                    "Generando memorandum de Advisory Board…"
                )
                st.rerun()
            if memo:
                st.markdown(f"""
                <div style="border-bottom:2px solid #B8904A;padding-bottom:16px;margin-bottom:24px;">
                    <div style="font-size:9px;color:#B8904A;letter-spacing:4px;text-transform:uppercase;
                                font-weight:600;margin-bottom:6px;">Osterling Advisory · Memorandum de Advisory Board</div>
                    <div style="font-size:20px;font-weight:700;color:#1E2D3D;">
                        {memo.get('titulo','Análisis Industrial')}</div>
                </div>""", unsafe_allow_html=True)

                st.markdown(
                    '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                    'font-weight:600;margin-bottom:8px;">Perfil del Activo</div>',
                    unsafe_allow_html=True
                )
                st.markdown(
                    f'<div style="font-size:13px;color:#1E2D3D;line-height:1.8;margin-bottom:20px;">'
                    f'{memo.get("perfil_activo","")}</div>',
                    unsafe_allow_html=True
                )

                _mc1, _mc2 = st.columns(2)
                with _mc1:
                    st.markdown(
                        '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                        'font-weight:600;margin-bottom:8px;">Indicadores Clave</div>',
                        unsafe_allow_html=True
                    )
                    st.markdown(
                        f'<div style="background:#F7F5F1;border-left:3px solid #B8904A;border-radius:4px;'
                        f'padding:14px 16px;font-size:13px;color:#1E2D3D;line-height:1.8;">'
                        f'{memo.get("indicadores_clave","")}</div>',
                        unsafe_allow_html=True
                    )
                with _mc2:
                    st.markdown(
                        '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                        'font-weight:600;margin-bottom:8px;">Posicionamiento de Mercado</div>',
                        unsafe_allow_html=True
                    )
                    st.markdown(
                        f'<div style="background:#F7F5F1;border-left:3px solid #1E2D3D;border-radius:4px;'
                        f'padding:14px 16px;font-size:13px;color:#1E2D3D;line-height:1.8;">'
                        f'{memo.get("posicionamiento_mercado","")}</div>',
                        unsafe_allow_html=True
                    )

                st.markdown('<div style="height:18px;"></div>', unsafe_allow_html=True)

                st.markdown(
                    '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                    'font-weight:600;margin-bottom:8px;margin-top:4px;">Estructura Financiera</div>',
                    unsafe_allow_html=True
                )
                st.markdown(
                    f'<div style="font-size:13px;color:#1E2D3D;line-height:1.8;margin-bottom:20px;">'
                    f'{memo.get("estructura_financiera","")}</div>',
                    unsafe_allow_html=True
                )

                _mf1, _mf2 = st.columns(2)
                with _mf1:
                    st.markdown(
                        '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                        'font-weight:600;margin-bottom:8px;">Factores Relevantes</div>',
                        unsafe_allow_html=True
                    )
                    for _fitem in (memo.get("factores_relevantes") or []):
                        st.markdown(
                            f'<div style="font-size:12px;color:#1E2D3D;padding:8px 12px;'
                            f'border-bottom:1px solid #E8E4DC;line-height:1.6;">→ {_fitem}</div>',
                            unsafe_allow_html=True
                        )
                with _mf2:
                    st.markdown(
                        '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                        'font-weight:600;margin-bottom:8px;">Consideraciones</div>',
                        unsafe_allow_html=True
                    )
                    for _citem in (memo.get("consideraciones") or []):
                        st.markdown(
                            f'<div style="font-size:12px;color:#1E2D3D;padding:8px 12px;'
                            f'border-bottom:1px solid #E8E4DC;line-height:1.6;">· {_citem}</div>',
                            unsafe_allow_html=True
                        )

                st.markdown('<div style="height:18px;"></div>', unsafe_allow_html=True)

                st.markdown(
                    f'<div style="background:#1E2D3D;border-radius:6px;padding:20px 24px;margin-top:4px;">'
                    f'<div style="font-size:9px;color:#B8904A;letter-spacing:2px;text-transform:uppercase;'
                    f'font-weight:600;margin-bottom:10px;">Síntesis del Análisis</div>'
                    f'<div style="font-size:13px;color:#FFFFFF;line-height:1.8;">'
                    f'{memo.get("sintesis","")}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

                if st.button("REGENERAR", key="btn_ind_rsm_regen"):
                    st.session_state.ind_resumen = None
                    st.rerun()

        # TAB 5: FACTIBILIDAD
        with ind_tabs[5]:
            fac = st.session_state.get("industrial_factibilidad")
            if not fac:
                st.markdown(
                    '<div style="background:#F7F5F1;border:1px solid #D8D4CC;border-radius:8px;'
                    'padding:36px 32px;text-align:center;margin-top:8px;">'
                    '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                    'font-weight:600;margin-bottom:12px;">Análisis Opcional</div>'
                    '<div style="font-size:16px;font-weight:600;color:#1E2D3D;margin-bottom:8px;">'
                    'Factibilidad Técnica y Legal</div>'
                    '<div style="width:36px;height:2px;background:#B8904A;margin:12px auto;"></div>'
                    '<div style="font-size:13px;color:#7A7268;line-height:1.7;max-width:480px;margin:0 auto;">'
                    'Adjunta la <strong>Partida Registral</strong>, el <strong>Certificado de Parámetros</strong> '
                    'y/o el <strong>Certificado de Zonificación y Vías</strong> en el panel izquierdo, '
                    'luego presiona <strong>ANALIZAR DOCUMENTOS</strong> para obtener:'
                    '<br><br>✓ Compatibilidad de zonificación con la actividad<br>'
                    '✓ Restricciones técnicas de acceso y altura<br>'
                    '✓ Estado registral: cargas, hipotecas, medidas cautelares'
                    '</div></div>',
                    unsafe_allow_html=True
                )
            else:
                sem_g = fac.get("semaforo_global", "amarillo").lower()
                sem_t = fac.get("semaforo_tecnico", "amarillo").lower()
                sem_l = fac.get("semaforo_legal", "amarillo").lower()

                _SEM = {
                    "verde":    ("#1A4731", "#E8F5EE"),
                    "amarillo": ("#7A4F1A", "#FFF8EE"),
                    "rojo":     ("#7A1A1A", "#FFF0F0"),
                }
                _ETIQ = {
                    "verde": "SIN ALERTAS CRÍTICAS",
                    "amarillo": "OBSERVACIONES",
                    "rojo": "ALERTAS CRÍTICAS",
                }
                gc, gbg = _SEM.get(sem_g, ("#1E2D3D", "#F5F2ED"))

                # Semáforo global
                st.markdown(f"""
                <div style="background:{gbg};border:1px solid {gc};border-left:4px solid {gc};
                            border-radius:8px;padding:20px 24px;margin-bottom:20px;">
                    <div style="font-size:9px;letter-spacing:3px;color:{gc};text-transform:uppercase;
                                font-weight:700;opacity:0.7;margin-bottom:6px;">Evaluación Global</div>
                    <div style="font-size:20px;font-weight:700;color:{gc};margin-bottom:10px;">{_ETIQ.get(sem_g,'—')}</div>
                    <div style="display:flex;gap:20px;flex-wrap:wrap;margin-bottom:12px;">
                        <div style="font-size:11px;color:{_SEM.get(sem_t,('',''))[0]};background:{_SEM.get(sem_t,('','#F5F2ED'))[1]};
                                    padding:4px 12px;border-radius:4px;font-weight:600;">
                            ● Técnico: {sem_t.upper()}</div>
                        <div style="font-size:11px;color:{_SEM.get(sem_l,('',''))[0]};background:{_SEM.get(sem_l,('','#F5F2ED'))[1]};
                                    padding:4px 12px;border-radius:4px;font-weight:600;">
                            ● Legal: {sem_l.upper()}</div>
                    </div>
                    <div style="font-size:13px;color:{gc};opacity:0.85;line-height:1.6;margin-bottom:6px;">
                        {fac.get('resumen_tecnico','—')}</div>
                    <div style="font-size:13px;color:{gc};opacity:0.85;line-height:1.6;">
                        {fac.get('resumen_legal','—')}</div>
                </div>""", unsafe_allow_html=True)

                # ANÁLISIS TÉCNICO
                st.markdown('<div class="section-title">Factibilidad Técnica</div>', unsafe_allow_html=True)
                ft1, ft2 = st.columns(2)
                _zon_cert = fac.get("zonificacion_certificada") or "—"
                _compat = fac.get("compatible_actividad")
                _compat_icon = "✓" if _compat is True else ("✗" if _compat is False else "—")
                _compat_col = "#1A4731" if _compat is True else ("#7A1A1A" if _compat is False else "#9A9080")
                ft1.metric("Zonificación certificada", _zon_cert)
                ft2.metric("Compatible con actividad", _compat_icon)

                nota_compat = fac.get("nota_compatibilidad", "")
                if nota_compat:
                    st.markdown(f'<div class="alert-gold">{nota_compat}</div>', unsafe_allow_html=True)

                # Actividades
                act_perm  = fac.get("actividades_permitidas", []) or []
                act_cond  = fac.get("actividades_condicionadas", []) or []
                act_proh  = fac.get("actividades_prohibidas", []) or []
                if act_perm or act_cond or act_proh:
                    fa1, fa2, fa3 = st.columns(3)
                    with fa1:
                        st.markdown('<div style="font-size:9px;color:#1A4731;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-bottom:6px;">Permitidas</div>', unsafe_allow_html=True)
                        for a in act_perm:
                            st.markdown(f'<div style="font-size:12px;color:#1E2D3D;padding:3px 0;">✓ {a}</div>', unsafe_allow_html=True)
                    with fa2:
                        st.markdown('<div style="font-size:9px;color:#7A4F1A;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-bottom:6px;">Condicionadas</div>', unsafe_allow_html=True)
                        for a in act_cond:
                            st.markdown(f'<div style="font-size:12px;color:#1E2D3D;padding:3px 0;">⚠ {a}</div>', unsafe_allow_html=True)
                    with fa3:
                        st.markdown('<div style="font-size:9px;color:#7A1A1A;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-bottom:6px;">Prohibidas</div>', unsafe_allow_html=True)
                        for a in act_proh:
                            st.markdown(f'<div style="font-size:12px;color:#1E2D3D;padding:3px 0;">✗ {a}</div>', unsafe_allow_html=True)

                # Restricciones
                restr = fac.get("restricciones_especiales", []) or []
                alt_max = fac.get("restricciones_altura_m")
                acceso_pesado = fac.get("acceso_vehiculos_pesados")
                vias = fac.get("vias_frente", []) or []
                if alt_max or acceso_pesado is not None or restr or vias:
                    st.markdown('<div class="section-title">Restricciones y Accesos</div>', unsafe_allow_html=True)
                    fr1, fr2, fr3 = st.columns(3)
                    fr1.metric("Altura máxima", f"{alt_max} m" if alt_max else "No especificada")
                    _acc_icon = "✓ Permitido" if acceso_pesado is True else ("✗ Restringido" if acceso_pesado is False else "—")
                    fr2.metric("Vehículos pesados", _acc_icon)
                    fr3.metric("Vías de frente", str(len(vias)) if vias else "—")
                    if vias:
                        for v in vias:
                            _ancho = f" · {v.get('ancho_ml')} ml" if v.get('ancho_ml') else ""
                            st.markdown(f'<div class="alert-legal">'
                                        f'<strong>{v.get("nombre","—")}</strong> — {v.get("tipo","—")}{_ancho}</div>',
                                        unsafe_allow_html=True)
                    for rs in restr:
                        st.markdown(f'<div class="alert-gold">⚠ {rs}</div>', unsafe_allow_html=True)

                # ── Identificación registral rápida ──────────
                prop_l = fac.get("propietarios_partida", []) or []
                area_r = fac.get("area_registral_m2")
                partida_n = fac.get("partida_numero")
                dir_r = fac.get("direccion_partida")
                if partida_n or area_r or prop_l:
                    st.markdown('<div class="section-title">Identificación Registral</div>', unsafe_allow_html=True)
                    fl1, fl2, fl3 = st.columns(3)
                    fl1.metric("N° Partida", partida_n or "—")
                    fl2.metric("Área registral", f"{area_r:,.2f} m²" if area_r else "—")
                    fl3.metric("Propietario(s)", str(len(prop_l)) if prop_l else "—")
                    if prop_l:
                        st.markdown('<div class="alert-legal"><strong>Titular(es):</strong> ' +
                                    ' · '.join(str(p.get("nombre", p) if isinstance(p, dict) else p) for p in prop_l) +
                                    '</div>', unsafe_allow_html=True)
                    if dir_r:
                        st.markdown(f'<div class="alert-legal"><strong>Dirección registral:</strong> {dir_r}</div>',
                                    unsafe_allow_html=True)

                # ── Checklist de hallazgos ────────────────────
                ind_hallazgos = fac.get("hallazgos") or []
                ind_completitud = fac.get("completitud") or {}
                if ind_hallazgos:
                    _icv = ind_completitud.get("verificados", len(ind_hallazgos))
                    _ict = ind_completitud.get("total", 16)
                    _icnv = ind_completitud.get("no_verificables") or []
                    _icv_pct = round(_icv / _ict * 100) if _ict else 0
                    _icv_color = "#1A4731" if _icv_pct >= 80 else ("#7A4F1A" if _icv_pct >= 60 else "#7A1A1A")
                    st.markdown(
                        f'<div style="display:flex;justify-content:space-between;align-items:center;'
                        f'background:#F7F5F1;border:1px solid #E4E0D8;border-radius:8px;'
                        f'padding:12px 20px;margin:16px 0;">'
                        f'<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:700;">Completitud del análisis</div>'
                        f'<div style="font-size:16px;font-weight:700;color:{_icv_color};">{_icv}/{_ict} puntos verificados</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                    _ISEV = {
                        "verde":         ("🟢", "#1A4731", "#E8F5EE", "#C3E6CB"),
                        "amarillo":      ("🟡", "#7A4F1A", "#FFF8EE", "#FFE0A0"),
                        "rojo":          ("🔴", "#7A1A1A", "#FFF0F0", "#F5C6CB"),
                        "no_verificable":("⚪", "#6A6A6A", "#F5F5F5", "#D8D4CC"),
                    }
                    ind_tecnico_h = [h for h in ind_hallazgos if h.get("categoria") == "tecnico"]
                    ind_legal_h   = [h for h in ind_hallazgos if h.get("categoria") == "legal"]

                    def _render_ind_hallazgos(h_list, section_lbl):
                        if not h_list:
                            return
                        st.markdown(f'<div class="section-title">{section_lbl}</div>', unsafe_allow_html=True)
                        for h in h_list:
                            sev = (h.get("severidad") or "no_verificable").lower()
                            icon, tcol, bg, brd = _ISEV.get(sev, _ISEV["no_verificable"])
                            num = h.get("numero", "")
                            punto = h.get("punto", "—")
                            hallazgo_txt = h.get("hallazgo", "—")
                            subsanacion = h.get("subsanacion")
                            sev_lbl = {"verde": "OK", "amarillo": "VERIFICAR", "rojo": "CRÍTICO", "no_verificable": "NO VERIFICABLE"}.get(sev, sev.upper())
                            sub_html = ""
                            if subsanacion:
                                sub_html = (
                                    f'<div style="margin-top:10px;padding-top:10px;border-top:1px solid {brd};">'
                                    f'<div style="font-size:9px;color:{tcol};letter-spacing:1.5px;text-transform:uppercase;font-weight:700;margin-bottom:4px;">Procedimiento de subsanación</div>'
                                    f'<div style="font-size:12px;color:#2A3A4A;line-height:1.6;">{subsanacion}</div>'
                                    f'</div>'
                                )
                            st.markdown(
                                f'<div style="background:{bg};border:1px solid {brd};border-left:4px solid {tcol};'
                                f'border-radius:6px;padding:14px 18px;margin-bottom:10px;">'
                                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px;">'
                                f'<div style="font-size:9px;color:{tcol};letter-spacing:1.5px;text-transform:uppercase;font-weight:700;">'
                                f'{icon} {num}. {punto}</div>'
                                f'<div style="font-size:8px;font-weight:700;color:{tcol};background:rgba(0,0,0,0.07);'
                                f'padding:2px 8px;border-radius:3px;white-space:nowrap;">{sev_lbl}</div>'
                                f'</div>'
                                f'<div style="font-size:12px;color:#1E2D3D;line-height:1.6;">{hallazgo_txt}</div>'
                                f'{sub_html}'
                                f'</div>',
                                unsafe_allow_html=True
                            )

                    _render_ind_hallazgos(ind_tecnico_h, "Análisis Técnico — Certificado de Parámetros / Zonificación")
                    _render_ind_hallazgos(ind_legal_h,   "Análisis Legal — Partida Registral SUNARP")

                    if _icnv:
                        with st.expander(f"Puntos no verificables ({len(_icnv)}) — documento no adjunto o ilegible"):
                            for nv in _icnv:
                                st.markdown(f'<div style="font-size:12px;color:#9A9080;padding:3px 0;">⚪ {nv}</div>', unsafe_allow_html=True)
                else:
                    # Fallback: legacy alertas
                    alertas_t = fac.get("alertas_tecnicas", []) or []
                    alertas_l = fac.get("alertas_legales", []) or []
                    if alertas_t or alertas_l:
                        st.markdown('<div class="section-title">Alertas</div>', unsafe_allow_html=True)
                        for al in alertas_t:
                            st.markdown(f'<div class="alert-gold">⚠ [TÉCNICO] {al}</div>', unsafe_allow_html=True)
                        for al in alertas_l:
                            _al_style = "alert-gold" if sem_l == "amarillo" else "alert-legal"
                            _icon_al = "🔴" if sem_l == "rojo" else "🟡"
                            st.markdown(f'<div class="{_al_style}">{_icon_al} [LEGAL] {al}</div>', unsafe_allow_html=True)

        st.markdown("""
<div style="border-top:1px solid #E8E0D4;margin-top:32px;padding-top:14px;">
  <div style="font-size:9px;color:#9A8A7A;line-height:1.7;text-align:center;">
    <span style="color:#B8904A;font-weight:700;letter-spacing:1px;">NOTA · </span>
    Esta IA de Análisis Inmobiliario debe utilizarse como herramienta complementaria al criterio profesional,
    permitiendo obtener resultados preliminares de manera rápida. Como paso final, el profesional podrá
    terminar de definir las tipologías, distribución por plantas y las modificaciones que considere pertinentes.
    La IA irá volviéndose más responsiva y alineada con la visión del profesional a medida que se retroalimenta
    con sus decisiones.
  </div>
</div>""", unsafe_allow_html=True)

    else:
        st.markdown(
            '<div style="border-radius:8px;min-height:420px;'
            'background:linear-gradient(160deg,#1A2737 0%,#1E2D3D 60%,#1A2737 100%);'
            'display:flex;align-items:center;justify-content:center;'
            'box-shadow:0 8px 32px rgba(30,45,61,0.18);padding:64px 48px;">'
            '<div style="max-width:600px;width:100%;text-align:center;">'
            '<div style="font-size:9px;color:#B8904A;letter-spacing:4px;text-transform:uppercase;font-weight:600;margin-bottom:16px;">Osterling Advisory</div>'
            '<div style="font-size:28px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;margin-bottom:8px;">Proyecto Logístico / Industrial</div>'
            '<div style="width:48px;height:2px;background:#B8904A;margin:16px auto;"></div>'
            '<div style="font-size:13px;color:#B0C0D0;line-height:1.7;margin-bottom:32px;">'
            'Evalúa la viabilidad de compra de un activo industrial o logístico. '
            'Ingresa los datos del terreno, tipo de nave y condiciones de financiamiento '
            'para obtener: costo total del proyecto, cuota mensual, yield, payback '
            'y comparativa compra vs. arrendamiento.'
            '</div>'
            '<div style="display:flex;gap:24px;justify-content:center;flex-wrap:wrap;">'
            '<div style="background:rgba(184,144,74,0.1);border:1px solid rgba(184,144,74,0.3);'
            'border-radius:6px;padding:14px 20px;min-width:120px;">'
            '<div style="font-size:20px;font-weight:700;color:#B8904A;">I1–I4</div>'
            '<div style="font-size:10px;color:#8AA8C0;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px;">Zonificaciones</div>'
            '</div>'
            '<div style="background:rgba(184,144,74,0.1);border:1px solid rgba(184,144,74,0.3);'
            'border-radius:6px;padding:14px 20px;min-width:120px;">'
            '<div style="font-size:20px;font-weight:700;color:#B8904A;">4</div>'
            '<div style="font-size:10px;color:#8AA8C0;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px;">Tipos de Nave</div>'
            '</div>'
            '<div style="background:rgba(184,144,74,0.1);border:1px solid rgba(184,144,74,0.3);'
            'border-radius:6px;padding:14px 20px;min-width:120px;">'
            '<div style="font-size:20px;font-weight:700;color:#B8904A;">Yield</div>'
            '<div style="font-size:10px;color:#8AA8C0;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px;">+ DSCR + Payback</div>'
            '</div>'
            '</div>'
            '</div></div>',
            unsafe_allow_html=True
        )

# ═══════════════════════════════════════════════════════
# MÓDULO 3: INMUEBLE RESIDENCIAL
# ═══════════════════════════════════════════════════════

elif tipo_op == "Inmueble Residencial":
    r = st.session_state.get("residencial_result")

    if r:
        _rc1, _rc2, _rc3 = st.columns([2, 1, 1])
        with _rc1:
            _res_cmp_label = st.text_input("Nombre del escenario", placeholder="Ej: Depto Miraflores 80m²",
                                            label_visibility="collapsed", key="res_cmp_label")
        with _rc2:
            if st.button("GUARDAR EN COMPARATIVA", use_container_width=True, key="btn_res_cmp"):
                _lbl = _res_cmp_label.strip() or f"Escenario {len(st.session_state.res_comparativa)+1}"
                _entry = {"label": _lbl, "r": dict(r)}
                existing = [e["label"] for e in st.session_state.res_comparativa]
                if _lbl in existing:
                    st.session_state.res_comparativa[existing.index(_lbl)] = _entry
                elif len(st.session_state.res_comparativa) < 3:
                    st.session_state.res_comparativa.append(_entry)
                else:
                    st.session_state.res_comparativa[2] = _entry
                st.toast(f"✓ '{_lbl}' guardado en comparativa")
        with _rc3:
            if st.session_state.res_comparativa and st.button("LIMPIAR COMPARATIVA", use_container_width=True, key="btn_res_cmp_clear"):
                st.session_state.res_comparativa = []
                st.rerun()

        # ── Property Hero Banner ─────────────────────────────
        import base64 as _b64h
        _hero_fotos = st.session_state.get("res_fotos_bytes") or []
        _hero_zona  = r.get("zona", "—")
        _hero_precio = r.get("precio", 0)
        _hero_m2    = r.get("m2", 0)
        _hero_dorm  = r.get("dormitorios", "—")
        _hero_ppm2  = r.get("precio_m2", 0)
        _hero_yield = r.get("yield_bruto", 0)

        if _hero_fotos:
            try:
                _b64_hero = _b64h.b64encode(_hero_fotos[0]).decode()
                _hero_photo_css = f"url('data:image/jpeg;base64,{_b64_hero}') center/cover no-repeat"
            except Exception:
                _hero_photo_css = "linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%)"
        else:
            _hero_photo_css = "linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%)"

        st.markdown(f"""
        <div style="position:relative;border-radius:16px;overflow:hidden;margin-bottom:20px;
                    box-shadow:0 6px 30px rgba(30,45,61,0.22);">
            <div style="background:{_hero_photo_css};height:220px;"></div>
            <div style="position:absolute;inset:0;background:linear-gradient(to bottom,
                        rgba(0,0,0,0.0) 0%,rgba(0,0,0,0.75) 100%);
                        display:flex;flex-direction:column;justify-content:flex-end;
                        padding:24px 28px;">
                <div style="font-size:9px;color:rgba(255,255,255,0.60);letter-spacing:3px;
                            text-transform:uppercase;margin-bottom:6px;">
                    Análisis de Inmueble Residencial · FACTIS
                </div>
                <div style="font-size:28px;font-weight:800;color:#FFFFFF;line-height:1.15;
                            text-shadow:0 2px 8px rgba(0,0,0,0.5);">
                    {_hero_zona} &nbsp;·&nbsp; {_hero_dorm}
                </div>
                <div style="display:flex;gap:20px;margin-top:12px;flex-wrap:wrap;">
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;
                                    letter-spacing:1px;">Precio</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">${_hero_precio:,}</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;
                                    letter-spacing:1px;">Área</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_hero_m2} m²</div>
                    </div>
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;
                                    letter-spacing:1px;">USD/m²</div>
                        <div style="font-size:20px;font-weight:700;color:#FFFFFF;">${_hero_ppm2:,}</div>
                    </div>
                    {"" if _hero_yield == 0 else f'<div><div style="font-size:9px;color:rgba(255,255,255,0.55);text-transform:uppercase;letter-spacing:1px;">Yield bruto</div><div style="font-size:20px;font-weight:700;color:#FFFFFF;">{_hero_yield:.1f}%</div></div>'}
                </div>
            </div>
        </div>""", unsafe_allow_html=True)

        res_tab_labels = ["Parámetros", "Financiero", "Inversión"] if r.get('uso') in ["Inversión para alquilar", "Evaluación para venta"] else ["Parámetros", "Financiero", "Escenarios"]
        res_tabs = st.tabs(res_tab_labels + ["Comparativa", "Amortización", "Legal", "Resumen", "Documentos"])

        # TAB 0: MERCADO
        with res_tabs[0]:
            # ── District Profile Hero ─────────────────────────────
            _zona_key    = r.get("zona", "")
            _m_data      = MERCADO.get(_zona_key, {})
            _p2br        = _m_data.get("precio_2br", 0)
            _p1br        = _m_data.get("precio_1br", 0)
            _p3br        = _m_data.get("precio_3br", 0)
            _yield_z     = _m_data.get("yield_mercado_pct", 0)
            _var_z       = _m_data.get("variacion_anual_pct", 0)
            _alq_z       = _m_data.get("alquiler_m2_mes", 0)

            # Tier classification
            if _p2br >= 2500:   _tier_name, _tier_col, _tier_bg = "PREMIUM",       "#FFFFFF", "#1E2D3D"
            elif _p2br >= 2000: _tier_name, _tier_col, _tier_bg = "ALTO",          "#FFFFFF", "#B8904A"
            elif _p2br >= 1500: _tier_name, _tier_col, _tier_bg = "RESIDENCIAL",   "#FFFFFF", "#4A7A6B"
            elif _p2br >= 1000: _tier_name, _tier_col, _tier_bg = "CONSOLIDADO",   "#FFFFFF", "#5A6A7A"
            else:               _tier_name, _tier_col, _tier_bg = "EMERGENTE",     "#FFFFFF", "#8A7A6A"

            # Investment grade
            if _yield_z >= 6 and _var_z >= 3:  _grade, _grade_col = "A",   "#1A4731"
            elif _yield_z >= 5.5 or _var_z >= 2: _grade, _grade_col = "B+", "#3A6A50"
            elif _yield_z >= 5 or _var_z >= 0:  _grade, _grade_col = "B",  "#7A5500"
            else:                                _grade, _grade_col = "C",  "#7A1A1A"

            _trend_icon  = "↑" if _var_z > 1 else ("→" if abs(_var_z) <= 1 else "↓")
            _trend_col   = "#1A4731" if _var_z > 1 else ("#7A5500" if abs(_var_z) <= 1 else "#7A1A1A")

            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%);
                        border-radius:12px;padding:24px 28px;margin-bottom:20px;
                        box-shadow:0 4px 20px rgba(30,45,61,0.20);">
                <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px;">
                    <div>
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);letter-spacing:3px;
                                    text-transform:uppercase;margin-bottom:6px;">Perfil de Distrito · Urbania 2025</div>
                        <div style="font-size:28px;font-weight:800;color:#FFFFFF;line-height:1.1;">{_zona_key}</div>
                        <div style="margin-top:8px;display:flex;gap:8px;flex-wrap:wrap;">
                            <span style="background:{_tier_bg};color:{_tier_col};font-size:9px;font-weight:700;
                                         padding:3px 10px;border-radius:12px;letter-spacing:1.5px;">{_tier_name}</span>
                            <span style="background:rgba(255,255,255,0.12);color:#FFFFFF;font-size:9px;font-weight:700;
                                         padding:3px 10px;border-radius:12px;letter-spacing:1px;">
                                         Alquiler ${_alq_z:.1f}/m²/mes</span>
                        </div>
                    </div>
                    <div style="text-align:center;background:rgba(255,255,255,0.10);border-radius:10px;
                                padding:14px 20px;min-width:80px;">
                        <div style="font-size:9px;color:rgba(255,255,255,0.55);letter-spacing:2px;
                                    text-transform:uppercase;margin-bottom:4px;">Inv. Grade</div>
                        <div style="font-size:36px;font-weight:900;color:#FFFFFF;line-height:1;">{_grade}</div>
                    </div>
                </div>
                <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:20px;">
                    <div style="background:rgba(255,255,255,0.08);border-radius:8px;padding:12px 14px;">
                        <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                                    letter-spacing:1px;margin-bottom:4px;">Precio 1D</div>
                        <div style="font-size:15px;font-weight:700;color:#FFFFFF;">${_p1br:,}/m²</div>
                    </div>
                    <div style="background:rgba(255,255,255,0.08);border-radius:8px;padding:12px 14px;">
                        <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                                    letter-spacing:1px;margin-bottom:4px;">Precio 2D</div>
                        <div style="font-size:15px;font-weight:700;color:#FFFFFF;">${_p2br:,}/m²</div>
                    </div>
                    <div style="background:rgba(255,255,255,0.08);border-radius:8px;padding:12px 14px;">
                        <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                                    letter-spacing:1px;margin-bottom:4px;">Yield Zona</div>
                        <div style="font-size:15px;font-weight:700;color:#FFFFFF;">{_yield_z:.1f}%</div>
                    </div>
                    <div style="background:rgba(255,255,255,0.08);border-radius:8px;padding:12px 14px;">
                        <div style="font-size:9px;color:rgba(255,255,255,0.5);text-transform:uppercase;
                                    letter-spacing:1px;margin-bottom:4px;">Variación 12m</div>
                        <div style="font-size:15px;font-weight:700;color:#FFFFFF;">{_trend_icon} {_var_z:+.1f}%</div>
                    </div>
                </div>
            </div>""", unsafe_allow_html=True)

            # ── Price Gauge ───────────────────────────────────────
            _ppm2_r = r.get("precio_m2", 0)
            _ref_r  = r.get("precio_m2_mercado", 0)
            if _ref_r > 0 and _ppm2_r > 0:
                _gauge_min = int(_ref_r * 0.60)
                _gauge_max = int(_ref_r * 1.50)
                fig_gauge = go.Figure(go.Indicator(
                    mode="gauge+number+delta",
                    value=_ppm2_r,
                    number={"prefix": "$", "suffix": "/m²", "font": {"size": 26, "color": "#1E2D3D"}},
                    delta={"reference": _ref_r, "relative": True, "valueformat": "+.1%",
                           "font": {"size": 13},
                           "increasing": {"color": "#7A1A1A"},
                           "decreasing": {"color": "#1A4731"}},
                    title={"text": "Posición de Precio vs. Mediana de Zona", "font": {"size": 11, "color": "#9A9080"}},
                    gauge={
                        "axis": {"range": [_gauge_min, _gauge_max], "tickformat": "$,.0f",
                                 "tickfont": {"size": 9, "color": "#9A9080"},
                                 "nticks": 6},
                        "bar": {"color": "#B8904A", "thickness": 0.25},
                        "bgcolor": "#F7F5F1",
                        "borderwidth": 0,
                        "steps": [
                            {"range": [_gauge_min, int(_ref_r * 0.92)],    "color": "#E8F5EE"},
                            {"range": [int(_ref_r * 0.92), int(_ref_r * 1.08)], "color": "#FFF8EE"},
                            {"range": [int(_ref_r * 1.08), _gauge_max],   "color": "#FDECEA"},
                        ],
                        "threshold": {"line": {"color": "#1E2D3D", "width": 3},
                                      "thickness": 0.85, "value": _ref_r},
                    }
                ))
                fig_gauge.update_layout(
                    height=240, margin=dict(t=50, b=10, l=30, r=30),
                    paper_bgcolor="rgba(0,0,0,0)", font=dict(family="Inter"),
                )
                st.plotly_chart(fig_gauge, use_container_width=True, config={"displayModeBar": False})
                st.caption("Aguja naranja = precio pagado · Línea negra = mediana de zona (Urbania nov-2025) · Verde = oportunidad · Rojo = sobre mercado")

            # ── Posición de precio (KPIs) ────────────────────────
            _ppm2_r = r.get("precio_m2", 0)
            _ref_r   = r.get("precio_m2_mercado", 0)
            _diff_r  = ((_ppm2_r - _ref_r) / _ref_r * 100) if _ref_r > 0 else 0
            st.markdown('<div class="section-title">Posición de Precio en el Mercado</div>', unsafe_allow_html=True)

            if abs(_diff_r) <= 8:
                _sem_color, _sem_bg, _sem_label = "#1A4731", "#E8F5EE", "EN LÍNEA CON EL MERCADO"
            elif _diff_r > 8:
                _sem_color, _sem_bg, _sem_label = "#7A1A1A", "#FDECEA", "SOBRE EL MERCADO"
            else:
                _sem_color, _sem_bg, _sem_label = "#1A4731", "#E8F5EE", "POR DEBAJO — OPORTUNIDAD"

            pm1, pm2, pm3 = st.columns(3)
            pm1.metric("Precio pagado / m²", f"${_ppm2_r:,.0f}/m²")
            pm2.metric("Mediana zona (Urbania)", f"${_ref_r:,}/m²",
                       delta=f"{_diff_r:+.1f}% vs. mercado")
            pm3.metric("Precio justo estimado", f"${int(r.get('m2',0) * _ref_r):,}")

            st.markdown(
                f'<div style="background:{_sem_bg};border:1px solid {_sem_color};border-left:4px solid {_sem_color};'
                f'border-radius:6px;padding:14px 20px;margin:12px 0;">'
                f'<div style="font-size:9px;color:{_sem_color};letter-spacing:2px;font-weight:700;text-transform:uppercase;">Posición de Mercado</div>'
                f'<div style="font-size:20px;font-weight:800;color:{_sem_color};margin:4px 0;">{_sem_label}</div>'
                f'<div style="font-size:12px;color:{_sem_color};opacity:0.85;">'
                f'El precio pagado (${_ppm2_r:,.0f}/m²) está {abs(_diff_r):.1f}% '
                f'{"sobre" if _diff_r > 0 else "bajo"} la mediana de {r.get("zona","la zona")} (${_ref_r:,}/m²). '
                f'{"Considera negociar." if _diff_r > 8 else ("Precio competitivo — buena entrada." if _diff_r < -8 else "Precio consistente con el mercado.")}'
                f'</div></div>',
                unsafe_allow_html=True
            )

            # ── Alquiler de mercado ─────────────────────────────
            st.markdown('<div class="section-title">Renta de Mercado — Zona</div>', unsafe_allow_html=True)
            _alq_ref_mes = r.get("alquiler_mercado_m2", 0) * r.get("m2", 0)
            _alq_actual = r.get("alquiler_mes", 0)

            am1, am2, am3 = st.columns(3)
            am1.metric("Alquiler mercado estimado", f"${_alq_ref_mes:,.0f}/mes")
            am2.metric("Alquiler ingresado", f"${_alq_actual:,.0f}/mes" if _alq_actual > 0 else "No aplica")
            if _alq_actual > 0 and _alq_ref_mes > 0:
                _diff_alq = (_alq_actual - _alq_ref_mes) / _alq_ref_mes * 100
                am3.metric("Diferencial vs. mercado", f"{_diff_alq:+.1f}%")

            # ── Yield benchmark ─────────────────────────────────
            st.markdown('<div class="section-title">Rentabilidad vs. Mercado</div>', unsafe_allow_html=True)
            _yield_mkt = r.get("yield_mercado_pct", 0)
            _yield_this = r.get("yield_bruto", 0)
            _yield_neto = r.get("yield_neto", 0)

            ym1, ym2, ym3, ym4 = st.columns(4)
            ym1.metric("Yield bruto este inmueble", f"{_yield_this:.1f}%")
            ym2.metric("Yield neto este inmueble", f"{_yield_neto:.1f}%")
            ym3.metric(f"Yield promedio zona", f"{_yield_mkt:.1f}%")
            if _yield_this > 0 and _yield_mkt > 0:
                _beat = _yield_this - _yield_mkt
                ym4.metric("Diferencial", f"{_beat:+.1f}pp",
                           delta="sobre mercado" if _beat >= 0 else "bajo mercado")

            # ── Tendencia del distrito ───────────────────────────
            st.markdown('<div class="section-title">Tendencia del Distrito — Últimos 12 Meses</div>', unsafe_allow_html=True)
            _var = r.get("variacion_anual_pct", 0)
            _var_color = "#1A4731" if _var >= 3 else ("#7A5500" if _var >= 0 else "#7A1A1A")
            _var_icon = "↑" if _var > 1 else ("→" if abs(_var) <= 1 else "↓")
            _var_label = "Distrito en alza" if _var >= 3 else ("Estable" if abs(_var) <= 1 else "Distrito en baja")

            td1, td2 = st.columns(2)
            td1.metric("Variación anual precio zona", f"{_var:+.1f}%")
            td2.metric("Inflación 12m (Perú)", "+1.4%")

            _real_var = _var - 1.4
            st.markdown(
                f'<div style="background:#F5F2ED;border-left:3px solid {_var_color};border-radius:4px;'
                f'padding:10px 16px;margin:8px 0;">'
                f'<span style="font-size:22px;color:{_var_color};">{_var_icon}</span> '
                f'<strong style="color:{_var_color};">{_var_label}</strong> — '
                f'Variación real (descontando inflación): <strong style="color:{_var_color};">{_real_var:+.1f}%</strong>. '
                f'{"La zona genera plusvalía real para el comprador." if _real_var > 0 else "El inmueble pierde valor real contra inflación — negociar precio."}'
                f'</div>',
                unsafe_allow_html=True
            )

            # ── Valor a futuro con tasa real ─────────────────────
            st.markdown('<div class="section-title">Proyección de Valor</div>', unsafe_allow_html=True)
            _tasa_apr = r.get("tasa_apreciacion_pct", 4.0)
            pv1, pv2, pv3 = st.columns(3)
            pv1.metric("Valor actual", f"${r.get('precio', 0):,.0f}")
            pv2.metric("Valor estimado 5 años", f"${r.get('valor_5', 0):,.0f}",
                       f"+${r.get('ganancia_capital_5', 0):,.0f} ({_tasa_apr:.1f}%/año)")
            pv3.metric("Valor estimado 10 años", f"${r.get('valor_10', 0):,.0f}",
                       f"+${r.get('ganancia_capital_10', 0):,.0f}")
            st.caption(f"Tasa de apreciación usada: {_tasa_apr:.1f}% anual (variación anual de la zona). Fuente: Urbania nov-25.")

        # TAB 1: CRÉDITO HIPOTECARIO
        with res_tabs[1]:
            st.markdown('<div class="section-title">Estructura del Financiamiento</div>', unsafe_allow_html=True)
            rc1, rc2, rc3, rc4 = st.columns(4)
            rc1.metric("Precio de Compra", f"${r.get('precio', 0):,.0f}")
            rc2.metric("Pago inicial (Capital Propio)", f"${r.get('pie', 0):,.0f}", f"{r.get('pct_pie', 0):.0f}%")
            rc3.metric("Monto del Crédito", f"${r.get('monto_credito', 0):,.0f}", f"{100-r.get('pct_pie', 0):.0f}%")
            rc4.metric("Cuota Mensual", f"${r.get('cuota_mensual', 0):,.0f}" if (r.get('cuota_mensual') or 0) > 0 else "Contado")

            st.markdown('<div class="section-title">Resumen del Crédito</div>', unsafe_allow_html=True)
            rr1, rr2, rr3 = st.columns(3)
            rr1.metric("Total a Pagar", f"${r.get('total_pagado', 0):,.0f}", f"Plazo {r.get('plazo_anos', 0)} años")
            rr2.metric("Total Intereses", f"${r.get('total_intereses', 0):,.0f}", f"{r.get('tasa_anual', 0):.2f}% TEA")
            rr3.metric("Ingreso Mínimo Sugerido", f"${r.get('ingreso_minimo', 0):,.0f}/mes", "Regla 30% ingresos")

            st.markdown(
                '<div class="alert-gold">'
                f'<strong>Ingreso mínimo recomendado:</strong> Los bancos generalmente exigen que la cuota hipotecaria '
                f'no supere el 30% del ingreso neto mensual. Con una cuota de <strong>${r.get("cuota_mensual", 0):,.0f}</strong>, '
                f'se recomienda acreditar ingresos mínimos de <strong>${r.get("ingreso_minimo", 0):,.0f}/mes</strong>.'
                '</div>',
                unsafe_allow_html=True
            )

            if (r.get('monto_credito') or 0) > 0:
                fig_pie_res = go.Figure(go.Pie(
                    labels=["Pago inicial (Capital Propio)", "Financiado por banco", "Intereses totales"],
                    values=[r.get('pie', 0), r.get('monto_credito', 0), r.get('total_intereses', 0)],
                    marker_colors=["#1E2D3D", "#B8904A", "#C8A86A"],
                    textfont=dict(size=11, color="#FFFFFF"),
                    hole=0.40,
                ))
                fig_pie_res.update_layout(
                    height=300, margin=dict(t=10, b=10, l=10, r=10),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    legend=dict(font=dict(color="#4A5568", size=11)),
                    font=dict(family="Inter"),
                )
                st.plotly_chart(fig_pie_res, use_container_width=True)

        # TAB 2: INVERSIÓN o ESCENARIOS
        with res_tabs[2]:
            if r.get('uso') in ["Inversión para alquilar", "Evaluación para venta"]:
                st.markdown('<div class="section-title">Análisis de Rentabilidad</div>', unsafe_allow_html=True)
                ri1, ri2, ri3, ri4 = st.columns(4)
                ri1.metric("Alquiler Mensual", f"${r.get('alquiler_mes', 0):,.0f}")
                ri2.metric("Yield Bruto Anual", f"{r.get('yield_bruto', 0):.1f}%")
                ri3.metric("Yield Neto Anual", f"{r.get('yield_neto', 0):.1f}%", f"-${r.get('gastos_mes', 0):,.0f}/mes gastos")
                if r.get('payback_anos'):
                    ri4.metric("Payback", f"{r['payback_anos']:.1f} años")
                else:
                    ri4.metric("Payback", "N/A")

                # ── Proyección indexada (solo contratos plurianuales) ──
                if r.get("tipo_contrato") == "Plurianual (3+ años)" and r.get("ajuste_anual_pct", 0) > 0:
                    _raj  = r["ajuste_anual_pct"]
                    _rini = r["inicio_ajuste_ano"]
                    st.markdown(
                        f'<div style="background:#F0F4FF;border-left:3px solid #3B5BDB;border-radius:6px;'
                        f'padding:10px 14px;margin:12px 0 6px;font-size:12px;">'
                        f'<strong>Contrato plurianual · ajuste +{_raj:.1f}% anual desde año {_rini}</strong><br>'
                        f'La renta crece mientras la cuota hipotecaria permanece fija — '
                        f'el flujo mejora año a año.'
                        f'</div>', unsafe_allow_html=True)
                    _rpi1, _rpi2, _rpi3, _rpi4 = st.columns(4)
                    _rpi1.metric("Alquiler año 1", f"${r['alquiler_mes']:,.0f}/mes")
                    _rpi2.metric("Alquiler año 3", f"${r['alquiler_ano3']:,.0f}/mes",
                                 f"+{((r['alquiler_ano3']/r['alquiler_mes']-1)*100):.1f}%" if r['alquiler_mes'] > 0 else "")
                    _rpi3.metric("Alquiler año 5", f"${r['alquiler_ano5']:,.0f}/mes",
                                 f"+{((r['alquiler_ano5']/r['alquiler_mes']-1)*100):.1f}%" if r['alquiler_mes'] > 0 else "")
                    if r.get("payback_indexado") and r.get("payback_anos"):
                        _rdpb = r["payback_anos"] - r["payback_indexado"]
                        _rpi4.metric("Payback indexado", f"{r['payback_indexado']} años",
                                     f"{_rdpb:.1f} años menos vs renta fija")
                    _rydata = {
                        "Año": ["Año 1 (base)", f"Año 3 (+{_raj:.1f}% × {max(3-_rini+1,0)})", f"Año 5 (+{_raj:.1f}% × {max(5-_rini+1,0)})"],
                        "Alquiler/mes": [f"${r['alquiler_mes']:,.0f}", f"${r['alquiler_ano3']:,.0f}", f"${r['alquiler_ano5']:,.0f}"],
                        "Yield neto": [f"{r['yield_neto']:.1f}%", f"{r['yield_neto_ano3']:.1f}%", f"{r['yield_neto_ano5']:.1f}%"],
                    }
                    st.table(pd.DataFrame(_rydata).set_index("Año"))

                if (r.get('cuota_mensual') or 0) > 0:
                    st.markdown('<div class="section-title">Flujo Mensual Neto</div>', unsafe_allow_html=True)
                    rf1, rf2, rf3 = st.columns(3)
                    flujo = r.get('flujo_mensual') or 0
                    rf1.metric("Renta Neta", f"${r.get('renta_neta_mes', 0):,.0f}/mes")
                    rf2.metric("Cuota Mensual", f"${r.get('cuota_mensual', 0):,.0f}/mes")
                    rf3.metric("Flujo Mensual", f"${flujo:,.0f}" if flujo >= 0 else f"-${abs(flujo):,.0f}",
                               "Positivo" if flujo >= 0 else "Déficit — evaluar pie")

                    st.markdown(
                        '<div class="alert-gold">'
                        f'<strong>Alquiler para autofinanciar:</strong> Necesitas una renta mínima de '
                        f'<strong>${r.get("alquiler_equilibrio", 0):,.0f}/mes</strong> para cubrir la cuota más los gastos operativos.'
                        '</div>',
                        unsafe_allow_html=True
                    )

                st.markdown('<div class="section-title">Apreciación del Capital (Lima +4% anual est.)</div>', unsafe_allow_html=True)
                rap1, rap2 = st.columns(2)
                rap1.metric("Valor Estimado a 5 años", f"${r.get('valor_5', 0):,.0f}", f"+${r.get('ganancia_capital_5', 0):,.0f}")
                rap2.metric("Valor Estimado a 10 años", f"${r.get('valor_10', 0):,.0f}", f"+${r.get('ganancia_capital_10', 0):,.0f}")

            else:
                st.markdown('<div class="section-title">Escenarios de Financiamiento</div>', unsafe_allow_html=True)
                precio_base = r.get('precio', 0)
                escenarios = [
                    ("Conservador", 10, 9.5, 25),
                    ("Estándar", 20, 8.5, 20),
                    ("Agresivo (comprador)", 30, 7.5, 15),
                ]
                esc_data = []
                for name, pct, tasa, plazo in escenarios:
                    c_res = calcular_residencial({"precio": precio_base, "pct_pie": pct, "tasa_anual": tasa, "plazo_anos": plazo, "uso": "Vivienda propia", "alquiler_mes": 0, "gastos_mes": 0})
                    esc_data.append({"Escenario": name, "Pago inicial %": f"{pct}%", "TEA %": f"{tasa}%", "Plazo": f"{plazo}a", "Cuota": f"${c_res['cuota_mensual']:,.0f}", "Ingreso Mín.": f"${c_res['ingreso_minimo']:,.0f}", "Total Pagado": f"${c_res['total_pagado']:,.0f}"})
                _esc_cols = ["Pago inicial %", "TEA %", "Plazo", "Cuota", "Ingreso Mín.", "Total Pagado"]
                _th = "".join(f'<th style="background:#1E2D3D;color:#FFFFFF;padding:10px 14px;font-size:10px;letter-spacing:1px;text-transform:uppercase;font-weight:700;border:1px solid #2A3D51;white-space:nowrap;">{c}</th>' for c in ["Escenario"] + _esc_cols)
                _trows = "".join(
                    '<tr>' + f'<td style="background:#F5F2ED;color:#1E2D3D;padding:9px 14px;font-size:12px;border:1px solid #E0DAD0;font-weight:600;">{d["Escenario"]}</td>' +
                    "".join(f'<td style="background:#FFFFFF;color:#1E2D3D;padding:9px 14px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">{d[c]}</td>' for c in _esc_cols) +
                    '</tr>'
                    for d in esc_data
                )
                st.markdown(f'<div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;">'
                            f'<thead><tr>{_th}</tr></thead><tbody>{_trows}</tbody></table></div>',
                            unsafe_allow_html=True)

                st.markdown('<div class="section-title">Apreciación del Capital (Lima +4% anual est.)</div>', unsafe_allow_html=True)
                rav1, rav2 = st.columns(2)
                rav1.metric("Valor Estimado a 5 años", f"${r.get('valor_5', 0):,.0f}", f"+${r.get('ganancia_capital_5', 0):,.0f}")
                rav2.metric("Valor Estimado a 10 años", f"${r.get('valor_10', 0):,.0f}", f"+${r.get('ganancia_capital_10', 0):,.0f}")

            # Multi-project comparison (residential)
            if st.session_state.res_comparativa and len(st.session_state.res_comparativa) > 1:
                st.markdown('<div class="section-title">Comparativa de Escenarios</div>', unsafe_allow_html=True)
                _res_cmp_metrics = ["precio", "cuota_mensual", "yield_neto", "payback_anos", "flujo_mensual", "ingreso_minimo"]
                _res_cmp_labels  = ["Precio", "Cuota Mensual", "Yield Neto %", "Payback (años)", "Flujo Mensual", "Ingreso Mínimo"]
                _res_cmp_fmts    = ["${:,.0f}", "${:,.0f}", "{:.1f}%", "{:.1f} a.", "${:,.0f}", "${:,.0f}"]

                _rth = '<th style="background:#1E2D3D;color:#FFFFFF;padding:9px 14px;font-size:10px;letter-spacing:1px;text-transform:uppercase;">Indicador</th>'
                for e in st.session_state.res_comparativa:
                    _rth += f'<th style="background:#1E2D3D;color:#B8904A;padding:9px 14px;font-size:10px;text-align:right;">{e["label"]}</th>'
                _rrows = ""
                for lbl, key, fmt in zip(_res_cmp_labels, _res_cmp_metrics, _res_cmp_fmts):
                    bg_row = "#FFFFFF" if _res_cmp_metrics.index(key) % 2 == 0 else "#F9F7F4"
                    _rrow = f'<td style="background:{bg_row};color:#4A5568;padding:8px 14px;font-size:11px;font-weight:600;border:1px solid #E0DAD0;">{lbl}</td>'
                    vals = [e["r"].get(key) for e in st.session_state.res_comparativa]
                    _hib = key in ["yield_neto", "flujo_mensual"]
                    _lib = key in ["precio", "cuota_mensual", "payback_anos", "ingreso_minimo"]
                    valid = [v for v in vals if v is not None]
                    best = max(valid) if _hib and valid else (min(valid) if _lib and valid else None)
                    for v in vals:
                        try: disp = fmt.format(v) if v is not None else "—"
                        except: disp = "—"
                        is_b = (v == best and best is not None)
                        _rrow += f'<td style="background:{bg_row};color:{"#1A4731" if is_b else "#1E2D3D"};padding:8px 14px;font-size:12px;font-weight:{"700" if is_b else "400"};border:1px solid #E0DAD0;text-align:right;">{disp}</td>'
                    _rrows += f"<tr>{_rrow}</tr>"
                st.markdown(f'<div style="overflow-x:auto;margin-bottom:16px;"><table style="width:100%;border-collapse:collapse;"><thead><tr>{_rth}</tr></thead><tbody>{_rrows}</tbody></table></div>', unsafe_allow_html=True)
                st.caption("Verde = mejor valor para ese indicador.")

        # TAB 3: COMPARATIVA DE INMUEBLES
        with res_tabs[3]:
            st.markdown('<div class="section-title">Comparativa de Inmuebles</div>', unsafe_allow_html=True)

            _comps = st.session_state.get("res_inmuebles_comp", [])

            if not _comps:
                st.markdown('<div class="alert-legal">Agrega inmuebles en el panel izquierdo (sección <strong>COMPARATIVA DE INMUEBLES</strong>) para comparar alternativas de mercado.</div>', unsafe_allow_html=True)
            else:
                import base64 as _b64

                # ── District tier helper ──────────────────────────
                def _tier_badge(zona_key: str) -> str:
                    _pm2 = MERCADO.get(zona_key, {}).get("precio_2br", 0)
                    if _pm2 >= 2500: return ('<span style="background:#1E2D3D;color:#F5F2ED;font-size:9px;font-weight:700;'
                                              'padding:2px 8px;border-radius:10px;letter-spacing:1px;">PREMIUM</span>')
                    if _pm2 >= 2000: return ('<span style="background:#B8904A;color:#FFFFFF;font-size:9px;font-weight:700;'
                                              'padding:2px 8px;border-radius:10px;letter-spacing:1px;">PREMIUM +</span>')
                    if _pm2 >= 1500: return ('<span style="background:#4A7A6B;color:#FFFFFF;font-size:9px;font-weight:700;'
                                              'padding:2px 8px;border-radius:10px;letter-spacing:1px;">RESIDENCIAL</span>')
                    if _pm2 >= 1000: return ('<span style="background:#5A6A7A;color:#FFFFFF;font-size:9px;font-weight:700;'
                                              'padding:2px 8px;border-radius:10px;letter-spacing:1px;">CONSOLIDADO</span>')
                    return ('<span style="background:#8A7A6A;color:#FFFFFF;font-size:9px;font-weight:700;'
                             'padding:2px 8px;border-radius:10px;letter-spacing:1px;">EMERGENTE</span>')

                # ── All properties including current ──────────────
                _all = [{"nombre": "Este inmueble", "is_main": True,
                         "precio": r.get("precio", 0) or 0, "m2": r.get("m2", 0),
                         "precio_m2": r.get("precio_m2", 0),
                         "alquiler": r.get("alquiler_mes", 0),
                         "yield_bruto": r.get("yield_bruto", 0),
                         "zona": r.get("zona", "—"),
                         "dormitorios": r.get("dormitorios", "—"),
                         "imagen_bytes": (st.session_state.get("res_fotos_bytes") or [None])[0],
                        }] + [dict(c, is_main=False) for c in _comps]

                _best_pm2   = min((c["precio_m2"] for c in _all if c["precio_m2"] > 0), default=0)
                _best_yield = max((c["yield_bruto"] for c in _all), default=0)
                _best_precio = min(c["precio"] for c in _all)

                # ── Property Cards Grid ────────────────────────────
                _GRAD_PALETTES = [
                    "linear-gradient(135deg,#1E2D3D 0%,#2A4060 100%)",
                    "linear-gradient(135deg,#2D3A2E 0%,#3D5C40 100%)",
                    "linear-gradient(135deg,#3A2A1E 0%,#5C4030 100%)",
                    "linear-gradient(135deg,#2A2A3A 0%,#3D3D5C 100%)",
                ]
                _cards_html = '<div style="display:grid;grid-template-columns:repeat(2,1fr);gap:16px;margin-top:8px;">'
                for _ci_idx, _ci in enumerate(_all):
                    _is_main = _ci.get("is_main", False)
                    _img_b = _ci.get("imagen_bytes")
                    if _img_b:
                        try:
                            _img_src = f"data:image/jpeg;base64,{_b64.b64encode(_img_b).decode()}"
                            _photo_css = f"background:url('{_img_src}') center/cover no-repeat;"
                        except Exception:
                            _img_src = None
                            _photo_css = _GRAD_PALETTES[_ci_idx % len(_GRAD_PALETTES)]
                    else:
                        _photo_css = _GRAD_PALETTES[_ci_idx % len(_GRAD_PALETTES)]

                    _border = "2px solid #B8904A" if _is_main else "1px solid #E4E0D8"
                    _label_badge = ('<div style="position:absolute;top:10px;left:10px;background:#B8904A;color:#FFFFFF;'
                                    'font-size:9px;font-weight:700;padding:3px 8px;border-radius:3px;letter-spacing:1px;">▶ ANALIZADO</div>') if _is_main else ""

                    _best_badges = ""
                    if _ci["precio_m2"] == _best_pm2 and _ci["precio_m2"] > 0:
                        _best_badges += '<span style="background:#1A4731;color:#FFFFFF;font-size:8px;font-weight:700;padding:2px 6px;border-radius:2px;margin-right:4px;">✓ MEJOR USD/m²</span>'
                    if _ci["yield_bruto"] == _best_yield and _ci["yield_bruto"] > 0:
                        _best_badges += '<span style="background:#1A4731;color:#FFFFFF;font-size:8px;font-weight:700;padding:2px 6px;border-radius:2px;">✓ MEJOR YIELD</span>'

                    _zona_disp = _ci.get("zona", "—")
                    _tier_html = _tier_badge(_zona_disp) if _zona_disp in MERCADO else ""
                    _dorm_disp = _ci.get("dormitorios", "—")
                    _precio_str = f"${_ci['precio']:,}"
                    _pm2_str    = f"${_ci['precio_m2']:,}/m²" if _ci["precio_m2"] > 0 else "—"
                    _alq_str    = f"${_ci['alquiler']:,}/mes" if _ci.get("alquiler", 0) > 0 else "—"
                    _yield_str  = f"{_ci['yield_bruto']:.1f}%" if _ci.get("yield_bruto", 0) > 0 else "—"
                    _m2_str     = f"{_ci['m2']:,} m²" if _ci.get("m2", 0) > 0 else "—"

                    _cards_html += f'''
                    <div style="border:{_border};border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(30,45,61,0.10);background:#FFFFFF;">
                        <div style="position:relative;height:160px;{_photo_css}">
                            {_label_badge}
                            <div style="position:absolute;bottom:0;left:0;right:0;height:70px;
                                        background:linear-gradient(transparent,rgba(0,0,0,0.65));
                                        padding:10px 14px 12px;display:flex;flex-direction:column;justify-content:flex-end;">
                                <div style="color:#FFFFFF;font-size:18px;font-weight:800;line-height:1.2;">{_precio_str}</div>
                                <div style="color:rgba(255,255,255,0.85);font-size:10px;margin-top:2px;">{_pm2_str} &nbsp;·&nbsp; {_m2_str}</div>
                            </div>
                        </div>
                        <div style="padding:14px 16px;">
                            <div style="font-size:13px;font-weight:700;color:#1E2D3D;margin-bottom:4px;">{_ci["nombre"]}</div>
                            <div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
                                <span style="font-size:11px;color:#6A7A8A;">{_zona_disp}</span>
                                {_tier_html}
                            </div>
                            <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px;">
                                <div style="background:#F7F5F1;border-radius:6px;padding:8px 10px;">
                                    <div style="font-size:9px;color:#9A9080;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px;">Tipología</div>
                                    <div style="font-size:12px;font-weight:600;color:#1E2D3D;">{_dorm_disp}</div>
                                </div>
                                <div style="background:#F7F5F1;border-radius:6px;padding:8px 10px;">
                                    <div style="font-size:9px;color:#9A9080;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px;">Alquiler est.</div>
                                    <div style="font-size:12px;font-weight:600;color:#1E2D3D;">{_alq_str}</div>
                                </div>
                                <div style="background:#F7F5F1;border-radius:6px;padding:8px 10px;">
                                    <div style="font-size:9px;color:#9A9080;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px;">Yield bruto</div>
                                    <div style="font-size:12px;font-weight:600;color:{"#1A4731" if _ci.get("yield_bruto",0)>0 else "#9A9080"};">{_yield_str}</div>
                                </div>
                                <div style="background:#F7F5F1;border-radius:6px;padding:8px 10px;">
                                    <div style="font-size:9px;color:#9A9080;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px;">USD/m²</div>
                                    <div style="font-size:12px;font-weight:600;color:#1E2D3D;">{_pm2_str}</div>
                                </div>
                            </div>
                            {f'<div style="margin-top:4px;">{_best_badges}</div>' if _best_badges else ""}
                        </div>
                    </div>'''

                _cards_html += '</div>'
                st.markdown(_cards_html, unsafe_allow_html=True)

                # ── Comparison chart ───────────────────────────────
                if any(c["precio_m2"] > 0 for c in _all):
                    _nombres = [("★ " + c["nombre"]) if c.get("is_main") else c["nombre"] for c in _all]
                    _pm2s = [c["precio_m2"] for c in _all]
                    _colors = ["#B8904A" if c.get("is_main") else ("#4A7A6B" if c["precio_m2"] == _best_pm2 else "#C8D4DE") for c in _all]
                    fig_comp = go.Figure(go.Bar(
                        x=_nombres, y=_pm2s, marker_color=_colors, marker_line_width=0,
                        text=[f"${v:,}/m²" for v in _pm2s], textposition="outside",
                        textfont=dict(size=10, color="#1E2D3D"),
                    ))
                    fig_comp.update_layout(
                        height=300, yaxis_title="USD/m²",
                        title=dict(text="Precio por m² — Comparativa de mercado", font=dict(size=11, color="#9A9080"), x=0),
                        margin=dict(t=40, b=10, l=0, r=0),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(color="#1E2D3D", family="Inter"),
                        xaxis=dict(showgrid=False),
                        yaxis=dict(tickformat="$,.0f", gridcolor="#EEEBE3"),
                        bargap=0.4,
                    )
                    st.plotly_chart(fig_comp, use_container_width=True, config={"displayModeBar": False})
                st.caption("Verde = mejor valor para ese indicador. Naranja = inmueble analizado.")

        # TAB 4: AMORTIZACIÓN
        with res_tabs[4]:
            st.markdown('<div class="section-title">Tabla de Amortización (primeros 10 años)</div>', unsafe_allow_html=True)
            if r.get('amort_tabla'):
                _amort_th = "".join(
                    f'<th style="background:#1E2D3D;color:#FFFFFF;padding:10px 14px;font-size:10px;'
                    f'letter-spacing:1px;text-transform:uppercase;font-weight:700;border:1px solid #2A3D51;text-align:right;">{h}</th>'
                    for h in ["Año", "Capital Pagado", "Intereses Pagados", "Saldo Restante"]
                )
                _amort_rows = ""
                for i, row in enumerate(r['amort_tabla']):
                    bg = "#FFFFFF" if i % 2 == 0 else "#F9F7F4"
                    _amort_rows += (
                        f'<tr>'
                        f'<td style="background:{bg};color:#1E2D3D;padding:9px 14px;font-size:12px;border:1px solid #E0DAD0;font-weight:700;text-align:center;">{row["año"]}</td>'
                        f'<td style="background:{bg};color:#1A4731;padding:9px 14px;font-size:12px;border:1px solid #E0DAD0;text-align:right;font-weight:600;">${row["capital"]:,.0f}</td>'
                        f'<td style="background:{bg};color:#7A4F1A;padding:9px 14px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">${row["interes"]:,.0f}</td>'
                        f'<td style="background:{bg};color:#1E2D3D;padding:9px 14px;font-size:12px;border:1px solid #E0DAD0;text-align:right;">${row["saldo"]:,.0f}</td>'
                        f'</tr>'
                    )
                st.markdown(
                    f'<div style="overflow-x:auto;margin-bottom:20px;">'
                    f'<table style="width:100%;border-collapse:collapse;">'
                    f'<thead><tr>{_amort_th}</tr></thead>'
                    f'<tbody>{_amort_rows}</tbody>'
                    f'</table></div>',
                    unsafe_allow_html=True
                )

                anos_a = [row['año'] for row in r['amort_tabla']]
                fig_amort = go.Figure()
                fig_amort.add_trace(go.Bar(name="Capital", x=anos_a, y=[row['capital'] for row in r['amort_tabla']], marker_color="#1E2D3D"))
                fig_amort.add_trace(go.Bar(name="Intereses", x=anos_a, y=[row['interes'] for row in r['amort_tabla']], marker_color="#B8904A"))
                fig_amort.update_layout(
                    barmode="stack", height=280, margin=dict(t=20, b=20, l=10, r=10),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(title=dict(text="Año", font=dict(color="#4A5568", size=11)), tickfont=dict(color="#4A5568")),
                    yaxis=dict(title=dict(text="USD", font=dict(color="#4A5568", size=11)), tickfont=dict(color="#4A5568"), tickformat="$,.0f"),
                    legend=dict(font=dict(color="#4A5568", size=11)),
                    font=dict(family="Inter"),
                )
                st.plotly_chart(fig_amort, use_container_width=True)
            else:
                st.markdown('<div class="alert-legal">Sin crédito — compra al contado.</div>', unsafe_allow_html=True)

        # TAB 5: LEGAL + DESCARGA
        with res_tabs[5]:
            st.markdown('<div class="section-title">Análisis Legal y Descarga de Informe</div>', unsafe_allow_html=True)
            _res_html = generar_informe_residencial_html(
                r, st.session_state.get("residencial_legal"),
                datetime.datetime.now().strftime("%d/%m/%Y"),
                distrito=st.session_state.get("_res_zona_val", r.get("zona", "")),
                m2=st.session_state.get("_res_m2_val", 0),
                antiguedad=st.session_state.get("_res_antiguedad_val", 0),
                fotos=st.session_state.get("res_fotos_bytes", []),
            )
            _res_nombre = (st.session_state.get("nombre_proyecto") or
                           st.session_state.get("_res_zona_val", "Proyecto"))
            st.download_button(
                "DESCARGAR INFORME RESIDENCIAL",
                data=_res_html.encode("utf-8"),
                file_name=f"Informe de Análisis Residencial - {_res_nombre}.html",
                mime="text/html",
                use_container_width=True,
            )
            st.caption("El archivo .html se abre en cualquier navegador y puede imprimirse como PDF (Cmd+P / Ctrl+P).")
            st.markdown("---")
            lg = st.session_state.get("residencial_legal")
            if not lg:
                st.markdown(
                    '<div style="background:#F7F5F1;border:1px solid #D8D4CC;border-radius:8px;'
                    'padding:36px 32px;text-align:center;margin-top:8px;">'
                    '<div style="font-size:9px;color:#9A9080;letter-spacing:3px;text-transform:uppercase;'
                    'font-weight:600;margin-bottom:12px;">Análisis Opcional</div>'
                    '<div style="font-size:16px;font-weight:600;color:#1E2D3D;margin-bottom:8px;">'
                    'Análisis Legal Registral</div>'
                    '<div style="width:36px;height:2px;background:#B8904A;margin:12px auto;"></div>'
                    '<div style="font-size:13px;color:#7A7268;line-height:1.7;max-width:480px;margin:0 auto;">'
                    'Adjunta la <strong>Partida Registral (SUNARP)</strong> y/o el <strong>PU/HR</strong> '
                    'en el panel izquierdo, luego presiona <strong>ANALIZAR DOCUMENTOS</strong> para verificar:'
                    '<br><br>✓ Titularidad y propietarios registrales<br>'
                    '✓ Cargas, hipotecas y medidas cautelares<br>'
                    '✓ Consistencia de áreas y direcciones entre documentos'
                    '</div></div>',
                    unsafe_allow_html=True
                )
            else:
                sem = lg.get("semaforo", "amarillo").lower()
                sem_cfg = {
                    "verde":    ("#1A4731", "#E8F5EE", "SIN ALERTAS CRÍTICAS"),
                    "amarillo": ("#7A4F1A", "#FFF8EE", "OBSERVACIONES MENORES"),
                    "rojo":     ("#7A1A1A", "#FFF0F0", "ALERTAS CRÍTICAS"),
                }.get(sem, ("#1E2D3D", "#F5F2ED", "INDETERMINADO"))
                sc, sbg, setiq = sem_cfg

                st.markdown(f"""
                <div style="background:{sbg};border:1px solid {sc};border-left:4px solid {sc};
                            border-radius:8px;padding:20px 24px;margin-bottom:20px;">
                    <div style="font-size:9px;letter-spacing:3px;color:{sc};text-transform:uppercase;
                                font-weight:700;opacity:0.7;margin-bottom:6px;">Estado Legal del Inmueble</div>
                    <div style="font-size:20px;font-weight:700;color:{sc};margin-bottom:10px;">{setiq}</div>
                    <div style="font-size:13px;color:{sc};opacity:0.85;line-height:1.6;">
                        {lg.get('resumen_legal','—')}</div>
                </div>""", unsafe_allow_html=True)

                # ── Checklist de hallazgos ───────────────────
                res_hallazgos = lg.get("hallazgos") or []
                res_completitud = lg.get("completitud") or {}
                if res_hallazgos:
                    _rcv = res_completitud.get("verificados", len(res_hallazgos))
                    _rct = res_completitud.get("total", 20)
                    _rcnv = res_completitud.get("no_verificables") or []
                    _rcv_pct = round(_rcv / _rct * 100) if _rct else 0
                    _rcv_color = "#1A4731" if _rcv_pct >= 80 else ("#7A4F1A" if _rcv_pct >= 60 else "#7A1A1A")
                    st.markdown(
                        f'<div style="display:flex;justify-content:space-between;align-items:center;'
                        f'background:#F7F5F1;border:1px solid #E4E0D8;border-radius:8px;'
                        f'padding:12px 20px;margin-bottom:16px;">'
                        f'<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:700;">Completitud del análisis</div>'
                        f'<div style="font-size:16px;font-weight:700;color:{_rcv_color};">{_rcv}/{_rct} puntos verificados</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                    _RSEV = {
                        "verde":         ("🟢", "#1A4731", "#E8F5EE", "#C3E6CB"),
                        "amarillo":      ("🟡", "#7A4F1A", "#FFF8EE", "#FFE0A0"),
                        "rojo":          ("🔴", "#7A1A1A", "#FFF0F0", "#F5C6CB"),
                        "no_verificable":("⚪", "#6A6A6A", "#F5F5F5", "#D8D4CC"),
                    }
                    _rpartida_h = [h for h in res_hallazgos if h.get("categoria") in ("partida", "legal")]
                    _rparams_h  = [h for h in res_hallazgos if h.get("categoria") in ("parametros", "tecnico")]
                    _rother_h   = [h for h in res_hallazgos if h.get("categoria") not in ("partida", "legal", "parametros", "tecnico")]

                    def _render_res_hallazgos(h_list, section_lbl):
                        if not h_list:
                            return
                        st.markdown(f'<div class="section-title">{section_lbl}</div>', unsafe_allow_html=True)
                        for h in h_list:
                            sev = (h.get("severidad") or "no_verificable").lower()
                            icon_h, tcol, bg, brd = _RSEV.get(sev, _RSEV["no_verificable"])
                            num = h.get("numero", "")
                            punto = h.get("punto", "—")
                            hallazgo_txt = h.get("hallazgo", "—")
                            subsanacion = h.get("subsanacion")
                            sev_lbl = {"verde": "OK", "amarillo": "VERIFICAR", "rojo": "CRÍTICO", "no_verificable": "NO VERIFICABLE"}.get(sev, sev.upper())
                            sub_html = ""
                            if subsanacion:
                                sub_html = (
                                    f'<div style="margin-top:10px;padding-top:10px;border-top:1px solid {brd};">'
                                    f'<div style="font-size:9px;color:{tcol};letter-spacing:1.5px;text-transform:uppercase;font-weight:700;margin-bottom:4px;">Procedimiento de subsanación</div>'
                                    f'<div style="font-size:12px;color:#2A3A4A;line-height:1.6;">{subsanacion}</div>'
                                    f'</div>'
                                )
                            st.markdown(
                                f'<div style="background:{bg};border:1px solid {brd};border-left:4px solid {tcol};'
                                f'border-radius:6px;padding:14px 18px;margin-bottom:10px;">'
                                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px;">'
                                f'<div style="font-size:9px;color:{tcol};letter-spacing:1.5px;text-transform:uppercase;font-weight:700;">'
                                f'{icon_h} {num}. {punto}</div>'
                                f'<div style="font-size:8px;font-weight:700;color:{tcol};background:rgba(0,0,0,0.07);'
                                f'padding:2px 8px;border-radius:3px;white-space:nowrap;">{sev_lbl}</div>'
                                f'</div>'
                                f'<div style="font-size:12px;color:#1E2D3D;line-height:1.6;">{hallazgo_txt}</div>'
                                f'{sub_html}'
                                f'</div>',
                                unsafe_allow_html=True
                            )

                    _render_res_hallazgos(_rpartida_h, "Análisis Registral — Partida SUNARP")
                    _render_res_hallazgos(_rparams_h,  "Análisis Urbanístico — Certificado de Parámetros")
                    if _rother_h:
                        _render_res_hallazgos(_rother_h, "Verificación Cruzada")
                    if _rcnv:
                        with st.expander(f"Puntos no verificables ({len(_rcnv)})"):
                            for nv in _rcnv:
                                st.markdown(f'<div style="font-size:12px;color:#9A9080;padding:3px 0;">⚪ {nv}</div>', unsafe_allow_html=True)
                else:
                    alertas = lg.get("alertas", []) or []
                    if alertas:
                        st.markdown('<div class="section-title">Alertas</div>', unsafe_allow_html=True)
                        for al in alertas:
                            icon = "🔴" if sem == "rojo" else ("🟡" if sem == "amarillo" else "🟢")
                            st.markdown(f'<div class="alert-gold">{icon} {al}</div>', unsafe_allow_html=True)

                st.markdown('<div class="section-title">Verificación Cruzada — Documentos</div>', unsafe_allow_html=True)

                def _check(ok):
                    if ok is True:  return "✓", "#1A4731"
                    if ok is False: return "✗", "#7A1A1A"
                    return "—", "#9A9080"

                def _vcard_r(label, row1_lbl, row1_val, row2_lbl, row2_val, note, icon, col):
                    note_html = ('<div style="font-size:11px;color:#7A4F1A;margin-top:8px;font-style:italic;">' + note + '</div>') if note else ''
                    return (
                        '<div style="background:#FFFFFF;border:1px solid #E4E0D8;border-radius:6px;padding:16px 20px;margin-bottom:10px;">'
                        '<div style="display:flex;justify-content:space-between;align-items:flex-start;">'
                        '<div style="flex:1;">'
                        '<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;margin-bottom:8px;">' + label + '</div>'
                        '<div style="font-size:12px;color:#1E2D3D;margin-bottom:4px;"><strong>' + row1_lbl + ':</strong> ' + str(row1_val) + '</div>'
                        '<div style="font-size:12px;color:#1E2D3D;"><strong>' + row2_lbl + ':</strong> ' + str(row2_val) + '</div>'
                        + note_html
                        + '</div>'
                        '<div style="font-size:28px;color:' + col + ';font-weight:700;margin-left:16px;">' + icon + '</div>'
                        '</div></div>'
                    )

                def _fmt_prop_r(p):
                    if not p or not isinstance(p, dict):
                        return str(p) if p else "—"
                    nombre = p.get("nombre") or "—"
                    dni    = p.get("dni")
                    pct    = p.get("porcentaje")
                    cond   = p.get("condicion")
                    tipo   = p.get("tipo_doc", "DNI")
                    parts  = [nombre]
                    if pct:  parts.append(f"({pct})")
                    if cond: parts.append(f"[{cond}]")
                    result = " ".join(parts)
                    if dni:
                        result += f' <span style="background:#E8F5EE;color:#1A4731;font-size:10px;font-weight:700;padding:2px 6px;border-radius:3px;border:1px solid #1A4731;">{tipo}: {dni}</span>'
                    else:
                        result += ' <span style="background:#FFF8EE;color:#7A4F1A;font-size:10px;padding:2px 6px;border-radius:3px;border:1px solid #B8904A;">DNI no encontrado</span>'
                    return result

                prop_p_raw = lg.get("propietarios_partida") or []
                prop_h_raw = lg.get("propietarios_puhr") or []

                if prop_p_raw and isinstance(prop_p_raw[0], dict):
                    prop_p_str = "<br>".join(_fmt_prop_r(p) for p in prop_p_raw)
                else:
                    prop_p_str = ", ".join(str(x) for x in prop_p_raw) if prop_p_raw else "—"

                if prop_h_raw and isinstance(prop_h_raw[0], dict):
                    prop_h_str = "<br>".join(_fmt_prop_r(p) for p in prop_h_raw)
                else:
                    prop_h_str = ", ".join(str(x) for x in prop_h_raw if x) if prop_h_raw else "—"

                icon_p, col_p = _check(lg.get("propietarios_coinciden"))
                note_p = lg.get("diferencias_propietarios", "") or ""
                note_p_html = ('<div style="font-size:11px;color:#7A4F1A;margin-top:8px;font-style:italic;">' + note_p + '</div>') if note_p else ""
                st.markdown(
                    '<div style="background:#FFFFFF;border:1px solid #E4E0D8;border-radius:6px;padding:16px 20px;margin-bottom:10px;">'
                    '<div style="display:flex;justify-content:space-between;align-items:flex-start;">'
                    '<div style="flex:1;">'
                    '<div style="font-size:9px;color:#9A9080;letter-spacing:2px;text-transform:uppercase;font-weight:600;margin-bottom:8px;">Titularidad / Propietarios</div>'
                    '<div style="font-size:12px;color:#1E2D3D;margin-bottom:6px;"><strong>Partida:</strong><br>' + prop_p_str + '</div>'
                    '<div style="font-size:12px;color:#1E2D3D;"><strong>PU/HR:</strong><br>' + prop_h_str + '</div>'
                    + note_p_html +
                    '</div>'
                    '<div style="font-size:28px;color:' + col_p + ';font-weight:700;margin-left:16px;">' + icon_p + '</div>'
                    '</div></div>',
                    unsafe_allow_html=True
                )

                dir_p = lg.get("direccion_partida", "—") or "—"
                dir_h = lg.get("direccion_puhr", "—") or "—"
                icon_d, col_d = _check(lg.get("direcciones_coinciden"))
                st.markdown(_vcard_r(
                    "Ubicación / Dirección",
                    "Partida", dir_p, "PU/HR", dir_h,
                    lg.get("diferencias_direccion", ""),
                    icon_d, col_d,
                ), unsafe_allow_html=True)

                area_r = lg.get("area_registral_m2")
                area_h_v = lg.get("area_puhr_m2")
                disc = lg.get("discrepancia_area_m2")
                icon_a, col_a = _check(lg.get("areas_coinciden"))
                disc_note = ("Discrepancia: <strong>" + f"{disc:+.2f} m²</strong>") if disc else ""
                st.markdown(_vcard_r(
                    "Área del Inmueble",
                    "Partida", f"{area_r:,.2f} m²" if area_r else "—",
                    "PU/HR",   f"{area_h_v:,.2f} m²" if area_h_v else "—",
                    disc_note, icon_a, col_a,
                ), unsafe_allow_html=True)

                _autoavaluo_r  = lg.get("valor_autoavaluo")
                _moneda_av_r   = lg.get("moneda_autoavaluo") or "PEN"
                _anio_av_r     = lg.get("anio_autoavaluo")
                _clasif_muni_r = lg.get("clasificacion_municipal")
                _cond_sat_r    = lg.get("condicion_propietario_sat")
                _uso_predio_r  = lg.get("uso_predio")
                if any([_autoavaluo_r, _clasif_muni_r, _cond_sat_r, _uso_predio_r]):
                    st.markdown('<div class="section-title">Datos Municipales — PU/HR</div>', unsafe_allow_html=True)
                    _mu_r_cols = st.columns(2)
                    if _autoavaluo_r:
                        _av_label_r = f"Autoavalúo {_anio_av_r}" if _anio_av_r else "Autoavalúo"
                        _av_val_r   = f"{_moneda_av_r} {_autoavaluo_r:,.0f}" if isinstance(_autoavaluo_r, (int, float)) else f"{_moneda_av_r} {_autoavaluo_r}"
                        _mu_r_cols[0].metric(_av_label_r, _av_val_r)
                    if _clasif_muni_r:
                        _mu_r_cols[1].metric("Clasificación Municipal", _clasif_muni_r)
                    if _cond_sat_r or _uso_predio_r:
                        _mu_r_cols2 = st.columns(2)
                        if _cond_sat_r:   _mu_r_cols2[0].metric("Condición (SAT)", _cond_sat_r)
                        if _uso_predio_r: _mu_r_cols2[1].metric("Uso del Predio", _uso_predio_r)

                hipotecas = lg.get("hipotecas_vigentes", []) or []
                cargas    = lg.get("cargas_vigentes", []) or []
                medidas   = lg.get("medidas_cautelares", []) or []
                if hipotecas or cargas or medidas:
                    st.markdown('<div class="section-title">Cargas, Gravámenes e Hipotecas</div>', unsafe_allow_html=True)
                    for h in hipotecas:
                        st.markdown(
                            f'<div style="background:#FFF0F0;border:1px solid #E8B4B4;border-left:3px solid #C0392B;'
                            f'border-radius:6px;padding:12px 16px;margin-bottom:8px;color:#1E2D3D;font-size:13px;">'
                            f'<strong>Hipoteca</strong> — {h.get("acreedor","—")} · {h.get("monto","—")} '
                            f'<span style="font-size:11px;color:#7A1A1A;">({h.get("estado","—")})</span></div>',
                            unsafe_allow_html=True)
                    for _cg_res in cargas:
                        st.markdown(
                            f'<div style="background:#FFF8EE;border:1px solid #DFC07A;border-left:3px solid #B8904A;'
                            f'border-radius:6px;padding:12px 16px;margin-bottom:8px;color:#1E2D3D;font-size:13px;">'
                            f'<strong>{_cg_res.get("tipo","Carga")}</strong> — {_cg_res.get("descripcion","—")}</div>',
                            unsafe_allow_html=True)
                    for _mc_res in medidas:
                        st.markdown(
                            f'<div style="background:#FFF0F0;border:1px solid #E8B4B4;border-left:3px solid #C0392B;'
                            f'border-radius:6px;padding:12px 16px;margin-bottom:8px;color:#1E2D3D;font-size:13px;">'
                            f'<strong>Medida Cautelar</strong> — {_mc_res.get("tipo","—")}: {_mc_res.get("descripcion","—")}</div>',
                            unsafe_allow_html=True)
                elif sem == "verde":
                    st.markdown('<div class="alert-legal">✓ Sin cargas, hipotecas ni medidas cautelares detectadas.</div>',
                                unsafe_allow_html=True)

        # TAB 6: RESUMEN IA (residencial)
        with res_tabs[6]:
            rsm_r = st.session_state.get("res_resumen")
            if not rsm_r:
                st.markdown(
                    '<div style="background:#F7F5F1;border:1px solid #D8D4CC;border-radius:8px;'
                    'padding:36px 32px;text-align:center;margin-top:8px;">'
                    '<div style="font-size:16px;font-weight:600;color:#1E2D3D;margin-bottom:8px;">Resumen Ejecutivo</div>'
                    '<div style="width:36px;height:2px;background:#B8904A;margin:12px auto;"></div>'
                    '<div style="font-size:13px;color:#7A7268;line-height:1.7;max-width:480px;margin:0 auto 24px;">'
                    'Análisis ejecutivo con recomendación, argumentos clave y riesgos para este inmueble residencial.'
                    '</div></div>', unsafe_allow_html=True)
            if st.button("GENERAR RESUMEN EJECUTIVO", use_container_width=True, type="primary", key="btn_res_rsm"):
                _r_copy = dict(r)
                st.session_state.res_resumen = _run_with_retry(
                    lambda _rc=_r_copy: generar_resumen_ejecutivo_ia("residencial", _rc),
                    "Generando resumen ejecutivo…"
                )
                st.rerun()
            if rsm_r:
                _rec = rsm_r.get("recomendacion", "evaluar_con_condiciones")
                _rec_cfg = {
                    "comprar":                ("#1A4731", "#E8F5EE", "RECOMENDADO — COMPRAR"),
                    "evaluar_con_condiciones":("#7A4F1A", "#FFF8EE", "EVALUAR CON CONDICIONES"),
                    "no_recomendado":         ("#7A1A1A", "#FFF0F0", "NO RECOMENDADO"),
                }.get(_rec, ("#1E2D3D", "#F5F2ED", "—"))
                _rc2, _rbg2, _retiq2 = _rec_cfg

                st.markdown(f"""
                <div style="background:{_rbg2};border:1px solid {_rc2};border-left:5px solid {_rc2};
                            border-radius:8px;padding:22px 28px;margin-bottom:20px;">
                    <div style="font-size:9px;letter-spacing:3px;color:{_rc2};text-transform:uppercase;
                                font-weight:700;opacity:0.7;margin-bottom:6px;">Recomendación</div>
                    <div style="font-size:20px;font-weight:700;color:{_rc2};margin-bottom:8px;">{_retiq2}</div>
                    <div style="font-size:15px;font-weight:600;color:{_rc2};margin-bottom:12px;">{rsm_r.get('titulo','')}</div>
                    <div style="font-size:13px;color:{_rc2};opacity:0.9;line-height:1.7;">{rsm_r.get('resumen','')}</div>
                </div>""", unsafe_allow_html=True)

                _rra1, _rra2 = st.columns(2)
                with _rra1:
                    st.markdown('<div class="section-title">A favor</div>', unsafe_allow_html=True)
                    for a in (rsm_r.get("argumentos_favor") or []):
                        st.markdown(f'<div class="alert-legal" style="margin-bottom:6px;">✓ {a}</div>', unsafe_allow_html=True)
                with _rra2:
                    st.markdown('<div class="section-title">Riesgos</div>', unsafe_allow_html=True)
                    for rk in (rsm_r.get("riesgos") or []):
                        st.markdown(f'<div class="alert-gold" style="margin-bottom:6px;">⚠ {rk}</div>', unsafe_allow_html=True)

                st.markdown(
                    f'<div style="background:#1E2D3D;border-radius:6px;padding:16px 22px;margin-top:16px;">'
                    f'<div style="font-size:9px;color:#B8904A;letter-spacing:2px;text-transform:uppercase;font-weight:600;margin-bottom:6px;">Conclusión</div>'
                    f'<div style="font-size:13px;color:#FFFFFF;line-height:1.7;">{rsm_r.get("conclusion","")}</div>'
                    f'</div>', unsafe_allow_html=True)

                if st.button("REGENERAR", key="btn_res_rsm_regen"):
                    st.session_state.res_resumen = None
                    st.rerun()

        # TAB 7: DOCUMENTOS
        with res_tabs[7]:
            st.markdown('<div class="section-title">Documentos Profesionales</div>', unsafe_allow_html=True)

            _doc_tipo = st.radio("Tipo de documento",
                                  ["Informe de Valoración", "Propuesta de Alquiler", "Propuesta de Compra", "Contraoferta"],
                                  horizontal=True)

            _dc1, _dc2 = st.columns(2)
            _doc_agente = _dc1.text_input("Nombre del agente", placeholder="Tu nombre completo")
            _doc_cliente = _dc2.text_input("Nombre del cliente / propietario", placeholder="A quién va dirigido")
            _doc_obs = st.text_area("Observaciones adicionales", placeholder="Condiciones especiales, notas del cliente, etc.", height=80)

            if st.button("GENERAR DOCUMENTO", use_container_width=True, type="primary"):
                # Generate document HTML
                _fecha_doc = datetime.datetime.now().strftime("%d/%m/%Y")
                _zona_doc = r.get("zona", "Lima")
                _precio_doc = r.get("precio", 0) or 0
                _m2_doc = r.get("m2", 0)
                _ppm2_doc = r.get("precio_m2", 0)
                _ref_m2_doc = r.get("precio_m2_mercado", 0)
                _diff_doc = ((_ppm2_doc - _ref_m2_doc) / _ref_m2_doc * 100) if _ref_m2_doc > 0 else 0
                _alq_doc = r.get("alquiler_mes", 0)
                _alq_mkt_doc = r.get("alquiler_mercado_m2", 0) * _m2_doc
                _yield_doc = r.get("yield_bruto", 0)
                _yield_mkt_doc = r.get("yield_mercado_pct", 0)
                _var_doc = r.get("variacion_anual_pct", 0)
                _cuota_doc = r.get("cuota_mensual", 0)
                _pie_doc = r.get("pie", 0)
                _ingreso_min_doc = r.get("ingreso_minimo", 0)

                _NAV = "#1E2D3D"
                _GOLD = "#B8904A"

                _posicion = ("por debajo del mercado — precio competitivo" if _diff_doc < -5
                             else ("en línea con el mercado" if abs(_diff_doc) <= 8
                                   else f"{_diff_doc:.1f}% sobre la mediana de zona"))

                _doc_html = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8">
<style>
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#EDEAE4;margin:0;padding:32px;color:{_NAV};}}
  .page{{background:#FFFFFF;max-width:820px;margin:0 auto;padding:48px 52px;border-radius:8px;box-shadow:0 4px 24px rgba(0,0,0,0.08);}}
  .header{{background:linear-gradient(135deg,#1E2D3D,#243850);padding:28px 32px;border-radius:6px;margin-bottom:32px;}}
  .header-title{{font-size:24px;font-weight:800;color:#FFFFFF;letter-spacing:-0.5px;}}
  .header-sub{{font-size:11px;color:{_GOLD};letter-spacing:3px;text-transform:uppercase;font-weight:600;margin-bottom:10px;}}
  .header-meta{{font-size:11px;color:#8AA8C0;margin-top:8px;}}
  .section-title{{font-size:9px;color:#9A9080;letter-spacing:2.5px;text-transform:uppercase;font-weight:700;border-bottom:1px solid #D8D4CC;padding-bottom:6px;margin:28px 0 14px;}}
  .metric-row{{display:flex;gap:20px;flex-wrap:wrap;margin-bottom:20px;}}
  .metric-box{{flex:1;min-width:140px;background:#F9F7F4;border:1px solid #E4E0D8;border-top:3px solid {_GOLD};border-radius:6px;padding:14px 16px;}}
  .metric-label{{font-size:9px;color:#9A9080;letter-spacing:1.5px;text-transform:uppercase;font-weight:600;margin-bottom:4px;}}
  .metric-value{{font-size:20px;font-weight:700;color:{_NAV};letter-spacing:-0.5px;}}
  .alert{{border-left:4px solid {_GOLD};background:#FFFBF3;border-radius:4px;padding:12px 16px;margin:12px 0;font-size:12px;color:#5C3D10;line-height:1.6;}}
  .market-box{{background:#F0F4F8;border:1px solid #C8D4DE;border-radius:6px;padding:16px 20px;margin:12px 0;}}
  table{{width:100%;border-collapse:collapse;margin:12px 0;}}
  th{{background:{_NAV};color:#FFFFFF;padding:9px 14px;font-size:10px;text-transform:uppercase;letter-spacing:1px;text-align:left;}}
  td{{padding:9px 14px;font-size:12px;border-bottom:1px solid #E8E4DC;color:{_NAV};}}
  tr:nth-child(even) td{{background:#F9F7F4;}}
  .footer{{margin-top:40px;padding-top:16px;border-top:1px solid #D8D4CC;font-size:10px;color:#9A9080;text-align:center;line-height:1.6;}}
  .gold{{color:{_GOLD};font-weight:700;}}
  .verde{{color:#1A4731;font-weight:700;}}
  .rojo{{color:#7A1A1A;font-weight:700;}}
  @media print{{body{{background:white;padding:0;}} .page{{box-shadow:none;}}}}
</style>
</head>
<body>
<div class="page">

<div class="header">
  <div class="header-sub">Osterling Advisory · FACTIS</div>
  <div class="header-title">{_doc_tipo}</div>
  <div class="header-meta">
    {'Preparado por: ' + _doc_agente + ' · ' if _doc_agente else ''}Zona: {_zona_doc} · Fecha: {_fecha_doc}
    {' · Para: ' + _doc_cliente if _doc_cliente else ''}
  </div>
</div>

<div class="section-title">I. Resumen del Inmueble</div>
<div class="metric-row">
  <div class="metric-box"><div class="metric-label">Precio</div><div class="metric-value">${_precio_doc:,}</div></div>
  <div class="metric-box"><div class="metric-label">Área</div><div class="metric-value">{_m2_doc} m²</div></div>
  <div class="metric-box"><div class="metric-label">USD / m²</div><div class="metric-value">${_ppm2_doc:,.0f}</div></div>
  <div class="metric-box"><div class="metric-label">Zona</div><div class="metric-value" style="font-size:14px;">{_zona_doc}</div></div>
</div>

<div class="section-title">II. Posición de Mercado</div>
<div class="market-box">
  <table>
    <tr><th>Indicador</th><th>Este Inmueble</th><th>Mercado Zona (Urbania nov-25)</th><th>Diferencial</th></tr>
    <tr>
      <td>Precio / m²</td>
      <td><strong>${_ppm2_doc:,.0f}/m²</strong></td>
      <td>${_ref_m2_doc:,}/m²</td>
      <td class="{'verde' if _diff_doc <= 0 else 'rojo'}">{_diff_doc:+.1f}%</td>
    </tr>
    {"" if not _alq_doc else f'<tr><td>Alquiler mensual</td><td><strong>${_alq_doc:,}/mes</strong></td><td>${_alq_mkt_doc:,.0f}/mes</td><td class="' + ("verde" if _alq_doc >= _alq_mkt_doc else "rojo") + f'">{((_alq_doc-_alq_mkt_doc)/_alq_mkt_doc*100) if _alq_mkt_doc > 0 else 0:+.1f}% vs. mercado</td></tr>'}
    {"" if not _yield_doc else f'<tr><td>Yield bruto anual</td><td><strong>{_yield_doc:.1f}%</strong></td><td>{_yield_mkt_doc:.1f}%</td><td class="' + ("verde" if _yield_doc >= _yield_mkt_doc else "rojo") + f'">{(_yield_doc-_yield_mkt_doc):+.1f}pp</td></tr>'}
    <tr>
      <td>Tendencia anual zona</td>
      <td colspan="2">{_var_doc:+.1f}% en 12 meses (Urbania Lima Index)</td>
      <td class="{'verde' if _var_doc >= 0 else 'rojo'}">{'↑ En alza' if _var_doc >= 2 else ('→ Estable' if abs(_var_doc) < 2 else '↓ En baja')}</td>
    </tr>
  </table>
</div>

<div class="alert">
  <strong>Posición de precio:</strong> El inmueble se encuentra <strong>{_posicion}</strong>.
  {"La variación anual de la zona (" + str(_var_doc) + "%) indica " + ("potencial de plusvalía." if _var_doc > 2 else ("mercado estable." if abs(_var_doc) <= 2 else "presión a la baja — oportunidad de negociación."))}
</div>

{'<div class="section-title">III. Análisis Financiero del Comprador</div><div class="metric-row"><div class="metric-box"><div class="metric-label">Pago inicial (' + str(int(r.get("pct_pie", 0))) + '%)</div><div class="metric-value">$' + f"{_pie_doc:,.0f}" + '</div></div><div class="metric-box"><div class="metric-label">Cuota mensual</div><div class="metric-value">$' + f"{_cuota_doc:,.0f}" + '</div></div><div class="metric-box"><div class="metric-label">Ingreso mínimo</div><div class="metric-value">$' + f"{_ingreso_min_doc:,.0f}" + '/mes</div></div></div>' if _cuota_doc > 0 else ""}

{'<div class="section-title">IV. Rentabilidad de la Inversión</div><div class="metric-row"><div class="metric-box"><div class="metric-label">Yield Bruto</div><div class="metric-value">' + f"{_yield_doc:.1f}%" + '</div></div><div class="metric-box"><div class="metric-label">Yield Neto</div><div class="metric-value">' + f"{r.get('yield_neto',0):.1f}%" + '</div></div><div class="metric-box"><div class="metric-label">Payback</div><div class="metric-value">' + (f"{r.get('payback_anos',0):.1f} años" if r.get('payback_anos') else "N/A") + '</div></div><div class="metric-box"><div class="metric-label">Yield mercado zona</div><div class="metric-value">' + f"{_yield_mkt_doc:.1f}%" + '</div></div></div>' if _yield_doc > 0 else ""}

<div class="section-title">V. Proyección de Valor (tasa zona: {r.get("tasa_apreciacion_pct", 4.0):.1f}%/año)</div>
<div class="metric-row">
  <div class="metric-box"><div class="metric-label">Valor a 5 años</div><div class="metric-value">${r.get('valor_5', 0):,.0f}</div></div>
  <div class="metric-box"><div class="metric-label">Ganancia capital 5a</div><div class="metric-value">+${r.get('ganancia_capital_5', 0):,.0f}</div></div>
  <div class="metric-box"><div class="metric-label">Valor a 10 años</div><div class="metric-value">${r.get('valor_10', 0):,.0f}</div></div>
</div>

{('<div class="section-title">VI. Observaciones</div><div class="alert">' + _doc_obs + '</div>') if _doc_obs else ""}

<div class="footer">
  Documento generado por FACTIS · Osterling Advisory<br>
  Los valores de mercado corresponden al Índice Urbania Lima — Noviembre 2025 · Tipo de cambio SUNAT: 3.42 S./USD<br>
  {_fecha_doc}<br><br>
  <span style="font-size:9px;color:#B8AA9A;">
  NOTA: Esta IA de Análisis Inmobiliario debe utilizarse como herramienta complementaria al criterio profesional,
  permitiendo obtener resultados preliminares de manera rápida. Como paso final, el profesional podrá terminar de
  definir las tipologías, distribución y las modificaciones pertinentes. La IA irá volviéndose más responsiva y
  alineada con la visión del profesional a medida que se retroalimenta con sus decisiones.
  </span>
</div>

</div>
</body>
</html>"""

                st.session_state["_res_doc_html"] = _doc_html
                st.success("Documento generado.")

            if st.session_state.get("_res_doc_html"):
                _doc_proy = st.session_state.get("nombre_proyecto") or "Proyecto"
                _doc_filename = f"{_doc_tipo} - {_doc_proy}.html"
                st.download_button(
                    f"DESCARGAR {_doc_tipo.upper()}",
                    data=st.session_state["_res_doc_html"].encode("utf-8"),
                    file_name=_doc_filename,
                    mime="text/html",
                    use_container_width=True,
                )
                st.caption("Abre en tu navegador y usa Cmd+P para imprimir a PDF o compartir.")

                # Preview
                st.markdown("---")
                st.markdown('<div class="section-title">Vista Previa</div>', unsafe_allow_html=True)
                st.components.v1.html(st.session_state["_res_doc_html"], height=700, scrolling=True)

        st.markdown("""
<div style="border-top:1px solid #E8E0D4;margin-top:32px;padding-top:14px;">
  <div style="font-size:9px;color:#9A8A7A;line-height:1.7;text-align:center;">
    <span style="color:#B8904A;font-weight:700;letter-spacing:1px;">NOTA · </span>
    Esta IA de Análisis Inmobiliario debe utilizarse como herramienta complementaria al criterio profesional,
    permitiendo obtener resultados preliminares de manera rápida. Como paso final, el profesional podrá
    terminar de definir las tipologías, distribución por plantas y las modificaciones que considere pertinentes.
    La IA irá volviéndose más responsiva y alineada con la visión del profesional a medida que se retroalimenta
    con sus decisiones.
  </div>
</div>""", unsafe_allow_html=True)

    else:
        st.markdown(
            '<div style="border-radius:8px;min-height:420px;'
            'background:linear-gradient(160deg,#1A2737 0%,#1E2D3D 60%,#1A2737 100%);'
            'display:flex;align-items:center;justify-content:center;'
            'box-shadow:0 8px 32px rgba(30,45,61,0.18);padding:64px 48px;">'
            '<div style="max-width:600px;width:100%;text-align:center;">'
            '<div style="font-size:9px;color:#B8904A;letter-spacing:4px;text-transform:uppercase;font-weight:600;margin-bottom:16px;">Osterling Advisory</div>'
            '<div style="font-size:28px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;margin-bottom:8px;">Análisis Residencial</div>'
            '<div style="width:48px;height:2px;background:#B8904A;margin:16px auto;"></div>'
            '<div style="font-size:13px;color:#B0C0D0;line-height:1.7;margin-bottom:32px;">'
            'Evalúa la compra de un inmueble residencial ya sea para uso propio o como inversión. '
            'Obtén la cuota mensual exacta, el ingreso mínimo recomendado para calificar al crédito, '
            'rentabilidad por alquiler, payback y proyección de apreciación de capital.'
            '</div>'
            '<div style="display:flex;gap:24px;justify-content:center;flex-wrap:wrap;">'
            '<div style="background:rgba(184,144,74,0.1);border:1px solid rgba(184,144,74,0.3);'
            'border-radius:6px;padding:14px 20px;min-width:120px;">'
            '<div style="font-size:18px;font-weight:700;color:#B8904A;">Cuota</div>'
            '<div style="font-size:10px;color:#8AA8C0;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px;">Hipotecaria Exacta</div>'
            '</div>'
            '<div style="background:rgba(184,144,74,0.1);border:1px solid rgba(184,144,74,0.3);'
            'border-radius:6px;padding:14px 20px;min-width:120px;">'
            '<div style="font-size:18px;font-weight:700;color:#B8904A;">Yield</div>'
            '<div style="font-size:10px;color:#8AA8C0;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px;">Bruto y Neto</div>'
            '</div>'
            '<div style="background:rgba(184,144,74,0.1);border:1px solid rgba(184,144,74,0.3);'
            'border-radius:6px;padding:14px 20px;min-width:120px;">'
            '<div style="font-size:18px;font-weight:700;color:#B8904A;">+4%</div>'
            '<div style="font-size:10px;color:#8AA8C0;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px;">Apreciación Lima</div>'
            '</div>'
            '</div>'
            '</div></div>',
            unsafe_allow_html=True
        )

# ═══════════════════════════════════════════════════════
# MÓDULO 5: PORTFOLIO
# ═══════════════════════════════════════════════════════
elif tipo_op == "Portfolio":
    st.markdown(
        '<div style="font-size:9px;color:#B8904A;letter-spacing:4px;text-transform:uppercase;'
        'font-weight:600;margin-bottom:4px;">Osterling Advisory</div>'
        '<div style="font-size:26px;font-weight:700;color:#FFFFFF;letter-spacing:-0.5px;">'
        '📁 Portfolio de Proyectos</div>'
        '<div style="font-size:13px;color:#B0C0D0;margin-top:6px;margin-bottom:20px;">'
        'Todos los proyectos guardados con KPIs consolidados e historial de versiones.</div>',
        unsafe_allow_html=True)

    _port_proyectos = listar_proyectos(con_resumen=True)

    if not _port_proyectos:
        st.info("No hay proyectos guardados aún. Analiza un proyecto y guárdalo para verlo aquí.")
    else:
        # Filtros
        _pf1, _pf2, _pf3 = st.columns([2, 1, 1])
        _port_tipos = sorted({getattr(p, "_tipo", "") for p in _port_proyectos if getattr(p, "_tipo", "")})
        _port_filtro_tipo = _pf1.selectbox("Filtrar por tipo", ["Todos"] + _port_tipos, key="port_filtro_tipo")
        _port_orden = _pf2.selectbox("Ordenar por", ["Fecha", "Margen", "TIR", "Utilidad"], key="port_orden")
        _port_buscar = _pf3.text_input("Buscar proyecto", key="port_buscar", placeholder="nombre...")

        _port_vis = [p for p in _port_proyectos if hasattr(p, "_resumen")]
        if _port_filtro_tipo != "Todos":
            _port_vis = [p for p in _port_vis if getattr(p, "_tipo", "") == _port_filtro_tipo]
        if _port_buscar:
            _port_vis = [p for p in _port_vis if _port_buscar.lower() in getattr(p, "_nombre", "").lower()]

        if _port_orden == "Margen":
            _port_vis.sort(key=lambda p: float(getattr(p, "_resumen", {}).get("margen_pct", 0) or 0), reverse=True)
        elif _port_orden == "TIR":
            _port_vis.sort(key=lambda p: float(getattr(p, "_resumen", {}).get("tir_anual_pct", 0) or 0), reverse=True)
        elif _port_orden == "Utilidad":
            _port_vis.sort(key=lambda p: float(getattr(p, "_resumen", {}).get("utilidad_neta", 0) or 0), reverse=True)

        st.markdown(f"**{len(_port_vis)} proyectos**")
        st.markdown("---")

        # Cards grid — 2 columnas
        _GOLD = "#B8904A"
        _DARK = "#0A1628"
        for i in range(0, len(_port_vis), 2):
            _pcols = st.columns(2, gap="medium")
            for j, p in enumerate(_port_vis[i:i+2]):
                _rs  = getattr(p, "_resumen", {})
                _tip = getattr(p, "_tipo", "—")
                _zon = getattr(p, "_zona", "—")
                _fec = getattr(p, "_fecha", "—")
                _nom = getattr(p, "_nombre", p.name)

                _mg  = _rs.get("margen_pct") or _rs.get("margen_neto") or 0
                _tir = _rs.get("tir_anual_pct") or 0
                _un  = _rs.get("utilidad_neta") or 0
                _ing = _rs.get("ingresos_brutos") or _rs.get("costo_total") or 0

                _mg_color = "#4CAF50" if float(_mg) >= 15 else ("#FFC107" if float(_mg) >= 10 else "#FF4444")
                _tipo_color = {"inmobiliario": "#1A3A6B", "industrial": "#6B3A1A", "residencial": "#1A6B3A"}.get(_tip, "#444")

                with _pcols[j]:
                    st.markdown(
                        f'<div style="background:linear-gradient(135deg,#1A2737,#1E2D3D);'
                        f'border-radius:12px;padding:18px 20px;border:1px solid rgba(184,144,74,0.2);'
                        f'margin-bottom:4px;">'
                        f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:10px;">'
                        f'<div style="font-size:13px;font-weight:700;color:#FFFFFF;line-height:1.3;max-width:70%;">{_nom}</div>'
                        f'<div style="font-size:9px;font-weight:700;color:#FFF;background:{_tipo_color};'
                        f'padding:3px 8px;border-radius:4px;text-transform:uppercase;letter-spacing:1px;">{_tip}</div>'
                        f'</div>'
                        f'<div style="font-size:10px;color:#8AA8C0;margin-bottom:12px;">{_zon} · {_fec}</div>'
                        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">'
                        f'<div style="background:rgba(255,255,255,0.04);border-radius:6px;padding:8px 10px;">'
                        f'<div style="font-size:16px;font-weight:800;color:{_mg_color};">{float(_mg):.1f}%</div>'
                        f'<div style="font-size:9px;color:#8AA8C0;text-transform:uppercase;letter-spacing:1px;">Margen neto</div>'
                        f'</div>'
                        f'<div style="background:rgba(255,255,255,0.04);border-radius:6px;padding:8px 10px;">'
                        f'<div style="font-size:16px;font-weight:800;color:#D4A853;">{float(_tir):.1f}%</div>'
                        f'<div style="font-size:9px;color:#8AA8C0;text-transform:uppercase;letter-spacing:1px;">TIR anual</div>'
                        f'</div>'
                        f'<div style="background:rgba(255,255,255,0.04);border-radius:6px;padding:8px 10px;">'
                        f'<div style="font-size:14px;font-weight:700;color:#FFFFFF;">${float(_un):,.0f}</div>'
                        f'<div style="font-size:9px;color:#8AA8C0;text-transform:uppercase;letter-spacing:1px;">Utilidad neta</div>'
                        f'</div>'
                        f'<div style="background:rgba(255,255,255,0.04);border-radius:6px;padding:8px 10px;">'
                        f'<div style="font-size:14px;font-weight:700;color:#FFFFFF;">${float(_ing):,.0f}</div>'
                        f'<div style="font-size:9px;color:#8AA8C0;text-transform:uppercase;letter-spacing:1px;">Ingresos / Costo</div>'
                        f'</div>'
                        f'</div></div>',
                        unsafe_allow_html=True)

        # ── Tabla comparativa de todos los proyectos ──
        if len(_port_vis) >= 2:
            st.markdown("---")
            st.markdown("#### Tabla comparativa")
            _port_df_rows = []
            for p in _port_vis:
                _rs = getattr(p, "_resumen", {})
                _port_df_rows.append({
                    "Proyecto":    getattr(p, "_nombre", p.name),
                    "Tipo":        getattr(p, "_tipo", "—"),
                    "Zona":        getattr(p, "_zona", "—"),
                    "Fecha":       getattr(p, "_fecha", "—"),
                    "Margen (%)":  round(float(_rs.get("margen_pct") or 0), 1),
                    "TIR (%)":     round(float(_rs.get("tir_anual_pct") or 0), 1),
                    "Util. Neta":  f"${float(_rs.get('utilidad_neta') or 0):,.0f}",
                    "Ingresos":    f"${float(_rs.get('ingresos_brutos') or 0):,.0f}",
                })
            st.dataframe(pd.DataFrame(_port_df_rows), hide_index=True, use_container_width=True)

